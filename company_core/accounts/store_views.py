from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from urllib.parse import urlparse
import re

from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import json
import requests

from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.core.cache import cache
from django.core.paginator import Paginator
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from django.db.models import (
    Q,
    Sum,
    Count,
    Max,
    Min,
    OuterRef,
    Subquery,
    Value,
    ExpressionWrapper,
    F,
    DecimalField,
    Case,
    When,
)
from django.db.models.functions import Coalesce, TruncMonth, Cast
from django.forms import modelformset_factory, inlineformset_factory
from django.template.loader import render_to_string
from django.utils import timezone
from django.urls import reverse, resolve
from django.urls.exceptions import Resolver404
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_GET, require_POST

try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False
    HTML = None
    CSS = None
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .decorators import customer_login_required
from .models import (
        Product,
        ProductStock,
        StorefrontCartItem,
        ProductAlternateSku,
        Category,
        CategoryGroup,
        CategoryAttribute,
        ProductBrand,
        ProductAttributeValue,
        StorefrontHeroShowcase,
        StorefrontHeroShowcaseItem,
        StorefrontHeroPackage,
        StorefrontMessageBanner,
        StorefrontFlyer,
        ProductModel,
        ProductVin,
        Customer,
        GroupedInvoice,
        IncomeRecord2,
        PendingInvoice,
        Payment,
        CustomerCredit,
        CustomerCreditItem,
        PAYMENT_LINK_PROVIDER_STRIPE,
        PAYMENT_LINK_PROVIDER_CLOVER,
        PAYMENT_LINK_PROVIDER_NONE,
        WorkOrder,
        Vehicle,
        VehicleMaintenanceTask,
        VehicleServicePortalOverride,
        JobHistory,
        Service,
        PublicBooking,
        PublicContactMessage,
        InvoiceActivity,
        calculate_tax_total,
    )
from . import paid_invoice_views
from .view_invoices import send_grouped_invoice_email, _build_invoice_context, _render_pdf
from .invoice_activity import log_invoice_activity
from .forms import (
    CustomerPortalProfileForm,
    VehicleForm,
    VehicleMaintenanceTaskForm,
    VehicleMaintenanceCompleteForm,
    CustomerStatementForm,
    CustomerPortalQuickMaintenanceForm,
    StorefrontHeroShowcaseForm,
    StorefrontHeroShowcaseItemForm,
    StorefrontHeroPackageForm,
    StorefrontMessageBannerForm,
    StorefrontFlyerForm,
    StorefrontPriceVisibilityForm,
)
from .utils import (
    apply_stock_fields,
    annotate_products_with_stock,
    get_business_user,
    get_customer_user_ids,
    get_product_user_ids,
    get_stock_owner,
    get_storefront_owner,
    get_storefront_profiles,
    is_parts_store_business,
    resolve_storefront_root_user,
    resolve_company_logo_url,
    resolve_storefront_category_flags,
    resolve_storefront_price_flags,
)
from .pdf_utils import apply_branding_defaults, render_template_to_pdf


STATEMENT_PDF_CSS = CSS(
    string='''
        @page { size: Letter; margin: 1.25cm; }
        body { font-family: "Helvetica", "Arial", sans-serif; font-size: 12px; color: #1f2933; }
        h1, h2, h3 { color: #0f172a; }
        table { width: 100%; border-collapse: collapse; margin-top: 18px; }
        th, td { padding: 8px 10px; border-bottom: 1px solid #d2d6dc; text-align: left; }
        th { background-color: #f1f5f9; font-weight: 600; }
        tfoot td { font-weight: 600; }
    '''
) if WEASYPRINT_AVAILABLE else None


def _customer_portal_nav_items(request, active_slug):
    """Return ordered navigation items for the customer portal layout."""

    is_parts_store = is_parts_store_business()
    dashboard_url = reverse('accounts:customer_dashboard')

    if is_parts_store:
        nav_items = [
            {
                'slug': 'shop',
                'label': 'Shop',
                'icon': 'fa-store',
                'url': reverse('accounts:store_product_list'),
                'is_external': True,
            },
            {
                'slug': 'invoices',
                'label': 'Invoices',
                'icon': 'fa-file-invoice-dollar',
                'url': reverse('accounts:customer_invoice_list'),
            },
            {
                'slug': 'returns',
                'label': 'Returns',
                'icon': 'fa-undo',
                'url': reverse('accounts:customer_returns'),
            },
            {
                'slug': 'statements',
                'label': 'Statements',
                'icon': 'fa-file-lines',
                'url': reverse('accounts:customer_invoice_statements'),
            },
            {
                'slug': 'settlements',
                'label': 'Balance',
                'icon': 'fa-scale-balanced',
                'url': reverse('accounts:customer_settlement_summary'),
            },
            {
                'slug': 'profile',
                'label': 'Settings',
                'icon': 'fa-gear',
                'url': reverse('accounts:customer_profile'),
            },
            {
                'slug': 'logout',
                'label': 'Logout',
                'icon': 'fa-arrow-right-from-bracket',
                'url': f"{reverse('accounts:logout')}?next={reverse('accounts:store_product_list')}",
            },
        ]
        active_slug = 'shop' if active_slug == 'overview' else active_slug
    else:
        nav_items = [
            {
                'slug': 'overview',
                'label': 'Home',
                'icon': 'fa-house',
                'url': dashboard_url,
            },
            {
                'slug': 'invoices',
                'label': 'Invoices',
                'icon': 'fa-file-invoice-dollar',
                'url': reverse('accounts:customer_invoice_list'),
            },
            {
                'slug': 'returns',
                'label': 'Returns',
                'icon': 'fa-undo',
                'url': reverse('accounts:customer_returns'),
            },
            {
                'slug': 'workorders',
                'label': 'Workorders',
                'icon': 'fa-screwdriver-wrench',
                'url': reverse('accounts:customer_workorder_list'),
            },
            {
                'slug': 'vehicles',
                'label': 'Vehicles',
                'icon': 'fa-truck-front',
                'url': reverse('accounts:customer_vehicle_overview'),
            },
            {
                'slug': 'maintenance',
                'label': 'Maintenance',
                'icon': 'fa-calendar-check',
                'url': reverse('accounts:customer_maintenance_list'),
            },
            {
                'slug': 'settlements',
                'label': 'Balances',
                'icon': 'fa-scale-unbalanced',
                'url': reverse('accounts:customer_settlement_summary'),
            },
            {
                'slug': 'statements',
                'label': 'Statements',
                'icon': 'fa-file-lines',
                'url': reverse('accounts:customer_invoice_statements'),
            },
            {
                'slug': 'shop',
                'label': 'Shop products',
                'icon': 'fa-store',
                'url': reverse('accounts:store_product_list'),
                'is_external': True,
            },
        ]

    for item in nav_items:
        item['active'] = item['slug'] == active_slug

    return nav_items


def _customer_portal_layout_context(request, active_slug):
    """Shared context variables for templates that use the customer portal shell."""

    customer_account = getattr(request.user, 'customer_portal', None)
    is_parts_store = is_parts_store_business()
    return {
        'customer_account': customer_account,
        'portal_nav_items': _customer_portal_nav_items(request, active_slug),
        'portal_active_item': active_slug,
        'portal_is_storefront': is_parts_store,
        'portal_base_template': (
            'store/customer_portal_storefront_base.html'
            if is_parts_store
            else 'store/customer_portal_base.html'
        ),
    }


def _annotate_invoice_credit_totals(queryset):
    amount_field = DecimalField(max_digits=10, decimal_places=2)
    credit_amount_expr = Case(
        When(customer_credit__tax_included=True, then=Cast('amount', amount_field)),
        default=ExpressionWrapper(
            Cast('amount', amount_field) + Cast('tax_paid', amount_field),
            output_field=amount_field,
        ),
        output_field=amount_field,
    )
    credit_total = CustomerCreditItem.objects.filter(
        source_invoice=OuterRef('pk')
    ).values('source_invoice').annotate(
        total=Coalesce(
            Sum(credit_amount_expr),
            Value(Decimal('0.00')),
            output_field=amount_field,
        )
    ).values('total')
    return queryset.annotate(
        credit_total=Coalesce(
            Subquery(credit_total, output_field=amount_field),
            Value(Decimal('0.00')),
            output_field=amount_field,
        )
    )


def _get_customer_credit_total(customer_account):
    amount_field = DecimalField(max_digits=10, decimal_places=2)
    credit_amount_expr = Case(
        When(customer_credit__tax_included=True, then=Cast('amount', amount_field)),
        default=ExpressionWrapper(
            Cast('amount', amount_field) + Cast('tax_paid', amount_field),
            output_field=amount_field,
        ),
        output_field=amount_field,
    )
    total = CustomerCreditItem.objects.filter(
        customer_credit__customer=customer_account,
        customer_credit__user=customer_account.user,
    ).aggregate(
        total=Coalesce(
            Sum(credit_amount_expr),
            Value(Decimal('0.00')),
            output_field=amount_field,
        )
    )['total']
    return total or Decimal('0.00')


def _build_customer_credit_rows(customer_account, *, start_date=None, end_date=None, limit=None):
    credits_qs = CustomerCredit.objects.filter(
        user=customer_account.user,
        customer=customer_account,
    ).prefetch_related(
        'items',
        'items__product',
        'items__source_invoice',
    ).order_by('-date', '-id')
    if start_date:
        credits_qs = credits_qs.filter(date__gte=start_date)
    if end_date:
        credits_qs = credits_qs.filter(date__lte=end_date)
    if limit:
        credits_qs = credits_qs[:limit]

    credit_rows = []
    total_credit = Decimal('0.00')
    for credit in credits_qs:
        total_incl_tax, total_tax, total_excl_tax = credit.calculate_totals()
        total_incl_tax = Decimal(str(total_incl_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_tax = Decimal(str(total_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_excl_tax = Decimal(str(total_excl_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_credit += total_incl_tax
        item_rows = []
        for item in credit.items.all():
            label = (item.description or '').strip()
            if not label:
                label = (item.part_no or '').strip()
            if not label and item.product:
                label = (item.product.name or '').strip()
            if not label:
                label = 'Item'
            item_rows.append({
                'qty': item.qty or 0,
                'label': label,
                'invoice_number': item.source_invoice.invoice_number if item.source_invoice else None,
            })
        credit_rows.append({
            'credit': credit,
            'items': item_rows,
            'total_incl_tax': total_incl_tax,
            'total_tax': total_tax,
            'total_excl_tax': total_excl_tax,
        })

    return credit_rows, total_credit


def _summarize_customer_invoices(customer_account, *, today=None):
    """Build summary stats and recent invoices for customer portal views."""

    today = today or timezone.localdate()
    invoices_qs = customer_account.invoices.all()
    invoice_totals = invoices_qs.aggregate(
        total_invoiced=Coalesce(Sum('total_amount'), Decimal('0.00')),
    )
    payment_totals = Payment.objects.filter(invoice__customer=customer_account).aggregate(
        total_paid=Coalesce(Sum('amount'), Decimal('0.00')),
    )

    total_invoiced = invoice_totals['total_invoiced'] or Decimal('0.00')
    total_paid = payment_totals['total_paid'] or Decimal('0.00')
    total_credit = _get_customer_credit_total(customer_account)
    outstanding_balance = total_invoiced - total_paid - total_credit

    invoices_qs = (
        customer_account.invoices.select_related('user')
        .prefetch_related('payments')
        .order_by('-date', '-id')
    )
    all_invoices = list(invoices_qs)
    recent_invoices = all_invoices[:3]

    overdue_balance = Decimal('0.00')
    overdue_count = 0
    open_count = 0
    for invoice in all_invoices:
        balance_due = invoice.balance_due()
        if balance_due <= Decimal('0.00'):
            continue
        open_count += 1
        due_date = invoice.due_date
        if due_date and due_date < today:
            overdue_count += 1
            overdue_balance += balance_due

    invoice_summary = {
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'total_credit': total_credit,
        'outstanding_balance': outstanding_balance,
        'invoice_count': len(all_invoices),
        'open_count': open_count,
        'overdue_count': overdue_count,
        'overdue_balance': overdue_balance,
        'pending_count': PendingInvoice.objects.filter(
            grouped_invoice__customer=customer_account
        ).count(),
    }

    return invoice_summary, recent_invoices


def _customer_invoice_pdf_response(invoice, request, *, inline=False):
    """Shared helper used for customer-portal invoice downloads/prints."""

    if invoice.payment_status == 'Paid':
        context = paid_invoice_views._invoice_context(invoice, request=request)
        template = 'invoices/paid_invoice_pdf.html'
        if not WEASYPRINT_AVAILABLE:
            raise ImportError("WeasyPrint is not available. PDF generation is disabled. Please install GTK+ libraries for Windows.")
        context = apply_branding_defaults(context)
        pdf_bytes = HTML(string=render_to_string(template, context)).write_pdf(
            stylesheets=[paid_invoice_views.PDF_CSS]
        )
        filename = f'paid_{invoice.invoice_number}.pdf'
    else:
        context = _build_invoice_context(invoice, request, profile=getattr(invoice.user, 'profile', None))
        pdf_bytes = _render_pdf('invoices/grouped_invoice_pdf.html', context)
        filename = f'Invoice_{invoice.invoice_number}.pdf'

    disposition = 'inline' if inline else 'attachment'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
    log_invoice_activity(
        invoice,
        event_type=InvoiceActivity.EVENT_VIEWED,
        request=request,
    )
    return response


def _customer_workorder_pdf_response(workorder, request):
    """Render a work order PDF for customers linked to the portal."""

    profile = getattr(workorder.user, 'profile', None)
    subtotal = workorder.records.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    company_logo_url = (
        resolve_company_logo_url(profile, request=request, for_pdf=True)
        if profile else None
    )
    context = {
        'workorder': workorder,
        'profile': profile,
        'subtotal': subtotal,
        'company_logo_url': company_logo_url,
    }
    context = apply_branding_defaults(context)
    html_string = render_to_string('workorders/workorder_pdf.html', context)
    if not WEASYPRINT_AVAILABLE:
        raise ImportError("WeasyPrint is not available. PDF generation is disabled. Please install GTK+ libraries for Windows.")
    pdf_bytes = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="WorkOrder_{workorder.id}.pdf"'
    return response


def _statement_logo_url(profile, request=None, *, for_pdf=False):
    """
    Return a logo URL for customer statements.

    Uses the same centralized logic as invoices/workorders, including a static
    fallback (2G logo) when no profile logo is available.
    """

    from .utils import resolve_company_logo_url

    if not profile:
        return ''
    return resolve_company_logo_url(profile, request=request, for_pdf=for_pdf) or ''


def _resolve_statement_period(period_key, reference_date):
    """Return (start, end, label) for the requested statement period."""

    reference_date = reference_date or timezone.localdate()

    if period_key == 'week':
        start = reference_date - timedelta(days=reference_date.weekday())
        end = start + timedelta(days=6)
        label = f"Week of {start:%b %d, %Y}"
    elif period_key == 'quarter':
        quarter = ((reference_date.month - 1) // 3) + 1
        start_month = 3 * (quarter - 1) + 1
        end_month = start_month + 2
        start = date(reference_date.year, start_month, 1)
        end_day = monthrange(reference_date.year, end_month)[1]
        end = date(reference_date.year, end_month, end_day)
        label = f"Q{quarter} {reference_date.year}"
    elif period_key == 'semiannual':
        if reference_date.month <= 6:
            start = date(reference_date.year, 1, 1)
            end = date(reference_date.year, 6, monthrange(reference_date.year, 6)[1])
            label = f"First half {reference_date.year}"
        else:
            start = date(reference_date.year, 7, 1)
            end = date(reference_date.year, 12, 31)
            label = f"Second half {reference_date.year}"
    elif period_key == 'annual':
        start = date(reference_date.year, 1, 1)
        end = date(reference_date.year, 12, 31)
        label = f"Fiscal year {reference_date.year}"
    else:  # month
        start = reference_date.replace(day=1)
        end = date(reference_date.year, reference_date.month, monthrange(reference_date.year, reference_date.month)[1])
        label = reference_date.strftime('%B %Y')

    return start, end, label


def _statement_totals(invoices):
    subtotal = sum((invoice.subtotal for invoice in invoices), Decimal('0.00'))
    tax = sum((invoice.tax_total for invoice in invoices), Decimal('0.00'))
    total = sum((invoice.total_amount for invoice in invoices), Decimal('0.00'))
    paid = sum(
        (invoice.total_paid or Decimal('0.00') for invoice in invoices),
        Decimal('0.00'),
    )

    return {
        'count': len(invoices),
        'subtotal': subtotal,
        'tax': tax,
        'total': total,
        'paid': paid,
    }


def _statement_pdf_response(customer_account, invoices, *, start_date, end_date, period_label, totals, request):
    profile = getattr(customer_account.user, 'profile', None)
    owner = getattr(customer_account, 'user', None)
    if profile and getattr(profile, 'company_name', None):
        business_name = profile.company_name
    elif owner:
        business_name = owner.get_full_name() or owner.get_username()
    else:
        business_name = customer_account.name
    context = {
        'customer': customer_account,
        'profile': profile,
        'invoices': invoices,
        'start_date': start_date,
        'end_date': end_date,
        'period_label': period_label,
        'totals': totals,
        'generated_on': timezone.now(),
        'logo_url': _statement_logo_url(profile, request=request),
        'logo_url_pdf': _statement_logo_url(profile, for_pdf=True),
        'business_name': business_name,
    }
    context = apply_branding_defaults(context)
    html = render_to_string('invoices/customer_statement_pdf.html', context)
    if not WEASYPRINT_AVAILABLE:
        raise ImportError("WeasyPrint is not available. PDF generation is disabled. Please install GTK+ libraries for Windows.")
    pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf(
        stylesheets=[STATEMENT_PDF_CSS]
    )
    filename = f'statement_paid_invoices_{start_date:%Y%m%d}_{end_date:%Y%m%d}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _statement_excel_response(customer_account, invoices, *, start_date, end_date, period_label, totals):
    profile = getattr(customer_account.user, 'profile', None)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Paid invoices'

    row = 1
    if profile and getattr(profile, 'company_name', None):
        cell = worksheet.cell(row=row, column=1, value=profile.company_name)
        cell.font = Font(bold=True, size=14)
        row += 1

    worksheet.cell(row=row, column=1, value=f'Paid invoice statement for {customer_account.name}').font = Font(bold=True)
    row += 1
    worksheet.cell(row=row, column=1, value=f'Period: {period_label} ({start_date:%b %d, %Y} – {end_date:%b %d, %Y})')
    row += 1
    worksheet.cell(row=row, column=1, value=f'Generated on: {timezone.now():%b %d, %Y %I:%M %p}')
    row += 2

    headers = ['Invoice #', 'Date', 'Subtotal', 'Tax', 'Total', 'Paid on', 'Payment method']
    for idx, heading in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=idx, value=heading)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for invoice in invoices:
        worksheet.cell(row=row, column=1, value=invoice.invoice_number)
        worksheet.cell(row=row, column=2, value=invoice.date)
        worksheet.cell(row=row, column=3, value=invoice.subtotal)
        worksheet.cell(row=row, column=4, value=invoice.tax_total)
        worksheet.cell(row=row, column=5, value=invoice.total_amount)
        last_payment = getattr(invoice, 'latest_payment', None)
        worksheet.cell(row=row, column=6, value=last_payment.date if last_payment else '')
        worksheet.cell(row=row, column=7, value=last_payment.method if last_payment else '')
        row += 1

    row += 1
    worksheet.cell(row=row, column=1, value='Totals').font = Font(bold=True)
    worksheet.cell(row=row, column=3, value=totals['subtotal'])
    worksheet.cell(row=row, column=4, value=totals['tax'])
    worksheet.cell(row=row, column=5, value=totals['total'])
    worksheet.cell(row=row, column=6, value=totals['paid'])

    for column in range(1, len(headers) + 1):
        column_letter = get_column_letter(column)
        max_length = 0
        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max(max_length + 2, 14)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f'statement_paid_invoices_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _get_customer_portal_account(request):
    user = getattr(request, "user", None)
    if not user or not getattr(user, "is_authenticated", False):
        return None
    try:
        return user.customer_portal
    except ObjectDoesNotExist:
        return None


def _sync_session_cart_to_db(request, customer_account, store_owner):
    if not request or not getattr(request, "session", None):
        return
    if getattr(request, "_storefront_cart_synced", False):
        return
    request._storefront_cart_synced = True

    session_cart = request.session.get("cart", {}) or {}
    if not session_cart or not customer_account or not store_owner:
        return

    quantities = {}
    for product_id, qty in session_cart.items():
        try:
            product_id_int = int(product_id)
            qty_int = int(qty)
        except (TypeError, ValueError):
            continue
        if qty_int <= 0:
            continue
        quantities[product_id_int] = quantities.get(product_id_int, 0) + qty_int

    if not quantities:
        request.session.pop("cart", None)
        return

    product_qs = _storefront_product_queryset(request, owner=store_owner).filter(
        id__in=list(quantities.keys())
    )
    product_map = {product.id: product for product in product_qs}

    for product_id, qty in quantities.items():
        product = product_map.get(product_id)
        if not product:
            continue
        available_qty = getattr(product, "stock_quantity", product.quantity_in_stock)
        qty = min(qty, available_qty)
        if qty <= 0:
            continue
        cart_item, created = StorefrontCartItem.objects.get_or_create(
            customer=customer_account,
            store_owner=store_owner,
            product=product,
            defaults={"quantity": qty},
        )
        if not created:
            available_qty = getattr(product, "stock_quantity", product.quantity_in_stock)
            new_qty = min(cart_item.quantity + qty, available_qty)
            if new_qty != cart_item.quantity:
                cart_item.quantity = new_qty
                cart_item.save(update_fields=["quantity", "updated_at"])

    request.session.pop("cart", None)


def _get_cart_count(customer_account, store_owner):
    if not customer_account or not store_owner:
        return 0
    total = StorefrontCartItem.objects.filter(
        customer=customer_account,
        store_owner=store_owner,
    ).aggregate(total=Coalesce(Sum("quantity"), 0))["total"]
    return int(total or 0)


def _clear_cart(customer_account, store_owner):
    if not customer_account or not store_owner:
        return
    StorefrontCartItem.objects.filter(
        customer=customer_account,
        store_owner=store_owner,
    ).delete()


def _get_cart(request):
    """Return the cart dict stored for the customer and storefront."""
    customer_account = _get_customer_portal_account(request)
    if not customer_account:
        return {}
    store_owner = get_storefront_owner(request)
    _sync_session_cart_to_db(request, customer_account, store_owner)
    cart_items = StorefrontCartItem.objects.filter(
        customer=customer_account,
        store_owner=store_owner,
    )
    return {str(item.product_id): item.quantity for item in cart_items}


def _get_cart_product_ids(request):
    """Return cart product ids as strings for template checks."""
    customer_account = _get_customer_portal_account(request)
    if not customer_account:
        return set()
    store_owner = get_storefront_owner(request)
    _sync_session_cart_to_db(request, customer_account, store_owner)
    product_ids = StorefrontCartItem.objects.filter(
        customer=customer_account,
        store_owner=store_owner,
    ).values_list("product_id", flat=True)
    return {str(product_id) for product_id in product_ids}


def _build_storefront_cart_items(cart, product_qs):
    """Build cart items with pricing metadata for storefront views."""
    items = []
    subtotal = Decimal('0.00')
    subtotal_before_discounts = Decimal('0.00')
    discount_total = Decimal('0.00')
    sellers = set()
    cart_quantities = {}

    for product_id, qty in cart.items():
        product = product_qs.filter(pk=product_id).first()
        if not product:
            continue
        if hasattr(product, "stock_quantity"):
            product.quantity_in_stock = product.stock_quantity
            product.reorder_level = product.stock_reorder
        try:
            quantity = int(qty)
        except (TypeError, ValueError):
            continue
        if quantity <= 0:
            continue
        available_qty = getattr(product, "stock_quantity", product.quantity_in_stock)
        quantity = min(quantity, available_qty)
        unit_price = _resolve_storefront_price(product)
        if unit_price is None:
            continue
        base_price = product.sale_price if product.sale_price is not None else unit_price
        line_subtotal = unit_price * quantity
        line_base_total = base_price * quantity
        line_discount = line_base_total - line_subtotal
        if line_discount < Decimal('0.00'):
            line_discount = Decimal('0.00')

        subtotal += line_subtotal
        subtotal_before_discounts += line_base_total
        discount_total += line_discount

        items.append({
            'product': product,
            'quantity': quantity,
            'subtotal': line_subtotal,
            'unit_price': unit_price,
            'original_unit_price': base_price,
            'discount_total': line_discount,
            'is_free': False,
        })
        sellers.add(product.user)
        cart_quantities[product.id] = quantity

    return items, subtotal, subtotal_before_discounts, discount_total, sellers, cart_quantities


def _resolve_storefront_free_items(store_owner, cart_quantities):
    """Return free bonus items for qualifying storefront packages."""
    if not store_owner or not cart_quantities:
        return []

    stock_owner = get_stock_owner(store_owner)
    packages = (
        StorefrontHeroPackage.objects.filter(user=store_owner, is_active=True)
        .select_related('primary_product', 'secondary_product', 'free_product')
    )
    free_items = []
    for package in packages:
        if not package.free_product_id:
            continue
        if not package.primary_product_id or not package.secondary_product_id:
            continue
        primary_qty = cart_quantities.get(package.primary_product_id, 0)
        secondary_qty = cart_quantities.get(package.secondary_product_id, 0)
        if primary_qty <= 0 or secondary_qty <= 0:
            continue
        free_qty = min(primary_qty, secondary_qty)
        free_product = package.free_product
        if not free_product or free_qty <= 0:
            continue
        if free_product.quantity_in_stock is not None:
            available_qty = free_product.quantity_in_stock
            if stock_owner:
                stock_record = ProductStock.objects.filter(product=free_product, user=stock_owner).first()
                available_qty = stock_record.quantity_in_stock if stock_record else 0
            free_qty = min(free_qty, available_qty)
            free_product.quantity_in_stock = available_qty
        if free_qty <= 0:
            continue

        base_price = free_product.sale_price if free_product.sale_price is not None else free_product.promotion_price
        if base_price is None:
            base_price = Decimal('0.00')
        free_value = base_price * free_qty
        free_items.append({
            'product': free_product,
            'quantity': free_qty,
            'subtotal': Decimal('0.00'),
            'unit_price': Decimal('0.00'),
            'original_unit_price': base_price,
            'discount_total': Decimal('0.00'),
            'free_value': free_value,
            'is_free': True,
            'package_title': package.title,
        })

    return free_items


@require_POST
def set_storefront_location(request):
    """Persist the selected storefront location in session."""

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")
    if not next_url or not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = reverse("accounts:store_product_list")

    root_user = resolve_storefront_root_user(request)
    profiles = list(get_storefront_profiles(root_user)) if root_user else []
    allowed_ids = {profile.user_id for profile in profiles}

    selected_raw = (request.POST.get("storefront_location") or "").strip()
    selected_id = None
    if selected_raw:
        try:
            selected_id = int(selected_raw)
        except (TypeError, ValueError):
            selected_id = None

    previous_id = request.session.get("storefront_owner_id")
    previous_id_int = None
    if previous_id is not None:
        try:
            previous_id_int = int(previous_id)
        except (TypeError, ValueError):
            previous_id_int = None

    previous_owner = None
    if previous_id_int in allowed_ids:
        for profile in profiles:
            if profile.user_id == previous_id_int:
                previous_owner = profile.user
                break

    customer_account = _get_customer_portal_account(request)

    if selected_id and selected_id in allowed_ids:
        if previous_id_int is not None and previous_id_int != selected_id:
            if customer_account and previous_owner:
                _sync_session_cart_to_db(request, customer_account, previous_owner)
            request.session.pop("cart", None)
        request.session["storefront_owner_id"] = selected_id
    else:
        if previous_id_int is not None:
            if customer_account and previous_owner:
                _sync_session_cart_to_db(request, customer_account, previous_owner)
            request.session.pop("cart", None)
        request.session.pop("storefront_owner_id", None)

    store_owner = get_storefront_owner(request)
    store_user_ids = get_product_user_ids(store_owner) if store_owner else []
    parsed_next = urlparse(next_url)
    path = parsed_next.path or ""
    if path:
        try:
            match = resolve(path)
        except Resolver404:
            match = None
        if match:
            url_name = match.url_name
            if url_name == "store_group_detail":
                group_id = match.kwargs.get("group_id")
                if group_id:
                    group_qs = CategoryGroup.objects.filter(id=group_id, is_active=True)
                    if store_user_ids:
                        group_qs = group_qs.filter(user__in=store_user_ids)
                    if not group_qs.exists():
                        messages.info(
                            request,
                            "That category group isn't available at this store location.",
                        )
                        next_url = reverse("accounts:store_product_list")
            elif url_name == "store_category_detail":
                category_id = match.kwargs.get("category_id")
                if category_id:
                    category_qs = Category.objects.filter(id=category_id, is_active=True)
                    if store_user_ids:
                        category_qs = category_qs.filter(user__in=store_user_ids)
                    if not category_qs.exists():
                        messages.info(
                            request,
                            "That category isn't available at this store location.",
                        )
                        next_url = reverse("accounts:store_product_list")
            elif url_name == "store_product_detail":
                product_id = match.kwargs.get("pk")
                if product_id:
                    product_qs = Product.objects.filter(
                        id=product_id,
                        is_published_to_store=True,
                    )
                    if store_user_ids:
                        product_qs = product_qs.filter(user__in=store_user_ids)
                    if not product_qs.exists():
                        messages.info(
                            request,
                            "That product isn't available at this store location.",
                        )
                        next_url = reverse("accounts:store_product_list")

    return redirect(next_url)


OPEN_METEO_GEOCODE_URL = "https://geocoding-api.open-meteo.com/v1/search"
OPEN_METEO_FORECAST_URL = "https://api.open-meteo.com/v1/forecast"

PROVINCE_NAME_MAP = {
    "AB": "Alberta",
    "BC": "British Columbia",
    "MB": "Manitoba",
    "NB": "New Brunswick",
    "NL": "Newfoundland and Labrador",
    "NT": "Northwest Territories",
    "NS": "Nova Scotia",
    "NU": "Nunavut",
    "ON": "Ontario",
    "PE": "Prince Edward Island",
    "QC": "Quebec",
    "SK": "Saskatchewan",
    "YT": "Yukon",
}

US_STATE_NAMES = {
    "AL": "Alabama",
    "AK": "Alaska",
    "AZ": "Arizona",
    "AR": "Arkansas",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DE": "Delaware",
    "FL": "Florida",
    "GA": "Georgia",
    "HI": "Hawaii",
    "ID": "Idaho",
    "IL": "Illinois",
    "IN": "Indiana",
    "IA": "Iowa",
    "KS": "Kansas",
    "KY": "Kentucky",
    "LA": "Louisiana",
    "ME": "Maine",
    "MD": "Maryland",
    "MA": "Massachusetts",
    "MI": "Michigan",
    "MN": "Minnesota",
    "MS": "Mississippi",
    "MO": "Missouri",
    "MT": "Montana",
    "NE": "Nebraska",
    "NV": "Nevada",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NY": "New York",
    "NC": "North Carolina",
    "ND": "North Dakota",
    "OH": "Ohio",
    "OK": "Oklahoma",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "RI": "Rhode Island",
    "SC": "South Carolina",
    "SD": "South Dakota",
    "TN": "Tennessee",
    "TX": "Texas",
    "UT": "Utah",
    "VT": "Vermont",
    "VA": "Virginia",
    "WA": "Washington",
    "WV": "West Virginia",
    "WI": "Wisconsin",
    "WY": "Wyoming",
}


def _normalize_weather_text(value):
    return re.sub(r"\s+", " ", value or "").strip()


def _normalize_weather_key(value):
    return _normalize_weather_text(value).lower()


def _extract_postal_code(address):
    if not address:
        return ""
    canada_match = re.search(r"[A-Z]\d[A-Z]\s?\d[A-Z]\d", address, re.IGNORECASE)
    if canada_match:
        return re.sub(r"\s+", " ", canada_match.group(0).upper()).strip()
    us_match = re.search(r"\b\d{5}(?:-\d{4})?\b", address)
    if us_match:
        return us_match.group(0)
    return ""


def _extract_province_code(address):
    if not address:
        return ""
    match = re.search(r"\b[A-Z]{2}\b", address.upper())
    if not match:
        return ""
    code = match.group(0)
    if code in PROVINCE_NAME_MAP or code in US_STATE_NAMES:
        return code
    return ""


def _extract_city_name(address):
    if not address:
        return ""
    parts = [part.strip() for part in re.split(r",|\n", address) if part.strip()]
    if len(parts) >= 2:
        return parts[1]
    return ""


def _compose_weather_address(street, city, province, postal):
    parts = []
    if street:
        parts.append(street)
    if city:
        parts.append(city)
    if province and (street or city or postal):
        parts.append(province)
    if postal:
        parts.append(postal)
    return ", ".join(parts)


def _build_store_weather_parts(profile):
    address_value = _normalize_weather_text(getattr(profile, "company_address", ""))
    street = _normalize_weather_text(getattr(profile, "street_address", ""))
    city = _normalize_weather_text(getattr(profile, "city", "")) or _extract_city_name(address_value)
    province = _normalize_weather_text(getattr(profile, "province", "")) or _extract_province_code(address_value)
    postal = _normalize_weather_text(getattr(profile, "postal_code", "")) or _extract_postal_code(address_value)
    composed_address = address_value or _compose_weather_address(street, city, province, postal)
    return {
        "address": composed_address,
        "street": street,
        "city": city,
        "province": province,
        "postal": postal,
    }


def _resolve_country_code(province):
    code = (province or "").upper()
    if code in PROVINCE_NAME_MAP:
        return "CA"
    if code in US_STATE_NAMES:
        return "US"
    return ""


def _resolve_province_name(province):
    code = (province or "").upper()
    if code in PROVINCE_NAME_MAP:
        return PROVINCE_NAME_MAP[code]
    if code in US_STATE_NAMES:
        return US_STATE_NAMES[code]
    return province or ""


def _geocode_query(query):
    params = {
        "name": query,
        "count": 10,
        "language": "en",
        "format": "json",
    }
    response = requests.get(OPEN_METEO_GEOCODE_URL, params=params, timeout=10)
    response.raise_for_status()
    payload = response.json()
    return payload.get("results") or []


def _pick_geocode_result(results, parts):
    if not results:
        return None
    country_code = _resolve_country_code(parts.get("province"))
    province_name = _normalize_weather_key(_resolve_province_name(parts.get("province")))
    city_name = _normalize_weather_key(parts.get("city"))

    candidates = results
    if country_code:
        candidates = [item for item in candidates if item.get("country_code") == country_code]
    if province_name:
        province_matches = [
            item for item in candidates
            if _normalize_weather_key(item.get("admin1")) == province_name
        ]
        if province_matches:
            candidates = province_matches
    if city_name:
        city_matches = [
            item for item in candidates
            if _normalize_weather_key(item.get("name")) == city_name
        ]
        if city_matches:
            candidates = city_matches
    return candidates[0] if candidates else results[0]


def _geocode_store(parts):
    queries = []
    city = parts.get("city")
    province = parts.get("province")
    postal = parts.get("postal")
    address = parts.get("address")
    if city and province:
        queries.append(f"{city}, {province}")
    if city:
        queries.append(city)
    if postal:
        queries.append(postal)
    if address:
        queries.append(address)

    seen = set()
    for query in queries:
        normalized = _normalize_weather_key(query)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        results = _geocode_query(query)
        picked = _pick_geocode_result(results, parts)
        if picked and picked.get("latitude") is not None and picked.get("longitude") is not None:
            return {
                "latitude": float(picked["latitude"]),
                "longitude": float(picked["longitude"]),
            }
    return None


def _fetch_weather(coords):
    params = {
        "latitude": f"{coords['latitude']:.4f}",
        "longitude": f"{coords['longitude']:.4f}",
        "current": "temperature_2m,weathercode,is_day,snowfall,precipitation",
        "hourly": "snowfall",
        "daily": "sunrise,sunset",
        "forecast_days": 1,
        "timezone": "auto",
        "temperature_unit": "celsius",
    }
    response = requests.get(OPEN_METEO_FORECAST_URL, params=params, timeout=10)
    response.raise_for_status()
    return response.json()


def _parse_weather_time(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def _find_closest_weather_index(times, target_time):
    if not times or not target_time:
        return None
    target_dt = _parse_weather_time(target_time)
    if not target_dt:
        return None
    best_index = None
    best_diff = None
    for index, value in enumerate(times):
        time_dt = _parse_weather_time(value)
        if not time_dt:
            continue
        diff = abs((time_dt - target_dt).total_seconds())
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_index = index
    return best_index


def _resolve_snowfall_amount(data):
    current = data.get("current") or data.get("current_weather") or {}
    snowfall = current.get("snowfall")
    if isinstance(snowfall, (int, float)):
        return float(snowfall)
    hourly = data.get("hourly") or {}
    snow = hourly.get("snowfall") or []
    times = hourly.get("time") or []
    if not snow or not times or len(snow) != len(times):
        return None
    target_time = current.get("time") or times[0]
    index = _find_closest_weather_index(times, target_time)
    if index is None:
        return None
    try:
        return float(snow[index])
    except (TypeError, ValueError):
        return None


def _resolve_snowfall_units(data):
    current_units = data.get("current_units") or {}
    hourly_units = data.get("hourly_units") or {}
    return current_units.get("snowfall") or hourly_units.get("snowfall") or ""


@require_GET
def storefront_weather(request):
    store_owner = get_storefront_owner(request)
    if not store_owner:
        return JsonResponse({"error": "store_missing"}, status=404)

    cache_key = f"storefront_weather:{store_owner.id}"
    cached = cache.get(cache_key)
    if cached:
        return JsonResponse(cached)

    profile = getattr(store_owner, "profile", None)
    if not profile:
        return JsonResponse({"error": "profile_missing"}, status=404)

    parts = _build_store_weather_parts(profile)
    if not parts["address"] and not parts["city"] and not parts["postal"]:
        return JsonResponse({"error": "address_missing"}, status=400)

    try:
        coords = _geocode_store(parts)
    except (requests.RequestException, ValueError):
        return JsonResponse({"error": "geocode_failed"}, status=502)
    if not coords:
        return JsonResponse({"error": "geocode_failed"}, status=404)

    try:
        data = _fetch_weather(coords)
    except (requests.RequestException, ValueError):
        return JsonResponse({"error": "weather_unavailable"}, status=502)

    current = data.get("current") or data.get("current_weather") or {}
    temperature = current.get("temperature_2m", current.get("temperature"))
    weather_code = current.get("weathercode")
    is_day_raw = current.get("is_day")
    if isinstance(is_day_raw, (int, float)):
        is_day = bool(int(is_day_raw))
    else:
        is_day = bool(is_day_raw)
    sunrise_list = (data.get("daily") or {}).get("sunrise") or []
    sunset_list = (data.get("daily") or {}).get("sunset") or []
    sunrise = sunrise_list[0] if sunrise_list else None
    sunset = sunset_list[0] if sunset_list else None
    payload = {
        "temperature": temperature,
        "weatherCode": weather_code,
        "isDay": is_day,
        "sunrise": sunrise,
        "sunset": sunset,
        "timeZone": data.get("timezone") or "UTC",
        "snowfall": _resolve_snowfall_amount(data),
        "snowfallUnit": _resolve_snowfall_units(data),
    }
    cache.set(cache_key, payload, 600)
    return JsonResponse(payload)


def _storefront_product_queryset(request=None, *, owner=None):
    """Base queryset for products that are eligible to appear in the public store."""

    store_owner = owner or get_storefront_owner(request)
    queryset = Product.objects.filter(is_published_to_store=True)
    if store_owner:
        store_user_ids = get_product_user_ids(store_owner)
        if store_user_ids:
            queryset = queryset.filter(user__in=store_user_ids)
        else:
            queryset = queryset.filter(user=store_owner)
        queryset = annotate_products_with_stock(queryset, store_owner)
    return queryset


PRODUCT_OVERVIEW_ATTRIBUTE_NAMES = {
    "overview",
    "product overview",
    "product summary",
    "summary",
    "product description",
    "description",
}

PRODUCT_FEATURE_ATTRIBUTE_NAMES = {
    "features",
    "features benefits",
    "features and benefits",
    "benefits",
    "key features",
    "product features",
    "product benefits",
    "feature highlights",
    "highlights",
}


def _normalize_attribute_label(label):
    if not label:
        return ""
    cleaned = re.sub(r"[^a-z0-9]+", " ", label.lower())
    return " ".join(cleaned.split())


def _split_feature_text(text):
    if not text:
        return []
    normalized = str(text).replace("\r\n", "\n").replace("\r", "\n")
    for bullet in ("\u2022", "\u2023", "\u2043", "\u00b7", "\u2219"):
        normalized = normalized.replace(bullet, "\n")
    items = []
    for line in normalized.split("\n"):
        for chunk in line.split(";"):
            item = chunk.strip()
            if not item:
                continue
            item = item.lstrip("-*").strip()
            if item:
                items.append(item)
    return items


def _resolve_product_detail_sections(product, attribute_values):
    overview_text = ""
    feature_items = []
    feature_seen = set()
    overview_attribute_ids = set()
    feature_attribute_ids = set()

    for value in attribute_values:
        display_value = value.get_display_value()
        if not display_value:
            continue
        normalized_name = _normalize_attribute_label(value.attribute.name)
        if not overview_text and normalized_name in PRODUCT_OVERVIEW_ATTRIBUTE_NAMES:
            overview_text = display_value
            overview_attribute_ids.add(value.attribute_id)
            continue
        if normalized_name in PRODUCT_FEATURE_ATTRIBUTE_NAMES:
            feature_attribute_ids.add(value.attribute_id)
            for item in _split_feature_text(display_value):
                key = item.casefold()
                if key in feature_seen:
                    continue
                feature_seen.add(key)
                feature_items.append(item)

    if not overview_text and product.description:
        overview_text = product.description

    filtered_attributes = [
        value
        for value in attribute_values
        if value.attribute_id not in overview_attribute_ids
        and value.attribute_id not in feature_attribute_ids
    ]
    return overview_text, feature_items, filtered_attributes


def _resolve_storefront_price(product):
    if not product:
        return None
    return product.promotion_price if product.promotion_price is not None else product.sale_price


def _calculate_discounted_price(sale_price, discount_percent):
    if sale_price is None or not discount_percent:
        return None
    try:
        percent = Decimal(str(discount_percent))
    except (TypeError, ValueError):
        return None
    if percent <= Decimal("0"):
        return None
    multiplier = (Decimal("100") - percent) / Decimal("100")
    return (sale_price * multiplier).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _apply_storefront_hero_discounts(hero_showcase, packages):
    if not hero_showcase:
        return

    discount_map = {}
    for slide in hero_showcase.slides.select_related("product"):
        if slide.discount_percent and slide.product_id:
            current = discount_map.get(slide.product_id, 0)
            discount_map[slide.product_id] = max(current, slide.discount_percent)

    for package in packages:
        if not package.is_active or not package.discount_percent:
            continue
        for product_id in (package.primary_product_id, package.secondary_product_id):
            if not product_id:
                continue
            current = discount_map.get(product_id, 0)
            discount_map[product_id] = max(current, package.discount_percent)

    if not discount_map:
        return

    products = Product.objects.filter(id__in=discount_map.keys())
    for product in products:
        discount_percent = discount_map.get(product.id)
        discounted_price = _calculate_discounted_price(product.sale_price, discount_percent)
        if discounted_price is None:
            continue
        if product.promotion_price != discounted_price:
            product.promotion_price = discounted_price
            product.save(update_fields=["promotion_price"])


def _build_storefront_marketing_context(request, available_products, *, store_owner=None):
    store_owner = store_owner or get_storefront_owner(request)
    featured_source = available_products.order_by('-updated_at')
    featured_qs = featured_source.filter(is_featured=True)
    featured_products = list(featured_qs[:8]) if featured_qs.exists() else list(featured_source[:8])

    hero_showcase = None
    hero_slides = []
    hero_banner = None
    hero_packages = []
    flyer = None
    flyer_slides = []
    flyer_packages = []
    flyer_has_content = False
    if store_owner:
        hero_showcase = StorefrontHeroShowcase.objects.filter(user=store_owner).first()
        if hero_showcase:
            hero_slides = list(
                hero_showcase.slides.select_related('product', 'product__brand')
                .filter(product__is_published_to_store=True)
                .order_by('id')
            )
        hero_banner = StorefrontMessageBanner.objects.filter(
            user=store_owner,
            is_active=True,
        ).first()
        if hero_banner and not (hero_banner.message or '').strip():
            hero_banner = None
        hero_packages = list(
            StorefrontHeroPackage.objects.filter(user=store_owner, is_active=True).order_by('id')
        )
        flyer = StorefrontFlyer.objects.filter(user=store_owner, is_active=True).first()
        if flyer:
            flyer_slides = hero_slides
            flyer_packages = hero_packages
            flyer_has_content = bool(flyer_slides or flyer_packages)

    stock_owner = get_stock_owner(store_owner) if store_owner else None
    if stock_owner:
        apply_stock_fields(featured_products)
        slide_product_ids = [slide.product_id for slide in hero_slides if slide.product_id]
        package_product_ids = []
        for package in hero_packages:
            for product_id in (package.primary_product_id, package.secondary_product_id, package.free_product_id):
                if product_id:
                    package_product_ids.append(product_id)
        stock_product_ids = set(slide_product_ids + package_product_ids)
        if stock_product_ids:
            stock_rows = ProductStock.objects.filter(
                user=stock_owner,
                product_id__in=stock_product_ids,
            )
            stock_map = {row.product_id: row for row in stock_rows}
            for slide in hero_slides:
                product = slide.product
                if not product:
                    continue
                stock_record = stock_map.get(product.id)
                product.quantity_in_stock = stock_record.quantity_in_stock if stock_record else 0
                product.reorder_level = stock_record.reorder_level if stock_record else 0
            for package in hero_packages:
                for product in (package.primary_product, package.secondary_product, package.free_product):
                    if not product:
                        continue
                    stock_record = stock_map.get(product.id)
                    product.quantity_in_stock = stock_record.quantity_in_stock if stock_record else 0
                    product.reorder_level = stock_record.reorder_level if stock_record else 0

    hero_cards = []
    for slide in hero_slides:
        hero_cards.append({'kind': 'product', 'product': slide.product, 'slide': slide})
    for package in hero_packages:
        hero_cards.append({'kind': 'package', 'package': package})

    price_flags = resolve_storefront_price_flags(request, store_owner)
    return {
        'featured_products': featured_products,
        'hero_showcase': hero_showcase,
        'hero_slides': hero_slides,
        'hero_cards': hero_cards,
        'hero_banner': hero_banner,
        'flyer': flyer if flyer_has_content else None,
        'flyer_slides': flyer_slides,
        'flyer_packages': flyer_packages,
        'flyer_has_content': flyer_has_content,
        'show_prices_hero': price_flags['hero'],
        'show_prices_featured': price_flags['featured'],
    }


def _customer_matches_store(customer_account, store_owner):
    if not customer_account or not store_owner:
        return False
    if customer_account.user_id == store_owner.id:
        return True
    customer_business = get_business_user(customer_account.user) or customer_account.user
    store_business = get_business_user(store_owner) or store_owner
    if customer_business and store_business and customer_business.id == store_business.id:
        return True
    if not customer_business or not store_business:
        return False
    shared_customer_user_ids = set(get_customer_user_ids(customer_business))
    return store_business.id in shared_customer_user_ids


def _storefront_categories_queryset(available_products, *, owner=None, include_empty=False):
    if include_empty and owner:
        store_user_ids = get_product_user_ids(owner)
        if store_user_ids:
            return Category.objects.filter(
                user__in=store_user_ids,
                is_active=True,
            ).select_related('group', 'parent')
        return Category.objects.filter(user=owner, is_active=True).select_related('group', 'parent')

    return (
        Category.objects.filter(products__in=available_products, is_active=True)
        .select_related('group', 'parent')
        .distinct()
    )


def _storefront_groups_queryset(available_products, *, owner=None, include_empty=False):
    queryset = CategoryGroup.objects.filter(is_active=True)
    if owner:
        store_user_ids = get_product_user_ids(owner)
        if store_user_ids:
            queryset = queryset.filter(user__in=store_user_ids)
        else:
            queryset = queryset.filter(user=owner)
    if include_empty:
        return queryset.order_by('sort_order', 'name')

    return (
        queryset.filter(
            categories__products__in=available_products,
            categories__is_active=True,
        )
        .distinct()
        .order_by('sort_order', 'name')
    )


def _build_category_tree(categories):
    by_parent = defaultdict(list)
    for category in categories:
        by_parent[category.parent_id].append(category)
    for children in by_parent.values():
        children.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))

    def build(parent_id=None):
        nodes = []
        for category in by_parent.get(parent_id, []):
            nodes.append({
                "category": category,
                "children": build(category.id),
            })
        return nodes

    return build(None)


def _flatten_category_tree(tree, active_path_ids):
    flattened = []
    active_ids = set(active_path_ids or [])

    def walk(nodes, depth=0):
        for node in nodes:
            category = node["category"]
            flattened.append({
                "category": category,
                "depth": depth,
                "is_current": category.id == (active_path_ids[-1] if active_path_ids else None),
                "is_active": category.id in active_ids,
                "has_children": bool(node["children"]),
            })
            walk(node["children"], depth + 1)

    walk(tree)
    return flattened


def _collect_descendant_ids(category, children_map):
    ids = [category.id]
    stack = list(children_map.get(category.id, []))
    while stack:
        child = stack.pop()
        ids.append(child.id)
        stack.extend(children_map.get(child.id, []))
    return ids


def _build_category_path(category):
    path = []
    current = category
    while current:
        path.append(current)
        current = current.parent
    return list(reversed(path))


def _build_storefront_breadcrumbs(group=None, category_path=None, *, label=None):
    home_url = reverse('accounts:public_home')
    breadcrumbs = [
        {"label": "Home", "url": home_url},
        {"label": "All Categories", "url": reverse('accounts:store_product_list')},
    ]
    if group:
        breadcrumbs.append({
            "label": group.name,
            "url": reverse('accounts:store_group_detail', args=[group.id]),
        })
    if category_path:
        for category in category_path:
            breadcrumbs.append({
                "label": category.name,
                "url": reverse('accounts:store_category_detail', args=[category.id]),
            })
    if label:
        breadcrumbs.append({"label": label, "url": None})
    return breadcrumbs


def _coerce_int(value, default, allowed):
    try:
        value = int(value)
    except (TypeError, ValueError):
        return default
    return value if value in allowed else default


def _build_product_list_context(
    request,
    available_products,
    *,
    group=None,
    category=None,
    include_descendants=False,
    include_empty_categories=False,
    owner=None,
):
    search_query = (request.GET.get('q') or '').strip()
    brand_ids = [value for value in request.GET.getlist('brand') if value]
    model_ids = [value for value in request.GET.getlist('model') if value]
    vin_ids = [value for value in request.GET.getlist('vin') if value]
    sort_value = (request.GET.get('sort') or 'popular').strip()
    view_mode = (request.GET.get('view') or 'grid').strip().lower()
    per_page = _coerce_int(request.GET.get('per_page'), 20, {20, 50, 100})
    view_mode = view_mode if view_mode in {'grid', 'list'} else 'grid'

    products = available_products
    if group:
        products = products.filter(category__group_id=group.id)

    categories = _storefront_categories_queryset(
        available_products,
        owner=owner,
        include_empty=include_empty_categories,
    )
    children_map = defaultdict(list)
    for cat in categories:
        children_map[cat.parent_id].append(cat)
    for children in children_map.values():
        children.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))

    descendant_ids = None
    if category:
        if include_descendants:
            descendant_ids = _collect_descendant_ids(category, children_map)
        else:
            descendant_ids = [category.id]
        products = products.filter(category_id__in=descendant_ids)

    if search_query:
        alternate_ids = ProductAlternateSku.objects.filter(
            sku__icontains=search_query,
            kind__in=['interchange', 'equivalent'],
        ).values('product_id')
        products = products.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(sku__icontains=search_query)
            | Q(category__name__icontains=search_query)
            | Q(brand__name__icontains=search_query)
            | Q(vehicle_model__name__icontains=search_query)
            | Q(vin_number__vin__icontains=search_query)
            | Q(pk__in=alternate_ids)
        )

    filter_base = products
    models = list(
        ProductModel.objects.filter(products__in=filter_base, is_active=True)
        .distinct()
        .order_by('sort_order', 'name')
    )
    vins = list(
        ProductVin.objects.filter(products__in=filter_base, is_active=True)
        .distinct()
        .order_by('sort_order', 'vin')
    )

    if brand_ids:
        products = products.filter(brand_id__in=brand_ids)
    if model_ids:
        products = products.filter(vehicle_model_id__in=model_ids)
    if vin_ids:
        products = products.filter(vin_number_id__in=vin_ids)

    attribute_filters = []
    if category:
        category_chain = []
        current = category
        while current:
            category_chain.append(current)
            current = current.parent

        attributes = (
            CategoryAttribute.objects.filter(
                category__in=category_chain,
                is_filterable=True,
                is_active=True,
            )
            .prefetch_related('options')
            .order_by('sort_order', 'name')
        )

        selected_attribute_values = {}
        for attribute in attributes:
            param_name = f"attr_{attribute.id}"
            selected_attribute_values[attribute.id] = (request.GET.get(param_name) or "").strip()

        def apply_attribute_filters(attribute_products, selected_values):
            filtered_products = attribute_products
            for attribute in attributes:
                selected_value = selected_values.get(attribute.id, "")
                if not selected_value:
                    continue
                if attribute.attribute_type == "select":
                    filtered_products = filtered_products.filter(
                        attribute_values__attribute=attribute,
                        attribute_values__option_id=selected_value,
                    )
                elif attribute.attribute_type == "boolean":
                    normalized = selected_value.lower()
                    if normalized in ("1", "true", "yes", "on"):
                        bool_value = True
                    elif normalized in ("0", "false", "no", "off"):
                        bool_value = False
                    else:
                        bool_value = None
                    if bool_value is not None:
                        filtered_products = filtered_products.filter(
                            attribute_values__attribute=attribute,
                            attribute_values__value_boolean=bool_value,
                        )
                elif attribute.attribute_type == "number":
                    try:
                        number_value = Decimal(selected_value)
                    except (ArithmeticError, ValueError, TypeError):
                        number_value = None
                    if number_value is not None:
                        filtered_products = filtered_products.filter(
                            attribute_values__attribute=attribute,
                            attribute_values__value_number=number_value,
                        )
                else:
                    filtered_products = filtered_products.filter(
                        attribute_values__attribute=attribute,
                        attribute_values__value_text__iexact=selected_value,
                    )
            return filtered_products

        def build_attribute_options(attribute_products):
            options_by_attribute = {}
            for attribute in attributes:
                options = []
                if attribute.attribute_type == "select":
                    option_ids = (
                        ProductAttributeValue.objects.filter(
                            product__in=attribute_products,
                            attribute=attribute,
                            option__isnull=False,
                        )
                        .values_list("option_id", flat=True)
                        .distinct()
                    )
                    option_qs = (
                        attribute.options.filter(is_active=True, id__in=option_ids)
                        .order_by("sort_order", "value")
                    )
                    options = [{"value": str(option.id), "label": option.value} for option in option_qs]
                elif attribute.attribute_type == "boolean":
                    bool_values = (
                        ProductAttributeValue.objects.filter(
                            product__in=attribute_products,
                            attribute=attribute,
                            value_boolean__isnull=False,
                        )
                        .values_list("value_boolean", flat=True)
                        .distinct()
                    )
                    if True in bool_values:
                        options.append({"value": "true", "label": "Yes"})
                    if False in bool_values:
                        options.append({"value": "false", "label": "No"})
                elif attribute.attribute_type == "number":
                    number_values = (
                        ProductAttributeValue.objects.filter(
                            product__in=attribute_products,
                            attribute=attribute,
                            value_number__isnull=False,
                        )
                        .values_list("value_number", flat=True)
                        .distinct()
                        .order_by("value_number")[:50]
                    )
                    for number_value in number_values:
                        if number_value is None:
                            continue
                        raw_value = str(number_value)
                        label = f"{raw_value} {attribute.value_unit}".strip() if attribute.value_unit else raw_value
                        options.append({"value": raw_value, "label": label})
                else:
                    text_values = (
                        ProductAttributeValue.objects.filter(
                            product__in=attribute_products,
                            attribute=attribute,
                        )
                        .exclude(value_text="")
                        .values_list("value_text", flat=True)
                        .distinct()
                        .order_by("value_text")[:50]
                    )
                    options = [{"value": value, "label": value} for value in text_values if value]
                options_by_attribute[attribute.id] = options
            return options_by_attribute

        def attribute_has_full_coverage(attribute_products, attribute, total_count):
            if total_count <= 0:
                return False
            coverage_qs = ProductAttributeValue.objects.filter(
                product__in=attribute_products,
                attribute=attribute,
            )
            if attribute.attribute_type == "select":
                coverage_qs = coverage_qs.filter(option__isnull=False)
            elif attribute.attribute_type == "boolean":
                coverage_qs = coverage_qs.filter(value_boolean__isnull=False)
            elif attribute.attribute_type == "number":
                coverage_qs = coverage_qs.filter(value_number__isnull=False)
            else:
                coverage_qs = coverage_qs.exclude(value_text="").exclude(value_text__isnull=True)
            return coverage_qs.values("product_id").distinct().count() == total_count

        attribute_products_base = products
        options_by_attribute = {}
        while True:
            filtered_products = apply_attribute_filters(attribute_products_base, selected_attribute_values)
            if attributes:
                filtered_products = filtered_products.distinct()

            filtered_count = filtered_products.count()
            options_by_attribute = build_attribute_options(filtered_products)
            auto_selected = None
            for attribute in attributes:
                if selected_attribute_values.get(attribute.id):
                    continue
                options = options_by_attribute.get(attribute.id, [])
                if len(options) == 1 and attribute_has_full_coverage(
                    filtered_products,
                    attribute,
                    filtered_count,
                ):
                    auto_selected = (attribute.id, options[0]["value"])
                    break

            if not auto_selected:
                products = filtered_products
                break
            selected_attribute_values[auto_selected[0]] = auto_selected[1]

        for attribute in attributes:
            param_name = f"attr_{attribute.id}"
            selected_value = selected_attribute_values.get(attribute.id, "")
            options = options_by_attribute.get(attribute.id, [])
            if options or selected_value:
                attribute_filters.append(
                    {
                        "attribute": attribute,
                        "param": param_name,
                        "selected": selected_value,
                        "options": options,
                    }
                )

    brand_filters = Q(products__in=products)
    if brand_ids:
        brand_filters |= Q(id__in=brand_ids, products__in=available_products)
    brands = list(
        ProductBrand.objects.filter(is_active=True)
        .filter(brand_filters)
        .distinct()
        .order_by('sort_order', 'name')
    )

    if sort_value in {"price_low", "price_high"}:
        price_expr = Coalesce('promotion_price', 'sale_price', 'cost_price')
        products = products.annotate(price_sort=price_expr)

    if sort_value == "name_asc":
        products = products.order_by('name')
    elif sort_value == "name_desc":
        products = products.order_by('-name')
    elif sort_value == "newest":
        products = products.order_by('-updated_at')
    elif sort_value == "price_low":
        products = products.order_by('price_sort', 'name')
    elif sort_value == "price_high":
        products = products.order_by('-price_sort', 'name')
    else:
        products = products.order_by('-is_featured', '-updated_at', 'name')

    paginator = Paginator(products, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    apply_stock_fields(page_obj.object_list)

    params = request.GET.copy()
    params.pop('page', None)
    params.pop('partial', None)
    for entry in attribute_filters:
        selected_value = (entry.get("selected") or "").strip()
        if selected_value:
            params[entry["param"]] = selected_value
        else:
            params.pop(entry["param"], None)
    query_string = params.urlencode()

    view_grid_params = params.copy()
    view_grid_params['view'] = 'grid'
    view_list_params = params.copy()
    view_list_params['view'] = 'list'

    has_attribute_filters = any(
        (entry.get("selected") or "") for entry in attribute_filters
    )
    has_active_filters = bool(
        search_query
        or brand_ids
        or model_ids
        or vin_ids
        or has_attribute_filters
    )

    return {
        "products": page_obj,
        "products_count": paginator.count,
        "brands": brands,
        "models": models,
        "vins": vins,
        "attribute_filters": attribute_filters,
        "search_query": search_query,
        "selected_brand_ids": [str(value) for value in brand_ids],
        "selected_model_ids": [str(value) for value in model_ids],
        "selected_vin_ids": [str(value) for value in vin_ids],
        "sort_value": sort_value,
        "view_mode": view_mode,
        "per_page": per_page,
        "query_string": query_string,
        "grid_view_url": f"?{view_grid_params.urlencode()}" if view_grid_params else "?view=grid",
        "list_view_url": f"?{view_list_params.urlencode()}" if view_list_params else "?view=list",
        "per_page_options": [20, 50, 100],
        "sort_options": [
            ("popular", "Sort by Popularity"),
            ("newest", "Newest"),
            ("name_asc", "Name: A-Z"),
            ("name_desc", "Name: Z-A"),
            ("price_low", "Price: Low to High"),
            ("price_high", "Price: High to Low"),
        ],
        "has_active_filters": has_active_filters,
    }


def _get_authenticated_customer_info(user):
    """Collect the logged-in user's details for checkout."""
    customer_profile = getattr(user, 'customer_portal', None)
    if customer_profile:
        name = customer_profile.name or user.get_full_name() or user.get_username()
        email = customer_profile.email or user.email or ''
        phone = customer_profile.phone_number or ''
        address = customer_profile.address or ''
        cc_emails = customer_profile.get_cc_emails()
        return {
            'name': name,
            'email': email,
            'phone': phone,
            'address': address,
            'cc_emails': cc_emails,
        }

    profile = getattr(user, 'profile', None)

    name_parts = [user.first_name, user.last_name]
    name = " ".join(part for part in name_parts if part).strip()
    if not name and profile and getattr(profile, 'company_name', None):
        name = profile.company_name
    if not name:
        name = user.get_username()

    email = user.email or ''
    if not email and profile and getattr(profile, 'company_email', None):
        email = profile.company_email

    phone = ''
    if profile and getattr(profile, 'company_phone', None):
        phone = profile.company_phone

    address = ''
    if profile:
        address = profile.company_address or ''
        if not address:
            address_parts = [
                getattr(profile, 'street_address', ''),
                getattr(profile, 'city', ''),
                getattr(profile, 'province', ''),
                getattr(profile, 'postal_code', ''),
            ]
            address = ", ".join(part for part in address_parts if part)

    return {
        'name': name,
        'email': email,
        'phone': phone,
        'address': address,
        'cc_emails': [],
    }


def _render_product_list_page(request, available_products, *, group=None, category=None, include_descendants=False):
    store_owner = get_storefront_owner(request)
    category_flags = resolve_storefront_category_flags(request, store_owner)
    show_empty_categories = category_flags["show_empty_categories"]
    context = _build_product_list_context(
        request,
        available_products,
        group=group,
        category=category,
        include_descendants=include_descendants,
        include_empty_categories=show_empty_categories,
        owner=store_owner,
    )
    context["customer_account"] = (
        getattr(request.user, 'customer_portal', None) if request.user.is_authenticated else None
    )
    context["cart_product_ids"] = (
        _get_cart_product_ids(request) if request.user.is_authenticated else set()
    )
    category_path = _build_category_path(category) if category else []
    breadcrumbs = _build_storefront_breadcrumbs(
        group=group,
        category_path=category_path,
        label=context["search_query"] if context["search_query"] else None,
    )
    if context["search_query"]:
        heading = f"Search results for \"{context['search_query']}\""
        subheading = f"{context['products_count']} results"
    elif category:
        heading = category.name
        subheading = category.description or None
    elif group:
        heading = group.name
        subheading = group.description or None
    else:
        heading = "All Products"
        subheading = None

    context.update({
        "breadcrumbs": breadcrumbs,
        "heading": heading,
        "subheading": subheading,
        "active_group": group,
        "active_category": category,
    })
    context.update(
        _build_storefront_marketing_context(
            request,
            available_products,
            store_owner=store_owner,
        )
    )
    store_owner = get_storefront_owner(request)
    price_flags = resolve_storefront_price_flags(request, store_owner)
    context["show_prices_catalog"] = price_flags["catalog"]
    template_name = (
        'store/product_list_public_content.html'
        if request.GET.get('partial') == '1'
        else 'store/product_list_public.html'
    )
    return render(request, template_name, context)


def product_list(request):
    """Display all category groups with their top-level categories."""
    available_products = _storefront_product_queryset(request)
    store_owner = get_storefront_owner(request)
    category_flags = resolve_storefront_category_flags(request, store_owner)
    show_empty_categories = category_flags["show_empty_categories"]
    category_groups = list(
        _storefront_groups_queryset(
            available_products,
            owner=store_owner,
            include_empty=show_empty_categories,
        )
    )
    categories = list(
        _storefront_categories_queryset(
            available_products,
            owner=store_owner,
            include_empty=show_empty_categories,
        )
    )

    grouped_categories = []
    for group in category_groups:
        group_categories = [
            category for category in categories
            if category.group_id == group.id and category.parent_id is None
        ]
        group_categories.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))
        grouped_categories.append({
            "group": group,
            "categories": group_categories,
        })

    brand_logo_qs = (
        ProductBrand.objects.filter(is_active=True)
        .exclude(logo__isnull=True)
        .exclude(logo="")
    )
    if store_owner:
        store_user_ids = get_product_user_ids(store_owner)
        if store_user_ids:
            brand_logo_qs = brand_logo_qs.filter(user__in=store_user_ids)
        else:
            brand_logo_qs = brand_logo_qs.filter(user=store_owner)
    brand_logos = list(brand_logo_qs.order_by("sort_order", "name"))

    context = {
        "breadcrumbs": _build_storefront_breadcrumbs(),
        "category_groups": category_groups,
        "grouped_categories": grouped_categories,
        "brand_logos": brand_logos,
        "search_query": (request.GET.get('q') or '').strip(),
        "customer_account": getattr(request.user, 'customer_portal', None) if request.user.is_authenticated else None,
        "cart_product_ids": _get_cart_product_ids(request) if request.user.is_authenticated else set(),
    }
    context.update(
        _build_storefront_marketing_context(
            request,
            available_products,
            store_owner=store_owner,
        )
    )
    return render(request, 'store/category_groups.html', context)


def store_group_detail(request, group_id):
    """Display a category group with its categories and navigation."""
    available_products = _storefront_product_queryset(request)
    store_owner = get_storefront_owner(request)
    category_flags = resolve_storefront_category_flags(request, store_owner)
    show_empty_categories = category_flags["show_empty_categories"]
    group_queryset = CategoryGroup.objects.filter(id=group_id, is_active=True)
    if store_owner:
        store_user_ids = get_product_user_ids(store_owner)
        if store_user_ids:
            group_queryset = group_queryset.filter(user__in=store_user_ids)
        else:
            group_queryset = group_queryset.filter(user=store_owner)
    group = get_object_or_404(group_queryset)

    categories = [
        category for category in _storefront_categories_queryset(
            available_products,
            owner=store_owner,
            include_empty=show_empty_categories,
        )
        if category.group_id == group.id
    ]
    top_categories = [cat for cat in categories if cat.parent_id is None]
    top_categories.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))

    category_tree = _build_category_tree(categories)
    category_tree_flat = _flatten_category_tree(category_tree, [])

    context = {
        "group": group,
        "categories": top_categories,
        "category_tree": category_tree_flat,
        "breadcrumbs": _build_storefront_breadcrumbs(group=group),
        "search_query": (request.GET.get('q') or '').strip(),
        "customer_account": getattr(request.user, 'customer_portal', None) if request.user.is_authenticated else None,
    }
    context.update(
        _build_storefront_marketing_context(
            request,
            available_products,
            store_owner=store_owner,
        )
    )
    return render(request, 'store/group_detail.html', context)


def store_category_detail(request, category_id):
    """Display a category with subcategories or its product list."""
    available_products = _storefront_product_queryset(request)
    store_owner = get_storefront_owner(request)
    category_flags = resolve_storefront_category_flags(request, store_owner)
    show_empty_categories = category_flags["show_empty_categories"]
    category_queryset = Category.objects.select_related('group', 'parent').filter(
        id=category_id,
        is_active=True,
    )
    if store_owner:
        store_user_ids = get_product_user_ids(store_owner)
        if store_user_ids:
            category_queryset = category_queryset.filter(user__in=store_user_ids)
        else:
            category_queryset = category_queryset.filter(user=store_owner)
    category = get_object_or_404(category_queryset)
    categories = list(
        _storefront_categories_queryset(
            available_products,
            owner=store_owner,
            include_empty=show_empty_categories,
        )
    )
    children = [cat for cat in categories if cat.parent_id == category.id]
    children.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))

    if not children:
        return _render_product_list_page(request, available_products, group=category.group, category=category)

    category_path = _build_category_path(category)
    active_path_ids = [cat.id for cat in category_path]
    tree_categories = [cat for cat in categories if cat.group_id == category.group_id]
    category_tree = _build_category_tree(tree_categories)
    category_tree_flat = _flatten_category_tree(category_tree, active_path_ids)

    context = {
        "group": category.group,
        "category": category,
        "subcategories": children,
        "category_tree": category_tree_flat,
        "breadcrumbs": _build_storefront_breadcrumbs(group=category.group, category_path=category_path),
        "search_query": (request.GET.get('q') or '').strip(),
        "customer_account": getattr(request.user, 'customer_portal', None) if request.user.is_authenticated else None,
    }
    context.update(
        _build_storefront_marketing_context(
            request,
            available_products,
            store_owner=store_owner,
        )
    )
    return render(request, 'store/category_detail.html', context)


def store_search(request):
    """Search across all published products."""
    available_products = _storefront_product_queryset(request)
    return _render_product_list_page(request, available_products)


def store_search_suggestions(request):
    """Return lightweight product suggestions for live storefront search."""
    query = (request.GET.get('q') or '').strip()
    if len(query) < 2:
        return JsonResponse({'query': query, 'results': [], 'categories': [], 'has_more': False})

    available_products = _storefront_product_queryset(request)
    products = available_products.select_related('category', 'brand')
    store_owner = get_storefront_owner(request)
    category_flags = resolve_storefront_category_flags(request, store_owner)
    show_empty_categories = category_flags["show_empty_categories"]

    alternate_ids = ProductAlternateSku.objects.filter(
        sku__icontains=query,
        kind__in=['interchange', 'equivalent'],
    ).values('product_id')
    products = products.filter(
        Q(name__icontains=query)
        | Q(description__icontains=query)
        | Q(sku__icontains=query)
        | Q(category__name__icontains=query)
        | Q(brand__name__icontains=query)
        | Q(vehicle_model__name__icontains=query)
        | Q(vin_number__vin__icontains=query)
        | Q(pk__in=alternate_ids)
    )

    brand_ids = [value for value in request.GET.getlist('brand') if value]
    model_ids = [value for value in request.GET.getlist('model') if value]
    vin_ids = [value for value in request.GET.getlist('vin') if value]
    if brand_ids:
        products = products.filter(brand_id__in=brand_ids)
    if model_ids:
        products = products.filter(vehicle_model_id__in=model_ids)
    if vin_ids:
        products = products.filter(vin_number_id__in=vin_ids)

    attr_params = {
        key: value
        for key, value in request.GET.items()
        if key.startswith('attr_') and value
    }
    if attr_params:
        attr_ids = []
        for key in attr_params.keys():
            try:
                attr_ids.append(int(key.split('_', 1)[1]))
            except (IndexError, ValueError, TypeError):
                continue
        attributes = CategoryAttribute.objects.filter(id__in=attr_ids, is_active=True)
        attr_map = {str(attribute.id): attribute for attribute in attributes}

        for param, selected_value in attr_params.items():
            attr_id = param.split('_', 1)[1]
            attribute = attr_map.get(attr_id)
            if not attribute:
                continue
            if attribute.attribute_type == 'select':
                products = products.filter(
                    attribute_values__attribute=attribute,
                    attribute_values__option_id=selected_value,
                )
            elif attribute.attribute_type == 'boolean':
                normalized = selected_value.lower()
                if normalized in ('1', 'true', 'yes', 'on'):
                    bool_value = True
                elif normalized in ('0', 'false', 'no', 'off'):
                    bool_value = False
                else:
                    bool_value = None
                if bool_value is not None:
                    products = products.filter(
                        attribute_values__attribute=attribute,
                        attribute_values__value_boolean=bool_value,
                    )
            elif attribute.attribute_type == 'number':
                try:
                    number_value = Decimal(selected_value)
                except (ArithmeticError, ValueError, TypeError):
                    number_value = None
                if number_value is not None:
                    products = products.filter(
                        attribute_values__attribute=attribute,
                        attribute_values__value_number=number_value,
                    )
            else:
                products = products.filter(
                    attribute_values__attribute=attribute,
                    attribute_values__value_text__iexact=selected_value,
                )

        products = products.distinct()

    try:
        limit = int(request.GET.get('limit', 8))
    except (TypeError, ValueError):
        limit = 8
    limit = max(1, min(limit, 12))
    category_limit = 5

    products = products.order_by('-is_featured', 'name')
    product_list = list(products[: limit + 1])
    has_more = len(product_list) > limit
    product_list = product_list[:limit]

    price_flags = resolve_storefront_price_flags(request, store_owner)
    show_prices = price_flags.get('catalog', False)

    results = []
    for product in product_list:
        price_display = ''
        price_badge = ''
        note = ''
        if show_prices:
            price_value = product.storefront_price
            if price_value is None:
                price_value = product.cost_price
            if price_value is not None:
                price_display = f"${price_value:.2f}"
            else:
                price_display = "Call for price"
            if product.has_promotion:
                if product.promotion_discount_percent:
                    price_badge = f"Save {product.promotion_discount_percent}%"
                else:
                    price_badge = "Promo"
        else:
            note = "Sign in to view price"

        available_qty = getattr(product, "stock_quantity", product.quantity_in_stock)
        if available_qty and available_qty > 0:
            stock_label = f"{available_qty} in stock"
        else:
            stock_label = "Out of stock"

        results.append({
            'id': product.id,
            'name': product.name,
            'sku': product.sku or '',
            'brand': product.brand.name if product.brand else '',
            'category': product.category.name if product.category else '',
            'image': product.image.url if product.image else '',
            'url': reverse('accounts:store_product_detail', args=[product.id]),
            'price': price_display,
            'price_badge': price_badge,
            'note': note,
            'stock_label': stock_label,
        })

    categories = (
        _storefront_categories_queryset(
            available_products,
            owner=store_owner,
            include_empty=show_empty_categories,
        )
        .filter(
            Q(name__icontains=query)
            | Q(group__name__icontains=query)
            | Q(parent__name__icontains=query)
        )
        .order_by('sort_order', 'name')
    )
    category_list = list(categories[:category_limit])
    category_results = []
    for category in category_list:
        category_path = _build_category_path(category)
        parent_labels = [cat.name for cat in category_path[:-1]]
        category_results.append({
            'id': category.id,
            'name': category.name,
            'group': category.group.name if category.group else '',
            'path': " / ".join(parent_labels),
            'url': reverse('accounts:store_category_detail', args=[category.id]),
        })

    return JsonResponse({
        'query': query,
        'results': results,
        'categories': category_results,
        'has_more': has_more,
    })


@customer_login_required
def customer_dashboard(request):
    """Display account information for customer portal users."""

    customer_account = request.user.customer_portal
    today = timezone.localdate()
    invoice_summary, recent_invoices = _summarize_customer_invoices(
        customer_account,
        today=today,
    )

    if is_parts_store_business():
        recent_credit_rows, recent_credit_total = _build_customer_credit_rows(
            customer_account,
            limit=3,
        )
        context = {
            'customer_account': customer_account,
            'invoice_summary': invoice_summary,
            'recent_invoices': recent_invoices,
            'recent_credit_rows': recent_credit_rows,
            'recent_credit_total': recent_credit_total,
        }
        context.update(_customer_portal_layout_context(request, active_slug='shop'))
        return render(request, 'store/customer_dashboard_storefront.html', context)

    workorders_qs = (
        customer_account.work_orders.select_related('vehicle')
        .prefetch_related('assignments__mechanic')
        .order_by('-date_created')
    )

    workorder_summary = {
        'total': workorders_qs.count(),
        'open': workorders_qs.filter(status__in=['pending', 'in_progress']).count(),
        'completed': workorders_qs.filter(status='completed').count(),
    }
    recent_workorders = list(workorders_qs[:6])
    for wo in recent_workorders:
        mechanic_names = [
            assignment.mechanic.name
            for assignment in wo.assignments.all()
            if assignment.mechanic_id
        ]
        wo.assigned_mechanic = ", ".join(mechanic_names) if mechanic_names else None

    vehicle_qs = customer_account.vehicles.all()
    vehicle_count = vehicle_qs.count()
    vehicle_preview = list(
        vehicle_qs.order_by('unit_number', 'make_model', 'vin_number')[:4]
    )

    active_statuses = VehicleMaintenanceTask.active_statuses()
    maintenance_qs = VehicleMaintenanceTask.objects.filter(
        vehicle__customer=customer_account,
        status__in=active_statuses,
    )
    maintenance_summary = {
        'active': maintenance_qs.count(),
        'overdue': maintenance_qs.filter(due_date__lt=today).count(),
    }

    quick_maintenance_form = CustomerPortalQuickMaintenanceForm(customer=customer_account)

    if request.method == 'POST' and request.POST.get('action') == 'quick_maintenance':
        quick_maintenance_form = CustomerPortalQuickMaintenanceForm(
            request.POST,
            customer=customer_account,
        )
        if quick_maintenance_form.is_valid():
            task = quick_maintenance_form.save(commit=False)
            task.user = customer_account.user
            task.save()
            messages.success(
                request,
                f"Maintenance '{task.title}' added for {task.vehicle}.",
            )
            return redirect('accounts:customer_dashboard')
        messages.error(
            request,
            'We could not add that maintenance reminder. Please fix the errors below and try again.',
        )

    maintenance_tasks = list(
        maintenance_qs.select_related('vehicle')
        .order_by('due_date', 'priority', 'title', 'pk')[:5]
    )

    vehicle_spending_raw = list(
        customer_account.invoices
        .values('vin_no', 'make_model', 'unit_no')
        .annotate(total=Coalesce(Sum('total_amount'), Decimal('0.00')))
        .order_by('-total')[:5]
    )

    vehicle_spending = [
        {
            'label': entry['make_model'] or entry['unit_no'] or entry['vin_no'] or 'Vehicle',
            'total': entry['total'],
        }
        for entry in vehicle_spending_raw
    ]

    vehicle_spending_chart_data = [
        {
            'label': vehicle['label'],
            'value': float(vehicle['total']),
        }
        for vehicle in vehicle_spending
    ]

    completed_tasks_qs = VehicleMaintenanceTask.objects.filter(
        vehicle__customer=customer_account,
        status=VehicleMaintenanceTask.STATUS_COMPLETED,
    )
    completed_total = completed_tasks_qs.count()
    on_time_completed = completed_tasks_qs.filter(
        Q(due_date__isnull=True) |
        Q(completed_date__isnull=False, completed_date__lte=F('due_date'))
    ).count() if completed_total else 0
    completion_rate = int(round((on_time_completed / completed_total) * 100)) if completed_total else 0

    maintenance_performance = {
        'completed': completed_total,
        'on_time': on_time_completed,
        'rate': completion_rate,
    }
    maintenance_performance['degree'] = float(maintenance_performance['rate']) * 3.6
    overdue_count = max(completed_total - on_time_completed, 0)
    maintenance_chart_data = [
        {'label': 'On time', 'value': float(on_time_completed)},
        {'label': 'Other', 'value': float(overdue_count)},
    ]
    monthly_spending_raw = list(
        customer_account.invoices.annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total=Coalesce(Sum('total_amount'), Decimal('0.00')))
        .order_by('month')
    )
    monthly_spending_raw = [entry for entry in monthly_spending_raw if entry['month']]
    recent_monthly_spending = monthly_spending_raw[-6:]
    monthly_spending_chart_data = [
        {
            'label': entry['month'].strftime('%b %Y'),
            'value': float(entry['total']),
        }
        for entry in recent_monthly_spending
    ]

    context = {
        'customer_account': customer_account,
        'invoice_summary': invoice_summary,
        'workorder_summary': workorder_summary,
        'recent_workorders': recent_workorders,
        'vehicle_count': vehicle_count,
        'maintenance_summary': maintenance_summary,
        'quick_maintenance_form': quick_maintenance_form,
        'maintenance_tasks': maintenance_tasks,
        'recent_invoices': recent_invoices,
        'vehicle_preview': vehicle_preview,
        'vehicle_preview_has_more': vehicle_count > len(vehicle_preview),
        'today': today,
        'vehicle_spending': vehicle_spending,
        'vehicle_spending_chart_data': vehicle_spending_chart_data,
        'maintenance_performance': maintenance_performance,
        'maintenance_chart_data': maintenance_chart_data,
        'monthly_spending_chart_data': monthly_spending_chart_data,
    }
    context.update(_customer_portal_layout_context(request, active_slug='overview'))
    return render(request, 'store/customer_dashboard.html', context)


def product_detail(request, pk):
    """Display a single product."""
    product = get_object_or_404(
        _storefront_product_queryset(request)
        .select_related('brand', 'category')
        .prefetch_related('alternate_skus'),
        pk=pk,
    )
    apply_stock_fields([product])
    attribute_values = list(
        product.attribute_values
        .select_related('attribute', 'option')
        .filter(attribute__is_active=True)
        .order_by('attribute__sort_order', 'attribute__name')
    )
    overview_text, feature_items, attribute_values = _resolve_product_detail_sections(
        product,
        attribute_values,
    )
    interchange_skus = list(
        product.alternate_skus
        .filter(kind__in=['interchange', 'equivalent'])
        .order_by('kind', 'sku')
        .values_list('sku', flat=True)
    )
    category_path = _build_category_path(product.category) if product.category else []
    active_group = product.category.group if product.category else None
    breadcrumbs = _build_storefront_breadcrumbs(
        group=active_group,
        category_path=category_path,
        label=product.name,
    )

    query_params = request.GET.copy()
    query_params.pop('partial', None)
    query_string = query_params.urlencode()
    if query_string:
        if category_path:
            back_url = f"{reverse('accounts:store_category_detail', args=[category_path[-1].id])}?{query_string}"
        else:
            back_url = f"{reverse('accounts:store_search')}?{query_string}"
    elif category_path:
        back_url = reverse('accounts:store_category_detail', args=[category_path[-1].id])
    elif active_group:
        back_url = reverse('accounts:store_group_detail', args=[active_group.id])
    else:
        back_url = reverse('accounts:store_product_list')

    context = {
        'product': product,
        'attribute_values': attribute_values,
        'product_overview': overview_text,
        'feature_items': feature_items,
        'interchange_skus': interchange_skus,
        'customer_account': getattr(request.user, 'customer_portal', None),
        'breadcrumbs': breadcrumbs,
        'back_url': back_url,
    }
    store_owner = get_storefront_owner(request)
    price_flags = resolve_storefront_price_flags(request, store_owner)
    context['show_prices_catalog'] = price_flags['catalog']
    if request.user.is_authenticated:
        cart_product_ids = _get_cart_product_ids(request)
        context['in_cart'] = str(product.id) in cart_product_ids
    else:
        context['in_cart'] = False
    return render(request, 'store/product_detail.html', context)


@customer_login_required
def add_to_cart(request, product_id):
    """Add a product to the customer's cart."""
    product = get_object_or_404(
        _storefront_product_queryset(request).filter(stock_quantity__gt=0),
        pk=product_id,
    )
    qty = int(request.POST.get('quantity', 1))
    available_qty = getattr(product, "stock_quantity", product.quantity_in_stock)
    qty = max(1, min(qty, available_qty))
    customer_account = request.user.customer_portal
    store_owner = get_storefront_owner(request)
    _sync_session_cart_to_db(request, customer_account, store_owner)

    cart_item = StorefrontCartItem.objects.filter(
        customer=customer_account,
        store_owner=store_owner,
        product=product,
    ).first()
    if cart_item:
        available_qty = getattr(product, "stock_quantity", product.quantity_in_stock)
        new_qty = min(cart_item.quantity + qty, available_qty)
        if new_qty != cart_item.quantity:
            cart_item.quantity = new_qty
            cart_item.save(update_fields=["quantity", "updated_at"])
    else:
        StorefrontCartItem.objects.create(
            customer=customer_account,
            store_owner=store_owner,
            product=product,
            quantity=qty,
        )
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        cart_count = _get_cart_count(customer_account, store_owner)
        return JsonResponse(
            {
                'status': 'success',
                'cart_count': cart_count,
                'product_id': product_id,
            }
        )
    return redirect(request.META.get('HTTP_REFERER') or reverse('accounts:store_cart'))


@customer_login_required
def cart_view(request):
    """Display cart contents."""
    cart = _get_cart(request)
    customer_account = request.user.customer_portal
    store_owner = get_storefront_owner(request)
    product_qs = _storefront_product_queryset(request, owner=store_owner)
    paid_items, total, subtotal_before_discounts, discount_total, _, cart_quantities = _build_storefront_cart_items(
        cart,
        product_qs,
    )
    free_items = _resolve_storefront_free_items(store_owner, cart_quantities)
    items = paid_items + free_items
    return render(
        request,
        'store/cart.html',
        {
            'items': items,
            'total': total,
            'customer_account': customer_account,
            'subtotal_before_discounts': subtotal_before_discounts,
            'discount_total': discount_total,
            'free_items': free_items,
        },
    )


@customer_login_required
def update_cart(request, product_id):
    """Update quantity or remove an item from the cart."""
    customer_account = request.user.customer_portal
    store_owner = get_storefront_owner(request)
    _sync_session_cart_to_db(request, customer_account, store_owner)
    cart_item = StorefrontCartItem.objects.filter(
        customer=customer_account,
        store_owner=store_owner,
        product_id=product_id,
    ).first()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'remove':
            if cart_item:
                cart_item.delete()
        elif action in {'increment', 'decrement'}:
            current_qty = cart_item.quantity if cart_item else 0

            product = _storefront_product_queryset(request, owner=store_owner).filter(pk=product_id).first()
            if not product:
                if cart_item:
                    cart_item.delete()
            else:
                if action == 'increment':
                    new_qty = current_qty + 1
                else:
                    new_qty = 1 if current_qty <= 1 else current_qty - 1

                available_qty = getattr(product, "stock_quantity", product.quantity_in_stock)
                new_qty = min(new_qty, available_qty)
                if new_qty <= 0:
                    if cart_item:
                        cart_item.delete()
                elif cart_item:
                    if new_qty != cart_item.quantity:
                        cart_item.quantity = new_qty
                        cart_item.save(update_fields=["quantity", "updated_at"])
                else:
                    StorefrontCartItem.objects.create(
                        customer=customer_account,
                        store_owner=store_owner,
                        product=product,
                        quantity=new_qty,
                    )
        else:
            try:
                qty = int(request.POST.get('quantity', 1))
                if qty > 0:
                    product = _storefront_product_queryset(request, owner=store_owner).filter(pk=product_id).first()
                    if product:
                        available_qty = getattr(product, "stock_quantity", product.quantity_in_stock)
                        qty = min(qty, available_qty)
                        if qty <= 0:
                            if cart_item:
                                cart_item.delete()
                        elif cart_item:
                            if qty != cart_item.quantity:
                                cart_item.quantity = qty
                                cart_item.save(update_fields=["quantity", "updated_at"])
                        else:
                            StorefrontCartItem.objects.create(
                                customer=customer_account,
                                store_owner=store_owner,
                                product=product,
                                quantity=qty,
                            )
                    else:
                        if cart_item:
                            cart_item.delete()
                else:
                    if cart_item:
                        cart_item.delete()
            except ValueError:
                pass
    return redirect('accounts:store_cart')


@customer_login_required
def checkout(request):
    """Checkout view that creates an invoice and sends it via email."""
    cart = _get_cart(request)
    if not cart:
        return redirect('accounts:public_home')

    store_owner = get_storefront_owner(request)
    product_qs = _storefront_product_queryset(request, owner=store_owner)
    paid_items, subtotal, subtotal_before_discounts, discount_total, sellers, cart_quantities = _build_storefront_cart_items(
        cart,
        product_qs,
    )
    free_items = _resolve_storefront_free_items(store_owner, cart_quantities)
    items = paid_items + free_items

    if not paid_items:
        messages.error(
            request,
            'Your cart no longer contains products that are available for purchase.',
        )
        _clear_cart(request.user.customer_portal, store_owner)
        request.session.pop("cart", None)
        return redirect('accounts:store_cart')

    customer_info = _get_authenticated_customer_info(request.user)
    customer_info['cc_emails_raw'] = ", ".join(customer_info.get('cc_emails', []))
    if request.method == 'POST':
        customer_info = customer_info.copy()
        customer_info['name'] = (request.POST.get('billing_name') or customer_info['name']).strip()
        customer_info['email'] = (request.POST.get('billing_email') or customer_info['email']).strip()
        customer_info['phone'] = (request.POST.get('billing_phone') or '').strip()
        customer_info['address'] = (request.POST.get('billing_address') or '').strip()
        cc_emails_raw = (request.POST.get('billing_cc_emails') or '').strip()
        cc_emails = Customer.parse_cc_emails(cc_emails_raw)
        customer_info['cc_emails'] = cc_emails
        customer_info['cc_emails_raw'] = ", ".join(cc_emails)
    missing_fields = {
        'email': not bool(customer_info.get('email')),
        'phone': not bool(customer_info.get('phone')),
        'address': not bool(customer_info.get('address')),
    }
    error = None
    customer_account = getattr(request.user, 'customer_portal', None)

    if not store_owner:
        if len(sellers) != 1:
            return render(request, 'store/checkout.html', {
                'items': items,
                'subtotal_before_discounts': subtotal_before_discounts,
                'discount_total': discount_total,
                'subtotal': subtotal,
                'customer_info': customer_info,
                'missing_fields': missing_fields,
                'error': 'Products from multiple sellers cannot be purchased together.',
                'customer_account': customer_account,
                'free_items': free_items,
            })
        seller = sellers.pop()
    else:
        seller = store_owner

    if missing_fields['email']:
        error = 'Please update your account email before completing checkout.'

    if customer_account and not _customer_matches_store(customer_account, seller):
        error = 'This account is not associated with this business.'

    tax_label = 'Tax'
    tax_total = Decimal('0.00')
    seller_profile = getattr(seller, 'profile', None)
    if seller_profile:
        tax_label = seller_profile.get_tax_name() if hasattr(seller_profile, 'get_tax_name') else 'Tax'
        if seller_profile.province:
            tax_total = calculate_tax_total(subtotal, seller_profile.province)
    total_due = subtotal + tax_total

    if request.method == 'POST' and not error:
        if not customer_account:
            error = 'Unable to locate your customer account. Please contact support.'
        else:
            updates = {}
            if customer_info['name'] and customer_account.name != customer_info['name']:
                updates['name'] = customer_info['name']
            if customer_account.address != customer_info['address']:
                updates['address'] = customer_info['address']
            if customer_account.phone_number != customer_info['phone']:
                updates['phone_number'] = customer_info['phone']
            if customer_account.email != customer_info['email']:
                updates['email'] = customer_info['email']
            if customer_account.cc_emails != customer_info.get('cc_emails_raw', ''):
                updates['cc_emails'] = customer_info.get('cc_emails_raw', '')
            if updates:
                for field, value in updates.items():
                    setattr(customer_account, field, value)
                customer_account.save(update_fields=list(updates.keys()))

            invoice = GroupedInvoice.objects.create(
                user=seller,
                customer=customer_account,
                date=timezone.now().date(),
                bill_to=customer_info['name'],
                bill_to_email=customer_info['email'] or None,
                bill_to_address=customer_info['address'] or None,
            )

            for item in paid_items:
                IncomeRecord2.objects.create(
                    grouped_invoice=invoice,
                    product=item['product'],
                    job=item['product'].name,
                    qty=Decimal(item['quantity']),
                    rate=item['unit_price'],
                )
            for item in free_items:
                bonus_label = item['product'].name
                if item.get('package_title'):
                    bonus_label = f"{bonus_label} (Free with {item['package_title']})"
                IncomeRecord2.objects.create(
                    grouped_invoice=invoice,
                    product=item['product'],
                    job=bonus_label,
                    qty=Decimal(item['quantity']),
                    rate=Decimal('0.00'),
                )

            invoice.ensure_inventory_transactions()
            invoice.recalculate_total_amount()
            invoice.create_online_payment_link()
            PendingInvoice.objects.get_or_create(grouped_invoice=invoice)

            original_user = request.user
            try:
                request.user = seller
                send_grouped_invoice_email(request, invoice.pk)
            finally:
                request.user = original_user

            _clear_cart(request.user.customer_portal, store_owner)
            request.session.pop("cart", None)
            return redirect('accounts:store_order_complete', invoice_number=invoice.invoice_number)

    return render(request, 'store/checkout.html', {
        'items': items,
        'subtotal_before_discounts': subtotal_before_discounts,
        'discount_total': discount_total,
        'subtotal': subtotal,
        'tax_label': tax_label,
        'tax_total': tax_total,
        'total_due': total_due,
        'customer_info': customer_info,
        'missing_fields': missing_fields,
        'error': error,
        'customer_account': customer_account,
        'free_items': free_items,
    })


@login_required
def store_hub(request):
    """Summary hub for managing the public storefront."""

    store_user_ids = get_product_user_ids(request.user)
    product_qs = Product.objects.filter(user__in=store_user_ids)
    group_qs = CategoryGroup.objects.filter(user__in=store_user_ids)
    category_qs = Category.objects.filter(user__in=store_user_ids)
    profile = getattr(request.user, 'profile', None)
    price_form = StorefrontPriceVisibilityForm(instance=profile) if profile else None

    if request.method == 'POST' and request.POST.get('save_price_settings') and profile:
        price_form = StorefrontPriceVisibilityForm(request.POST, instance=profile)
        if price_form.is_valid():
            price_form.save()
            messages.success(request, 'Storefront price visibility updated successfully.')
        else:
            messages.error(request, 'Please review the price visibility settings below.')

    context = {
        'public_store_url': request.build_absolute_uri(reverse('accounts:store_product_list')),
        'product_count': product_qs.count(),
        'published_product_count': product_qs.filter(is_published_to_store=True).count(),
        'featured_product_count': product_qs.filter(is_featured=True).count(),
        'promotion_product_count': product_qs.filter(promotion_price__isnull=False).count(),
        'group_count': group_qs.count(),
        'active_group_count': group_qs.filter(is_active=True).count(),
        'category_count': category_qs.count(),
        'active_category_count': category_qs.filter(is_active=True).count(),
        'price_form': price_form,
    }
    return render(request, 'store/hub.html', context)


@login_required
def manage_storefront(request):
    """Allow business users to decide which products appear in the public store."""

    ProductFormSet = modelformset_factory(
        Product,
        fields=['is_published_to_store', 'is_featured', 'promotion_price'],
        extra=0,
        widgets={
            'promotion_price': forms.NumberInput(attrs={'class': 'form-control form-control-sm', 'min': '0', 'step': '0.01'}),
        },
    )
    store_user_ids = get_product_user_ids(request.user)
    queryset = annotate_products_with_stock(
        Product.objects.filter(user__in=store_user_ids).order_by('name'),
        request.user,
    )
    formset = ProductFormSet(request.POST or None, queryset=queryset)
    apply_stock_fields([form.instance for form in formset.forms])

    for form in formset.forms:
        field = form.fields.get('is_published_to_store')
        if field:
            field.widget.attrs.setdefault('class', 'form-check-input')
        featured_field = form.fields.get('is_featured')
        if featured_field:
            featured_field.widget.attrs.setdefault('class', 'form-check-input')
        promo_field = form.fields.get('promotion_price')
        if promo_field:
            promo_field.widget.attrs.setdefault('class', 'form-control form-control-sm')

    if request.method == 'POST':
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Storefront settings updated successfully.')
            return redirect('accounts:store_manage')
        messages.error(request, 'Please review the errors highlighted below.')

    public_store_url = request.build_absolute_uri(reverse('accounts:store_product_list'))
    product_qs = Product.objects.filter(user__in=store_user_ids)
    group_qs = CategoryGroup.objects.filter(user__in=store_user_ids).order_by('sort_order', 'name')
    category_qs = Category.objects.filter(user__in=store_user_ids).select_related('group').order_by('sort_order', 'name')

    context = {
        'formset': formset,
        'product_count': formset.total_form_count(),
        'public_store_url': public_store_url,
        'group_count': group_qs.count(),
        'active_group_count': group_qs.filter(is_active=True).count(),
        'category_count': category_qs.count(),
        'active_category_count': category_qs.filter(is_active=True).count(),
        'published_product_count': product_qs.filter(is_published_to_store=True).count(),
        'featured_product_count': product_qs.filter(is_featured=True).count(),
        'promotion_product_count': product_qs.filter(promotion_price__isnull=False).count(),
        'group_options': group_qs,
        'category_options': category_qs,
    }
    return render(request, 'store/manage_storefront.html', context)


@login_required
def manage_storefront_hero(request):
    """Manage the storefront promotions section."""

    business_user = get_business_user(request.user) or request.user
    hero, _ = StorefrontHeroShowcase.objects.get_or_create(user=business_user)
    banner, _ = StorefrontMessageBanner.objects.get_or_create(user=business_user)
    flyer, _ = StorefrontFlyer.objects.get_or_create(user=business_user)
    package_queryset = StorefrontHeroPackage.objects.filter(user=business_user).order_by('id')

    SlideFormSet = inlineformset_factory(
        StorefrontHeroShowcase,
        StorefrontHeroShowcaseItem,
        form=StorefrontHeroShowcaseItemForm,
        extra=6,
        can_delete=True,
    )
    PackageFormSet = modelformset_factory(
        StorefrontHeroPackage,
        form=StorefrontHeroPackageForm,
        extra=3,
        can_delete=True,
    )

    if request.method == 'POST':
        hero_form = StorefrontHeroShowcaseForm(request.POST, instance=hero)
        slide_formset = SlideFormSet(request.POST, instance=hero, prefix='slides')
        banner_form = StorefrontMessageBannerForm(request.POST, instance=banner)
        flyer_form = StorefrontFlyerForm(request.POST, instance=flyer)
        package_formset = PackageFormSet(request.POST, queryset=package_queryset, prefix='packages')
    else:
        hero_form = StorefrontHeroShowcaseForm(instance=hero)
        slide_formset = SlideFormSet(instance=hero, prefix='slides')
        banner_form = StorefrontMessageBannerForm(instance=banner)
        flyer_form = StorefrontFlyerForm(instance=flyer)
        package_formset = PackageFormSet(queryset=package_queryset, prefix='packages')

    product_queryset = Product.objects.filter(
        user=business_user,
        is_published_to_store=True,
    ).select_related('category', 'brand').order_by('name')
    for form in slide_formset.forms:
        if 'product' in form.fields:
            form.fields['product'].queryset = product_queryset
    for form in package_formset.forms:
        for field_name in ('primary_product', 'secondary_product', 'free_product'):
            if field_name in form.fields:
                form.fields[field_name].queryset = product_queryset

    if request.method == 'POST':
        forms_valid = (
            hero_form.is_valid()
            and slide_formset.is_valid()
            and banner_form.is_valid()
            and flyer_form.is_valid()
            and package_formset.is_valid()
        )
        if forms_valid:
            hero_instance = hero_form.save(commit=False)
            hero_instance.user = business_user
            hero_instance.save()
            banner_instance = banner_form.save(commit=False)
            banner_instance.user = business_user
            banner_instance.save()
            flyer_instance = flyer_form.save(commit=False)
            flyer_instance.user = business_user
            flyer_instance.save()
            slide_formset.save()
            package_instances = package_formset.save(commit=False)
            for package in package_instances:
                package.user = business_user
                package.save()
            for package in package_formset.deleted_objects:
                package.delete()
            _apply_storefront_hero_discounts(
                hero_instance,
                StorefrontHeroPackage.objects.filter(user=business_user),
            )
            messages.success(request, 'Storefront promotions updated successfully.')
            return redirect('accounts:store_hero')
        messages.error(request, 'Please fix the errors below and try again.')

    product_payload = []
    for product in product_queryset:
        if product.storefront_price is not None:
            price_display = f"${product.storefront_price:.2f}"
        else:
            price_display = "Call for price"
        product_payload.append({
            'id': product.id,
            'name': product.name,
            'image': product.image.url if product.image else '',
            'price': price_display,
            'category': product.category.name if product.category else 'Uncategorized',
            'brand': product.brand.name if product.brand else '',
            'sku': product.sku or '',
            'source_name': product.source_name or '',
            'source_product_id': product.source_product_id or '',
            'sale_price': f"{product.sale_price:.2f}" if product.sale_price is not None else None,
            'promotion_price': (
                f"{product.promotion_price:.2f}" if product.promotion_price is not None else None
            ),
            'has_promotion': product.has_promotion,
            'promotion_discount_percent': product.promotion_discount_percent,
            'detail_url': reverse('accounts:store_product_detail', args=[product.id]),
        })

    price_flags = resolve_storefront_price_flags(request, business_user)
    context = {
        'hero_form': hero_form,
        'slide_formset': slide_formset,
        'banner_form': banner_form,
        'flyer_form': flyer_form,
        'package_formset': package_formset,
        'public_home_url': request.build_absolute_uri(reverse('accounts:public_home')),
        'hero_product_data': product_payload,
        'show_prices_hero': price_flags['hero'],
    }
    return render(request, 'store/storefront_hero.html', context)


def storefront_flyer_pdf(request):
    store_owner = get_storefront_owner(request)
    if not store_owner:
        return HttpResponse("Flyer unavailable.", status=404)

    flyer = StorefrontFlyer.objects.filter(user=store_owner, is_active=True).first()
    if not flyer:
        return HttpResponse("Flyer unavailable.", status=404)

    available_products = _storefront_product_queryset(request, owner=store_owner)
    marketing_context = _build_storefront_marketing_context(
        request,
        available_products,
        store_owner=store_owner,
    )
    if not marketing_context.get('flyer_has_content'):
        return HttpResponse("Flyer unavailable.", status=404)

    profile = getattr(store_owner, 'profile', None)
    context = {
        **marketing_context,
        'flyer': flyer,
        'profile': profile,
        'business_user': store_owner,
        'store_url': request.build_absolute_uri(reverse('accounts:store_product_list')),
        'contact_url': request.build_absolute_uri(reverse('accounts:public_contact')),
        'request': request,
    }
    try:
        pdf_bytes = render_template_to_pdf(
            'store/storefront_flyer_pdf.html',
            context,
            base_url=request.build_absolute_uri('/'),
        )
    except ImportError as exc:
        return HttpResponse(str(exc), status=501)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="storefront_flyer.pdf"'
    return response


@customer_login_required
def order_complete(request, invoice_number):
    """Simple confirmation page after order submission."""
    return render(
        request,
        'store/order_complete.html',
        {
            'invoice_number': invoice_number,
            'customer_account': request.user.customer_portal,
        },
    )


@customer_login_required
def customer_profile(request):
    """Allow customer portal users to review and update their contact details."""

    customer_account = request.user.customer_portal
    form = CustomerPortalProfileForm(request.POST or None, instance=customer_account)

    if request.method == 'POST' and form.is_valid():
        form.save()
        email = form.cleaned_data.get('email')
        if email and request.user.email != email:
            request.user.email = email
            request.user.save(update_fields=['email'])
        messages.success(request, 'Your profile has been updated successfully.')
        return redirect('accounts:customer_profile')

    context = {
        'form': form,
    }
    context.update(_customer_portal_layout_context(request, active_slug='profile'))
    return render(request, 'store/customer_profile.html', context)


@customer_login_required
def customer_invoice_list(request):
    """Display all invoices associated with the logged-in customer."""

    customer_account = request.user.customer_portal
    search_query = (request.GET.get('search') or '').strip()
    invoices_qs = (
        customer_account.invoices
        .select_related('user')
        .prefetch_related('payments')
        .annotate(
            total_paid_amount=Coalesce(
                Sum('payments__amount'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )
    )
    invoices_qs = _annotate_invoice_credit_totals(invoices_qs).annotate(
        balance_due_amount=ExpressionWrapper(
            F('total_amount') - F('total_paid_amount') - F('credit_total'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    )

    if search_query:
        invoices_qs = invoices_qs.filter(
            Q(invoice_number__icontains=search_query) |
            Q(bill_to__icontains=search_query) |
            Q(payment_status__icontains=search_query) |
            Q(total_amount__icontains=search_query)
        )

    invoices_qs = invoices_qs.order_by('-date', '-id')
    paginator = Paginator(invoices_qs, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()
    params.pop('page', None)
    credit_rows, credit_total = _build_customer_credit_rows(customer_account)
    context = {
        'invoices': page_obj,
        'page_obj': page_obj,
        'query_string': params.urlencode(),
        'search_query': search_query,
        'credit_rows': credit_rows,
        'credit_total': credit_total,
    }
    context.update(_customer_portal_layout_context(request, active_slug='invoices'))
    return render(request, 'store/customer_invoices.html', context)


@customer_login_required
def customer_returns(request):
    """Display customer credits for returned parts."""

    customer_account = request.user.customer_portal
    credit_rows, total_credit = _build_customer_credit_rows(customer_account)
    context = {
        'credit_rows': credit_rows,
        'total_credit': total_credit,
    }
    context.update(_customer_portal_layout_context(request, active_slug='returns'))
    return render(request, 'store/customer_returns.html', context)


@customer_login_required
def customer_invoice_download(request, invoice_id):
    """Allow customer portal users to download individual invoices as PDFs."""

    customer_account = request.user.customer_portal
    invoice = get_object_or_404(
        GroupedInvoice.objects.select_related('user__profile', 'customer').prefetch_related('income_records', 'payments'),
        pk=invoice_id,
        customer=customer_account,
    )

    return _customer_invoice_pdf_response(invoice, request, inline=False)


@customer_login_required
def customer_invoice_print(request, invoice_id):
    """Return the invoice PDF inline so the browser can open its print dialog."""

    customer_account = request.user.customer_portal
    invoice = get_object_or_404(
        GroupedInvoice.objects.select_related('user__profile', 'customer').prefetch_related('income_records', 'payments'),
        pk=invoice_id,
        customer=customer_account,
    )

    return _customer_invoice_pdf_response(invoice, request, inline=True)


@customer_login_required
def customer_invoice_statements(request):
    """Allow customers to export paid-invoice statements for accountant use."""

    customer_account = request.user.customer_portal
    today = timezone.localdate()

    if request.method == 'POST':
        form = CustomerStatementForm(request.POST)
    else:
        form = CustomerStatementForm(request.GET or None, initial={'period': 'month', 'reference_date': today})

    invoices = []
    totals = {'count': 0, 'subtotal': Decimal('0.00'), 'tax': Decimal('0.00'), 'total': Decimal('0.00'), 'paid': Decimal('0.00')}
    start_date = end_date = None
    period_label = ''
    period_value = form['period'].value() if 'period' in form.fields else None
    current_period = period_value or form.fields['period'].initial or 'month'
    reference_raw = form['reference_date'].value() if 'reference_date' in form.fields else None
    if isinstance(reference_raw, str) and reference_raw:
        try:
            current_reference_date = datetime.strptime(reference_raw, '%Y-%m-%d').date()
        except ValueError:
            current_reference_date = today
    elif isinstance(reference_raw, date):
        current_reference_date = reference_raw
    else:
        initial_reference = form.initial.get('reference_date') if getattr(form, 'initial', None) else None
        current_reference_date = initial_reference or today

    if form.is_valid():
        period = form.cleaned_data['period']
        reference_date = form.cleaned_data.get('reference_date') or today
        start_date, end_date, period_label = _resolve_statement_period(period, reference_date)
        current_period = period
        current_reference_date = reference_date

        invoices_qs = (
            customer_account.invoices
            .filter(date__gte=start_date, date__lte=end_date)
            .annotate(
                subtotal=Coalesce(Sum('income_records__amount'), Value(Decimal('0.00')), output_field=DecimalField()),
                tax_total=Coalesce(Sum('income_records__tax_collected'), Value(Decimal('0.00')), output_field=DecimalField()),
                total_paid=Coalesce(Sum('payments__amount'), Value(Decimal('0.00')), output_field=DecimalField()),
            )
        )
        invoices_qs = invoices_qs.annotate(
            balance=ExpressionWrapper(
                F('total_amount') - F('total_paid'),
                output_field=DecimalField(),
            )
        )
        invoices_qs = (
            invoices_qs
            .filter(balance__lte=Decimal('0.01'))
            .select_related('user__profile')
            .prefetch_related('payments')
            .order_by('date', 'invoice_number')
        )
        invoices = list(invoices_qs)
        for invoice in invoices:
            payments = sorted(
                invoice.payments.all(),
                key=lambda payment: ((payment.date or date.min), payment.pk or 0)
            )
            invoice.latest_payment = payments[-1] if payments else None
        totals = _statement_totals(invoices)

        export_format = request.POST.get('export_format')
        if request.method == 'POST' and export_format in {'pdf', 'excel'}:
            if not invoices:
                messages.info(request, 'No paid invoices were found for the selected period.')
            elif export_format == 'pdf':
                return _statement_pdf_response(
                    customer_account,
                    invoices,
                    start_date=start_date,
                    end_date=end_date,
                    period_label=period_label,
                    totals=totals,
                    request=request,
                )
            else:
                return _statement_excel_response(
                    customer_account,
                    invoices,
                    start_date=start_date,
                    end_date=end_date,
                    period_label=period_label,
                    totals=totals,
                )

    profile = getattr(customer_account.user, 'profile', None)
    owner = getattr(customer_account, 'user', None)
    if profile and getattr(profile, 'company_name', None):
        business_name = profile.company_name
    elif owner:
        business_name = owner.get_full_name() or owner.get_username()
    else:
        business_name = customer_account.name

    context = {
        'form': form,
        'invoices': invoices,
        'totals': totals,
        'start_date': start_date,
        'end_date': end_date,
        'period_label': period_label,
        'logo_url': _statement_logo_url(getattr(customer_account.user, 'profile', None), request=request),
        'profile': profile,
        'business_name': business_name,
        'current_period': current_period,
        'current_reference_date': current_reference_date,
    }
    context.update(_customer_portal_layout_context(request, active_slug='statements'))
    return render(request, 'store/customer_statements.html', context)


@customer_login_required
def customer_workorder_list(request):
    """List work orders for the logged-in customer."""

    customer_account = request.user.customer_portal
    workorders = list(
        customer_account.work_orders
        .select_related('vehicle')
        .prefetch_related('assignments__mechanic', 'records__product__supplier')
        .order_by('-date_created')
    )

    for workorder in workorders:
        assignments = list(workorder.assignments.all())
        workorder.technician_names = [
            assignment.mechanic.name
            for assignment in assignments
            if assignment.mechanic and assignment.mechanic.name
        ]

        parts_replaced = []
        parts_suppliers = []
        for record in workorder.records.all():
            product = record.product
            if not product:
                continue
            if product.name and product.name not in parts_replaced:
                parts_replaced.append(product.name)
            supplier = product.supplier
            if supplier and supplier.name and supplier.name not in parts_suppliers:
                parts_suppliers.append(supplier.name)

        workorder.parts_replaced = parts_replaced
        workorder.parts_suppliers = parts_suppliers
        workorder.work_performed = workorder.correction or workorder.description or ""

    context = {
        'workorders': workorders,
    }
    context.update(_customer_portal_layout_context(request, active_slug='workorders'))
    return render(request, 'store/customer_workorders.html', context)


@customer_login_required
def customer_workorder_download(request, workorder_id):
    """Allow a portal customer to download a completed workorder PDF."""

    customer_account = request.user.customer_portal
    workorder = get_object_or_404(
        WorkOrder.objects.select_related('user__profile').prefetch_related('records'),
        pk=workorder_id,
        customer=customer_account,
    )
    if workorder.status != 'completed':
        return HttpResponseForbidden("Work order must be completed before downloading.")

    return _customer_workorder_pdf_response(workorder, request)


@customer_login_required
def customer_vehicle_overview(request):
    """Display all vehicles associated with the logged-in customer."""

    customer_account = request.user.customer_portal
    vehicle_form = VehicleForm()

    if request.method == 'POST' and request.POST.get('action') == 'add_vehicle':
        vehicle_form = VehicleForm(request.POST)
        if vehicle_form.is_valid():
            vehicle = vehicle_form.save(commit=False)
            vehicle.customer = customer_account
            try:
                vehicle.save()
            except IntegrityError:
                vehicle_form.add_error(
                    'vin_number',
                    'A vehicle with this VIN is already on file.',
                )
            else:
                customer_account.update_vehicle_count()
                messages.success(
                    request,
                    'Vehicle added successfully. You can now track maintenance reminders for it.',
                )
                return redirect('accounts:customer_vehicle_overview')
        else:
            messages.error(request, 'Please review the highlighted fields and try again.')

    active_statuses = VehicleMaintenanceTask.active_statuses()
    today = timezone.localdate()

    latest_job_date = JobHistory.objects.filter(
        vehicle=OuterRef('pk')
    ).order_by('-job_date').values('job_date')[:1]

    next_due_task = VehicleMaintenanceTask.objects.filter(
        vehicle=OuterRef('pk'),
        status__in=active_statuses,
    ).order_by('due_date', 'priority', 'title', 'pk')

    vehicles_qs = (
        customer_account.vehicles.annotate(
            last_job_date=Subquery(latest_job_date),
            job_count=Count('job_history', distinct=True),
            upcoming_maintenance=Count(
                'maintenance_tasks',
                filter=Q(maintenance_tasks__status__in=active_statuses),
                distinct=True,
            ),
            overdue_maintenance=Count(
                'maintenance_tasks',
                filter=Q(
                    maintenance_tasks__status__in=active_statuses,
                    maintenance_tasks__due_date__lt=today,
                ),
                distinct=True,
            ),
            next_due_date=Subquery(next_due_task.values('due_date')[:1]),
            next_due_title=Subquery(next_due_task.values('title')[:1]),
        )
        .order_by('unit_number', 'vin_number')
    )

    vehicles = list(vehicles_qs)
    service_columns = list(
        Service.objects.filter(
            user=customer_account.user,
            is_active=True,
            show_on_customer_portal=True,
        )
        .select_related('job_name')
        .order_by('job_name__name', 'name')[:4]
    )

    vehicle_ids = [vehicle.id for vehicle in vehicles]
    service_ids = [service.id for service in service_columns]
    service_maps = {}
    override_map = {}
    if service_ids and vehicle_ids:
        override_map = {
            (override.vehicle_id, override.service_id): override
            for override in VehicleServicePortalOverride.objects.filter(
                vehicle_id__in=vehicle_ids,
                service_id__in=service_ids,
            )
        }
    if service_columns and vehicle_ids:
        for service in service_columns:
            terms = [service.name]
            if service.job_name_id and service.job_name and service.job_name.name:
                terms.append(service.job_name.name)
            normalized_terms = []
            for term in terms:
                trimmed = (term or '').strip()
                if trimmed and trimmed not in normalized_terms:
                    normalized_terms.append(trimmed)

            task_match = Q()
            history_match = Q()
            for term in normalized_terms:
                task_match |= Q(title__icontains=term) | Q(description__icontains=term)
                history_match |= Q(description__icontains=term)

            completed_map = {}
            upcoming_map = {}
            history_map = {}

            if normalized_terms:
                completed_rows = (
                    VehicleMaintenanceTask.objects.filter(
                        vehicle_id__in=vehicle_ids,
                        status=VehicleMaintenanceTask.STATUS_COMPLETED,
                    )
                    .filter(task_match)
                    .values('vehicle_id')
                    .annotate(
                        last_completed=Max('completed_date'),
                        last_mileage=Max('actual_mileage'),
                    )
                    .values_list('vehicle_id', 'last_completed', 'last_mileage')
                )
                completed_map = {
                    vehicle_id: {'date': last_completed, 'mileage': last_mileage}
                    for vehicle_id, last_completed, last_mileage in completed_rows
                }

                upcoming_rows = (
                    VehicleMaintenanceTask.objects.filter(
                        vehicle_id__in=vehicle_ids,
                        status__in=active_statuses,
                    )
                    .filter(task_match)
                    .values('vehicle_id')
                    .annotate(
                        next_due_date=Min('due_date'),
                        next_due_mileage=Min('due_mileage'),
                    )
                    .values_list('vehicle_id', 'next_due_date', 'next_due_mileage')
                )
                upcoming_map = {
                    vehicle_id: {'date': next_due_date, 'mileage': next_due_mileage}
                    for vehicle_id, next_due_date, next_due_mileage in upcoming_rows
                }

                history_rows = (
                    JobHistory.objects.filter(
                        vehicle_id__in=vehicle_ids,
                    )
                    .filter(history_match)
                    .values('vehicle_id')
                    .annotate(last_job=Max('job_date'))
                    .values_list('vehicle_id', 'last_job')
                )
                history_map = {
                    vehicle_id: last_job
                    for vehicle_id, last_job in history_rows
                }

            service_maps[service.id] = {
                'completed': completed_map,
                'upcoming': upcoming_map,
                'history': history_map,
            }

    for vehicle in vehicles:
        portal_cells = []
        for service in service_columns:
            maps = service_maps.get(service.id, {})
            completed_entry = maps.get('completed', {}).get(vehicle.id, {})
            upcoming_entry = maps.get('upcoming', {}).get(vehicle.id, {})
            history_entry = maps.get('history', {}).get(vehicle.id)
            last_date = completed_entry.get('date') or history_entry
            last_mileage = completed_entry.get('mileage')
            next_due_date = upcoming_entry.get('date')
            next_due_mileage = upcoming_entry.get('mileage')
            override = override_map.get((vehicle.id, service.id))
            if override:
                if override.last_service_date is not None:
                    last_date = override.last_service_date
                if override.last_service_mileage is not None:
                    last_mileage = override.last_service_mileage
                if override.next_service_date is not None:
                    next_due_date = override.next_service_date
                if override.next_service_mileage is not None:
                    next_due_mileage = override.next_service_mileage
            portal_cells.append(
                {
                    'last_date': last_date,
                    'last_mileage': last_mileage,
                    'next_due_date': next_due_date,
                    'next_due_mileage': next_due_mileage,
                }
            )
        vehicle.portal_service_cells = portal_cells
    summary = {
        'vehicle_count': len(vehicles),
        'upcoming_total': sum(v.upcoming_maintenance or 0 for v in vehicles),
        'overdue_total': sum(v.overdue_maintenance or 0 for v in vehicles),
    }

    next_due_dates = sorted(
        [v.next_due_date for v in vehicles if v.next_due_date]
    )
    summary['next_due_date'] = next_due_dates[0] if next_due_dates else None

    context = {
        'vehicles': vehicles,
        'service_columns': service_columns,
        'summary': summary,
        'today': today,
        'vehicle_form': vehicle_form,
        'vehicle_status_choices': Vehicle.STATUS_CHOICES,
    }
    context.update(_customer_portal_layout_context(request, active_slug='vehicles'))
    return render(request, 'store/customer_vehicle_list.html', context)


@customer_login_required
@require_POST
def customer_vehicle_inline_update(request, vehicle_id):
    customer_account = request.user.customer_portal
    vehicle = get_object_or_404(
        Vehicle,
        pk=vehicle_id,
        customer=customer_account,
    )

    payload = {}
    if request.content_type and 'application/json' in request.content_type:
        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except json.JSONDecodeError:
            return JsonResponse(
                {'success': False, 'message': 'Invalid request payload.'},
                status=400,
            )

    if not isinstance(payload, dict):
        payload = {}

    vehicle_payload = payload.get('vehicle') or {}
    if not isinstance(vehicle_payload, dict):
        vehicle_payload = {}

    allowed_fields = {
        'unit_number',
        'vin_number',
        'make_model',
        'year',
        'license_plate',
        'status',
        'start_date_in_service',
        'assigned_to',
    }
    form_data = {}
    for field in VehicleForm.Meta.fields:
        value = getattr(vehicle, field, '')
        if isinstance(value, date):
            value = value.isoformat()
        if value is None:
            value = ''
        form_data[field] = value
    for field in allowed_fields:
        if field in vehicle_payload:
            form_data[field] = vehicle_payload.get(field) or ''

    form = VehicleForm(form_data, instance=vehicle)
    if not form.is_valid():
        return JsonResponse(
            {'success': False, 'errors': form.errors, 'message': 'Please correct the highlighted fields.'},
            status=400,
        )

    services_payload = payload.get('services') or {}
    if not isinstance(services_payload, dict):
        services_payload = {}

    service_ids = []
    for raw_id in services_payload.keys():
        try:
            service_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    allowed_service_ids = set(
        Service.objects.filter(
            user=customer_account.user,
            show_on_customer_portal=True,
            is_active=True,
            id__in=service_ids,
        ).values_list('id', flat=True)
    )

    service_errors = {}
    service_updates = []

    def parse_date_value(value):
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return 'invalid'

    for raw_id, service_payload in services_payload.items():
        try:
            service_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if service_id not in allowed_service_ids:
            continue
        if not isinstance(service_payload, dict):
            continue

        last_date_raw = service_payload.get('last_date') or ''
        next_date_raw = service_payload.get('next_date') or ''
        last_date = parse_date_value(last_date_raw)
        next_date = parse_date_value(next_date_raw)

        if last_date == 'invalid' or next_date == 'invalid':
            service_errors[service_id] = 'Enter valid dates for service updates.'
            continue

        service_updates.append(
            {
                'service_id': service_id,
                'last_date': last_date,
                'next_date': next_date,
            }
        )

    if service_errors:
        return JsonResponse(
            {'success': False, 'errors': service_errors, 'message': 'Please correct the service dates.'},
            status=400,
        )

    vehicle = form.save()

    service_responses = {}
    for update in service_updates:
        service_id = update['service_id']
        last_date = update['last_date']
        next_date = update['next_date']
        if last_date or next_date:
            override, _ = VehicleServicePortalOverride.objects.get_or_create(
                vehicle=vehicle,
                service_id=service_id,
            )
            override.last_service_date = last_date
            override.next_service_date = next_date
            override.save()
        else:
            VehicleServicePortalOverride.objects.filter(
                vehicle=vehicle,
                service_id=service_id,
            ).delete()

        service_responses[service_id] = {
            'last_date': last_date.isoformat() if last_date else '',
            'next_date': next_date.isoformat() if next_date else '',
        }

    response = {
        'success': True,
        'vehicle': {
            'unit_number': vehicle.unit_number or '',
            'vin_number': vehicle.vin_number or '',
            'make_model': vehicle.make_model or '',
            'year': vehicle.year or '',
            'license_plate': vehicle.license_plate or '',
            'status': vehicle.status or '',
            'status_label': vehicle.get_status_display() if vehicle.status else '',
            'start_date_in_service': vehicle.start_date_in_service.isoformat() if vehicle.start_date_in_service else '',
            'assigned_to': vehicle.assigned_to or '',
        },
        'services': service_responses,
    }
    return JsonResponse(response)


@customer_login_required
def customer_vehicle_detail(request, vehicle_id):
    """Detailed dashboard for a single vehicle in the customer portal."""

    customer_account = request.user.customer_portal
    vehicle = get_object_or_404(
        Vehicle.objects.select_related('customer'),
        pk=vehicle_id,
        customer=customer_account,
    )

    active_statuses = VehicleMaintenanceTask.active_statuses()
    today = timezone.localdate()

    upcoming_tasks = list(
        vehicle.maintenance_tasks.filter(status__in=active_statuses)
        .select_related('work_order')
        .order_by('due_date', 'priority', 'title', 'pk')
    )
    completed_tasks = list(
        vehicle.maintenance_tasks.filter(status=VehicleMaintenanceTask.STATUS_COMPLETED)
        .select_related('work_order')
        .order_by('-completed_date', '-updated_at', '-pk')[:10]
    )

    work_orders = list(
        vehicle.work_orders.select_related('invoice')
        .order_by('-scheduled_date', '-id')[:15]
    )

    job_history_qs = (
        vehicle.job_history.select_related('invoice', 'source_income_record')
        .order_by('-job_date', '-id')
    )
    job_history = list(job_history_qs[:25])
    job_totals = job_history_qs.aggregate(
        total_service_cost=Coalesce(Sum('service_cost'), Value(Decimal('0.00'))),
        total_tax_amount=Coalesce(Sum('tax_amount'), Value(Decimal('0.00'))),
        total_overall_cost=Coalesce(Sum('total_job_cost'), Value(Decimal('0.00'))),
        job_count=Count('id'),
    )

    related_invoices = list(
        vehicle.customer.invoices.filter(job_history_entries__vehicle=vehicle)
        .distinct()
        .order_by('-date', '-id')[:10]
    )

    overdue_count = sum(1 for task in upcoming_tasks if task.is_overdue)
    next_due = upcoming_tasks[0] if upcoming_tasks else None
    last_completed = completed_tasks[0] if completed_tasks else None

    summary = {
        'upcoming_count': len(upcoming_tasks),
        'overdue_count': overdue_count,
        'completed_count': len(completed_tasks),
        'next_due_task': next_due,
        'last_completed_task': last_completed,
    }

    status_choices = [
        (code, label)
        for code, label in VehicleMaintenanceTask.STATUS_CHOICES
        if code != VehicleMaintenanceTask.STATUS_COMPLETED
    ]

    context = {
        'vehicle': vehicle,
        'summary': summary,
        'upcoming_tasks': upcoming_tasks,
        'completed_tasks': completed_tasks,
        'work_orders': work_orders,
        'job_history': job_history,
        'job_totals': job_totals,
        'related_invoices': related_invoices,
        'maintenance_form': VehicleMaintenanceTaskForm(),
        'status_choices': status_choices,
        'today': today,
    }
    context.update(_customer_portal_layout_context(request, active_slug='vehicles'))
    return render(request, 'store/customer_vehicle_detail.html', context)


@customer_login_required
def customer_maintenance_list(request):
    """Display upcoming maintenance tasks for all vehicles of the logged-in customer."""

    customer_account = request.user.customer_portal
    active_statuses = VehicleMaintenanceTask.active_statuses()
    today = timezone.localdate()

    maintenance_tasks = list(
        VehicleMaintenanceTask.objects.filter(
            vehicle__customer=customer_account,
            status__in=active_statuses,
        )
        .select_related('vehicle')
        .order_by('due_date', 'priority', 'title', 'pk')
    )

    context = {
        'maintenance_tasks': maintenance_tasks,
        'today': today,
    }
    context.update(_customer_portal_layout_context(request, active_slug='maintenance'))
    return render(request, 'store/customer_maintenance_list.html', context)


@customer_login_required
@require_POST
def customer_vehicle_add_maintenance(request, vehicle_id):
    """Allow the customer to add their own maintenance reminder."""

    customer_account = request.user.customer_portal
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id, customer=customer_account)

    form = VehicleMaintenanceTaskForm(request.POST)
    if form.is_valid():
        task = form.save(commit=False)
        task.vehicle = vehicle
        task.user = vehicle.customer.user
        task.save()
        messages.success(
            request,
            f"Maintenance '{task.title}' added for {vehicle}.",
        )
    else:
        messages.error(request, 'Unable to add maintenance. Please review the highlighted fields.')

    return redirect('accounts:customer_vehicle_detail', vehicle_id=vehicle.id)


@customer_login_required
@require_POST
def customer_vehicle_update_status(request, task_id):
    """Update the status of a maintenance task for the logged-in customer."""

    customer_account = request.user.customer_portal
    task = get_object_or_404(
        VehicleMaintenanceTask,
        pk=task_id,
        vehicle__customer=customer_account,
    )

    new_status = request.POST.get('status')
    allowed_statuses = {
        code for code, _ in VehicleMaintenanceTask.STATUS_CHOICES
        if code != VehicleMaintenanceTask.STATUS_COMPLETED
    }

    if new_status not in allowed_statuses:
        messages.error(request, 'Please choose a valid status option.')
        return redirect('accounts:customer_vehicle_detail', vehicle_id=task.vehicle_id)

    task.status = new_status
    if new_status in {VehicleMaintenanceTask.STATUS_PLANNED, VehicleMaintenanceTask.STATUS_CANCELLED}:
        task.work_order = None
    task.save(update_fields=['status', 'work_order', 'updated_at'])

    messages.success(
        request,
        f"Updated maintenance '{task.title}' to {task.get_status_display()}.",
    )
    return redirect('accounts:customer_vehicle_detail', vehicle_id=task.vehicle_id)


@customer_login_required
@require_POST
def customer_vehicle_complete_maintenance(request, task_id):
    """Capture completion details for a maintenance task."""

    customer_account = request.user.customer_portal
    task = get_object_or_404(
        VehicleMaintenanceTask,
        pk=task_id,
        vehicle__customer=customer_account,
    )

    form = VehicleMaintenanceCompleteForm(
        request.POST,
        instance=task,
        vehicle=task.vehicle,
        user=task.vehicle.customer.user,
    )

    if form.is_valid():
        completed_task = form.save(commit=False)
        completed_task.status = VehicleMaintenanceTask.STATUS_COMPLETED
        if not completed_task.completed_date:
            completed_task.completed_date = timezone.localdate()
        completed_task.save()

        if completed_task.actual_mileage is not None:
            vehicle = completed_task.vehicle
            if vehicle.current_mileage is None or completed_task.actual_mileage > vehicle.current_mileage:
                vehicle.current_mileage = completed_task.actual_mileage
                vehicle.save(update_fields=['current_mileage'])

        messages.success(request, f"Marked '{completed_task.title}' as completed.")
    else:
        messages.error(request, 'Please review the completion details and try again.')

    return redirect('accounts:customer_vehicle_detail', vehicle_id=task.vehicle_id)


@customer_login_required
def customer_settlement_summary(request):
    """Provide a snapshot of paid and outstanding balances for the customer."""

    customer_account = request.user.customer_portal
    invoice_totals = customer_account.invoices.aggregate(
        total_invoiced=Coalesce(Sum('total_amount'), Decimal('0.00')),
    )
    payment_totals = Payment.objects.filter(invoice__customer=customer_account).aggregate(
        total_paid=Coalesce(Sum('amount'), Decimal('0.00')),
    )

    total_invoiced = invoice_totals['total_invoiced'] or Decimal('0.00')
    total_paid = payment_totals['total_paid'] or Decimal('0.00')
    credit_rows, total_credit = _build_customer_credit_rows(customer_account)
    outstanding_balance = total_invoiced - total_paid - total_credit

    outstanding_qs = (
        customer_account.invoices
        .select_related('user__profile')
        .prefetch_related('payments')
        .annotate(
            total_paid_amount=Coalesce(
                Sum('payments__amount'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )
    )
    outstanding_qs = _annotate_invoice_credit_totals(outstanding_qs).annotate(
        balance_due_amount=ExpressionWrapper(
            F('total_amount') - F('total_paid_amount') - F('credit_total'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    ).filter(balance_due_amount__gt=Decimal('0.00')).order_by('-date', '-id')
    paginator = Paginator(outstanding_qs, 100)
    page_number = request.GET.get('page')
    outstanding_page = paginator.get_page(page_number)
    outstanding_params = request.GET.copy()
    outstanding_params.pop('page', None)
    outstanding_query_string = outstanding_params.urlencode()
    recent_payments = Payment.objects.filter(invoice__customer=customer_account).order_by('-date', '-id')[:10]

    profile = getattr(customer_account.user, 'profile', None)
    contact_email = None
    if profile:
        contact_email = profile.interact_email or profile.company_email

    has_outstanding = outstanding_page.paginator.count > 0
    provider = getattr(profile, "payment_link_provider", PAYMENT_LINK_PROVIDER_STRIPE)
    if provider not in {
        PAYMENT_LINK_PROVIDER_STRIPE,
        PAYMENT_LINK_PROVIDER_CLOVER,
        PAYMENT_LINK_PROVIDER_NONE,
    }:
        provider = PAYMENT_LINK_PROVIDER_STRIPE

    has_online_links = False
    if has_outstanding:
        if provider == PAYMENT_LINK_PROVIDER_CLOVER:
            has_online_links = outstanding_qs.filter(
                Q(clover_payment_link__isnull=False) & ~Q(clover_payment_link="")
            ).exists()
        elif provider == PAYMENT_LINK_PROVIDER_NONE:
            has_online_links = False
        else:
            has_online_links = outstanding_qs.filter(
                (Q(stripe_payment_link__isnull=False) & ~Q(stripe_payment_link="")) |
                (Q(stripe_subscription_link__isnull=False) & ~Q(stripe_subscription_link=""))
            ).exists()

    show_action_column = has_online_links
    show_interac_notice = has_outstanding and not has_online_links

    context = {
        'invoice_summary': {
            'total_invoiced': total_invoiced,
            'total_paid': total_paid,
            'total_credit': total_credit,
            'outstanding_balance': outstanding_balance,
        },
        'contact_email': contact_email,
        'show_action_column': show_action_column,
        'show_interac_notice': show_interac_notice,
        'outstanding_invoices': outstanding_page,
        'outstanding_page': outstanding_page,
        'outstanding_query_string': outstanding_query_string,
        'recent_payments': recent_payments,
        'credit_rows': credit_rows,
    }
    context.update(_customer_portal_layout_context(request, active_slug='settlements'))
    return render(request, 'store/customer_settlements.html', context)
