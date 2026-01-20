"""Helpers for styling exported Excel templates."""

from __future__ import annotations

from typing import Iterable, Mapping, Optional, Union

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_INSTRUCTION_FILL = PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid")
_INSTRUCTION_FONT = Font(color="3F3F3F", bold=False)
_INSTRUCTION_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALT_ROW_FILL = PatternFill(start_color="F3F6FC", end_color="F3F6FC", fill_type="solid")
_DATA_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def apply_template_styling(
    worksheet: Worksheet,
    headers: Iterable[str],
    header_row_index: int,
    instruction_row_index: Optional[int] = None,
    column_width_overrides: Optional[Mapping[Union[str, int], float]] = None,
    minimum_width: float = 12.0,
) -> None:
    """Apply consistent styling to exported Excel templates."""

    column_width_overrides = column_width_overrides or {}
    headers = list(headers)

    if instruction_row_index:
        _style_instruction_row(worksheet, instruction_row_index)

    _style_header_row(worksheet, header_row_index)
    _style_data_rows(worksheet, header_row_index)
    _auto_size_columns(
        worksheet,
        headers=headers,
        header_row_index=header_row_index,
        instruction_row_index=instruction_row_index,
        column_width_overrides=column_width_overrides,
        minimum_width=minimum_width,
    )


def _style_instruction_row(worksheet: Worksheet, row_index: int) -> None:
    cells = worksheet[row_index]
    worksheet.row_dimensions[row_index].height = 48
    for cell in cells:
        if isinstance(cell, MergedCell):
            continue
        cell.fill = _INSTRUCTION_FILL
        cell.font = _INSTRUCTION_FONT
        cell.alignment = _INSTRUCTION_ALIGNMENT
        cell.border = _BORDER


def _style_header_row(worksheet: Worksheet, row_index: int) -> None:
    cells = worksheet[row_index]
    worksheet.row_dimensions[row_index].height = 26
    for cell in cells:
        if isinstance(cell, MergedCell):
            continue
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _BORDER


def _style_data_rows(worksheet: Worksheet, header_row_index: int) -> None:
    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        row = worksheet[row_index]
        if all((cell.value in (None, "")) for cell in row):
            continue
        max_line_count = 1
        use_alt_fill = (row_index - header_row_index) % 2 == 1
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, ""):
                text = str(cell.value)
                # Normalize newlines so splitlines counts every explicit line break
                normalized_text = text.replace("\r\n", "\n").replace("\r", "\n")
                line_count = max(1, len(normalized_text.split("\n")))
                if line_count > max_line_count:
                    max_line_count = line_count
            cell.alignment = _DATA_ALIGNMENT
            cell.border = _BORDER
            if use_alt_fill:
                cell.fill = _ALT_ROW_FILL
        base_height = 22
        worksheet.row_dimensions[row_index].height = base_height * max_line_count


def _auto_size_columns(
    worksheet: Worksheet,
    headers: Iterable[str],
    header_row_index: int,
    instruction_row_index: Optional[int],
    column_width_overrides: Mapping[Union[str, int], float],
    minimum_width: float,
) -> None:
    headers = list(headers)
    for index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(index)
        explicit_width = column_width_overrides.get(header)
        if explicit_width is None:
            explicit_width = column_width_overrides.get(index)
        if explicit_width is not None:
            worksheet.column_dimensions[column_letter].width = float(explicit_width)
            continue

        max_length = len(str(header)) + 2
        for row_index in range(header_row_index + 1, worksheet.max_row + 1):
            if instruction_row_index and row_index == instruction_row_index:
                continue
            value = worksheet.cell(row=row_index, column=index).value
            if value in (None, ""):
                continue
            value_length = len(str(value))
            if value_length > max_length:
                max_length = value_length
        width = max(minimum_width, min(max_length + 2, 60))
        worksheet.column_dimensions[column_letter].width = float(width)
