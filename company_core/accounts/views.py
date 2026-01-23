import json
import itertools
import re
from io import BytesIO
from typing import Set
from collections import defaultdict
from .decorators import subscription_required, activation_required, accountant_login_required
from .utils import (
    verify_stripe_account,
    format_currency,
    get_default_store_owner,
    get_storefront_owner,
    get_primary_business_user,
    get_business_user,
    get_business_user_ids,
    get_connected_business_group,
    get_customer_user_ids,
    get_product_user_ids,
    get_stock_owner,
    is_parts_store_business,
    resolve_storefront_price_flags,
    resolve_storefront_category_flags,
    notify_mechanic_assignment,
    sync_workorder_assignments,
    resolve_company_logo_url,
    apply_stock_fields,
    upsert_product_stock,
)
from .service_reminders import (
    create_maintenance_tasks_from_services,
    parse_service_entries,
)
from .view_workorder import (
    generate_pm_inspection_pdf,
    DEFAULT_PM_BUSINESS_INFO,
    _serialize_vehicle_history,
    _serialize_vehicle_maintenance,
)
from .view_invoices import send_grouped_invoice_email
from .paid_invoice_views import send_paid_invoice_email
from .invoice_activity import log_invoice_activity, build_email_open_tracking_url
from .pdf_utils import apply_branding_defaults, render_template_to_pdf
from django.http import FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
from .excel_formatting import apply_template_styling
import shutil
import pytz
import os
import base64
import datetime
from django.views import View
from django.contrib.auth import get_user_model
from django.core.serializers.json import DjangoJSONEncoder
from django.utils.safestring import mark_safe
from django.db import models, transaction, IntegrityError
from django.urls import reverse_lazy, reverse
from .forms import (
    PublicContactForm,
    PortalContactForm,
    BusinessBookingSettingsForm,
    BusinessHolidayForm,
    CustomerSignupForm,
    StaffSignupForm,
    DisplayPreferencesForm,
    ConnectedBusinessGroupForm,
    FlyerEmailForm,
)
from .models import (
    PublicBooking,
    EmergencyRequest,
    PublicContactMessage,
    BusinessBookingSettings,
    BusinessHoliday,
    Service,
    ServiceJobName,
    Product,
    Category,
    CategoryGroup,
    CategoryAttribute,
    ProductBrand,
    InventoryTransaction,
    MechExpense,
    MechExpenseItem,
)
from django.conf import settings
from django.core.mail import send_mail, EmailMultiAlternatives
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.views import LoginView
from django.contrib.auth import login
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import EmailMessage
from django.core.validators import validate_email
from .utils import build_cc_list
from django.contrib.auth.forms import SetPasswordForm
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.utils.encoding import force_bytes, force_str
from django.db.models.functions import (
    TruncMonth,
    TruncDate,
    Lower,
    Coalesce,
    Reverse,
    StrIndex,
    Substr,
    Length,
    Cast,
    NullIf,
)
from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden, HttpResponseRedirect, Http404
from django.utils.http import urlsafe_base64_encode
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from urllib.parse import urlencode, unquote
from collections.abc import Mapping
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, FloatField, Value, Q, Subquery, OuterRef, Case, When, Max, CharField, IntegerField, DateTimeField
from django.utils.dateparse import parse_datetime
from .forms import (
    VehicleForm,
    VehicleMaintenanceTaskForm,
    VehicleMaintenanceCompleteForm,
    VehicleQuickWorkOrderForm,
    EstimateRecordFormSet,
    GroupedEstimateForm,
    EstimateRecordForm,
    CustomerForm,
    AccountantPortalForm,
    NoteForm,
    PaymentForm,
    InvoiceCustomizationForm,
    InvoiceSequenceForm,
    QRCodeStyleForm,
    ProfileForm,
    SignUpForm,
    ExpenseForm,
    IncomeForm,
    InvoiceForm,
    GroupedInvoiceForm,
    IncomeRecord2FormSet,
    MechExpenseForm,
    MechExpenseItemFormSet,
    PROVINCE_CHOICES,
    IncomeRecord2Form,
    FleetVehicleForm,
    MaintenanceRecordForm,
    CategoryForm,
    ServiceForm,
    BankingIntegrationSettingsForm,
    QuickBooksSettingsForm,
    QuickProductCreateForm,
    SupplierForm,
    SupplierPortalForm,
    SupplierCreditForm,
    SupplierCreditItemFormSet,
    CustomerCreditForm,
    CustomerCreditItemFormSet,
    SupplierChequeForm,
)
from .models import (
    ReminderLog,
    Vehicle,
    VehicleMaintenanceTask,
    JobHistory,
    ensure_decimal,
    GroupedEstimate,
    EstimateRecord,
    Driver,
    Product,
    InventoryTransaction,
    Customer,
    TERM_CHOICES,
    TransportCustomer,
    TransportTrip,
    WorkOrder,
    WorkOrderRecord,
    WorkOrderAssignment,
    Mechanic,
    Payment,
    Note,
    UserStripeAccount,
    ExpenseRecord,
    IncomeRecord,
    InvoiceDetail,
    IncomeRecord2,
    GroupedInvoice,
    Profile,
    PendingInvoice,
    PaidInvoice,
    MechExpense,
    MechExpensePayment,
    MechExpenseItem,
    SupplierCredit,
    SupplierCreditItem,
    CustomerCredit,
    CustomerCreditItem,
    SupplierCheque,
    SupplierChequeLine,
    BusinessBankAccount,
    PROVINCE_TAX_RATES,
    get_tax_component_rates,
    FleetVehicle,
    MaintenanceRecord,
    Category,
    CategoryGroup,
    CategoryAttribute,
    CategoryAttributeOption,
    Supplier,
    ProductBrand,
    ProductModel,
    ProductVin,
    InventoryLocation,
    ProductAttributeValue,
    ProductAlternateSku,
    TaxExemptionReason,
    QuickBooksSettings,
    ActivityLog,
    InvoiceActivity,
    PAYMENT_LINK_PROVIDER_STRIPE,
    PAYMENT_LINK_PROVIDER_CLOVER,
    CloverConnection,
    QuickBooksConnection,
    QuickBooksImportMap,
    BankingIntegrationSettings,
    BankConnection,
    BankTransaction,
    ConnectedBusinessGroup,
)
from .quickbooks_desktop_service import QuickBooksDesktopService
from .quickbooks_service import QuickBooksService, QuickBooksIntegrationError
from django.contrib.auth.mixins import LoginRequiredMixin
import logging
from django.core.mail import send_mail
from django.core import signing
import stripe
from django.contrib.auth.views import PasswordChangeView
from django.utils.deprecation import MiddlewareMixin
from django.contrib.auth import logout
from django.contrib.auth import authenticate
from django.utils.decorators import method_decorator
from django.utils.html import escape, strip_tags
from django.utils import timezone
from django.utils.timezone import now
from .signup_process import (
    signup, payment, signup_thankyou,
    stripe_webhook, activate, activation_complete,
    activate_features, resend_activation_email,
    activation_invalid, activation_require
)
import secrets
from django.utils.text import slugify
from datetime import datetime
from datetime import timedelta
from django.utils.timezone import make_aware
import xlwt
from xhtml2pdf import pisa
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False
    HTML = None
    CSS = None
from django import forms
from .utils import generate_invoice_pdf, calculate_next_occurrence
import datetime
from decimal import Decimal
from django.forms.models import model_to_dict
from django.forms import inlineformset_factory
from django.core.exceptions import ValidationError
from dateutil.relativedelta import relativedelta
from .utils import get_overdue_total_balance


def _resolve_primary_business_user():
    return get_primary_business_user() or get_default_store_owner()


def _get_business_display_name(user):
    default_name = getattr(settings, "DEFAULT_BUSINESS_NAME", "Truck Zone") or "Truck Zone"
    if not user:
        return default_name
    profile = getattr(user, "profile", None)
    if profile and getattr(profile, "company_name", None):
        return profile.company_name
    return default_name


def _log_admin_activity(request, *, action, object_type, description, object_id=None, metadata=None):
    """Record an activity log entry for the acting staff member."""

    actor = getattr(request, "actual_user", None)
    if not actor or not actor.is_authenticated:
        actor = getattr(request, "user", None)

    if not actor or not actor.is_authenticated:
        return

    business_user = get_business_user(getattr(request, "user", None))
    if not business_user:
        return

    ActivityLog.objects.create(
        business=business_user,
        actor=actor,
        action=action,
        object_type=object_type,
        object_id=str(object_id or ""),
        description=description,
        metadata=metadata or {},
    )


def to_decimal(value):
    if value is None:
        return Decimal('0.00')
    return Decimal(value)

# views.py
def get_invoice_context(invoice, request):
    """
    Prepares the context dictionary for invoice PDF generation.

    :param invoice: The GroupedInvoice instance.
    :param request: The HTTP request object.
    :return: Dictionary containing context data.
    """
    profile = invoice.user.profile
    records = list(invoice.income_records.all())

    subtotal = sum((record.amount for record in records), Decimal('0.00'))
    tax = sum((record.tax_collected for record in records), Decimal('0.00'))
    total_amount = subtotal + tax

    parts_subtotal = sum((record.amount for record in records if record.product_id), Decimal('0.00'))
    labor_subtotal = sum((record.amount for record in records if not record.product_id), Decimal('0.00'))
    show_parts_breakdown = parts_subtotal > Decimal('0.00')

    # Generate absolute URL for company logo
    company_logo_url = resolve_company_logo_url(profile, request=request)

    payment_link = invoice.payment_link
    payment_provider = invoice.get_payment_link_provider()
    invoice_note = (invoice.notes or '').strip()
    if invoice_note:
        note = invoice_note
    elif profile.show_note:
        note = (
            generate_dynamic_invoice_note(invoice)
            if profile.use_dynamic_note else
            (profile.note or '')
        )
    else:
        note = ''

    return {
        'object': invoice,  # Ensure your template uses 'object' to refer to the invoice
        'profile': profile,
        'subtotal': subtotal,
        'tax': tax,
        'total_amount': total_amount,
        'company_logo_url': company_logo_url,  # Add the absolute URL to context
        'payment_link': payment_link,
        'payment_provider': payment_provider,
        'parts_subtotal': parts_subtotal,
        'labor_subtotal': labor_subtotal,
        'show_parts_breakdown': show_parts_breakdown,
        'note': note,
        'invoice_open_tracking_url': build_email_open_tracking_url(
            invoice,
            request=request,
        ),
    }

def generate_invoice_pdf(context, request):
    """
    Generates a PDF for the given invoice context using WeasyPrint.

    :param context: Dictionary containing context data for the invoice.
    :param request: The HTTP request object.
    :return: Bytes of the generated PDF.
    """
    # Render the HTML template with context
    context_for_pdf = context.copy()
    profile = context.get('profile')
    if profile is not None:
        context_for_pdf['company_logo_url'] = resolve_company_logo_url(
            profile,
            for_pdf=True,
        )

    html_string = render_to_string('invoices/invoice.html', context_for_pdf, request=request)

    if not WEASYPRINT_AVAILABLE:
        raise ImportError("WeasyPrint is not available. PDF generation is disabled. Please install GTK+ libraries for Windows.")

    # Create a WeasyPrint HTML object
    html = HTML(string=html_string, base_url=request.build_absolute_uri())

    # Generate PDF
    pdf = html.write_pdf()

    return pdf

def generate_workorder_pdf(workorder, request):
    """Generate PDF bytes for a completed work order."""
    profile = workorder.user.profile
    subtotal = workorder.records.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    if not isinstance(subtotal, Decimal):
        subtotal = Decimal(str(subtotal))

    grand_total = subtotal
    tax_total = Decimal('0.00')

    invoice = getattr(workorder, 'invoice', None)
    if invoice and getattr(invoice, 'total_amount', None) is not None:
        invoice_total = invoice.total_amount
        if not isinstance(invoice_total, Decimal):
            try:
                invoice_total = Decimal(str(invoice_total))
            except (InvalidOperation, TypeError, ValueError):
                invoice_total = subtotal

        grand_total = invoice_total.quantize(Decimal('0.01'))
        difference = (grand_total - subtotal).quantize(Decimal('0.01'))
        if difference > Decimal('0.00'):
            tax_total = difference
        else:
            grand_total = subtotal
            tax_total = Decimal('0.00')

    logo_url = resolve_company_logo_url(profile, for_pdf=True)
    context = {
        'workorder': workorder,
        'profile': profile,
        'subtotal': subtotal,
        'tax_total': tax_total,
        'grand_total': grand_total,
        'company_logo_url': logo_url,
    }
    context = apply_branding_defaults(context)
    html_string = render_to_string('workorders/workorder_pdf.html', context)
    if not WEASYPRINT_AVAILABLE:
        raise ImportError("WeasyPrint is not available. PDF generation is disabled. Please install GTK+ libraries for Windows.")
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    return pdf

@login_required
def print_invoice(request, pk):
    """
    Renders the invoice in an HTML template and triggers the print dialog.

    :param request: The HTTP request object.
    :param pk: Primary key of the GroupedInvoice.
    :return: HTTP response rendering the invoice template.
    """
    invoice = get_object_or_404(GroupedInvoice, pk=pk, user=request.user)
    context = get_invoice_context(invoice, request)  # Pass request to context

    # Add a flag to indicate that this is the print view
    context['print_view'] = True

    return render(request, 'invoices/invoice.html', context)

@login_required
def download_invoice(request, pk):
    """
    Generates and serves the invoice PDF as an attachment for download.

    :param request: The HTTP request object.
    :param pk: Primary key of the GroupedInvoice.
    :return: HTTP response with PDF content.
    """
    invoice = get_object_or_404(GroupedInvoice, pk=pk, user=request.user)
    context = get_invoice_context(invoice, request)  # Pass request to context
    pdf = generate_invoice_pdf(context, request)

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.invoice_number}.pdf"'
    return response

@login_required
def email_invoice(request, pk):
    """
    Generates the invoice PDF and sends it via email to the customer and notifies the user.

    :param request: The HTTP request object.
    :param pk: Primary key of the GroupedInvoice.
    :return: HTTP response indicating success or failure.
    """
    invoice = get_object_or_404(GroupedInvoice, pk=pk, user=request.user)
    include_workorder = request.GET.get('include_workorder') == '1'
    context = get_invoice_context(invoice, request)  # Pass request to context

    # Generate the PDF invoice
    pdf = generate_invoice_pdf(context, request)

    # Company details
    company_email = invoice.user.profile.company_email
    company_name = invoice.user.profile.company_name
    customer_email = invoice.bill_to_email
    customer_cc_emails = invoice.customer.get_cc_emails() if invoice.customer else []

    # Email subjects
    customer_email_subject = f"Invoice from {company_name} - Invoice #{invoice.invoice_number}"
    user_email_subject = f"Invoice #{invoice.invoice_number} sent to {invoice.bill_to}"

    # Prepare email contexts
    customer_context = context.copy()
    customer_context['recipient_name'] = invoice.bill_to

    user_context = {
        'invoice': invoice,
        'customer': invoice.bill_to,
        'customer_email': invoice.bill_to_email,
        'company_name': company_name,
        'profile': invoice.user.profile,
        'total_amount': context['total_amount'],
    }

    # Render email bodies using templates
    customer_context = apply_branding_defaults(customer_context)
    customer_email_body = render_to_string('emails/invoice_email_customer.html', customer_context)
    user_context = apply_branding_defaults(user_context)
    user_email_body = render_to_string('emails/invoice_email_user.html', user_context)

    # Prepare the customer email
    customer_cc = build_cc_list(
        company_email,
        *customer_cc_emails,
        exclude=[customer_email],
    )
    customer_email_message = EmailMessage(
        subject=customer_email_subject,
        body=customer_email_body,
        from_email=company_email,
        to=[customer_email] if customer_email else [],
        cc=customer_cc or None,
    )
    customer_email_message.content_subtype = 'html'  # Set email content to HTML
    customer_email_message.attach(f"Invoice_{invoice.invoice_number}.pdf", pdf, 'application/pdf')
    if include_workorder and invoice.work_order:
        wo_pdf = generate_workorder_pdf(invoice.work_order, request)
        customer_email_message.attach(
            f"WorkOrder_{invoice.work_order.id}.pdf", wo_pdf, 'application/pdf'
        )
        pm_pdf = None
        pm_inspection = getattr(invoice.work_order, 'pm_inspection', None)
        if pm_inspection:
            pm_pdf = generate_pm_inspection_pdf(pm_inspection, request)
            customer_email_message.attach(
                f"WorkOrder_{invoice.work_order.id}_PM.pdf", pm_pdf, 'application/pdf'
            )

    # Prepare the user email
    user_email_message = EmailMessage(
        subject=user_email_subject,
        body=user_email_body,
        from_email=company_email,
        to=[request.user.email],
    )
    user_email_message.content_subtype = 'html'  # Set email content to HTML
    if include_workorder and invoice.work_order:
        wo_pdf = wo_pdf if 'wo_pdf' in locals() else generate_workorder_pdf(invoice.work_order, request)
        user_email_message.attach(
            f"WorkOrder_{invoice.work_order.id}.pdf", wo_pdf, 'application/pdf'
        )
        if 'pm_pdf' in locals():
            pm_pdf_bytes = pm_pdf
        else:
            pm_inspection = getattr(invoice.work_order, 'pm_inspection', None)
            pm_pdf_bytes = generate_pm_inspection_pdf(pm_inspection, request) if pm_inspection else None
        if pm_pdf_bytes:
            user_email_message.attach(
                f"WorkOrder_{invoice.work_order.id}_PM.pdf", pm_pdf_bytes, 'application/pdf'
            )

    try:
        # Send both emails
        customer_email_message.send()
        user_email_message.send()
        log_invoice_activity(
            invoice,
            event_type=InvoiceActivity.EVENT_EMAIL_SENT,
            request=request,
        )
        return HttpResponse('Emails sent successfully', status=200)
    except Exception as e:
        # Log the error as needed
        return HttpResponse(f'Failed to send emails: {str(e)}', status=500)


stripe.api_key = settings.STRIPE_SECRET_KEY

logger = logging.getLogger(__name__)
User = get_user_model()


def _get_or_create_category_for_user(user, category_name):
    """Return a category for the provided user, creating it when necessary."""
    name = (category_name or 'Parts').strip() or 'Parts'
    category = Category.objects.filter(user=user, name__iexact=name).first()
    if category:
        return category
    return Category.objects.create(user=user, name=name)


def _get_or_create_supplier_for_user(user, supplier_name):
    """Return a supplier for the provided user, creating it when necessary."""
    normalized_name = (supplier_name or "").strip()
    if not normalized_name:
        return None
    supplier = Supplier.objects.filter(user=user, name__iexact=normalized_name).first()
    if supplier:
        return supplier
    return Supplier.objects.create(user=user, name=normalized_name)


def _get_vendor_options_for_user(user):
    """Return vendor names drawn from expenses and inventory suppliers."""

    if not user or not getattr(user, "is_authenticated", False):
        return []

    vendor_names = set()

    expense_vendors = (
        MechExpense.objects.filter(user=user)
        .values_list("vendor", flat=True)
        .distinct()
    )
    supplier_names = Supplier.objects.filter(user=user).values_list("name", flat=True)

    for raw_name in list(expense_vendors) + list(supplier_names):
        normalized = (raw_name or "").strip()
        if normalized:
            vendor_names.add(normalized)

    return sorted(vendor_names, key=lambda name: name.lower())


def _ensure_inventory_product_for_item(
    *,
    user,
    category_name,
    existing_product,
    create_inventory_product,
    part_no,
    description,
    price_decimal,
    has_price,
    supplier_name,
):
    """Resolve the inventory product to use for an expense line."""
    supplier = _get_or_create_supplier_for_user(user, supplier_name)

    if not create_inventory_product:
        if existing_product and supplier and existing_product.supplier_id != supplier.id:
            existing_product.supplier = supplier
            existing_product.save(update_fields=["supplier"])
        return existing_product

    normalized_part_no = (part_no or '').strip()
    normalized_description = (description or '').strip()
    product = existing_product

    category = _get_or_create_category_for_user(user, category_name)

    if not product and normalized_part_no:
        product = Product.objects.filter(
            user=user,
            sku__iexact=normalized_part_no,
        ).first()

    if not product and normalized_description:
        product = Product.objects.filter(
            user=user,
            name__iexact=normalized_description,
        ).first()

    if not product:
        default_name = normalized_description or normalized_part_no or 'New Product'
        description_value = normalized_description or default_name
        new_product = Product.objects.create(
            user=user,
            sku=normalized_part_no or None,
            name=default_name,
            description=description_value,
            category=category,
            cost_price=price_decimal if has_price else Decimal('0.00'),
            sale_price=price_decimal if has_price else Decimal('0.00'),
            supplier=supplier,
        )

        return new_product

    if normalized_part_no:
        conflict = Product.objects.filter(
            user=user,
            sku__iexact=normalized_part_no,
        ).exclude(pk=product.pk).first()
        if conflict:
            product = conflict

    fields_to_update = []

    if normalized_part_no and product.sku != normalized_part_no:
        product.sku = normalized_part_no
        fields_to_update.append('sku')

    if normalized_description and product.description != normalized_description:
        product.description = normalized_description
        fields_to_update.append('description')

    if has_price and (
        product.cost_price != price_decimal
        or product.sale_price != price_decimal
    ):
        product.cost_price = price_decimal
        product.sale_price = price_decimal
        fields_to_update.extend(['cost_price', 'sale_price'])

    if not product.category_id:
        product.category = category
        fields_to_update.append('category')

    if supplier and product.supplier_id != supplier.id:
        product.supplier = supplier
        fields_to_update.append('supplier')

    if fields_to_update:
        product.save(update_fields=list(dict.fromkeys(fields_to_update)))

    return product


def verify_stripe_account(user):
    try:
        UserStripeAccount.objects.get(user=user)
        return True
    except UserStripeAccount.DoesNotExist:
        return False

@login_required
def profile_completion_view(request):
    profile = request.user.profile

    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('accounts:home')
    else:
        form = ProfileForm(instance=profile)

    completion_percentage = profile.profile_completion()
    context = {
        'form': form,
        'completion_percentage': completion_percentage,
    }
    return render(request, 'profile/profile_completion.html', context)

def login_redirect(request):
    profile = request.user.profile
    if profile.profile_completion() < 100:
        return redirect('accounts:profile_completion')
    else:
        return redirect('accounts:home')

@login_required
def payment_cancel(request):
    # You may pass relevant context if needed, like invoice details
    invoice_number = request.GET.get('invoice_number', 'Unknown')
    return render(request, 'payment/payment_cancel.html', {'invoice_number': invoice_number})


@login_required
def payment_success(request):
    # You may pass relevant context if needed, like invoice details
    invoice_number = request.GET.get('invoice_number', 'Unknown')
    return render(request, 'payment/payment_success.html', {'invoice_number': invoice_number})

@login_required
def payment_already_made(request):
    # You may pass relevant context if needed, like invoice details
    invoice_number = request.GET.get('invoice_number', 'Unknown')
    return render(request, 'payment/payment_already_made.html', {'invoice_number': invoice_number})

@login_required
def error_page(request):
    return render(request, 'payment/error_page.html')

"""
Public transport forms/views
"""
def public_booking(request):
    """Redirect legacy service booking to the parts contact flow."""
    return redirect('accounts:public_contact')


def _coerce_time_config(value: object, fallback: datetime.time) -> datetime.time:
    """Convert a settings value into a ``datetime.time`` object."""
    if isinstance(value, datetime.time):
        return value
    if isinstance(value, datetime.datetime):
        return value.time()
    if isinstance(value, str):
        try:
            return datetime.datetime.strptime(value.strip(), '%H:%M').time()
        except ValueError:
            pass
    return fallback


def _get_booking_settings_instance() -> BusinessBookingSettings:
    """Return the singleton booking settings, tolerating missing tables before migrations."""

    try:
        return BusinessBookingSettings.get_solo()
    except Exception:
        return BusinessBookingSettings(
            start_time=datetime.time(hour=9, minute=0),
            end_time=datetime.time(hour=17, minute=0),
            slot_interval_minutes=60,
        )


def _is_holiday(target_date: datetime.date) -> tuple[bool, str | None]:
    """Return a tuple indicating whether ``target_date`` is a holiday and the public reason."""

    try:
        holiday = BusinessHoliday.objects.filter(date=target_date).first()
    except Exception:
        holiday = None

    if holiday:
        reason = holiday.reason.strip() if holiday.reason else None
        return True, reason

    holidays = getattr(settings, 'BOOKING_BUSINESS_HOLIDAYS', []) or []
    for holiday in holidays:
        reason = None
        date_value = None
        if isinstance(holiday, dict):
            raw_date = holiday.get('date') or holiday.get('day')
            reason = holiday.get('reason') or holiday.get('label')
            if isinstance(raw_date, datetime.date):
                date_value = raw_date
            elif isinstance(raw_date, datetime.datetime):
                date_value = raw_date.date()
            elif isinstance(raw_date, str):
                try:
                    date_value = datetime.datetime.strptime(raw_date.strip(), '%Y-%m-%d').date()
                except ValueError:
                    date_value = None
        elif isinstance(holiday, datetime.date):
            date_value = holiday
        elif isinstance(holiday, datetime.datetime):
            date_value = holiday.date()
        elif isinstance(holiday, str):
            try:
                date_value = datetime.datetime.strptime(holiday.strip(), '%Y-%m-%d').date()
            except ValueError:
                date_value = None

        if date_value and date_value == target_date:
            cleaned_reason = reason.strip() if isinstance(reason, str) and reason.strip() else None
            return True, cleaned_reason

    return False, None


def booking_slots(request):
    """Return available time slots for a given date and service.

    Slots are generated from the business start/end configuration and remain
    available unless the day is configured as a holiday. This allows businesses
    with multiple mechanics to offer every slot within their operating hours.
    """
    date_str = request.GET.get('date')  # YYYY-MM-DD
    service = request.GET.get('service')
    if not date_str:
        return JsonResponse({ 'slots': [] })
    try:
        target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({ 'slots': [] })

    is_holiday, holiday_reason = _is_holiday(target_date)
    if is_holiday:
        return JsonResponse({
            'slots': [],
            'service': service,
            'is_holiday': True,
            'holiday_reason': holiday_reason,
        })

    default_start = datetime.time(hour=9, minute=0)
    default_end = datetime.time(hour=17, minute=0)

    settings_instance = _get_booking_settings_instance()

    start_config = settings_instance.start_time or getattr(settings, 'BOOKING_BUSINESS_START', default_start)
    end_config = settings_instance.end_time or getattr(settings, 'BOOKING_BUSINESS_END', default_end)
    start_time = _coerce_time_config(start_config, default_start)
    end_time = _coerce_time_config(end_config, default_end)

    # Ensure the end is after the start; fall back to defaults otherwise.
    start_dt = datetime.datetime.combine(target_date, start_time)
    end_dt = datetime.datetime.combine(target_date, end_time)
    default_end_dt = datetime.datetime.combine(target_date, default_end)
    default_start_dt = datetime.datetime.combine(target_date, default_start)
    if end_dt <= start_dt:
        start_time = default_start
        end_time = default_end
        start_dt = default_start_dt
        end_dt = default_end_dt

    try:
        step_minutes = int(settings_instance.slot_interval_minutes)
        if step_minutes <= 0:
            raise ValueError
    except (TypeError, ValueError):
        try:
            step_minutes = int(getattr(settings, 'BOOKING_SLOT_INTERVAL_MINUTES', 60))
            if step_minutes <= 0:
                raise ValueError
        except (TypeError, ValueError):
            step_minutes = 60

    all_slots = []
    slot_delta = datetime.timedelta(minutes=step_minutes)
    current_dt = start_dt
    while current_dt < end_dt:
        all_slots.append(current_dt.time().strftime('%H:%M'))
        current_dt += slot_delta

    return JsonResponse({
        'slots': all_slots,
        'service': service,
        'is_holiday': False,
        'holiday_reason': None,
    })


def public_emergency(request):
    """Redirect legacy emergency requests to the parts contact flow."""
    return redirect('accounts:public_contact')


def public_contact_form(request):
    if request.method == 'POST':
        form = PublicContactForm(request.POST)
        if form.is_valid():
            message_obj = form.save(commit=False)
            message_obj.source = PublicContactMessage.SOURCE_PUBLIC
            if not message_obj.subject:
                message_obj.subject = "Parts inquiry"
            message_obj.save()
            return render(request, 'public/contact_success.html', { 'message_obj': message_obj })
    else:
        form = PublicContactForm()
    return render(request, 'public/contact_form.html', { 'form': form })


def pm_inspection_download(request):
    """Redirect legacy inspection downloads to the parts contact flow."""
    return redirect('accounts:public_contact')
@login_required
def check_payment_status(request, invoice_number):
    grouped_invoice = get_object_or_404(GroupedInvoice, invoice_number=invoice_number)

    link = grouped_invoice.payment_link
    if not link:
        return redirect('accounts:error_page')  # Redirect to an error page if no checkout link exists

    provider = grouped_invoice.get_payment_link_provider()
    if provider == PAYMENT_LINK_PROVIDER_STRIPE:
        stripe.api_key = settings.STRIPE_SECRET_KEY
        session = stripe.checkout.Session.retrieve(link)
        payment_intent_id = session.payment_intent
        payment_intent = stripe.PaymentIntent.retrieve(payment_intent_id)

        if payment_intent.status == 'succeeded':
            return render(request, 'payment/payment_already_made.html', {'invoice': grouped_invoice})
        return redirect(link)

    if grouped_invoice.payment_status == 'Paid':
        return render(request, 'payment/payment_already_made.html', {'invoice': grouped_invoice})
    return redirect(link)

@login_required
def account_setup_complete(request):
    return render(request, 'registration/account_setup_complete.html')

class ReauthForm(forms.Form):
    password = forms.CharField(widget=forms.PasswordInput)


@login_required
def reauth(request):
    if request.method == 'POST':
        form = ReauthForm(request.POST)
        if form.is_valid():
            user = authenticate(request, username=request.user.username, password=form.cleaned_data['password'])
            if user is not None:
                login(request, user)  # Re-log the user in after authentication
                messages.success(request, 'You have been successfully reauthenticated!')
                return redirect('accounts:home')  # Redirect to the page they were trying to access
            else:
                messages.error(request, 'Your password was incorrect. Please try again.')
    else:
        form = ReauthForm()

    return render(request, 'registration/reauth.html', {'form': form})

@login_required
def connect_stripe_account(request):
    # Create a new Stripe account if it doesn't exist
    try:
        user_stripe_account = UserStripeAccount.objects.get(user=request.user)
    except UserStripeAccount.DoesNotExist:
        account = stripe.Account.create(
            type='express',
            country='CA',  # Correct country code for Canada
            email=request.user.email,
            business_type='individual',
        )
        user_stripe_account = UserStripeAccount.objects.create(
            user=request.user,
            stripe_account_id=account.id
        )

    # Create an account link for the user to complete the setup
    account_link = stripe.AccountLink.create(
        account=user_stripe_account.stripe_account_id,
        refresh_url='https://www.smart-invoices.com/reauth/',
        return_url='https://www.smart-invoices.com/account-setup-complete/',
        type='account_onboarding',
    )

    return redirect(account_link.url)


@login_required
def disconnect_stripe_account(request):
    account = UserStripeAccount.objects.filter(user=request.user).first()
    if account:
        account.stripe_account_id = None
        account.is_verified = False
        account.save(update_fields=['stripe_account_id', 'is_verified'])
    messages.success(request, 'Stripe disconnected.')
    return redirect('accounts:account_settings')

@login_required
def customize_invoice(request):
    profile = request.user.profile

    if request.method == 'POST':
        if 'reset' in request.POST:
            # Reset to default values
            profile.invoice_header_color = '#007bff'  # Default color
            profile.invoice_font_size = 16  # Default font size
            profile.show_logo = True
            profile.show_address = True
            profile.term = 'net_30'  # Default to Net 30
            profile.show_note = True  # Default to show note
            profile.note = 'Please ensure payment is made within the specified period to avoid any potential overdue charges.'  # Default note
            profile.invoice_sequence_next = None
            profile.payment_link_provider = PAYMENT_LINK_PROVIDER_STRIPE
            profile.save()
            return redirect('accounts:home')  # Redirect to refresh the page

        form = InvoiceCustomizationForm(request.POST)
        if form.is_valid():
            invoice_sequence_next = form.cleaned_data.get('invoice_sequence_next')
            if invoice_sequence_next:
                candidate = GroupedInvoice._format_invoice_number(request.user, invoice_sequence_next)
                if GroupedInvoice.objects.filter(user=request.user, invoice_number=candidate).exists():
                    form.add_error(
                        'invoice_sequence_next',
                        'That invoice number is already in use. Choose a higher number.',
                    )
                    context = {
                        'form': form,
                        'profile': profile,
                    }
                    return render(request, 'accounts/customize_invoice.html', context)

            # Save form data to profile
            profile.invoice_header_color = form.cleaned_data['invoice_header_color']
            profile.invoice_font_size = form.cleaned_data['invoice_font_size']
            profile.show_logo = form.cleaned_data['show_logo']
            profile.show_address = form.cleaned_data['show_address']
            profile.term = form.cleaned_data['term']
            payment_link_provider = form.cleaned_data.get('payment_link_provider')
            profile.show_note = form.cleaned_data['show_note']
            profile.invoice_sequence_next = invoice_sequence_next

            # Update note only if show_note is True
            if profile.show_note:
                profile.note = form.cleaned_data['note']
            else:
                profile.note = ''  # Clear the note if not shown

            if payment_link_provider:
                profile.payment_link_provider = payment_link_provider

            profile.save()
            return redirect('accounts:home')  # Redirect to refresh the page
    else:
        # Initialize form with profile data
        initial_data = {
            'invoice_header_color': profile.invoice_header_color,
            'invoice_font_size': profile.invoice_font_size,
            'show_logo': profile.show_logo,
            'show_address': profile.show_address,
            'term': profile.term,
            'show_note': profile.show_note,
            'note': profile.note,
            'invoice_sequence_next': profile.invoice_sequence_next,
            'payment_link_provider': profile.payment_link_provider,
        }
        form = InvoiceCustomizationForm(initial=initial_data)

    # Context to include profile information
    context = {
        'form': form,
        'profile': profile,
    }

    return render(request, 'accounts/customize_invoice.html', context)


@login_required
def qr_code_style(request):
    profile = request.user.profile
    sample_product = (
        Product.objects.filter(user__in=get_product_user_ids(request.user))
        .order_by('-updated_at', '-id')
        .first()
    )

    sample_qr_url = ""
    if sample_product:
        sample_qr_url = request.build_absolute_uri(
            reverse('accounts:qr_stock_in', args=[sample_product.id])
        )

    if request.method == 'POST':
        form = QRCodeStyleForm(request.POST)
        if form.is_valid():
            profile.qr_code_font_scale = form.cleaned_data['qr_code_font_scale']
            profile.qr_show_name = form.cleaned_data['show_product_name']
            profile.qr_show_description = form.cleaned_data['show_product_description']
            profile.qr_show_sku = form.cleaned_data['show_product_sku']
            profile.save(
                update_fields=[
                    'qr_code_font_scale',
                    'qr_show_name',
                    'qr_show_description',
                    'qr_show_sku',
                ]
            )
            messages.success(request, 'QR code label preferences saved.')
            return redirect('accounts:qr_code_style')
    else:
        form = QRCodeStyleForm(
            initial={
                'qr_code_font_scale': profile.qr_code_font_scale,
                'show_product_name': profile.qr_show_name,
                'show_product_description': profile.qr_show_description,
                'show_product_sku': profile.qr_show_sku,
            }
        )

    context = {
        'form': form,
        'profile': profile,
        'sample_product': sample_product,
        'sample_qr_url': sample_qr_url,
    }

    return render(request, 'accounts/qr_code_style.html', context)


@login_required
def edit_profile(request):
    profile, created = Profile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('accounts:profile_view')
    else:
        form = ProfileForm(instance=profile)
    return render(request, 'accounts/edit_profile.html', {'form': form})

@login_required
def profile_view(request):
    profile = Profile.objects.get(user=request.user)
    return render(request, 'accounts/profile_view.html', {'profile': profile})

@login_required
def manage_models(request):
    return render(request, 'app/manage_models.html')


class MechExpenseListView(LoginRequiredMixin, ListView):
    model = MechExpense
    template_name = 'app/mechexpense_list.html'
    context_object_name = 'object_list'
    paginate_by = 100

    def _get_filter_params(self):
        period_value = self.request.GET.get('period')
        if period_value is None:
            period_value = '1y'
        return {
            'search': (self.request.GET.get('search') or '').strip(),
            'period': (period_value or '').strip().lower(),
            'supplier': (self.request.GET.get('supplier') or '').strip(),
            'status': (self.request.GET.get('status') or '').strip().lower(),
            'sort': (self.request.GET.get('sort') or 'date').strip(),
            'direction': (self.request.GET.get('direction') or 'desc').strip().lower(),
        }

    def _normalize_sort(self, sort_value, direction_value):
        sort_map = {
            'receipt': 'receipt_no',
            'date': 'date',
            'supplier': 'vendor',
            'category': 'categorie',
            'total': 'total_incl_tax',
            'status': 'payment_rank',
        }
        sort_key = sort_value if sort_value in sort_map else 'date'
        direction = direction_value if direction_value in {'asc', 'desc'} else 'desc'
        return sort_key, direction, sort_map[sort_key]

    def get_queryset(self):
        params = self._get_filter_params()
        query = params['search']
        period = params['period']
        supplier = params['supplier']
        status = params['status']
        user = self.request.user
        sort_key, direction, sort_field = self._normalize_sort(params['sort'], params['direction'])
        params['sort'] = sort_key
        params['direction'] = direction
        self.filter_params = params

        items_queryset = (
            MechExpenseItem.objects.filter(mech_expense=OuterRef('pk'))
            .values('mech_expense')
        )
        items_subtotal = items_queryset.annotate(
            total=Coalesce(Sum('amount'), Value(0.0))
        ).values('total')
        items_tax = items_queryset.annotate(
            total=Coalesce(Sum('tax_paid'), Value(0.0))
        ).values('total')
        credit_items = (
            SupplierCreditItem.objects.filter(source_expense=OuterRef('pk'))
            .values('source_expense')
        )
        credit_amount_expr = Case(
            When(supplier_credit__tax_included=True, then=F('amount')),
            default=ExpressionWrapper(F('amount') + F('tax_paid'), output_field=FloatField()),
            output_field=FloatField(),
        )
        credit_total = credit_items.annotate(
            total=Coalesce(Sum(credit_amount_expr), Value(0.0))
        ).values('total')

        queryset = (
            MechExpense.objects.filter(user=user)
            .annotate(paid_total=Coalesce(Sum('payments__amount'), Value(Decimal('0.00'))))
            .annotate(
                items_subtotal=Coalesce(Subquery(items_subtotal, output_field=FloatField()), Value(0.0)),
                items_tax=Coalesce(Subquery(items_tax, output_field=FloatField()), Value(0.0)),
                credit_total=Coalesce(Subquery(credit_total, output_field=FloatField()), Value(0.0)),
            )
            .annotate(
                total_incl_tax=Case(
                    When(tax_included=True, then=F('items_subtotal')),
                    default=F('items_subtotal') + F('items_tax'),
                    output_field=FloatField(),
                ),
                payment_rank=Case(
                    When(paid=True, then=Value(2)),
                    When(
                        paid=False,
                        then=Case(
                            When(
                                Q(paid_total__gt=0) | Q(credit_total__gt=0),
                                then=Value(1),
                            ),
                            default=Value(0),
                            output_field=IntegerField(),
                        ),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                ),
            )
            .prefetch_related('mechexpenseitem_set', 'payments')
        )

        if query:
            queryset = queryset.filter(
                Q(vendor__icontains=query) |
                Q(date__icontains=query) |
                Q(receipt_no__icontains=query)
            )

        if supplier:
            queryset = queryset.filter(vendor__iexact=supplier)

        if status:
            if status == 'paid':
                queryset = queryset.filter(paid=True)
            elif status == 'partial':
                queryset = queryset.filter(
                    paid=False
                ).filter(Q(paid_total__gt=0) | Q(credit_total__gt=0))
            elif status == 'unpaid':
                queryset = queryset.filter(paid=False, paid_total=0, credit_total=0)

        if period:
            today = datetime.date.today()
            start_date = None
            end_date = None
            if period == 'today':
                start_date = today
                end_date = today
            elif period == 'yesterday':
                start_date = today - datetime.timedelta(days=1)
                end_date = start_date
            elif period == 'this_week':
                start_date = today - datetime.timedelta(days=today.weekday())
                end_date = start_date + datetime.timedelta(days=6)
            elif period == 'last_week':
                end_date = today - datetime.timedelta(days=today.weekday() + 1)
                start_date = end_date - datetime.timedelta(days=6)
            elif period == 'this_month':
                start_date = today.replace(day=1)
                end_date = today
            elif period == 'last_month':
                first_day_this_month = today.replace(day=1)
                end_date = first_day_this_month - datetime.timedelta(days=1)
                start_date = end_date.replace(day=1)
            elif period == 'this_year':
                start_date = today.replace(month=1, day=1)
                end_date = today
            elif period == 'last_year':
                start_date = today.replace(year=today.year - 1, month=1, day=1)
                end_date = today.replace(year=today.year - 1, month=12, day=31)
            elif period == '1y':
                start_date = today - relativedelta(years=1)
                end_date = today

            if start_date and end_date:
                queryset = queryset.filter(date__range=(start_date, end_date))
            elif start_date:
                queryset = queryset.filter(date__gte=start_date)
            elif end_date:
                queryset = queryset.filter(date__lte=end_date)

        ordering = sort_field if direction == 'asc' else f'-{sort_field}'
        queryset = queryset.order_by(ordering, '-pk')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = getattr(self, 'filter_params', None) or self._get_filter_params()
        supplier_options = (
            MechExpense.objects.filter(user=self.request.user)
            .exclude(vendor__isnull=True)
            .exclude(vendor__exact='')
            .values_list('vendor', flat=True)
            .distinct()
            .order_by('vendor')
        )
        query_params = self.request.GET.copy()
        query_params.pop('page', None)
        query_string = query_params.urlencode()
        context.update({
            'search_query': params['search'],
            'selected_period': params['period'],
            'selected_supplier': params['supplier'],
            'selected_status': params['status'],
            'sort': params['sort'],
            'direction': params['direction'],
            'supplier_options': supplier_options,
            'query_params': f'&{query_string}' if query_string else '',
        })
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string('app/mechexpense_list_content.html', context, request=self.request)
            return JsonResponse({'html': html})
        return super().render_to_response(context, **response_kwargs)


@login_required
@require_POST
@login_required
@require_POST
def mark_mech_expenses_paid(request):
    """Mark one or more mechanic expenses as paid for the current user."""

    expense_ids = []
    mark_all = False
    record_payments = False
    payment_method = None
    payment_notes = None

    try:
        if request.headers.get('Content-Type', '').startswith('application/json'):
            payload = json.loads(request.body.decode('utf-8') or '{}')
            expense_ids = payload.get('expense_ids', [])
            mark_all = bool(payload.get('mark_all'))
            record_payments = bool(payload.get('record_payments'))
            payment_method = (payload.get('payment_method') or '').strip() or None
            payment_notes = (payload.get('payment_notes') or '').strip() or None
        else:
            expense_ids = request.POST.getlist('expense_ids') or request.POST.getlist('selected_expenses')
            mark_all = (request.POST.get('mark_all') or '').lower() in {'1', 'true', 'yes', 'on'}
            record_payments = (request.POST.get('record_payments') or '').lower() in {'1', 'true', 'yes', 'on'}
            payment_method = (request.POST.get('payment_method') or '').strip() or None
            payment_notes = (request.POST.get('payment_notes') or '').strip() or None
    except json.JSONDecodeError:
        expense_ids = []

    if mark_all and not record_payments:
        record_payments = True

    if not mark_all and not expense_ids:
        message = 'Please select at least one expense to mark as paid.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': message}, status=400)
        messages.warning(request, message)
        return redirect(request.POST.get('next') or reverse('accounts:mechexpense_list'))

    user_expenses = (
        MechExpense.objects.filter(user=request.user)
        if mark_all
        else MechExpense.objects.filter(user=request.user, pk__in=expense_ids)
    )
    expenses = list(
        user_expenses
        .prefetch_related(
            'mechexpenseitem_set',
            'supplier_credit_items__supplier_credit',
        )
        .annotate(
            paid_total=Coalesce(
                Sum('payments__amount'),
                Value(Decimal('0.00')),
                output_field=DecimalField(),
            )
        )
    )

    if not expenses:
        message = 'We could not find the selected expenses.' if not mark_all else 'No expenses were found to update.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': message}, status=404)
        messages.warning(request, message)
        return redirect(request.POST.get('next') or reverse('accounts:mechexpense_list'))

    expenses_to_mark = []
    expenses_to_pay = []
    for expense in expenses:
        remaining = expense.remaining_balance
        if remaining > 0:
            expenses_to_mark.append(expense)
            expenses_to_pay.append((expense, remaining))
        elif not expense.paid:
            expenses_to_mark.append(expense)

    updated_ids = [expense.pk for expense in expenses_to_mark]
    updated_count = len(updated_ids)
    already_paid_count = len(expenses) - updated_count

    payments_created = 0
    if record_payments and expenses_to_pay:
        if not payment_method:
            payment_method = "Bulk update"
        if payment_notes is None:
            payment_notes = "Marked as paid in bulk."

        payments_to_create = []
        for expense, remaining in expenses_to_pay:
            payments_to_create.append(
                MechExpensePayment(
                    mech_expense=expense,
                    amount=remaining.quantize(Decimal('0.01')),
                    method=payment_method,
                    notes=payment_notes,
                    recorded_by=request.user,
                )
            )
        if payments_to_create:
            MechExpensePayment.objects.bulk_create(payments_to_create)
            payments_created = len(payments_to_create)

    if updated_ids:
        MechExpense.objects.filter(pk__in=updated_ids).update(paid=True)

    target_label = "expenses"
    target_prefix = "all" if mark_all else "selected"
    if updated_count:
        success_message = (
            f"{updated_count} expense{'s' if updated_count != 1 else ''} marked as paid."
        )
        if payments_created:
            success_message += (
                f" {payments_created} payment record{'s' if payments_created != 1 else ''} created."
            )
    else:
        success_message = f"{target_prefix.capitalize()} {target_label} are already marked as paid."

    if already_paid_count and updated_count:
        success_message += f" {already_paid_count} expense{'s' if already_paid_count != 1 else ''} were already paid."
    elif already_paid_count and not updated_count:
        success_message = f"{target_prefix.capitalize()} {target_label} were already marked as paid."

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': success_message,
            'updated_ids': updated_ids,
        })

    if updated_count:
        messages.success(request, success_message)
    else:
        messages.info(request, success_message)

    next_url = request.POST.get('next') or reverse('accounts:mechexpense_list')
    return redirect(next_url)


@login_required
@require_POST
def toggle_mech_expense_status(request, pk):
    """Toggle the paid status for a single mechanic expense."""

    expense = get_object_or_404(MechExpense, pk=pk, user=request.user)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        payload = {}

    set_to = payload.get('set_to') if isinstance(payload, dict) else None
    amount_raw = payload.get('amount')
    method = (payload.get('method') or '').strip()
    notes = payload.get('notes')
    delete_history = bool(payload.get('delete_history'))

    total_amount_incl_tax, _, _ = expense.calculate_totals()
    total_amount_incl_tax = Decimal(str(total_amount_incl_tax or 0))

    normalized = set_to.lower() if isinstance(set_to, str) else None
    wants_unpaid = normalized in {'unpaid', 'false', '0'} or set_to is False

    if wants_unpaid:
        expense.paid = False
        if delete_history:
            expense.payments.all().delete()
        expense.save(update_fields=['paid'])

        message = (
            'Expense marked as unpaid. Payment history has been cleared.'
            if delete_history else 'Expense marked as unpaid.'
        )

        response_payload = {
            'success': True,
            'paid': expense.paid,
            'payment_status': expense.payment_status,
            'message': message,
            'total_paid': str(expense.total_paid_amount),
            'remaining_balance': str(expense.remaining_balance),
        }

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(response_payload)

        messages.info(request, message)
        return redirect('accounts:mechexpense_list')

    try:
        if amount_raw in (None, ''):
            payment_amount = expense.remaining_balance
        else:
            payment_amount = Decimal(str(amount_raw))
            payment_amount = payment_amount.quantize(Decimal('0.01'))
    except (InvalidOperation, TypeError):
        return JsonResponse({'success': False, 'message': 'Please enter a valid payment amount.'}, status=400)

    if payment_amount <= 0:
        return JsonResponse({'success': False, 'message': 'Payment amount must be greater than zero.'}, status=400)

    if payment_amount > expense.remaining_balance:
        return JsonResponse({
            'success': False,
            'message': 'Payment amount cannot exceed the remaining balance.',
        }, status=400)

    payment_method = method or 'Cash'
    MechExpensePayment.objects.create(
        mech_expense=expense,
        amount=payment_amount,
        method=payment_method,
        notes=notes,
        recorded_by=request.user,
    )

    total_paid = expense.total_paid_amount
    expense.paid = (total_paid + expense.total_credit_amount) >= total_amount_incl_tax
    expense.save(update_fields=['paid'])

    message = (
        'Expense marked as paid.' if expense.paid else 'Payment recorded. Expense is partially paid.'
    )

    response_payload = {
        'success': True,
        'paid': expense.paid,
        'payment_status': expense.payment_status,
        'message': message,
        'total_paid': str(total_paid),
        'remaining_balance': str(expense.remaining_balance),
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse(response_payload)

    if expense.paid:
        messages.success(request, message)
    else:
        messages.info(request, message)

    return redirect('accounts:mechexpense_list')

class GroupedInvoiceCreateView(LoginRequiredMixin, CreateView):
    model = GroupedInvoice
    form_class = GroupedInvoiceForm
    template_name = 'app/groupedinvoice_form.html'
    success_url = reverse_lazy('accounts:groupedinvoice_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['incomerecord2_formset'] = IncomeRecord2FormSet(self.request.POST, form_kwargs={'user': self.request.user})
        else:
            context['incomerecord2_formset'] = IncomeRecord2FormSet(form_kwargs={'user': self.request.user})
            context['generated_invoice_number'] = GroupedInvoice.generate_invoice_number(
                self.request.user,
                commit=False,
            )
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        incomerecord2_formset = context['incomerecord2_formset']
        if form.is_valid() and incomerecord2_formset.is_valid():
            self.object = form.save()
            incomerecord2_formset.instance = self.object
            incomerecord2_formset.save()

            # Ensure inventory transactions exist for each sold product
            self.object.ensure_inventory_transactions()

            first_record = self.object.income_records.order_by('line_order', 'id').first()
            if first_record and self.object.customer and first_record.rate is not None:
                if self.object.customer.charge_rate != first_record.rate:
                    self.object.customer.charge_rate = first_record.rate
                    self.object.customer.save(update_fields=['charge_rate'])
            messages.success(self.request, f'Invoice {self.object.invoice_number} created successfully with {incomerecord2_formset.total_form_count()} jobs.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class GroupedInvoiceUpdateView(LoginRequiredMixin, UpdateView):
    model = GroupedInvoice
    form_class = GroupedInvoiceForm
    template_name = 'app/groupedinvoice_form.html'
    success_url = reverse_lazy('accounts:groupedinvoice_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['incomerecord2_formset'] = IncomeRecord2FormSet(self.request.POST, instance=self.object, form_kwargs={'user': self.request.user})
        else:
            context['incomerecord2_formset'] = IncomeRecord2FormSet(instance=self.object, form_kwargs={'user': self.request.user})
        context['generated_invoice_number'] = self.object.invoice_number
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['incomerecord2_formset']
        if form.is_valid() and formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # Ensure inventory transactions exist for each sold product
            self.object.ensure_inventory_transactions()

            first_record = self.object.income_records.order_by('line_order', 'id').first()
            if first_record and self.object.customer and first_record.rate is not None:
                if self.object.customer.charge_rate != first_record.rate:
                    self.object.customer.charge_rate = first_record.rate
                    self.object.customer.save(update_fields=['charge_rate'])
            # Recalculate the total and update payment link after changes
            self.object.recalculate_total_amount()
            self.object.save()  # This will trigger create_payment_link() if needed
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class GroupedInvoiceDeleteView(LoginRequiredMixin, DeleteView):
    model = GroupedInvoice
    template_name = 'app/groupedinvoice_confirm_delete.html'
    success_url = reverse_lazy('accounts:groupedinvoice_list')


@login_required
@require_POST
def quick_invoice_create(request):
    """
    Lightweight invoice creation endpoint for the dashboard "Quick Invoice" form.
    Returns JSON with redirect_url when called via AJAX.
    """
    from django.urls import reverse
    from django.utils.dateparse import parse_date
    from django.db import transaction

    customer_id = request.POST.get('customer') or ''
    vehicle_id = (request.POST.get('vehicle') or '').strip()
    raw_date = request.POST.get('date') or ''
    bill_to_email = (request.POST.get('bill_to_email') or '').strip() or None

    # Support multiple invoice lines by allowing repeated fields (getlist),
    # while staying backwards compatible with single-value posts.
    line_sources = [((v or '').strip().lower()) for v in request.POST.getlist('line_source')]
    product_ids = [((v or '').strip()) for v in request.POST.getlist('product_id')]
    service_ids = [((v or '').strip()) for v in request.POST.getlist('service_id')]
    jobs = [((v or '').strip()) for v in request.POST.getlist('job')]
    qty_raw_list = [((v or '').strip()) for v in request.POST.getlist('qty')]
    rate_raw_list = [((v or '').strip()) for v in request.POST.getlist('rate')]

    # If the client posted a single value for these fields (older UI), Django's
    # getlist still returns a single-item list as long as the key exists. If the
    # keys were omitted entirely, normalize to a single empty row.
    if not any([line_sources, product_ids, service_ids, jobs, qty_raw_list, rate_raw_list]):
        line_sources = ['']
        product_ids = ['']
        service_ids = ['']
        jobs = ['']
        qty_raw_list = ['']
        rate_raw_list = ['']

    tax_exempt = (request.POST.get('tax_exempt') in ('1', 'true', 'on', 'yes'))
    post_action = (request.POST.get('post_action') or 'view').strip().lower()

    def _json_error(message: str, *, status=400):
        payload = {
            'success': False,
            'errors': {'form': {'__all__': [message]}},
        }
        return JsonResponse(payload, status=status)

    if not customer_id:
        return _json_error('Please select a customer.')

    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, pk=customer_id, user__in=business_user_ids)

    invoice_date = parse_date(raw_date) if raw_date else None
    if not invoice_date:
        invoice_date = timezone.now().date()

    vehicle = None
    if vehicle_id:
        try:
            vehicle_pk = int(vehicle_id)
        except (TypeError, ValueError):
            vehicle_pk = None
        if vehicle_pk:
            vehicle = Vehicle.objects.filter(
                pk=vehicle_pk,
                customer_id=customer.pk,
                customer__user__in=business_user_ids,
            ).first()

    def _parse_decimal(value: str):
        value = (value or '').strip()
        if value == '':
            return None
        return Decimal(value)

    with transaction.atomic():
        invoice = GroupedInvoice.objects.create(
            user=request.user,
            customer=customer,
            date=invoice_date,
            bill_to=customer.name,
            bill_to_address=customer.address or '',
            bill_to_email=bill_to_email or (customer.email or None),
            vin_no=(vehicle.vin_number if vehicle and vehicle.vin_number else None),
            mileage=(float(vehicle.current_mileage) if vehicle and vehicle.current_mileage is not None else None),
            unit_no=(vehicle.unit_number if vehicle and vehicle.unit_number else None),
            make_model=(vehicle.make_model if vehicle and vehicle.make_model else None),
            license_plate=(vehicle.license_plate if vehicle and vehicle.license_plate else None),
            tax_exempt=tax_exempt,
        )

        pending_invoice, _ = PendingInvoice.objects.get_or_create(grouped_invoice=invoice)

        def _at(lst, idx, default=''):
            try:
                return lst[idx]
            except Exception:
                return default

        line_count = max(
            len(line_sources),
            len(product_ids),
            len(service_ids),
            len(jobs),
            len(qty_raw_list),
            len(rate_raw_list),
        )
        if line_count <= 0:
            line_count = 1

        for idx in range(line_count):
            job = (_at(jobs, idx, '') or '').strip()
            line_source = (_at(line_sources, idx, '') or '').strip().lower()
            product_id = (_at(product_ids, idx, '') or '').strip()
            service_id = (_at(service_ids, idx, '') or '').strip()

            qty_raw = (_at(qty_raw_list, idx, '') or '').strip()
            rate_raw = (_at(rate_raw_list, idx, '') or '').strip()

            try:
                qty = _parse_decimal(qty_raw)
                rate = _parse_decimal(rate_raw)
            except (InvalidOperation, TypeError, ValueError):
                return _json_error(f'Line {idx + 1}: Qty and rate must be valid numbers.')

            resolved_product = None
            if product_id:
                try:
                    resolved_product = Product.objects.filter(
                        user__in=get_product_user_ids(request.user),
                        pk=int(product_id),
                    ).first()
                except (TypeError, ValueError):
                    resolved_product = None

            resolved_service = None
            if service_id:
                try:
                    resolved_service = Service.objects.filter(user=request.user, pk=int(service_id), is_active=True).first()
                except (TypeError, ValueError):
                    resolved_service = None

            # If caller says "product/service" but the target isn't resolvable, fall back gracefully.
            if line_source == 'product' and not resolved_product:
                line_source = 'custom'
            if line_source == 'service' and not resolved_service:
                line_source = 'custom'

            should_create_line = bool(job) or bool(resolved_product) or bool(resolved_service) or (rate is not None)
            if not should_create_line:
                continue

            # Prefer service defaults if a service is selected and qty/rate were not supplied.
            if resolved_service:
                if qty is None and resolved_service.fixed_hours is not None:
                    qty = Decimal(str(resolved_service.fixed_hours))
                if rate is None and resolved_service.fixed_rate is not None:
                    rate = Decimal(str(resolved_service.fixed_rate))

            IncomeRecord2.objects.create(
                grouped_invoice=invoice,
                pending_invoice=pending_invoice,
                product=resolved_product,
                job=job or (resolved_product.name if resolved_product else 'Service'),
                qty=qty if qty is not None else Decimal('1'),
                rate=rate if rate is not None else Decimal('0'),
                date=invoice_date,
            )

        # Keep totals/payment links consistent with the rest of the app.
        invoice.recalculate_total_amount()

    redirect_url = reverse('accounts:groupedinvoice_detail', args=[invoice.pk])
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        if post_action == 'continue':
            return JsonResponse({'success': True, 'invoice_id': invoice.pk})
        return JsonResponse({'success': True, 'redirect_url': redirect_url, 'invoice_id': invoice.pk})
    if post_action == 'continue':
        return redirect('accounts:home')
    return redirect(redirect_url)

class MechExpenseDetailView(LoginRequiredMixin, DetailView):
    model = MechExpense
    template_name = 'app/mechexpense_detail.html'
    context_object_name = 'object'

    def get_queryset(self):
        queryset = super().get_queryset()
        accountant_profile = getattr(self.request.user, "accountant_portal", None)
        if accountant_profile:
            detail_access = accountant_profile.accountant_access_level in ("full", "read_only")
            if not detail_access:
                raise Http404
            return queryset.filter(user=accountant_profile.user)
        return queryset.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        accountant_profile = getattr(self.request.user, "accountant_portal", None)
        context['portal_read_only'] = bool(accountant_profile)
        expense = self.get_object()
        total_amount_incl_tax, total_tax, total_amount_excl_tax = expense.calculate_totals()
        context['total_amount_incl_tax'] = total_amount_incl_tax
        context['total_tax'] = total_tax
        context['total_amount_excl_tax'] = total_amount_excl_tax
        context['tax_included'] = expense.tax_included
        context['tax_name'] = expense.get_tax_label()
        context['province'] = expense.get_province_display()
        context['required_categories'] = ['Parts', 'Supplies', 'Miscellaneous', 'Tools']
        context['payment_status'] = expense.payment_status
        context['total_paid'] = expense.total_paid_amount
        context['remaining_balance'] = expense.remaining_balance
        context['payments'] = expense.payments.order_by('-created_at')

        # Handle Fuel category for fuel economy calculation
        if expense.categorie == 'Fuel' and expense.odometer_reading:
            # Get the previous fuel expense with an odometer reading
            previous_expense_qs = MechExpense.objects.filter(
                user=expense.user,
                categorie='Fuel',
                date__lt=expense.date,
                odometer_reading__isnull=False
            )
            unit_number = (expense.unit_number or '').strip()
            if unit_number:
                previous_expense_qs = previous_expense_qs.filter(unit_number__iexact=unit_number)
            previous_expense = previous_expense_qs.order_by('-date').first()
            if previous_expense:
                distance = expense.odometer_reading - previous_expense.odometer_reading
                if distance > 0:
                    # Calculate total liters of fuel purchased
                    total_liters = sum(item.qty for item in expense.mechexpenseitem_set.all())
                    if total_liters > 0:
                        fuel_economy = distance / total_liters  # km per liter
                        context['fuel_economy'] = round(fuel_economy, 2)
        elif expense.categorie in ['Rent', 'Insurance']:
            # Handle start and end dates
            context['start_date'] = expense.start_date
            context['end_date'] = expense.end_date
            if expense.start_date and expense.end_date:
                rental_duration = (expense.end_date - expense.start_date).days
                context['rental_duration'] = rental_duration

        # Recurring information
        context['is_recurring'] = expense.is_recurring
        context['frequency'] = expense.get_frequency_display()
        context['next_occurrence'] = expense.next_occurrence

        return context


class MechExpensePaymentMixin:
    def record_inline_payment(self, expense):
        if not getattr(expense, "pk", None):
            return
        record_payment = self.request.POST.get('record_payment') == 'true'
        if not record_payment:
            return

        total_amount_incl_tax, _, _ = expense.calculate_totals()
        total_amount_decimal = Decimal(str(total_amount_incl_tax or 0))
        if total_amount_decimal <= Decimal('0'):
            return

        payment_amount_raw = self.request.POST.get('payment_amount')
        try:
            payment_amount = (
                Decimal(str(payment_amount_raw))
                if payment_amount_raw not in (None, '')
                else total_amount_decimal
            )
            payment_amount = payment_amount.quantize(Decimal('0.01'))
        except (InvalidOperation, TypeError):
            return

        if payment_amount <= 0:
            return
        if payment_amount > total_amount_decimal:
            payment_amount = total_amount_decimal

        payment_method = self.request.POST.get('payment_method') or 'Cash'
        payment_notes = self.request.POST.get('payment_notes') or ''
        create_cheque = self.request.POST.get('payment_create_cheque') == 'true'
        cheque = None

        if payment_method == 'Cheque' and create_cheque:
            cheque_number = (self.request.POST.get('payment_cheque_number') or '').strip()
            bank_account = (self.request.POST.get('payment_bank_account') or '').strip()
            cheque_memo = (self.request.POST.get('payment_cheque_memo') or '').strip()
            cheque_date_raw = self.request.POST.get('payment_cheque_date')
            cheque_date = expense.date or timezone.localdate()
            if cheque_date_raw:
                try:
                    cheque_date = datetime.datetime.strptime(cheque_date_raw, "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    cheque_date = expense.date or timezone.localdate()

            if cheque_number:
                supplier_name = (expense.vendor or '').strip()
                supplier = _get_or_create_supplier_for_user(self.request.user, supplier_name)
                cheque = SupplierCheque.objects.create(
                    user=self.request.user,
                    supplier=supplier,
                    supplier_name=supplier_name,
                    cheque_number=cheque_number,
                    bank_account=bank_account,
                    date=cheque_date,
                    memo=cheque_memo or payment_notes or None,
                )
            else:
                messages.warning(self.request, 'Cheque number missing. Payment recorded without creating a cheque.')

        notes = payment_notes
        if cheque and cheque.cheque_number:
            cheque_note = f"Cheque #{cheque.cheque_number}"
            notes = f"{cheque_note} - {notes}" if notes else cheque_note

        payment = MechExpensePayment.objects.create(
            mech_expense=expense,
            amount=payment_amount,
            method=payment_method,
            notes=notes,
            recorded_by=self.request.user,
            cheque=cheque,
        )

        if cheque:
            SupplierChequeLine.objects.create(
                cheque=cheque,
                mech_expense=expense,
                amount=payment.amount,
                memo=cheque.memo,
            )

        total_paid = expense.total_paid_amount
        expense.paid = (total_paid + expense.total_credit_amount) >= total_amount_decimal
        expense.save(update_fields=['paid'])


def _record_inline_payment_from_request(expense, request):
    """
    Record a payment when submitted from function-based views (e.g., add_records).
    Mirrors MechExpensePaymentMixin.record_inline_payment.
    """
    if not getattr(expense, "pk", None):
        return
    record_payment = request.POST.get('record_payment') == 'true'
    if not record_payment:
        return

    total_amount_incl_tax, _, _ = expense.calculate_totals()
    total_amount_decimal = Decimal(str(total_amount_incl_tax or 0))
    if total_amount_decimal <= Decimal('0'):
        return

    payment_amount_raw = request.POST.get('payment_amount')
    try:
        payment_amount = (
            Decimal(str(payment_amount_raw))
            if payment_amount_raw not in (None, '')
            else total_amount_decimal
        )
        payment_amount = payment_amount.quantize(Decimal('0.01'))
    except (InvalidOperation, TypeError):
        return

    if payment_amount <= 0:
        return
    if payment_amount > total_amount_decimal:
        payment_amount = total_amount_decimal

    payment_method = request.POST.get('payment_method') or 'Cash'
    payment_notes = request.POST.get('payment_notes') or ''
    create_cheque = request.POST.get('payment_create_cheque') == 'true'
    cheque = None

    if payment_method == 'Cheque' and create_cheque:
        cheque_number = (request.POST.get('payment_cheque_number') or '').strip()
        bank_account = (request.POST.get('payment_bank_account') or '').strip()
        cheque_memo = (request.POST.get('payment_cheque_memo') or '').strip()
        cheque_date_raw = request.POST.get('payment_cheque_date')
        cheque_date = expense.date or timezone.localdate()
        if cheque_date_raw:
            try:
                cheque_date = datetime.datetime.strptime(cheque_date_raw, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                cheque_date = expense.date or timezone.localdate()

        if cheque_number:
            supplier_name = (expense.vendor or '').strip()
            supplier = _get_or_create_supplier_for_user(request.user, supplier_name)
            cheque = SupplierCheque.objects.create(
                user=request.user,
                supplier=supplier,
                supplier_name=supplier_name,
                cheque_number=cheque_number,
                bank_account=bank_account,
                date=cheque_date,
                memo=cheque_memo or payment_notes or None,
            )
        else:
            messages.warning(request, 'Cheque number missing. Payment recorded without creating a cheque.')

    notes = payment_notes
    if cheque and cheque.cheque_number:
        cheque_note = f"Cheque #{cheque.cheque_number}"
        if notes:
            notes = f"{cheque_note} - {notes}"
        else:
            notes = cheque_note

    payment = MechExpensePayment.objects.create(
        mech_expense=expense,
        amount=payment_amount,
        method=payment_method,
        notes=notes,
        recorded_by=request.user,
        cheque=cheque,
    )

    if cheque:
        SupplierChequeLine.objects.create(
            cheque=cheque,
            mech_expense=expense,
            amount=payment.amount,
            memo=cheque.memo,
        )

    total_paid = expense.total_paid_amount
    expense.paid = (total_paid + expense.total_credit_amount) >= total_amount_decimal
    expense.save(update_fields=['paid'])


class MechExpenseAjaxResponseMixin:
    """Provide lightweight JSON responses for AJAX submissions."""

    def _is_ajax(self):
        return self.request.headers.get('x-requested-with') == 'XMLHttpRequest'

    def _serialize_errors(self, form=None, formset=None, status=400):
        if not self._is_ajax():
            return None

        errors = {}
        if form is not None:
            errors['form'] = form.errors.get_json_data()

        if formset is not None:
            item_errors = []
            for idx, inline_form in enumerate(formset.forms):
                if inline_form.errors:
                    item_errors.append({
                        "index": idx,
                        "errors": inline_form.errors.get_json_data(),
                    })
            non_form_errors = list(formset.non_form_errors())
            if item_errors:
                errors['items'] = item_errors
            if non_form_errors:
                errors['formset'] = non_form_errors

        return JsonResponse({"success": False, "errors": errors}, status=status)

    def _success_response(self, expense):
        if not self._is_ajax():
            return None

        return JsonResponse({
            "success": True,
            "expense_id": expense.pk,
            "receipt_no": expense.receipt_no,
            "redirect": self.get_success_url(),
        })


class MechExpenseInventoryMixin:
    """Sync mech expense item changes to inventory transactions."""

    @staticmethod
    def _find_or_create_product(expense, item):
        user = expense.user
        part = (item.part_no or "").strip()
        desc = (item.description or "").strip()
        supplier = _get_or_create_supplier_for_user(user, expense.vendor)
        qs = Product.objects.filter(user=user)
        product = None
        if part:
            product = qs.filter(models.Q(sku__iexact=part) | models.Q(name__iexact=part)).first()
        if not product and desc:
            product = qs.filter(name__iexact=desc).first()
        if product:
            if supplier and product.supplier_id != supplier.id:
                product.supplier = supplier
                product.save(update_fields=["supplier"])
            return product

        # Create a minimal product if allowed
        category = _get_or_create_category_for_user(user, expense.categorie or "Parts")
        try:
            product = Product.objects.create(
                user=user,
                name=desc or part or "Expense Item",
                sku=part or None,
                description=desc,
                cost_price=item.price or 0,
                supplier=supplier,
                category=category,
            )
        except Exception as exc:
            logger.warning("Could not create product for expense %s item %s: %s", expense.id, getattr(item, "id", None), exc)
            return None
        return product

    @classmethod
    def _log_transactions(cls, expense, items, transaction_type, remark_suffix):
        for item in items:
            product = cls._find_or_create_product(expense, item)
            if not product:
                continue
            try:
                qty_decimal = Decimal(str(item.qty or 0)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                qty_int = int(qty_decimal) if qty_decimal > 0 else 0
                if qty_int <= 0:
                    continue
                remarks = f"Expense {expense.receipt_no or expense.id} - {remark_suffix}"
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type=transaction_type,
                    quantity=qty_int,
                    remarks=remarks,
                    user=expense.user,
                )
            except Exception as exc:  # best-effort logging
                logger.warning("Inventory sync skipped for expense %s item %s: %s", expense.id, getattr(item, "id", None), exc)

    def _apply_inventory_sync(self, expense, previous_items=None):
        prev = previous_items or []
        if prev:
            self._log_transactions(expense, prev, 'OUT', 'reverse edit/delete')
        current_items = expense.mechexpenseitem_set.all()
        self._log_transactions(expense, current_items, 'IN', 'stock in from expense')

class MechExpenseCreateView(MechExpenseAjaxResponseMixin, MechExpensePaymentMixin, MechExpenseInventoryMixin, LoginRequiredMixin, CreateView):
    model = MechExpense
    form_class = MechExpenseForm
    template_name = 'accounts/mach/add_records.html'
    success_url = reverse_lazy('accounts:mechexpense_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_products_qs = Product.objects.filter(user__in=get_product_user_ids(self.request.user)).order_by('name')
        inventory_products_data = {}
        vendor_product_map = defaultdict(list)
        all_product_ids = []
        for product in inventory_products_qs:
            product_id_str = str(product.id)
            supplier_name = (product.supplier.name or '').strip() if product.supplier else ''
            normalized_supplier = supplier_name.lower()
            display_name = f"{product.name} ({product.sku})" if product.sku else product.name
            inventory_products_data[product_id_str] = {
                'name': product.name,
                'sku': product.sku or '',
                'description': product.description or '',
                'cost_price': float(product.cost_price) if product.cost_price is not None else 0.0,
                'sale_price': float(product.sale_price) if product.sale_price is not None else 0.0,
                'supplier': supplier_name,
                'category_id': product.category_id,
                'category_name': product.category.name if product.category else '',
                'display_name': display_name,
            }
            all_product_ids.append(product_id_str)
            if normalized_supplier:
                vendor_product_map[normalized_supplier].append(product_id_str)
            else:
                vendor_product_map['__unassigned__'].append(product_id_str)
        if self.request.POST:
            categorie = self.request.POST.get('categorie', 'Parts')
            context['mechexpenseitem_formset'] = MechExpenseItemFormSet(
                self.request.POST,
                prefix='mechexpenseitem_set',
                categorie=categorie,
                user=self.request.user,
            )
        else:
            categorie = context.get('form').initial.get('categorie', 'Parts') if context.get('form') else 'Parts'
            context['mechexpenseitem_formset'] = MechExpenseItemFormSet(
                prefix='mechexpenseitem_set',
                categorie=categorie,
                user=self.request.user,
            )
            context['mechexpenseitem_formset'].extra = 1
        context['distinct_vendors'] = _get_vendor_options_for_user(self.request.user)
        context['inventory_products'] = inventory_products_qs
        vendor_product_map.setdefault('__unassigned__', [])
        context['inventory_products_data'] = json.dumps(inventory_products_data)
        context['vendor_product_map'] = json.dumps(dict(vendor_product_map))
        context['all_inventory_product_ids'] = json.dumps(all_product_ids)
        context['province_tax_rates'] = json.dumps(PROVINCE_TAX_RATES)
        context['mech_expense_form'] = context.get('form')
        context['mech_expense_item_formset'] = context.get('mechexpenseitem_formset')
        context['category_form'] = CategoryForm(user=self.request.user)
        context['quick_product_form'] = QuickProductCreateForm(user=self.request.user)
        context['supplier_form'] = SupplierForm(user=self.request.user)
        context['is_edit'] = False
        context['recurring_worker_enabled'] = getattr(settings, "RECURRING_WORKER_ENABLED", False)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        mechexpenseitem_formset = context['mechexpenseitem_formset']
        if not mechexpenseitem_formset.is_valid():
            error_response = self._serialize_errors(form=form, formset=mechexpenseitem_formset)
            if error_response:
                return error_response
            return self.render_to_response(self.get_context_data(form=form))

        self.object = form.save()
        mechexpenseitem_formset.instance = self.object
        mechexpenseitem_formset.save()
        self.record_inline_payment(self.object)
        record_inventory = self.object.record_in_inventory
        if record_inventory:
            self._apply_inventory_sync(self.object)

        messages.success(
            self.request,
            f'Expense {self.object.vendor} created successfully with {mechexpenseitem_formset.total_form_count()} products.',
        )
        success_response = self._success_response(self.object)
        if success_response:
            return success_response
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        context = self.get_context_data(form=form)
        formset = context.get('mechexpenseitem_formset')
        error_response = self._serialize_errors(form=form, formset=formset)
        if error_response:
            return error_response
        return self.render_to_response(context)

class MechExpenseUpdateView(MechExpenseAjaxResponseMixin, MechExpensePaymentMixin, MechExpenseInventoryMixin, LoginRequiredMixin, UpdateView):
    model = MechExpense
    form_class = MechExpenseForm
    template_name = 'accounts/mach/add_records.html'
    success_url = reverse_lazy('accounts:mechexpense_list')
    context_object_name = 'mechexpense'

    def get_form_kwargs(self):
        """
        Pass the current user to the form.
        """
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        """
        Add the formset and distinct vendors to the context data.
        """
        context = super().get_context_data(**kwargs)
        inventory_products_qs = Product.objects.filter(user__in=get_product_user_ids(self.request.user)).order_by('name')
        inventory_products_data = {}
        vendor_product_map = defaultdict(list)
        all_product_ids = []
        for product in inventory_products_qs:
            product_id_str = str(product.id)
            supplier_name = (product.supplier.name or '').strip() if product.supplier else ''
            normalized_supplier = supplier_name.lower()
            display_name = f"{product.name} ({product.sku})" if product.sku else product.name
            inventory_products_data[product_id_str] = {
                'name': product.name,
                'sku': product.sku or '',
                'description': product.description or '',
                'cost_price': float(product.cost_price) if product.cost_price is not None else 0.0,
                'sale_price': float(product.sale_price) if product.sale_price is not None else 0.0,
                'supplier': supplier_name,
                'category_id': product.category_id,
                'category_name': product.category.name if product.category else '',
                'display_name': display_name,
            }
            all_product_ids.append(product_id_str)
            if normalized_supplier:
                vendor_product_map[normalized_supplier].append(product_id_str)
            else:
                vendor_product_map['__unassigned__'].append(product_id_str)
        if self.request.POST:
            categorie = self.request.POST.get('categorie', 'Parts')
            context['mechexpenseitem_formset'] = MechExpenseItemFormSet(
                self.request.POST,
                instance=self.object,
                prefix='mechexpenseitem_set',
                categorie=categorie,
                user=self.request.user,
            )
        else:
            categorie = self.object.categorie or 'Parts'
            context['mechexpenseitem_formset'] = MechExpenseItemFormSet(
                instance=self.object,
                prefix='mechexpenseitem_set',
                categorie=categorie,
                user=self.request.user,
            )
        context['distinct_vendors'] = _get_vendor_options_for_user(self.request.user)
        context['inventory_products'] = inventory_products_qs
        vendor_product_map.setdefault('__unassigned__', [])
        context['inventory_products_data'] = json.dumps(inventory_products_data)
        context['vendor_product_map'] = json.dumps(dict(vendor_product_map))
        context['all_inventory_product_ids'] = json.dumps(all_product_ids)
        context['province_tax_rates'] = json.dumps(PROVINCE_TAX_RATES)
        context['mech_expense_form'] = context.get('form')
        context['mech_expense_item_formset'] = context.get('mechexpenseitem_formset')
        context['category_form'] = CategoryForm(user=self.request.user)
        context['quick_product_form'] = QuickProductCreateForm(user=self.request.user)
        context['supplier_form'] = SupplierForm(user=self.request.user)
        context['is_edit'] = True
        context['recurring_worker_enabled'] = getattr(settings, "RECURRING_WORKER_ENABLED", False)
        return context

    def form_valid(self, form):
        """
        Process the form and formset. After saving the expense record and its items,
        check if inventory recording is selected and create an InventoryTransaction entry.
        """
        context = self.get_context_data()
        mechexpenseitem_formset = context['mechexpenseitem_formset']
        previous_items = list(self.object.mechexpenseitem_set.all())
        if not mechexpenseitem_formset.is_valid():
            error_response = self._serialize_errors(form=form, formset=mechexpenseitem_formset)
            if error_response:
                return error_response
            messages.error(self.request, 'Please correct the errors below.')
            return self.render_to_response(self.get_context_data(form=form))

        self.object = form.save(commit=False)
        self.object.save()
        mechexpenseitem_formset.instance = self.object
        mechexpenseitem_formset.save()

        self.record_inline_payment(self.object)
        record_inventory = self.object.record_in_inventory
        try:
            if record_inventory:
                self._apply_inventory_sync(self.object, previous_items=previous_items)
            elif previous_items:
                # If inventory sync is disabled during edit, reverse any prior stock in.
                self._log_transactions(self.object, previous_items, 'OUT', 'expense edit reversal')
        except Exception:
            logger.warning("Inventory sync on edit failed for expense %s", self.object.id)

        messages.success(
            self.request,
            f'Receipt No. {self.object.receipt_no} updated successfully.'
        )
        success_response = self._success_response(self.object)
        if success_response:
            return success_response
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        context = self.get_context_data(form=form)
        formset = context.get('mechexpenseitem_formset')
        error_response = self._serialize_errors(form=form, formset=formset)
        if error_response:
            return error_response
        messages.error(self.request, 'Please correct the errors below.')
        return self.render_to_response(context)




class MechExpenseDeleteView(MechExpenseInventoryMixin, LoginRequiredMixin, DeleteView):
    model = MechExpense
    template_name = 'app/mechexpense_confirm_delete.html'
    success_url = reverse_lazy('accounts:mechexpense_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        previous_items = list(self.object.mechexpenseitem_set.all())
        response = super().delete(request, *args, **kwargs)
        try:
            self._log_transactions(self.object, previous_items, 'OUT', 'expense deleted')
        except Exception:
            logger.warning("Inventory sync on delete failed for expense %s", getattr(self.object, 'id', None))
        vendor_label = getattr(self.object, "vendor", "")
        messages.success(request, f'Expense for {vendor_label} deleted successfully.')
        return response


def _update_expense_paid_status(expense):
    total_amount_incl_tax, _, _ = expense.calculate_totals()
    total_amount_incl_tax = Decimal(str(total_amount_incl_tax or 0))
    total_paid = expense.total_paid_amount + expense.total_credit_amount
    expense.paid = total_paid >= total_amount_incl_tax
    expense.save(update_fields=['paid'])


def _update_invoice_paid_status(invoice):
    total_amount = Decimal(str(invoice.total_amount or 0))
    total_paid = ensure_decimal(invoice.total_paid())
    pending_invoice = PendingInvoice.objects.filter(grouped_invoice=invoice).first()
    is_paid = total_paid >= total_amount
    if pending_invoice and pending_invoice.is_paid != is_paid:
        pending_invoice.is_paid = is_paid
        pending_invoice.save(update_fields=['is_paid'])
    invoice.update_date_fully_paid()


def _annotate_invoice_credit_totals(queryset, *, invoice_field='pk'):
    credit_items = (
        CustomerCreditItem.objects.filter(source_invoice=OuterRef(invoice_field))
        .values('source_invoice')
    )
    amount_field = DecimalField(max_digits=10, decimal_places=2)
    credit_amount_expr = Case(
        When(customer_credit__tax_included=True, then=Cast('amount', amount_field)),
        default=ExpressionWrapper(
            Cast('amount', amount_field) + Cast('tax_paid', amount_field),
            output_field=amount_field,
        ),
        output_field=amount_field,
    )
    credit_total = credit_items.annotate(
        total=Coalesce(Sum(credit_amount_expr), Value(Decimal('0.00')), output_field=amount_field)
    ).values('total')
    return queryset.annotate(
        credit_total=Coalesce(
            Subquery(credit_total, output_field=amount_field),
            Value(Decimal('0.00')),
            output_field=amount_field,
        )
    )

def _allocate_supplier_payment(
    *,
    user,
    supplier_name,
    total_amount,
    method,
    notes=None,
    cheque=None,
):
    remaining_amount = Decimal(str(total_amount or 0)).quantize(
        Decimal('0.01'),
        rounding=ROUND_HALF_UP,
    )
    if remaining_amount <= 0:
        return Decimal('0.00'), []

    expenses = (
        MechExpense.objects.filter(user=user, vendor__iexact=supplier_name)
        .prefetch_related('payments', 'supplier_credit_items')
        .order_by('date', 'id')
    )

    applied_total = Decimal('0.00')
    applied_expenses = []

    for expense in expenses:
        remaining = expense.remaining_balance
        if remaining <= 0:
            continue
        allocation = remaining if remaining_amount >= remaining else remaining_amount
        allocation = Decimal(str(allocation)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if allocation <= 0:
            continue

        MechExpensePayment.objects.create(
            mech_expense=expense,
            cheque=cheque,
            amount=allocation,
            method=method,
            notes=notes,
            recorded_by=user,
        )
        if cheque:
            SupplierChequeLine.objects.create(
                cheque=cheque,
                mech_expense=expense,
                amount=allocation,
                memo=notes,
            )

        applied_total += allocation
        remaining_amount -= allocation
        _update_expense_paid_status(expense)
        applied_expenses.append(expense)

        if remaining_amount <= 0:
            break

    return applied_total, applied_expenses


class SupplierCreditInventoryMixin:
    """Sync supplier credit items to inventory transactions."""

    @staticmethod
    def _find_or_create_product(credit, item):
        if item.product:
            return item.product

        part = (item.part_no or "").strip()
        desc = (item.description or "").strip()
        supplier_name = credit.supplier_name or (credit.supplier.name if credit.supplier else "")

        if not part and not desc:
            return None

        qs = Product.objects.filter(user=credit.user)
        product = None
        if part:
            product = qs.filter(Q(sku__iexact=part) | Q(name__iexact=part)).first()
        if not product and desc:
            product = qs.filter(name__iexact=desc).first()

        if product and supplier_name and product.supplier_id != getattr(credit.supplier, 'id', None):
            product.supplier = credit.supplier
            product.save(update_fields=['supplier'])
        return product

    @classmethod
    def _log_transactions(cls, credit, items, transaction_type, remark_suffix):
        for item in items:
            product = cls._find_or_create_product(credit, item)
            if not product:
                continue
            try:
                qty_decimal = Decimal(str(item.qty or 0)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                qty_int = int(qty_decimal) if qty_decimal > 0 else 0
                if qty_int <= 0:
                    continue
                remarks = f"Supplier credit {credit.credit_no or credit.id} - {remark_suffix}"
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type=transaction_type,
                    quantity=qty_int,
                    remarks=remarks,
                    user=credit.user,
                )
            except Exception as exc:  # best-effort logging
                logger.warning("Inventory sync skipped for credit %s item %s: %s", credit.id, getattr(item, "id", None), exc)

    def _apply_inventory_sync(self, credit, previous_items=None):
        prev = previous_items or []
        if prev:
            self._log_transactions(credit, prev, 'IN', 'reverse edit/delete')
        current_items = credit.items.all()
        self._log_transactions(credit, current_items, 'OUT', 'stock out from supplier credit')


class CustomerCreditInventoryMixin:
    """Sync customer credit items to inventory transactions."""

    @staticmethod
    def _find_or_create_product(credit, item):
        if item.product:
            return item.product

        part = (item.part_no or "").strip()
        desc = (item.description or "").strip()

        if not part and not desc:
            return None

        qs = Product.objects.filter(user=credit.user)
        product = None
        if part:
            product = qs.filter(Q(sku__iexact=part) | Q(name__iexact=part)).first()
        if not product and desc:
            product = qs.filter(name__iexact=desc).first()

        return product

    @classmethod
    def _log_transactions(cls, credit, items, transaction_type, remark_suffix):
        for item in items:
            product = cls._find_or_create_product(credit, item)
            if not product:
                continue
            try:
                qty_decimal = Decimal(str(item.qty or 0)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                qty_int = int(qty_decimal) if qty_decimal > 0 else 0
                if qty_int <= 0:
                    continue
                remarks = f"Customer credit {credit.credit_no or credit.id} - {remark_suffix}"
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type=transaction_type,
                    quantity=qty_int,
                    remarks=remarks,
                    user=credit.user,
                )
            except Exception as exc:  # best-effort logging
                logger.warning("Inventory sync skipped for credit %s item %s: %s", credit.id, getattr(item, "id", None), exc)

    def _apply_inventory_sync(self, credit, previous_items=None):
        prev = previous_items or []
        if prev:
            self._log_transactions(credit, prev, 'OUT', 'reverse edit/delete')
        current_items = credit.items.all()
        self._log_transactions(credit, current_items, 'IN', 'stock in from customer credit')


class SupplierCreditListView(LoginRequiredMixin, ListView):
    model = SupplierCredit
    template_name = 'app/supplier_credit_list.html'
    context_object_name = 'object_list'
    paginate_by = 100

    def _get_filter_params(self):
        period_value = self.request.GET.get('period')
        if period_value is None:
            period_value = '1y'
        return {
            'search': (self.request.GET.get('search') or '').strip(),
            'period': (period_value or '').strip().lower(),
            'supplier': (self.request.GET.get('supplier') or '').strip(),
        }

    def get_queryset(self):
        params = self._get_filter_params()
        query = params['search']
        period = params['period']
        supplier = params['supplier']
        user = self.request.user

        items_queryset = (
            SupplierCreditItem.objects.filter(supplier_credit=OuterRef('pk'))
            .values('supplier_credit')
        )
        items_subtotal = items_queryset.annotate(
            total=Coalesce(Sum('amount'), Value(0.0))
        ).values('total')
        items_tax = items_queryset.annotate(
            total=Coalesce(Sum('tax_paid'), Value(0.0))
        ).values('total')
        applied_amount_expr = Case(
            When(supplier_credit__tax_included=True, then=F('amount')),
            default=ExpressionWrapper(F('amount') + F('tax_paid'), output_field=FloatField()),
            output_field=FloatField(),
        )
        items_applied = items_queryset.annotate(
            total=Coalesce(
                Sum(applied_amount_expr, filter=Q(source_expense__isnull=False)),
                Value(0.0),
            )
        ).values('total')

        queryset = (
            SupplierCredit.objects.filter(user=user)
            .annotate(
                items_subtotal=Coalesce(Subquery(items_subtotal, output_field=FloatField()), Value(0.0)),
                items_tax=Coalesce(Subquery(items_tax, output_field=FloatField()), Value(0.0)),
                applied_total=Coalesce(Subquery(items_applied, output_field=FloatField()), Value(0.0)),
            )
            .annotate(
                total_incl_tax=Case(
                    When(tax_included=True, then=F('items_subtotal')),
                    default=F('items_subtotal') + F('items_tax'),
                    output_field=FloatField(),
                ),
                available_total=ExpressionWrapper(
                    F('total_incl_tax') - F('applied_total'),
                    output_field=FloatField(),
                ),
            )
            .select_related('supplier')
            .prefetch_related('items')
        )

        if query:
            queryset = queryset.filter(
                Q(credit_no__icontains=query) |
                Q(supplier_name__icontains=query) |
                Q(supplier__name__icontains=query)
            )

        if supplier:
            queryset = queryset.filter(
                Q(supplier__name__iexact=supplier) | Q(supplier_name__iexact=supplier)
            )

        if period:
            today = datetime.date.today()
            start_date = None
            end_date = None
            if period == 'today':
                start_date = today
                end_date = today
            elif period == 'yesterday':
                start_date = today - datetime.timedelta(days=1)
                end_date = start_date
            elif period == 'this_week':
                start_date = today - datetime.timedelta(days=today.weekday())
                end_date = start_date + datetime.timedelta(days=6)
            elif period == 'last_week':
                end_date = today - datetime.timedelta(days=today.weekday() + 1)
                start_date = end_date - datetime.timedelta(days=6)
            elif period == 'this_month':
                start_date = today.replace(day=1)
                end_date = today
            elif period == 'last_month':
                first_day_this_month = today.replace(day=1)
                end_date = first_day_this_month - datetime.timedelta(days=1)
                start_date = end_date.replace(day=1)
            elif period == 'this_year':
                start_date = today.replace(month=1, day=1)
                end_date = today
            elif period == 'last_year':
                start_date = today.replace(year=today.year - 1, month=1, day=1)
                end_date = today.replace(year=today.year - 1, month=12, day=31)
            elif period == '1y':
                start_date = today - relativedelta(years=1)
                end_date = today

            if start_date and end_date:
                queryset = queryset.filter(date__range=(start_date, end_date))
            elif start_date:
                queryset = queryset.filter(date__gte=start_date)
            elif end_date:
                queryset = queryset.filter(date__lte=end_date)

        queryset = queryset.order_by('-date', '-pk')
        self.filter_params = params
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = getattr(self, 'filter_params', None) or self._get_filter_params()
        supplier_options = (
            SupplierCredit.objects.filter(user=self.request.user)
            .exclude(supplier_name__exact='')
            .values_list('supplier_name', flat=True)
            .distinct()
            .order_by('supplier_name')
        )
        context.update({
            'search_query': params['search'],
            'selected_period': params['period'],
            'selected_supplier': params['supplier'],
            'supplier_options': supplier_options,
        })
        return context


class SupplierCreditDetailView(LoginRequiredMixin, DetailView):
    model = SupplierCredit
    template_name = 'app/supplier_credit_detail.html'

    def get_queryset(self):
        return SupplierCredit.objects.filter(user=self.request.user).prefetch_related('items', 'items__source_expense')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        credit = self.object
        total_amount_incl_tax, total_tax, total_amount_excl_tax = credit.calculate_totals()
        context.update({
            'total_amount_incl_tax': total_amount_incl_tax,
            'total_tax': total_tax,
            'total_amount_excl_tax': total_amount_excl_tax,
            'applied_amount': credit.applied_amount,
            'available_amount': credit.available_amount,
        })
        return context


class SupplierCreditCreateView(SupplierCreditInventoryMixin, LoginRequiredMixin, CreateView):
    model = SupplierCredit
    form_class = SupplierCreditForm
    template_name = 'accounts/mach/add_supplier_credit.html'
    success_url = reverse_lazy('accounts:supplier_credit_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_products_qs = Product.objects.filter(user__in=get_product_user_ids(self.request.user)).order_by('name')
        inventory_products_data = {}
        vendor_product_map = defaultdict(list)
        all_product_ids = []
        for product in inventory_products_qs:
            product_id_str = str(product.id)
            supplier_name = (product.supplier.name or '').strip() if product.supplier else ''
            normalized_supplier = supplier_name.lower()
            display_name = f"{product.name} ({product.sku})" if product.sku else product.name
            inventory_products_data[product_id_str] = {
                'name': product.name,
                'sku': product.sku or '',
                'description': product.description or '',
                'cost_price': float(product.cost_price) if product.cost_price is not None else 0.0,
                'supplier': supplier_name,
                'display_name': display_name,
            }
            all_product_ids.append(product_id_str)
            if normalized_supplier:
                vendor_product_map[normalized_supplier].append(product_id_str)
            else:
                vendor_product_map['__unassigned__'].append(product_id_str)

        if self.request.POST:
            context['supplier_credit_item_formset'] = SupplierCreditItemFormSet(
                self.request.POST,
                prefix='credit_items',
                user=self.request.user,
            )
        else:
            context['supplier_credit_item_formset'] = SupplierCreditItemFormSet(
                prefix='credit_items',
                user=self.request.user,
            )
            context['supplier_credit_item_formset'].extra = 1

        context['supplier_credit_form'] = context.get('form')
        vendor_product_map.setdefault('__unassigned__', [])
        context['inventory_products'] = inventory_products_qs
        context['inventory_products_data'] = json.dumps(inventory_products_data)
        context['vendor_product_map'] = json.dumps(dict(vendor_product_map))
        context['all_inventory_product_ids'] = json.dumps(all_product_ids)
        context['supplier_receipts_url'] = reverse('accounts:supplier_credit_receipts')
        context['province_tax_rates'] = json.dumps(PROVINCE_TAX_RATES)
        context['is_edit'] = False
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['supplier_credit_item_formset']
        if not formset.is_valid():
            messages.error(self.request, 'Please correct the errors below in the items.')
            return self.render_to_response(self.get_context_data(form=form))

        self.object = form.save()
        formset.instance = self.object

        items_to_refresh = set()
        for inline_form in formset.forms:
            if not hasattr(inline_form, 'cleaned_data'):
                continue
            if inline_form.cleaned_data.get('DELETE'):
                continue
            item = inline_form.save(commit=False)
            selected_product = inline_form.cleaned_data.get('product')
            create_inventory_product = inline_form.cleaned_data.get('create_inventory_product')
            part_no = inline_form.cleaned_data.get('part_no')
            description = inline_form.cleaned_data.get('description')
            price = inline_form.cleaned_data.get('price')

            price_decimal = Decimal('0.00')
            has_price = False
            if price not in (None, ''):
                try:
                    price_decimal = Decimal(str(price)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    has_price = True
                except (InvalidOperation, TypeError, ValueError):
                    price_decimal = Decimal('0.00')

            product = _ensure_inventory_product_for_item(
                user=self.request.user,
                category_name="Returns",
                existing_product=selected_product,
                create_inventory_product=create_inventory_product,
                part_no=part_no,
                description=description,
                price_decimal=price_decimal,
                has_price=has_price,
                supplier_name=self.object.supplier_name or (self.object.supplier.name if self.object.supplier else ''),
            )
            item.product = product

            if not item.part_no and product and product.sku:
                item.part_no = product.sku
            if not item.description and product:
                item.description = product.description or product.name

            if item.source_expense_item and not item.source_expense:
                item.source_expense = item.source_expense_item.mech_expense
            if item.source_expense_item and item.source_expense_item.mech_expense.user_id != self.request.user.id:
                item.source_expense_item = None
            if item.source_expense and item.source_expense.user_id != self.request.user.id:
                item.source_expense = None
                item.source_expense_item = None
            if item.source_expense_item and item.source_expense and item.source_expense_item.mech_expense_id != item.source_expense_id:
                item.source_expense_item = None
            supplier_name = (self.object.supplier_name or '').strip().lower()
            if supplier_name and item.source_expense and (item.source_expense.vendor or '').strip().lower() != supplier_name:
                item.source_expense = None
                item.source_expense_item = None

            item.supplier_credit = self.object
            item.save()

            if item.source_expense:
                items_to_refresh.add(item.source_expense)

        deleted_instances = []
        if hasattr(formset, 'deleted_objects'):
            deleted_instances = formset.deleted_objects
        elif hasattr(formset, 'deleted_forms'):
            deleted_instances = [
                form.instance
                for form in formset.deleted_forms
                if getattr(form, 'instance', None) and form.instance.pk
            ]

        for deleted in deleted_instances:
            if deleted.source_expense:
                items_to_refresh.add(deleted.source_expense)
            deleted.delete()

        record_inventory = self.object.record_in_inventory
        if record_inventory:
            self._apply_inventory_sync(self.object)

        for expense in items_to_refresh:
            _update_expense_paid_status(expense)

        messages.success(self.request, f'Supplier credit {self.object.credit_no} created successfully.')
        return redirect(self.get_success_url())


class SupplierCreditUpdateView(SupplierCreditInventoryMixin, LoginRequiredMixin, UpdateView):
    model = SupplierCredit
    form_class = SupplierCreditForm
    template_name = 'accounts/mach/add_supplier_credit.html'
    success_url = reverse_lazy('accounts:supplier_credit_list')

    def get_queryset(self):
        return SupplierCredit.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_products_qs = Product.objects.filter(user__in=get_product_user_ids(self.request.user)).order_by('name')
        inventory_products_data = {}
        vendor_product_map = defaultdict(list)
        all_product_ids = []
        for product in inventory_products_qs:
            product_id_str = str(product.id)
            supplier_name = (product.supplier.name or '').strip() if product.supplier else ''
            normalized_supplier = supplier_name.lower()
            display_name = f"{product.name} ({product.sku})" if product.sku else product.name
            inventory_products_data[product_id_str] = {
                'name': product.name,
                'sku': product.sku or '',
                'description': product.description or '',
                'cost_price': float(product.cost_price) if product.cost_price is not None else 0.0,
                'supplier': supplier_name,
                'display_name': display_name,
            }
            all_product_ids.append(product_id_str)
            if normalized_supplier:
                vendor_product_map[normalized_supplier].append(product_id_str)
            else:
                vendor_product_map['__unassigned__'].append(product_id_str)

        if self.request.POST:
            context['supplier_credit_item_formset'] = SupplierCreditItemFormSet(
                self.request.POST,
                instance=self.object,
                prefix='credit_items',
                user=self.request.user,
            )
        else:
            context['supplier_credit_item_formset'] = SupplierCreditItemFormSet(
                instance=self.object,
                prefix='credit_items',
                user=self.request.user,
            )

        context['supplier_credit_form'] = context.get('form')
        vendor_product_map.setdefault('__unassigned__', [])
        context['inventory_products'] = inventory_products_qs
        context['inventory_products_data'] = json.dumps(inventory_products_data)
        context['vendor_product_map'] = json.dumps(dict(vendor_product_map))
        context['all_inventory_product_ids'] = json.dumps(all_product_ids)
        context['supplier_receipts_url'] = reverse('accounts:supplier_credit_receipts')
        context['province_tax_rates'] = json.dumps(PROVINCE_TAX_RATES)
        context['is_edit'] = True
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['supplier_credit_item_formset']
        previous_items = list(self.object.items.all())
        was_recorded = self.object.record_in_inventory
        affected_expenses = {item.source_expense for item in previous_items if item.source_expense}

        if not formset.is_valid():
            messages.error(self.request, 'Please correct the errors below in the items.')
            return self.render_to_response(self.get_context_data(form=form))

        self.object = form.save()
        formset.instance = self.object

        for inline_form in formset.forms:
            if not hasattr(inline_form, 'cleaned_data'):
                continue
            if inline_form.cleaned_data.get('DELETE'):
                continue
            item = inline_form.save(commit=False)
            selected_product = inline_form.cleaned_data.get('product')
            create_inventory_product = inline_form.cleaned_data.get('create_inventory_product')
            part_no = inline_form.cleaned_data.get('part_no')
            description = inline_form.cleaned_data.get('description')
            price = inline_form.cleaned_data.get('price')

            price_decimal = Decimal('0.00')
            has_price = False
            if price not in (None, ''):
                try:
                    price_decimal = Decimal(str(price)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    has_price = True
                except (InvalidOperation, TypeError, ValueError):
                    price_decimal = Decimal('0.00')

            product = _ensure_inventory_product_for_item(
                user=self.request.user,
                category_name="Returns",
                existing_product=selected_product,
                create_inventory_product=create_inventory_product,
                part_no=part_no,
                description=description,
                price_decimal=price_decimal,
                has_price=has_price,
                supplier_name=self.object.supplier_name or (self.object.supplier.name if self.object.supplier else ''),
            )
            item.product = product

            if not item.part_no and product and product.sku:
                item.part_no = product.sku
            if not item.description and product:
                item.description = product.description or product.name

            if item.source_expense_item and not item.source_expense:
                item.source_expense = item.source_expense_item.mech_expense
            if item.source_expense_item and item.source_expense_item.mech_expense.user_id != self.request.user.id:
                item.source_expense_item = None
            if item.source_expense and item.source_expense.user_id != self.request.user.id:
                item.source_expense = None
                item.source_expense_item = None
            if item.source_expense_item and item.source_expense and item.source_expense_item.mech_expense_id != item.source_expense_id:
                item.source_expense_item = None
            supplier_name = (self.object.supplier_name or '').strip().lower()
            if supplier_name and item.source_expense and (item.source_expense.vendor or '').strip().lower() != supplier_name:
                item.source_expense = None
                item.source_expense_item = None

            item.supplier_credit = self.object
            item.save()

            if item.source_expense:
                affected_expenses.add(item.source_expense)

        deleted_instances = []
        if hasattr(formset, 'deleted_objects'):
            deleted_instances = formset.deleted_objects
        elif hasattr(formset, 'deleted_forms'):
            deleted_instances = [
                form.instance
                for form in formset.deleted_forms
                if getattr(form, 'instance', None) and form.instance.pk
            ]

        for deleted in deleted_instances:
            if deleted.source_expense:
                affected_expenses.add(deleted.source_expense)
            deleted.delete()

        record_inventory = self.object.record_in_inventory
        if record_inventory:
            self._apply_inventory_sync(
                self.object,
                previous_items=previous_items if was_recorded else [],
            )
        elif was_recorded:
            self._log_transactions(self.object, previous_items, 'IN', 'credit edit reversal')

        for expense in affected_expenses:
            _update_expense_paid_status(expense)

        messages.success(self.request, f'Supplier credit {self.object.credit_no} updated successfully.')
        return redirect(self.get_success_url())


class SupplierCreditDeleteView(SupplierCreditInventoryMixin, LoginRequiredMixin, DeleteView):
    model = SupplierCredit
    template_name = 'app/supplier_credit_confirm_delete.html'
    success_url = reverse_lazy('accounts:supplier_credit_list')

    def get_queryset(self):
        return SupplierCredit.objects.filter(user=self.request.user)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        previous_items = list(self.object.items.all())
        affected_expenses = {item.source_expense for item in previous_items if item.source_expense}
        record_inventory = self.object.record_in_inventory
        response = super().delete(request, *args, **kwargs)
        if record_inventory:
            try:
                self._log_transactions(self.object, previous_items, 'IN', 'credit deleted')
            except Exception:
                logger.warning("Inventory sync on delete failed for credit %s", getattr(self.object, 'id', None))
        for expense in affected_expenses:
            _update_expense_paid_status(expense)
        messages.success(request, f'Supplier credit {self.object.credit_no} deleted successfully.')
        return response


class CustomerCreditListView(LoginRequiredMixin, ListView):
    model = CustomerCredit
    template_name = 'app/customer_credit_list.html'
    context_object_name = 'object_list'
    paginate_by = 100

    def _get_filter_params(self):
        period_value = self.request.GET.get('period')
        if period_value is None:
            period_value = '1y'
        return {
            'search': (self.request.GET.get('search') or '').strip(),
            'period': (period_value or '').strip().lower(),
            'customer': (self.request.GET.get('customer') or '').strip(),
        }

    def get_queryset(self):
        params = self._get_filter_params()
        query = params['search']
        period = params['period']
        customer = params['customer']
        user = self.request.user

        items_queryset = (
            CustomerCreditItem.objects.filter(customer_credit=OuterRef('pk'))
            .values('customer_credit')
        )
        items_subtotal = items_queryset.annotate(
            total=Coalesce(Sum('amount'), Value(0.0))
        ).values('total')
        items_tax = items_queryset.annotate(
            total=Coalesce(Sum('tax_paid'), Value(0.0))
        ).values('total')
        applied_amount_expr = Case(
            When(customer_credit__tax_included=True, then=F('amount')),
            default=ExpressionWrapper(F('amount') + F('tax_paid'), output_field=FloatField()),
            output_field=FloatField(),
        )
        items_applied = items_queryset.annotate(
            total=Coalesce(
                Sum(applied_amount_expr, filter=Q(source_invoice__isnull=False)),
                Value(0.0),
            )
        ).values('total')

        queryset = (
            CustomerCredit.objects.filter(user=user)
            .annotate(
                items_subtotal=Coalesce(Subquery(items_subtotal, output_field=FloatField()), Value(0.0)),
                items_tax=Coalesce(Subquery(items_tax, output_field=FloatField()), Value(0.0)),
                applied_total=Coalesce(Subquery(items_applied, output_field=FloatField()), Value(0.0)),
            )
            .annotate(
                total_incl_tax=Case(
                    When(tax_included=True, then=F('items_subtotal')),
                    default=F('items_subtotal') + F('items_tax'),
                    output_field=FloatField(),
                ),
                available_total=ExpressionWrapper(
                    F('total_incl_tax') - F('applied_total'),
                    output_field=FloatField(),
                ),
            )
            .select_related('customer')
            .prefetch_related('items')
        )

        if query:
            queryset = queryset.filter(
                Q(credit_no__icontains=query) |
                Q(customer_name__icontains=query) |
                Q(customer__name__icontains=query)
            )

        if customer:
            queryset = queryset.filter(
                Q(customer__name__iexact=customer) | Q(customer_name__iexact=customer)
            )

        if period:
            today = datetime.date.today()
            start_date = None
            end_date = None
            if period == 'today':
                start_date = today
                end_date = today
            elif period == 'yesterday':
                start_date = today - datetime.timedelta(days=1)
                end_date = start_date
            elif period == 'this_week':
                start_date = today - datetime.timedelta(days=today.weekday())
                end_date = start_date + datetime.timedelta(days=6)
            elif period == 'last_week':
                end_date = today - datetime.timedelta(days=today.weekday() + 1)
                start_date = end_date - datetime.timedelta(days=6)
            elif period == 'this_month':
                start_date = today.replace(day=1)
                end_date = today
            elif period == 'last_month':
                first_day_this_month = today.replace(day=1)
                end_date = first_day_this_month - datetime.timedelta(days=1)
                start_date = end_date.replace(day=1)
            elif period == 'this_year':
                start_date = today.replace(month=1, day=1)
                end_date = today
            elif period == 'last_year':
                start_date = today.replace(year=today.year - 1, month=1, day=1)
                end_date = today.replace(year=today.year - 1, month=12, day=31)
            elif period == '1y':
                start_date = today - relativedelta(years=1)
                end_date = today

            if start_date and end_date:
                queryset = queryset.filter(date__range=(start_date, end_date))
            elif start_date:
                queryset = queryset.filter(date__gte=start_date)
            elif end_date:
                queryset = queryset.filter(date__lte=end_date)

        queryset = queryset.order_by('-date', '-pk')
        self.filter_params = params
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = getattr(self, 'filter_params', None) or self._get_filter_params()
        customer_options = (
            CustomerCredit.objects.filter(user=self.request.user)
            .exclude(customer_name__exact='')
            .values_list('customer_name', flat=True)
            .distinct()
            .order_by('customer_name')
        )
        context.update({
            'search_query': params['search'],
            'selected_period': params['period'],
            'selected_customer': params['customer'],
            'customer_options': customer_options,
        })
        return context


class CustomerCreditDetailView(LoginRequiredMixin, DetailView):
    model = CustomerCredit
    template_name = 'app/customer_credit_detail.html'

    def get_queryset(self):
        return CustomerCredit.objects.filter(user=self.request.user).prefetch_related(
            'items',
            'items__source_invoice',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        credit = self.object
        total_amount_incl_tax, total_tax, total_amount_excl_tax = credit.calculate_totals()
        context.update({
            'total_amount_incl_tax': total_amount_incl_tax,
            'total_tax': total_tax,
            'total_amount_excl_tax': total_amount_excl_tax,
            'applied_amount': credit.applied_amount,
            'available_amount': credit.available_amount,
        })
        return context


def _build_customer_credit_context(credit, *, request=None):
    profile = getattr(credit.user, 'profile', None)
    total_amount_incl_tax, total_tax, total_amount_excl_tax = credit.calculate_totals()
    total_amount_incl_tax = Decimal(str(total_amount_incl_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_tax = Decimal(str(total_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_amount_excl_tax = Decimal(str(total_amount_excl_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    company_logo_url = resolve_company_logo_url(profile, request=request)
    company_logo_url_pdf = resolve_company_logo_url(profile, for_pdf=True)

    context = {
        'credit': credit,
        'profile': profile,
        'customer': credit.customer,
        'customer_name': credit.customer_name or (credit.customer.name if credit.customer else ''),
        'items': list(credit.items.select_related('product', 'source_invoice').all()),
        'total_amount_incl_tax': total_amount_incl_tax,
        'total_tax': total_tax,
        'total_amount_excl_tax': total_amount_excl_tax,
        'company_logo_url': company_logo_url,
        'company_logo_url_pdf': company_logo_url_pdf,
        'generated_on': timezone.localdate(),
    }
    return apply_branding_defaults(context)


@login_required
def customer_credit_pdf(request, pk):
    credit = get_object_or_404(CustomerCredit, pk=pk, user=request.user)
    context = _build_customer_credit_context(credit, request=request)
    pdf_bytes = render_template_to_pdf('invoices/customer_credit_pdf.html', context)
    filename = f"Customer_Credit_{credit.credit_no or credit.pk}.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def customer_credit_print(request, pk):
    credit = get_object_or_404(CustomerCredit, pk=pk, user=request.user)
    context = _build_customer_credit_context(credit, request=request)
    pdf_bytes = render_template_to_pdf('invoices/customer_credit_pdf.html', context)
    filename = f"Customer_Credit_{credit.credit_no or credit.pk}.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
@require_POST
def customer_credit_email(request, pk):
    credit = get_object_or_404(CustomerCredit, pk=pk, user=request.user)
    customer = credit.customer
    customer_email = (customer.email if customer else '') or ''
    if not customer_email:
        messages.error(request, "Customer email is missing. Update the customer record and try again.")
        return redirect('accounts:customer_credit_detail', pk=credit.pk)

    context = _build_customer_credit_context(credit, request=request)
    pdf_bytes = render_template_to_pdf('invoices/customer_credit_pdf.html', context)
    email_html = render_to_string('emails/customer_credit_email.html', context)

    profile = getattr(credit.user, 'profile', None)
    company_name = context.get('business_name') or (getattr(profile, 'company_name', '') if profile else '')
    subject = f"Customer Credit Memo #{credit.credit_no or credit.pk}"
    if company_name:
        subject = f"{subject} from {company_name}"

    customer_cc_emails = customer.get_cc_emails() if customer else []
    cc_recipients = build_cc_list(
        getattr(request.user, 'email', None),
        getattr(profile, "company_email", None),
        *customer_cc_emails,
        exclude=[customer_email],
    )

    email_message = EmailMultiAlternatives(
        subject=subject,
        body="Please find your customer credit memo attached.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[customer_email],
        cc=cc_recipients or None,
    )
    email_message.attach_alternative(email_html, "text/html")
    email_message.attach(
        f"Customer_Credit_{credit.credit_no or credit.pk}.pdf",
        pdf_bytes,
        'application/pdf',
    )
    email_message.send()

    messages.success(request, f"Customer credit memo sent to {customer_email}.")
    return redirect('accounts:customer_credit_detail', pk=credit.pk)


class CustomerCreditCreateView(CustomerCreditInventoryMixin, LoginRequiredMixin, CreateView):
    model = CustomerCredit
    form_class = CustomerCreditForm
    template_name = 'accounts/mach/add_customer_credit.html'
    success_url = reverse_lazy('accounts:customer_credit_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_products_qs = Product.objects.filter(user__in=get_product_user_ids(self.request.user)).order_by('name')
        inventory_products_data = {}
        for product in inventory_products_qs:
            product_id_str = str(product.id)
            display_name = f"{product.name} ({product.sku})" if product.sku else product.name
            inventory_products_data[product_id_str] = {
                'name': product.name,
                'sku': product.sku or '',
                'description': product.description or '',
                'cost_price': float(product.cost_price) if product.cost_price is not None else 0.0,
                'sale_price': float(product.sale_price) if product.sale_price is not None else 0.0,
                'display_name': display_name,
            }

        if self.request.POST:
            context['customer_credit_item_formset'] = CustomerCreditItemFormSet(
                self.request.POST,
                prefix='credit_items',
                user=self.request.user,
            )
        else:
            context['customer_credit_item_formset'] = CustomerCreditItemFormSet(
                prefix='credit_items',
                user=self.request.user,
            )
            context['customer_credit_item_formset'].extra = 1

        context['customer_credit_form'] = context.get('form')
        context['inventory_products_data'] = json.dumps(inventory_products_data)
        context['customer_invoices_url'] = reverse('accounts:customer_credit_invoices')
        context['province_tax_rates'] = json.dumps(PROVINCE_TAX_RATES)
        context['is_edit'] = False
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['customer_credit_item_formset']

        if not formset.is_valid():
            messages.error(self.request, 'Please correct the errors below in the items.')
            return self.render_to_response(self.get_context_data(form=form))

        self.object = form.save()
        formset.instance = self.object

        for inline_form in formset.forms:
            if not hasattr(inline_form, 'cleaned_data'):
                continue
            if inline_form.cleaned_data.get('DELETE'):
                continue
            item = inline_form.save(commit=False)
            selected_product = inline_form.cleaned_data.get('product')
            create_inventory_product = inline_form.cleaned_data.get('create_inventory_product')
            part_no = inline_form.cleaned_data.get('part_no')
            description = inline_form.cleaned_data.get('description')
            price = inline_form.cleaned_data.get('price')

            price_decimal = Decimal('0.00')
            has_price = False
            if price not in (None, ''):
                try:
                    price_decimal = Decimal(str(price)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    has_price = True
                except (InvalidOperation, TypeError, ValueError):
                    price_decimal = Decimal('0.00')

            product = _ensure_inventory_product_for_item(
                user=self.request.user,
                category_name="Returns",
                existing_product=selected_product,
                create_inventory_product=create_inventory_product,
                part_no=part_no,
                description=description,
                price_decimal=price_decimal,
                has_price=has_price,
                supplier_name='',
            )
            item.product = product

            if not item.part_no and product and product.sku:
                item.part_no = product.sku
            if not item.description and product:
                item.description = product.description or product.name

            if item.source_invoice_item and not item.source_invoice:
                item.source_invoice = item.source_invoice_item.grouped_invoice
            if item.source_invoice_item and item.source_invoice_item.grouped_invoice.user_id != self.request.user.id:
                item.source_invoice_item = None
            if item.source_invoice and item.source_invoice.user_id != self.request.user.id:
                item.source_invoice = None
                item.source_invoice_item = None
            if item.source_invoice_item and item.source_invoice and item.source_invoice_item.grouped_invoice_id != item.source_invoice_id:
                item.source_invoice_item = None

            customer_name = (self.object.customer_name or '').strip().lower()
            if customer_name and item.source_invoice:
                invoice_customer_name = ''
                if item.source_invoice.customer:
                    invoice_customer_name = (item.source_invoice.customer.name or '').strip().lower()
                else:
                    invoice_customer_name = (item.source_invoice.bill_to or '').strip().lower()
                if invoice_customer_name and invoice_customer_name != customer_name:
                    item.source_invoice = None
                    item.source_invoice_item = None

            item.customer_credit = self.object
            item.save()

        deleted_instances = []
        if hasattr(formset, 'deleted_objects'):
            deleted_instances = formset.deleted_objects
        elif hasattr(formset, 'deleted_forms'):
            deleted_instances = [
                form.instance
                for form in formset.deleted_forms
                if getattr(form, 'instance', None) and form.instance.pk
            ]

        for deleted in deleted_instances:
            deleted.delete()

        record_inventory = self.object.record_in_inventory
        if record_inventory:
            self._apply_inventory_sync(self.object)

        messages.success(self.request, f'Customer credit {self.object.credit_no} created successfully.')
        return redirect(self.get_success_url())


class CustomerCreditUpdateView(CustomerCreditInventoryMixin, LoginRequiredMixin, UpdateView):
    model = CustomerCredit
    form_class = CustomerCreditForm
    template_name = 'accounts/mach/add_customer_credit.html'
    success_url = reverse_lazy('accounts:customer_credit_list')

    def get_queryset(self):
        return CustomerCredit.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_products_qs = Product.objects.filter(user__in=get_product_user_ids(self.request.user)).order_by('name')
        inventory_products_data = {}
        for product in inventory_products_qs:
            product_id_str = str(product.id)
            display_name = f"{product.name} ({product.sku})" if product.sku else product.name
            inventory_products_data[product_id_str] = {
                'name': product.name,
                'sku': product.sku or '',
                'description': product.description or '',
                'cost_price': float(product.cost_price) if product.cost_price is not None else 0.0,
                'sale_price': float(product.sale_price) if product.sale_price is not None else 0.0,
                'display_name': display_name,
            }

        if self.request.POST:
            context['customer_credit_item_formset'] = CustomerCreditItemFormSet(
                self.request.POST,
                instance=self.object,
                prefix='credit_items',
                user=self.request.user,
            )
        else:
            context['customer_credit_item_formset'] = CustomerCreditItemFormSet(
                instance=self.object,
                prefix='credit_items',
                user=self.request.user,
            )

        context['customer_credit_form'] = context.get('form')
        context['inventory_products_data'] = json.dumps(inventory_products_data)
        context['customer_invoices_url'] = reverse('accounts:customer_credit_invoices')
        context['province_tax_rates'] = json.dumps(PROVINCE_TAX_RATES)
        context['is_edit'] = True
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['customer_credit_item_formset']
        previous_items = list(self.object.items.all())
        was_recorded = self.object.record_in_inventory
        if not formset.is_valid():
            messages.error(self.request, 'Please correct the errors below in the items.')
            return self.render_to_response(self.get_context_data(form=form))

        self.object = form.save()
        formset.instance = self.object

        for inline_form in formset.forms:
            if not hasattr(inline_form, 'cleaned_data'):
                continue
            if inline_form.cleaned_data.get('DELETE'):
                continue
            item = inline_form.save(commit=False)
            selected_product = inline_form.cleaned_data.get('product')
            create_inventory_product = inline_form.cleaned_data.get('create_inventory_product')
            part_no = inline_form.cleaned_data.get('part_no')
            description = inline_form.cleaned_data.get('description')
            price = inline_form.cleaned_data.get('price')

            price_decimal = Decimal('0.00')
            has_price = False
            if price not in (None, ''):
                try:
                    price_decimal = Decimal(str(price)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    has_price = True
                except (InvalidOperation, TypeError, ValueError):
                    price_decimal = Decimal('0.00')

            product = _ensure_inventory_product_for_item(
                user=self.request.user,
                category_name="Returns",
                existing_product=selected_product,
                create_inventory_product=create_inventory_product,
                part_no=part_no,
                description=description,
                price_decimal=price_decimal,
                has_price=has_price,
                supplier_name='',
            )
            item.product = product

            if not item.part_no and product and product.sku:
                item.part_no = product.sku
            if not item.description and product:
                item.description = product.description or product.name

            if item.source_invoice_item and not item.source_invoice:
                item.source_invoice = item.source_invoice_item.grouped_invoice
            if item.source_invoice_item and item.source_invoice_item.grouped_invoice.user_id != self.request.user.id:
                item.source_invoice_item = None
            if item.source_invoice and item.source_invoice.user_id != self.request.user.id:
                item.source_invoice = None
                item.source_invoice_item = None
            if item.source_invoice_item and item.source_invoice and item.source_invoice_item.grouped_invoice_id != item.source_invoice_id:
                item.source_invoice_item = None

            customer_name = (self.object.customer_name or '').strip().lower()
            if customer_name and item.source_invoice:
                invoice_customer_name = ''
                if item.source_invoice.customer:
                    invoice_customer_name = (item.source_invoice.customer.name or '').strip().lower()
                else:
                    invoice_customer_name = (item.source_invoice.bill_to or '').strip().lower()
                if invoice_customer_name and invoice_customer_name != customer_name:
                    item.source_invoice = None
                    item.source_invoice_item = None

            item.customer_credit = self.object
            item.save()

        deleted_instances = []
        if hasattr(formset, 'deleted_objects'):
            deleted_instances = formset.deleted_objects
        elif hasattr(formset, 'deleted_forms'):
            deleted_instances = [
                form.instance
                for form in formset.deleted_forms
                if getattr(form, 'instance', None) and form.instance.pk
            ]

        for deleted in deleted_instances:
            deleted.delete()

        record_inventory = self.object.record_in_inventory
        if record_inventory:
            self._apply_inventory_sync(
                self.object,
                previous_items=previous_items if was_recorded else [],
            )
        elif was_recorded:
            self._log_transactions(self.object, previous_items, 'OUT', 'credit edit reversal')

        messages.success(self.request, f'Customer credit {self.object.credit_no} updated successfully.')
        return redirect(self.get_success_url())


class CustomerCreditDeleteView(CustomerCreditInventoryMixin, LoginRequiredMixin, DeleteView):
    model = CustomerCredit
    template_name = 'app/customer_credit_confirm_delete.html'
    success_url = reverse_lazy('accounts:customer_credit_list')

    def get_queryset(self):
        return CustomerCredit.objects.filter(user=self.request.user)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Customer credit {self.object.credit_no} deleted successfully.')
        return response


@login_required
def customer_credit_invoices(request):
    customer_id = request.GET.get('customer_id')
    if not customer_id:
        return JsonResponse({'invoices': []})
    customer_user_ids = get_customer_user_ids(request.user)
    store_user_ids = get_business_user_ids(request.user)
    customer = get_object_or_404(Customer, pk=customer_id, user__in=customer_user_ids)
    invoices = (
        GroupedInvoice.objects.filter(user__in=store_user_ids, customer=customer)
        .prefetch_related('income_records', 'income_records__product')
        .order_by('-date', '-id')
    )
    invoice_payload = []
    for invoice in invoices:
        items_payload = []
        for item in invoice.income_records.all():
            product = item.product
            part_no = product.sku if product and product.sku else ''
            description = item.job or (product.description if product else '') or (product.name if product else '')
            items_payload.append({
                'id': item.id,
                'part_no': part_no,
                'description': description or '',
                'qty': float(item.qty or 0),
                'price': float(item.rate or 0),
                'amount': float(item.amount or 0),
                'product_id': str(product.id) if product else '',
                'invoice_id': invoice.id,
            })
        invoice_payload.append({
            'id': invoice.id,
            'invoice_number': invoice.invoice_number or f"Invoice {invoice.id}",
            'date': invoice.date.isoformat() if invoice.date else '',
            'items': items_payload,
        })
    return JsonResponse({'invoices': invoice_payload})


@login_required
def supplier_credit_receipts(request):
    supplier_id = request.GET.get('supplier_id')
    if not supplier_id:
        return JsonResponse({'receipts': []})
    supplier = get_object_or_404(Supplier, pk=supplier_id, user=request.user)
    supplier_name = (supplier.name or '').strip()
    receipts = (
        MechExpense.objects.filter(user=request.user, vendor__iexact=supplier_name)
        .prefetch_related('mechexpenseitem_set')
        .order_by('-date', '-id')
    )
    receipt_payload = []
    for receipt in receipts:
        items_payload = []
        for item in receipt.mechexpenseitem_set.all():
            product = None
            part = (item.part_no or '').strip()
            desc = (item.description or '').strip()
            if part:
                product = Product.objects.filter(
                    user=request.user,
                    sku__iexact=part,
                    supplier=supplier,
                ).first()
            if not product and desc:
                product = Product.objects.filter(
                    user=request.user,
                    name__iexact=desc,
                    supplier=supplier,
                ).first()
            items_payload.append({
                'id': item.id,
                'part_no': item.part_no or '',
                'description': item.description or '',
                'qty': float(item.qty or 0),
                'price': float(item.price or 0),
                'amount': float(item.amount or 0),
                'product_id': str(product.id) if product else '',
                'expense_id': receipt.id,
            })
        total_amount_incl_tax, _, _ = receipt.calculate_totals()
        receipt_payload.append({
            'id': receipt.id,
            'receipt_no': receipt.receipt_no,
            'date': receipt.date.isoformat() if receipt.date else '',
            'total': float(total_amount_incl_tax or 0),
            'items': items_payload,
        })
    return JsonResponse({'receipts': receipt_payload})

@login_required
def supplier_cheque_list(request):
    query = (request.GET.get('q', '') or '').strip()
    date_range_key, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )
    date_range_key, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )
    date_range_key, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )
    date_range_key, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )
    date_range_key, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )
    cheques = SupplierCheque.objects.filter(user=request.user).select_related('supplier')
    if query:
        cheques = cheques.filter(
            Q(cheque_number__icontains=query) |
            Q(supplier_name__icontains=query) |
            Q(supplier__name__icontains=query)
        )
    cheques = cheques.order_by('-date', '-id').prefetch_related('lines')
    return render(request, 'app/supplier_cheque_list.html', {
        'cheques': cheques,
        'query': query,
    })


@login_required
def supplier_cheque_detail(request, pk):
    cheque = get_object_or_404(
        SupplierCheque.objects.filter(user=request.user).select_related('supplier'),
        pk=pk,
    )
    lines = cheque.lines.select_related('mech_expense')
    return render(request, 'app/supplier_cheque_detail.html', {
        'cheque': cheque,
        'lines': lines,
    })


@login_required
def supplier_cheque_create(request):
    if request.method == 'POST':
        form = SupplierChequeForm(request.POST, user=request.user)
        if form.is_valid():
            cheque = form.save(commit=False)
            cheque.user = request.user
            cheque.save()

            selected_ids = request.POST.getlist('selected_expenses')
            total_amount_raw = request.POST.get('total_amount')
            if not selected_ids:
                total_amount = None
                if total_amount_raw not in (None, ''):
                    try:
                        total_amount = Decimal(str(total_amount_raw)).quantize(
                            Decimal('0.01'),
                            rounding=ROUND_HALF_UP,
                        )
                    except (InvalidOperation, TypeError):
                        total_amount = None

                if not total_amount or total_amount <= 0:
                    messages.error(
                        request,
                        'Please select expenses or enter a total amount to pay.',
                    )
                    cheque.delete()
                    return redirect('accounts:supplier_cheque_add')

                supplier = cheque.supplier
                supplier_name = (supplier.name or '').strip() if supplier else ''
                if not supplier_name:
                    messages.error(request, 'Please select a supplier before applying a cheque amount.')
                    cheque.delete()
                    return redirect('accounts:supplier_cheque_add')

                memo_text = (cheque.memo or '').strip()
                notes = f"Cheque #{cheque.cheque_number}"
                if memo_text:
                    notes = f"{notes} - {memo_text}"

                applied_total, _ = _allocate_supplier_payment(
                    user=request.user,
                    supplier_name=supplier_name,
                    total_amount=total_amount,
                    method='Cheque',
                    notes=notes,
                    cheque=cheque,
                )
                if applied_total <= 0:
                    cheque.delete()
                    messages.error(request, 'No outstanding expenses were found for this supplier.')
                    return redirect('accounts:supplier_cheque_add')
                if applied_total < total_amount:
                    messages.info(
                        request,
                        f'Only ${applied_total:,.2f} was applied because the supplier has a smaller outstanding balance.',
                    )
                messages.success(
                    request,
                    f'Cheque {cheque.cheque_number} created for ${applied_total:,.2f}.',
                )
                return redirect('accounts:supplier_cheque_detail', pk=cheque.pk)

            supplier = cheque.supplier
            supplier_name = (supplier.name or '').strip() if supplier else ''
            expenses = (
                MechExpense.objects.filter(user=request.user, id__in=selected_ids)
                .prefetch_related('mechexpenseitem_set', 'payments', 'supplier_credit_items')
            )
            expenses_by_id = {str(expense.id): expense for expense in expenses}

            lines_created = []
            total_paid = Decimal('0.00')
            for expense_id in selected_ids:
                expense = expenses_by_id.get(str(expense_id))
                if not expense:
                    continue
                if supplier_name and (expense.vendor or '').strip().lower() != supplier_name.lower():
                    continue

                amount_raw = request.POST.get(f'amount_{expense_id}', '')
                try:
                    amount = Decimal(str(amount_raw)).quantize(Decimal('0.01'))
                except (InvalidOperation, TypeError):
                    continue

                if amount <= 0:
                    continue

                remaining = expense.remaining_balance
                if amount > remaining:
                    amount = remaining

                if amount <= 0:
                    continue

                SupplierChequeLine.objects.create(
                    cheque=cheque,
                    mech_expense=expense,
                    amount=amount,
                )
                MechExpensePayment.objects.create(
                    mech_expense=expense,
                    cheque=cheque,
                    amount=amount,
                    method='Cheque',
                    notes=f"Cheque #{cheque.cheque_number}",
                    recorded_by=request.user,
                )
                total_paid += amount
                _update_expense_paid_status(expense)
                lines_created.append(expense_id)

            if not lines_created:
                cheque.delete()
                messages.error(request, 'No valid expenses were selected to pay.')
                return redirect('accounts:supplier_cheque_add')

            messages.success(
                request,
                f'Cheque {cheque.cheque_number} created for ${total_paid:,.2f}.',
            )
            return redirect('accounts:supplier_cheque_detail', pk=cheque.pk)
    else:
        form = SupplierChequeForm(user=request.user)

    return render(request, 'app/supplier_cheque_form.html', {
        'form': form,
        'expenses_url': reverse('accounts:supplier_cheque_expenses'),
    })


@login_required
def supplier_cheque_expenses(request):
    supplier_id = request.GET.get('supplier_id')
    if not supplier_id:
        return JsonResponse({'expenses': []})
    supplier = get_object_or_404(Supplier, pk=supplier_id, user=request.user)
    supplier_name = (supplier.name or '').strip()
    expenses = (
        MechExpense.objects.filter(user=request.user, vendor__iexact=supplier_name)
        .prefetch_related('mechexpenseitem_set', 'payments', 'supplier_credit_items')
        .order_by('-date', '-id')
    )
    payload = []
    for expense in expenses:
        remaining = expense.remaining_balance
        if remaining <= 0:
            continue
        total_amount_incl_tax, _, _ = expense.calculate_totals()
        payload.append({
            'id': expense.id,
            'receipt_no': expense.receipt_no,
            'date': expense.date.isoformat() if expense.date else '',
            'total': float(total_amount_incl_tax or 0),
            'remaining': float(remaining),
        })
    return JsonResponse({'expenses': payload})


@login_required
def supplier_cheque_pdf(request, pk):
    cheque = get_object_or_404(
        SupplierCheque.objects.filter(user=request.user).select_related('supplier'),
        pk=pk,
    )
    lines = cheque.lines.select_related('mech_expense')
    context = {
        'cheque': cheque,
        'lines': lines,
        'profile': getattr(request.user, 'profile', None),
    }
    pdf_data = render_template_to_pdf('cheques/supplier_cheque_pdf.html', context)
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=\"Cheque_{cheque.cheque_number}.pdf\"'
    return response


@login_required
@require_POST
def supplier_make_payment(request):
    next_url = request.POST.get('next') or reverse('accounts:supplier_list')
    supplier_id = request.POST.get('supplier_id')
    supplier_name_raw = (request.POST.get('supplier_name') or '').strip()

    supplier = None
    if supplier_id:
        supplier = Supplier.objects.filter(user=request.user, id=supplier_id).first()
    if not supplier and supplier_name_raw:
        supplier = _get_or_create_supplier_for_user(request.user, supplier_name_raw)

    if not supplier:
        messages.error(request, 'Please select a valid supplier.')
        return redirect(next_url)

    supplier_name = (supplier.name or supplier_name_raw).strip()

    amount_raw = request.POST.get('amount') or request.POST.get('payment_amount')
    try:
        amount = Decimal(str(amount_raw)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError):
        amount = Decimal('0.00')

    if amount <= 0:
        messages.error(request, 'Please enter a valid payment amount.')
        return redirect(next_url)

    payment_type = (request.POST.get('payment_type') or 'other').strip().lower()

    if payment_type == 'cheque':
        cheque_number = (request.POST.get('cheque_number') or '').strip()
        if not cheque_number:
            messages.error(request, 'Cheque number is required.')
            return redirect(next_url)

        bank_account = (request.POST.get('bank_account') or '').strip()
        memo_text = (request.POST.get('memo') or '').strip()
        date_raw = request.POST.get('cheque_date') or request.POST.get('date')
        cheque_date = timezone.localdate()
        if date_raw:
            try:
                cheque_date = datetime.datetime.strptime(date_raw, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                messages.error(request, 'Please provide a valid cheque date (YYYY-MM-DD).')
                return redirect(next_url)

        cheque = SupplierCheque.objects.create(
            user=request.user,
            supplier=supplier,
            cheque_number=cheque_number,
            bank_account=bank_account,
            date=cheque_date,
            memo=memo_text or None,
        )

        notes = f"Cheque #{cheque.cheque_number}"
        if memo_text:
            notes = f"{notes} - {memo_text}"

        applied_total, _ = _allocate_supplier_payment(
            user=request.user,
            supplier_name=supplier_name,
            total_amount=amount,
            method='Cheque',
            notes=notes,
            cheque=cheque,
        )
        if applied_total <= 0:
            cheque.delete()
            messages.error(request, 'No outstanding expenses were found for this supplier.')
            return redirect(next_url)
        if applied_total < amount:
            messages.info(
                request,
                f'Only ${applied_total:,.2f} was applied because the supplier has a smaller outstanding balance.',
            )
        messages.success(
            request,
            f'Cheque {cheque.cheque_number} created for ${applied_total:,.2f}.',
        )
        return redirect(next_url)

    payment_method = (request.POST.get('payment_method') or request.POST.get('method') or 'Cash').strip() or 'Cash'
    payment_notes = (request.POST.get('payment_notes') or request.POST.get('notes') or '').strip() or None

    applied_total, _ = _allocate_supplier_payment(
        user=request.user,
        supplier_name=supplier_name,
        total_amount=amount,
        method=payment_method,
        notes=payment_notes,
    )
    if applied_total <= 0:
        messages.error(request, 'No outstanding expenses were found for this supplier.')
        return redirect(next_url)
    if applied_total < amount:
        messages.info(
            request,
            f'Only ${applied_total:,.2f} was applied because the supplier has a smaller outstanding balance.',
        )
    messages.success(
        request,
        f'Recorded ${applied_total:,.2f} payment for {supplier_name}.',
    )
    return redirect(next_url)


@login_required
@require_POST
def bank_account_create(request):
    name = (request.POST.get('name') or '').strip()
    account_number = (request.POST.get('account_number') or '').strip()
    if not name:
        return JsonResponse({'status': 'error', 'message': 'Bank account name is required.'}, status=400)

    account, created = BusinessBankAccount.objects.get_or_create(
        user=request.user,
        name=name,
        account_number=account_number,
        defaults={'is_active': True},
    )
    if not created and not account.is_active:
        account.is_active = True
        account.save(update_fields=['is_active'])

    return JsonResponse({
        'status': 'success',
        'label': account.display_label,
        'value': account.display_label,
        'id': account.id,
    })

# ==============================
# Authenticated Appointments/Contacts (Public forms backend)
# ==============================
class AppointmentListView(LoginRequiredMixin, ListView):
    model = PublicBooking
    template_name = 'accounts/mach/appointments_list.html'
    context_object_name = 'bookings'
    paginate_by = 100

    def get_queryset(self):
        return PublicBooking.objects.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        settings_instance = _get_booking_settings_instance()
        context.setdefault('settings_form', BusinessBookingSettingsForm(instance=settings_instance))
        context.setdefault('holiday_form', BusinessHolidayForm())
        context['booking_settings'] = settings_instance
        context['holidays'] = BusinessHoliday.objects.order_by('date')
        return context

    def post(self, request, *args, **kwargs):
        form_type = request.POST.get('form_type')
        settings_instance = BusinessBookingSettings.get_solo()

        if form_type == 'settings':
            settings_form = BusinessBookingSettingsForm(request.POST, instance=settings_instance)
            holiday_form = BusinessHolidayForm()
            if settings_form.is_valid():
                settings_form.save()
                messages.success(request, 'Booking hours updated successfully.')
                return redirect('accounts:appointments_list')
            context = self.get_context_data(settings_form=settings_form, holiday_form=holiday_form)
            return self.render_to_response(context)

        if form_type == 'holiday':
            holiday_form = BusinessHolidayForm(request.POST)
            settings_form = BusinessBookingSettingsForm(instance=settings_instance)
            if holiday_form.is_valid():
                holiday_form.save()
                messages.success(request, 'Holiday added successfully.')
                return redirect('accounts:appointments_list')
            context = self.get_context_data(settings_form=settings_form, holiday_form=holiday_form)
            return self.render_to_response(context)

        if form_type == 'delete_holiday':
            holiday_id = request.POST.get('holiday_id')
            if holiday_id:
                try:
                    BusinessHoliday.objects.get(id=holiday_id).delete()
                    messages.success(request, 'Holiday removed successfully.')
                except BusinessHoliday.DoesNotExist:
                    messages.error(request, 'The selected holiday could not be found.')
            return redirect('accounts:appointments_list')

        return redirect('accounts:appointments_list')


class AppointmentDetailView(LoginRequiredMixin, DetailView):
    model = PublicBooking
    template_name = 'accounts/mach/appointments_detail.html'
    context_object_name = 'booking'

    def get_object(self, queryset=None):
        booking = super().get_object(queryset)
        if booking.status == PublicBooking.STATUS_NEW:
            booking.status = PublicBooking.STATUS_SEEN
            booking.save(update_fields=['status'])
            _log_admin_activity(
                self.request,
                action="appointment_opened",
                object_type="appointment",
                description=f"Opened appointment request from {booking.full_name}",
                object_id=booking.pk,
                metadata={
                    "full_name": booking.full_name,
                    "service_type": booking.service_type,
                    "previous_status": PublicBooking.STATUS_NEW,
                },
            )
        return booking


class ContactMessageListView(LoginRequiredMixin, ListView):
    model = PublicContactMessage
    template_name = 'accounts/mach/contacts_list.html'
    context_object_name = 'contact_messages'
    paginate_by = 100

    def get_queryset(self):
        return PublicContactMessage.objects.order_by('-created_at')


class ContactMessageDetailView(LoginRequiredMixin, DetailView):
    model = PublicContactMessage
    template_name = 'accounts/mach/contacts_detail.html'
    context_object_name = 'message_obj'

    KEYWORD_REFERENCE_PATTERN = re.compile(
        r"(?i)(?:work\s*order|workorder|invoice|maintenance\s*task|task|job|wo)\s*(?:#|no\.?|number|:)?\s*([A-Za-z0-9-]+)"
    )

    def get_object(self, queryset=None):
        message = super().get_object(queryset)
        if message.status == PublicContactMessage.STATUS_NEW:
            message.status = PublicContactMessage.STATUS_SEEN
            message.save(update_fields=['status'])
            _log_admin_activity(
                self.request,
                action="contact_message_opened",
                object_type="contact_message",
                description=f"Opened contact message from {message.full_name}",
                object_id=message.pk,
                metadata={
                    "full_name": message.full_name,
                    "email": message.email,
                    "message_type": message.message_type,
                    "source": message.source,
                    "reference_code": message.reference_code,
                    "previous_status": PublicContactMessage.STATUS_NEW,
                },
            )
        return message

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        message: PublicContactMessage = self.object

        reference_tokens: Set[str] = set()
        numeric_candidates: Set[int] = set()

        def _record_token(raw_token: str) -> None:
            cleaned = raw_token.strip().strip('#').strip()
            if not cleaned:
                return
            reference_tokens.add(cleaned)
            digit_matches = re.findall(r'\d+', cleaned)
            for match in digit_matches:
                try:
                    numeric_candidates.add(int(match))
                except (TypeError, ValueError):
                    continue

        if message.reference_code:
            _record_token(message.reference_code)

        if message.message:
            for token in self.KEYWORD_REFERENCE_PATTERN.findall(message.message):
                _record_token(token)
            for loose_match in re.findall(r'#(\d+)', message.message):
                try:
                    numeric_candidates.add(int(loose_match))
                except (TypeError, ValueError):
                    continue

        user = self.request.user

        matched_workorders_qs = WorkOrder.objects.none()
        if numeric_candidates:
            matched_workorders_qs = WorkOrder.objects.filter(user=user, pk__in=numeric_candidates)

        invoice_query = Q()
        if numeric_candidates:
            invoice_query |= Q(pk__in=numeric_candidates)
        for token in reference_tokens:
            invoice_query |= Q(invoice_number__iexact=token)

        matched_invoices_qs = GroupedInvoice.objects.none()
        if invoice_query:
            matched_invoices_qs = GroupedInvoice.objects.filter(user=user).filter(invoice_query)

        matched_workorders = list(matched_workorders_qs)
        matched_invoices = list(matched_invoices_qs)

        context['matched_workorders'] = matched_workorders
        context['matched_invoices'] = matched_invoices
        context['reference_tokens'] = sorted(reference_tokens)
        context['reference_numbers'] = sorted(numeric_candidates)
        context['has_direct_matches'] = bool(matched_workorders or matched_invoices)
        return context


class CommunicationHubView(LoginRequiredMixin, TemplateView):
    """Unified dashboard for bookings, contact messages, outbound emails, and AI chats."""

    template_name = 'accounts/communications/communication_hub.html'

    def dispatch(self, request, *args, **kwargs):
        # Customer-portal users should never access the business communication hub.
        if getattr(request.user, 'customer_portal', None):
            raise Http404
        return super().dispatch(request, *args, **kwargs)

    def _business_user(self):
        profile = getattr(self.request.user, "profile", None)
        if profile and hasattr(profile, "get_business_user"):
            try:
                return profile.get_business_user()
            except Exception:
                return self.request.user
        return self.request.user

    def _email_logs(self, business_user):
        logs_qs = ActivityLog.objects.filter(business=business_user).filter(
            Q(object_type__icontains='email') | Q(action__icontains='email')
        ).order_by('-created_at')[:15]
        entries = []
        for log in logs_qs:
            meta = log.metadata or {}
            status = meta.get("status") or ("Failed" if "fail" in (log.action or "").lower() else "Sent")
            entries.append({
                "subject": meta.get("subject") or log.description,
                "recipient": meta.get("email") or meta.get("recipient") or meta.get("to") or "",
                "status": status,
                "created_at": log.created_at,
                "channel": meta.get("channel") or "Email",
                "error": meta.get("error") or meta.get("message") or "",
            })
        return entries

    def _ai_conversations(self, business_user):
        convo_qs = ActivityLog.objects.filter(business=business_user).filter(
            Q(object_type__icontains='ai') | Q(action__icontains='ai')
        ).order_by('-created_at')[:10]
        conversations = []
        for log in convo_qs:
            meta = log.metadata or {}
            conversations.append({
                "title": meta.get("title") or meta.get("topic") or log.description,
                "source": meta.get("source") or ("Public site" if "public" in (log.object_type or "").lower() else "Customer portal"),
                "created_at": log.created_at,
                "status": meta.get("status") or "Complete",
                "summary": meta.get("summary") or meta.get("prompt") or "",
            })
        return conversations

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        business_user = self._business_user()
        today = timezone.localdate()

        bookings_qs = PublicBooking.objects.order_by('-created_at')
        contacts_qs = PublicContactMessage.objects.order_by('-created_at')

        bookings = list(bookings_qs[:6])
        contact_messages = list(contacts_qs[:6])

        booking_stats = {
            "new": bookings_qs.filter(status=PublicBooking.STATUS_NEW).count(),
            "scheduled": bookings_qs.filter(status=PublicBooking.STATUS_SCHEDULED).count(),
            "processing": bookings_qs.filter(status=PublicBooking.STATUS_PROCESSING).count(),
            "today": bookings_qs.filter(created_at__date=today).count(),
            "total": bookings_qs.count(),
        }
        contact_stats = {
            "new": contacts_qs.filter(status=PublicContactMessage.STATUS_NEW).count(),
            "today": contacts_qs.filter(created_at__date=today).count(),
            "total": contacts_qs.count(),
        }

        email_logs = self._email_logs(business_user)
        email_status_counts = defaultdict(int)
        for entry in email_logs:
            email_status_counts[entry.get("status") or "Sent"] += 1
        email_status_summary = {
            "sent": email_status_counts.get("Sent", 0) + email_status_counts.get("sent", 0),
            "failed": email_status_counts.get("Failed", 0) + email_status_counts.get("failed", 0),
        }
        email_status_summary["total"] = sum(email_status_counts.values())

        ai_conversations = self._ai_conversations(business_user)

        booking_settings = _get_booking_settings_instance()
        upcoming_holiday = None
        try:
            upcoming_holiday = BusinessHoliday.objects.filter(date__gte=today).order_by('date').first()
        except Exception:
            upcoming_holiday = None

        context.update({
            "bookings": bookings,
            "contact_messages": contact_messages,
            "booking_stats": booking_stats,
            "contact_stats": contact_stats,
            "email_logs": email_logs,
            "email_status_counts": email_status_counts,
            "email_status_summary": email_status_summary,
            "ai_conversations": ai_conversations,
            "booking_settings": booking_settings,
            "upcoming_holiday": upcoming_holiday,
        })
        return context


def _flyer_template_library(request):
    store_url = request.build_absolute_uri(reverse('accounts:store_product_list'))
    services_url = request.build_absolute_uri(reverse('accounts:public_services'))
    booking_url = request.build_absolute_uri(reverse('accounts:public_booking'))
    contact_url = request.build_absolute_uri(reverse('accounts:public_contact'))

    return {
        "holiday": {
            "label": "Holiday readiness",
            "description": "Seasonal inspections, extended hours, and priority slots.",
            "badge": "Holiday",
            "template": "emails/flyers/holiday_flyer.html",
            "eyebrow": "Holiday special",
            "theme": {
                "primary": "#0f766e",
                "accent": "#f59e0b",
                "background": "#f3f4f6",
                "surface": "#ffffff",
            },
            "defaults": {
                "subject": "Holiday readiness for your fleet",
                "preheader": "Priority service slots and winter safety checks.",
                "headline": "Holiday readiness, handled.",
                "subheadline": "Schedule priority inspections before peak travel weeks.",
                "message": (
                    "Our team is opening extra service slots for holiday travel. "
                    "Bring your fleet in for a fast inspection so your drivers stay safe and on schedule."
                ),
                "highlights": [
                    "Winter tire and brake safety checks",
                    "Battery and electrical diagnostics",
                    "24/7 roadside response planning",
                ],
                "cta_text": "Book a holiday slot",
                "cta_url": booking_url,
                "footer_note": "Holiday availability is limited. Reserve early.",
            },
        },
        "weekly": {
            "label": "Weekly spotlight",
            "description": "Keep customers posted with a weekly service highlight.",
            "badge": "Weekly",
            "template": "emails/flyers/weekly_flyer.html",
            "eyebrow": "Weekly spotlight",
            "stats_left": "Fast scheduling",
            "stats_right": "Fleet-ready service",
            "theme": {
                "primary": "#1d4ed8",
                "accent": "#22c55e",
                "background": "#f1f5f9",
                "surface": "#ffffff",
            },
            "defaults": {
                "subject": "Weekly service spotlight",
                "preheader": "This week: priority maintenance and fast turnaround.",
                "headline": "This week at the shop",
                "subheadline": "Priority maintenance and quick turnarounds.",
                "message": (
                    "Stay ahead of downtime with a weekly maintenance check-in. "
                    "We are highlighting quick inspections and same-week scheduling for returning customers."
                ),
                "highlights": [
                    "Priority diagnostics for repeat customers",
                    "Same-week booking windows",
                    "Parts availability updates",
                ],
                "cta_text": "See weekly services",
                "cta_url": services_url,
                "footer_note": "Reply with your preferred day and time.",
            },
        },
        "promotion": {
            "label": "Promotion launch",
            "description": "Announce new promos with bold callouts and clear CTAs.",
            "badge": "Promo",
            "template": "emails/flyers/promo_flyer.html",
            "promo_badge": "Limited time",
            "promo_detail": "Bundle maintenance services and save on labor.",
            "theme": {
                "primary": "#111827",
                "accent": "#ea580c",
                "background": "#fff7ed",
                "surface": "#ffffff",
            },
            "defaults": {
                "subject": "New maintenance promotion",
                "preheader": "Bundle services and save on labor this month.",
                "headline": "Save on bundled maintenance",
                "subheadline": "Limited-time promotion for fleet partners.",
                "message": (
                    "We are rolling out a fresh promotion for fleet customers. "
                    "Book a bundled maintenance visit and reduce service downtime."
                ),
                "highlights": [
                    "Labor savings on bundled inspections",
                    "Priority scheduling for fleets",
                    "Dedicated service advisor support",
                ],
                "cta_text": "Claim the promotion",
                "cta_url": contact_url,
                "footer_note": "Offer valid while appointment slots last.",
            },
        },
    }


def _split_highlights(raw_text):
    if not raw_text:
        return []
    lines = re.split(r"[\r\n]+", str(raw_text))
    return [line.strip() for line in lines if line.strip()]


def _resolve_flyer_branding(profile, request):
    company_name = (
        getattr(profile, "company_name", None)
        or getattr(settings, "DEFAULT_BUSINESS_NAME", "Truck Zone")
        or "Truck Zone"
    )
    contact_email = (
        getattr(profile, "company_email", None)
        or getattr(settings, "DEFAULT_BUSINESS_EMAIL", None)
        or settings.DEFAULT_FROM_EMAIL
    )
    if contact_email:
        try:
            validate_email(contact_email)
        except ValidationError:
            contact_email = settings.DEFAULT_FROM_EMAIL
    business_phone = (
        getattr(profile, "company_phone", None)
        or getattr(settings, "DEFAULT_BUSINESS_PHONE", "")
    )
    business_address = (
        getattr(profile, "company_address", None)
        or getattr(settings, "DEFAULT_BUSINESS_ADDRESS", "")
    )
    company_logo_url = resolve_company_logo_url(profile, request=request)
    return {
        "company_name": company_name,
        "contact_email": contact_email,
        "business_phone": business_phone,
        "business_address": business_address,
        "company_logo_url": company_logo_url,
    }


@login_required
def flyer_campaigns(request):
    if getattr(request.user, "customer_portal", None):
        raise Http404

    business_user = get_business_user(request.user) or request.user
    business_user_ids = get_customer_user_ids(request.user)
    profile = getattr(business_user, "profile", None)

    template_library = _flyer_template_library(request)
    template_choices = [(key, data["label"]) for key, data in template_library.items()]
    default_key = next(iter(template_library))
    selected_key = request.POST.get("template_key") or request.GET.get("template") or default_key
    if selected_key not in template_library:
        selected_key = default_key
    selected_template = template_library[selected_key]

    customers_qs = Customer.objects.filter(user__in=business_user_ids).exclude(email__isnull=True).exclude(email__exact='')
    recipient_count = customers_qs.count()

    template_defaults = {}
    template_options = []
    for key, data in template_library.items():
        defaults = data.get("defaults", {})
        template_defaults[key] = {
            "subject": defaults.get("subject", ""),
            "preheader": defaults.get("preheader", ""),
            "headline": defaults.get("headline", ""),
            "subheadline": defaults.get("subheadline", ""),
            "message": defaults.get("message", ""),
            "highlights": "\n".join(defaults.get("highlights", [])),
            "cta_text": defaults.get("cta_text", ""),
            "cta_url": defaults.get("cta_url", ""),
            "footer_note": defaults.get("footer_note", ""),
        }
        template_options.append({
            "key": key,
            "label": data.get("label", key.title()),
            "description": data.get("description", ""),
            "badge": data.get("badge", "Template"),
        })

    initial_defaults = template_defaults.get(selected_key, {})
    initial = {
        "template_key": selected_key,
        "audience": "test",
        **initial_defaults,
    }

    preview_html = None
    if request.method == "POST":
        form = FlyerEmailForm(request.POST, template_choices=template_choices)
        action = (request.POST.get("action") or "preview").strip()
        if form.is_valid():
            cleaned = form.cleaned_data
            selected_key = cleaned.get("template_key") or selected_key
            selected_template = template_library.get(selected_key, selected_template)
            template_path = selected_template["template"]
            defaults = selected_template.get("defaults", {})
            branding = _resolve_flyer_branding(profile, request)

            base_context = {
                "profile": profile,
                "business_user": business_user,
                "company_name": branding["company_name"],
                "company_logo_url": branding["company_logo_url"],
                "contact_email": branding["contact_email"],
                "business_phone": branding["business_phone"],
                "business_address": branding["business_address"],
                "subject": cleaned.get("subject", "").strip(),
                "preheader": cleaned.get("preheader", "").strip(),
                "headline": cleaned.get("headline", "").strip(),
                "subheadline": cleaned.get("subheadline", "").strip(),
                "message": cleaned.get("message", "").strip(),
                "highlights": _split_highlights(cleaned.get("highlights")),
                "cta_text": cleaned.get("cta_text", "").strip(),
                "cta_url": cleaned.get("cta_url", "").strip(),
                "footer_note": (cleaned.get("footer_note") or "").strip() or defaults.get("footer_note", ""),
                "eyebrow": selected_template.get("eyebrow"),
                "promo_badge": selected_template.get("promo_badge"),
                "promo_detail": selected_template.get("promo_detail"),
                "stats_left": selected_template.get("stats_left"),
                "stats_right": selected_template.get("stats_right"),
                "theme": selected_template.get("theme", {}),
                "current_year": timezone.now().year,
            }
            base_context = apply_branding_defaults(base_context)

            preview_name = (
                request.user.get_full_name().strip()
                or request.user.username
                or "there"
            )
            preview_context = {**base_context, "recipient_name": preview_name}
            preview_html = render_to_string(template_path, preview_context)

            if action == "send":
                audience = cleaned.get("audience")
                recipients = []
                if audience == "test":
                    if not request.user.email:
                        form.add_error("audience", "Add an email to your user profile to send a test.")
                    else:
                        recipients = [(request.user.email, preview_name)]
                else:
                    seen = set()
                    for customer in customers_qs:
                        email = (customer.email or "").strip()
                        if not email:
                            continue
                        key = email.lower()
                        if key in seen:
                            continue
                        try:
                            validate_email(email)
                        except ValidationError:
                            continue
                        seen.add(key)
                        recipients.append((email, customer.name or "there"))
                    if not recipients:
                        form.add_error("audience", "No customer emails are available yet.")

                if not form.errors and recipients:
                    subject = base_context["subject"] or defaults.get("subject", "Flyer update")
                    from_email = branding["contact_email"] or settings.DEFAULT_FROM_EMAIL
                    reply_to = [branding["contact_email"]] if branding["contact_email"] else None
                    sent_count = 0
                    failed_count = 0

                    for email, recipient_name in recipients:
                        send_context = {**base_context, "recipient_name": recipient_name}
                        html_content = render_to_string(template_path, send_context)
                        text_content = strip_tags(html_content)
                        try:
                            message = EmailMultiAlternatives(
                                subject=subject,
                                body=text_content,
                                from_email=from_email,
                                to=[email],
                                reply_to=reply_to,
                            )
                            message.attach_alternative(html_content, "text/html")
                            message.send()
                            sent_count += 1
                        except Exception as exc:
                            failed_count += 1
                            _log_admin_activity(
                                request,
                                action="email_flyer_failed",
                                object_type="email_flyer",
                                description=f"Flyer email failed for {email}",
                                metadata={
                                    "status": "Failed",
                                    "email": email,
                                    "subject": subject,
                                    "template": selected_key,
                                    "error": str(exc),
                                },
                            )

                    _log_admin_activity(
                        request,
                        action="email_flyer_sent",
                        object_type="email_flyer",
                        description=f"Flyer campaign sent: {subject}",
                        metadata={
                            "status": "Sent" if sent_count else "Failed",
                            "subject": subject,
                            "template": selected_key,
                            "audience": audience,
                            "recipient_count": len(recipients),
                            "sent_count": sent_count,
                            "failed_count": failed_count,
                        },
                    )
                    if failed_count:
                        messages.warning(
                            request,
                            f"Flyer sent to {sent_count} customers. {failed_count} failed.",
                        )
                    else:
                        messages.success(request, f"Flyer sent to {sent_count} customers.")
                    return redirect("accounts:flyer_campaigns")
    else:
        form = FlyerEmailForm(initial=initial, template_choices=template_choices)
        template_path = selected_template["template"]
        branding = _resolve_flyer_branding(profile, request)
        defaults = selected_template.get("defaults", {})
        base_context = {
            "profile": profile,
            "business_user": business_user,
            "company_name": branding["company_name"],
            "company_logo_url": branding["company_logo_url"],
            "contact_email": branding["contact_email"],
            "business_phone": branding["business_phone"],
            "business_address": branding["business_address"],
            "subject": defaults.get("subject", ""),
            "preheader": defaults.get("preheader", ""),
            "headline": defaults.get("headline", ""),
            "subheadline": defaults.get("subheadline", ""),
            "message": defaults.get("message", ""),
            "highlights": defaults.get("highlights", []),
            "cta_text": defaults.get("cta_text", ""),
            "cta_url": defaults.get("cta_url", ""),
            "footer_note": defaults.get("footer_note", ""),
            "eyebrow": selected_template.get("eyebrow"),
            "promo_badge": selected_template.get("promo_badge"),
            "promo_detail": selected_template.get("promo_detail"),
            "stats_left": selected_template.get("stats_left"),
            "stats_right": selected_template.get("stats_right"),
            "theme": selected_template.get("theme", {}),
            "current_year": timezone.now().year,
        }
        base_context = apply_branding_defaults(base_context)
        preview_name = (
            request.user.get_full_name().strip()
            or request.user.username
            or "there"
        )
        preview_html = render_to_string(
            template_path,
            {**base_context, "recipient_name": preview_name},
        )

    context = {
        "form": form,
        "preview_html": preview_html,
        "selected_template_label": selected_template.get("label"),
        "template_defaults": template_defaults,
        "template_options": template_options,
        "recipient_count": recipient_count,
    }
    return render(request, "accounts/communications/flyer_campaigns.html", context)


def _build_category_path(category):
    path = []
    current = category
    while current:
        path.append(current)
        current = current.parent
    return list(reversed(path))


def _collect_descendant_ids(category, children_map):
    ids = [category.id]
    stack = list(children_map.get(category.id, []))
    while stack:
        child = stack.pop()
        ids.append(child.id)
        stack.extend(children_map.get(child.id, []))
    return ids


def _build_category_tree(categories):
    by_parent = defaultdict(list)
    for category in categories:
        by_parent[category.parent_id].append(category)
    for children in by_parent.values():
        children.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))

    def build(parent_id=None):
        nodes = []
        for category in by_parent.get(parent_id, []):
            nodes.append({
                "category": category,
                "children": build(category.id),
            })
        return nodes

    return build(None)


def _flatten_category_tree(tree, active_path_ids):
    flattened = []
    active_ids = set(active_path_ids or [])

    def walk(nodes, depth=0):
        for node in nodes:
            category = node["category"]
            flattened.append({
                "category": category,
                "depth": depth,
                "is_current": category.id == (active_path_ids[-1] if active_path_ids else None),
                "is_active": category.id in active_ids,
                "has_children": bool(node["children"]),
            })
            walk(node["children"], depth + 1)

    walk(tree)
    return flattened


def _build_storefront_context(request, available_products):
    products = available_products
    search_query = (request.GET.get('q') or '').strip()
    category_ids = [value for value in request.GET.getlist('category') if value]
    group_ids = [value for value in request.GET.getlist('group') if value]
    brand_ids = [value for value in request.GET.getlist('brand') if value]
    customer_account = getattr(request.user, 'customer_portal', None) if request.user.is_authenticated else None
    store_owner = get_storefront_owner(request)
    category_flags = resolve_storefront_category_flags(request, store_owner)
    show_empty_categories = category_flags["show_empty_categories"]

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(sku__icontains=search_query)
            | Q(category__name__icontains=search_query)
            | Q(brand__name__icontains=search_query)
            | Q(vehicle_model__name__icontains=search_query)
            | Q(vin_number__vin__icontains=search_query)
        )

    if group_ids:
        products = products.filter(category__group_id__in=group_ids)

    categories_qs = Category.objects.filter(is_active=True)
    if store_owner:
        categories_qs = categories_qs.filter(user=store_owner)
    if not show_empty_categories:
        categories_qs = categories_qs.filter(products__in=available_products)
    if group_ids:
        categories_qs = categories_qs.filter(group_id__in=group_ids)
    categories = list(
        categories_qs.select_related('group', 'parent').distinct()
    )

    children_map = defaultdict(list)
    for category in categories:
        children_map[category.parent_id].append(category)
    for children in children_map.values():
        children.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))

    selected_category = None
    if len(category_ids) == 1:
        category_lookup = Category.objects.select_related('group', 'parent').filter(
            id=category_ids[0],
            is_active=True,
        )
        if store_owner:
            category_lookup = category_lookup.filter(user=store_owner)
        selected_category = category_lookup.first()

    active_group = None
    if len(group_ids) == 1:
        group_lookup = CategoryGroup.objects.filter(id=group_ids[0], is_active=True)
        if store_owner:
            group_lookup = group_lookup.filter(user=store_owner)
        active_group = group_lookup.first()
    if not active_group and selected_category and selected_category.group_id:
        active_group = selected_category.group

    descendant_ids = None
    if category_ids:
        if selected_category:
            descendant_ids = _collect_descendant_ids(selected_category, children_map)
            products = products.filter(category_id__in=descendant_ids)
        else:
            products = products.filter(category_id__in=category_ids)

    if brand_ids:
        products = products.filter(brand_id__in=brand_ids)

    attribute_filters = []
    if selected_category:
        category_chain = []
        current = selected_category
        while current:
            category_chain.append(current)
            current = current.parent

        attributes = (
            CategoryAttribute.objects.filter(
                category__in=category_chain,
                is_filterable=True,
                is_active=True,
            )
            .prefetch_related("options")
            .order_by("sort_order", "name")
        )
        for attribute in attributes:
            param_name = f"attr_{attribute.id}"
            selected_value = (request.GET.get(param_name) or "").strip()
            if selected_value:
                if attribute.attribute_type == "select":
                    products = products.filter(
                        attribute_values__attribute=attribute,
                        attribute_values__option_id=selected_value,
                    )
                elif attribute.attribute_type == "boolean":
                    normalized = selected_value.lower()
                    if normalized in ("1", "true", "yes", "on"):
                        bool_value = True
                    elif normalized in ("0", "false", "no", "off"):
                        bool_value = False
                    else:
                        bool_value = None
                    if bool_value is not None:
                        products = products.filter(
                            attribute_values__attribute=attribute,
                            attribute_values__value_boolean=bool_value,
                        )
                elif attribute.attribute_type == "number":
                    try:
                        number_value = Decimal(selected_value)
                    except (ArithmeticError, ValueError):
                        number_value = None
                    if number_value is not None:
                        products = products.filter(
                            attribute_values__attribute=attribute,
                            attribute_values__value_number=number_value,
                        )
                else:
                    products = products.filter(
                        attribute_values__attribute=attribute,
                        attribute_values__value_text__icontains=selected_value,
                    )

            attribute_filters.append(
                {
                    "attribute": attribute,
                    "param": param_name,
                    "selected": selected_value,
                    "options": list(attribute.options.filter(is_active=True).order_by("sort_order", "value")),
                }
            )
        if attributes:
            products = products.distinct()

    brand_products = available_products
    if group_ids:
        brand_products = brand_products.filter(category__group_id__in=group_ids)
    if category_ids:
        if descendant_ids:
            brand_products = brand_products.filter(category_id__in=descendant_ids)
        else:
            brand_products = brand_products.filter(category_id__in=category_ids)
    brands = list(
        ProductBrand.objects.filter(
            products__in=brand_products,
            is_active=True,
        )
        .distinct()
        .order_by('sort_order', 'name')
    )

    group_queryset = CategoryGroup.objects.filter(is_active=True)
    if store_owner:
        group_queryset = group_queryset.filter(user=store_owner)
    if not show_empty_categories:
        group_queryset = group_queryset.filter(
            categories__products__in=available_products,
            categories__is_active=True,
        ).distinct()
    all_category_groups = list(group_queryset.order_by('sort_order', 'name'))

    has_filter = bool(search_query or group_ids or category_ids or brand_ids)
    display_groups = [active_group] if active_group and has_filter else all_category_groups
    grouped_categories = []
    for group in display_groups:
        if not group:
            continue
        group_categories = [
            category for category in categories
            if category.group_id == group.id and category.parent_id is None
        ]
        group_categories.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))
        grouped_categories.append({
            "group": group,
            "categories": group_categories,
        })

    group_categories = []
    if active_group:
        group_categories = [
            category for category in categories
            if category.group_id == active_group.id and category.parent_id is None
        ]
        group_categories.sort(key=lambda cat: (cat.sort_order, cat.name.lower()))

    selected_category_children = []
    if selected_category:
        selected_category_children = children_map.get(selected_category.id, [])
        if not selected_category_children:
            child_queryset = Category.objects.filter(
                parent=selected_category,
                is_active=True,
            )
            if store_owner:
                child_queryset = child_queryset.filter(user=store_owner)
            if not show_empty_categories:
                child_queryset = child_queryset.filter(products__in=available_products)
            selected_category_children = list(
                child_queryset.distinct().order_by('sort_order', 'name')
            )

    show_results = has_filter
    if selected_category:
        show_products = bool(search_query) or not selected_category_children
    elif group_ids:
        show_products = bool(search_query) or bool(brand_ids)
    else:
        show_products = bool(search_query) or bool(brand_ids)

    active_path = _build_category_path(selected_category) if selected_category else []
    active_path_ids = [category.id for category in active_path]

    category_tree_flat = []
    if show_results:
        tree_categories = categories
        if active_group:
            tree_categories = [category for category in categories if category.group_id == active_group.id]
        category_tree = _build_category_tree(tree_categories)
        category_tree_flat = _flatten_category_tree(category_tree, active_path_ids)

    home_url = reverse('accounts:public_home')
    store_url = reverse('accounts:store_product_list')
    breadcrumbs = [{"label": "Home", "url": home_url}]
    if show_results:
        breadcrumbs.append({"label": "All Categories", "url": store_url})
        if active_group:
            breadcrumbs.append({
                "label": active_group.name,
                "url": reverse('accounts:store_group_detail', args=[active_group.id]),
            })
        for category in active_path:
            breadcrumbs.append({
                "label": category.name,
                "url": reverse('accounts:store_category_detail', args=[category.id]),
            })
        if search_query:
            breadcrumbs.append({"label": f"Search: {search_query}", "url": None})

    if search_query:
        heading = f"Search results for \"{search_query}\""
        subheading = f"{products.count()} results"
    elif selected_category:
        heading = selected_category.name
        subheading = selected_category.description or None
    elif active_group:
        heading = active_group.name
        subheading = active_group.description or None
    else:
        heading = "All Categories"
        subheading = None

    paginator = Paginator(products.order_by('name'), 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    params = request.GET.copy()
    params.pop('page', None)

    category_params = request.GET.copy()
    category_params.pop('category', None)
    category_params.pop('page', None)
    category_query = category_params.urlencode()
    category_query_prefix = f"?{category_query}&" if category_query else "?"

    group_params = request.GET.copy()
    group_params.pop('group', None)
    group_params.pop('page', None)
    group_query = group_params.urlencode()
    group_query_prefix = f"?{group_query}&" if group_query else "?"

    base_params = request.GET.copy()
    base_params.pop('group', None)
    base_params.pop('category', None)
    base_params.pop('page', None)
    base_query = base_params.urlencode()
    base_query_prefix = f"?{base_query}" if base_query else ""

    return {
        'storefront_products': page_obj,
        'storefront_products_count': paginator.count,
        'storefront_categories': categories,
        'storefront_category_groups': all_category_groups,
        'storefront_grouped_categories': grouped_categories,
        'storefront_group_categories': group_categories,
        'storefront_subcategories': selected_category_children,
        'storefront_category_tree': category_tree_flat,
        'storefront_selected_category': selected_category,
        'storefront_active_group': active_group,
        'storefront_breadcrumbs': breadcrumbs,
        'storefront_heading': heading,
        'storefront_subheading': subheading,
        'storefront_search_query': search_query,
        'storefront_selected_category_ids': [str(value) for value in category_ids],
        'storefront_selected_group_ids': [str(value) for value in group_ids],
        'storefront_selected_brand_ids': [str(value) for value in brand_ids],
        'storefront_attribute_filters': attribute_filters,
        'storefront_brands': brands,
        'storefront_show_results': show_results,
        'storefront_show_products': show_products,
        'storefront_query_string': params.urlencode(),
        'storefront_category_query_prefix': category_query_prefix,
        'storefront_group_query_prefix': group_query_prefix,
        'storefront_base_query_prefix': base_query_prefix,
        'customer_account': customer_account,
    }


def public_home(request):
    group_id = request.GET.get('group')
    category_id = request.GET.get('category')
    search_query = (request.GET.get('q') or '').strip()

    if category_id and category_id.isdigit():
        params = request.GET.copy()
        params.pop('category', None)
        params.pop('group', None)
        query_string = params.urlencode()
        category_url = reverse('accounts:store_category_detail', args=[int(category_id)])
        if query_string:
            category_url = f"{category_url}?{query_string}"
        return redirect(category_url)

    if group_id and group_id.isdigit():
        params = request.GET.copy()
        params.pop('group', None)
        params.pop('category', None)
        query_string = params.urlencode()
        group_url = reverse('accounts:store_group_detail', args=[int(group_id)])
        if query_string:
            group_url = f"{group_url}?{query_string}"
        return redirect(group_url)

    if search_query:
        params = request.GET.copy()
        params.pop('group', None)
        params.pop('category', None)
        query_string = params.urlencode()
        search_url = reverse('accounts:store_search')
        if query_string:
            search_url = f"{search_url}?{query_string}"
        return redirect(search_url)

    return redirect('accounts:store_product_list')

def public_about(request):
    return render(request, 'public_about.html')

def public_services(request):
    return redirect('accounts:public_home')


def service_engine(request):
    return redirect('accounts:public_home')


def service_transmission(request):
    return redirect('accounts:public_home')


def service_brakes(request):
    return redirect('accounts:public_home')


def service_electrical(request):
    return redirect('accounts:public_home')


def service_maintenance(request):
    return redirect('accounts:public_home')


def service_dot(request):
    return redirect('accounts:public_home')


def service_dpf(request):
    return redirect('accounts:public_home')

def service_tires(request):
    return redirect('accounts:public_home')

def service_road_service(request):
    return redirect('accounts:public_home')

def public_contact(request):
    # Unified "Send Us a Message" form
    initial = {}

    if request.method == 'POST':
        form = PublicContactForm(request.POST)
        if form.is_valid():
            message_obj = form.save(commit=False)
            message_obj.source = PublicContactMessage.SOURCE_PUBLIC
            message_obj.message_type = PublicContactMessage.TYPE_GENERAL
            if not message_obj.subject:
                message_obj.subject = "Parts inquiry"
            message_obj.save()
            context = {'message_obj': message_obj}
            return render(request, 'public/contact_success.html', context)
    else:
        form = PublicContactForm(initial=initial)

    return render(request, 'public/contact_form.html', { 'form': form })


def public_faq(request):
    return render(request, 'public/faq.html')

@login_required
def home(request):
    # --- Profile and Template Determination ---
    import json
    from decimal import Decimal
    from datetime import timedelta
    from django.shortcuts import render
    from django.db.models import Sum, Value, F, Count, Q, DecimalField
    from django.db.models.functions import Coalesce, TruncWeek, TruncMonth, TruncYear, TruncDate
    from django.db.utils import OperationalError as DbOperationalError
    from django.utils.dateparse import parse_date
    from django.db.models.expressions import ExpressionWrapper
    from django.utils import timezone
    from django.utils.safestring import mark_safe
    from django.core.serializers.json import DjangoJSONEncoder
    from .utils import get_overdue_total_balance
    import pytz

    # --- Profile and Template Determination ---
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        # IMPORTANT: Redirect to a profile creation/setup page if profile is critical
        # return redirect('accounts:your_profile_setup_url_name') # Example
        # For now, if it's not found, we can't proceed with logic depending on profile.term etc.
        # You need to decide how to handle this gracefully.
        # If profile is optional for some parts, more complex handling is needed.
        # Assuming for now that if profile doesn't exist, we cannot render the dashboard.
        # You could render a simple message or redirect.
        return render(request, 'accounts/profile_missing.html') # Create this template


    # Choose dashboard layout by occupation.
    occupation = (getattr(profile, 'occupation', '') or '').strip().lower()
    if occupation == 'towing':
        template_name = 'accounts/towing/home.html'
    elif occupation == 'parts_store':
        template_name = 'accounts/parts_store/home.html'
    else:
        template_name = 'accounts/mach/home.html'

    business_user = profile.get_business_user() if hasattr(profile, 'get_business_user') else request.user
    business_user_ids = get_customer_user_ids(request.user)
    actual_user = getattr(request, 'actual_user', request.user)
    show_recent_activity = actual_user.is_authenticated and actual_user.is_superuser
    recent_activity_logs = []
    has_additional_activity = False
    if show_recent_activity:
        activity_log_qs = ActivityLog.objects.filter(
            business=business_user
        ).select_related('actor').order_by('-created_at')
        recent_activity_logs = list(activity_log_qs[:5])
        has_additional_activity = activity_log_qs.count() > 5

    # --- Timezone and Date Setup ---
    user_timezone_str = getattr(settings, 'TIME_ZONE', 'UTC')
    user_timezone = pytz.timezone(user_timezone_str)
    today = timezone.now().astimezone(user_timezone).date()
    today_tasks_count = (
        WorkOrder.objects.filter(user=request.user, scheduled_date=today)
        .exclude(status='completed')
        .count()
    )

    # --- Towing/Invoice snapshot metrics ---
    week_start = today - timedelta(days=today.weekday())  # Monday
    this_week_invoices_count = GroupedInvoice.objects.filter(
        user=request.user,
        date__gte=week_start,
        date__isnull=False,
    ).count()
    todays_invoices_count = GroupedInvoice.objects.filter(
        user=request.user,
        date=today,
    ).count()
    this_week_earnings_paid = (
        GroupedInvoice.objects.filter(
            user=request.user,
            date__gte=week_start,
            date__isnull=False,
            paid_invoice__isnull=False,
        ).aggregate(total=Coalesce(Sum('total_amount'), Value(Decimal('0.00'))))['total']
        or Decimal('0.00')
    )

    # --- Financial Calculations (Invoices, Overdue) ---
    term_days = TERM_CHOICES.get(profile.term, 30)

    pending_invoices_qs = PendingInvoice.objects.filter(
        is_paid=False,
        grouped_invoice__user=request.user
    ).select_related('grouped_invoice').prefetch_related('grouped_invoice__payments').annotate(
        total_paid=Coalesce(
            Sum('grouped_invoice__payments__amount'),
            Value(Decimal('0.00')),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )
    pending_invoices_qs = _annotate_invoice_credit_totals(
        pending_invoices_qs,
        invoice_field='grouped_invoice_id',
    ).annotate(
        balance_due=ExpressionWrapper(
            F('grouped_invoice__total_amount') - F('total_paid') - F('credit_total'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )

    if term_days > 0:
        overdue_threshold_date = today - timedelta(days=term_days)
        overdue_filter_lookup = 'grouped_invoice__date__lt'
    else:
        overdue_threshold_date = today
        overdue_filter_lookup = 'grouped_invoice__date__lte'

    overdue_invoices_qs = pending_invoices_qs.filter(
        **{overdue_filter_lookup: overdue_threshold_date}
    )
    overdue_total_balance = overdue_invoices_qs.aggregate(
        total_overdue=Sum('balance_due')
    )['total_overdue'] or Decimal('0.00')

    grouped_payments_sum = Coalesce(
        Sum('payments__amount'),
        Value(Decimal('0.00')),
        output_field=DecimalField(max_digits=10, decimal_places=2)
    )

    # --- Overdue customers summary for dashboard ---
    # Aggregate overdue balance per customer, sorted by highest first.
    # Also include "reminder sent today" flag.
    overdue_customer_rows = []
    try:
        # Use a subquery for invoice total_paid so we can safely Sum(balance_due)
        # in the customer aggregation (Django disallows summing an aggregate annotation).
        invoice_paid_subquery = (
            Payment.objects.filter(invoice=OuterRef('pk'))
            .values('invoice')
            .annotate(
                total=Coalesce(
                    Sum('amount'),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            )
            .values('total')[:1]
        )
        outstanding_customer_rows = list(
            _annotate_invoice_credit_totals(
                GroupedInvoice.objects.filter(
                    user=request.user,
                    customer__isnull=False,
                ).annotate(
                    total_paid=Coalesce(
                        Subquery(invoice_paid_subquery),
                        Value(Decimal('0.00')),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    )
                )
            )
            .annotate(
                balance_due=ExpressionWrapper(
                    F('total_amount') - F('total_paid') - F('credit_total'),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
            )
            .filter(balance_due__gt=Decimal('0.00'))
            .values('customer_id')
            .annotate(total_outstanding=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField()))
        )
        outstanding_by_customer = {
            row.get('customer_id'): (row.get('total_outstanding') or Decimal('0.00'))
            for row in outstanding_customer_rows
        }
        overdue_customer_rows = list(
            _annotate_invoice_credit_totals(
                GroupedInvoice.objects.filter(
                    user=request.user,
                    customer__isnull=False,
                ).annotate(
                    total_paid=Coalesce(
                        Subquery(invoice_paid_subquery),
                        Value(Decimal('0.00')),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    )
                )
            )
            .annotate(
                balance_due=ExpressionWrapper(
                    F('total_amount') - F('total_paid') - F('credit_total'),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
            )
            .filter(
                balance_due__gt=Decimal('0.00'),
                date__isnull=False,
                **({'date__lt': overdue_threshold_date} if term_days > 0 else {'date__lte': overdue_threshold_date}),
            )
            .values('customer_id', 'customer__name')
            .annotate(total_overdue=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField()))
            .order_by('-total_overdue', 'customer__name')
        )
    except DbOperationalError:
        overdue_customer_rows = []

    reminder_rows = list(
        ReminderLog.objects.filter(customer__user__in=business_user_ids)
        .values('customer_id')
        .annotate(
            last_sent=Max('sent_at'),
            reminder_count=Count('id'),
        )
    )
    reminder_by_customer = {
        row['customer_id']: {
            'last_sent': row.get('last_sent'),
            'reminder_count': row.get('reminder_count') or 0,
        }
        for row in reminder_rows
    }

    customer_meta_by_id = {}
    if overdue_customer_rows:
        customer_ids = [row.get('customer_id') for row in overdue_customer_rows if row.get('customer_id')]
        if customer_ids:
            customer_meta_by_id = {
                row['id']: row
                for row in Customer.objects.filter(user__in=business_user_ids, id__in=customer_ids).values(
                    'id',
                    'phone_number',
                    'next_followup',
                    'collection_notes',
                )
            }

    overdue_customers_payload = []
    total_overdue_customers_amount = Decimal('0.00')
    for row in overdue_customer_rows:
        cid = row.get('customer_id')
        amount = row.get('total_overdue') or Decimal('0.00')
        total_overdue_customers_amount += Decimal(str(amount))
        outstanding_amount = outstanding_by_customer.get(cid) or Decimal('0.00')
        reminder_info = reminder_by_customer.get(cid) or {}
        last_sent = reminder_info.get('last_sent')
        sent_today = bool(last_sent and last_sent.date() == today)
        customer_meta = customer_meta_by_id.get(cid) or {}
        next_followup = customer_meta.get('next_followup')
        overdue_customers_payload.append(
            {
                'customer_id': cid,
                'customer_name': row.get('customer__name') or '',
                'balance_due': float(outstanding_amount),
                'outstanding_due': float(outstanding_amount),
                'sent_today': sent_today,
                'last_sent': last_sent.isoformat() if last_sent else '',
                'reminder_count': int(reminder_info.get('reminder_count') or 0),
                'customer_phone': customer_meta.get('phone_number') or '',
                'next_followup': next_followup.isoformat() if next_followup else '',
                'collection_notes': customer_meta.get('collection_notes') or '',
            }
        )

    overdue_customers_count = len(overdue_customers_payload)
    grouped_pending_invoices_qs = GroupedInvoice.objects.filter(user=request.user).annotate(
        total_paid=grouped_payments_sum,
    )
    grouped_pending_invoices_qs = _annotate_invoice_credit_totals(
        grouped_pending_invoices_qs
    ).annotate(
        balance_due=ExpressionWrapper(
            F('total_amount') - F('total_paid') - F('credit_total'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).filter(balance_due__gt=Decimal('0.00'))
    pending_invoices_count = grouped_pending_invoices_qs.count()
    pending_invoices_total = grouped_pending_invoices_qs.aggregate(
        total=Sum('balance_due')
    )['total'] or Decimal('0.00')
    current_year = today.year
    pending_invoices_this_year_count = grouped_pending_invoices_qs.filter(
        date__isnull=False,
        date__year=current_year,
    ).count()
    last_365_start = today - timedelta(days=365)
    if term_days > 0:
        overdue_grouped_qs = grouped_pending_invoices_qs.filter(
            date__isnull=False,
            date__lt=overdue_threshold_date,
        )
    else:
        overdue_grouped_qs = grouped_pending_invoices_qs.filter(
            date__isnull=False,
            date__lte=overdue_threshold_date,
        )
    overdue_balance_last_365_days = overdue_grouped_qs.filter(
        date__gte=last_365_start,
    ).aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
    total_pending_balance = pending_invoices_total

    expenses_agg = MechExpenseItem.objects.filter(mech_expense__user=request.user).aggregate(total=Sum('amount'))
    income_agg = Payment.objects.filter(invoice__user=request.user).aggregate(total=Sum('amount'))
    expense_total = expenses_agg['total'] or Decimal('0.00')
    income_total = income_agg['total'] or Decimal('0.00')

    # --- Graph Data Calculations ---
    # ***** CORRECTION APPLIED HERE: Removed tzinfo from Trunc functions for DateFields *****

    # Monthly Data
    # Assuming Payment.date and MechExpense.date (via mech_expense__date) are DateFields
    monthly_income_data = list(Payment.objects.filter(invoice__user=request.user).annotate(
        period=TruncMonth('date') # REMOVED tzinfo
    ).values('period').annotate(total=Sum('amount')).order_by('period').values('period', 'total'))
    for item in monthly_income_data: item['period'] = item['period'].strftime('%B %Y') # Formatting remains

    monthly_expenses_data = list(MechExpenseItem.objects.filter(mech_expense__user=request.user).annotate(
        period=TruncMonth('mech_expense__date') # REMOVED tzinfo
    ).values('period').annotate(total=Sum('amount')).order_by('period').values('period', 'total'))
    for item in monthly_expenses_data: item['period'] = item['period'].strftime('%B %Y')

    # Weekly Data
    weekly_income_data = list(Payment.objects.filter(invoice__user=request.user).annotate(
        period=TruncWeek('date') # REMOVED tzinfo
    ).values('period').annotate(total=Sum('amount')).order_by('period').values('period', 'total'))
    for item in weekly_income_data: item['period'] = item['period'].strftime('Week of %b %d, %Y')

    weekly_expenses_data = list(MechExpenseItem.objects.filter(mech_expense__user=request.user).annotate(
        period=TruncWeek('mech_expense__date') # REMOVED tzinfo
    ).values('period').annotate(total=Sum('amount')).order_by('period').values('period', 'total'))
    for item in weekly_expenses_data: item['period'] = item['period'].strftime('Week of %b %d, %Y')

    # Yearly Data
    yearly_income_data = list(Payment.objects.filter(invoice__user=request.user).annotate(
        period=TruncYear('date') # REMOVED tzinfo
    ).values('period').annotate(total=Sum('amount')).order_by('period').values('period', 'total'))
    for item in yearly_income_data: item['period'] = item['period'].strftime('%Y')

    yearly_expenses_data = list(MechExpenseItem.objects.filter(mech_expense__user=request.user).annotate(
        period=TruncYear('mech_expense__date') # REMOVED tzinfo
    ).values('period').annotate(total=Sum('amount')).order_by('period').values('period', 'total'))
    for item in yearly_expenses_data: item['period'] = item['period'].strftime('%Y')

    has_graph_data = any([len(monthly_income_data) > 0, len(weekly_income_data) > 0, len(yearly_income_data) > 0])
    default_interval = "none"
    if len(monthly_income_data) >= 1:
        default_interval = "monthly"
    elif len(weekly_income_data) >= 1:
        default_interval = "weekly"
    elif len(yearly_income_data) >= 1:
        default_interval = "yearly"

    # --- Notes & Profile Completion ---
    recent_notes = Note.objects.filter(user=request.user).order_by('-pinned', '-created_at')[:5]
    note_form = NoteForm()
    service_count = Service.objects.filter(user=request.user, is_active=True).count()

    required_profile_fields = [
        'company_name', 'company_address', 'company_email', 'company_phone',
        'gst_hst_number', 'occupation', 'street_address', 'city', 'postal_code'
    ]
    filled_count = 0
    for field_name in required_profile_fields:
        if getattr(profile, field_name, None):
            filled_count += 1
    completion_percentage = int(filled_count / len(required_profile_fields) * 100) if len(required_profile_fields) > 0 else 100

    # --- Initial Notification Counts for Badge ---
    initial_low_stock_count = Product.get_low_stock_products(request.user).count()
    total_initial_notification_count = initial_low_stock_count

    new_appointments_count = PublicBooking.objects.filter(
        status=PublicBooking.STATUS_NEW
    ).count()
    new_contact_messages_count = PublicContactMessage.objects.filter(
        status=PublicContactMessage.STATUS_NEW
    ).count()
    open_workorders_count = WorkOrder.objects.filter(
        user=request.user
    ).exclude(status='completed').count()
    active_mechanics_count = Mechanic.objects.filter(user=request.user).count()

    upcoming_maintenance = list(
        VehicleMaintenanceTask.objects.filter(
            user=request.user,
            status__in=VehicleMaintenanceTask.ACTIVE_STATUSES,
        )
        .select_related('vehicle')
        .order_by('due_date', 'priority', 'title')[:6]
    )
    recent_workorders = list(
        WorkOrder.objects.filter(user=request.user)
        .select_related('customer', 'vehicle')
        .prefetch_related('assignments__mechanic')
        .order_by('-date_created')[:6]
    )
    for wo in recent_workorders:
        mechanic_names = [
            assignment.mechanic.name
            for assignment in wo.assignments.all()
            if assignment.mechanic_id
        ]
        wo.assigned_mechanic = ", ".join(mechanic_names) if mechanic_names else None
        wo.display_unit = (
            wo.unit_no
            or (wo.vehicle.unit_number if getattr(wo, "vehicle", None) else None)
            or None
        )

    recent_invoices = list(
        GroupedInvoice.objects.filter(user=request.user)
        .select_related('customer')
        .prefetch_related('payments')
        .order_by('-date', '-id')[:6]
    )
    for inv in recent_invoices:
        payment_qs = inv.payments.all() if hasattr(inv, 'payments') else []
        total_paid = sum((p.amount or Decimal('0.00')) for p in payment_qs)
        try:
            total_paid = Decimal(str(total_paid))
        except Exception:
            total_paid = Decimal('0.00')
        total_amount = inv.total_amount or Decimal('0.00')
        credit_total = ensure_decimal(getattr(inv, 'total_credit_amount', Decimal('0.00')))
        balance_due = total_amount - total_paid - credit_total
        if total_paid >= total_amount:
            inv.display_status = 'Paid'
        elif total_paid > Decimal('0.00'):
            inv.display_status = 'Partially Paid'
        else:
            inv.display_status = 'Unpaid'
    top_pending_invoices = list(pending_invoices_qs[:6])
    pending_invoices_count = pending_invoices_qs.count()
    recent_bills_total = (
        MechExpenseItem.objects.filter(
            mech_expense__user=request.user,
            mech_expense__date__gte=today - timedelta(days=30),
        ).aggregate(total=Sum('amount'))['total']
        or Decimal('0.00')
    )
    quick_customers = Customer.objects.filter(user__in=business_user_ids).order_by('name')[:25]
    quick_vehicles = Vehicle.objects.filter(customer__user__in=business_user_ids).select_related('customer').order_by('customer__name', 'unit_number', 'make_model')
    quick_mechanics = Mechanic.objects.filter(user=request.user).order_by('name')
    supplier_names = list(
        Supplier.objects.filter(user=request.user)
        .order_by('name')
        .values_list('name', flat=True)
    )
    sales_window_start = today - timedelta(days=30)
    pending_customer_approvals_qs = (
        Customer.objects.filter(
            user__in=business_user_ids,
            portal_signup_status=Customer.PORTAL_STATUS_PENDING,
            portal_user__isnull=False,
        )
        .select_related('portal_user')
        .order_by('portal_user__date_joined', 'name')
    )
    pending_customer_approvals_count = pending_customer_approvals_qs.count()
    pending_customer_approvals = list(pending_customer_approvals_qs[:5])
    stock_owner = get_stock_owner(request.user)
    inventory_value_expr = ExpressionWrapper(
        F('products__cost_price') * F('products__stock_levels__quantity_in_stock'),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    top_suppliers = list(
        Supplier.objects.filter(user=request.user)
        .annotate(
            total_inventory_value=Coalesce(
                Sum(
                    inventory_value_expr,
                    filter=Q(products__stock_levels__user=stock_owner),
                ),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .order_by('-total_inventory_value', 'name')[:5]
    )
    category_group_sales_rows = (
        IncomeRecord2.objects.filter(
            grouped_invoice__user=request.user,
            grouped_invoice__date__gte=sales_window_start,
            product__isnull=False,
        )
        .values('product__category__group__name')
        .annotate(
            total_amount=Coalesce(
                Sum('amount'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .order_by('-total_amount')
    )
    category_group_sales = []
    for row in category_group_sales_rows:
        label = row['product__category__group__name'] or 'Uncategorized'
        amount = row['total_amount'] or Decimal('0.00')
        category_group_sales.append({
            'label': label,
            'value': float(amount),
        })
    if len(category_group_sales) > 6:
        top_slices = category_group_sales[:6]
        other_total = sum(item['value'] for item in category_group_sales[6:])
        if other_total > 0:
            top_slices.append({'label': 'Other', 'value': float(other_total)})
        category_group_sales = top_slices
    category_group_sales_total = sum(item['value'] for item in category_group_sales)
    inventory_products_data = [
        {
            "id": p.id,
            "name": p.name,
            "sku": p.sku or "",
            "cost_price": float(p.cost_price) if p.cost_price is not None else 0.0,
            "sale_price": float(p.sale_price) if p.sale_price is not None else None,
            "description": p.description or "",
            "supplier": (p.supplier.name if p.supplier else "") or "",
        }
        for p in Product.objects.filter(user__in=get_product_user_ids(request.user)).select_related('supplier').order_by('name')
    ]
    workorder_30d_start = today - timedelta(days=29)
    volume_map = {}
    try:
        volume_rows = (
            WorkOrder.objects.filter(
                user=request.user,
                scheduled_date__gte=workorder_30d_start,
                scheduled_date__lte=today,
            )
            .annotate(day=TruncDate('scheduled_date'))
            .values('day')
            .annotate(total=Count('id'))
            .order_by('day')
        )
        volume_map = {row['day']: row['total'] for row in volume_rows}
    except DbOperationalError:
        # SQLite sometimes raises "user-defined function raised exception" when truncating;
        # fallback to Python aggregation to keep the dashboard working.
        raw_dates = WorkOrder.objects.filter(
            user=request.user,
            scheduled_date__gte=workorder_30d_start,
            scheduled_date__lte=today,
        ).values_list('scheduled_date', flat=True)
        for raw in raw_dates:
            day = raw
            if not day and raw:
                day = parse_date(str(raw))
            if not day:
                continue
            volume_map[day] = volume_map.get(day, 0) + 1
    workorder_volume_chart = [
        {
            "label": (workorder_30d_start + timedelta(days=offset)).strftime('%b %d'),
            "value": volume_map.get(workorder_30d_start + timedelta(days=offset), 0),
        }
        for offset in range(30)
    ]

    revenue_window_start = today - timedelta(days=30)
    revenue_rows = (
        IncomeRecord2.objects.filter(
            grouped_invoice__user=request.user,
            grouped_invoice__date__gte=revenue_window_start,
        )
        .values('job', 'product__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    revenue_pairs = []
    for row in revenue_rows:
        label = row['job'] or row['product__name'] or 'Other'
        amount = row['total'] or Decimal('0.00')
        revenue_pairs.append((label, amount))
    service_revenue_chart = []
    if revenue_pairs:
        top_slices = revenue_pairs[:5]
        remainder_total = sum(amount for _, amount in revenue_pairs[5:])
        service_revenue_chart = [
            {"label": label, "value": float(amount)} for label, amount in top_slices
        ]
        if remainder_total > 0:
            service_revenue_chart.append({"label": "Other", "value": float(remainder_total)})
    province_code = getattr(profile, 'province', None) or 'ON'
    default_tax_rate = float(PROVINCE_TAX_RATES.get(province_code, 0) or 0)

    quickbooks_settings = QuickBooksSettings.objects.filter(user=request.user).first()
    quickbooks_status = 'not_configured'
    if quickbooks_settings:
        if quickbooks_settings.is_configured:
            quickbooks_status = 'configured'
            if quickbooks_settings.has_valid_access_token():
                quickbooks_status = 'connected'
        else:
            quickbooks_status = 'incomplete'

    pending_admin_profiles = []
    can_manage_admins = False

    business_scope_user = business_user
    if actual_user and actual_user.is_authenticated:
        if actual_user.is_superuser or actual_user == business_scope_user:
            can_manage_admins = True
            pending_admin_profiles = list(
                Profile.objects.filter(
                    business_owner=business_scope_user,
                    is_business_admin=True,
                    admin_approved=False,
                )
                .select_related("user")
                .order_by("user__date_joined")
            )

    superuser_summary_cards = []
    admin_shortcuts = {}
    if actual_user.is_superuser:
        total_active_users = User.objects.filter(is_active=True).count()
        approved_admin_count = Profile.objects.filter(
            is_business_admin=True,
            admin_approved=True,
        ).count()
        pending_admin_count = len(pending_admin_profiles)
        todays_new_users = User.objects.filter(date_joined__date=today).count()

        superuser_summary_cards = [
            {
                'label': 'Active portal users',
                'value': total_active_users,
                'icon': 'fa-users',
            },
            {
                'label': 'Approved staff accounts',
                'value': approved_admin_count,
                'icon': 'fa-user-shield',
            },
            {
                'label': 'Pending staff approvals',
                'value': pending_admin_count,
                'icon': 'fa-user-clock',
            },
            {
                'label': 'New users today',
                'value': todays_new_users,
                'icon': 'fa-magic',
            },
        ]

        admin_shortcuts = {
            'django_admin': reverse('admin:index'),
            'user_directory': reverse('admin:auth_user_changelist'),
        }

    quick_service_catalog, _service_desc_strings, _job_name_choices = build_service_job_catalog(request.user)

    context = {
        'total_pending_balance': total_pending_balance,
        'overdue_total_balance': overdue_total_balance,
        'expense_total': expense_total,
        'income_total': income_total,
        'payment_methods': PAYMENT_METHOD_OPTIONS,
        'monthly_income': mark_safe(json.dumps(monthly_income_data, cls=DjangoJSONEncoder)),
        'monthly_expenses': mark_safe(json.dumps(monthly_expenses_data, cls=DjangoJSONEncoder)),
        'weekly_income': mark_safe(json.dumps(weekly_income_data, cls=DjangoJSONEncoder)),
        'weekly_expenses': mark_safe(json.dumps(weekly_expenses_data, cls=DjangoJSONEncoder)),
        'yearly_income': mark_safe(json.dumps(yearly_income_data, cls=DjangoJSONEncoder)),
        'yearly_expenses': mark_safe(json.dumps(yearly_expenses_data, cls=DjangoJSONEncoder)),
        'default_interval': default_interval,
        'has_graph_data': has_graph_data,
        'recent_notes': recent_notes,
        'note_form': note_form,
        'completion_percentage': completion_percentage,
        'term_days': term_days,
        'total_initial_notification_count': total_initial_notification_count,
        'new_appointments_count': new_appointments_count,
        'new_contact_messages_count': new_contact_messages_count,
        'quickbooks_settings': quickbooks_settings,
        'quickbooks_status': quickbooks_status,
        'recent_activity_logs': recent_activity_logs,
        'has_additional_activity': has_additional_activity,
        'show_recent_activity': show_recent_activity,
        'actual_user': actual_user,
        'pending_admin_profiles': pending_admin_profiles,
        'can_manage_admins': can_manage_admins,
        'superuser_summary_cards': superuser_summary_cards,
        'admin_shortcuts': admin_shortcuts,
        'service_count': service_count,
        'open_workorders_count': open_workorders_count,
        'active_mechanics_count': active_mechanics_count,
        'upcoming_maintenance': upcoming_maintenance,
        'recent_workorders': recent_workorders,
        'recent_invoices': recent_invoices,
        'top_pending_invoices': top_pending_invoices,
        'pending_invoices_count': pending_invoices_count,
        'pending_invoices_total': pending_invoices_total,
        'overdue_balance_last_365_days': overdue_balance_last_365_days,
        'pending_invoices_this_year_count': pending_invoices_this_year_count,
        'current_year': current_year,
        'recent_bills_total': recent_bills_total,
        'quick_customers': quick_customers,
        'quick_vehicles': quick_vehicles,
        'quick_mechanics': quick_mechanics,
        'today': today,
        'today_tasks_count': today_tasks_count,
        'this_week_invoices_count': this_week_invoices_count,
        'this_week_earnings_paid': this_week_earnings_paid,
        'todays_invoices_count': todays_invoices_count,
        'supplier_names': supplier_names,
        'pending_customer_approvals': pending_customer_approvals,
        'pending_customer_approvals_count': pending_customer_approvals_count,
        'top_suppliers': top_suppliers,
        'category_group_sales': category_group_sales,
        'category_group_sales_total': category_group_sales_total,
        'inventory_products_data': inventory_products_data,
        'default_tax_rate': default_tax_rate,
        'tax_label': province_code,
        'workorder_volume_chart': mark_safe(json.dumps(workorder_volume_chart, cls=DjangoJSONEncoder)),
        'service_revenue_chart': mark_safe(json.dumps(service_revenue_chart, cls=DjangoJSONEncoder)),
        'quick_vehicles_json': json.dumps([
            {
                "id": v.id,
                "label": (v.unit_number or "Unit") + " " + (v.make_model or v.vin_number or ""),
                "customer_id": v.customer_id,
                "unit_number": v.unit_number,
                "make_model": v.make_model,
                "vin": v.vin_number,
            }
            for v in quick_vehicles
        ]),
        'quick_customers_json': json.dumps(
            [
                {
                    "id": c.id,
                    "name": c.name,
                    "email": c.email or "",
                    "address": c.address or "",
                }
                for c in quick_customers
            ],
            cls=DjangoJSONEncoder,
        ),
        'quick_service_catalog_json': json.dumps(
            quick_service_catalog,
            cls=DjangoJSONEncoder,
        ),
        'overdue_customers_json': json.dumps(
            overdue_customers_payload,
            cls=DjangoJSONEncoder,
        ),
        'overdue_customers_count': overdue_customers_count,
        'overdue_customers_total': float(total_overdue_customers_amount),
        'google_maps_api_key': getattr(settings, 'GOOGLE_MAPS_API_KEY', ''),
    }
    return render(request, template_name, context)


ACTIVITY_CATEGORY_CHOICES = (
    ('', 'All activity'),
    ('invoices', 'Invoices'),
    ('workorders', 'Work Orders'),
    ('inventory', 'Inventory'),
    ('payments', 'Payments'),
    ('customers', 'Customers'),
    ('system', 'System events'),
)

ACTIVITY_CATEGORY_QUERIES = {
    'invoices': Q(object_type__icontains='invoice') | Q(object_type__icontains='estimate'),
    'workorders': Q(object_type__icontains='workorder') | Q(object_type__icontains='work order'),
    'inventory': Q(object_type__icontains='inventory') | Q(object_type__icontains='product') | Q(object_type__icontains='part'),
    'payments': Q(object_type__icontains='payment') | Q(object_type__icontains='payout'),
    'customers': Q(object_type__icontains='customer') | Q(object_type__icontains='client') | Q(object_type__icontains='contact'),
    'system': Q(actor__isnull=True),
}


@login_required
def activity_log_list(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        return render(request, 'accounts/profile_missing.html')

    business_user = profile.get_business_user() if hasattr(profile, 'get_business_user') else request.user

    base_logs = ActivityLog.objects.filter(business=business_user)
    actor_ids = (
        base_logs.exclude(actor__isnull=True)
        .values_list('actor_id', flat=True)
        .distinct()
    )
    actors = User.objects.filter(id__in=actor_ids).order_by('first_name', 'last_name', 'username')

    logs = base_logs.select_related('actor').order_by('-created_at')

    query = request.GET.get('q', '').strip()
    if query:
        logs = logs.filter(
            Q(description__icontains=query)
            | Q(action__icontains=query)
            | Q(object_type__icontains=query)
        )

    actor_filter = request.GET.get('actor', '').strip()
    if actor_filter:
        logs = logs.filter(actor_id=actor_filter)

    category_filter = request.GET.get('category', '').strip()
    if category_filter and category_filter in ACTIVITY_CATEGORY_QUERIES:
        logs = logs.filter(ACTIVITY_CATEGORY_QUERIES[category_filter])

    start_filter = request.GET.get('start', '').strip()
    if start_filter:
        start_dt = parse_datetime(start_filter)
        if start_dt:
            if timezone.is_naive(start_dt):
                start_dt = timezone.make_aware(start_dt, timezone.get_current_timezone())
            logs = logs.filter(created_at__gte=start_dt)

    end_filter = request.GET.get('end', '').strip()
    if end_filter:
        end_dt = parse_datetime(end_filter)
        if end_dt:
            if timezone.is_naive(end_dt):
                end_dt = timezone.make_aware(end_dt, timezone.get_current_timezone())
            logs = logs.filter(created_at__lte=end_dt)

    total_count = logs.count()

    paginator = Paginator(logs, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    for log in page_obj.object_list:
        if isinstance(log.metadata, Mapping):
            log.metadata_items = list(log.metadata.items())
            log.metadata_text = None
        else:
            log.metadata_items = None
            log.metadata_text = log.metadata

    preserved_params = request.GET.copy()
    preserved_params.pop('page', None)
    filters_querystring = preserved_params.urlencode()

    context = {
        'activity_logs': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_count': total_count,
        'actors': actors,
        'category_choices': ACTIVITY_CATEGORY_CHOICES,
        'filters': {
            'q': query,
            'actor': actor_filter,
            'category': category_filter,
            'start': start_filter,
            'end': end_filter,
        },
        'filters_querystring': filters_querystring,
    }
    return render(request, 'accounts/mach/activity_log_list.html', context)


@login_required
@require_POST
def approve_admin_profile(request, profile_id):
    """Approve a pending staff request."""

    actual_user = getattr(request, 'actual_user', request.user)
    business_user = get_business_user(request.user)

    if not (actual_user and (actual_user.is_superuser or actual_user == business_user)):
        return HttpResponseForbidden("You do not have permission to approve staff accounts.")

    profile = get_object_or_404(
        Profile.objects.select_related('user'),
        pk=profile_id,
        is_business_admin=True,
        business_owner=business_user,
    )

    if profile.admin_approved:
        messages.info(request, f"{profile.user.get_full_name() or profile.user.username} is already approved.")
    else:
        profile.admin_approved = True
        profile.save(update_fields=['admin_approved'])
        messages.success(
            request,
            f"{profile.user.get_full_name() or profile.user.username} has been approved as a staff member.",
        )

    return redirect('accounts:home')


@login_required
def get_report_data(
    request,
    year_param,
    start_month_param,
    end_month_param,
    month_selected,
    specific_date=None,
    start_date_param=None,
    end_date_param=None,
    report_user=None,
):
    report_user = report_user or request.user
    # Initial QuerySets filtered by user
    expenses = ExpenseRecord.objects.filter(user=report_user)
    mach_expense_items = (
        MechExpenseItem.objects.filter(
            mech_expense__user=report_user,
            mech_expense__paid=True,
        )
        .annotate(last_paid_at=Max('mech_expense__payments__created_at'))
        .annotate(
            report_date=Coalesce(F('mech_expense__date'), TruncDate('last_paid_at')),
            report_province=Coalesce(
                F('mech_expense__province'),
                F('mech_expense__user__profile__province'),
            ),
        )
    )
    province_rate_cases = [
        When(report_province=code, then=Value(float(rate)))
        for code, rate in PROVINCE_TAX_RATES.items()
        if code != "CU"
    ]
    tax_rate_expression = Case(
        When(
            report_province="CU",
            then=Coalesce(Cast("mech_expense__custom_tax_rate", FloatField()), Value(0.0)),
        ),
        *province_rate_cases,
        default=Value(0.0),
        output_field=FloatField(),
    )
    tax_multiplier = ExpressionWrapper(Value(1.0) + tax_rate_expression, output_field=FloatField())
    tax_ratio = ExpressionWrapper(tax_rate_expression / tax_multiplier, output_field=FloatField())
    report_tax_paid_expression = Case(
        When(
            mech_expense__tax_included=True,
            then=ExpressionWrapper(F("amount") * tax_ratio, output_field=FloatField()),
        ),
        default=ExpressionWrapper(F("amount") * tax_rate_expression, output_field=FloatField()),
        output_field=FloatField(),
    )
    mach_expense_items = mach_expense_items.annotate(
        report_tax_rate=tax_rate_expression,
        report_tax_paid=report_tax_paid_expression,
    )
    incomes = IncomeRecord.objects.filter(user=report_user)
    incomes_mach = GroupedInvoice.objects.filter(user=report_user)

    date_range_applied = False
    date_range_label = None

    if start_date_param or end_date_param:
        try:
            start_value = start_date_param or end_date_param
            end_value = end_date_param or start_date_param
            start_date = datetime.datetime.strptime(start_value, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_value, '%Y-%m-%d').date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
            date_range_applied = True
            if start_date == end_date:
                date_range_label = start_date.strftime('%b %d, %Y')
            else:
                date_range_label = f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
            expenses = expenses.filter(date__gte=start_date, date__lte=end_date)
            mach_expense_items = mach_expense_items.filter(
                report_date__gte=start_date,
                report_date__lte=end_date,
            )
            incomes = incomes.filter(date__gte=start_date, date__lte=end_date)
            incomes_mach = incomes_mach.filter(
                date_fully_paid__gte=start_date,
                date_fully_paid__lte=end_date,
            )
        except ValueError:
            date_range_applied = True
            date_range_label = "Invalid Range"

    if not date_range_applied and isinstance(specific_date, datetime.date):
        date_range_applied = True
        date_range_label = specific_date.strftime('%b %d, %Y')
        expenses = expenses.filter(date=specific_date)
        mach_expense_items = mach_expense_items.filter(report_date=specific_date)
        incomes = incomes.filter(date=specific_date)
        incomes_mach = incomes_mach.filter(date_fully_paid=specific_date)

    # Apply year filter if provided
    if year_param:
        try:
            year = int(year_param)
            expenses = expenses.filter(date__year=year)
            mach_expense_items = mach_expense_items.filter(report_date__year=year)
            incomes = incomes.filter(date__year=year)
            incomes_mach = incomes_mach.filter(date_fully_paid__year=year)
        except ValueError:
            year = None
    else:
        year = None

        # Annotate total_paid for each GroupedInvoice using a Subquery
    payments_subquery = Payment.objects.filter(invoice=OuterRef('pk')).values('invoice').annotate(
        total_paid=Sum('amount')
    ).values('total_paid')

    # Subquery to annotate total_tax_collected per invoice
    tax_collected_subquery = IncomeRecord2.objects.filter(
        grouped_invoice=OuterRef('pk')
    ).values('grouped_invoice').annotate(
        total_tax_collected=Sum('tax_collected')
    ).values('total_tax_collected')



    incomes_mach = incomes_mach.annotate(
        total_paid=Coalesce(Subquery(payments_subquery, output_field=DecimalField()), Decimal('0.00')),
        total_tax_collected=Coalesce(Subquery(tax_collected_subquery, output_field=DecimalField()), Decimal('0.00'))
    )


    # Apply month range filter if provided
    if not date_range_applied and start_month_param and end_month_param:
        try:
            start_date = datetime.datetime.strptime(start_month_param, '%Y-%m').date()
            end_date = datetime.datetime.strptime(end_month_param, '%Y-%m').date()
             # Validate that start_date <= end_date
            if start_date > end_date:
                raise ValueError("Start month cannot be after end month.")
            # Adjust dates to fit within the selected year if applicable
            if year:
                if start_date.year != year:
                    start_date = datetime.datetime(year, start_date.month, 1).date()
                if end_date.year != year:
                    next_month = end_date.month % 12 + 1
                    next_year = end_date.year + (end_date.month // 12)
                    last_day = (datetime.datetime(next_year, next_month, 1) - timedelta(days=1)).day
                    end_date = datetime.datetime(year, end_date.month, last_day).date()
            # Filter QuerySets by the date range
            expenses = expenses.filter(date__gte=start_date, date__lte=end_date)
            mach_expense_items = mach_expense_items.filter(
                report_date__gte=start_date,
                report_date__lte=end_date
            )
            incomes = incomes.filter(date__gte=start_date, date__lte=end_date)
            incomes_mach = incomes_mach.filter(
                date_fully_paid__gte=start_date,
                date_fully_paid__lte=end_date
            )

            formatted_range = f"{start_date.strftime('%B')} - {end_date.strftime('%B')} {year}" if year else "Selected Range"
            is_all_months = False
            is_all_data = False
        except ValueError:
            formatted_range = "Invalid Range"
            is_all_months = False
            is_all_data = False
    elif not date_range_applied and year_param:
        # If only year is selected, show all months in that year
        formatted_range = f"All Months {year}"
        is_all_months = True
        is_all_data = False
    elif not date_range_applied:
        # If no year is selected, show all data
        formatted_range = "All Data"
        is_all_months = False
        is_all_data = True
    else:
        formatted_range = date_range_label or "Selected Date"
        is_all_months = False
        is_all_data = False

    # Apply month_selected filter if provided
    if not date_range_applied and month_selected:
        try:

            selected_month_date = datetime.datetime.strptime(month_selected, '%Y-%m')
            # Filter QuerySets for the selected month
            incomes_mach = incomes_mach.filter(
                date_fully_paid__year=selected_month_date.year,
                date_fully_paid__month=selected_month_date.month
            )
            mach_expense_items = mach_expense_items.filter(
                report_date__year=selected_month_date.year,
                report_date__month=selected_month_date.month
            )
            formatted_range = selected_month_date.strftime('%B %Y')
            is_all_months = False
            is_all_data = False
        except ValueError:
            formatted_range = "Invalid Month Selected"
            is_all_months = False
            is_all_data = False

    if date_range_applied:
        grouped_incomes = incomes_mach.values('date_fully_paid').annotate(
            total_invoices=Count('id', distinct=True),
            total_amount=Sum('total_paid'),
            total_tax_collected=Sum('total_tax_collected')
        ).order_by('date_fully_paid')

        grouped_expenses = mach_expense_items.values(
            mech_expense__date=F('report_date')
        ).annotate(
            total_entries=Count('id', distinct=True),
            total_amount=Sum('amount'),
            total_tax_paid=Sum('report_tax_paid')
        ).order_by('mech_expense__date')
    else:
        # Grouped incomes with aggregation
        grouped_incomes = incomes_mach.values('date_fully_paid').annotate(
            total_invoices=Count('id', distinct=True),
            total_amount=Sum('total_paid'),
            total_tax_collected=Sum('total_tax_collected')
        ).order_by('date_fully_paid')



        # Grouped expenses with aggregation
        grouped_expenses = mach_expense_items.values(
            mech_expense__date=F('report_date')
        ).annotate(
            total_entries=Count('id', distinct=True),
            total_amount=Sum('amount'),
            total_tax_paid=Sum('report_tax_paid')
        ).order_by('mech_expense__date')


    # Calculate various totals using utility functions
    total_expense = to_decimal(expenses.aggregate(total=Sum('total'))['total'])
    total_income = to_decimal(incomes.aggregate(total=Sum('amount'))['total'])
    incomes_mach_total = to_decimal(
        incomes_mach.aggregate(total=Sum('total_paid'))['total']
    )
    total_tax_paid = to_decimal(expenses.aggregate(total=Sum('tax_paid'))['total'])
    total_tax_paid_mach = to_decimal(mach_expense_items.aggregate(total=Sum('report_tax_paid'))['total'])
    total_tax_collected = to_decimal(incomes.aggregate(total=Sum('tax_collected'))['total'])
    total_tax_collected_mach = to_decimal(
        incomes_mach.aggregate(total=Sum('total_tax_collected'))['total']
    )
    total_tax_difference = total_tax_collected - total_tax_paid
    total_tax_difference_mach = total_tax_collected_mach - total_tax_paid_mach

    # Aggregating expenses and incomes
    expense_totals = expenses.aggregate(
        total_fuel=Sum('fuel'),
        total_plates=Sum('plates'),
        total_wsib=Sum('wsib'),
        total_repairs=Sum('repairs'),
        total_parking=Sum('parking'),
        total_wash=Sum('wash'),
        total_def=Sum('def_fluid'),
        total_insurance=Sum('insurance'),
        total_expense=Sum('total'),
        total_tax_paid=Sum('tax_paid')
    )
    expense_totals = {k: to_decimal(v) for k, v in expense_totals.items()}

    mach_total_expenses = mach_expense_items.aggregate(
        total_entries=Count('id', distinct=True),
        total_amount=Sum('amount'),
        total_tax_paid=Sum('report_tax_paid')
    )
    mach_total_expenses = {k: to_decimal(v) for k, v in mach_total_expenses.items()}


    income_totals = incomes.aggregate(
        total_invoices=Count('id', distinct=True),
        total_amount=Sum('amount'),
        total_tax_collected=Sum('tax_collected')
    )
    income_totals = {k: to_decimal(v) for k, v in income_totals.items()}



    # Extract distinct years for filtering options
    expense_years = expenses.dates('date', 'year', order='DESC')
    mach_report_dates = list(mach_expense_items.values_list('report_date', flat=True).distinct())
    mach_expense_years = sorted(
        {datetime.date(d.year, 1, 1) for d in mach_report_dates if d},
        reverse=True,
    )
    income_years = incomes.dates('date', 'year', order='DESC')
    income_mach_years = incomes_mach.dates('date_fully_paid', 'year', order='DESC')

    all_years = sorted(
        set(expense_years) | set(mach_expense_years) | set(income_years) | set(income_mach_years),
        reverse=True
    )

    # Generate a list of month objects with pre-formatted 'value' keys
    month_list = [
        {'number': 1, 'name': 'January'},
        {'number': 2, 'name': 'February'},
        {'number': 3, 'name': 'March'},
        {'number': 4, 'name': 'April'},
        {'number': 5, 'name': 'May'},
        {'number': 6, 'name': 'June'},
        {'number': 7, 'name': 'July'},
        {'number': 8, 'name': 'August'},
        {'number': 9, 'name': 'September'},
        {'number': 10, 'name': 'October'},
        {'number': 11, 'name': 'November'},
        {'number': 12, 'name': 'December'},
    ]

    # Add 'value' key to each month
    for month in month_list:
        if year_param:
            month['value'] = f"{year_param}-{month['number']:02d}"
        else:
            month['value'] = f"{month['number']:02d}"

    # Monthly data logic (grouped by month)
    expense_months = expenses.annotate(month=TruncMonth('date')).values_list('month', flat=True).distinct()
    mach_expense_months = {
        datetime.date(d.year, d.month, 1)
        for d in mach_report_dates
        if d is not None
    }
    income_months = incomes.annotate(month=TruncMonth('date')).values_list('month', flat=True).distinct()
    income_mach_months = incomes_mach.annotate(month=TruncMonth('date_fully_paid')).values_list('month', flat=True).distinct()

    all_months = sorted(
        month for month in (set(expense_months) | set(mach_expense_months) | set(income_months) | set(income_mach_months))
        if month is not None
    )

    monthly_data = []
    for m in all_months:
        month_str = m.strftime('%Y-%m')
        month_incomes_mach = incomes_mach.filter(
            date_fully_paid__year=m.year,
            date_fully_paid__month=m.month
        )

        this_month_incomes_mach = to_decimal(
            month_incomes_mach.aggregate(total=Sum('total_paid'))['total']
        )
        this_month_tax_collected_mach = to_decimal(
            month_incomes_mach.aggregate(total=Sum('total_tax_collected'))['total']
        )
        this_month_expenses_mach = to_decimal(
            mach_expense_items.filter(
                report_date__year=m.year,
                report_date__month=m.month
            ).aggregate(total=Sum('amount'))['total']
        )
        this_month_tax_paid_mach = to_decimal(
            mach_expense_items.filter(
                report_date__year=m.year,
                report_date__month=m.month
            ).aggregate(total=Sum('report_tax_paid'))['total']
        )

        margin_mach = this_month_incomes_mach - this_month_expenses_mach
        tax_difference_mach = this_month_tax_collected_mach - this_month_tax_paid_mach

        monthly_data.append({
            'month': month_str,
            'total_mach_income': format_currency(this_month_incomes_mach),
            'total_mach_expense': format_currency(this_month_expenses_mach),
            'margin_mach': format_currency(margin_mach),
            'total_tax_collected_mach': format_currency(this_month_tax_collected_mach),
            'total_tax_paid_mach': format_currency(this_month_tax_paid_mach),
            'tax_difference_mach': format_currency(tax_difference_mach),
        })


    # Calculate total_mac_margin
    total_mac_margin = incomes_mach_total - mach_total_expenses.get('total_amount', Decimal('0.00'))

    # Assemble the context dictionary
    context = {
        'expenses': expenses,
        'mach_expense_items': mach_expense_items,
        'incomes': incomes,
        'incomes_mach': incomes_mach,
        'expense_totals': {k: format_currency(v) if 'total' in k else v for k, v in expense_totals.items()},
        'mach_total_expenses': mach_total_expenses,
        'total_mach_income': format_currency(incomes_mach_total),
        'total_mac_margin': format_currency(total_mac_margin),
        'total_mach_expenses': format_currency(mach_total_expenses.get('total_amount', Decimal('0.00'))),
        'total_tax_paid_mach': format_currency(total_tax_paid_mach),
        'total_tax_collected_mach': format_currency(total_tax_collected_mach),
        'total_tax_difference_mach': format_currency(total_tax_difference_mach),
        'formatted_month': formatted_range,
        'months': month_list,
        'income_totals': income_totals,
        'monthly_data': monthly_data,
        'years': all_years,
        'selected_year': year_param,
        'selected_start_month': start_month_param,
        'selected_end_month': end_month_param,
        'is_all_months': is_all_months,
        'is_all_data': is_all_data,
        'month_selected': month_selected,
        'grouped_incomes': grouped_incomes,
        'grouped_expenses': grouped_expenses,
        'specific_date': specific_date,
    }

    return context

@login_required
def income_details_by_date(request):
    date_str = request.GET.get('date')
    month_selected = request.GET.get('month')
    try:
        specific_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return HttpResponse("Invalid date.")

    report_user = request.user
    portal_read_only = False
    accountant_profile = getattr(request.user, "accountant_portal", None)
    if accountant_profile:
        detail_access = accountant_profile.accountant_access_level in ("full", "read_only")
        if not detail_access:
            raise Http404
        report_user = accountant_profile.user
        portal_read_only = True

    # Reuse get_report_data to get data for the specific date
    context = get_report_data(
        request,
        year_param=None,
        start_month_param=None,
        end_month_param=None,
        month_selected=month_selected,
        specific_date=specific_date,
        report_user=report_user,
    )

    if specific_date:
        detailed_incomes = (
            GroupedInvoice.objects.filter(
                user=report_user,
                date_fully_paid=specific_date,
            )
            .annotate(total_tax_collected=Sum('income_records__tax_collected'))
            .order_by('date_fully_paid')
        )
        context['grouped_incomes'] = detailed_incomes
    context['portal_read_only'] = portal_read_only

    return render(request, 'accounts/mach/income_details_by_date.html', context)

@login_required
def expense_details_by_date(request):
    date_str = request.GET.get('date')
    month_selected = request.GET.get('month')
    try:
        specific_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return HttpResponse("Invalid date.")

    report_user = request.user
    portal_read_only = False
    accountant_profile = getattr(request.user, "accountant_portal", None)
    if accountant_profile:
        detail_access = accountant_profile.accountant_access_level in ("full", "read_only")
        if not detail_access:
            raise Http404
        report_user = accountant_profile.user
        portal_read_only = True

    # Reuse get_report_data to get data for the specific date
    context = get_report_data(
        request,
        year_param=None,
        start_month_param=None,
        end_month_param=None,
        month_selected=month_selected,
        specific_date=specific_date,
        report_user=report_user,
    )
    expense_items = context.get('mach_expense_items')
    if expense_items is not None:
        context['expense_receipts'] = (
            expense_items.values('mech_expense_id')
            .annotate(
                receipt_no=Max('mech_expense__receipt_no'),
                receipt_date=Max('mech_expense__date'),
                total_tax=Coalesce(Sum('report_tax_paid'), Value(0.0)),
                total_amount=Coalesce(Sum('amount'), Value(0.0)),
            )
            .order_by('receipt_date', 'mech_expense_id')
        )
    else:
        context['expense_receipts'] = []
    context['portal_read_only'] = portal_read_only

    return render(request, 'accounts/mach/expense_details_by_date.html', context)


@login_required
def tables(request):
    template_name = 'accounts/mach/tables.html'
    year_param = request.GET.get('year', '')
    start_month_param = request.GET.get('start_month', '')
    end_month_param = request.GET.get('end_month', '')
    month_selected = request.GET.get('month', '')  # For detailed view

    context = get_report_data(request, year_param, start_month_param, end_month_param, month_selected)

    return render(request, template_name, context)

@login_required
def download_report(request, report_user=None):
    # Extract parameters from GET request
    format_param = request.GET.get('format', 'pdf').lower()
    report_types = request.GET.getlist('report_type')
    date_param = request.GET.get('date')
    month_selected = request.GET.get('month')
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')

    # Extract filter parameters
    year_param = request.GET.get('year', '')
    start_month_param = request.GET.get('start_month', '')
    end_month_param = request.GET.get('end_month', '')

    # Handle date-specific reports
    specific_date = None
    if date_param:
        try:
            specific_date = datetime.datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD.")

    # Fetch data using your existing logic
    if report_user is None:
        accountant_profile = getattr(request.user, "accountant_portal", None)
        if accountant_profile:
            report_user = accountant_profile.user
        else:
            report_user = request.user
    context = get_report_data(
        request,
        year_param=year_param,
        start_month_param=start_month_param,
        end_month_param=end_month_param,
        month_selected=month_selected,
        specific_date=specific_date,
        start_date_param=start_date_param,
        end_date_param=end_date_param,
        report_user=report_user,
    )

    if not report_types:
        return HttpResponse("No report type selected.")

    # Determine the filename based on available parameters
    if specific_date:
        filename_base = f"report_{specific_date}"
    elif month_selected:
        filename_base = f"report_{month_selected}"
    else:
        filename_base = "report"

    if format_param == 'pdf':
        # Generate PDF using WeasyPrint
        template_path = 'accounts/report_pdf_template.html'  # Ensure this path is correct
        response = HttpResponse(content_type='application/pdf')
        filename = f"{filename_base}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Render the HTML template with context
        html_string = render_to_string(
            template_path,
            {**context, 'report_types': report_types},
            request=request,
        )

        if not WEASYPRINT_AVAILABLE:
            raise ImportError("WeasyPrint is not available. PDF generation is disabled. Please install GTK+ libraries for Windows.")

        # Create WeasyPrint HTML object
        html = HTML(string=html_string, base_url=request.build_absolute_uri())

        # Generate PDF
        try:
            html.write_pdf(target=response)
        except Exception as e:
            return HttpResponse(f'WeasyPrint Error: {str(e)}')

        return response

    elif format_param == 'excel':
        # Generate Excel using xlwt
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = f"{filename_base}.xls"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = xlwt.Workbook(encoding='utf-8')
        base_border = 'borders: left thin, right thin, top thin, bottom thin;'
        header_style = xlwt.easyxf(
            'font: bold on; align: horiz center, vert center; '
            'pattern: pattern solid, fore_colour ice_blue; '
            f'{base_border}'
        )
        text_style = xlwt.easyxf(f'align: horiz left, vert center; {base_border}')
        int_style = xlwt.easyxf(f'align: horiz right, vert center; {base_border}', num_format_str='0')
        money_style = xlwt.easyxf(f'align: horiz right, vert center; {base_border}', num_format_str='$#,##0.00')
        total_style = xlwt.easyxf(f'font: bold on; align: horiz right, vert center; {base_border}', num_format_str='$#,##0.00')

        def normalize_text(value):
            if value is None:
                return ''
            if isinstance(value, (int, float, Decimal)):
                return f"{value:,.2f}"
            return str(value)

        def track_width(widths, col_idx, value):
            text_value = normalize_text(value)
            widths[col_idx] = max(widths[col_idx], len(text_value))

        def write_cell(ws, widths, row_idx, col_idx, value, style):
            ws.write(row_idx, col_idx, value, style)
            track_width(widths, col_idx, value)

        def apply_widths(ws, widths):
            for idx, width in enumerate(widths):
                adjusted = min(60, max(10, width + 2))
                ws.col(idx).width = int(adjusted * 256)

        for report_type in report_types:
            if report_type == 'income_expense':
                ws = wb.add_sheet('Income & Expenses')
                columns = ['Month', 'Total Income', 'Total Expenses', 'Net Income']
                widths = [len(col) for col in columns]
                ws.panes_frozen = True
                ws.horz_split_pos = 1
                for col_num, column_title in enumerate(columns):
                    write_cell(ws, widths, 0, col_num, column_title, header_style)

                monthly_data = context.get('monthly_data', [])
                for row_num, data in enumerate(monthly_data, start=1):
                    write_cell(ws, widths, row_num, 0, data['month'], text_style)
                    write_cell(ws, widths, row_num, 1, float(data['total_mach_income'].replace(',', '').replace('$', '')), money_style)
                    write_cell(ws, widths, row_num, 2, float(data['total_mach_expense'].replace(',', '').replace('$', '')), money_style)
                    write_cell(ws, widths, row_num, 3, float(data['margin_mach'].replace(',', '').replace('$', '')), money_style)
                apply_widths(ws, widths)

            elif report_type == 'tax_report':
                ws = wb.add_sheet('Tax Reports')
                columns = ['Month', 'Tax Collected', 'Tax Paid', 'Difference']
                widths = [len(col) for col in columns]
                ws.panes_frozen = True
                ws.horz_split_pos = 1
                for col_num, column_title in enumerate(columns):
                    write_cell(ws, widths, 0, col_num, column_title, header_style)

                monthly_data = context.get('monthly_data', [])
                for row_num, data in enumerate(monthly_data, start=1):
                    write_cell(ws, widths, row_num, 0, data['month'], text_style)
                    write_cell(ws, widths, row_num, 1, float(data['total_tax_collected_mach'].replace(',', '').replace('$', '')), money_style)
                    write_cell(ws, widths, row_num, 2, float(data['total_tax_paid_mach'].replace(',', '').replace('$', '')), money_style)
                    write_cell(ws, widths, row_num, 3, float(data['tax_difference_mach'].replace(',', '').replace('$', '')), money_style)
                apply_widths(ws, widths)

            elif report_type in ['detailed_income_date', 'detailed_income']:
                ws = wb.add_sheet('Detailed Incomes')
                columns = ['Paid Date', 'Number of Invoices', 'Total Tax', 'Total Amount']
                widths = [len(col) for col in columns]
                ws.panes_frozen = True
                ws.horz_split_pos = 1
                for col_num, column_title in enumerate(columns):
                    write_cell(ws, widths, 0, col_num, column_title, header_style)

                grouped_incomes = context.get('grouped_incomes', [])
                for row_num, income in enumerate(grouped_incomes, start=1):
                    paid_date = income.get('date_fully_paid')
                    paid_date_display = paid_date.strftime('%Y-%m-%d') if paid_date else 'N/A'
                    write_cell(ws, widths, row_num, 0, paid_date_display, text_style)
                    write_cell(ws, widths, row_num, 1, income['total_invoices'], int_style)
                    write_cell(ws, widths, row_num, 2, float(income['total_tax_collected']) if income['total_tax_collected'] else 0.00, money_style)
                    write_cell(ws, widths, row_num, 3, float(income['total_amount']) if income['total_amount'] else 0.00, money_style)
                apply_widths(ws, widths)

            elif report_type in ['detailed_expense_date', 'detailed_expense']:
                ws = wb.add_sheet('Detailed Expenses')
                columns = ['Date', 'Vendor', 'Part No', 'Price', 'Qty', 'Tax', 'Total']
                widths = [len(col) for col in columns]
                ws.panes_frozen = True
                ws.horz_split_pos = 1
                for col_num, column_title in enumerate(columns):
                    write_cell(ws, widths, 0, col_num, column_title, header_style)

                expense_items = context.get('mach_expense_items', [])
                for row_num, expense_item in enumerate(expense_items, start=1):
                    report_date = getattr(expense_item, 'report_date', None)
                    expense_date = expense_item.mech_expense.date
                    date_display = (expense_date or report_date).strftime('%Y-%m-%d') if (expense_date or report_date) else 'N/A'
                    write_cell(ws, widths, row_num, 0, date_display, text_style)
                    write_cell(ws, widths, row_num, 1, expense_item.mech_expense.vendor, text_style)
                    write_cell(ws, widths, row_num, 2, expense_item.part_no, text_style)
                    write_cell(ws, widths, row_num, 3, float(expense_item.price), money_style)
                    write_cell(ws, widths, row_num, 4, expense_item.qty, int_style)
                    tax_paid_value = getattr(expense_item, 'report_tax_paid', expense_item.tax_paid)
                    write_cell(ws, widths, row_num, 5, float(tax_paid_value or 0), money_style)
                    write_cell(ws, widths, row_num, 6, float(expense_item.amount), money_style)
                apply_widths(ws, widths)

            # Add additional report types here as needed

        try:
            wb.save(response)
        except IndexError as e:
            return HttpResponse(f"Error generating Excel report: {str(e)}")
        return response

    else:
        return HttpResponse("Invalid format selected. Please choose 'pdf' or 'excel'.")


def _connected_business_display_name(user):
    profile = getattr(user, "profile", None)
    if profile and getattr(profile, "company_name", None):
        return profile.company_name
    return user.get_full_name() or user.get_username()


def _build_connected_business_context(request):
    business_user = get_business_user(request.user) or request.user
    group = get_connected_business_group(business_user)
    members = []
    if group:
        members = list(group.members.select_related("profile").order_by("username"))

    connected_members = []
    for member in members:
        connected_members.append(
            {
                "id": member.id,
                "username": member.get_username(),
                "display_name": _connected_business_display_name(member),
                "is_current": member.id == business_user.id,
            }
        )

    can_manage = request.user.is_superuser or request.user == business_user
    return {
        "connected_group": group,
        "connected_members": connected_members,
        "connected_business_user": business_user,
        "connected_can_manage": can_manage,
    }


def _can_manage_connected_business(request, business_user):
    return request.user.is_superuser or request.user == business_user


def _first_form_error(form, fallback):
    if not form.errors:
        return fallback
    for field, errors in form.errors.items():
        if not errors:
            continue
        label = form.fields.get(field).label if field != "__all__" else "Form"
        return f"{label}: {errors[0]}"
    return fallback


def _copy_inventory_catalog(source_user, target_user):
    results = {
        "groups_created": 0,
        "categories_created": 0,
        "attributes_created": 0,
        "options_created": 0,
        "suppliers_created": 0,
        "brands_created": 0,
        "models_created": 0,
        "vins_created": 0,
        "locations_created": 0,
        "products_created": 0,
        "products_skipped": 0,
        "alternate_skus_created": 0,
        "attribute_values_created": 0,
    }

    def _key(value):
        return (value or "").strip().lower()

    group_map = {}
    target_groups = {
        _key(group.name): group
        for group in CategoryGroup.objects.filter(user=target_user)
    }
    for source_group in CategoryGroup.objects.filter(user=source_user):
        key = _key(source_group.name)
        target_group = target_groups.get(key)
        if not target_group:
            target_group = CategoryGroup.objects.create(
                user=target_user,
                name=source_group.name,
                description=source_group.description,
                image=source_group.image,
                sort_order=source_group.sort_order,
                is_active=source_group.is_active,
            )
            target_groups[key] = target_group
            results["groups_created"] += 1
        group_map[source_group.id] = target_group

    category_map = {}
    pending_parents = []
    target_categories = {
        _key(category.name): category
        for category in Category.objects.filter(user=target_user)
    }
    for source_category in Category.objects.filter(user=source_user).select_related("group", "parent"):
        key = _key(source_category.name)
        target_category = target_categories.get(key)
        if not target_category:
            target_category = Category.objects.create(
                user=target_user,
                name=source_category.name,
                description=source_category.description,
                image=source_category.image,
                sort_order=source_category.sort_order,
                is_active=source_category.is_active,
                group=group_map.get(source_category.group_id),
            )
            target_categories[key] = target_category
            results["categories_created"] += 1
            if source_category.parent_id:
                pending_parents.append((target_category, source_category.parent_id))
        category_map[source_category.id] = target_category

    for target_category, source_parent_id in pending_parents:
        parent = category_map.get(source_parent_id)
        if parent and target_category.parent_id != parent.id:
            target_category.parent = parent
            target_category.save(update_fields=["parent"])

    attribute_map = {}
    target_attributes = defaultdict(dict)
    for attr in CategoryAttribute.objects.filter(user=target_user).select_related("category"):
        target_attributes[attr.category_id][_key(attr.name)] = attr

    for source_attr in CategoryAttribute.objects.filter(user=source_user).select_related("category"):
        target_category = category_map.get(source_attr.category_id)
        if not target_category:
            continue
        key = _key(source_attr.name)
        target_attr = target_attributes[target_category.id].get(key)
        if not target_attr:
            target_attr = CategoryAttribute.objects.create(
                user=target_user,
                category=target_category,
                name=source_attr.name,
                description=source_attr.description,
                value_unit=source_attr.value_unit,
                attribute_type=source_attr.attribute_type,
                is_filterable=source_attr.is_filterable,
                is_comparable=source_attr.is_comparable,
                sort_order=source_attr.sort_order,
                is_active=source_attr.is_active,
            )
            target_attributes[target_category.id][key] = target_attr
            results["attributes_created"] += 1
        attribute_map[source_attr.id] = target_attr

    option_map = {}
    target_options = defaultdict(dict)
    for option in CategoryAttributeOption.objects.filter(attribute__user=target_user).select_related("attribute"):
        target_options[option.attribute_id][_key(option.value)] = option

    for source_option in CategoryAttributeOption.objects.filter(attribute__user=source_user).select_related("attribute"):
        target_attr = attribute_map.get(source_option.attribute_id)
        if not target_attr:
            continue
        key = _key(source_option.value)
        target_option = target_options[target_attr.id].get(key)
        if not target_option:
            target_option = CategoryAttributeOption.objects.create(
                attribute=target_attr,
                value=source_option.value,
                sort_order=source_option.sort_order,
                is_active=source_option.is_active,
            )
            target_options[target_attr.id][key] = target_option
            results["options_created"] += 1
        option_map[source_option.id] = target_option

    supplier_map = {}
    target_suppliers = {
        _key(supplier.name): supplier
        for supplier in Supplier.objects.filter(user=target_user)
    }
    for source_supplier in Supplier.objects.filter(user=source_user):
        key = _key(source_supplier.name)
        target_supplier = target_suppliers.get(key)
        if not target_supplier:
            target_supplier = Supplier.objects.create(
                user=target_user,
                name=source_supplier.name,
                contact_person=source_supplier.contact_person,
                email=source_supplier.email,
                phone_number=source_supplier.phone_number,
                address=source_supplier.address,
            )
            target_suppliers[key] = target_supplier
            results["suppliers_created"] += 1
        supplier_map[source_supplier.id] = target_supplier

    brand_map = {}
    target_brands = {
        _key(brand.name): brand
        for brand in ProductBrand.objects.filter(user=target_user)
    }
    for source_brand in ProductBrand.objects.filter(user=source_user):
        key = _key(source_brand.name)
        target_brand = target_brands.get(key)
        if not target_brand:
            target_brand = ProductBrand.objects.create(
                user=target_user,
                name=source_brand.name,
                description=source_brand.description,
                logo=source_brand.logo,
                sort_order=source_brand.sort_order,
                is_active=source_brand.is_active,
            )
            target_brands[key] = target_brand
            results["brands_created"] += 1
        brand_map[source_brand.id] = target_brand

    model_map = {}
    target_models = {
        _key(model.name): model
        for model in ProductModel.objects.filter(user=target_user)
    }
    for source_model in ProductModel.objects.filter(user=source_user).select_related("brand"):
        key = _key(source_model.name)
        target_model = target_models.get(key)
        if not target_model:
            target_model = ProductModel.objects.create(
                user=target_user,
                brand=brand_map.get(source_model.brand_id),
                name=source_model.name,
                description=source_model.description,
                year_start=source_model.year_start,
                year_end=source_model.year_end,
                sort_order=source_model.sort_order,
                is_active=source_model.is_active,
            )
            target_models[key] = target_model
            results["models_created"] += 1
        model_map[source_model.id] = target_model

    vin_map = {}
    target_vins = {
        _key(vin.vin): vin
        for vin in ProductVin.objects.filter(user=target_user)
    }
    for source_vin in ProductVin.objects.filter(user=source_user):
        key = _key(source_vin.vin)
        target_vin = target_vins.get(key)
        if not target_vin:
            target_vin = ProductVin.objects.create(
                user=target_user,
                vin=source_vin.vin,
                description=source_vin.description,
                sort_order=source_vin.sort_order,
                is_active=source_vin.is_active,
            )
            target_vins[key] = target_vin
            results["vins_created"] += 1
        vin_map[source_vin.id] = target_vin

    target_locations = {
        _key(location.name): location
        for location in InventoryLocation.objects.filter(user=target_user)
    }
    for source_location in InventoryLocation.objects.filter(user=source_user):
        key = _key(source_location.name)
        target_location = target_locations.get(key)
        if not target_location:
            target_location = InventoryLocation.objects.create(
                user=target_user,
                name=source_location.name,
            )
            target_locations[key] = target_location
            results["locations_created"] += 1

    target_products_by_sku = {}
    target_products_by_name = {}
    for target_product in Product.objects.filter(user=target_user):
        if target_product.sku:
            target_products_by_sku[_key(target_product.sku)] = target_product
        name_key = _key(target_product.name)
        if name_key and name_key not in target_products_by_name:
            target_products_by_name[name_key] = target_product

    source_products_qs = (
        Product.objects.filter(user=source_user)
        .select_related("category", "supplier", "brand", "vehicle_model", "vin_number")
        .prefetch_related("alternate_skus", "attribute_values", "attribute_values__option")
    )
    for source_product in source_products_qs:
        sku_key = _key(source_product.sku) if source_product.sku else ""
        name_key = _key(source_product.name)
        target_product = None
        if sku_key:
            target_product = target_products_by_sku.get(sku_key)
        if not target_product and name_key:
            target_product = target_products_by_name.get(name_key)

        if target_product:
            results["products_skipped"] += 1
            continue

        sale_price = source_product.sale_price
        margin = source_product.margin
        promotion_price = source_product.promotion_price
        if sale_price is None and margin is None:
            sale_price = source_product.cost_price
        if promotion_price is not None and sale_price is not None and promotion_price > sale_price:
            promotion_price = sale_price

        target_product = Product.objects.create(
            user=target_user,
            sku=source_product.sku or None,
            name=source_product.name,
            description=source_product.description,
            item_type=source_product.item_type,
            category=category_map.get(source_product.category_id),
            supplier=supplier_map.get(source_product.supplier_id),
            brand=brand_map.get(source_product.brand_id),
            vehicle_model=model_map.get(source_product.vehicle_model_id),
            vin_number=vin_map.get(source_product.vin_number_id),
            source_name=source_product.source_name,
            source_url=source_product.source_url,
            source_product_id=source_product.source_product_id,
            cost_price=source_product.cost_price,
            sale_price=sale_price,
            promotion_price=promotion_price,
            margin=margin,
            quantity_in_stock=source_product.quantity_in_stock,
            reorder_level=source_product.reorder_level,
            image=source_product.image,
            is_published_to_store=source_product.is_published_to_store,
            is_featured=source_product.is_featured,
            warranty_expiry_date=source_product.warranty_expiry_date,
            warranty_length=source_product.warranty_length,
            location=source_product.location,
        )
        upsert_product_stock(
            target_product,
            target_user,
            quantity_in_stock=source_product.quantity_in_stock or 0,
            reorder_level=source_product.reorder_level or 0,
        )
        results["products_created"] += 1
        if sku_key:
            target_products_by_sku[sku_key] = target_product
        if name_key:
            target_products_by_name[name_key] = target_product

        for alt_sku in source_product.alternate_skus.all():
            if ProductAlternateSku.objects.filter(
                product=target_product,
                sku__iexact=alt_sku.sku,
            ).exists():
                continue
            ProductAlternateSku.objects.create(
                product=target_product,
                sku=alt_sku.sku,
                kind=alt_sku.kind,
                source_name=alt_sku.source_name,
            )
            results["alternate_skus_created"] += 1

        for attr_value in source_product.attribute_values.all():
            target_attr = attribute_map.get(attr_value.attribute_id)
            if not target_attr:
                continue
            option = option_map.get(attr_value.option_id) if attr_value.option_id else None
            _, created = ProductAttributeValue.objects.get_or_create(
                product=target_product,
                attribute=target_attr,
                defaults={
                    "option": option,
                    "value_text": attr_value.value_text,
                    "value_number": attr_value.value_number,
                    "value_boolean": attr_value.value_boolean,
                },
            )
            if created:
                results["attribute_values_created"] += 1

    return results


@login_required
def account_settings(request):
    user = request.user
    if not user.is_active:
        return redirect('accounts:subscription_cancelled')

    profile = user.profile
    invoice_sequence_form = InvoiceSequenceForm(
        initial={'invoice_sequence_next': profile.invoice_sequence_next}
    )
    qb_settings = QuickBooksSettings.objects.filter(user=user).first()
    clover_connection = CloverConnection.objects.filter(user=user).first()
    banking_settings = BankingIntegrationSettings.objects.filter(user=user).first()
    stripe_account = UserStripeAccount.objects.filter(user=user).first()

    stripe_configured = bool(
        getattr(settings, "STRIPE_SECRET_KEY", "") and getattr(settings, "STRIPE_PUBLISHABLE_KEY", "")
    )
    clover_configured = bool(
        getattr(settings, "CLOVER_CLIENT_ID", "") and getattr(settings, "CLOVER_REDIRECT_URI", "")
    )
    banking_configured = bool(
        getattr(settings, "FLINKS_CLIENT_ID", "") and getattr(settings, "FLINKS_CLIENT_SECRET", "")
    )
    qb_defaults = getattr(settings, "QUICKBOOKS_DEFAULTS", {}) or {}
    quickbooks_configured = bool(
        qb_defaults.get("client_id") and qb_defaults.get("client_secret") and qb_defaults.get("redirect_uri")
    )

    stripe_connected = bool(stripe_account and stripe_account.stripe_account_id)
    clover_connected = bool(clover_connection and clover_connection.is_configured)
    banking_connected = bool(
        banking_settings
        and banking_settings.enabled
        and BankConnection.objects.filter(user=user, status=BankConnection.STATUS_CONNECTED).exists()
    )
    quickbooks_connected = bool(qb_settings and qb_settings.is_configured)

    context = {
        'profile': profile,
        'quickbooks_settings': qb_settings,
        'clover_connection': clover_connection,
        'banking_settings': banking_settings,
        'stripe_account': stripe_account,
        'stripe_configured': stripe_configured,
        'clover_configured': clover_configured,
        'banking_configured': banking_configured,
        'quickbooks_configured': quickbooks_configured,
        'stripe_connected': stripe_connected,
        'clover_connected': clover_connected,
        'banking_connected': banking_connected,
        'quickbooks_connected': quickbooks_connected,
        'invoice_sequence_form': invoice_sequence_form,
    }
    context.update(_build_connected_business_context(request))

    return render(
        request,
        'registration/account_settings.html',
        context,
    )


@login_required
@require_POST
def connected_business_update(request):
    business_user = get_business_user(request.user) or request.user
    if not _can_manage_connected_business(request, business_user):
        messages.error(request, "Only the business owner can manage connected businesses.")
        return redirect('accounts:account_settings')

    group = get_connected_business_group(business_user)
    form = ConnectedBusinessGroupForm(request.POST, instance=group)
    if form.is_valid():
        connected_group = form.save(commit=False)
        is_new = connected_group.pk is None
        connected_group.save()
        if not connected_group.members.filter(pk=business_user.pk).exists():
            connected_group.members.add(business_user)
        if is_new:
            messages.success(request, "Connected business group created.")
        else:
            messages.success(request, "Connected business settings updated.")
    else:
        messages.error(
            request,
            _first_form_error(form, "Unable to update connected business settings."),
        )
    return redirect('accounts:account_settings')


@login_required
@require_POST
def connected_business_add_member(request):
    business_user = get_business_user(request.user) or request.user
    if not _can_manage_connected_business(request, business_user):
        messages.error(request, "Only the business owner can manage connected businesses.")
        return redirect('accounts:account_settings')

    identifier = (request.POST.get("identifier") or "").strip()
    if not identifier:
        messages.error(request, "Enter a username or email to connect.")
        return redirect('accounts:account_settings')

    UserModel = get_user_model()
    if "@" in identifier:
        target_user = UserModel.objects.filter(email__iexact=identifier).first()
    else:
        target_user = UserModel.objects.filter(username__iexact=identifier).first()

    if not target_user:
        messages.error(request, "No business account was found with that identifier.")
        return redirect('accounts:account_settings')

    target_business_user = get_business_user(target_user) or target_user
    if target_business_user == business_user:
        messages.info(request, "That business is already connected.")
        return redirect('accounts:account_settings')

    group = get_connected_business_group(business_user)
    if not group:
        group = ConnectedBusinessGroup.objects.create(name="")
        group.members.add(business_user)

    if group.members.filter(pk=target_business_user.pk).exists():
        messages.info(request, "That business is already in the connected group.")
        return redirect('accounts:account_settings')

    if ConnectedBusinessGroup.objects.filter(members=target_business_user).exclude(pk=group.pk).exists():
        messages.error(request, "That business already belongs to another connected group.")
        return redirect('accounts:account_settings')

    group.members.add(target_business_user)
    messages.success(request, "Business added to the connected group.")
    return redirect('accounts:account_settings')


@login_required
@require_POST
def connected_business_remove_member(request, member_id):
    business_user = get_business_user(request.user) or request.user
    if not _can_manage_connected_business(request, business_user):
        messages.error(request, "Only the business owner can manage connected businesses.")
        return redirect('accounts:account_settings')

    group = get_connected_business_group(business_user)
    if not group:
        messages.error(request, "No connected business group found.")
        return redirect('accounts:account_settings')

    member = get_object_or_404(User, pk=member_id)
    if not group.members.filter(pk=member.pk).exists():
        messages.error(request, "That business is not in your connected group.")
        return redirect('accounts:account_settings')

    group.members.remove(member)
    if group.members.count() == 0:
        group.delete()
        messages.success(request, "Connected business group removed.")
    else:
        messages.success(request, "Business removed from the connected group.")
    return redirect('accounts:account_settings')


@login_required
@require_POST
def connected_business_copy_inventory(request):
    business_user = get_business_user(request.user) or request.user
    if not _can_manage_connected_business(request, business_user):
        messages.error(request, "Only the business owner can copy inventory between businesses.")
        return redirect('accounts:account_settings')

    group = get_connected_business_group(business_user)
    if not group:
        messages.error(request, "Connect businesses before copying inventory.")
        return redirect('accounts:account_settings')

    try:
        source_id = int(request.POST.get("source_business_id", ""))
        target_id = int(request.POST.get("target_business_id", ""))
    except (TypeError, ValueError):
        messages.error(request, "Select both a source and destination business.")
        return redirect('accounts:account_settings')

    if source_id == target_id:
        messages.error(request, "Choose two different businesses to copy inventory.")
        return redirect('accounts:account_settings')

    member_ids = set(group.members.values_list("id", flat=True))
    if source_id not in member_ids or target_id not in member_ids:
        messages.error(request, "Both businesses must belong to the connected group.")
        return redirect('accounts:account_settings')

    source_user = get_object_or_404(User, pk=source_id)
    target_user = get_object_or_404(User, pk=target_id)

    try:
        with transaction.atomic():
            results = _copy_inventory_catalog(source_user, target_user)
    except Exception as exc:
        logger.exception("Failed to copy inventory from %s to %s", source_id, target_id)
        messages.error(request, f"Inventory copy failed: {exc}")
        return redirect('accounts:account_settings')

    total_created = sum(
        results[key]
        for key in (
            "groups_created",
            "categories_created",
            "attributes_created",
            "options_created",
            "suppliers_created",
            "brands_created",
            "models_created",
            "vins_created",
            "locations_created",
            "products_created",
            "alternate_skus_created",
            "attribute_values_created",
        )
    )

    if total_created == 0:
        messages.info(request, "No new inventory items were copied.")
        return redirect('accounts:account_settings')

    source_name = _connected_business_display_name(source_user)
    target_name = _connected_business_display_name(target_user)
    messages.success(
        request,
        (
            f"Inventory copied from {source_name} to {target_name}. "
            f"Products: {results['products_created']} added "
            f"({results['products_skipped']} already existed), "
            f"categories: {results['categories_created']}, "
            f"suppliers: {results['suppliers_created']}."
        ),
    )
    return redirect('accounts:account_settings')


@login_required
@require_POST
def update_invoice_sequence(request):
    profile = request.user.profile
    form = InvoiceSequenceForm(request.POST)
    if form.is_valid():
        invoice_sequence_next = form.cleaned_data.get('invoice_sequence_next')
        if invoice_sequence_next:
            candidate = GroupedInvoice._format_invoice_number(
                request.user,
                invoice_sequence_next,
            )
            if GroupedInvoice.objects.filter(
                user=request.user,
                invoice_number=candidate,
            ).exists():
                messages.error(
                    request,
                    'That invoice number is already in use. Choose a higher number.',
                )
                return redirect('accounts:account_settings')
        profile.invoice_sequence_next = invoice_sequence_next
        profile.save(update_fields=['invoice_sequence_next'])
        if invoice_sequence_next:
            messages.success(
                request,
                f'Invoice sequence updated. Next invoice will start at {invoice_sequence_next}.',
            )
        else:
            messages.success(
                request,
                'Invoice sequence reset to automatic numbering.',
            )
    else:
        messages.error(request, 'Enter a valid invoice sequence number.')
    return redirect('accounts:account_settings')


@login_required
def display_preferences(request):
    """Allow users to adjust their interface scale preferences."""
    profile = getattr(request.user, 'profile', None)
    if profile is None:
        messages.error(request, 'No profile is associated with this account.')
        return redirect('accounts:account_settings')

    if request.method == 'POST':
        form = DisplayPreferencesForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Interface scale updated successfully.')
            return redirect('accounts:display_preferences')
    else:
        form = DisplayPreferencesForm(instance=profile)

    return render(
        request,
        'registration/display_preferences.html',
        {
            'form': form,
            'current_portal_scale': profile.ui_scale_percentage,
            'current_public_scale': profile.ui_scale_public_percentage,
        },
    )


@login_required
def quickbooks_settings(request):
    """Display and update QuickBooks Online integration credentials."""

    settings_instance = QuickBooksSettings.objects.filter(user=request.user).first()

    if request.method == 'POST':
        form = QuickBooksSettingsForm(request.POST, instance=settings_instance)
        if form.is_valid():
            quickbooks_settings = form.save(commit=False)
            quickbooks_settings.user = request.user
            quickbooks_settings.save()

            action = request.POST.get('action')
            if quickbooks_settings.integration_type == QuickBooksSettings.INTEGRATION_DESKTOP:
                if action == 'refresh_tokens':
                    messages.info(request, 'QuickBooks Desktop does not require token refresh. Settings saved successfully.')
                elif action == 'clear_tokens':
                    messages.info(request, 'QuickBooks Desktop does not cache access tokens.')
                else:
                    messages.success(request, 'QuickBooks Desktop settings saved successfully.')
            else:
                if action == 'refresh_tokens':
                    try:
                        QuickBooksService(quickbooks_settings).ensure_access_token()
                        messages.success(request, 'QuickBooks access token refreshed successfully.')
                    except QuickBooksIntegrationError as exc:
                        messages.error(request, str(exc))
                elif action == 'clear_tokens':
                    quickbooks_settings.clear_cached_tokens()
                    messages.info(request, 'Cached QuickBooks tokens cleared. They will refresh on the next sync.')
                else:
                    messages.success(request, 'QuickBooks settings saved successfully.')

            return redirect('accounts:quickbooks_settings')
    else:
        defaults = getattr(settings, 'QUICKBOOKS_DEFAULTS', {})
        initial = {}
        quickbooks_fields = [
            'integration_type',
            'client_id',
            'client_secret',
            'realm_id',
            'redirect_uri',
            'environment',
            'refresh_token',
            'desktop_company_name',
            'auto_sync_enabled',
        ]
        if settings_instance:
            for field in quickbooks_fields:
                if field == 'auto_sync_enabled':
                    # Respect the stored checkbox value for existing configurations.
                    continue
                current_value = getattr(settings_instance, field, None)
                if current_value in (None, '') and field in defaults and defaults[field]:
                    initial[field] = defaults[field]
        else:
            initial = {field: defaults.get(field) for field in quickbooks_fields if defaults.get(field) is not None}

        form = QuickBooksSettingsForm(instance=settings_instance, initial=initial)

    sync_summary = request.session.pop('quickbooks_sync_summary', None)

    context = {
        'form': form,
        'quickbooks_settings': settings_instance,
        'quickbooks_sync_summary': sync_summary,
    }
    return render(request, 'accounts/quickbooks/settings.html', context)


@login_required
def quickbooks_disconnect(request):
    QuickBooksSettings.objects.filter(user=request.user).delete()
    QuickBooksConnection.objects.filter(user=request.user).delete()
    QuickBooksImportMap.objects.filter(user=request.user).delete()
    messages.success(request, 'QuickBooks disconnected.')
    return redirect('accounts:account_settings')


def _create_expense_from_bank_transaction(transaction, user):
    vendor_raw = transaction.merchant_name or transaction.description or "Bank transaction"
    vendor_limit = MechExpense._meta.get_field("vendor").max_length
    vendor = vendor_raw[:vendor_limit]
    expense_date = transaction.posted_at or transaction.authorized_at or timezone.localdate()
    amount = abs(Decimal(transaction.amount))
    profile = getattr(user, "profile", None)
    province = getattr(profile, "province", None)

    expense = MechExpense.objects.create(
        user=user,
        vendor=vendor,
        date=expense_date,
        categorie="Miscellaneous",
        paid=True,
        tax_included=True,
        province=province,
        record_in_inventory=False,
    )
    MechExpenseItem.objects.create(
        mech_expense=expense,
        description=transaction.description or vendor_raw,
        qty=1,
        price=float(amount),
    )
    MechExpensePayment.objects.create(
        mech_expense=expense,
        amount=amount,
        method="Bank Transfer",
        notes="Imported from banking transaction",
        recorded_by=user,
    )
    return expense


@login_required
def banking_settings(request):
    settings_instance = BankingIntegrationSettings.objects.filter(user=request.user).first()

    if request.method == 'POST':
        form = BankingIntegrationSettingsForm(request.POST, instance=settings_instance)
        if form.is_valid():
            banking_settings = form.save(commit=False)
            banking_settings.user = request.user
            if not banking_settings.provider:
                banking_settings.provider = BankingIntegrationSettings.PROVIDER_FLINKS
            banking_settings.save()
            messages.success(request, 'Banking integration settings saved successfully.')
            return redirect('accounts:banking_settings')
    else:
        form = BankingIntegrationSettingsForm(instance=settings_instance)

    pending_count = BankTransaction.objects.filter(
        user=request.user,
        status=BankTransaction.STATUS_PENDING,
    ).count()

    return render(
        request,
        'accounts/banking/settings.html',
        {
            'form': form,
            'banking_settings': settings_instance,
            'pending_count': pending_count,
        },
    )


@login_required
def banking_disconnect(request):
    BankingIntegrationSettings.objects.filter(user=request.user).update(enabled=False)
    BankConnection.objects.filter(user=request.user).delete()
    messages.success(request, 'Banking disconnected.')
    return redirect('accounts:account_settings')


@login_required
def banking_transactions(request):
    settings_instance = BankingIntegrationSettings.objects.filter(user=request.user).first()

    if request.method == 'POST':
        transaction_id = request.POST.get('transaction_id')
        action = request.POST.get('action')
        transaction_obj = get_object_or_404(
            BankTransaction,
            pk=transaction_id,
            user=request.user,
        )

        if transaction_obj.status != BankTransaction.STATUS_PENDING:
            messages.info(request, 'That transaction has already been reviewed.')
            return redirect('accounts:banking_transactions')

        if action == 'approve':
            try:
                expense = transaction_obj.linked_expense
                if expense is None:
                    expense = _create_expense_from_bank_transaction(transaction_obj, request.user)
                transaction_obj.linked_expense = expense
                transaction_obj.status = BankTransaction.STATUS_APPROVED
                transaction_obj.reviewed_at = timezone.now()
                transaction_obj.save(update_fields=['linked_expense', 'status', 'reviewed_at', 'updated_at'])
                messages.success(request, 'Transaction approved and expense created.')
            except Exception as exc:
                logger.exception("Failed to create expense from bank transaction %s", transaction_obj.id)
                messages.error(request, f'Unable to create expense: {exc}')
        elif action == 'ignore':
            transaction_obj.status = BankTransaction.STATUS_IGNORED
            transaction_obj.reviewed_at = timezone.now()
            transaction_obj.save(update_fields=['status', 'reviewed_at', 'updated_at'])
            messages.info(request, 'Transaction ignored.')
        else:
            messages.error(request, 'Unknown action requested.')
        return redirect('accounts:banking_transactions')

    transactions = BankTransaction.objects.filter(
        user=request.user,
        status=BankTransaction.STATUS_PENDING,
    ).order_by('-posted_at', '-created_at')

    return render(
        request,
        'accounts/banking/transactions.html',
        {
            'banking_settings': settings_instance,
            'transactions': transactions,
        },
    )


@login_required
@require_POST
def quickbooks_sync_action(request):
    """Handle QuickBooks integration actions triggered from the dashboard."""

    action = request.POST.get('action')
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('accounts:home')
    settings_instance = QuickBooksSettings.objects.filter(user=request.user).first()

    if not settings_instance or not settings_instance.is_configured:
        messages.error(request, 'Please configure your QuickBooks credentials before syncing.')
        return redirect(next_url)

    try:
        if settings_instance.integration_type == QuickBooksSettings.INTEGRATION_DESKTOP:
            service = QuickBooksDesktopService(settings_instance)
            if action == 'refresh':
                messages.info(request, 'QuickBooks Desktop does not use OAuth tokens. No refresh is necessary.')
            elif action == 'import':
                uploaded = request.FILES.get('import_file')
                if not uploaded:
                    messages.error(request, 'Please upload the IIF file exported from QuickBooks Desktop.')
                else:
                    file_contents = uploaded.read().decode('utf-8-sig')
                    created, updated = service.import_invoices(request.user, file_contents, uploaded.name)
                    messages.success(
                        request,
                        f'Imported {created} new and {updated} updated invoices from QuickBooks Desktop.',
                    )
            elif action in {'export', 'sync'}:
                if action == 'sync':
                    summary = service.sync_invoices(request.user)
                    request.session['quickbooks_sync_summary'] = {
                        'exported_new': summary['exported_new'],
                        'exported_updated': summary['exported_updated'],
                    }
                    filename = summary['export_filename']
                    payload = summary['export_payload']
                else:
                    created, updated, filename, payload = service.export_invoices(request.user)
                    request.session['quickbooks_sync_summary'] = {
                        'exported_new': created,
                        'exported_updated': updated,
                    }
                response = HttpResponse(payload, content_type='text/plain')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            else:
                messages.error(request, 'Unknown QuickBooks action requested.')
        else:
            service = QuickBooksService(settings_instance)
            if action == 'refresh':
                service.ensure_access_token()
                messages.success(request, 'QuickBooks access token refreshed successfully.')
            elif action == 'import':
                created, updated = service.import_invoices(request.user)
                messages.success(
                    request,
                    f'Imported {created} new and {updated} updated invoices from QuickBooks.',
                )
            elif action == 'export':
                created, updated = service.export_invoices(request.user)
                messages.success(
                    request,
                    f'Exported {created} new and {updated} existing invoices to QuickBooks.',
                )
            elif action == 'sync':
                summary = service.sync_invoices(request.user)
                messages.success(
                    request,
                    'QuickBooks synchronisation completed successfully.',
                )
                request.session['quickbooks_sync_summary'] = summary
            else:
                messages.error(request, 'Unknown QuickBooks action requested.')
    except QuickBooksIntegrationError as exc:
        messages.error(request, str(exc))
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception('Unexpected QuickBooks integration failure')
        messages.error(request, 'An unexpected error occurred while communicating with QuickBooks.')

    return redirect(next_url)


class DeactivateUserMiddleware(MiddlewareMixin):
    def process_request(self, request):
        if request.user.is_authenticated:
            profile = request.user.profile
            if not request.user.is_active or (profile.deactivation_date and now() >= profile.deactivation_date):
                request.user.is_active = False
                request.user.save()
                logout(request)
                return redirect('accounts:subscription_cancelled')

@login_required
def change_card(request):
    user = request.user
    profile = user.profile

    card_info = None
    last4 = None
    brand = None

    # We'll create these variables to hold the current placeholder data
    name_on_card = ""
    card_expiry = ""
    address_line = ""
    city = ""
    postal_code = ""

    try:
        customer = stripe.Customer.retrieve(profile.stripe_customer_id)

        # Retrieve the default PaymentMethod info
        payment_methods = stripe.PaymentMethod.list(
            customer=profile.stripe_customer_id,
            type="card"
        )

        if payment_methods.data:
            card_info = payment_methods.data[0]  # Assuming the user has only one card on file
            last4 = card_info['card']['last4']
            brand = card_info['card']['brand']

            # Expiration date
            exp_month = card_info['card']['exp_month']
            exp_year = card_info['card']['exp_year']
            # Format your expiration date as you like: "MM/YY"
            card_expiry = f"{exp_month}/{str(exp_year)[-2:]}"

            # Billing details
            billing_details = card_info['billing_details']
            name_on_card = billing_details.get('name', "")

            # Address details (if any)
            address_data = billing_details.get('address', {})
            address_line = address_data.get('line1', "")
            city = address_data.get('city', "")
            postal_code = address_data.get('postal_code', "")

    except stripe.error.StripeError as e:
        messages.error(request, f"An error occurred: {e}")

    if request.method == 'POST':
        # If you're not using a Django Form class, you can manually grab POST data:
        token = request.POST.get('stripeToken', None)
        name_on_card = request.POST.get('name_on_card', "")
        address_line = request.POST.get('address', "")
        city = request.POST.get('city', "")
        postal_code = request.POST.get('postal_code', "")
        expiry = request.POST.get('expiry', "")

        # Do any required parsing for expiry or other fields here
        # ...

        if token:
            try:
                # Update the customer's default payment method in Stripe
                stripe.Customer.modify(
                    profile.stripe_customer_id,
                    invoice_settings={'default_payment_method': token},
                )
                # Potentially update billing details on that payment method
                # but this requires an extra call to stripe.PaymentMethod.modify():
                # stripe.PaymentMethod.modify(
                #     token,
                #     billing_details={
                #         'name': name_on_card,
                #         'address': {
                #             'line1': address_line,
                #             'city': city,
                #             'postal_code': postal_code,
                #             'country': 'CA'
                #         }
                #     }
                # )

                messages.success(request, 'Your card has been updated successfully.')
                return redirect('accounts:account_settings')
            except stripe.error.StripeError as e:
                messages.error(request, f"An error occurred: {e}")
                return redirect('accounts:change_card')

    context = {
        'last4': last4,
        'brand': brand,
        'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY,
        'card_expiry': card_expiry,
        'name_on_card': name_on_card,
        'address_line': address_line,
        'city': city,
        'postal_code': postal_code,
    }

    return render(request, 'registration/change_card.html', context)




@login_required
def cancel_subscription(request):
    user = request.user
    profile = user.profile

    if request.method == 'POST':
        try:
            subscription = stripe.Subscription.retrieve(profile.stripe_subscription_id)
            in_trial_period = subscription.status == 'trial period'

            if in_trial_period:
                # Cancel immediately
                stripe.Subscription.delete(subscription.id)
                profile.stripe_subscription_id = None
                profile.save()

                # Deactivate the user immediately
                user.is_active = False
                user.save()
                logout(request)
                messages.success(request, 'Your subscription has been canceled immediately, and your account has been deactivated.')

                return redirect('accounts:subscription_cancelled')  # Updated to the correct pattern
            else:
                # Set cancel at the end of the billing period
                billing_end_date = datetime.datetime.fromtimestamp(subscription.current_period_end)
                stripe.Subscription.modify(subscription.id, cancel_at_period_end=True)
                messages.success(request, f'Your subscription has been canceled. You will still have access until {billing_end_date}. After that, your account will be deactivated.')

                # Schedule account deactivation at the end of the billing period
                profile.deactivation_date = billing_end_date
                profile.save()

            return render(request, 'registration/subscription_cancelled.html', {
                'billing_end_date': billing_end_date,
            })
        except stripe.error.StripeError as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect('accounts:cancel_subscription')
    else:
        try:
            subscription = stripe.Subscription.retrieve(profile.stripe_subscription_id)
            in_trial_period = subscription.status == 'trialing'
        except stripe.error.StripeError:
            in_trial_period = False

    return render(request, 'registration/cancel_subscription.html', {
        'in_trial_period': in_trial_period,
    })


def subscription_cancelled(request):
    return render(request, 'registration/subscription_cancelled.html')

@login_required
def subscription_details(request):
    user = request.user
    profile = user.profile

    subscription = None
    product_name = None
    current_period_end = None
    invoices = []
    if profile.stripe_subscription_id:
        try:
            # Retrieve the subscription object
            subscription = stripe.Subscription.retrieve(profile.stripe_subscription_id)

            # Debug: print the subscription structure to see its content
            print(subscription)

            # Access the subscription items correctly
            items = subscription.get('items', None)
            if items:
                item_data = items.get('data', None)
                if item_data:
                    first_item = item_data[0]
                    product_id = first_item.get('price', {}).get('product', None)
                    if product_id:
                        product = stripe.Product.retrieve(product_id)
                        product_name = product.name

            current_period_end = datetime.datetime.fromtimestamp(subscription.get('current_period_end', 0))

            # Retrieve all invoices for the customer
            invoice_list = stripe.Invoice.list(customer=profile.stripe_customer_id, limit=12)
            # Extract invoice details including URL
            invoices = [{
                'number': invoice.number,
                'amount_due': invoice.amount_due / 100,  # Convert cents to dollars
                'currency': invoice.currency.upper(),
                'status': invoice.status,
                'invoice_url': invoice.hosted_invoice_url,
                'date': datetime.datetime.fromtimestamp(invoice.created)
            } for invoice in invoice_list]
        except stripe.error.StripeError as e:
            logger.error(f"Stripe error: {e}")
            messages.error(request, "An error occurred while retrieving your subscription details.")
            return redirect('accounts:account_settings')

    context = {
        'subscription': subscription,
        'product_name': product_name,
        'current_period_end': current_period_end,
        'invoices': invoices,
    }
    return render(request, 'registration/subscription_details.html', context)


@login_required
def change_subscription_plan(request):
    user = request.user
    profile = user.profile

    if request.method == 'POST':
        new_plan = request.POST.get('plan')
        plan_id = settings.STRIPE_PLANS.get(new_plan)

        if not plan_id:
            messages.error(request, "Invalid plan selected.")
            return redirect('accounts:subscription_details')

        try:
            subscription = stripe.Subscription.retrieve(profile.stripe_subscription_id)
            subscription.items = [{
                'id': subscription['items']['data'][0].id,
                'price': plan_id,
            }]
            subscription.save()

            messages.success(request, f"Your subscription has been updated to {new_plan}.")
            return redirect('accounts:subscription_details')
        except stripe.error.StripeError as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect('accounts:change_subscription_plan')

    return render(request, 'registration/change_subscription_plan.html', {
        'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY
    })


@require_POST
@csrf_exempt
def check_username(request):
    data = json.loads(request.body)
    username = data.get('username', "")
    if User.objects.filter(username=username).exists():
        return JsonResponse({'message': 'Username already taken'}, status=200)
    else:
        return JsonResponse({'message': 'Username available'}, status=200)

@activation_required
@subscription_required
@login_required
def add_records(request):
    inventory_products_qs = Product.objects.filter(
        user__in=get_product_user_ids(request.user)
    ).order_by('name')

    if request.method == 'POST':
        mech_expense_form = MechExpenseForm(request.POST, user=request.user)
        if mech_expense_form.is_valid():
            mech_expense = mech_expense_form.save(commit=False)
            mech_expense.user = request.user
            categorie = mech_expense_form.cleaned_data.get('categorie')
            mech_expense_item_formset = MechExpenseItemFormSet(
                request.POST,
                instance=mech_expense,
                categorie=categorie,
                prefix='mechexpenseitem_set',
                user=request.user,
            )
            if mech_expense_item_formset.is_valid():
                mech_expense.save()
                mech_expense_item_formset.save()

                # Process inventory recording if selected
                if mech_expense.record_in_inventory:
                    category_name = mech_expense.categorie or 'Parts'
                    for form in mech_expense_item_formset.forms:
                        if not hasattr(form, 'cleaned_data'):
                            continue
                        if form.cleaned_data.get('DELETE'):
                            continue
                        qty = form.cleaned_data.get('qty') or 0
                        try:
                            quantity_decimal = Decimal(str(qty))
                        except (ValueError, TypeError, InvalidOperation):
                            quantity_decimal = Decimal('0')
                        if quantity_decimal <= 0:
                            continue
                        quantity_int = int(quantity_decimal.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
                        if quantity_int <= 0:
                            continue

                        selected_product = form.cleaned_data.get('inventory_product')
                        create_inventory_product = form.cleaned_data.get('create_inventory_product')
                        part_no = form.cleaned_data.get('part_no')
                        description = form.cleaned_data.get('description')
                        price = form.cleaned_data.get('price')

                        price_decimal = Decimal('0.00')
                        has_price = price not in (None, '')
                        if has_price:
                            try:
                                price_decimal = Decimal(str(price)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                            except (InvalidOperation, ValueError):
                                price_decimal = Decimal('0.00')
                                has_price = False

                        product = _ensure_inventory_product_for_item(
                            user=request.user,
                            category_name=category_name,
                            existing_product=selected_product,
                            create_inventory_product=create_inventory_product,
                            part_no=part_no,
                            description=description,
                            price_decimal=price_decimal,
                            has_price=has_price,
                            supplier_name=mech_expense.vendor,
                        )

                        if product and quantity_int > 0:
                            InventoryTransaction.objects.create(
                                product=product,
                                transaction_type='IN',
                                quantity=quantity_int,
                                transaction_date=mech_expense.date,
                                remarks=f"Inventory recorded from expense receipt {mech_expense.receipt_no}",
                                user=request.user,
                            )

                # Record inline payment if submitted
                _record_inline_payment_from_request(mech_expense, request)

                messages.success(
                    request,
                    f'Receipt No. {mech_expense.receipt_no} created successfully.'
                )

                if 'save_continue' in request.POST:
                    messages.info(request, 'You can continue adding another expense.')
                    return redirect('accounts:add_records')

                edit_url = reverse('accounts:mechexpense_edit', args=[mech_expense.pk])
                messages.info(
                    request,
                    mark_safe(
                        f'Need to make a change? <a href="{edit_url}">Edit this expense</a> from the manage expenses page.'
                    )
                )
                return redirect('accounts:mechexpense_list')
            else:
                messages.error(request, 'Please correct the errors below in the items.')
        else:
            # When expense form is invalid, ensure the formset is built with posted data
            categorie = request.POST.get('categorie', 'Parts')
            mech_expense_item_formset = MechExpenseItemFormSet(
                request.POST,
                instance=MechExpense(),
                categorie=categorie,
                prefix='mechexpenseitem_set',
                user=request.user,
            )
            messages.error(request, 'Please correct the errors below.')
    else:
        mech_expense_form = MechExpenseForm(user=request.user)
        initial_categorie = 'Parts'  # Default category
        mech_expense_item_formset = MechExpenseItemFormSet(
            queryset=MechExpenseItem.objects.none(),
            categorie=initial_categorie,
            prefix='mechexpenseitem_set',
            user=request.user,
        )
        mech_expense_item_formset.extra = 1

    # Fetch distinct vendors for the datalist
    distinct_vendors = _get_vendor_options_for_user(request.user)

    inventory_products_data = {}
    vendor_product_map = defaultdict(list)
    all_product_ids = []
    for product in inventory_products_qs:
        product_id_str = str(product.id)
        supplier_name = (product.supplier.name or '').strip() if product.supplier else ''
        normalized_supplier = supplier_name.lower()
        display_name = f"{product.name} ({product.sku})" if product.sku else product.name
        inventory_products_data[product_id_str] = {
            'name': product.name,
            'sku': product.sku or '',
            'description': product.description or '',
            'cost_price': float(product.cost_price) if product.cost_price is not None else 0.0,
            'sale_price': float(product.sale_price) if product.sale_price is not None else 0.0,
            'supplier': supplier_name,
            'category_id': product.category_id,
            'category_name': product.category.name if product.category else '',
            'display_name': display_name,
        }
        all_product_ids.append(product_id_str)
        if normalized_supplier:
            vendor_product_map[normalized_supplier].append(product_id_str)
        else:
            vendor_product_map['__unassigned__'].append(product_id_str)
    vendor_product_map.setdefault('__unassigned__', [])

    context = {
        'mech_expense_form': mech_expense_form,
        'mech_expense_item_formset': mech_expense_item_formset,
        'distinct_vendors': distinct_vendors,
        'category_form': CategoryForm(user=request.user),
        'inventory_products': inventory_products_qs,
        'inventory_products_data': json.dumps(inventory_products_data),
        'vendor_product_map': json.dumps(dict(vendor_product_map)),
        'all_inventory_product_ids': json.dumps(all_product_ids),
        'province_tax_rates': json.dumps(PROVINCE_TAX_RATES),
        'quick_product_form': QuickProductCreateForm(user=request.user),
        'supplier_form': SupplierForm(user=request.user),
        'is_edit': False,
        'bank_accounts': BusinessBankAccount.objects.filter(
            user=request.user,
            is_active=True,
        ).order_by('name', 'id'),
        'recurring_worker_enabled': getattr(settings, "RECURRING_WORKER_ENABLED", False),
    }
    return render(request, 'accounts/mach/add_records.html', context)


@login_required
@require_POST
def quick_create_inventory_product(request):
    product_id = request.POST.get('product_id')
    supplier_name = request.POST.get('supplier_name')
    instance = None
    if product_id:
        instance = get_object_or_404(Product, pk=product_id, user=request.user)
    form = QuickProductCreateForm(request.POST, user=request.user, instance=instance)
    if form.is_valid():
        stock_owner = get_stock_owner(request.user)
        product = form.save(commit=False)
        if not product.user_id:
            product.user = stock_owner or request.user
        supplier = _get_or_create_supplier_for_user(request.user, supplier_name)
        if supplier and product.supplier_id != supplier.id:
            product.supplier = supplier
        item_type = (request.POST.get('item_type') or product.item_type or 'inventory').strip()
        if item_type not in dict(Product._meta.get_field('item_type').choices):
            item_type = 'inventory'
        product.item_type = item_type
        raw_stock = request.POST.get('quantity_in_stock')
        if product.item_type == 'inventory':
            if raw_stock not in (None, ''):
                try:
                    product.quantity_in_stock = Decimal(raw_stock)
                except (ArithmeticError, ValueError, TypeError):
                    product.quantity_in_stock = product.quantity_in_stock or 0
            if product.quantity_in_stock is None:
                product.quantity_in_stock = 0
        else:
            product.quantity_in_stock = 0
            product.reorder_level = 0
        if product.reorder_level is None:
            product.reorder_level = 0
        product.save()
        stock_quantity = product.quantity_in_stock if product.item_type == 'inventory' else 0
        stock_reorder = product.reorder_level if product.item_type == 'inventory' else 0
        upsert_product_stock(
            product,
            request.user,
            quantity_in_stock=stock_quantity,
            reorder_level=stock_reorder,
        )

        display_name = f"{product.name} ({product.sku})" if product.sku else product.name
        category_id = product.category_id
        category_name = product.category.name if product.category else ''

        product_data = {
            'id': str(product.id),
            'name': product.name,
            'sku': product.sku or '',
            'description': product.description or '',
            'cost_price': str(product.cost_price),
            'sale_price': str(product.sale_price),
            'item_type': product.item_type,
            'display_name': display_name,
            'category_id': category_id,
            'category_name': category_name,
            'supplier': product.supplier.name if product.supplier else '',
        }
        return JsonResponse({'success': True, 'product': product_data})

    errors = {field: [str(error) for error in field_errors] for field, field_errors in form.errors.items()}
    return JsonResponse({'success': False, 'errors': errors}, status=400)


def _calculate_grouped_doc_totals(records):
    base_subtotal = Decimal('0.00')
    shop_supply_total = Decimal('0.00')
    discount_total = Decimal('0.00')
    interest_total = Decimal('0.00')
    tax_total = Decimal('0.00')

    for record in records:
        amount = ensure_decimal(getattr(record, 'amount', None))
        tax_amount = ensure_decimal(getattr(record, 'tax_collected', None))
        job_label = (getattr(record, 'job', '') or '').strip().lower()

        if job_label.startswith('shop supply'):
            shop_supply_total += amount
        elif job_label.startswith('discount'):
            discount_total += amount
        elif job_label.startswith('interest'):
            interest_total += amount
        else:
            base_subtotal += amount

        if not job_label.startswith('interest'):
            tax_total += tax_amount

    pre_tax_subtotal = base_subtotal + shop_supply_total + discount_total
    subtotal_after_interest = pre_tax_subtotal + interest_total
    grand_total = pre_tax_subtotal + tax_total + interest_total

    def quantize_amount(value):
        return ensure_decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    return {
        'base_subtotal': quantize_amount(base_subtotal),
        'shop_supply_total': quantize_amount(shop_supply_total),
        'discount_total': quantize_amount(abs(discount_total)),
        'interest_total': quantize_amount(interest_total),
        'pre_tax_subtotal': quantize_amount(pre_tax_subtotal),
        'subtotal_after_interest': quantize_amount(subtotal_after_interest),
        'tax_total': quantize_amount(tax_total),
        'grand_total': quantize_amount(grand_total),
    }


@activation_required
@subscription_required
@login_required
def add_dailylog(request):
    documentType = request.GET.get('type', 'invoice')
    default_tmpl = 'accounts/mach/daily_workslip.html'

    # --- choose form & formset based on type ---
    if documentType == 'estimate':
        GroupedForm = GroupedEstimateForm
        FormsetClass = EstimateRecordFormSet
        record_qs = EstimateRecord.objects.none()
        detail_url_name = 'accounts:estimate_detail'
        edit_url_name = 'accounts:estimate_edit'
        list_url_name = 'accounts:estimate_list'
        record_model = EstimateRecord
        parent_field = 'grouped_estimate__user'
    else:  # 'invoice' (default)
        GroupedForm = GroupedInvoiceForm
        FormsetClass = IncomeRecord2FormSet
        record_qs = IncomeRecord2.objects.none()
        detail_url_name = 'accounts:groupedinvoice_detail'
        edit_url_name = 'accounts:groupedinvoice_edit'
        list_url_name = 'accounts:groupedinvoice_list'
        record_model = IncomeRecord2
        parent_field = 'grouped_invoice__user'

    template_name = default_tmpl

    selected_vehicle = None

    # --- handle POST ---
    if request.method == 'POST':
        grouped_form = GroupedForm(request.POST, user=request.user)
        formset = FormsetClass(
            request.POST,
            prefix='income_records',
            queryset=record_qs,
            form_kwargs={'user': request.user}
        )
        # Initialize modal forms for re-rendering context on error
        customer_form = CustomerForm(request.POST or None, user=request.user)
        vehicle_form = VehicleForm(request.POST or None) # Add user if needed by VehicleForm

        service_entries = []
        if documentType == 'invoice':
            selected_vehicle = _resolve_invoice_vehicle_from_post(request.POST, request.user)
            service_entries = parse_service_entries(
                request.POST,
                prefix='income_records',
                user=request.user,
            )

        if grouped_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    gi = grouped_form.save(commit=False)
                    gi.user = request.user

                    cust = grouped_form.cleaned_data.get('customer')
                    if cust:
                        gi.bill_to = cust.name
                        gi.bill_to_email = cust.email
                        gi.bill_to_address = cust.address
                    # Ensure vehicle details from form are saved if applicable
                    # Example: gi.vin_no = grouped_form.cleaned_data.get('vin_no')
                    # ... add other fields if they are part of GroupedForm but not auto-handled
                    gi.save()  # Save the main grouped invoice/estimate object first

                    if getattr(gi, 'tax_exempt', False) and getattr(gi, 'tax_exempt_reason', ''):
                        TaxExemptionReason.objects.get_or_create(
                            user=request.user,
                            reason=gi.tax_exempt_reason
                        )

                    formset.instance = gi  # Associate formset with the saved instance
                    formset.save()  # Save the related income/estimate records

                    # If it's an invoice, update the customer rate, create a pending invoice,
                    # and ensure inventory transactions exist for each product sold.
                    if documentType == 'invoice' and isinstance(gi, GroupedInvoice):
                        gi.ensure_inventory_transactions()

                        first_record = gi.income_records.order_by('line_order', 'id').first()
                        if first_record and gi.customer and first_record.rate is not None:
                            if gi.customer.charge_rate != first_record.rate:
                                gi.customer.charge_rate = first_record.rate
                                gi.customer.save(update_fields=['charge_rate'])

                        PendingInvoice.objects.get_or_create(grouped_invoice=gi)

                        if selected_vehicle and service_entries:
                            invoice_mileage = grouped_form.cleaned_data.get('mileage')
                            if invoice_mileage in (None, '', []):
                                invoice_mileage = getattr(selected_vehicle, 'current_mileage', None)
                            invoice_date = grouped_form.cleaned_data.get('date') or timezone.localdate()
                            source_label = (
                                f"Invoice {gi.invoice_number}"
                                if gi.invoice_number
                                else f"Invoice #{gi.pk}"
                            )
                            created_tasks = create_maintenance_tasks_from_services(
                                service_entries,
                                vehicle=selected_vehicle,
                                user=request.user,
                                base_date=invoice_date,
                                base_mileage=invoice_mileage,
                                source_label=source_label,
                                grouped_invoice=gi,
                            )
                            if created_tasks:
                                task_count = len(created_tasks)
                                messages.info(
                                    request,
                                    f"Added {task_count} maintenance reminder"
                                    f"{'s' if task_count != 1 else ''} for {selected_vehicle}.",
                                )

            except Exception as e:
                messages.error(request, f"An error occurred while saving: {e}")
                logger.error(f"Error adding daily log for user {request.user.username}, type {documentType}", exc_info=True)
                # Fall through to re-render form with errors (context is prepared below)
            else: # Success case
                list_url = reverse(list_url_name)
                edit_url = reverse(edit_url_name, kwargs={'pk': gi.pk})
                detail_url = reverse(detail_url_name, kwargs={'pk': gi.pk})
                new_form_url = f"{reverse('accounts:add_dailylog')}?type={documentType}"
                redirect_dashboard_flag = request.POST.get('redirect_dashboard')
                save_send_flag = request.POST.get('save_send')

                # --- REDIRECTION LOGIC ---
                if redirect_dashboard_flag:
                    messages.success(request, f"{documentType.capitalize()} saved successfully.")
                    return redirect(list_url)
                if save_send_flag:
                    if documentType == 'invoice':
                        try:
                            if getattr(gi, 'payment_status', None) == 'Paid':
                                send_paid_invoice_email(gi, request=request, include_workorder=False)
                            else:
                                send_grouped_invoice_email(request, gi.pk)
                            messages.success(request, "Invoice saved and email sent successfully.")
                        except Exception as e:
                            messages.error(request, f"Invoice saved, but email failed to send: {e}")
                    else:
                        messages.success(request, f"{documentType.capitalize()} saved successfully.")
                    return redirect(edit_url)
                if 'save_stay' in request.POST:
                    messages.success(request, f"{documentType.capitalize()} saved successfully.")
                    return redirect(edit_url)
                if 'save_continue' in request.POST:
                    messages.success(request, f"{documentType.capitalize()} saved successfully. Ready for a new one.")
                    return redirect(new_form_url)

                elif 'save_send_email' in request.POST:
                    if documentType == 'invoice':
                        try:
                            if getattr(gi, 'payment_status', None) == 'Paid':
                                send_paid_invoice_email(gi, request=request, include_workorder=False)
                            else:
                                send_grouped_invoice_email(request, gi.pk)
                            messages.success(request, "Invoice saved and email sent successfully.")
                        except Exception as e:
                            messages.error(request, f"Invoice saved, but email failed to send: {e}")
                        return redirect(edit_url)
                    messages.success(request, f"{documentType.capitalize()} saved successfully.")
                    messages.info(request, "Email sending is typically configured for invoices. Taking you to estimate details.")
                    return redirect(reverse(detail_url_name, kwargs={'pk': gi.pk}))

                else: # Default action (e.g., "Submit Log" / "Save" button)
                    if 'submit_income' in request.POST:
                        messages.success(request, f"{documentType.capitalize()} saved successfully.")
                        return redirect(list_url)
                    messages.success(request, f"{documentType.capitalize()} submitted successfully.")
                    return redirect(detail_url)

        else: # Forms are not valid
            # Prepare detailed error messages
            error_messages = []
            for field, errors in grouped_form.errors.items():
                label = grouped_form.fields[field].label if field != '__all__' and field in grouped_form.fields else 'Form'
                error_messages.append(f"{label}: {'; '.join(errors)}")
            for i, form_errors in enumerate(formset.errors):
                if form_errors:
                    for field, errors in form_errors.items():
                        label = formset.forms[i].fields[field].label if field != '__all__' and field in formset.forms[i].fields else 'Item'
                        error_messages.append(f"Item {i+1} - {label}: {'; '.join(errors)}")
            if formset.non_form_errors():
                error_messages.append(f"General item errors: {'; '.join(formset.non_form_errors())}")

            if error_messages:
                messages.error(request, "Please correct the errors below: " + "; ".join(error_messages))
            else:
                messages.error(request, "Please correct the validation errors.")
            # Fall through to re-render the page with forms containing errors

    # --- handle GET or re-render on POST error ---
    else: # GET request
        grouped_form = GroupedForm(user=request.user)
        formset = FormsetClass(
            prefix='income_records',
            queryset=record_qs, # Empty queryset for new form
            form_kwargs={'user': request.user}
        )
        customer_form = CustomerForm(user=request.user) # For the modal
        vehicle_form = VehicleForm() # For the modal

    # --- common context preparation (for GET or POST with errors) ---
    business_user_ids = get_customer_user_ids(request.user)
    customers = Customer.objects.filter(user__in=business_user_ids)
    user_profile = getattr(request.user, 'profile', None)
    if user_profile:
        province = getattr(user_profile, 'province', None)
        tax_name = user_profile.get_tax_name() if hasattr(user_profile, 'get_tax_name') else "Tax"
    else:
        province = None
        tax_name = "Tax"
    tax_rate = PROVINCE_TAX_RATES.get(province, 0.00)
    tax_components_json = json.dumps(get_tax_component_rates(province))

    default_reasons = [
        "Indigenous Status",
        "Government Purchase",
        "Resale",
        "Export",
    ]
    saved_reasons = list(
        TaxExemptionReason.objects.filter(user=request.user).values_list('reason', flat=True)
    )
    exemption_reasons = list(dict.fromkeys(default_reasons + saved_reasons))

    # Fetching existing data for suggestions (simplified)
    filter_kwargs = {parent_field: request.user}
    existing_jobsites = list(record_model.objects.filter(**filter_kwargs).values_list('jobsite', flat=True).distinct().exclude(jobsite__isnull=True).exclude(jobsite__exact=''))
    existing_trucks = list(record_model.objects.filter(**filter_kwargs).values_list('truck', flat=True).distinct().exclude(truck__isnull=True).exclude(truck__exact=''))
    existing_jobs = list(record_model.objects.filter(**filter_kwargs).values_list('job', flat=True).distinct().exclude(job__isnull=True).exclude(job__exact=''))
    job_catalog, service_description_strings, _job_name_choices = build_service_job_catalog(request.user)
    existing_jobs = sorted(set(existing_jobs) | set(service_description_strings))

    product_user_ids = get_product_user_ids(request.user)
    products = Product.objects.filter(user__in=product_user_ids)
    categories = Category.objects.filter(user__in=product_user_ids).order_by('name')
    product_data = {
        str(p.id): {
            "name": p.name,
            "sku": p.sku or '',
            "description": p.description or '',
            "price": str(p.sale_price),
            "item_type": p.item_type,
        } for p in products
    }
    product_data_json = json.dumps(product_data)

    has_driver = Driver.objects.filter(user=request.user).exists()
    drivers = Driver.objects.filter(user=request.user) if has_driver else None
    fleet_vehicles = FleetVehicle.objects.filter(user=request.user)

    if documentType == 'estimate':
        generated_number = GroupedEstimate.generate_estimate_number(request.user)
    else:
        generated_number = GroupedInvoice.generate_invoice_number(request.user, commit=False)

    resolved_vehicle = selected_vehicle or getattr(getattr(grouped_form, 'instance', None), 'vehicle', None)
    vehicle_history_payload = _serialize_vehicle_history(resolved_vehicle)
    vehicle_maintenance_payload = _serialize_vehicle_maintenance(resolved_vehicle)

    invoice_is_paid = False
    invoice_total_paid_amount = Decimal('0.00')
    invoice_balance_due_amount = Decimal('0.00')
    payment_history = []
    payment_history_count = 0
    payment_history_last_date = None
    if documentType == 'invoice' and getattr(grouped_form, 'instance', None) and grouped_form.instance.pk:
        total_amount = grouped_form.instance.total_amount or Decimal('0.00')
        total_paid = grouped_form.instance.total_paid()
        credit_total = ensure_decimal(grouped_form.instance.total_credit_amount)
        total_settled = total_paid
        tolerance = Decimal('0.01')
        pending_invoice = getattr(grouped_form.instance, 'pending_invoice', None)
        pending_is_paid = bool(pending_invoice and getattr(pending_invoice, 'is_paid', False))
        invoice_is_paid = pending_is_paid or bool(getattr(grouped_form.instance, 'paid_invoice', None)) or (total_amount > Decimal('0.00') and total_settled + tolerance >= total_amount)
        if invoice_is_paid:
            invoice_total_paid_amount = total_amount
            invoice_balance_due_amount = Decimal('0.00')
        else:
            invoice_total_paid_amount = total_settled
            invoice_balance_due_amount = (total_amount - total_paid - credit_total).quantize(Decimal('0.01'))
            if invoice_balance_due_amount < Decimal('0.00'):
                invoice_balance_due_amount = Decimal('0.00')
        payment_history = list(grouped_form.instance.payments.order_by('date', 'id'))
        payment_history_count = len(payment_history)
        if payment_history:
            payment_history_last_date = payment_history[-1].date

    document_records = []
    if getattr(grouped_form, 'instance', None) and grouped_form.instance.pk:
        if documentType == 'invoice':
            document_records = list(grouped_form.instance.income_records.all())
        elif documentType == 'estimate':
            document_records = list(grouped_form.instance.estimate_records.all())
    document_totals = _calculate_grouped_doc_totals(document_records)

    context = {
        'grouped_invoice_form': grouped_form,
        'formset': formset,
        'customer_form': customer_form, # Modal form
        'vehicle_form': vehicle_form,   # Modal form
        'customers': customers,         # For main dropdown
        'jobsite_list': existing_jobsites,
        'truck_list': existing_trucks,
        'job_list': existing_jobs,
        'job_catalog_json': job_catalog,
        'products': products,
        'categories': categories,
        'product_data_json': product_data_json,
        'has_driver': has_driver,
        'drivers': drivers,
        'fleet_vehicles': fleet_vehicles,
        'tax_name': tax_name,
        'tax_rate': tax_rate,
        'tax_components_json': tax_components_json,
        'exemption_reasons': exemption_reasons,
        'documentType': documentType,
        'formset_prefix': 'income_records',
        'generated_invoice_number': generated_number,
        'payment_methods': PAYMENT_METHOD_OPTIONS,
        'invoice_is_paid': invoice_is_paid,
        'invoice_total_paid_amount': invoice_total_paid_amount,
        'invoice_balance_due_amount': invoice_balance_due_amount,
        'payment_history': payment_history,
        'payment_history_count': payment_history_count,
        'payment_history_last_date': payment_history_last_date,
        'base_subtotal_amount': document_totals['base_subtotal'],
        'shop_supply_total_amount': document_totals['shop_supply_total'],
        'discount_total_amount': document_totals['discount_total'],
        'interest_total_amount': document_totals['interest_total'],
        'pre_tax_subtotal_amount': document_totals['pre_tax_subtotal'],
        'subtotal_after_interest_amount': document_totals['subtotal_after_interest'],
        'tax_total_amount': document_totals['tax_total'],
        'grand_total_amount': document_totals['grand_total'],
        'vehicle_history': vehicle_history_payload,
        'vehicle_history_json': json.dumps(vehicle_history_payload, cls=DjangoJSONEncoder),
        'vehicle_history_api': reverse('accounts:vehicle_history_summary'),
        'vehicle_maintenance': vehicle_maintenance_payload,
        'vehicle_maintenance_json': json.dumps(vehicle_maintenance_payload, cls=DjangoJSONEncoder),
        'vehicle_maintenance_api': reverse('accounts:vehicle_maintenance_summary'),
    }
    return render(request, template_name, context)





@login_required
@require_POST
def add_customer(request):
    form = CustomerForm(request.POST, user=request.user)
    if form.is_valid():
        customer = form.save()
        data = {
            'success': True,
            'customer': model_to_dict(
                customer,
                fields=['id', 'name', 'email', 'cc_emails', 'address', 'phone_number'],
            )
        }
        return JsonResponse(data)
    else:
        errors = form.errors.get_json_data()
        return JsonResponse({'success': False, 'errors': errors})

@login_required
def get_customer_details(request):
    customer_id = request.GET.get('customer_id')
    try:
        business_user_ids = get_customer_user_ids(request.user)
        customer = Customer.objects.get(id=customer_id, user__in=business_user_ids)
        data = {
            'name': customer.name,
            'email': customer.email,
            'cc_emails': customer.cc_emails or '',
            'address': customer.address,
            'portal_enabled': bool(customer.portal_user and customer.portal_user.is_active),
            'portal_username': customer.portal_user.username if customer.portal_user else '',
        }
        return JsonResponse(data)
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)

def contact_support(request):
    valid_types = {choice[0] for choice in PublicContactMessage.MESSAGE_TYPE_CHOICES}
    requested_type = request.GET.get('type') or None
    initial_type = requested_type if requested_type in valid_types else PublicContactMessage.TYPE_SUPPORT

    supplier_portal_request = False
    supplier_account = None
    if request.user.is_authenticated and request.GET.get('source') == 'supplier_portal':
        try:
            supplier_account = request.user.supplier_portal
            supplier_portal_request = supplier_account is not None
        except Supplier.DoesNotExist:
            supplier_account = None

    default_name = request.user.get_full_name() or request.user.get_username()
    default_email = request.user.email

    initial = {
        'full_name': request.GET.get('name') or default_name,
        'email': request.GET.get('email') or default_email,
        'message': request.GET.get('message') or '',
        'message_type': initial_type,
        'reference_code': request.GET.get('reference') or '',
        'subject': request.GET.get('subject') or request.GET.get('type') or '',
    }

    service_hint = request.POST.get('service_type') or request.GET.get('service') or ''

    if request.method == 'POST':
        form = PortalContactForm(request.POST)
        if form.is_valid():
            message_obj = form.save(commit=False)
            message_obj.source = PublicContactMessage.SOURCE_PORTAL
            message_obj.submitted_by = request.user
            try:
                portal_customer = request.user.customer_portal
            except Customer.DoesNotExist:
                portal_customer = None
            if portal_customer:
                message_obj.customer = portal_customer

            service_type = request.POST.get('service_type') or service_hint
            if service_type and service_type in dict(PublicContactMessage.SERVICE_TYPES):
                message_obj.service_type = service_type

            if (
                message_obj.message_type == PublicContactMessage.TYPE_MAINTENANCE
                and not message_obj.service_type
            ):
                if 'maintenance' in dict(PublicContactMessage.SERVICE_TYPES):
                    message_obj.service_type = 'maintenance'

            message_obj.status = PublicContactMessage.STATUS_NEW
            message_obj.save()
            business_name = (getattr(settings, "DEFAULT_BUSINESS_NAME", "") or "").strip()
            team_label = f"the {business_name} team" if business_name else "the team"
            messages.success(request, f"Your message has been submitted to {team_label}.")
            redirect_url = reverse('accounts:contact_support')
            if supplier_portal_request:
                redirect_url = f"{redirect_url}?source=supplier_portal"
            return HttpResponseRedirect(redirect_url)
    else:
        form = PortalContactForm(initial=initial)

    context = {
        'form': form,
        'service_type': service_hint,
    }

    template_name = 'accounts/contact_support.html'
    if supplier_portal_request and supplier_account:
        from .supplier_views import build_supplier_portal_context

        context.update(build_supplier_portal_context(request))
        template_name = 'suppliers/contact_support.html'

    return render(request, template_name, context)


def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            associated_users = User.objects.filter(email=data)
            if associated_users.exists():
                for user in associated_users:
                    subject = "Password Reset Requested"
                    email_template_name = "accounts/password_reset_email.html"
                    c = {
                        "email": user.email,
                        'site_name': 'Smart Invoices',
                        'domain': settings.SITE_URL,
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        "user": user,
                        'token': default_token_generator.make_token(user),
                    }
                    email = render_to_string(email_template_name, c)
                    email_message = EmailMessage(subject, email, to=[user.email])
                    email_message.content_subtype = "html"  # Set the email content type to HTML
                    email_message.send()
                messages.success(request, 'A message with reset password instructions has been sent to your inbox.')
                return redirect("accounts:password_reset_done")
            messages.error(request, 'An invalid email has been entered.')
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="accounts/password_reset.html", context={"password_reset_form": password_reset_form})

def password_reset_done(request):
    return render(request=request, template_name="accounts/password_reset_done.html")

def password_reset_confirm(request, uidb64=None, token=None):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Your password has been set. You may go ahead and log in now.')
                return redirect('accounts:password_reset_complete')
        else:
            form = SetPasswordForm(user)
        return render(request, 'accounts/password_reset_confirm.html', {'form': form})
    else:
        messages.error(request, 'The reset password link is no longer valid.')
        return redirect('accounts:password_reset_complete')

def password_reset_complete(request):
    return render(request=request, template_name="accounts/password_reset_complete.html")



@require_POST
@csrf_exempt
@login_required
def update_invoice_status(request):
    data = json.loads(request.body)
    invoice_id = data.get('invoice_id')
    is_paid = data.get('is_paid')

    try:
        grouped_invoice = GroupedInvoice.objects.get(id=invoice_id, user=request.user)
        pending_invoice = PendingInvoice.objects.filter(grouped_invoice=grouped_invoice).first()

        if pending_invoice:
            pending_invoice.is_paid = is_paid
            pending_invoice.save()
        else:
            return JsonResponse({'success': False, 'message': 'Pending invoice not found'})

        return JsonResponse({'success': True})
    except GroupedInvoice.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Grouped invoice not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Common manual payment method options for dropdowns.
PAYMENT_METHOD_OPTIONS = [
    "Cash",
    "Credit Card",
    "Debit",
    "E-Transfer",
    "ACH",
    "Cheque",
    "Manual",
    "Other",
]

DATE_RANGE_OPTIONS = [
    ("today", "Today"),
    ("yesterday", "Yesterday"),
    ("week", "This week"),
    ("month", "This month"),
    ("6m", "Last 6 months"),
    ("year", "This year"),
    ("1y", "Last 12 months"),
    ("all", "All time"),
    ("custom", "Custom range"),
]
DATE_RANGE_COOKIE_NAME = "invoice_date_range"

def _get_invoice_date_range_options(user=None):
    if user:
        years = GroupedInvoice.objects.filter(
            user=user,
            date__isnull=False,
        ).dates('date', 'year', order='DESC')
        year_options = [(str(item.year), str(item.year)) for item in years]
    else:
        current_year = timezone.localdate().year
        year_options = [(str(current_year - offset), str(current_year - offset)) for offset in range(0, 3)]

    return [
        ("today", "Today"),
        ("yesterday", "Yesterday"),
        ("week", "This week"),
        ("month", "This month"),
        ("last_month", "Last month"),
        ("30d", "Last 30 days"),
        ("this_quarter", "This quarter"),
        ("last_quarter", "Last quarter"),
        ("3m", "Last 3 months"),
        ("6m", "Last 6 months"),
        ("1y", "Last 12 months"),
        ("year", "Year to date"),
        ("this_year", "This year"),
    ] + year_options + [
        ("all", "All time"),
        ("custom", "Custom range"),
    ]


def _get_date_range_bounds(request, default: str = "1y", cookie_key: str | None = None):
    """
    Parse `?date_range=` and return (key, start_date, end_date).
    - key: one of today/yesterday/week/month/last_month/30d/3m/6m/year/this_year/1y/this_quarter/last_quarter/all/custom (fallback to default)
    - start_date/end_date: date objects or None
    """
    from django.utils import timezone

    def _parse_date(value):
        if not value:
            return None
        try:
            return datetime.datetime.strptime(value, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return None

    raw_key = (request.GET.get("date_range") or "").strip()
    if not raw_key and cookie_key:
        raw_key = (request.COOKIES.get(cookie_key) or "").strip()
    key = raw_key.lower()
    if not key:
        key = (default or "").strip().lower()
    valid_keys = {
        "today",
        "yesterday",
        "week",
        "month",
        "last_month",
        "30d",
        "3m",
        "6m",
        "year",
        "this_year",
        "1y",
        "this_quarter",
        "last_quarter",
        "all",
        "custom",
    }
    year_match = re.fullmatch(r"\d{4}", key or "")
    if not year_match and key not in valid_keys:
        key = default
        year_match = re.fullmatch(r"\d{4}", key or "")

    today = timezone.localdate()
    start_date = None
    end_date = today

    if year_match:
        year_value = int(year_match.group(0))
        if 1900 <= year_value <= 2100:
            start_date = datetime(year_value, 1, 1).date()
            end_date = datetime(year_value, 12, 31).date()
            return key, start_date, end_date
        key = default
        year_match = re.fullmatch(r"\d{4}", key or "")

    if key == "today":
        start_date = today
        end_date = today
    elif key == "yesterday":
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif key == "week":
        start_date = today - timedelta(days=today.weekday())  # Monday
    elif key == "month":
        start_date = today.replace(day=1)
    elif key == "last_month":
        first_day_this_month = today.replace(day=1)
        end_date = first_day_this_month - timedelta(days=1)
        start_date = end_date.replace(day=1)
    elif key == "30d":
        start_date = today - timedelta(days=30)
    elif key == "this_quarter":
        quarter = ((today.month - 1) // 3) + 1
        quarter_start_month = 3 * (quarter - 1) + 1
        start_date = today.replace(month=quarter_start_month, day=1)
    elif key == "last_quarter":
        quarter = ((today.month - 1) // 3) + 1
        quarter_start_month = 3 * (quarter - 1) + 1
        current_quarter_start = today.replace(month=quarter_start_month, day=1)
        start_date = current_quarter_start - relativedelta(months=3)
        end_date = current_quarter_start - timedelta(days=1)
    elif key == "3m":
        start_date = today - relativedelta(months=3)
    elif key == "6m":
        start_date = today - relativedelta(months=6)
    elif key == "year":
        start_date = today.replace(month=1, day=1)
    elif key == "this_year":
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    elif key == "1y":
        start_date = today - relativedelta(years=1)
    elif key == "all":
        start_date = None
        end_date = None
    elif key == "custom":
        start_date = _parse_date(request.GET.get("start_date"))
        end_date = _parse_date(request.GET.get("end_date"))
        if start_date and not end_date:
            end_date = start_date
        if end_date and not start_date:
            start_date = end_date

    return key, start_date, end_date


def _calc_percent(part: Decimal, total: Decimal) -> float:
    if not total or total <= Decimal("0.00"):
        return 0
    percent = (part / total) * Decimal("100")
    if percent < 0:
        percent = Decimal("0")
    if percent > 100:
        percent = Decimal("100")
    return float(percent.quantize(Decimal("0.1")))


def _resolve_sorting(request, *, default_sort=None, default_order='asc'):
    sort_by = (request.GET.get('sort_by') or '').strip().lower()
    order = (request.GET.get('order') or '').strip().lower()
    if not sort_by and default_sort:
        sort_by = default_sort
    if order not in {'asc', 'desc'}:
        order = default_order if default_order in {'asc', 'desc'} else 'asc'
    return sort_by, order


def _annotate_invoice_number_sort(queryset, field_name, *, alias='invoice_number_sort'):
    dash_alias = f"{alias}_dash_pos"
    queryset = queryset.annotate(**{
        dash_alias: StrIndex(Reverse(F(field_name)), Value('-')),
    })
    seq_start = Length(F(field_name)) - F(dash_alias) + Value(2)
    seq_str = Substr(F(field_name), seq_start)
    seq_num = Cast(NullIf(seq_str, Value('')), IntegerField())
    return queryset.annotate(**{
        alias: Coalesce(seq_num, Value(0)),
    })


class PendingInvoiceListView(LoginRequiredMixin, ListView):
    model = GroupedInvoice
    template_name = 'app/pendinginvoice_list_content.html'  # Ensure correct path
    context_object_name = 'pending_invoices'
    paginate_by = 100

    def _get_view_mode(self):
        mode = (self.request.GET.get('view') or 'invoice').strip().lower()
        return 'customer' if mode == 'customer' else 'invoice'

    def _get_invoice_queryset(self):
        if hasattr(self, '_pending_invoice_queryset'):
            return self._pending_invoice_queryset

        query = (self.request.GET.get('search') or '').strip()
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        user = self.request.user

        invoices = GroupedInvoice.objects.filter(user=user).select_related('customer').annotate(
            total_paid=Coalesce(
                Sum('payments__amount'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )
        invoices = _annotate_invoice_credit_totals(invoices).annotate(
            balance_due=ExpressionWrapper(
                F('total_amount') - F('total_paid') - F('credit_total'),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )

        # Only include invoices with a balance due greater than zero
        invoices = invoices.filter(balance_due__gt=Decimal('0.00'))

        if start_date:
            invoices = invoices.filter(date__isnull=False, date__gte=start_date)
        if end_date:
            invoices = invoices.filter(date__isnull=False, date__lte=end_date)

        if query:
            invoices = invoices.filter(
                Q(invoice_number__icontains=query) |
                Q(bill_to__icontains=query) |
                Q(date__icontains=query) |
                Q(total_amount__icontains=query)
            )

        sort_by, order = _resolve_sorting(
            self.request,
            default_sort='invoice_number',
            default_order='desc',
        )
        sort_map = {
            'invoice_number': 'invoice_number',
            'date': 'date',
            'bill_to': 'bill_to',
            'total_amount': 'total_amount',
            'total_paid': 'total_paid',
            'balance_due': 'balance_due',
        }
        if sort_by == 'invoice_number':
            invoices = _annotate_invoice_number_sort(invoices, 'invoice_number')
            if order == 'desc':
                invoices = invoices.order_by('-invoice_number_sort', '-invoice_number')
            else:
                invoices = invoices.order_by('invoice_number_sort', 'invoice_number')
        else:
            sort_field = sort_map.get(sort_by)
            if sort_field:
                if order == 'desc':
                    invoices = invoices.order_by(f'-{sort_field}')
                else:
                    invoices = invoices.order_by(sort_field)
            else:
                invoices = invoices.order_by('-date', '-id')

        self._pending_invoice_queryset = invoices
        return self._pending_invoice_queryset

    def _get_customer_rows(self):
        if hasattr(self, '_pending_customer_rows'):
            return self._pending_customer_rows

        query = (self.request.GET.get('search') or '').strip()
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        user = self.request.user

        invoice_paid_subquery = (
            Payment.objects.filter(invoice=OuterRef('pk'))
            .values('invoice')
            .annotate(
                total=Coalesce(
                    Sum('amount'),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            )
            .values('total')[:1]
        )
        invoices = _annotate_invoice_credit_totals(
            GroupedInvoice.objects.filter(
                user=user,
                customer__isnull=False,
            ).annotate(
                total_paid=Coalesce(
                    Subquery(invoice_paid_subquery),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
            )
        ).annotate(
            balance_due=ExpressionWrapper(
                F('total_amount') - F('total_paid') - F('credit_total'),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
        ).filter(balance_due__gt=Decimal('0.00'))

        if start_date:
            invoices = invoices.filter(date__isnull=False, date__gte=start_date)
        if end_date:
            invoices = invoices.filter(date__isnull=False, date__lte=end_date)

        if query:
            invoices = invoices.filter(
                Q(customer__name__icontains=query) |
                Q(customer__email__icontains=query)
            )

        rows = invoices.values('customer_id', 'customer__name').annotate(
            total_amount=Coalesce(Sum('total_amount'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
            total_paid=Coalesce(Sum('total_paid'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
            balance_due=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
            last_statement=Max('date'),
        )

        sort_by, order = _resolve_sorting(self.request, default_sort='balance_due', default_order='desc')
        sort_map = {
            'customer_name': 'customer__name',
            'last_statement': 'last_statement',
            'total_amount': 'total_amount',
            'total_paid': 'total_paid',
            'balance_due': 'balance_due',
        }
        sort_field = sort_map.get(sort_by)
        if sort_field:
            if order == 'desc':
                rows = rows.order_by(f'-{sort_field}', 'customer__name')
            else:
                rows = rows.order_by(sort_field, 'customer__name')
        else:
            rows = rows.order_by('-balance_due', 'customer__name')

        total_pending = invoices.aggregate(
            total=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2))
        )['total'] or Decimal('0.00')
        self._pending_customer_summary = {
            'total_pending': total_pending,
            'customer_count': rows.count(),
        }

        self._pending_customer_rows = rows
        return self._pending_customer_rows

    def get_queryset(self):
        view_mode = self._get_view_mode()
        if view_mode == 'customer':
            return self._get_customer_rows()
        return self._get_invoice_queryset()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        view_mode = self._get_view_mode()
        context['view_mode'] = view_mode

        if view_mode == 'customer':
            customer_rows = context.get('pending_invoices') or context.get('object_list')
            summary = getattr(self, '_pending_customer_summary', None) or {}
            total_pending = summary.get('total_pending', Decimal('0.00'))
            pending_count = summary.get('customer_count', len(customer_rows or []))
            context['pending_customers'] = customer_rows
            context['pending_invoices'] = []
            context['pending_count_label'] = 'Customers'
        else:
            queryset = self._get_invoice_queryset()
            total_pending = queryset.aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
            pending_count = queryset.count()
            context['pending_invoices'] = context.get('pending_invoices') or context.get('object_list')
            context['pending_count_label'] = 'Invoices'

        context['total_pending'] = total_pending
        context['pending_count'] = pending_count
        context['pending_invoice_count'] = pending_count if view_mode == 'invoice' else 0
        context['payment_methods'] = PAYMENT_METHOD_OPTIONS
        context['date_range_options'] = _get_invoice_date_range_options(self.request.user)
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        context['current_date_range'] = date_range_key
        context['selected_start_date'] = start_date
        context['selected_end_date'] = end_date
        params = self.request.GET.copy()
        params.pop('page', None)
        sort_params = params.copy()
        sort_params.pop('sort_by', None)
        sort_params.pop('order', None)
        context['query_string'] = params.urlencode()
        context['sort_query_string'] = sort_params.urlencode()
        context['search_query'] = self.request.GET.get('search', '')
        sort_default = 'invoice_number' if view_mode == 'invoice' else 'balance_due'
        sort_order = 'desc' if view_mode == 'invoice' else 'desc'
        sort_by, order = _resolve_sorting(
            self.request,
            default_sort=sort_default,
            default_order=sort_order,
        )
        context['sort_by'] = sort_by
        context['order'] = order
        context['today_date'] = timezone.localdate()
        context['payment_methods'] = PAYMENT_METHOD_OPTIONS

        def format_dt(value):
            if not value:
                return ""
            if timezone.is_aware(value):
                value = timezone.localtime(value)
            return value.strftime("%b %d, %Y %I:%M %p").lstrip("0")

        def payment_history_title(invoice):
            payments = list(getattr(invoice, "payments", []).all())
            if not payments:
                return "No payments recorded."
            lines = ["Payment history:"]
            for payment in payments:
                pay_date = payment.date.strftime("%b %d, %Y") if payment.date else "Unknown date"
                amount = f"{payment.amount:,.2f}"
                method = f" ({payment.method})" if payment.method else ""
                lines.append(f"{pay_date}: ${amount}{method}")
            return "\n".join(lines)

        status_class_map = {
            "Paid": "success",
            "Partial": "warning",
            "Pending": "primary",
        }

        for invoice in context.get("object_list", []):
            status_label = getattr(invoice, "status_label", "Pending")
            tone = status_class_map.get(status_label, "primary")
            invoice.email_button_class = f"btn-outline-{tone}"

            sent_at = getattr(invoice, "latest_email_sent_at", None)
            opened_at = getattr(invoice, "latest_email_opened_at", None)
            viewed_at = getattr(invoice, "latest_portal_viewed_at", None)

            last_seen_at = None
            if opened_at and viewed_at:
                last_seen_at = opened_at if opened_at >= viewed_at else viewed_at
            else:
                last_seen_at = opened_at or viewed_at

            invoice.email_status_sent_at = sent_at
            invoice.email_status_opened_at = opened_at
            invoice.email_status_viewed_at = viewed_at
            invoice.email_status_seen_at = last_seen_at

            if status_label == "Paid":
                invoice.email_button_label = "Email receipt"
                invoice.email_button_title = payment_history_title(invoice)
                continue

            if sent_at:
                if last_seen_at:
                    invoice.email_button_label = "Email sent and seen"
                    lines = [f"Sent: {format_dt(sent_at)}"]
                    if opened_at:
                        lines.append(f"Email opened: {format_dt(opened_at)}")
                    if viewed_at:
                        lines.append(f"Portal viewed: {format_dt(viewed_at)}")
                    invoice.email_button_title = "\n".join(lines)
                else:
                    invoice.email_button_label = "Email sent"
                    invoice.email_button_title = f"Sent: {format_dt(sent_at)}"
            else:
                invoice.email_button_label = "Email Invoice"
                invoice.email_button_title = "Not sent yet."

        return context

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        # Check for AJAX request by looking at the X-Requested-With header
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string('app/pending_invoices_table.html', context, request=request)
            total_pending = context.get('total_pending') or Decimal('0.00')
            pending_count = context.get('pending_count') or 0
            start_date = context.get('selected_start_date')
            end_date = context.get('selected_end_date')
            return JsonResponse({
                'html': html,
                'total_pending': f"${total_pending:,.2f}",
                'pending_count': pending_count,
                'pending_count_label': context.get('pending_count_label', 'Invoices'),
                'view_mode': context.get('view_mode', 'invoice'),
                'start_date': start_date.isoformat() if start_date else '',
                'end_date': end_date.isoformat() if end_date else '',
                'current_date_range': context.get('current_date_range', ''),
            })

        return self.render_to_response(context)

class MarkInvoicePaidView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        next_url = request.POST.get('next', 'accounts:pending_invoice_list')

        def respond_with_message(level: str, message: str, status: int = 400):
            getattr(messages, level)(request, message)
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                payload_status = 'error' if level == 'error' else level
                return JsonResponse({'status': payload_status, 'message': message}, status=status)
            return redirect(next_url)

        pending_invoice = None
        grouped_invoice = None
        grouped_invoice_id = (request.POST.get('grouped_invoice_id') or '').strip()
        if grouped_invoice_id:
            grouped_invoice = (
                GroupedInvoice.objects.filter(pk=grouped_invoice_id, user=request.user).first()
            )
            if grouped_invoice:
                pending_invoice = PendingInvoice.objects.filter(grouped_invoice=grouped_invoice).first()
        else:
            pending_invoice = (
                PendingInvoice.objects.select_related('grouped_invoice')
                .filter(pk=pk, grouped_invoice__user=request.user)
                .first()
            )
            if pending_invoice:
                grouped_invoice = pending_invoice.grouped_invoice
            else:
                grouped_invoice = GroupedInvoice.objects.filter(pk=pk, user=request.user).first()
                if grouped_invoice:
                    pending_invoice = PendingInvoice.objects.filter(grouped_invoice=grouped_invoice).first()

        if not grouped_invoice:
            return respond_with_message(
                'error',
                "We couldn't find that pending invoice. It may already be marked as paid.",
                status=404,
            )

        balance_due = grouped_invoice.balance_due()

        # Extract manual payment details
        amount_raw = request.POST.get('amount') or request.POST.get('payment_amount')
        method = request.POST.get('method') or request.POST.get('payment_method') or 'Manual'
        notes = request.POST.get('notes') or request.POST.get('payment_notes') or 'Marked as paid via MarkInvoicePaidView'
        payment_date_raw = request.POST.get('payment_date')

        if balance_due <= Decimal('0.00'):
            return respond_with_message(
                'info',
                f"Invoice {grouped_invoice.invoice_number} is already fully paid.",
            )

        try:
            payment_amount = Decimal(amount_raw) if amount_raw else balance_due
        except (InvalidOperation, TypeError):
            return respond_with_message(
                'error',
                "Please enter a valid payment amount.",
            )

        if payment_amount <= Decimal('0.00'):
            return respond_with_message(
                'error',
                "Payment amount must be greater than zero.",
            )

        if payment_amount > balance_due:
            return respond_with_message(
                'error',
                f"Payment cannot exceed the balance due (${balance_due}).",
            )

        payment_date = None
        if payment_date_raw:
            try:
                payment_date = datetime.datetime.strptime(payment_date_raw, "%Y-%m-%d").date()
            except ValueError:
                return respond_with_message(
                    'error',
                    "Please provide a valid payment date (YYYY-MM-DD).",
                )

        Payment.objects.create(
            invoice=grouped_invoice,
            amount=payment_amount,
            method=method,
            notes=notes,
            date=payment_date,
        )

        # Update PendingInvoice state only when payments cover the invoice.
        total_amount = grouped_invoice.total_amount or Decimal('0.00')
        total_paid = grouped_invoice.total_paid()
        if total_paid + Decimal('0.01') >= total_amount and pending_invoice:
            pending_invoice.is_paid = True
            pending_invoice.save()

        messages.success(
            request,
            f"Recorded payment of ${payment_amount:,.2f} for Invoice {grouped_invoice.invoice_number}.",
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {
                    'status': 'success',
                    'message': 'Payment recorded successfully.',
                    'balance_due': f"${grouped_invoice.balance_due():,.2f}",
                }
            )

        return redirect(next_url)

@login_required
@require_POST
def record_customer_payment(request, customer_id: int):
    """
    Record a customer-level payment and allocate it across unpaid invoices
    (oldest -> newest). If the payment doesn't fully cover the outstanding
    balance, the newest impacted invoice remains partially paid.
    """
    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)

    def json_error(message: str, status: int = 400):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': message}, status=status)
        messages.error(request, message)
        return redirect(request.META.get('HTTP_REFERER') or reverse('accounts:customer_list'))

    amount_raw = request.POST.get('amount') or request.POST.get('payment_amount')
    method = request.POST.get('method') or request.POST.get('payment_method') or 'Manual'
    notes = request.POST.get('notes') or request.POST.get('payment_notes') or 'Customer payment'
    payment_date_raw = request.POST.get('payment_date')

    payment_date = None
    if payment_date_raw:
        try:
            payment_date = datetime.datetime.strptime(payment_date_raw, "%Y-%m-%d").date()
        except ValueError:
            return json_error("Please provide a valid payment date (YYYY-MM-DD).")

    # Use subquery/annotations to avoid N+1 when determining balances.
    invoice_paid_subquery = (
        Payment.objects.filter(invoice=OuterRef('pk'))
        .values('invoice')
        .annotate(
            total=Coalesce(
                Sum('amount'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            )
        )
        .values('total')[:1]
    )
    unpaid_qs = (
        _annotate_invoice_credit_totals(
            GroupedInvoice.objects.filter(
                user=request.user,
                customer=customer,
            ).annotate(
                total_paid_calc=Coalesce(
                    Subquery(invoice_paid_subquery),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
            )
        )
        .annotate(
            balance_due_calc=ExpressionWrapper(
                F('total_amount') - F('total_paid_calc') - F('credit_total'),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
        )
        .filter(balance_due_calc__gt=Decimal('0.00'))
        .order_by('date', 'id')
    )
    unpaid_invoices = list(unpaid_qs)
    if not unpaid_invoices:
        return json_error(f"{customer.name} has no outstanding balance to pay.")

    outstanding_before = sum((inv.balance_due_calc for inv in unpaid_invoices), Decimal('0.00'))

    try:
        payment_amount = Decimal(amount_raw) if amount_raw else outstanding_before
    except (InvalidOperation, TypeError):
        return json_error("Please enter a valid payment amount.")

    if payment_amount <= Decimal('0.00'):
        return json_error("Payment amount must be greater than zero.")

    if payment_amount > outstanding_before + Decimal('0.01'):
        return json_error(
            f"Payment cannot exceed the outstanding balance (${outstanding_before:,.2f})."
        )

    # Capture overdue amounts before/after so the dashboard can update immediately.
    term_days = 30
    try:
        term_days = TERM_CHOICES.get(getattr(request.user.profile, 'term', None), 30)
    except Exception:
        term_days = 30
    today = timezone.now().date()
    overdue_threshold_date = today - timedelta(days=int(term_days or 0))
    overdue_filter = (
        {'date__lt': overdue_threshold_date}
        if int(term_days or 0) > 0
        else {'date__lte': overdue_threshold_date}
    )

    def customer_overdue_total() -> Decimal:
        rows = list(
            _annotate_invoice_credit_totals(
                GroupedInvoice.objects.filter(
                    user=request.user,
                    customer=customer,
                    date__isnull=False,
                    **overdue_filter,
                ).annotate(
                    total_paid_calc=Coalesce(
                        Subquery(invoice_paid_subquery),
                        Value(Decimal('0.00')),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    ),
                )
            )
            .annotate(
                balance_due_calc=ExpressionWrapper(
                    F('total_amount') - F('total_paid_calc') - F('credit_total'),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
            )
            .filter(balance_due_calc__gt=Decimal('0.00'))
            .values('id')
            .annotate(total=Coalesce(Sum('balance_due_calc'), Value(Decimal('0.00')), output_field=DecimalField()))
        )
        return (rows[0]['total'] if rows else Decimal('0.00')) or Decimal('0.00')

    overdue_before = customer_overdue_total()

    allocations = []
    remaining = payment_amount
    applied_total = Decimal('0.00')

    with transaction.atomic():
        for inv in unpaid_invoices:
            if remaining <= Decimal('0.00'):
                break

            inv_balance = (inv.balance_due_calc or Decimal('0.00')).quantize(Decimal('0.01'))
            if inv_balance <= Decimal('0.00'):
                continue

            apply_amount = min(inv_balance, remaining).quantize(Decimal('0.01'))
            if apply_amount <= Decimal('0.00'):
                continue

            Payment.objects.create(
                invoice_id=inv.pk,
                amount=apply_amount,
                method=method,
                notes=notes,
                date=payment_date,
            )
            remaining -= apply_amount
            applied_total += apply_amount

            # Refresh invoice payment state for accurate status.
            fresh_invoice = GroupedInvoice.objects.get(pk=inv.pk, user=request.user)
            balance_after = fresh_invoice.balance_due()
            total_amount = fresh_invoice.total_amount or Decimal('0.00')
            total_paid = fresh_invoice.total_paid()
            status = 'paid' if total_paid + Decimal('0.01') >= total_amount else 'partial'

            if status == 'paid':
                pending = PendingInvoice.objects.filter(grouped_invoice_id=inv.pk).first()
                if pending:
                    pending.is_paid = True
                    pending.save()

            allocations.append(
                {
                    'invoice_id': inv.pk,
                    'invoice_number': inv.invoice_number,
                    'applied': float(apply_amount),
                    'status': status,
                    'balance_remaining': float(max(balance_after, Decimal('0.00'))),
                }
            )

    outstanding_after = max(outstanding_before - applied_total, Decimal('0.00'))
    overdue_after = customer_overdue_total()

    # Statement links (no date range required; defaults to full range).
    download_base = reverse('accounts:download_invoice_pdf', args=[customer.id])
    print_base = reverse('accounts:print_invoice_pdf', args=[customer.id])
    email_url = reverse('accounts:send_invoice_statement', args=[customer.id])

    statement_links = {
        'pending': {
            'download': f"{download_base}?invoice_type=pending",
            'print': f"{print_base}?invoice_type=pending",
        },
        'paid': {
            'download': f"{download_base}?invoice_type=paid",
            'print': f"{print_base}?invoice_type=paid",
        },
        'all': {
            'download': f"{download_base}?invoice_type=all",
            'print': f"{print_base}?invoice_type=all",
        },
        'overdue': {
            'download': f"{download_base}?invoice_type=overdue",
            'print': f"{print_base}?invoice_type=overdue",
        },
    }

    msg = (
        f"Recorded ${applied_total:,.2f} payment for {customer.name}. "
        f"Outstanding ${outstanding_after:,.2f} remaining."
    )
    messages.success(request, msg)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse(
            {
                'status': 'success',
                'message': msg,
                'customer_id': customer.id,
                'customer_name': customer.name,
                'amount_received': float(applied_total),
                'outstanding_before': float(outstanding_before),
                'outstanding_after': float(outstanding_after),
                'overdue_before': float(overdue_before),
                'overdue_after': float(overdue_after),
                'allocations': allocations,
                'statement_links': statement_links,
                'statement_email_url': email_url,
            }
        )

    return redirect(request.META.get('HTTP_REFERER') or reverse('accounts:customer_list'))


@login_required
@require_POST
def update_overdue_followup(request, customer_id: int):
    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
    update_fields = []

    next_followup_raw = request.POST.get('next_followup') if 'next_followup' in request.POST else None
    notes_raw = request.POST.get('notes') if 'notes' in request.POST else None

    if next_followup_raw is not None:
        from django.utils.dateparse import parse_date

        next_followup_raw = next_followup_raw.strip()
        if not next_followup_raw:
            customer.next_followup = None
        else:
            parsed = parse_date(next_followup_raw)
            if not parsed:
                return JsonResponse(
                    {'status': 'error', 'message': 'Please provide a valid follow-up date (YYYY-MM-DD).'},
                    status=400,
                )
            customer.next_followup = parsed
        update_fields.append('next_followup')

    if notes_raw is not None:
        customer.collection_notes = notes_raw
        update_fields.append('collection_notes')

    if update_fields:
        customer.save(update_fields=update_fields)

    return JsonResponse(
        {
            'status': 'success',
            'next_followup': customer.next_followup.isoformat() if customer.next_followup else '',
            'collection_notes': customer.collection_notes or '',
        }
    )


class MarkInvoiceUnpaidView(LoginRequiredMixin, View):
    """Reverse a paid invoice back to unpaid by deleting recorded payments."""

    def post(self, request, pk, *args, **kwargs):
        invoice = get_object_or_404(GroupedInvoice, pk=pk, user=request.user)

        next_url = (
            request.POST.get('next')
            or request.META.get('HTTP_REFERER')
            or reverse('accounts:groupedinvoice_list')
        )

        payments_to_remove = list(invoice.payments.all().order_by('-date', '-pk'))

        if not payments_to_remove:
            message = "No recorded payments were found for this invoice."
            messages.error(request, message)
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': message}, status=400)
            return redirect(next_url)

        removed_total = sum((p.amount for p in payments_to_remove), Decimal('0.00'))

        with transaction.atomic():
            for payment in payments_to_remove:
                payment.delete()

            # Remove PaidInvoice record if it exists and recreate a PendingInvoice.
            PaidInvoice.objects.filter(grouped_invoice=invoice).delete()
            pending_invoice, _ = PendingInvoice.objects.get_or_create(grouped_invoice=invoice)
            if pending_invoice.is_paid:
                pending_invoice.is_paid = False
                pending_invoice.save(update_fields=['is_paid'])

            # Re-associate line items with the pending invoice state.
            invoice.income_records.update(pending_invoice=pending_invoice, paid_invoice=None)

            # Ensure "fully paid" metadata is cleared.
            invoice.update_date_fully_paid()
            invoice.refresh_from_db(fields=['date_fully_paid'])

        invoice.refresh_from_db()
        message = (
            f"Invoice {invoice.invoice_number} marked as unpaid. "
            f"Removed ${removed_total:,.2f} and cleared the payment history."
        )
        messages.success(request, message)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {
                    'status': 'success',
                    'message': message,
                    'balance_due': f"${invoice.balance_due():,.2f}",
                }
            )

        return redirect(next_url)


@login_required
def invoice_payment_history_manage(request, pk):
    """Return payment history HTML with edit/delete controls (AJAX)."""
    invoice = get_object_or_404(GroupedInvoice, pk=pk, user=request.user)
    payments = invoice.payments.all().order_by('-date', '-pk')
    html = render_to_string(
        'app/payment_history_table.html',
        {'payments': payments, 'show_actions': True},
        request=request,
    )
    return JsonResponse({'status': 'success', 'html': html})


class InvoicePaymentDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        payment = get_object_or_404(Payment.objects.select_related('invoice'), pk=pk, invoice__user=request.user)
        invoice = payment.invoice
        amount = payment.amount

        payment.delete()
        invoice.update_date_fully_paid()
        invoice.refresh_from_db()

        message = f"Deleted payment of ${amount:,.2f} from Invoice {invoice.invoice_number}."
        messages.success(request, message)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {
                    'status': 'success',
                    'message': message,
                    'balance_due': f"${invoice.balance_due():,.2f}",
                }
            )
        return redirect(request.META.get('HTTP_REFERER') or reverse('accounts:groupedinvoice_list'))


class InvoicePaymentUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        payment = get_object_or_404(Payment.objects.select_related('invoice'), pk=pk, invoice__user=request.user)
        invoice = payment.invoice

        amount_raw = request.POST.get('amount')
        method = request.POST.get('method') or 'Manual'
        notes = request.POST.get('notes') or ''
        payment_date_raw = request.POST.get('payment_date')

        try:
            new_amount = Decimal(amount_raw)
        except (InvalidOperation, TypeError):
            return JsonResponse({'status': 'error', 'message': 'Please enter a valid payment amount.'}, status=400)

        if new_amount <= Decimal('0.00'):
            return JsonResponse({'status': 'error', 'message': 'Payment amount must be greater than zero.'}, status=400)

        payment_date = payment.date
        if payment_date_raw:
            try:
                payment_date = datetime.datetime.strptime(payment_date_raw, "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'Please provide a valid payment date (YYYY-MM-DD).'}, status=400)

        other_total = (
            invoice.payments.exclude(pk=payment.pk)
            .aggregate(total=Coalesce(Sum('amount'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)))
            .get('total')
            or Decimal('0.00')
        )
        # Prevent overpay by more than 1 cent.
        if other_total + new_amount > invoice.total_amount + Decimal('0.01'):
            return JsonResponse({'status': 'error', 'message': 'Updated payment would exceed the invoice total.'}, status=400)

        payment.amount = new_amount
        payment.method = method
        payment.notes = notes
        payment.date = payment_date
        payment.save()

        invoice.update_date_fully_paid()
        invoice.refresh_from_db()

        message = f"Updated payment for Invoice {invoice.invoice_number}."
        messages.success(request, message)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {
                    'status': 'success',
                    'message': message,
                    'balance_due': f"${invoice.balance_due():,.2f}",
                }
            )
        return redirect(request.META.get('HTTP_REFERER') or reverse('accounts:groupedinvoice_list'))

@method_decorator(csrf_protect, name='dispatch')
class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    redirect_authenticated_user = True

    def form_invalid(self, form):
        logger.error('Login form invalid: %s', form.errors)
        return super().form_invalid(form)

    def form_valid(self, form):
        user = form.get_user()

        if not user.is_active:
            logout(self.request)
            return redirect('accounts:choose_plan')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        business_user = _resolve_primary_business_user()
        context['business_name'] = _get_business_display_name(business_user)
        context['next'] = self.request.GET.get('next') or self.request.POST.get('next')
        context['portal_switch_options'] = []
        context['active_portal'] = 'unified'
        return context

    def get_success_url(self):
        """Send each authenticated user to the correct dashboard automatically."""
        redirect_url = self.get_redirect_url()
        if redirect_url:
            return redirect_url

        user = self.request.user
        supplier_portal = getattr(user, 'supplier_portal', None)
        if supplier_portal:
            return reverse('accounts:supplier_dashboard')

        mechanic_portal = getattr(user, 'mechanic_portal', None)
        if mechanic_portal:
            return reverse('accounts:mechanic_portal_dashboard')

        accountant_portal = getattr(user, 'accountant_portal', None)
        if accountant_portal:
            return reverse('accounts:accountant_portal_dashboard')

        customer_portal = getattr(user, 'customer_portal', None)
        if customer_portal:
            if is_parts_store_business():
                return reverse('accounts:store_product_list')
            return reverse('accounts:customer_dashboard')

        return super().get_success_url()


@method_decorator(csrf_protect, name='dispatch')
class AccountantPortalLoginView(CustomLoginView):
    template_name = 'accountant/login.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_portal'] = 'accountant'
        context['portal_switch_options'] = []
        return context

def signup_options(request):
    """Display available signup paths for the business ecosystem."""

    business_user = _resolve_primary_business_user()
    context = {
        'business_user': business_user,
        'business_name': _get_business_display_name(business_user),
    }
    return render(request, 'registration/signup_options.html', context)


def admin_signup(request):
    """Collect staff signup requests that require approval by the owner."""

    business_user = _resolve_primary_business_user()
    if business_user is None:
        return render(
            request,
            'registration/admin_signup.html',
            {
                'form': None,
                'business_name': _get_business_display_name(business_user),
                'signup_disabled': True,
            },
            status=503,
        )

    if request.method == 'POST':
        form = StaffSignupForm(request.POST, business_owner=business_user)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                'Your staff signup request has been submitted for approval. '
                'You will receive an email once the account is activated.',
            )
            return redirect('accounts:login')
    else:
        form = StaffSignupForm(business_owner=business_user)

    context = {
        'form': form,
        'business_name': _get_business_display_name(business_user),
        'signup_disabled': False,
    }
    return render(request, 'registration/admin_signup.html', context)


CUSTOMER_SIGNUP_INVITE_SALT = "customer-signup-invite"
CUSTOMER_SIGNUP_INVITE_MAX_AGE_SECONDS = 60 * 60 * 24 * 14


def _build_customer_signup_invite_token(customer):
    if not customer or not customer.email:
        return ""
    signer = signing.TimestampSigner(salt=CUSTOMER_SIGNUP_INVITE_SALT)
    payload = f"{customer.id}:{customer.email.strip().lower()}"
    return signer.sign(payload)


def _resolve_customer_signup_invite(token, *, business_user):
    if not token:
        return None
    signer = signing.TimestampSigner(salt=CUSTOMER_SIGNUP_INVITE_SALT)
    try:
        payload = signer.unsign(token, max_age=CUSTOMER_SIGNUP_INVITE_MAX_AGE_SECONDS)
    except (signing.BadSignature, signing.SignatureExpired):
        return None
    parts = payload.split(":", 1)
    if len(parts) != 2:
        return None
    customer_id, email = parts[0].strip(), parts[1].strip().lower()
    if not customer_id.isdigit():
        return None
    business_user_ids = get_customer_user_ids(business_user)
    return Customer.objects.filter(
        id=int(customer_id),
        user__in=business_user_ids,
        email__iexact=email,
    ).first()


def customer_signup(request):
    """Allow shoppers to create their own customer portal account."""

    business_user = get_default_store_owner()
    if business_user is None:
        return render(
            request,
            'store/customer_signup.html',
            {
                'form': None,
                'business_name': 'our store',
                'signup_disabled': True,
            },
            status=503,
        )

    invite_token = request.GET.get("invite") or request.POST.get("invite")
    invite_customer = _resolve_customer_signup_invite(invite_token, business_user=business_user)
    approval_required = invite_customer is None
    if not invite_customer:
        invite_token = ""

    initial = {}
    if request.method == 'GET':
        prefill_email = request.GET.get('email')
        if invite_customer and invite_customer.email:
            prefill_email = invite_customer.email
        if prefill_email:
            initial['email'] = prefill_email
        if invite_customer:
            initial.update(
                {
                    "name": invite_customer.name,
                    "phone_number": invite_customer.phone_number,
                    "address": invite_customer.address,
                    "gst_hst_number": invite_customer.gst_hst_number,
                    "collect_gst_hst": invite_customer.collect_gst_hst,
                    "cc_emails": invite_customer.cc_emails,
                }
            )

    if request.method == 'POST':
        form = CustomerSignupForm(
            request.POST,
            business_user=business_user,
            approval_required=approval_required,
            invite_customer=invite_customer,
        )
        if form.is_valid():
            user = form.save()
            if approval_required:
                messages.success(
                    request,
                    'Thanks! Your customer portal request was submitted and is pending approval.',
                )
                return redirect('accounts:customer_signup_pending')
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, 'Your customer account has been created successfully.')
            next_url = request.POST.get('next') or request.GET.get('next')
            return redirect(next_url or 'accounts:public_home')
    else:
        form = CustomerSignupForm(
            business_user=business_user,
            initial=initial,
            approval_required=approval_required,
            invite_customer=invite_customer,
        )

    business_name = ''
    profile = getattr(business_user, 'profile', None)
    if profile and getattr(profile, 'company_name', None):
        business_name = profile.company_name
    if not business_name:
        business_name = business_user.get_full_name() or business_user.username

    next_param = request.POST.get('next') if request.method == 'POST' else request.GET.get('next', '')
    context = {
        'form': form,
        'business_name': business_name,
        'next': next_param,
        'signup_disabled': False,
        'approval_required': approval_required,
        'invite_token': invite_token,
    }
    return render(request, 'store/customer_signup.html', context)


def customer_signup_pending(request):
    """Landing page shown after a public customer signup request."""
    business_user = get_default_store_owner()
    business_name = _get_business_display_name(business_user)
    return render(
        request,
        'store/customer_signup_pending.html',
        {
            'business_name': business_name,
        },
    )

# Send self-signup invite for customers
@login_required
@require_POST
def send_customer_signup_invite(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        payload = request.POST

    customer_id = payload.get('customer_id')
    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
    if not customer.email:
        return JsonResponse({'success': False, 'message': 'Customer has no email on file.'}, status=400)

    business_profile = getattr(request.user, 'profile', None)
    business_name = getattr(business_profile, 'company_name', '') or request.user.get_full_name() or request.user.username
    signup_url = request.build_absolute_uri(reverse('accounts:customer_signup'))
    invite_token = _build_customer_signup_invite_token(customer)
    query_params = {"email": customer.email}
    if invite_token:
        query_params["invite"] = invite_token
    signup_url = f"{signup_url}?{urlencode(query_params)}"

    context = {
        'customer': customer,
        'business_name': business_name,
        'signup_url': signup_url,
    }
    subject = f"{business_name} - create your customer portal access"
    context = apply_branding_defaults(context)
    html_body = render_to_string('emails/customer_signup_invite.html', context)
    text_body = strip_tags(html_body)
    from_email = (
        getattr(settings, 'DEFAULT_FROM_EMAIL', None)
        or getattr(settings, 'EMAIL_HOST_USER', None)
        or getattr(business_profile, 'company_email', None)
        or request.user.email
    )
    if not from_email:
        logger.error("Signup invite missing from_email configuration")
        return JsonResponse({'success': False, 'message': 'Email sending is not configured.'}, status=200)

    email = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=from_email,
        to=[customer.email],
    )
    email.attach_alternative(html_body, "text/html")
    try:
        email.send(fail_silently=True)
    except Exception as exc:
        logger.error("Failed sending signup invite: %s", exc, exc_info=True)
        return JsonResponse({'success': False, 'message': 'Could not send email right now.'}, status=200)

    return JsonResponse({'success': True, 'message': 'Signup email sent.'})

# views.py


@login_required
def customer_approvals(request):
    """Redirect to the dashboard approvals panel."""
    return redirect(f"{reverse('accounts:home')}#customer-approvals")


@login_required
@require_POST
def approve_customer_signup(request, customer_id):
    """Approve a pending customer portal signup."""
    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(
        Customer.objects.select_related('portal_user'),
        id=customer_id,
        user__in=business_user_ids,
    )
    if customer.portal_signup_status != Customer.PORTAL_STATUS_PENDING:
        messages.info(request, f"{customer.name} is already approved.")
        return redirect('accounts:customer_approvals')

    portal_user = customer.portal_user
    if not portal_user:
        messages.error(request, "No portal account found for this customer.")
        return redirect('accounts:customer_approvals')

    if not portal_user.is_active:
        portal_user.is_active = True
        portal_user.save(update_fields=['is_active'])

    customer.portal_signup_status = Customer.PORTAL_STATUS_APPROVED
    customer.save(update_fields=['portal_signup_status'])
    _log_admin_activity(
        request,
        action="approve",
        object_type="customer_portal",
        object_id=customer.id,
        description=f"Approved customer portal access for {customer.name}.",
    )
    messages.success(request, f"{customer.name} has been approved for portal access.")
    return redirect('accounts:customer_approvals')


@login_required
def customer_list(request):
    from decimal import Decimal
    from django.core.paginator import Paginator
    from django.db.models import Sum, Value as V, DecimalField, F, Q
    from django.db.models.functions import Coalesce

    search_query = (request.GET.get('q', '') or '').strip()
    date_range_filter, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )
    show_overdue_only = request.GET.get('overdue') == '1'

    today = timezone.localdate()

    business_user_ids = get_customer_user_ids(request.user)
    customers_qs = Customer.objects.filter(user__in=business_user_ids)
    if search_query:
        customers_qs = customers_qs.filter(Q(name__icontains=search_query) | Q(email__icontains=search_query))

    # Build period totals for ALL matching customers (so ranking is correct for the selected period).
    customer_rows = list(customers_qs.values('id', 'name'))
    all_customer_ids = [row['id'] for row in customer_rows]

    totals_by_customer = {}
    paid_by_customer = {}
    credit_by_customer = {}
    overdue_invoice_totals_by_customer = {}
    overdue_paid_by_customer = {}
    overdue_credit_by_customer = {}

    if all_customer_ids:
        invoices_all_qs = GroupedInvoice.objects.filter(
            user=request.user,
            customer_id__in=all_customer_ids,
            customer__isnull=False,
        )
        if start_date:
            invoices_all_qs = invoices_all_qs.filter(date__gte=start_date)
        if end_date:
            invoices_all_qs = invoices_all_qs.filter(date__lte=end_date)
        invoices_all_qs = _annotate_invoice_credit_totals(invoices_all_qs)

        invoice_totals_rows = invoices_all_qs.values('customer_id').annotate(
            total_amount=Coalesce(
                Sum('total_amount'),
                V(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        totals_by_customer = {
            row['customer_id']: (row['total_amount'] or Decimal('0.00'))
            for row in invoice_totals_rows
        }

        payment_totals_rows = Payment.objects.filter(
            invoice__user=request.user,
            invoice__customer_id__in=all_customer_ids,
        )
        if start_date:
            payment_totals_rows = payment_totals_rows.filter(invoice__date__gte=start_date)
        if end_date:
            payment_totals_rows = payment_totals_rows.filter(invoice__date__lte=end_date)
        payment_totals_rows = payment_totals_rows.values('invoice__customer_id').annotate(
            total_paid=Coalesce(
                Sum('amount'),
                V(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        paid_by_customer = {
            row['invoice__customer_id']: (row['total_paid'] or Decimal('0.00'))
            for row in payment_totals_rows
        }

        credit_totals_rows = invoices_all_qs.values('customer_id').annotate(
            total_credit=Coalesce(
                Sum('credit_total'),
                V(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        credit_by_customer = {
            row['customer_id']: (row['total_credit'] or Decimal('0.00'))
            for row in credit_totals_rows
        }

        profile = getattr(request.user, 'profile', None)
        term_days = TERM_CHOICES.get(getattr(profile, 'term', None), 30)
        overdue_threshold_date = today - timedelta(days=term_days) if term_days > 0 else today
        overdue_date_filter = {'date__lt': overdue_threshold_date} if term_days > 0 else {'date__lte': overdue_threshold_date}

        overdue_invoices_all_qs = invoices_all_qs.filter(date__isnull=False, **overdue_date_filter)
        overdue_invoice_totals_rows = overdue_invoices_all_qs.values('customer_id').annotate(
            total_amount=Coalesce(
                Sum('total_amount'),
                V(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        overdue_invoice_totals_by_customer = {
            row['customer_id']: (row['total_amount'] or Decimal('0.00'))
            for row in overdue_invoice_totals_rows
        }

        overdue_credit_totals_rows = overdue_invoices_all_qs.values('customer_id').annotate(
            total_credit=Coalesce(
                Sum('credit_total'),
                V(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        overdue_credit_by_customer = {
            row['customer_id']: (row['total_credit'] or Decimal('0.00'))
            for row in overdue_credit_totals_rows
        }

        overdue_payment_totals_rows = Payment.objects.filter(
            invoice__user=request.user,
            invoice__customer_id__in=all_customer_ids,
            invoice__date__isnull=False,
            **{f"invoice__{k}": v for k, v in overdue_date_filter.items()},
        )
        if start_date:
            overdue_payment_totals_rows = overdue_payment_totals_rows.filter(invoice__date__gte=start_date)
        if end_date:
            overdue_payment_totals_rows = overdue_payment_totals_rows.filter(invoice__date__lte=end_date)
        overdue_payment_totals_rows = overdue_payment_totals_rows.values('invoice__customer_id').annotate(
            total_paid=Coalesce(
                Sum('amount'),
                V(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        overdue_paid_by_customer = {
            row['invoice__customer_id']: (row['total_paid'] or Decimal('0.00'))
            for row in overdue_payment_totals_rows
        }

    def _overdue_total_for_customer(cid: int) -> Decimal:
        overdue_invoice_total = overdue_invoice_totals_by_customer.get(cid, Decimal('0.00'))
        overdue_paid_total = overdue_paid_by_customer.get(cid, Decimal('0.00'))
        overdue_credit_total = overdue_credit_by_customer.get(cid, Decimal('0.00'))
        val = overdue_invoice_total - overdue_paid_total - overdue_credit_total
        return val if val > Decimal('0.00') else Decimal('0.00')

    # Apply overdue-only filter based on the selected period (and term rules).
    if show_overdue_only:
        filtered_rows = []
        for row in customer_rows:
            if _overdue_total_for_customer(row['id']) > Decimal('0.00'):
                filtered_rows.append(row)
        customer_rows = filtered_rows

    # Rank customers by selected period totals (desc).
    customer_rows.sort(
        key=lambda r: (
            -(totals_by_customer.get(r['id'], Decimal('0.00'))),
            (r.get('name') or '').lower(),
            r['id'],
        )
    )
    ranked_customer_ids = [row['id'] for row in customer_rows]

    paginator = Paginator(ranked_customer_ids, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    page_ids = list(page_obj.object_list)
    customers_map = Customer.objects.filter(id__in=page_ids, user__in=business_user_ids).in_bulk(page_ids)

    customer_data = []
    for cid in page_ids:
        customer = customers_map.get(cid)
        if not customer:
            continue
        total_amount = totals_by_customer.get(cid, Decimal('0.00'))
        paid_total = paid_by_customer.get(cid, Decimal('0.00'))
        credit_total = credit_by_customer.get(cid, Decimal('0.00'))
        pending_total = total_amount - paid_total - credit_total
        if pending_total < Decimal('0.00'):
            pending_total = Decimal('0.00')

        overdue_total = _overdue_total_for_customer(cid)

        customer_data.append(
            {
                'customer': customer,
                'pending_total': pending_total,
                'paid_total': paid_total,
                'total_amount': total_amount,
                'overdue_total': overdue_total,
            }
        )

    params = request.GET.copy()
    params.pop('page', None)
    query_string = params.urlencode()
    rank_offset = (page_obj.number - 1) * paginator.per_page

    merge_customers = (
        Customer.objects.filter(user__in=business_user_ids)
        .order_by(Lower('name'), 'id')
        .only('id', 'name', 'email')
    )

    context = {
        'customers': customer_data,
        'page_obj': page_obj,
        'query_string': query_string,
        'rank_offset': rank_offset,
        'search_query': search_query,
        'customer_form': CustomerForm(user=request.user),
        'merge_customers': merge_customers,
        'date_range_options': _get_invoice_date_range_options(request.user),
        'current_date_range': date_range_filter,
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'show_overdue_only': show_overdue_only,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('app/customer_list_content.html', context, request=request)
        pagination_html = render_to_string('app/customer_list_pagination.html', context, request=request)
        return JsonResponse({
            'html': html,
            'pagination_html': pagination_html,
            'has_results': bool(page_obj.object_list),
        })

    return render(request, 'app/customer_list.html', context)


@login_required
@require_POST
def merge_customers(request):
    source_id = (request.POST.get('source_customer') or '').strip()
    target_id = (request.POST.get('target_customer') or '').strip()

    if not source_id or not target_id:
        messages.error(request, 'Select two contacts to merge.')
        return redirect('accounts:customer_list')

    if source_id == target_id:
        messages.error(request, 'Choose two different contacts to merge.')
        return redirect('accounts:customer_list')

    business_user_ids = get_customer_user_ids(request.user)
    source_customer = get_object_or_404(Customer, id=source_id, user__in=business_user_ids)
    target_customer = get_object_or_404(Customer, id=target_id, user__in=business_user_ids)
    source_name = source_customer.name
    target_name = target_customer.name

    with transaction.atomic():
        GroupedInvoice.objects.filter(user__in=business_user_ids, customer=source_customer).update(
            customer=target_customer
        )
        GroupedEstimate.objects.filter(user__in=business_user_ids, customer=source_customer).update(
            customer=target_customer
        )
        WorkOrder.objects.filter(user__in=business_user_ids, customer=source_customer).update(
            customer=target_customer
        )
        Vehicle.objects.filter(customer=source_customer).update(customer=target_customer)
        ReminderLog.objects.filter(customer=source_customer).update(customer=target_customer)
        TransportTrip.objects.filter(user__in=business_user_ids, customer=source_customer).update(
            customer=target_customer
        )

        source_transport = TransportCustomer.objects.filter(customer=source_customer).first()
        target_transport = TransportCustomer.objects.filter(customer=target_customer).first()
        if source_transport:
            if target_transport:
                source_transport.delete()
            else:
                source_transport.customer = target_customer
                source_transport.save(update_fields=['customer'])

        source_customer.delete()
        target_customer.update_vehicle_count()

    messages.success(
        request,
        f"Merged {source_name} into {target_name}.",
    )
    return redirect('accounts:customer_list')


@login_required
@require_POST
def create_customer_portal_credentials(request):
    """Create or update portal credentials for a customer and optionally email them."""
    try:
        customer_id = request.POST.get('customer_id')
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password')
        send_email = request.POST.get('send_email') == 'true'

        business_user_ids = get_customer_user_ids(request.user)
        customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
        if not username:
            base = slugify(customer.name or customer.email or f"customer-{customer.id}")[:20] or f"customer-{customer.id}"
            username = base
            counter = 1
            while User.objects.filter(username=username).exclude(id=getattr(customer.portal_user, 'id', None)).exists():
                counter += 1
                username = f"{base}{counter}"

        if not password:
            password = secrets.token_urlsafe(8)

        portal_user = getattr(customer, 'portal_user', None)
        if portal_user:
            portal_user.username = username
            portal_user.email = customer.email or portal_user.email
        else:
            portal_user = User.objects.create_user(username=username, email=customer.email or '')
            customer.portal_user = portal_user

        portal_user.is_active = True
        portal_user.set_password(password)
        portal_user.save()
        customer.save(update_fields=['portal_user'])

        email_sent = False
        if send_email and (customer.email or portal_user.email):
            business_profile = getattr(request.user, 'profile', None)
            business_name = getattr(business_profile, 'company_name', '') or request.user.get_full_name() or request.user.username
            context = {
                'customer': customer,
                'business_name': business_name,
                'username': username,
                'password': password,
                'portal_url': request.build_absolute_uri(reverse('accounts:customer_dashboard')),
            }
            subject = f"{business_name} - your customer portal access"
            context = apply_branding_defaults(context)
            html_body = render_to_string('emails/customer_portal_credentials.html', context)
            text_body = strip_tags(html_body)
            email = EmailMultiAlternatives(
                subject=subject,
                body=text_body,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None) or settings.EMAIL_HOST_USER,
                to=[customer.email or portal_user.email],
            )
            email.attach_alternative(html_body, "text/html")
            try:
                email.send()
                email_sent = True
            except Exception as e:
                logger.warning("Customer portal email send failed: %s", e)
                email_sent = False

        return JsonResponse({
            'success': True,
            'username': username,
            'password': password,
            'email_sent': email_sent,
            'email': customer.email or portal_user.email,
        })
    except Exception as exc:
        logger.error("Failed creating portal credentials: %s", exc, exc_info=True)
        return JsonResponse({'success': False, 'message': 'Unable to create credentials right now.'}, status=400)


@login_required
def customer_vehicle_list(request, customer_id):
    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids) # Ensure customer belongs to user

    # Subquery to get the latest job date for each vehicle
    latest_job_date_subquery = JobHistory.objects.filter(
        vehicle=OuterRef('pk')
    ).order_by('-job_date').values('job_date')[:1]

    active_statuses = VehicleMaintenanceTask.active_statuses()
    today = timezone.localdate()
    next_due_subquery = VehicleMaintenanceTask.objects.filter(
        vehicle=OuterRef('pk'),
        status__in=active_statuses,
    ).order_by('due_date', 'priority', 'title', 'pk')

    vehicles = (
        customer.vehicles.all()
        .annotate(
            last_job_date=Subquery(latest_job_date_subquery),
            job_count=Count('job_history'),
            upcoming_maintenance=Count(
                'maintenance_tasks',
                filter=Q(maintenance_tasks__status__in=active_statuses),
                distinct=True,
            ),
            overdue_maintenance=Count(
                'maintenance_tasks',
                filter=Q(maintenance_tasks__status__in=active_statuses, maintenance_tasks__due_date__lt=today),
                distinct=True,
            ),
            next_due_date=Subquery(next_due_subquery.values('due_date')[:1]),
            next_due_title=Subquery(next_due_subquery.values('title')[:1]),
        )
        .order_by('unit_number', 'vin_number')
    )

    context = {
        'customer': customer,
        'vehicles': vehicles,
    }
    return render(request, 'vehicles/customer_vehicle_list.html', context)

@login_required
def customer_vehicles(request, pk):
    """Return JSON list of vehicles that belong to customer <pk>."""
    business_user_ids = get_customer_user_ids(request.user)
    latest_job_date = JobHistory.objects.filter(
        vehicle=OuterRef("pk")
    ).order_by("-job_date").values("job_date")[:1]

    active_statuses = VehicleMaintenanceTask.active_statuses()
    today = timezone.localdate()
    next_due_subquery = VehicleMaintenanceTask.objects.filter(
        vehicle=OuterRef('pk'),
        status__in=active_statuses,
    ).order_by('due_date', 'priority', 'title', 'pk')

    qs = (
        Vehicle.objects.filter(customer__pk=pk, customer__user__in=business_user_ids)
        .annotate(
            last_job_date=Subquery(latest_job_date),
            job_count=Count("job_history"),
            upcoming_maintenance=Count(
                'maintenance_tasks',
                filter=Q(maintenance_tasks__status__in=active_statuses),
                distinct=True,
            ),
            overdue_maintenance=Count(
                'maintenance_tasks',
                filter=Q(maintenance_tasks__status__in=active_statuses, maintenance_tasks__due_date__lt=today),
                distinct=True,
            ),
            next_due_date=Subquery(next_due_subquery.values('due_date')[:1]),
            next_due_title=Subquery(next_due_subquery.values('title')[:1]),
        )
        .order_by("unit_number", "vin_number")
    )

    data = [
        {
            "id": v.id,
            "unit_no": v.unit_number or "",
            "vin_no": v.vin_number or "",
            "license_plate": v.license_plate or "",
            "make_model": v.make_model or "",
            "year": v.year or "",
            "current_mileage": v.current_mileage,
            "last_job_date": v.last_job_date.isoformat() if v.last_job_date else "",
            "job_count": v.job_count,
            "upcoming_maintenance": v.upcoming_maintenance,
            "overdue_maintenance": v.overdue_maintenance,
            "next_due_date": v.next_due_date.isoformat() if v.next_due_date else "",
            "next_due_title": v.next_due_title or "",
            "detail_url": reverse('accounts:vehicle_detail_dashboard', args=[v.id]),
        }
        for v in qs
    ]
    return JsonResponse({"vehicles": data})


@login_required
def vehicle_detail_dashboard(request, vehicle_id):
    business_user_ids = get_customer_user_ids(request.user)
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, customer__user__in=business_user_ids)

    upcoming_qs = (
        vehicle.maintenance_tasks.filter(status__in=VehicleMaintenanceTask.active_statuses())
        .select_related('work_order')
        .prefetch_related('work_order__assignments__mechanic')
        .order_by('due_date', 'priority', 'title', 'pk')
    )
    completed_qs = (
        vehicle.maintenance_tasks.filter(status=VehicleMaintenanceTask.STATUS_COMPLETED)
        .select_related('work_order')
        .order_by('-completed_date', '-updated_at', '-pk')
    )
    cancelled_qs = (
        vehicle.maintenance_tasks.filter(status=VehicleMaintenanceTask.STATUS_CANCELLED)
        .select_related('work_order')
        .order_by('-updated_at', '-pk')
    )

    upcoming_tasks = list(upcoming_qs)
    completed_tasks = list(completed_qs)
    cancelled_tasks = list(cancelled_qs[:10])

    work_orders = vehicle.work_orders.select_related('customer').order_by('-scheduled_date', '-id')
    open_work_orders = work_orders.exclude(status='completed')

    job_history_qs = (
        vehicle.job_history.all()
        .order_by('-job_date', '-id')
        .select_related('invoice', 'source_income_record')
    )
    job_totals = job_history_qs.aggregate(
        total_service_cost=Coalesce(Sum('service_cost'), Value(Decimal('0.00'))),
        total_tax_amount=Coalesce(Sum('tax_amount'), Value(Decimal('0.00'))),
        total_overall_cost=Coalesce(Sum('total_job_cost'), Value(Decimal('0.00'))),
        job_count=Count('id'),
    )

    parts_usage = list(
        WorkOrderRecord.objects
        .filter(work_order__vehicle=vehicle, product__isnull=False)
        .select_related('product', 'work_order')
        .order_by('-work_order__scheduled_date', '-id')[:20]
    )

    maintenance_form = VehicleMaintenanceTaskForm()
    completion_form = VehicleMaintenanceCompleteForm(vehicle=vehicle, user=request.user)
    quick_workorder_form = VehicleQuickWorkOrderForm(user=request.user, vehicle=vehicle)

    overdue_count = sum(1 for task in upcoming_tasks if task.is_overdue)
    last_completed = completed_tasks[0] if completed_tasks else None

    summary = {
        'upcoming_count': len(upcoming_tasks),
        'overdue_count': overdue_count,
        'open_workorders': open_work_orders.count(),
        'last_completed_date': getattr(last_completed, 'completed_date', None),
        'last_completed_task': last_completed,
    }

    context = {
        'vehicle': vehicle,
        'upcoming_tasks': upcoming_tasks,
        'completed_tasks': completed_tasks[:20],
        'cancelled_tasks': cancelled_tasks,
        'work_orders': work_orders[:20],
        'open_work_orders': open_work_orders,
        'job_history_list': job_history_qs[:50],
        'job_totals': job_totals,
        'parts_usage': parts_usage,
        'maintenance_form': maintenance_form,
        'completion_form': completion_form,
        'quick_workorder_form': quick_workorder_form,
        'summary': summary,
    }
    return render(request, 'vehicles/vehicle_detail_dashboard.html', context)


@login_required
@require_POST
def vehicle_add_maintenance(request, vehicle_id):
    business_user_ids = get_customer_user_ids(request.user)
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, customer__user__in=business_user_ids)
    form = VehicleMaintenanceTaskForm(request.POST)
    if form.is_valid():
        task = form.save(commit=False)
        task.vehicle = vehicle
        task.user = request.user
        task.save()
        messages.success(request, f"Maintenance '{task.title}' added for {vehicle}.")
    else:
        messages.error(request, "Unable to add maintenance task. Please correct the highlighted fields.")
    return redirect('accounts:vehicle_detail_dashboard', vehicle_id=vehicle.id)


@login_required
@require_POST
def vehicle_update_maintenance_status(request, task_id):
    business_user_ids = get_customer_user_ids(request.user)
    task = get_object_or_404(
        VehicleMaintenanceTask,
        id=task_id,
        vehicle__customer__user__in=business_user_ids,
    )
    new_status = request.POST.get('status')
    valid_statuses = {choice[0] for choice in VehicleMaintenanceTask.STATUS_CHOICES}
    if new_status not in valid_statuses:
        messages.error(request, 'Invalid maintenance status provided.')
        return redirect('accounts:vehicle_detail_dashboard', vehicle_id=task.vehicle_id)
    if new_status == VehicleMaintenanceTask.STATUS_COMPLETED:
        messages.info(request, 'Use the completion form to record final details for this maintenance item.')
        return redirect('accounts:vehicle_detail_dashboard', vehicle_id=task.vehicle_id)

    task.status = new_status
    if new_status in {VehicleMaintenanceTask.STATUS_PLANNED, VehicleMaintenanceTask.STATUS_CANCELLED}:
        task.work_order = None
    task.save(update_fields=['status', 'work_order', 'updated_at'])
    messages.success(request, f"Updated maintenance '{task.title}' to {task.get_status_display()}.")
    return redirect('accounts:vehicle_detail_dashboard', vehicle_id=task.vehicle_id)


@login_required
@require_POST
def vehicle_complete_maintenance(request, task_id):
    business_user_ids = get_customer_user_ids(request.user)
    task = get_object_or_404(
        VehicleMaintenanceTask,
        id=task_id,
        vehicle__customer__user__in=business_user_ids,
    )
    form = VehicleMaintenanceCompleteForm(request.POST, instance=task, vehicle=task.vehicle, user=request.user)
    if form.is_valid():
        task = form.save(commit=False)
        task.status = VehicleMaintenanceTask.STATUS_COMPLETED
        if not task.completed_date:
            task.completed_date = timezone.localdate()
        task.save()
        if task.actual_mileage is not None:
            vehicle = task.vehicle
            if vehicle.current_mileage is None or task.actual_mileage > vehicle.current_mileage:
                vehicle.current_mileage = task.actual_mileage
                vehicle.save(update_fields=['current_mileage'])
        messages.success(request, f"Marked '{task.title}' as completed.")
    else:
        messages.error(request, 'Please review the completion details and try again.')
    return redirect('accounts:vehicle_detail_dashboard', vehicle_id=task.vehicle_id)


@login_required
@require_POST
def vehicle_quick_workorder(request, vehicle_id):
    business_user_ids = get_customer_user_ids(request.user)
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, customer__user__in=business_user_ids)
    form = VehicleQuickWorkOrderForm(request.POST, user=request.user, vehicle=vehicle)
    if not form.is_valid():
        messages.error(request, 'Unable to create work order. Please select at least one mechanic.')
        return redirect('accounts:vehicle_detail_dashboard', vehicle_id=vehicle.id)

    mechanics = list(form.cleaned_data['mechanics'])
    maintenance_task = form.cleaned_data.get('maintenance_task')
    scheduled_date = form.cleaned_data['scheduled_date']

    if maintenance_task and maintenance_task.work_order_id:
        workorder = maintenance_task.work_order
        assignments = list(workorder.assignments.select_related('mechanic'))
        assigned_names = ', '.join(a.mechanic.name for a in assignments) or 'Unassigned'
        messages.info(
            request,
            f"Maintenance task already linked to Work order #{workorder.id} (Status: {workorder.get_status_display()}). "
            f"Assigned to: {assigned_names}."
        )
        return redirect('accounts:vehicle_detail_dashboard', vehicle_id=vehicle.id)

    if maintenance_task and maintenance_task.description:
        description = f"{maintenance_task.title} - {maintenance_task.description}"
    elif maintenance_task:
        description = maintenance_task.title
    else:
        parts = [vehicle.make_model or '', vehicle.unit_number or '', vehicle.vin_number or '']
        friendly_name = ' '.join(part for part in parts if part).strip()
        description = f"Maintenance for {friendly_name or f'Vehicle #{vehicle.id}'}"

    try:
        with transaction.atomic():
            workorder = WorkOrder.objects.create(
                user=request.user,
                customer=vehicle.customer,
                vehicle=vehicle,
                scheduled_date=scheduled_date,
                vehicle_vin=vehicle.vin_number,
                mileage=vehicle.current_mileage,
                unit_no=vehicle.unit_number,
                make_model=vehicle.make_model,
                description=description[:500],
                status='pending',
            )
            sync_result = sync_workorder_assignments(workorder, mechanics)
            for assignment in sync_result['created']:
                notify_mechanic_assignment(workorder, assignment, request=request)
            if maintenance_task:
                maintenance_task.status = VehicleMaintenanceTask.STATUS_IN_PROGRESS
                maintenance_task.work_order = workorder
                maintenance_task.save(update_fields=['status', 'work_order', 'updated_at'])
    except Exception as exc:
        messages.error(request, f'Failed to create work order: {exc}')
        return redirect('accounts:vehicle_detail_dashboard', vehicle_id=vehicle.id)

    assigned_names = ', '.join(mechanic.name for mechanic in mechanics)
    messages.success(request, f"Work order #{workorder.id} created and assigned to {assigned_names}.")
    return redirect('accounts:vehicle_detail_dashboard', vehicle_id=vehicle.id)

@login_required
@require_POST # Ensures this view only accepts POST requests
def add_vehicle(request):
    form = VehicleForm(request.POST, request.FILES or None)
    customer_id = request.POST.get('customer')
    business_user_ids = get_customer_user_ids(request.user)

    # Log incoming data for debugging
    logger.info(f"add_vehicle attempt by user {request.user.username}. POST data: {request.POST}")

    if not customer_id:
        logger.warning(f"add_vehicle: No customer_id provided by user {request.user.username}.")
        return JsonResponse({
            'success': False,
            'errors': {'customer': [{'message': 'Customer ID was not provided in the request.'}]}
        }, status=400)

    try:
        # Fetch the customer instance. Ensure it belongs to the current user for security.
        customer_instance = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
        logger.info(f"add_vehicle: Customer '{customer_instance}' found for user {request.user.username}.")
    except Customer.DoesNotExist:
        logger.error(f"add_vehicle: Invalid or unauthorized customer_id '{customer_id}' for user {request.user.username}.")
        return JsonResponse({
            'success': False,
            'errors': {'customer': [{'message': 'Invalid Customer ID or customer does not belong to this user.'}]}
        }, status=400)
    except Exception as e:
        logger.error(f"add_vehicle: Error retrieving customer '{customer_id}' for user {request.user.username}: {str(e)}")
        return JsonResponse({
            'success': False,
            'errors': {'customer': [{'message': f'Error retrieving customer: {str(e)}'}]}
        }, status=500)

    if form.is_valid():
        try:
            vehicle = form.save(commit=False)
            vehicle.customer = customer_instance  # *** Assign the customer instance ***
            vehicle.user = request.user # Assuming your Vehicle model has a 'user' field
            vehicle.save()
            logger.info(f"add_vehicle: Vehicle '{vehicle}' saved successfully for customer '{customer_instance}' by user {request.user.username}.")
            return JsonResponse({
                'success': True,
                'vehicle': {
                    'id': vehicle.pk,
                    'unit_no': vehicle.unit_number,  # Use the actual model field name
                    'vin_no': vehicle.vin_number,    # Use the actual model field name
                    'license_plate': vehicle.license_plate,
                    'make_model': vehicle.make_model, # This one seems to be correct
                    'year': vehicle.year,
                    'current_mileage': vehicle.current_mileage,
                }
            })
        except Exception as e:
            logger.error(f"add_vehicle: Error during vehicle.save() for user {request.user.username}: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'errors': {'__all__': [{'message': f'An error occurred while saving the vehicle: {str(e)}'}]}
            }, status=500)
    else:
        logger.warning(f"add_vehicle: Form validation failed for user {request.user.username}. Errors: {form.errors.as_json()}")
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)

# --- PLACEHOLDER VIEWS FOR VEHICLE CRUD (implement later) ---
@login_required
@require_POST # Ensure this view only accepts POST requests for safety
def edit_vehicle(request, vehicle_id):
    # Ensure the vehicle belongs to a customer of the logged-in user for security
    business_user_ids = get_customer_user_ids(request.user)
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, customer__user__in=business_user_ids)

    if request.method == 'POST':
        form = VehicleForm(request.POST, instance=vehicle)
        if form.is_valid():
            form.save()
            # Optionally, prepare data for the updated row if you want to send it back
            # For now, just success
            return JsonResponse({'success': True,
                                 'vehicle': {
                                     'id': vehicle.id,
                                     'unit_number': vehicle.unit_number or "N/A",
                                     'vin_number': vehicle.vin_number,
                                     'license_plate': vehicle.license_plate or "",
                                     'make_model': vehicle.make_model or "N/A",
                                     'year': vehicle.year,
                                 }})
        else:
            # Collect form errors
            errors = {field: error[0] for field, error in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)

    # GET request not typically used for AJAX save, but could return form if needed
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
@require_POST # Important for delete operations
def delete_vehicle(request, vehicle_id):
    # Ensure the vehicle belongs to a customer of the logged-in user for security
    business_user_ids = get_customer_user_ids(request.user)
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, customer__user__in=business_user_ids)

    try:
        vehicle_name = str(vehicle) # For confirmation message
        vehicle.delete()
        return JsonResponse({'success': True, 'message': f'Vehicle {vehicle_name} deleted successfully.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# accounts/views.py
# ... (other imports) ...
from django.db.models import Sum, Count # Ensure Sum is imported

@login_required
def vehicle_job_history_list(request, vehicle_id):
    # Ensure the vehicle belongs to a customer of the logged-in user
    business_user_ids = get_customer_user_ids(request.user)
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, customer__user__in=business_user_ids)

    job_history = vehicle.job_history.all().order_by('-job_date', '-id').select_related('invoice', 'source_income_record') # Get all jobs for this vehicle

    # Calculate sums
    totals = job_history.aggregate(
        total_service_cost=Coalesce(Sum('service_cost'), Value(Decimal('0.00'))),
        total_tax_amount=Coalesce(Sum('tax_amount'), Value(Decimal('0.00'))),
        total_overall_cost=Coalesce(Sum('total_job_cost'), Value(Decimal('0.00'))),
        job_count=Count('id')
    )

    context = {
        'vehicle': vehicle,
        'job_history_list': job_history,
        'totals': totals,
    }
    return render(request, 'vehicles/vehicle_job_history_list.html', context)


@login_required
def vehicle_management(request):
    """Dashboard page to manage vehicles grouped by customer."""
    business_user_ids = get_customer_user_ids(request.user)
    customers = Customer.objects.filter(user__in=business_user_ids).order_by('name')
    customer_id = request.GET.get('customer')
    selected_customer = None
    if customer_id:
        selected_customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
    customer_form = CustomerForm(instance=selected_customer, user=request.user)
    context = {
        'customers': customers,
        'selected_customer': selected_customer,
        'customer_form': customer_form,
    }
    return render(request, 'accounts/mach/vehicle_management.html', context)


@login_required
@require_POST
def edit_customer_ajax(request, customer_id):
    """Inline edit for customer details via AJAX."""
    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
    form = CustomerForm(request.POST, instance=customer, user=request.user)
    if form.is_valid():
        try:
            customer = form.save()
        except IntegrityError:
            return JsonResponse(
                {
                    'success': False,
                    'errors': {
                        'portal_username': 'This portal username is already taken.'
                    }
                },
                status=400,
            )
        if form.cleaned_data.get('register_portal') and not customer.portal_user:
            return JsonResponse(
                {
                    'success': False,
                    'errors': {
                        'portal_username': 'Customer portal access could not be enabled.'
                    }
                },
                status=400,
            )
        return JsonResponse({
            'success': True,
            'customer': {
                'id': customer.id,
                'name': customer.name,
                'email': customer.email,
                'address': customer.address,
            }
        })
    else:
        errors = {f: e[0] for f, e in form.errors.items()}
        return JsonResponse({'success': False, 'errors': errors}, status=400)

VEHICLE_TEMPLATE_HEADERS = [
    "VIN Number",
    "Unit Number",
    "License Plate",
    "Make & Model",
    "Current Mileage",
]

VEHICLE_MAINTENANCE_TEMPLATE_HEADERS = [
    "Title",
    "Description",
    "Due Date",
    "Due Mileage",
    "Mileage Interval",
    "Priority",
    "Status",
    "Last Reminder Sent",
]

VEHICLE_MAINTENANCE_PRIORITY_MAP = {}
for value, label in VehicleMaintenanceTask.PRIORITY_CHOICES:
    VEHICLE_MAINTENANCE_PRIORITY_MAP[value.lower()] = value
    VEHICLE_MAINTENANCE_PRIORITY_MAP[label.lower()] = value

VEHICLE_MAINTENANCE_STATUS_MAP = {}
for value, label in VehicleMaintenanceTask.STATUS_CHOICES:
    VEHICLE_MAINTENANCE_STATUS_MAP[value.lower()] = value
    VEHICLE_MAINTENANCE_STATUS_MAP[label.lower()] = value

VEHICLE_MAINTENANCE_INTERVAL_MAP = {}
for value, label in VehicleMaintenanceTask.mileage_interval_choices():
    VEHICLE_MAINTENANCE_INTERVAL_MAP[label.lower()] = value
    VEHICLE_MAINTENANCE_INTERVAL_MAP[str(value)] = value
    VEHICLE_MAINTENANCE_INTERVAL_MAP[str(value).lower()] = value
    formatted_value = f"{value:,}"
    VEHICLE_MAINTENANCE_INTERVAL_MAP[formatted_value] = value
    VEHICLE_MAINTENANCE_INTERVAL_MAP[formatted_value.lower()] = value


def _normalize_excel_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _sanitize_filename_segment(value):
    safe = re.sub(r"[^0-9a-zA-Z_-]+", "_", str(value or "export").strip())
    safe = safe.strip("_")
    return safe or "export"


def _parse_excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_excel_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        if timezone.is_naive(value):
            return timezone.make_aware(value, timezone.get_current_timezone())
        return timezone.localtime(value)
    if isinstance(value, datetime.date):
        combined = datetime.datetime.combine(value, datetime.time())
        return timezone.make_aware(combined, timezone.get_current_timezone())

    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    ):
        try:
            parsed = datetime.datetime.strptime(text, fmt)
        except ValueError:
            continue
        if timezone.is_naive(parsed):
            return timezone.make_aware(parsed, timezone.get_current_timezone())
        return timezone.localtime(parsed)

    parsed_date = _parse_excel_date(value)
    if parsed_date is None:
        return None
    combined = datetime.datetime.combine(parsed_date, datetime.time())
    return timezone.make_aware(combined, timezone.get_current_timezone())


def _get_header_indexes(header_row, expected_headers, optional_headers=None):
    optional_headers = set(optional_headers or [])
    normalized = [_normalize_excel_text(cell) for cell in header_row]
    missing = [
        header
        for header in expected_headers
        if header not in normalized and header not in optional_headers
    ]
    if missing:
        return None, missing
    return {
        header: normalized.index(header)
        for header in expected_headers
        if header in normalized
    }, None


def _get_cell(row, index):
    if row and index < len(row):
        return row[index]
    return None


def _normalize_choice(value, mapping, default):
    normalized = _normalize_excel_text(value)
    if not normalized:
        return default
    return mapping.get(normalized.lower(), default)


@login_required
def export_vehicles_template(request):
    customer_id = request.GET.get("customer_id")
    if not customer_id:
        messages.error(request, "Select a customer before downloading the vehicle template.")
        return redirect("accounts:vehicle_management")

    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Vehicles"
    worksheet.append(VEHICLE_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    vehicles = Vehicle.objects.filter(customer=customer).order_by("vin_number")
    for vehicle in vehicles:
        worksheet.append(
            [
                vehicle.vin_number or "",
                vehicle.unit_number or "",
                vehicle.license_plate or "",
                vehicle.make_model or "",
                vehicle.current_mileage if vehicle.current_mileage is not None else "",
            ]
        )

    apply_template_styling(
        worksheet,
        headers=VEHICLE_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={
            "VIN Number": 26,
            "Unit Number": 18,
            "License Plate": 18,
            "Make & Model": 34,
            "Current Mileage": 20,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"vehicles_{_sanitize_filename_segment(customer.name)}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def import_vehicles_from_excel(request):
    customer_id = request.POST.get("customer_id")
    if not customer_id:
        messages.error(request, "Select a customer before importing vehicles.")
        return redirect("accounts:vehicle_management")

    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please choose an Excel file to import.")
        return redirect(reverse("accounts:vehicle_management") + f"?customer={customer.id}")

    try:
        workbook = load_workbook(upload, data_only=True)
    except InvalidFileException:
        messages.error(request, "Please upload a valid .xlsx file.")
        return redirect(reverse("accounts:vehicle_management") + f"?customer={customer.id}")
    except Exception:
        messages.error(request, "Unable to read the uploaded file.")
        return redirect(reverse("accounts:vehicle_management") + f"?customer={customer.id}")

    worksheet = workbook.active
    try:
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        messages.error(request, "The uploaded file is empty.")
        return redirect(reverse("accounts:vehicle_management") + f"?customer={customer.id}")

    header_indexes, missing_headers = _get_header_indexes(header_row, VEHICLE_TEMPLATE_HEADERS)
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: " + ", ".join(missing_headers),
        )
        return redirect(reverse("accounts:vehicle_management") + f"?customer={customer.id}")

    created = updated = 0
    errors = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell in (None, "") for cell in row):
            continue

        vin_value = _get_cell(row, header_indexes["VIN Number"])
        vin = _normalize_excel_text(vin_value)
        unit_number = _normalize_excel_text(_get_cell(row, header_indexes["Unit Number"]))
        if not vin and not unit_number:
            errors.append(f"Row {row_index}: Provide at least VIN Number or Unit Number.")
            continue

        license_plate = _normalize_excel_text(_get_cell(row, header_indexes["License Plate"]))
        make_model = _normalize_excel_text(_get_cell(row, header_indexes["Make & Model"]))
        mileage_raw = _get_cell(row, header_indexes["Current Mileage"])
        mileage = None
        if mileage_raw not in (None, ""):
            try:
                mileage = int(float(mileage_raw))
                if mileage < 0:
                    raise ValueError
            except (TypeError, ValueError):
                errors.append(f"Row {row_index}: Invalid mileage '{mileage_raw}'.")
                continue

        existing_vehicle = None
        if vin:
            existing_vehicle = (
                Vehicle.objects.filter(vin_number__iexact=vin)
                .select_related("customer")
                .first()
            )
        elif unit_number:
            # Fallback match when VIN is missing: match within this customer by Unit Number.
            existing_vehicle = (
                Vehicle.objects.filter(customer=customer, unit_number__iexact=unit_number)
                .select_related("customer")
                .first()
            )
        if existing_vehicle and existing_vehicle.customer.user_id not in business_user_ids:
            errors.append(f"Row {row_index}: VIN '{vin}' belongs to another account.")
            continue

        try:
            if existing_vehicle:
                existing_vehicle.customer = customer
                existing_vehicle.unit_number = unit_number or None
                existing_vehicle.license_plate = license_plate or None
                existing_vehicle.make_model = make_model or None
                existing_vehicle.current_mileage = mileage
                existing_vehicle.save()
                updated += 1
            else:
                Vehicle.objects.create(
                    customer=customer,
                    vin_number=vin or None,
                    unit_number=unit_number or None,
                    license_plate=license_plate or None,
                    make_model=make_model or None,
                    current_mileage=mileage,
                )
                created += 1
        except IntegrityError:
            if vin:
                errors.append(f"Row {row_index}: A vehicle with VIN '{vin}' already exists.")
            else:
                errors.append(f"Row {row_index}: A vehicle with Unit '{unit_number}' already exists.")

    if created or updated:
        summary_parts = []
        if created:
            summary_parts.append(f"{created} vehicle{'s' if created != 1 else ''} created")
        if updated:
            summary_parts.append(f"{updated} vehicle{'s' if updated != 1 else ''} updated")
        messages.success(
            request,
            "Vehicle import complete: " + " and ".join(summary_parts),
        )
    else:
        messages.info(request, "No vehicles were added or updated.")

    if errors:
        preview = "; ".join(errors[:4])
        if len(errors) > 4:
            preview += f"; and {len(errors) - 4} more issue(s)"
        messages.warning(request, f"Issues processing the import: {preview}")

    return redirect(reverse("accounts:vehicle_management") + f"?customer={customer.id}")


@login_required
def export_vehicle_maintenance_template(request, vehicle_id):
    business_user_ids = get_customer_user_ids(request.user)
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, customer__user__in=business_user_ids)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Maintenance Tasks"
    worksheet.append(VEHICLE_MAINTENANCE_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    def add_dropdown(values, column_letter):
        if not values:
            return
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(values)}"',
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
        )
        validation.errorTitle = "Invalid selection"
        validation.error = "Please choose a value from the list."
        validation.promptTitle = "Select a value"
        validation.prompt = "Choose from the available options."
        worksheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}1048576")

    add_dropdown(
        [label for _, label in VehicleMaintenanceTask.mileage_interval_choices()],
        "E",
    )
    add_dropdown([label for _, label in VehicleMaintenanceTask.PRIORITY_CHOICES], "F")
    add_dropdown([label for _, label in VehicleMaintenanceTask.STATUS_CHOICES], "G")

    tasks = vehicle.maintenance_tasks.order_by("due_date", "title")
    for task in tasks:
        worksheet.append(
            [
                task.title or "",
                task.description or "",
                task.due_date.isoformat() if task.due_date else "",
                task.due_mileage if task.due_mileage is not None else "",
                task.mileage_interval_display or "",
                task.get_priority_display() if hasattr(task, "get_priority_display") else (task.priority or ""),
                task.get_status_display() if hasattr(task, "get_status_display") else (task.status or ""),
                (
                    timezone.localtime(task.last_reminder_sent).isoformat()
                    if task.last_reminder_sent
                    else ""
                ),
            ]
        )

    apply_template_styling(
        worksheet,
        headers=VEHICLE_MAINTENANCE_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={
            "Title": 32,
            "Description": 54,
            "Due Date": 20,
            "Due Mileage": 20,
            "Mileage Interval": 22,
            "Priority": 20,
            "Status": 20,
            "Last Reminder Sent": 26,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"maintenance_{_sanitize_filename_segment(vehicle.vin_number or vehicle.id)}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def import_vehicle_maintenance_from_excel(request, vehicle_id):
    business_user_ids = get_customer_user_ids(request.user)
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, customer__user__in=business_user_ids)
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please choose an Excel file to import.")
        return redirect("accounts:vehicle_detail_dashboard", vehicle_id=vehicle.id)

    try:
        workbook = load_workbook(upload, data_only=True)
    except InvalidFileException:
        messages.error(request, "Please upload a valid .xlsx file.")
        return redirect("accounts:vehicle_detail_dashboard", vehicle_id=vehicle.id)
    except Exception:
        messages.error(request, "Unable to read the uploaded file.")
        return redirect("accounts:vehicle_detail_dashboard", vehicle_id=vehicle.id)

    worksheet = workbook.active
    try:
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        messages.error(request, "The uploaded file is empty.")
        return redirect("accounts:vehicle_detail_dashboard", vehicle_id=vehicle.id)

    header_indexes, missing_headers = _get_header_indexes(
        header_row,
        VEHICLE_MAINTENANCE_TEMPLATE_HEADERS,
        optional_headers=["Last Reminder Sent"],
    )
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: " + ", ".join(missing_headers),
        )
        return redirect("accounts:vehicle_detail_dashboard", vehicle_id=vehicle.id)

    created = 0
    errors = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell in (None, "") for cell in row):
            continue

        title = _normalize_excel_text(_get_cell(row, header_indexes["Title"]))
        if not title:
            errors.append(f"Row {row_index}: Title is required.")
            continue

        description = _normalize_excel_text(_get_cell(row, header_indexes["Description"]))
        due_date_raw = _get_cell(row, header_indexes["Due Date"])
        due_date = _parse_excel_date(due_date_raw)
        if due_date_raw not in (None, "") and due_date is None:
            errors.append(f"Row {row_index}: Unable to parse Due Date '{due_date_raw}'.")
            continue

        due_mileage_raw = _get_cell(row, header_indexes["Due Mileage"])
        due_mileage = None
        if due_mileage_raw not in (None, ""):
            try:
                due_mileage = int(float(due_mileage_raw))
            except (TypeError, ValueError):
                errors.append(f"Row {row_index}: Invalid Due Mileage '{due_mileage_raw}'.")
                continue

        mileage_interval_raw = _get_cell(row, header_indexes["Mileage Interval"])
        mileage_interval = None
        if mileage_interval_raw not in (None, ""):
            normalized_interval = str(mileage_interval_raw).strip()
            mileage_interval = VEHICLE_MAINTENANCE_INTERVAL_MAP.get(normalized_interval.lower())
            if mileage_interval is None:
                try:
                    numeric_value = normalized_interval.replace(",", "")
                    mileage_interval = int(float(numeric_value))
                except (TypeError, ValueError):
                    errors.append(
                        f"Row {row_index}: Invalid Mileage Interval '{mileage_interval_raw}'."
                    )
                    continue

        priority = _normalize_choice(
            _get_cell(row, header_indexes["Priority"]),
            VEHICLE_MAINTENANCE_PRIORITY_MAP,
            VehicleMaintenanceTask.PRIORITY_MEDIUM,
        )
        status = _normalize_choice(
            _get_cell(row, header_indexes["Status"]),
            VEHICLE_MAINTENANCE_STATUS_MAP,
            VehicleMaintenanceTask.STATUS_PLANNED,
        )

        last_reminder = None
        if "Last Reminder Sent" in header_indexes:
            last_reminder_raw = _get_cell(row, header_indexes["Last Reminder Sent"])
            last_reminder = _parse_excel_datetime(last_reminder_raw)
            if last_reminder_raw not in (None, "") and last_reminder is None:
                errors.append(
                    f"Row {row_index}: Unable to parse Last Reminder Sent '{last_reminder_raw}'."
                )
                continue

        VehicleMaintenanceTask.objects.create(
            vehicle=vehicle,
            user=request.user,
            title=title,
            description=description or "",
            due_date=due_date,
            due_mileage=due_mileage,
            mileage_interval=mileage_interval,
            priority=priority,
            status=status,
            last_reminder_sent=last_reminder,
        )
        created += 1

    if created:
        messages.success(
            request,
            f"Imported {created} maintenance task{'s' if created != 1 else ''}.",
        )
    else:
        messages.info(request, "No maintenance tasks were added.")

    if errors:
        preview = "; ".join(errors[:4])
        if len(errors) > 4:
            preview += f"; and {len(errors) - 4} more issue(s)"
        messages.warning(request, f"Issues processing the import: {preview}")

    return redirect("accounts:vehicle_detail_dashboard", vehicle_id=vehicle.id)

@login_required
def customer_overdue_list(request):
    """
    Displays a table of customers with overdue amounts,
    and incorporates reminder counts and last reminder dates.
    """
    search_query = request.GET.get('q', '')

    # Ensure user profile exists, or handle appropriately based on your app structure
    try:
        user_profile = request.user.profile
    except Profile.DoesNotExist:
        # Handle case where profile doesn't exist, e.g., create one or redirect
        # For this example, we'll assume it exists or is handled elsewhere.
        # If you have a signal to create Profile on User creation, this might not be an issue.
        # As a fallback for demonstration, you might pass a default or raise an error.
        # This depends on your application's design.
        # For now, let's assume it's there or your app handles its absence.
        user_profile = None # Or some default if necessary for TERM_CHOICES logic

    business_user_ids = get_customer_user_ids(request.user)
    store_user_ids = get_business_user_ids(request.user)
    customers_query = Customer.objects.filter(user__in=business_user_ids)
    if search_query:
        customers_query = customers_query.filter(name__icontains=search_query)

    current_date = timezone.now().date()
    overdue_customers_data = []

    for customer in customers_query:
        # Calculate balance_due for each invoice of the customer
        invoices_with_balance = customer.invoices.filter(
            user__in=store_user_ids,
        ).annotate(
            total_paid=Coalesce(
                Sum('payments__amount', distinct=True), # Ensure payments aren't double-counted if joins occur
                Value(Decimal('0.00')),
                output_field=DecimalField()
            )
        )
        invoices_with_balance = _annotate_invoice_credit_totals(invoices_with_balance).annotate(
            balance_due=F('total_amount') - F('total_paid') - F('credit_total')
        )

        # Calculate total overdue amount for this specific customer
        overdue_total_for_customer = Decimal('0.00')
        for invoice in invoices_with_balance:
            if invoice.due_date and invoice.due_date < current_date and invoice.balance_due > 0:
                overdue_total_for_customer += invoice.balance_due

        if overdue_total_for_customer > 0:
            # Fetch reminder info for this customer using the ReminderLog model
            customer_reminders = customer.reminder_logs.all()  # .all() respects Meta ordering
            reminder_count = customer_reminders.count()
            last_reminder = customer_reminders.first() # Gets the most recent
            last_reminder_sent_on = last_reminder.sent_at if last_reminder else None

            overdue_customers_data.append({
                'customer': customer,
                'overdue_total': overdue_total_for_customer,
                'reminder_count': reminder_count,
                'last_reminder_sent_on': last_reminder_sent_on,
            })

    # Calculate the sum of overdue amounts for the customers being displayed
    total_overdue_for_displayed_customers = sum(d['overdue_total'] for d in overdue_customers_data)

    # The 'overdue_total_balance' for the top card could be this sum,
    # or your get_overdue_total_balance might provide a system-wide figure.
    # Here, we use the sum of displayed customers, which seems consistent with the table.
    # If get_overdue_total_balance is meant to be the definitive source for the card, use that:
    # term_days = TERM_CHOICES.get(user_profile.term, 30) if user_profile else 30
    # card_total_overdue = get_overdue_total_balance(request.user, term_days)

    overdue_customers_data.sort(key=lambda x: x['overdue_total'], reverse=True)

    context = {
        'overdue_customers': overdue_customers_data,
        'overdue_total_balance': total_overdue_for_displayed_customers, # This total is for the card
        'search_query': search_query,
        'today_date': current_date, # For template comparison ("Today" vs. date)
    }
    return render(request, 'app/customer_overdue_list.html', context)

@login_required
def customer_detail(request, customer_id, invoice_type):
    """
    Displays customer detail with invoices filtered by invoice_type.
    invoice_type can be 'pending', 'paid', 'all', or 'overdue'.
    For 'overdue', only invoices with due_date earlier than today are shown.
    """
    business_user_ids = get_customer_user_ids(request.user)
    store_user_ids = get_business_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
    invoices_qs = customer.invoices.filter(user__in=store_user_ids).annotate(
        total_paid=Coalesce(
            Sum('payments__amount'),
            Value(Decimal('0.00')),
            output_field=DecimalField()
        )
    )
    invoices_qs = _annotate_invoice_credit_totals(invoices_qs).annotate(
        balance_due=F('total_amount') - F('total_paid') - F('credit_total')
    )
    if invoice_type == 'paid':
        invoices = invoices_qs.filter(balance_due__lte=Decimal('0.00'))
    elif invoice_type == 'pending':
        invoices = invoices_qs.filter(balance_due__gt=Decimal('0.00'))
    elif invoice_type == 'overdue':
        pending_invoices = invoices_qs.filter(balance_due__gt=Decimal('0.00'))
        today = timezone.now().date()
        invoices = [inv for inv in pending_invoices if inv.due_date and inv.due_date < today]
    else:  # 'all'
        invoices = invoices_qs

    if invoice_type in ['pending', 'paid', 'all']:
        total_invoice_amount = invoices.aggregate(
            total=Coalesce(Sum('total_amount'), Value(Decimal('0.00')), output_field=DecimalField())
        )['total'] or Decimal('0.00')
    elif invoice_type == 'overdue':
        total_invoice_amount = sum(inv.total_amount for inv in invoices)

    credit_entries = []
    credits_qs = CustomerCredit.objects.filter(
        user__in=business_user_ids,
        customer=customer,
    ).prefetch_related(
        'items',
        'items__product',
        'items__source_invoice',
    ).order_by('-date', '-id')
    for credit in credits_qs:
        total_incl_tax, total_tax, total_excl_tax = credit.calculate_totals()
        total_incl_tax = Decimal(str(total_incl_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_tax = Decimal(str(total_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_excl_tax = Decimal(str(total_excl_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        credit_entries.append({
            'credit': credit,
            'items': list(credit.items.all()),
            'total_incl_tax': total_incl_tax,
            'total_excl_tax': total_excl_tax,
            'total_tax': total_tax,
            'display_total': -total_incl_tax,
        })

    context = {
        'customer': customer,
        'invoices': invoices,
        'total_invoice_amount': total_invoice_amount,
        'invoice_type': invoice_type,
        'credit_entries': credit_entries,
    }
    return render(request, 'app/customer_jobs.html', context)




@login_required
def customer_edit(request, customer_id):
    business_user = get_business_user(request.user) or request.user
    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer, user=business_user)
        if form.is_valid():
            try:
                customer = form.save()
            except IntegrityError:
                form.add_error('portal_username', 'This portal username is already taken.')
                messages.error(request, 'Customer portal access could not be enabled.')
            else:
                if form.cleaned_data.get('register_portal') and not customer.portal_user:
                    form.add_error(None, 'Customer portal access could not be enabled. Please verify the portal details.')
                    messages.error(request, 'Customer portal access could not be enabled.')
                else:
                    messages.success(request, 'Customer details updated successfully.')
                    return redirect('accounts:customer_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomerForm(instance=customer, user=business_user)

    return render(request, 'app/customer_edit.html', {'form': form, 'customer': customer})

@login_required
def delete_customer(request, customer_id):
    business_user_ids = get_customer_user_ids(request.user)
    customer = get_object_or_404(Customer, id=customer_id, user__in=business_user_ids)

    # Simply delete the customer (the template ensures only users with no balance can delete)
    customer.delete()
    messages.success(request, f'Customer {customer.name} has been successfully deleted.')

    return redirect('accounts:customer_list')

@csrf_exempt
def send_invoice_email(request):
    if request.method == 'POST':
        try:
            # Get form data
            email = request.POST['email']
            subject = request.POST['subject']
            body = request.POST['body']
            include_workorder = request.POST.get('include_workorder') == '1'
            invoice_number = request.POST.get('invoice_number')
            attachment = request.FILES.get('attachment')
            cc_business_owner = request.POST.get('cc_business_owner') == '1'
            
            # Parse recipients and CC recipients if provided
            recipients = json.loads(request.POST.get('recipients', '[]')) if request.POST.get('recipients') else [email]
            cc_recipients = json.loads(request.POST.get('cc_recipients', '[]')) if request.POST.get('cc_recipients') else []

            # Always CC the business profile/company email (if available)
            business_email = getattr(getattr(request.user, "profile", None), "company_email", None)

            invoice = None
            if invoice_number:
                invoice = GroupedInvoice.objects.filter(invoice_number=invoice_number).first()
            customer_cc_emails = []
            if invoice and invoice.customer:
                customer_cc_emails = invoice.customer.get_cc_emails()

            cc_recipients = build_cc_list(
                *cc_recipients,
                business_email,
                *customer_cc_emails,
                exclude=recipients,
            )

            # Create email message with TO and CC recipients
            all_recipients = recipients.copy()
            email_message = EmailMessage(
                subject,
                body,
                settings.SUPPORT_EMAIL,
                recipients,  # TO recipients
                cc=cc_recipients if cc_recipients else None  # CC recipients
            )
            
            # Attach the main PDF
            if attachment:
                email_message.attach(attachment.name, attachment.read(), 'application/pdf')
                
            # Attach work order if requested
            if include_workorder and invoice and invoice.work_order:
                wo_pdf = generate_workorder_pdf(invoice.work_order, request)
                email_message.attach(
                    f"WorkOrder_{invoice.work_order.id}.pdf", wo_pdf, 'application/pdf'
                )
                
            email_message.content_subtype = "html"  # Main content is now text/html
            email_message.send()
            if invoice:
                log_invoice_activity(
                    invoice,
                    event_type=InvoiceActivity.EVENT_EMAIL_SENT,
                    request=request,
                )

            # Log the email sending
            recipient_info = f"TO: {', '.join(recipients)}"
            if cc_recipients:
                recipient_info += f", CC: {', '.join(cc_recipients)}"
            
            logger.info("Email sent successfully - %s", recipient_info)
            return JsonResponse({
                'status': 'success', 
                'message': f'Email sent successfully to {len(recipients + cc_recipients)} recipient(s).'
            })
        except json.JSONDecodeError as e:
            logger.error("Error parsing recipients JSON: %s", str(e))
            return JsonResponse({'status': 'failed', 'error': 'Invalid recipient data format'}, status=400)
        except Exception as e:
            logger.error("Error sending email: %s", str(e))
            return JsonResponse({'status': 'failed', 'error': str(e)}, status=400)
    return JsonResponse({'status': 'failed'}, status=400)

def success_page(request):
    return render(request, 'app/success.html', {'message': 'Email sent successfully!'})

@login_required
def vendor_list(request):
    if request.method == 'POST':
        supplier_id = request.POST.get('supplier_id')
        supplier = get_object_or_404(Supplier, id=supplier_id, user=request.user)
        form = SupplierPortalForm(request.POST, supplier=supplier, prefix=str(supplier.id))
        if form.is_valid():
            form.save()
            if form.cleaned_data.get('register_portal'):
                messages.success(request, f'Supplier login enabled for {supplier.name}.')
            else:
                messages.success(request, f'Supplier login disabled for {supplier.name}.')
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
        return redirect('accounts:supplier_list')

    query = (request.GET.get('q', '') or '').strip()
    date_range_key, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )

    supplier_filter = Q(user=request.user)
    if query:
        supplier_filter &= Q(vendor__icontains=query)
    if start_date:
        supplier_filter &= Q(date__gte=start_date)
    if end_date:
        supplier_filter &= Q(date__lte=end_date)

    expense_queryset = (
        MechExpense.objects.filter(supplier_filter)
        .select_related('user__profile')
        .prefetch_related('mechexpenseitem_set', 'supplier_credit_items__supplier_credit')
        .annotate(paid_total=Coalesce(Sum('payments__amount'), Value(Decimal('0.00'))))
    )

    supplier_totals = {}
    supplier_cache = {}

    for expense in expense_queryset:
        total_incl_tax, total_tax, total_excl_tax = expense.calculate_totals()
        total_incl_tax = Decimal(str(total_incl_tax)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_tax = Decimal(str(total_tax)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_excl_tax = Decimal(str(total_excl_tax)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        raw_name = (expense.vendor or '').strip()
        cache_key = raw_name.lower()
        supplier = None
        if raw_name:
            supplier = supplier_cache.get(cache_key)
            if supplier is None:
                supplier = _get_or_create_supplier_for_user(request.user, raw_name)
                supplier_cache[cache_key] = supplier

        display_name = supplier.name if supplier else (raw_name or 'Unassigned Supplier')
        totals_key = (supplier.id if supplier else None, cache_key or '__unassigned__')

        data = supplier_totals.setdefault(
            totals_key,
            {
                'supplier': supplier,
                'supplier_name': display_name,
                'lookup_value': raw_name,
                'total_amount': Decimal('0.00'),
                'total_tax': Decimal('0.00'),
                'total_incl_tax': Decimal('0.00'),
                'total_paid': Decimal('0.00'),
                'credit_total_excl_tax': Decimal('0.00'),
                'credit_total_tax': Decimal('0.00'),
                'credit_total_incl_tax': Decimal('0.00'),
                'remaining_total': Decimal('0.00'),
                'credit_available_incl_tax': Decimal('0.00'),
            },
        )
        data['lookup_value'] = raw_name
        data['total_amount'] += total_excl_tax
        data['total_tax'] += total_tax
        data['total_incl_tax'] += total_incl_tax
        paid_total = Decimal(str(getattr(expense, 'paid_total', None) or 0)).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP,
        )
        data['total_paid'] += paid_total
        data['remaining_total'] += expense.remaining_balance

    credit_queryset = SupplierCredit.objects.filter(user=request.user).select_related('supplier').prefetch_related('items')
    if query:
        credit_queryset = credit_queryset.filter(
            Q(supplier_name__icontains=query) | Q(supplier__name__icontains=query)
        )
    if start_date:
        credit_queryset = credit_queryset.filter(date__gte=start_date)
    if end_date:
        credit_queryset = credit_queryset.filter(date__lte=end_date)

    for credit in credit_queryset:
        total_incl_tax, total_tax, total_excl_tax = credit.calculate_totals()
        total_incl_tax = Decimal(str(total_incl_tax)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_tax = Decimal(str(total_tax)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_excl_tax = Decimal(str(total_excl_tax)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        raw_name = (credit.supplier_name or '').strip()
        cache_key = raw_name.lower()
        supplier = credit.supplier
        if not supplier and raw_name:
            supplier = supplier_cache.get(cache_key)
            if supplier is None:
                supplier = _get_or_create_supplier_for_user(request.user, raw_name)
                supplier_cache[cache_key] = supplier

        display_name = supplier.name if supplier else (raw_name or 'Unassigned Supplier')
        totals_key = (supplier.id if supplier else None, cache_key or '__unassigned__')

        data = supplier_totals.setdefault(
            totals_key,
            {
                'supplier': supplier,
                'supplier_name': display_name,
                'lookup_value': raw_name,
                'total_amount': Decimal('0.00'),
                'total_tax': Decimal('0.00'),
                'total_incl_tax': Decimal('0.00'),
                'total_paid': Decimal('0.00'),
                'credit_total_excl_tax': Decimal('0.00'),
                'credit_total_tax': Decimal('0.00'),
                'credit_total_incl_tax': Decimal('0.00'),
                'credit_available_incl_tax': Decimal('0.00'),
                'remaining_total': Decimal('0.00'),
            },
        )
        data['lookup_value'] = raw_name
        data['credit_total_excl_tax'] += total_excl_tax
        data['credit_total_tax'] += total_tax
        data['credit_total_incl_tax'] += total_incl_tax
        available_amount = Decimal(str(credit.available_amount or 0)).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP,
        )
        data['credit_available_incl_tax'] += available_amount

    suppliers = sorted(
        supplier_totals.values(),
        key=lambda record: record['total_incl_tax'],
        reverse=True,
    )

    for index, record in enumerate(suppliers, start=1):
        record['rank'] = index
        record['balance_incl_tax'] = (
            record['remaining_total'] - record['credit_available_incl_tax']
        ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    overall_total_incl_tax = sum(
        (record['balance_incl_tax'] for record in suppliers),
        Decimal('0.00'),
    )

    chart_labels = [record['supplier_name'] for record in suppliers]
    chart_data = [float(record['balance_incl_tax']) for record in suppliers]
    bank_accounts = BusinessBankAccount.objects.filter(
        user=request.user,
        is_active=True,
    ).order_by('name', 'id')

    context = {
        'suppliers': suppliers,
        'query': query,
        'overall_total': overall_total_incl_tax,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'supplier_form': SupplierForm(user=request.user),
        'bank_accounts': bank_accounts,
        'date_range_options': _get_invoice_date_range_options(request.user),
        'current_date_range': date_range_key,
        'selected_start_date': start_date,
        'selected_end_date': end_date,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('app/supplier_list_content.html', context, request=request)
        return JsonResponse({
            'html': html,
            'chart_labels': context['chart_labels'],
            'chart_data': context['chart_data'],
        })

    return render(request, 'app/supplier_list.html', context)



def vendor_detail(request, vendor_name=None, supplier_name=None):
    lookup_name = (supplier_name or vendor_name or "").strip()

    supplier = _get_or_create_supplier_for_user(request.user, lookup_name)
    supplier_name = supplier.name if supplier else lookup_name
    if not (supplier_name or '').strip():
        supplier_name = 'Unassigned Supplier'

    query = (request.GET.get('q') or '').strip()
    active_view = (request.GET.get('view') or 'receipts').strip().lower()
    if active_view not in {'receipts', 'credits', 'all'}:
        active_view = 'receipts'
    status_filter = (request.GET.get('status') or 'all').strip().lower()
    if status_filter not in {'all', 'paid', 'partial', 'unpaid', 'open'}:
        status_filter = 'all'

    date_range_key, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )

    base_expense_qs = MechExpense.objects.filter(user=request.user, vendor=lookup_name)
    if query:
        base_expense_qs = base_expense_qs.filter(
            Q(mechexpenseitem__part_no__icontains=query) |
            Q(mechexpenseitem__description__icontains=query) |
            Q(mechexpenseitem__qty__icontains=query) |
            Q(mechexpenseitem__price__icontains=query) |
            Q(mechexpenseitem__amount__icontains=query) |
            Q(mechexpenseitem__tax_paid__icontains=query)
        ).distinct()
    if start_date:
        base_expense_qs = base_expense_qs.filter(date__gte=start_date)
    if end_date:
        base_expense_qs = base_expense_qs.filter(date__lte=end_date)

    expenses = base_expense_qs.prefetch_related(
        'mechexpenseitem_set',
        'payments',
        'supplier_credit_items__supplier_credit',
    )

    total_amount = sum(expense.calculate_totals()[0] for expense in expenses)
    total_tax = sum(expense.calculate_totals()[1] for expense in expenses)
    total_paid = sum(expense.total_paid_amount for expense in expenses)
    total_remaining = sum(expense.remaining_balance for expense in expenses)
    total_amount = Decimal(str(total_amount or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_tax = Decimal(str(total_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_paid = Decimal(str(total_paid or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_remaining = Decimal(str(total_remaining or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    credit_queryset = SupplierCredit.objects.filter(user=request.user).select_related('supplier').prefetch_related('items')
    if supplier:
        credit_queryset = credit_queryset.filter(
            Q(supplier=supplier) | Q(supplier_name__iexact=lookup_name)
        )
    else:
        credit_queryset = credit_queryset.filter(supplier_name__iexact=lookup_name)
    if query:
        credit_queryset = credit_queryset.filter(
            Q(credit_no__icontains=query) |
            Q(memo__icontains=query) |
            Q(items__description__icontains=query) |
            Q(items__part_no__icontains=query)
        ).distinct()
    if start_date:
        credit_queryset = credit_queryset.filter(date__gte=start_date)
    if end_date:
        credit_queryset = credit_queryset.filter(date__lte=end_date)

    credit_entries = []
    credit_total_incl_tax = Decimal('0.00')
    credit_total_tax = Decimal('0.00')
    credit_available_total = Decimal('0.00')
    for credit in credit_queryset:
        credit_total, credit_tax, credit_excl = credit.calculate_totals()
        credit_total = Decimal(str(credit_total or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        credit_tax = Decimal(str(credit_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        credit_excl = Decimal(str(credit_excl or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        available_amount = Decimal(str(credit.available_amount or 0)).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP,
        )
        credit_entries.append({
            'credit': credit,
            'total_incl_tax': credit_total,
            'total_tax': credit_tax,
            'total_excl_tax': credit_excl,
        })
        credit_total_incl_tax += credit_total
        credit_total_tax += credit_tax
        credit_available_total += available_amount

    balance_incl_tax = (total_remaining - credit_available_total).quantize(
        Decimal('0.01'),
        rounding=ROUND_HALF_UP,
    )

    receipt_rows = []
    for expense in expenses:
        status_value = expense.payment_status
        if status_filter == 'paid' and status_value != 'paid':
            continue
        if status_filter == 'partial' and status_value != 'partial':
            continue
        if status_filter == 'unpaid' and status_value != 'unpaid':
            continue
        if status_filter == 'open' and status_value not in {'unpaid', 'partial'}:
            continue
        total_incl_tax, _, _ = expense.calculate_totals()
        receipt_rows.append({
            'expense': expense,
            'total_incl_tax': Decimal(str(total_incl_tax or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
            'status': status_value,
            'remaining': expense.remaining_balance,
        })

    base_params = request.GET.copy()
    base_params.pop('view', None)
    receipts_params = base_params.copy()
    receipts_params['view'] = 'receipts'
    credits_params = base_params.copy()
    credits_params['view'] = 'credits'
    all_params = base_params.copy()
    all_params['view'] = 'all'
    portal_enabled = bool(
        supplier
        and getattr(supplier, 'portal_user', None)
        and supplier.portal_user.is_active
    )

    combined_entries = []
    for row in receipt_rows:
        combined_entries.append({
            'type': 'receipt',
            'date': row['expense'].date,
            'ref': row['expense'].receipt_no,
            'category': row['expense'].categorie or 'N/A',
            'amount': row['total_incl_tax'],
            'balance': row['expense'].remaining_balance,
            'status': row['status'],
            'url': reverse('accounts:mechexpense_detail', args=[row['expense'].pk]),
        })
    for entry in credit_entries:
        combined_entries.append({
            'type': 'credit',
            'date': entry['credit'].date,
            'ref': entry['credit'].credit_no,
            'category': entry['credit'].memo or 'Supplier credit',
            'amount': entry['total_incl_tax'] * Decimal('-1'),
            'balance': entry['credit'].available_amount,
            'status': 'credit',
            'url': reverse('accounts:supplier_credit_detail', args=[entry['credit'].pk]),
        })
    combined_entries.sort(
        key=lambda item: (
            item['date'] or datetime.date.min,
            item['ref'] or '',
        ),
        reverse=True,
    )

    context = {
        'expenses': expenses,
        'receipt_rows': receipt_rows,
        'supplier': supplier,
        'supplier_name': supplier_name,
        'total_amount': total_amount,
        'total_tax': total_tax,
        'total_paid': total_paid,
        'credit_entries': credit_entries,
        'credit_total': credit_total_incl_tax,
        'credit_total_tax': credit_total_tax,
        'balance_total': balance_incl_tax,
        'query': query,
        'active_view': active_view,
        'status_filter': status_filter,
        'combined_entries': combined_entries,
        'date_range_options': _get_invoice_date_range_options(request.user),
        'current_date_range': date_range_key,
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'receipts_query': receipts_params.urlencode(),
        'credits_query': credits_params.urlencode(),
        'all_query': all_params.urlencode(),
        'portal_enabled': portal_enabled,
        'portal_username': supplier.portal_user.username if portal_enabled else '',
    }

    return render(request, 'app/supplier_detail.html', context)


@login_required
def vendor_entries(request, vendor_name=None, supplier_name=None):
    lookup_name = (supplier_name or vendor_name or "").strip()
    supplier = _get_or_create_supplier_for_user(request.user, lookup_name)
    supplier_display = supplier.name if supplier else lookup_name
    if not (supplier_display or '').strip():
        supplier_display = 'Unassigned Supplier'

    query = (request.GET.get('q') or '').strip()
    selected_category = (request.GET.get('category') or '').strip()
    date_range_key, start_date, end_date = _get_date_range_bounds(
        request,
        default="1y",
        cookie_key=DATE_RANGE_COOKIE_NAME,
    )

    sort = (request.GET.get('sort') or 'date').strip().lower()
    direction = (request.GET.get('direction') or 'desc').strip().lower()
    sort_map = {
        'receipt': 'mech_expense__receipt_no',
        'date': 'mech_expense__date',
        'category': 'mech_expense__categorie',
        'part_no': 'part_no',
        'description': 'description',
        'qty': 'qty',
        'price': 'price',
        'amount': 'amount',
        'tax': 'tax_paid',
    }
    if sort not in sort_map:
        sort = 'date'
    if direction not in {'asc', 'desc'}:
        direction = 'desc'
    order_field = sort_map[sort]
    ordering = order_field if direction == 'asc' else f'-{order_field}'

    vendor_filter = Q(mech_expense__vendor__iexact=lookup_name)
    if supplier and supplier.name and supplier.name.strip().lower() != lookup_name.lower():
        vendor_filter |= Q(mech_expense__vendor__iexact=supplier.name)

    entries_qs = (
        MechExpenseItem.objects.select_related('mech_expense')
        .filter(mech_expense__user=request.user)
        .filter(vendor_filter)
    )
    if start_date:
        entries_qs = entries_qs.filter(mech_expense__date__gte=start_date)
    if end_date:
        entries_qs = entries_qs.filter(mech_expense__date__lte=end_date)
    if selected_category:
        entries_qs = entries_qs.filter(mech_expense__categorie__iexact=selected_category)
    if query:
        search_filter = (
            Q(part_no__icontains=query)
            | Q(description__icontains=query)
            | Q(mech_expense__receipt_no__icontains=query)
            | Q(mech_expense__categorie__icontains=query)
        )
        numeric_value = None
        try:
            numeric_value = Decimal(str(query))
        except (InvalidOperation, TypeError, ValueError):
            numeric_value = None
        if numeric_value is not None:
            search_filter |= Q(qty=numeric_value) | Q(price=numeric_value) | Q(amount=numeric_value) | Q(tax_paid=numeric_value)
        try:
            parsed_date = datetime.datetime.strptime(query, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            parsed_date = None
        if parsed_date:
            search_filter |= Q(mech_expense__date=parsed_date)
        entries_qs = entries_qs.filter(search_filter)

    entries = entries_qs.order_by(ordering, '-id')

    category_options = (
        MechExpense.objects.filter(user=request.user)
        .exclude(categorie__isnull=True)
        .exclude(categorie__exact='')
        .values_list('categorie', flat=True)
        .distinct()
        .order_by('categorie')
    )

    params = request.GET.copy()
    params.pop('sort', None)
    params.pop('direction', None)
    query_string = params.urlencode()
    query_prefix = f"{query_string}&" if query_string else ""

    context = {
        'supplier': supplier,
        'supplier_name': supplier_display,
        'entries': entries,
        'query': query,
        'selected_category': selected_category,
        'category_options': category_options,
        'date_range_options': _get_invoice_date_range_options(request.user),
        'current_date_range': date_range_key,
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'sort': sort,
        'direction': direction,
        'query_prefix': query_prefix,
    }

    return render(request, 'app/supplier_entries.html', context)


@login_required
@require_POST
def vendor_delete(request, vendor_name):
    decoded_vendor = unquote(vendor_name)
    expenses_qs = MechExpense.objects.filter(user=request.user, vendor=decoded_vendor)
    deleted_count = expenses_qs.count()

    if deleted_count == 0:
        messages.warning(request, f'No expenses found for supplier "{decoded_vendor or "Unassigned"}" to delete.')
    else:
        expenses_qs.delete()
        messages.success(request, f'Deleted {deleted_count} expense(s) for supplier "{decoded_vendor or "Unassigned"}".')

    return redirect('accounts:supplier_list')


@login_required
def category_list(request):
    """List and create user-defined categories."""
    categories = Category.objects.filter(user=request.user)
    if request.method == 'POST':
        form = CategoryForm(request.POST, user=request.user)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            messages.success(request, 'Category added successfully.')
            return redirect('accounts:category_list')
    else:
        form = CategoryForm(user=request.user)
    return render(request, 'app/category_list.html', {'categories': categories, 'form': form})


@login_required
def delete_user_category(request, pk):
    """Delete a user-owned category."""
    category = get_object_or_404(Category, pk=pk, user=request.user)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Category deleted successfully.')
    return redirect('accounts:category_list')


def _serialize_fixed_hours(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        normalized = value.normalize()
        text = format(normalized, 'f')
    else:
        try:
            text = format(Decimal(str(value)).normalize(), 'f')
        except (InvalidOperation, ValueError, TypeError):
            text = str(value)
    if '.' in text:
        text = text.rstrip('0').rstrip('.')
    return text or '0'


def _clean_fixed_hours_input(value):
    if value in (None, '', [], {}):
        return None
    try:
        hours = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        raise ValueError('Fixed hours must be a number.')
    if hours < 0:
        raise ValueError('Fixed hours cannot be negative.')
    return hours.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def _serialize_fixed_rate(value):
    if value is None:
        return None
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        amount = None
    if amount is None:
        text = str(value)
    else:
        normalized = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        text = format(normalized, 'f')
    if '.' in text:
        text = text.rstrip('0').rstrip('.')
    return text or '0'

def _clean_fixed_rate_input(value):
    if value in (None, '', [], {}):
        return None
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        raise ValueError('Fixed rate must be a number.')
    if amount < 0:
        raise ValueError('Fixed rate cannot be negative.')
    return amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def _clean_positive_integer_input(value, field_label):
    if value in (None, '', [], {}):
        return None
    try:
        numeric_value = int(str(value).strip())
    except (TypeError, ValueError, AttributeError):
        raise ValueError(f"{field_label} must be a whole number.")
    if numeric_value < 0:
        raise ValueError(f"{field_label} cannot be negative.")
    return numeric_value


SERVICE_TEMPLATE_HEADERS = [
    "Job name",
    "Job description",
    "More about this job",
    "Fixed hours",
    "Fixed rate",
    "Due after (km)",
    "Due after (months)",
]

OLD_SERVICE_TEMPLATE_HEADERS = [
    "Services with descriptions",
    "More about this service",
]


def build_service_job_catalog(user):
    """Return structured job data for the given user."""

    job_names = (
        ServiceJobName.objects.filter(user=user, is_active=True)
        .prefetch_related('services')
        .order_by(Lower('name'))
    )

    catalog = []
    description_strings = []
    job_name_choices = []

    for job in job_names:
        services = [
            {
                'id': service.id,
                'text': service.name,
                'notes': service.description or '',
                'fixed_hours': _serialize_fixed_hours(service.fixed_hours),
                'fixed_rate': _serialize_fixed_rate(service.fixed_rate),
                'due_after_kilometers': service.due_after_kilometers,
                'due_after_months': service.due_after_months,
            }
            for service in job.services.filter(is_active=True).order_by(Lower('name'))
        ]

        if not services:
            continue

        job_name_choices.append(job.name)
        catalog.append(
            {
                'id': job.id,
                'name': job.name,
                'descriptions': services,
            }
        )
        description_strings.extend(item['text'] for item in services)

    return catalog, description_strings, job_name_choices


def _resolve_invoice_vehicle_from_post(post_data, user):
    """Resolve the selected vehicle ID from the add invoice form."""

    if not user:
        return None

    raw_vehicle_id = (post_data.get('vehicle') or '').strip()
    if not raw_vehicle_id:
        return None
    try:
        vehicle_id = int(raw_vehicle_id)
    except (TypeError, ValueError):
        return None
    return Vehicle.objects.filter(pk=vehicle_id, customer__user=user).first()


@login_required
@require_POST
def create_service_description(request):
    """Persist a new job description under the selected job name for the current user."""

    payload = {}
    if request.body:
        try:
            payload = json.loads(request.body.decode('utf-8'))
            if not isinstance(payload, dict):
                payload = {}
        except (json.JSONDecodeError, UnicodeDecodeError):
            payload = {}

    if not payload:
        payload = request.POST

    raw_job_name_id = payload.get('job_name_id')
    raw_job_name_label = payload.get('job_name')
    raw_description = payload.get('description')
    raw_notes = payload.get('notes')
    raw_fixed_hours = payload.get('fixed_hours')
    raw_fixed_rate = payload.get('fixed_rate')
    raw_due_after_kilometers = payload.get('due_after_kilometers')
    raw_due_after_months = payload.get('due_after_months')
    raw_service_id = payload.get('service_id')

    payload_keys = set(payload.keys()) if hasattr(payload, 'keys') else set()
    notes_provided = 'notes' in payload_keys
    fixed_hours_provided = 'fixed_hours' in payload_keys
    fixed_rate_provided = 'fixed_rate' in payload_keys
    due_kilometers_provided = 'due_after_kilometers' in payload_keys
    due_months_provided = 'due_after_months' in payload_keys

    job_name_label = Service._normalize_job_name(raw_job_name_label)
    normalized_description = Service._normalize_text(raw_description)
    notes = (raw_notes or '').strip()

    service_to_update = None
    service_id = None
    if raw_service_id not in (None, ''):
        try:
            service_id = int(raw_service_id)
        except (TypeError, ValueError):
            service_id = None
    if service_id:
        service_to_update = Service.objects.filter(user=request.user, pk=service_id).first()
        if not service_to_update:
            return JsonResponse(
                {'success': False, 'error': 'The selected service could not be found.'},
                status=404,
            )

    try:
        fixed_hours = _clean_fixed_hours_input(raw_fixed_hours)
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    try:
        fixed_rate = _clean_fixed_rate_input(raw_fixed_rate)
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    try:
        due_after_kilometers = _clean_positive_integer_input(raw_due_after_kilometers, 'Due after (km)')
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    try:
        due_after_months = _clean_positive_integer_input(raw_due_after_months, 'Due after (months)')
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    if not normalized_description:
        return JsonResponse({'success': False, 'error': 'Please provide a job description.'}, status=400)

    job_group = None
    job_name_id = None

    if raw_job_name_id not in (None, ''):
        try:
            job_name_id = int(raw_job_name_id)
        except (TypeError, ValueError):
            job_name_id = None

    if job_name_id:
        job_group = ServiceJobName.objects.filter(user=request.user, pk=job_name_id).first()

    if not job_group:
        if not job_name_label:
            if service_to_update:
                job_group = service_to_update.job_name
            else:
                return JsonResponse({'success': False, 'error': 'Please select a job name.'}, status=400)

        if not job_group:
            job_group = ServiceJobName.objects.filter(user=request.user, name__iexact=job_name_label).first()
            if not job_group:
                job_group = ServiceJobName(user=request.user, name=job_name_label)
                try:
                    job_group.save()
                except IntegrityError:
                    job_group = ServiceJobName.objects.filter(user=request.user, name__iexact=job_name_label).first()
                if not job_group:
                    return JsonResponse({'success': False, 'error': 'We could not save that job name.'}, status=400)

    if service_to_update:
        duplicate_exists = Service.objects.filter(
            user=request.user,
            job_name=job_group,
            name__iexact=normalized_description,
        ).exclude(pk=service_to_update.pk).exists()
        if duplicate_exists:
            return JsonResponse(
                {'success': False, 'error': 'That job description already exists for this job name.'},
                status=409,
            )

        updates = []
        if service_to_update.job_name_id != job_group.id:
            service_to_update.job_name = job_group
            updates.append('job_name')
        if service_to_update.name != normalized_description:
            service_to_update.name = normalized_description
            updates.append('name')
        if notes_provided and (service_to_update.description or '') != notes:
            service_to_update.description = notes
            updates.append('description')
        if fixed_hours_provided and service_to_update.fixed_hours != fixed_hours:
            service_to_update.fixed_hours = fixed_hours
            updates.append('fixed_hours')
        if fixed_rate_provided and service_to_update.fixed_rate != fixed_rate:
            service_to_update.fixed_rate = fixed_rate
            updates.append('fixed_rate')
        if due_kilometers_provided and service_to_update.due_after_kilometers != due_after_kilometers:
            service_to_update.due_after_kilometers = due_after_kilometers
            updates.append('due_after_kilometers')
        if due_months_provided and service_to_update.due_after_months != due_after_months:
            service_to_update.due_after_months = due_after_months
            updates.append('due_after_months')
        if not service_to_update.is_active:
            service_to_update.is_active = True
            updates.append('is_active')

        if updates:
            try:
                service_to_update.save(update_fields=updates)
            except IntegrityError:
                return JsonResponse(
                    {'success': False, 'error': 'That job description already exists for this job name.'},
                    status=409,
                )
        service = service_to_update
        created = False
    else:
        existing = Service.objects.filter(
            user=request.user,
            job_name=job_group,
            name__iexact=normalized_description,
        ).first()

        service = Service.record_service(
            user=request.user,
            name=normalized_description,
            job_name=job_group,
            description=notes,
            fixed_hours=fixed_hours,
            fixed_rate=fixed_rate,
            due_after_kilometers=due_after_kilometers,
            due_after_months=due_after_months,
        )

        if not service:
            return JsonResponse({'success': False, 'error': 'We could not save that job description.'}, status=400)

        created = existing is None or existing.pk != service.pk

    response_data = {
        'success': True,
        'created': created,
        'service': {
            'id': service.pk,
            'job_name': {
                'id': job_group.pk,
                'name': job_group.name,
            },
            'description': service.name,
            'notes': service.description or '',
            'fixed_hours': _serialize_fixed_hours(service.fixed_hours),
            'fixed_rate': _serialize_fixed_rate(service.fixed_rate),
            'due_after_kilometers': service.due_after_kilometers,
            'due_after_months': service.due_after_months,
        },
    }

    return JsonResponse(response_data)


@login_required
def service_list(request):
    """Manage reusable service descriptions for invoice and work order suggestions."""
    date_range_key, start_date, end_date = _get_date_range_bounds(request, default="1y")

    services_qs = (
        Service.objects.filter(user=request.user)
        .select_related('job_name')
        .order_by(Lower('job_name__name'), Lower('name'))
    )
    services = list(services_qs)

    grouped_services = []
    for (job_name_id, job_name_label), service_group in itertools.groupby(
        services, key=lambda svc: (svc.job_name_id, svc.job_name.name)
    ):
        group_services = list(service_group)
        grouped_services.append(
            {
                'job_name_id': job_name_id,
                'job_name': job_name_label,
                'services': group_services,
            }
        )

    _job_catalog, _catalog_descriptions, job_name_choices = build_service_job_catalog(
        request.user
    )

    return render(
        request,
        'app/service_list.html',
        {
            'grouped_services': grouped_services,
            'job_name_choices': job_name_choices,
            'services_count': len(services),
            'date_range_options': DATE_RANGE_OPTIONS,
            'current_date_range': date_range_key,
            'selected_start_date': start_date,
            'selected_end_date': end_date,
        },
    )


@login_required
@require_POST
def save_service_inline(request):
    """Create or update a reusable service description via inline editing."""

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (TypeError, ValueError, json.JSONDecodeError):
        return JsonResponse({'error': 'Invalid data submitted.'}, status=400)

    form_data = {
        'job_name': payload.get('job_name', ''),
        'name': payload.get('name', ''),
        'description': payload.get('description', ''),
        'fixed_hours': payload.get('fixed_hours', ''),
        'fixed_rate': payload.get('fixed_rate', ''),
        'due_after_kilometers': payload.get('due_after_kilometers', ''),
        'due_after_months': payload.get('due_after_months', ''),
        'show_on_customer_portal': 'on' if payload.get('show_on_customer_portal') else '',
    }

    service_id_raw = payload.get('id')
    service_instance = None
    if service_id_raw:
        try:
            service_pk = int(service_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Invalid service identifier provided.'}, status=400)
        service_instance = get_object_or_404(Service, pk=service_pk, user=request.user)
        form = ServiceForm(form_data, instance=service_instance, user=request.user)
    else:
        form = ServiceForm(form_data, user=request.user)

    if not form.is_valid():
        return JsonResponse({'errors': form.errors}, status=400)

    service = form.save(commit=False)
    service.user = request.user
    service.is_active = True
    service.save()

    response_data = {
        'service': {
            'id': service.pk,
            'name': service.name,
            'description': service.description,
            'job_name': {
                'id': service.job_name_id,
                'name': service.job_name.name,
            },
            'fixed_hours': _serialize_fixed_hours(service.fixed_hours),
            'fixed_rate': _serialize_fixed_rate(service.fixed_rate),
            'due_after_kilometers': service.due_after_kilometers,
            'due_after_months': service.due_after_months,
            'show_on_customer_portal': service.show_on_customer_portal,
        }
    }

    status_code = 200 if service_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def rename_service_job_name(request, pk):
    """Rename a job name grouping, merging into an existing one if needed."""

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (TypeError, ValueError, json.JSONDecodeError):
        return JsonResponse({'error': 'Invalid data submitted.'}, status=400)

    new_name_raw = (payload.get('name') or '').strip()
    if not new_name_raw:
        return JsonResponse({'error': 'Please provide a job name.'}, status=400)

    normalized_name = re.sub(r'\s+', ' ', new_name_raw).strip()
    if not normalized_name:
        return JsonResponse({'error': 'Please provide a job name.'}, status=400)

    job_group = get_object_or_404(ServiceJobName, pk=pk, user=request.user)

    if normalized_name.lower() == job_group.name.lower():
        return JsonResponse(
            {
                'job_name': {
                    'id': job_group.pk,
                    'name': job_group.name,
                },
                'merged': False,
            }
        )

    existing_group = (
        ServiceJobName.objects.filter(user=request.user, name__iexact=normalized_name)
        .exclude(pk=job_group.pk)
        .first()
    )

    merged = False
    if existing_group:
        Service.objects.filter(user=request.user, job_name=job_group).update(
            job_name=existing_group
        )
        job_group.delete()
        job_group = existing_group
        merged = True
    else:
        job_group.name = normalized_name
        job_group.save(update_fields=['name'])

    return JsonResponse(
        {
            'job_name': {
                'id': job_group.pk,
                'name': job_group.name,
            },
            'merged': merged,
        }
    )


@login_required
@require_POST
def delete_service(request, pk):
    service = get_object_or_404(Service, pk=pk, user=request.user)
    service.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in (
        request.headers.get('Accept', '') or ''
    ):
        return JsonResponse({'deleted': True})

    messages.success(request, 'Service removed successfully.')
    return redirect('accounts:service_list')


@login_required
@require_POST
def bulk_delete_services(request):
    service_ids = request.POST.getlist('service_ids')

    if not service_ids:
        messages.info(request, 'Select at least one service to delete.')
        return redirect('accounts:service_list')

    services_qs = Service.objects.filter(user=request.user, pk__in=service_ids)
    deleted_count = services_qs.count()

    if not deleted_count:
        messages.warning(request, 'No matching services were found to delete.')
        return redirect('accounts:service_list')

    services_qs.delete()
    messages.success(
        request,
        f"Deleted {deleted_count} service{'s' if deleted_count != 1 else ''}.",
    )
    return redirect('accounts:service_list')


@login_required
def export_services_template(request):
    """Export the user's services to an Excel template."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Services"
    worksheet.append(SERVICE_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    services = (
        Service.objects.filter(user=request.user)
        .select_related('job_name')
        .order_by(Lower('job_name__name'), Lower('name'))
    )
    for service in services:
        worksheet.append(
            [
                service.job_name.name if service.job_name_id else "",
                service.name or "",
                service.description or "",
                float(service.fixed_hours) if service.fixed_hours is not None else "",
                float(service.fixed_rate) if service.fixed_rate is not None else "",
                service.due_after_kilometers or "",
                service.due_after_months or "",
            ]
        )

        apply_template_styling(
            worksheet,
            headers=SERVICE_TEMPLATE_HEADERS,
            header_row_index=1,
            column_width_overrides={
            "Job name": 32,
            "Job description": 48,
            "More about this job": 60,
            "Fixed hours": 14,
            "Fixed rate": 14,
            "Due after (km)": 18,
            "Due after (months)": 20,
            4: 14,
            5: 14,
            6: 18,
            7: 20,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="services_template.xlsx"'
    return response


@login_required
@require_POST
def import_services_from_excel(request):
    """Import or update services from an uploaded Excel file."""

    upload = request.FILES.get('file')
    if not upload:
        messages.error(request, 'Please choose an Excel file to import.')
        return redirect('accounts:service_list')

    try:
        workbook = load_workbook(upload, data_only=True)
    except InvalidFileException:
        messages.error(request, 'The uploaded file could not be read. Please upload a valid .xlsx file.')
        return redirect('accounts:service_list')
    except Exception:
        messages.error(request, 'We were unable to process that file. Please try again with a new export.')
        return redirect('accounts:service_list')

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        messages.error(request, 'The uploaded file is empty.')
        return redirect('accounts:service_list')

    def _normalize_header(value):
        if value is None:
            return ''
        return str(value).strip().lower()

    header = [_normalize_header(cell) for cell in rows[0][: len(SERVICE_TEMPLATE_HEADERS)]]
    expected_new = [column.lower() for column in SERVICE_TEMPLATE_HEADERS]
    expected_old = [column.lower() for column in OLD_SERVICE_TEMPLATE_HEADERS]
    if header == expected_new:
        using_new_template = True
    elif header[: len(expected_old)] == expected_old and len(header) == len(expected_old):
        using_new_template = False
    else:
        messages.error(request, 'The header row does not match the provided service template.')
        return redirect('accounts:service_list')

    created_count = 0
    updated_count = 0
    missing_name_rows = []
    invalid_fixed_hour_rows = []
    invalid_fixed_rate_rows = []
    invalid_due_kilometer_rows = []
    invalid_due_month_rows = []

    for index, row in enumerate(rows[1:], start=2):
        if not row or all(cell in (None, '') for cell in row):
            continue

        if using_new_template:
            job_name_cell = row[0] if len(row) > 0 else None
            name_cell = row[1] if len(row) > 1 else None
            description_cell = row[2] if len(row) > 2 else None
            fixed_hours_cell = row[3] if len(row) > 3 else None
            fixed_rate_cell = row[4] if len(row) > 4 else None
            due_after_kilometers_cell = row[5] if len(row) > 5 else None
            due_after_months_cell = row[6] if len(row) > 6 else None
        else:
            job_name_cell = row[0] if len(row) > 0 else None
            name_cell = row[0] if len(row) > 0 else None
            description_cell = row[1] if len(row) > 1 else None
            fixed_hours_cell = None
            fixed_rate_cell = None
            due_after_kilometers_cell = None
            due_after_months_cell = None

        name = ''
        if name_cell is not None:
            name = str(name_cell).strip()

        if not name:
            missing_name_rows.append(index)
            continue

        job_name_value = ''
        if job_name_cell is not None:
            job_name_value = str(job_name_cell).strip()
        if not job_name_value:
            job_name_value = name or 'General'

        description = ''
        if description_cell is not None:
            description = str(description_cell).strip()

        fixed_hours = None
        fixed_hours_provided = using_new_template
        fixed_rate = None
        fixed_rate_provided = using_new_template
        if using_new_template:
            if fixed_hours_cell in (None, ''):
                fixed_hours = None
            else:
                try:
                    fixed_hours = _clean_fixed_hours_input(fixed_hours_cell)
                except ValueError:
                    invalid_fixed_hour_rows.append(index)
                    fixed_hours_provided = False
                    fixed_hours = None
            if fixed_rate_cell in (None, ''):
                fixed_rate = None
            else:
                try:
                    fixed_rate = _clean_fixed_rate_input(fixed_rate_cell)
                except ValueError:
                    invalid_fixed_rate_rows.append(index)
                    fixed_rate_provided = False
                    fixed_rate = None
        else:
            fixed_hours_provided = False
            fixed_rate_provided = False

        due_after_kilometers = None
        due_after_months = None
        if using_new_template:
            if due_after_kilometers_cell not in (None, ''):
                try:
                    due_after_kilometers = _clean_positive_integer_input(
                        due_after_kilometers_cell, 'Due after (km)'
                    )
                except ValueError:
                    invalid_due_kilometer_rows.append(index)
                    due_after_kilometers = None
            if due_after_months_cell not in (None, ''):
                try:
                    due_after_months = _clean_positive_integer_input(
                        due_after_months_cell, 'Due after (months)'
                    )
                except ValueError:
                    invalid_due_month_rows.append(index)
                    due_after_months = None

        job_group = ServiceJobName.objects.filter(user=request.user, name__iexact=job_name_value).first()
        if not job_group:
            job_group = ServiceJobName.objects.create(user=request.user, name=job_name_value)

        service = Service.objects.filter(user=request.user, job_name=job_group, name__iexact=name).first()
        if service:
            updated = False
            if description != (service.description or ''):
                service.description = description
                updated = True
            if fixed_hours_provided and service.fixed_hours != fixed_hours:
                service.fixed_hours = fixed_hours
                updated = True
            if fixed_rate_provided and service.fixed_rate != fixed_rate:
                service.fixed_rate = fixed_rate
                updated = True
            if due_after_kilometers is not None and service.due_after_kilometers != due_after_kilometers:
                service.due_after_kilometers = due_after_kilometers
                updated = True
            if due_after_months is not None and service.due_after_months != due_after_months:
                service.due_after_months = due_after_months
                updated = True
            if not service.is_active:
                service.is_active = True
                updated = True
            if updated:
                service.save()
                updated_count += 1
        else:
            Service.objects.create(
                user=request.user,
                job_name=job_group,
                name=name,
                description=description,
                is_active=True,
                fixed_hours=fixed_hours if fixed_hours_provided else None,
                fixed_rate=fixed_rate if fixed_rate_provided else None,
                due_after_kilometers=due_after_kilometers,
                due_after_months=due_after_months,
            )
            created_count += 1

    summary_parts = []
    if created_count:
        summary_parts.append(f"{created_count} new service{'s' if created_count != 1 else ''}")
    if updated_count:
        summary_parts.append(f"{updated_count} existing service{'s' if updated_count != 1 else ''}")

    if summary_parts:
        if len(summary_parts) == 2:
            message_body = ' and '.join(summary_parts)
        else:
            message_body = summary_parts[0]
        messages.success(request, f"Import complete: {message_body}.")
    else:
        messages.info(request, 'The file was processed, but no changes were made.')

    if missing_name_rows:
        row_list = ', '.join(str(row_number) for row_number in missing_name_rows)
        messages.warning(request, f"Skipped rows {row_list} because no service name was provided.")

    if invalid_fixed_hour_rows:
        row_list = ', '.join(str(row_number) for row_number in invalid_fixed_hour_rows)
        messages.warning(request, f"Ignored fixed hours on rows {row_list} because the values were not valid numbers.")
    if invalid_fixed_rate_rows:
        row_list = ', '.join(str(row_number) for row_number in invalid_fixed_rate_rows)
        messages.warning(request, f"Ignored fixed rates on rows {row_list} because the values were not valid numbers.")

    if invalid_due_kilometer_rows:
        row_list = ', '.join(str(row_number) for row_number in invalid_due_kilometer_rows)
        messages.warning(request, f"Ignored kilometer intervals on rows {row_list} because the values were not valid numbers.")

    if invalid_due_month_rows:
        row_list = ', '.join(str(row_number) for row_number in invalid_due_month_rows)
        messages.warning(request, f"Ignored month intervals on rows {row_list} because the values were not valid numbers.")

    return redirect('accounts:service_list')

class RedirectAfterLogoutMiddleware:
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        response = self.get_response(request)
        if request.session.pop('redirect_after_logout', False):
            return redirect('accounts:choose_plan')
        return response

class GroupedInvoiceDeleteView(View):
    """
    Deletes a GroupedInvoice and its IncomeRecord2 + PendingInvoice
    in bulk, checking that the current user owns it (or is superuser).
    """
    success_url = reverse_lazy("accounts:groupedinvoice_list")
    template_name = "app/groupedinvoice_confirm_delete.html"

    def get_object(self, pk):
        # Direct lookup, ignoring any queryset filters
        invoice = get_object_or_404(GroupedInvoice, pk=pk)
        # Permission check
        if invoice.user != self.request.user and not self.request.user.is_superuser:
            raise Http404
        return invoice

    def get(self, request, pk, *args, **kwargs):
        # Show a confirmation page if you want, else shortcut to delete
        invoice = self.get_object(pk)
        return render(request, self.template_name, {"object": invoice})

    def post(self, request, pk, *args, **kwargs):
        invoice = self.get_object(pk)
        inv_num = invoice.invoice_number

        with transaction.atomic():
            # Delete line items individually so inventory reversals occur
            for line in IncomeRecord2.objects.filter(grouped_invoice=invoice):
                line.delete()
            PendingInvoice.objects.filter(grouped_invoice=invoice).delete()
            invoice.delete()

        messages.success(request, f"Invoice {inv_num} and its associated records have been deleted.")
        return redirect(self.success_url)


class GroupedInvoiceListView(LoginRequiredMixin, ListView):
    model = GroupedInvoice
    template_name = 'app/groupedinvoice_list.html'
    context_object_name = 'object_list'
    paginate_by = 100

    def _get_sorting(self):
        return _resolve_sorting(
            self.request,
            default_sort='invoice_number',
            default_order='desc',
        )

    def get_queryset(self):
        if hasattr(self, '_groupedinvoice_queryset'):
            return self._groupedinvoice_queryset

        query = self.request.GET.get('search')
        status_filter = (self.request.GET.get('status') or '').strip().lower()
        sort_by, order = self._get_sorting()
        user = self.request.user
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        queryset = GroupedInvoice.objects.filter(user=user).prefetch_related('payments')
        if start_date:
            queryset = queryset.filter(date__isnull=False, date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__isnull=False, date__lte=end_date)

        # Annotate total_paid + balance_due for filtering/status without N+1.
        invoice_paid_subquery = (
            Payment.objects.filter(invoice=OuterRef('pk'))
            .values('invoice')
            .annotate(
                total=Coalesce(
                    Sum('amount'),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            )
            .values('total')[:1]
        )
        queryset = _annotate_invoice_credit_totals(
            queryset.annotate(
                total_paid=Coalesce(
                    Subquery(invoice_paid_subquery),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
            )
        ).annotate(
            balance_due=ExpressionWrapper(
                F('total_amount') - F('total_paid') - F('credit_total'),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
            total_settled=ExpressionWrapper(
                F('total_paid'),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
        )

        # Overdue threshold based on profile term (same rule as dashboard).
        profile = getattr(user, 'profile', None)
        term_days = TERM_CHOICES.get(getattr(profile, 'term', None), 30)
        today = timezone.localdate()
        queryset = queryset.annotate(
            status_label=Case(
                When(
                    total_paid__gte=F('total_amount'),
                    then=Value('Paid'),
                ),
                When(
                    total_paid__gt=Decimal('0.00'),
                    then=Value('Partial'),
                ),
                default=Value('Pending'),
                output_field=CharField(),
            )
        )

        activity_sent_subquery = (
            InvoiceActivity.objects.filter(
                invoice=OuterRef('pk'),
                event_type=InvoiceActivity.EVENT_EMAIL_SENT,
            )
            .order_by('-created_at')
            .values('created_at')[:1]
        )
        activity_opened_subquery = (
            InvoiceActivity.objects.filter(
                invoice=OuterRef('pk'),
                event_type=InvoiceActivity.EVENT_EMAIL_OPENED,
            )
            .order_by('-created_at')
            .values('created_at')[:1]
        )
        activity_viewed_subquery = (
            InvoiceActivity.objects.filter(
                invoice=OuterRef('pk'),
                event_type=InvoiceActivity.EVENT_VIEWED,
            )
            .order_by('-created_at')
            .values('created_at')[:1]
        )

        queryset = queryset.annotate(
            latest_email_sent_at=Subquery(
                activity_sent_subquery,
                output_field=DateTimeField(),
            ),
            latest_email_opened_at=Subquery(
                activity_opened_subquery,
                output_field=DateTimeField(),
            ),
            latest_portal_viewed_at=Subquery(
                activity_viewed_subquery,
                output_field=DateTimeField(),
            ),
        )

        # Apply status filter
        if status_filter == 'paid':
            # Paid should include fully paid and partially paid invoices.
            queryset = queryset.filter(total_settled__gt=Decimal('0.00'))
        elif status_filter == 'pending':
            # Pending should include unpaid and partial invoices (any balance due).
            queryset = queryset.filter(balance_due__gt=Decimal('0.00'))
        elif status_filter == 'partial':
            queryset = queryset.filter(
                total_settled__gt=Decimal('0.00'),
                balance_due__gt=Decimal('0.00'),
            )

        if query:
            filters = (
                Q(invoice_number__icontains=query) |
                Q(bill_to__icontains=query) |
                Q(date__icontains=query) |
                Q(total_amount__icontains=query) |
                Q(vin_no__icontains=query) | Q(unit_no__icontains=query) | Q(make_model__icontains=query)
            )
            queryset = queryset.filter(filters)

        sort_map = {
            'invoice_number': 'invoice_number',
            'date': 'date',
            'bill_to': 'bill_to',
            'vin_no': 'vin_no',
            'unit_no': 'unit_no',
            'status': 'status_label',
            'total_amount': 'total_amount',
            'balance_due': 'balance_due',
        }
        if sort_by == 'invoice_number':
            queryset = _annotate_invoice_number_sort(queryset, 'invoice_number')
            if order == 'desc':
                queryset = queryset.order_by('-invoice_number_sort', '-invoice_number')
            else:
                queryset = queryset.order_by('invoice_number_sort', 'invoice_number')
        else:
            sort_field = sort_map.get(sort_by)
            if sort_field:
                if order == 'desc':
                    queryset = queryset.order_by(f'-{sort_field}')
                else:
                    queryset = queryset.order_by(sort_field)
            else:
                queryset = queryset.order_by('-date')

        self._groupedinvoice_queryset = queryset
        return self._groupedinvoice_queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # IMPORTANT: The summary bars should always reflect the whole business snapshot,
        # but they should respect the selected date_range filter.
        user = self.request.user
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        base_qs = GroupedInvoice.objects.filter(user=user)
        if start_date:
            base_qs = base_qs.filter(date__isnull=False, date__gte=start_date)
        if end_date:
            base_qs = base_qs.filter(date__isnull=False, date__lte=end_date)

        invoice_paid_subquery = (
            Payment.objects.filter(invoice=OuterRef('pk'))
            .values('invoice')
            .annotate(
                total=Coalesce(
                    Sum('amount'),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            )
            .values('total')[:1]
        )
        def annotate_with_payments(qs):
            return _annotate_invoice_credit_totals(
                qs.annotate(
                    total_paid=Coalesce(
                        Subquery(invoice_paid_subquery),
                        Value(Decimal('0.00')),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    ),
                )
            ).annotate(
                balance_due=ExpressionWrapper(
                    F('total_amount') - F('total_paid') - F('credit_total'),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
                total_settled=ExpressionWrapper(
                    F('total_paid'),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
            )

        snapshot_qs = annotate_with_payments(base_qs)
        totals = snapshot_qs.aggregate(
            total_invoices=Count('id'),
            total_amount=Coalesce(Sum('total_amount'), Value(Decimal('0.00')), output_field=DecimalField()),
            total_paid_amount=Coalesce(Sum('total_paid'), Value(Decimal('0.00')), output_field=DecimalField()),
            total_pending_amount=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField()),
            total_pending_count=Count('id', filter=Q(balance_due__gt=Decimal('0.00'))),
        )

        total_invoices = totals.get('total_invoices') or 0
        total_amount = totals.get('total_amount') or Decimal('0.00')
        total_paid_amount = totals.get('total_paid_amount') or Decimal('0.00')
        total_pending_amount = totals.get('total_pending_amount') or Decimal('0.00')
        total_pending_count = totals.get('total_pending_count') or 0

        context['total_invoices'] = total_invoices
        context['total_amount'] = total_amount
        context['total_paid_amount'] = total_paid_amount
        context['total_pending_amount'] = total_pending_amount
        context['total_pending_count'] = total_pending_count

        today = timezone.localdate()
        profile = getattr(user, 'profile', None)
        term_days = TERM_CHOICES.get(getattr(profile, 'term', None), 30)
        overdue_threshold = today - timedelta(days=term_days)
        pending_bound_start = today - timedelta(days=365)
        paid_bound_start = today - timedelta(days=30)

        pending_queryset = annotate_with_payments(
            GroupedInvoice.objects.filter(
                user=user,
                date__isnull=False,
                date__gte=pending_bound_start,
                date__lte=today,
            )
        ).filter(balance_due__gt=Decimal('0.00'))

        pending_unpaid_amount = pending_queryset.aggregate(
            total=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField())
        )['total'] or Decimal('0.00')
        pending_overdue_amount = pending_queryset.filter(date__lte=overdue_threshold).aggregate(
            total=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField())
        )['total'] or Decimal('0.00')
        pending_not_due_amount = pending_unpaid_amount - pending_overdue_amount
        if pending_not_due_amount < Decimal('0.00'):
            pending_not_due_amount = Decimal('0.00')

        paid_invoices_last_30 = GroupedInvoice.objects.filter(
            user=user,
            date__isnull=False,
            date_fully_paid__isnull=False,
            date_fully_paid__gte=paid_bound_start,
            date_fully_paid__lte=today,
        )
        total_days = 0
        on_time_count = 0
        paid_count = 0
        for issued_date, paid_date in paid_invoices_last_30.values_list('date', 'date_fully_paid'):
            if not issued_date or not paid_date:
                continue
            delta_days = (paid_date - issued_date).days
            if delta_days < 0:
                delta_days = 0
            total_days += delta_days
            paid_count += 1
            if delta_days <= 30:
                on_time_count += 1

        avg_days_to_pay = round(total_days / paid_count, 1) if paid_count else 0
        on_time_percent = round((on_time_count / paid_count) * 100, 1) if paid_count else 0

        context['pending_unpaid_amount'] = pending_unpaid_amount
        context['pending_overdue_amount'] = pending_overdue_amount
        context['pending_not_due_amount'] = pending_not_due_amount
        context['pending_overdue_percent'] = _calc_percent(pending_overdue_amount, pending_unpaid_amount)
        context['avg_days_to_pay'] = avg_days_to_pay
        context['on_time_percent'] = on_time_percent
        context['paid_fill_percent'] = on_time_percent

        # Pass sorting parameters to template
        sort_by, order = self._get_sorting()
        context['sort_by'] = sort_by
        context['order'] = order
        context['documentType'] = 'invoice'
        context['search_query'] = self.request.GET.get('search', '')
        context['status_filter'] = (self.request.GET.get('status') or '').strip().lower()
        context['date_range_options'] = _get_invoice_date_range_options(self.request.user)
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        context['current_date_range'] = date_range_key
        context['selected_start_date'] = start_date
        context['selected_end_date'] = end_date
        params = self.request.GET.copy()
        params.pop('page', None)
        sort_params = params.copy()
        sort_params.pop('sort_by', None)
        sort_params.pop('order', None)
        context['query_string'] = params.urlencode()
        context['sort_query_string'] = sort_params.urlencode()
        context['payment_methods'] = PAYMENT_METHOD_OPTIONS

        def format_dt(value):
            if not value:
                return ""
            if timezone.is_aware(value):
                value = timezone.localtime(value)
            return value.strftime("%b %d, %Y %I:%M %p").lstrip("0")

        def payment_history_title(invoice):
            payments = list(getattr(invoice, "payments", []).all())
            if not payments:
                return "No payments recorded."
            lines = ["Payment history:"]
            for payment in payments:
                pay_date = payment.date.strftime("%b %d, %Y") if payment.date else "Unknown date"
                amount = f"{payment.amount:,.2f}"
                method = f" ({payment.method})" if payment.method else ""
                lines.append(f"{pay_date}: ${amount}{method}")
            return "\n".join(lines)

        status_class_map = {
            "Paid": "success",
            "Partial": "warning",
            "Pending": "primary",
        }

        for invoice in context.get("object_list", []):
            status_label = getattr(invoice, "status_label", "Pending")
            tone = status_class_map.get(status_label, "primary")
            invoice.email_button_class = f"btn-outline-{tone}"

            sent_at = getattr(invoice, "latest_email_sent_at", None)
            opened_at = getattr(invoice, "latest_email_opened_at", None)
            viewed_at = getattr(invoice, "latest_portal_viewed_at", None)

            last_seen_at = None
            if opened_at and viewed_at:
                last_seen_at = opened_at if opened_at >= viewed_at else viewed_at
            else:
                last_seen_at = opened_at or viewed_at

            invoice.email_status_sent_at = sent_at
            invoice.email_status_opened_at = opened_at
            invoice.email_status_viewed_at = viewed_at
            invoice.email_status_seen_at = last_seen_at

            if status_label == "Paid":
                invoice.email_button_label = "Email receipt"
                invoice.email_button_title = payment_history_title(invoice)
                continue

            if sent_at:
                if last_seen_at:
                    invoice.email_button_label = "Email sent and seen"
                    lines = [f"Sent: {format_dt(sent_at)}"]
                    if opened_at:
                        lines.append(f"Email opened: {format_dt(opened_at)}")
                    if viewed_at:
                        lines.append(f"Portal viewed: {format_dt(viewed_at)}")
                    invoice.email_button_title = "\n".join(lines)
                else:
                    invoice.email_button_label = "Email sent"
                    invoice.email_button_title = f"Sent: {format_dt(sent_at)}"
            else:
                invoice.email_button_label = "Email Invoice"
                invoice.email_button_title = "Not sent yet."
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string('app/groupedinvoice_list_content.html', context)
            # Keep list updates lightweight; summary cards should not change on filter/search.
            return JsonResponse({'html': html})
        return super().render_to_response(context, **response_kwargs)


class GroupedEstimateListView(LoginRequiredMixin, ListView):
    model = GroupedEstimate
    template_name = 'app/estimate_list.html'
    context_object_name = 'object_list'

    def get_queryset(self):
        query = self.request.GET.get('search')
        sort_by = self.request.GET.get('sort_by')
        order = self.request.GET.get('order', 'asc')
        user = self.request.user

        queryset = GroupedEstimate.objects.filter(user=user)
        if query:
            filters = (
                Q(estimate_number__icontains=query) |
                Q(bill_to__icontains=query) |
                Q(date__icontains=query) |
                Q(total_amount__icontains=query) |
                Q(vin_no__icontains=query) |
                Q(unit_no__icontains=query) |
                Q(make_model__icontains=query)
            )
            queryset = queryset.filter(filters)

        allowed_sort_fields = ['estimate_number', 'date', 'bill_to', 'vin_no', 'unit_no', 'total_amount']
        if sort_by in allowed_sort_fields:
            if order == 'desc':
                queryset = queryset.order_by('-' + sort_by)
            else:
                queryset = queryset.order_by(sort_by)
        else:
            queryset = queryset.order_by('-date')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        total_invoices = queryset.count()
        total_amount = queryset.aggregate(total=Sum('total_amount'))['total'] or 0

        context['total_invoices'] = total_invoices
        context['total_amount'] = total_amount
        context['sort_by'] = self.request.GET.get('sort_by', '')
        context['order'] = self.request.GET.get('order', '')
        context['documentType'] = 'estimate'
        context['search_query'] = self.request.GET.get('search', '')
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string('app/estimate_list_content.html', context)
            return JsonResponse({'html': html})
        return super().render_to_response(context, **response_kwargs)


from .ai_service import generate_dynamic_invoice_note

# accounts/views.py

import logging
import stripe

from decimal import Decimal

from django.conf import settings
from django.shortcuts import get_object_or_404
from django.views.generic import DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum

from .models import GroupedInvoice, Profile, UserStripeAccount

logger = logging.getLogger(__name__)
stripe.api_key = settings.STRIPE_SECRET_KEY


class GroupedInvoiceDetailView(LoginRequiredMixin, DetailView):
    model = GroupedInvoice
    template_name = 'app/groupedinvoice_detail.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        accountant_profile = getattr(self.request.user, "accountant_portal", None)
        if accountant_profile:
            detail_access = accountant_profile.accountant_access_level in ("full", "read_only")
            if not detail_access:
                raise Http404
            return queryset.filter(user=accountant_profile.user)
        return queryset.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        invoice = self.get_object()
        accountant_profile = getattr(self.request.user, "accountant_portal", None)
        portal_read_only = bool(accountant_profile)
        context['portal_read_only'] = portal_read_only

        # Basic invoice info
        context['invoice'] = invoice

        # Quantities & totals with interest handled separately
        records = list(invoice.income_records.all())
        total_qty = sum(r.qty or 0 for r in records)
        interest_total = sum(
            r.amount for r in records
            if r.job and str(r.job).lower().startswith('interest')
        )
        taxable_records = [
            r for r in records
            if not (r.job and str(r.job).lower().startswith('interest'))
        ]
        pre_tax_subtotal = sum(r.amount for r in taxable_records)
        tax = sum(r.tax_collected for r in taxable_records)
        subtotal = pre_tax_subtotal + interest_total
        total_amount = pre_tax_subtotal + tax + interest_total

        context.update({
            'interest_total': interest_total,
            'pre_tax_subtotal': pre_tax_subtotal,
        })

        context.update({
            'total_qty':    total_qty,
            'subtotal':     subtotal,
            'tax':          tax,
            'total_amount': total_amount,
            'documentType': 'invoice',
            'balance_due': invoice.balance_due(),
        })

        # Profile & tax visibility
        profile = get_object_or_404(Profile, user=invoice.user)
        context['profile'] = profile

        context['show_tax'] = True

        # Logo URL for PDFs/emails
        context['company_logo_url'] = resolve_company_logo_url(
            profile,
            request=self.request,
        )

        # Back-button source
        source = self.request.GET.get('source', 'all')
        context['source'] = source
        if source == 'customer_jobs':
            context['customer'] = self.request.GET.get('customer')

        invoice_note = (invoice.notes or '').strip()
        if invoice_note:
            context['note'] = invoice_note
        elif profile.show_note:
            context['note'] = (
                generate_dynamic_invoice_note(invoice)
                if profile.use_dynamic_note else
                profile.note
            )
        else:
            context['note'] = ''

        # Stripe "receipt PDF" link
        pdf_url = None
        if invoice.payment_status == 'Paid' and invoice.stripe_invoice_id:
            try:
                stripe_inv = stripe.Invoice.retrieve(
                    invoice.stripe_invoice_id,
                    stripe_account=invoice.user.userstripeaccount.stripe_account_id
                )
                pdf_url = stripe_inv.get("invoice_pdf")
            except Exception as e:
                logger.warning(
                    "Could not retrieve invoice_pdf for %s: %s",
                    invoice.invoice_number, e
                )
        context['stripe_invoice_pdf_url'] = pdf_url

        # ————————————————————————————————
        # Show "Pay In-Person"?
        # ————————————————————————————————
        user = invoice.user

        provider = invoice.get_payment_link_provider()

        # 1) Do they have a connected Stripe account?
        has_connected_acct = (
            provider == PAYMENT_LINK_PROVIDER_STRIPE
            and hasattr(user, 'userstripeaccount')
            and bool(user.userstripeaccount.stripe_account_id)
        )

        # 2) Is this invoice unpaid and has an online checkout link?
        checkout_link = invoice.payment_link
        can_pay_inperson = (
            invoice.payment_status != 'Paid'
            and bool(checkout_link)
            and has_connected_acct
        )

        # 3) Is Terminal *enabled* on this user's profile?
        terminal_enabled = getattr(user.profile, 'terminal_enabled', False)

        # 4) Only show the button when *all* conditions are met:
        context['show_inperson_btn'] = can_pay_inperson and terminal_enabled and not portal_read_only

        # Expose the available online checkout link to templates
        context['payment_link'] = checkout_link
        context['payment_provider'] = provider

        context['payment_methods'] = PAYMENT_METHOD_OPTIONS

        context['can_reverse_payment'] = (
            invoice.payment_status == 'Paid'
            and invoice.payments.filter(
                notes__icontains='Marked as paid via MarkInvoicePaidView'
            ).exists()
        )

        activity_entries = list(invoice.activity_entries.all())
        sent_entries = [
            entry for entry in activity_entries
            if entry.event_type == InvoiceActivity.EVENT_EMAIL_SENT
        ]
        opened_entries = [
            entry for entry in activity_entries
            if entry.event_type == InvoiceActivity.EVENT_EMAIL_OPENED
        ]
        viewed_entries = [
            entry for entry in activity_entries
            if entry.event_type == InvoiceActivity.EVENT_VIEWED
        ]
        context['invoice_activity'] = {
            'sent_entries': sent_entries,
            'opened_entries': opened_entries,
            'viewed_entries': viewed_entries,
            'sent_count': len(sent_entries),
            'opened_count': len(opened_entries),
            'viewed_count': len(viewed_entries),
        }

        return context





@login_required
def generate_payment_link_view(request, invoice_id):
    invoice = get_object_or_404(GroupedInvoice, id=invoice_id, user=request.user)

    provider = invoice.get_payment_link_provider()

    if provider == PAYMENT_LINK_PROVIDER_STRIPE and not verify_stripe_account(request.user):
        messages.error(request, 'No valid Stripe account found.')
        return render(request, 'payment/payment_error.html', {'message': 'No valid Stripe account found.'})

    if provider == PAYMENT_LINK_PROVIDER_CLOVER:
        from .models import CloverConnection
        if not CloverConnection.objects.filter(user=request.user).exists():
            messages.error(request, 'No valid Clover connection found.')
            return render(
                request,
                'payment/payment_error.html',
                {'message': 'No valid Clover connection found.'},
            )

    # Recalculate total amount and trigger payment link update.
    invoice.recalculate_total_amount()
    invoice.save()  # Model's save() will trigger create_online_payment_link if needed.

    checkout_link = invoice.payment_link
    if checkout_link:
        messages.success(request, 'Payment link created/updated successfully.')
        return render(request, 'payment/payment_link.html', {'payment_link': checkout_link})

    messages.error(request, 'Failed to create/update payment link.')
    return render(
        request,
        'payment/payment_error.html',
        {'message': 'Failed to create/update payment link.'},
    )



def generate_invoices_detail(request):
    return render(request, 'public/generate_invoices_detail.html')

def track_expenses_detail(request):
    return render(request, 'public/track_expenses_detail.html')

def manage_inventory_detail(request):
    return render(request, 'public/manage_inventory_detail.html')

def search_purchases_detail(request):
    return render(request, 'public/search_purchases_detail.html')

def income_expense_table_detail(request):
    return render(request, 'public/income_expense_table_detail.html')

def process_payments_detail(request):
    return render(request, 'public/process_payments_detail.html')

class AddPaymentView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        invoice_id = request.POST.get('invoice_id')
        amount = request.POST.get('amount')
        method = request.POST.get('method', 'Manual')
        notes = request.POST.get('notes', '')

        invoice = get_object_or_404(GroupedInvoice, pk=invoice_id, user=request.user)
        try:
            amount = Decimal(amount)
        except:
            messages.error(request, "Invalid amount entered.")
            return redirect('accounts:pending_invoice_list')

        if amount <= 0:
            messages.error(request, "Amount must be greater than zero.")
            return redirect('accounts:pending_invoice_list')

        balance_due = invoice.balance_due()

        if amount > balance_due + Decimal('0.01'):
            messages.error(request, f"Amount cannot exceed the balance due (${balance_due}).")
            return redirect('accounts:pending_invoice_list')

        # Create the payment
        payment = Payment.objects.create(
            invoice=invoice,
            amount=amount,
            method=method,
            notes=notes
        )

        # Update the invoice's payment status
        invoice.update_date_fully_paid()

        messages.success(request, f"Payment of ${amount} added to Invoice {invoice.invoice_number}.")
        return redirect('accounts:pending_invoice_list')


@login_required
def payment_history(request):
    invoice_id = request.GET.get('invoice_id')
    try:
        invoice = GroupedInvoice.objects.get(pk=invoice_id, user=request.user)
        payments = invoice.payments.all().order_by('-date')

        html = render_to_string('app/payment_history_content.html', {'payments': payments}, request=request)
        return JsonResponse({'html': html})
    except GroupedInvoice.DoesNotExist:
        return JsonResponse({'html': '<p class="text-danger">Invoice not found.</p>'})
    except Exception as e:
        return JsonResponse({'html': f'<p class="text-danger">An error occurred: {str(e)}</p>'})

class OverdueInvoiceListView(LoginRequiredMixin, ListView):
    model = PendingInvoice
    template_name = 'app/overdueinvoice_list.html'
    context_object_name = 'overdue_invoices'
    paginate_by = 100

    def _get_view_mode(self):
        mode = (self.request.GET.get('view') or 'invoice').strip().lower()
        return 'customer' if mode == 'customer' else 'invoice'

    def _get_invoice_queryset(self):
        if hasattr(self, '_overdue_invoice_queryset'):
            return self._overdue_invoice_queryset

        user = self.request.user
        profile = getattr(user, 'profile', None)
        term_days = TERM_CHOICES.get(getattr(profile, 'term', None), 30)
        today = timezone.localdate()
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        query = (self.request.GET.get('search') or '').strip()

        qs = PendingInvoice.objects.select_related('grouped_invoice', 'grouped_invoice__customer').filter(
            is_paid=False,
            grouped_invoice__user=user,
            grouped_invoice__date__lt=(today - timedelta(days=term_days))
        ).annotate(
            total_paid=Coalesce(
                Sum('grouped_invoice__payments__amount'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )
        qs = _annotate_invoice_credit_totals(
            qs,
            invoice_field='grouped_invoice_id',
        ).annotate(
            balance_due=ExpressionWrapper(
                F('grouped_invoice__total_amount') - F('total_paid') - F('credit_total'),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        ).filter(balance_due__gt=0)

        if start_date:
            qs = qs.filter(grouped_invoice__date__isnull=False, grouped_invoice__date__gte=start_date)
        if end_date:
            qs = qs.filter(grouped_invoice__date__isnull=False, grouped_invoice__date__lte=end_date)

        if query:
            qs = qs.filter(
                Q(grouped_invoice__invoice_number__icontains=query) |
                Q(grouped_invoice__bill_to__icontains=query) |
                Q(grouped_invoice__date__icontains=query) |
                Q(grouped_invoice__total_amount__icontains=query) |
                Q(grouped_invoice__customer__name__icontains=query)
            )

        sort_by, order = _resolve_sorting(
            self.request,
            default_sort='invoice_number',
            default_order='desc',
        )
        sort_map = {
            'invoice_number': 'grouped_invoice__invoice_number',
            'date': 'grouped_invoice__date',
            'bill_to': 'grouped_invoice__bill_to',
            'total_amount': 'grouped_invoice__total_amount',
            'total_paid': 'total_paid',
            'balance_due': 'balance_due',
        }
        if sort_by == 'invoice_number':
            qs = _annotate_invoice_number_sort(qs, 'grouped_invoice__invoice_number')
            if order == 'desc':
                qs = qs.order_by('-invoice_number_sort', '-grouped_invoice__invoice_number')
            else:
                qs = qs.order_by('invoice_number_sort', 'grouped_invoice__invoice_number')
        else:
            sort_field = sort_map.get(sort_by)
            if sort_field:
                if order == 'desc':
                    qs = qs.order_by(f'-{sort_field}')
                else:
                    qs = qs.order_by(sort_field)
            else:
                qs = qs.order_by('-grouped_invoice__date')

        self._overdue_invoice_queryset = qs
        return self._overdue_invoice_queryset

    def _get_customer_rows(self):
        if hasattr(self, '_overdue_customer_rows'):
            return self._overdue_customer_rows

        user = self.request.user
        profile = getattr(user, 'profile', None)
        term_days = TERM_CHOICES.get(getattr(profile, 'term', None), 30)
        today = timezone.localdate()
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        query = (self.request.GET.get('search') or '').strip()

        overdue_filter = {'date__lt': today - timedelta(days=term_days)}

        invoice_paid_subquery = (
            Payment.objects.filter(invoice=OuterRef('pk'))
            .values('invoice')
            .annotate(
                total=Coalesce(
                    Sum('amount'),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            )
            .values('total')[:1]
        )
        invoices = _annotate_invoice_credit_totals(
            GroupedInvoice.objects.filter(
                user=user,
                customer__isnull=False,
                date__isnull=False,
                **overdue_filter,
            ).annotate(
                total_paid=Coalesce(
                    Subquery(invoice_paid_subquery),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
            )
        ).annotate(
            balance_due=ExpressionWrapper(
                F('total_amount') - F('total_paid') - F('credit_total'),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
        ).filter(balance_due__gt=Decimal('0.00'))

        if start_date:
            invoices = invoices.filter(date__gte=start_date)
        if end_date:
            invoices = invoices.filter(date__lte=end_date)

        if query:
            invoices = invoices.filter(
                Q(customer__name__icontains=query) |
                Q(customer__email__icontains=query)
            )

        reminder_last_sent = ReminderLog.objects.filter(
            customer_id=OuterRef('customer_id')
        ).order_by('-sent_at').values('sent_at')[:1]
        reminder_count = ReminderLog.objects.filter(
            customer_id=OuterRef('customer_id')
        ).values('customer_id').annotate(total=Count('id')).values('total')[:1]

        rows = invoices.values('customer_id', 'customer__name').annotate(
            total_amount=Coalesce(Sum('total_amount'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
            total_paid=Coalesce(Sum('total_paid'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
            balance_due=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
            last_statement=Subquery(reminder_last_sent, output_field=DateTimeField()),
            reminder_count=Coalesce(Subquery(reminder_count, output_field=IntegerField()), Value(0)),
        )

        sort_by, order = _resolve_sorting(self.request, default_sort='balance_due', default_order='desc')
        sort_map = {
            'customer_name': 'customer__name',
            'last_statement': 'last_statement',
            'total_amount': 'total_amount',
            'total_paid': 'total_paid',
            'balance_due': 'balance_due',
        }
        sort_field = sort_map.get(sort_by)
        if sort_field:
            if order == 'desc':
                rows = rows.order_by(f'-{sort_field}', 'customer__name')
            else:
                rows = rows.order_by(sort_field, 'customer__name')
        else:
            rows = rows.order_by('-balance_due', 'customer__name')

        total_overdue = invoices.aggregate(
            total=Coalesce(Sum('balance_due'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2))
        )['total'] or Decimal('0.00')
        self._overdue_customer_summary = {
            'total_overdue': total_overdue,
            'customer_count': rows.count(),
        }

        self._overdue_customer_rows = rows
        return self._overdue_customer_rows

    def get_queryset(self):
        view_mode = self._get_view_mode()
        if view_mode == 'customer':
            return self._get_customer_rows()
        return self._get_invoice_queryset()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        view_mode = self._get_view_mode()
        context['view_mode'] = view_mode

        if view_mode == 'customer':
            customer_rows = context.get('overdue_invoices') or context.get('object_list')
            summary = getattr(self, '_overdue_customer_summary', None) or {}
            total_overdue_balance = summary.get('total_overdue', Decimal('0.00'))
            overdue_count = summary.get('customer_count', len(customer_rows or []))
            context['overdue_customers'] = customer_rows
            context['overdue_invoices'] = []
            context['overdue_count_label'] = 'Customers'
        else:
            queryset = self._get_invoice_queryset()
            total_overdue_balance = queryset.aggregate(
                total=Sum('balance_due')
            )['total'] or Decimal('0.00')
            overdue_count = queryset.count()
            context['overdue_invoices'] = context.get('overdue_invoices') or context.get('object_list')
            context['overdue_count_label'] = 'Invoices'

        context['total_overdue_balance'] = total_overdue_balance
        context['overdue_count'] = overdue_count
        context['date_range_options'] = _get_invoice_date_range_options(self.request.user)
        date_range_key, start_date, end_date = _get_date_range_bounds(
            self.request,
            default="1y",
            cookie_key=DATE_RANGE_COOKIE_NAME,
        )
        context['current_date_range'] = date_range_key
        context['selected_start_date'] = start_date
        context['selected_end_date'] = end_date
        params = self.request.GET.copy()
        params.pop('page', None)
        sort_params = params.copy()
        sort_params.pop('sort_by', None)
        sort_params.pop('order', None)
        context['query_string'] = params.urlencode()
        context['sort_query_string'] = sort_params.urlencode()
        context['search_query'] = self.request.GET.get('search', '')
        sort_default = 'invoice_number' if view_mode == 'invoice' else 'balance_due'
        sort_order = 'desc' if view_mode == 'invoice' else 'desc'
        sort_by, order = _resolve_sorting(
            self.request,
            default_sort=sort_default,
            default_order=sort_order,
        )
        context['sort_by'] = sort_by
        context['order'] = order
        context['today_date'] = timezone.localdate()
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string('app/overdue_invoices_table.html', context, request=self.request)
            total_overdue_balance = context.get('total_overdue_balance') or Decimal('0.00')
            start_date = context.get('selected_start_date')
            end_date = context.get('selected_end_date')
            return JsonResponse({
                'html': html,
                'total_overdue_balance': f"${total_overdue_balance:,.2f}",
                'overdue_count': context.get('overdue_count', 0),
                'overdue_count_label': context.get('overdue_count_label', 'Invoices'),
                'view_mode': context.get('view_mode', 'invoice'),
                'start_date': start_date.isoformat() if start_date else '',
                'end_date': end_date.isoformat() if end_date else '',
                'current_date_range': context.get('current_date_range', ''),
            })
        return super().render_to_response(context, **response_kwargs)



@login_required
def choose_plan(request):
    if request.method == 'POST':
        plan_name = request.POST.get('plan_name')
        if plan_name in settings.STRIPE_PLANS:
            # Store the selected plan in the session
            request.session['selected_plan'] = plan_name
            return redirect('accounts:payment')
        else:
            messages.error(request, "Invalid plan selected.")
    return render(request, 'accounts/choose_plan.html')

@login_required
def edit_dailylog(request, pk):
    # Get the invoice and ensure it belongs to the logged-in user
    invoice = get_object_or_404(GroupedInvoice, pk=pk, user=request.user)
    documentType = 'invoice'  # Hardcoded for this example

    # Only truck mechanic templates remain
    template_name = 'accounts/mach/daily_workslip.html'

    FormsetClass = IncomeRecord2FormSet  # Specific to invoice editing
    GroupedForm = GroupedInvoiceForm     # Specific to invoice editing

    if request.method == 'POST':
        grouped_form = GroupedForm(
            request.POST,
            instance=invoice,
            user=request.user
        )
        formset = FormsetClass(
            request.POST,
            instance=invoice,
            prefix='income_records',  # Important for formset processing
            form_kwargs={'user': request.user}
        )

        if grouped_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    grouped_invoice = grouped_form.save(commit=False)
                    grouped_invoice.user = request.user

                    customer = grouped_form.cleaned_data.get('customer')
                    if customer:
                        grouped_invoice.bill_to = customer.name
                        grouped_invoice.bill_to_email = customer.email
                        grouped_invoice.bill_to_address = customer.address
                    else:
                        grouped_invoice.bill_to = grouped_form.cleaned_data.get('bill_to', None)
                        grouped_invoice.bill_to_email = grouped_form.cleaned_data.get('bill_to_email', None)
                        grouped_invoice.bill_to_address = grouped_form.cleaned_data.get('bill_to_address', None)

                    grouped_invoice.save()
                    formset.save()

                    # Ensure inventory transactions are present for all items
                    if documentType == 'invoice':
                        grouped_invoice.ensure_inventory_transactions()

                    # Update customer rate if needed and handle pending invoice logic
                    first_record = grouped_invoice.income_records.order_by('line_order', 'id').first()
                    if first_record and grouped_invoice.customer and first_record.rate is not None:
                        if grouped_invoice.customer.charge_rate != first_record.rate:
                            grouped_invoice.customer.charge_rate = first_record.rate
                            grouped_invoice.customer.save(update_fields=['charge_rate'])

                    pending_invoice, created = PendingInvoice.objects.get_or_create(grouped_invoice=grouped_invoice)

                    messages.success(request, f"Invoice {grouped_invoice.invoice_number} updated successfully")
            except Exception as e:
                messages.error(request, f"An error occurred during update: {e}")
                logger.error(f"Error editing daily log (pk={pk}): {e}", exc_info=True)

            redirect_dashboard_flag = request.POST.get('redirect_dashboard')
            save_send_flag = request.POST.get('save_send')
            list_url = reverse('accounts:groupedinvoice_list')
            edit_url = reverse('accounts:groupedinvoice_edit', kwargs={'pk': grouped_invoice.pk})

            if redirect_dashboard_flag:
                messages.success(request, "Invoice updated successfully.")
                return redirect(list_url)
            if save_send_flag:
                try:
                    if getattr(grouped_invoice, 'payment_status', None) == 'Paid':
                        send_paid_invoice_email(grouped_invoice, request=request, include_workorder=False)
                    else:
                        send_grouped_invoice_email(request, grouped_invoice.pk)
                    messages.success(request, f"Invoice {grouped_invoice.invoice_number} saved and email sent successfully.")
                except Exception as e:
                    messages.error(request, f"Invoice saved, but email failed to send: {e}")
                return redirect(edit_url)
            if 'save_stay' in request.POST:
                messages.success(request, f"Invoice {grouped_invoice.invoice_number} updated successfully.")
                return redirect(edit_url)
            if 'save_continue' in request.POST:
                return redirect(f"{reverse('accounts:add_dailylog')}?type=invoice")
            if 'submit_income' in request.POST:
                return redirect(list_url)
            else:
                return redirect(edit_url)
        else:
            error_list = []
            if grouped_form.errors:
                error_list.extend([f"{field}: {err[0]}" for field, err in grouped_form.errors.items()])
            if formset.errors:
                for i, form_errors in enumerate(formset.errors):
                    if form_errors:
                        error_list.extend([f"Item {i+1} - {field}: {err[0]}" for field, err in form_errors.items()])
            if formset.non_form_errors():
                error_list.extend(formset.non_form_errors())
            messages.error(request, 'Please correct the errors below: ' + "; ".join(error_list))
    else:
        grouped_form = GroupedForm(instance=invoice, user=request.user)
        formset = FormsetClass(
            instance=invoice,
            prefix='income_records',
            form_kwargs={'user': request.user}
        )

    # Prepare additional context variables
    business_user_ids = get_customer_user_ids(request.user)
    customer_form = CustomerForm(user=request.user)  # For modal add
    customers = Customer.objects.filter(user__in=business_user_ids)
    product_user_ids = get_product_user_ids(request.user)
    products = Product.objects.filter(user__in=product_user_ids)
    product_data = {
        str(p.id): {
            "name": p.name,
            "sku": p.sku or '',
            "description": p.description or '',
            "price": str(p.sale_price),
            "item_type": p.item_type,
        }
        for p in products
    }
    product_data_json = json.dumps(product_data)
    has_driver = Driver.objects.filter(user=request.user).exists()
    # IMPORTANT: Pass the full drivers queryset for use in the drop down
    drivers = Driver.objects.filter(user=request.user)
    fleet_vehicles = FleetVehicle.objects.filter(user=request.user)

    user_profile = getattr(request.user, 'profile', None)
    if user_profile:
        province = getattr(user_profile, 'province', None)
        tax_name = user_profile.get_tax_name() if hasattr(user_profile, 'get_tax_name') else "Tax"
    else:
        province = None
        tax_name = "Tax"
    tax_rate = PROVINCE_TAX_RATES.get(province, 0.00)
    tax_components_json = json.dumps(get_tax_component_rates(province))

    existing_jobsites = IncomeRecord2.objects.filter(grouped_invoice__user=request.user)\
        .values_list('jobsite', flat=True).distinct()
    existing_trucks = IncomeRecord2.objects.filter(grouped_invoice__user=request.user)\
        .values_list('truck', flat=True).distinct()
    existing_jobs = IncomeRecord2.objects.filter(grouped_invoice__user=request.user)\
        .values_list('job', flat=True).distinct()
    job_values = list(existing_jobs.exclude(job__isnull=True).exclude(job__exact=''))
    job_catalog, service_description_strings, _job_name_choices = build_service_job_catalog(request.user)
    job_list = sorted(set(job_values) | set(service_description_strings))

    resolved_vehicle = getattr(invoice, 'vehicle', None)
    vehicle_history_payload = _serialize_vehicle_history(resolved_vehicle)
    vehicle_maintenance_payload = _serialize_vehicle_maintenance(resolved_vehicle)

    total_amount = invoice.total_amount or Decimal('0.00')
    total_paid = invoice.total_paid()
    credit_total = ensure_decimal(invoice.total_credit_amount)
    total_settled = total_paid
    tolerance = Decimal('0.01')
    pending_invoice = getattr(invoice, 'pending_invoice', None)
    pending_is_paid = bool(pending_invoice and getattr(pending_invoice, 'is_paid', False))
    invoice_is_paid = pending_is_paid or bool(getattr(invoice, 'paid_invoice', None)) or (total_amount > Decimal('0.00') and total_settled + tolerance >= total_amount)
    if invoice_is_paid:
        invoice_total_paid_amount = total_amount
        invoice_balance_due_amount = Decimal('0.00')
    else:
        invoice_total_paid_amount = total_settled
        invoice_balance_due_amount = (total_amount - total_paid - credit_total).quantize(Decimal('0.01'))
        if invoice_balance_due_amount < Decimal('0.00'):
            invoice_balance_due_amount = Decimal('0.00')

    payment_history = list(invoice.payments.order_by('date', 'id'))
    payment_history_count = len(payment_history)
    payment_history_last_date = payment_history[-1].date if payment_history else None

    document_totals = _calculate_grouped_doc_totals(invoice.income_records.all())

    context = {
        'invoice': invoice,
        'grouped_invoice_form': grouped_form,
        'formset': formset,
        'customer_form': customer_form,
        'customers': customers,
        'products': products,
        'product_data_json': product_data_json,
        'has_driver': has_driver,
        'drivers': drivers,  # <<--- Added here to supply driver data
        'fleet_vehicles': fleet_vehicles,
        'jobsite_list': list(existing_jobsites.exclude(jobsite__isnull=True).exclude(jobsite__exact='')),
        'truck_list': list(existing_trucks.exclude(truck__isnull=True).exclude(truck__exact='')),
        'job_list': job_list,
        'job_catalog_json': job_catalog,
        'documentType': documentType,
        'formset_prefix': 'income_records',
        'generated_invoice_number': invoice.invoice_number,
        'payment_methods': PAYMENT_METHOD_OPTIONS,
        'invoice_is_paid': invoice_is_paid,
        'invoice_total_paid_amount': invoice_total_paid_amount,
        'invoice_balance_due_amount': invoice_balance_due_amount,
        'payment_history': payment_history,
        'payment_history_count': payment_history_count,
        'payment_history_last_date': payment_history_last_date,
        'base_subtotal_amount': document_totals['base_subtotal'],
        'shop_supply_total_amount': document_totals['shop_supply_total'],
        'discount_total_amount': document_totals['discount_total'],
        'interest_total_amount': document_totals['interest_total'],
        'pre_tax_subtotal_amount': document_totals['pre_tax_subtotal'],
        'subtotal_after_interest_amount': document_totals['subtotal_after_interest'],
        'tax_total_amount': document_totals['tax_total'],
        'grand_total_amount': document_totals['grand_total'],
        'tax_name': tax_name,
        'tax_rate': tax_rate,
        'tax_components_json': tax_components_json,
        'vehicle_history': vehicle_history_payload,
        'vehicle_history_json': json.dumps(vehicle_history_payload, cls=DjangoJSONEncoder),
        'vehicle_history_api': reverse('accounts:vehicle_history_summary'),
        'vehicle_maintenance': vehicle_maintenance_payload,
        'vehicle_maintenance_json': json.dumps(vehicle_maintenance_payload, cls=DjangoJSONEncoder),
        'vehicle_maintenance_api': reverse('accounts:vehicle_maintenance_summary'),
    }
    return render(request, template_name, context)


@login_required
def maintenance_center(request):
    """Unified maintenance hub with tabbed task lists and email trigger."""
    customer_id = request.GET.get('customer') or None
    vehicle_id = request.GET.get('vehicle') or None
    business_user_ids = get_customer_user_ids(request.user)

    base_qs = VehicleMaintenanceTask.objects.select_related(
        'vehicle',
        'vehicle__customer',
    ).filter(user=request.user)

    if customer_id:
        base_qs = base_qs.filter(vehicle__customer_id=customer_id)
    if vehicle_id:
        base_qs = base_qs.filter(vehicle_id=vehicle_id)

    upcoming_tasks = base_qs.filter(
        status__in=VehicleMaintenanceTask.active_statuses()
    ).order_by('due_date', 'priority', 'title', 'pk')
    completed_tasks = base_qs.filter(
        status=VehicleMaintenanceTask.STATUS_COMPLETED
    ).order_by('-completed_date', '-updated_at', '-pk')[:150]
    cancelled_tasks = base_qs.filter(
        status=VehicleMaintenanceTask.STATUS_CANCELLED
    ).order_by('-updated_at', '-pk')[:150]

    customers = Customer.objects.filter(user__in=business_user_ids).order_by('name')
    vehicles = Vehicle.objects.filter(customer__user__in=business_user_ids).order_by('unit_number', 'make_model', 'vin_number')
    if customer_id:
        vehicles = vehicles.filter(customer_id=customer_id)

    context = {
        'upcoming_tasks': upcoming_tasks,
        'completed_tasks': completed_tasks,
        'cancelled_tasks': cancelled_tasks,
        'customers': customers,
        'vehicles': vehicles,
        'active_customer_id': int(customer_id) if customer_id else None,
        'active_vehicle_id': int(vehicle_id) if vehicle_id else None,
    }
    return render(request, 'accounts/mach/maintenance_center.html', context)


@login_required
@require_POST
def send_maintenance_reminders(request):
    """Trigger upcoming maintenance emails to customers for this business."""
    days_ahead = 7
    try:
        days_ahead = max(1, min(30, int(request.POST.get('days', 7))))
    except (TypeError, ValueError):
        days_ahead = 7

    business_user_ids = get_customer_user_ids(request.user)

    customer_id = request.POST.get('customer') or None
    vehicle_id = request.POST.get('vehicle') or None

    today = timezone.localdate()
    window_end = today + timedelta(days=days_ahead)
    now = timezone.now()

    task_qs = VehicleMaintenanceTask.objects.select_related(
        'vehicle__customer',
        'vehicle__customer__portal_user',
        'user__profile',
    ).filter(
        user=request.user,
        status__in=VehicleMaintenanceTask.active_statuses(),
        due_date__isnull=False,
        due_date__lte=window_end,
    )

    if customer_id:
        task_qs = task_qs.filter(vehicle__customer_id=customer_id)
    if vehicle_id:
        task_qs = task_qs.filter(vehicle_id=vehicle_id)

    if not task_qs.exists():
        messages.info(request, "No upcoming maintenance tasks found in the selected window.")
        return redirect('accounts:maintenance_center')

    # Last service dates map for vehicles
    last_service_map = dict(
        VehicleMaintenanceTask.objects.filter(
            vehicle__customer__user__in=business_user_ids,
            status=VehicleMaintenanceTask.STATUS_COMPLETED,
        ).values('vehicle_id').annotate(last_completed=Max('completed_date')).values_list('vehicle_id', 'last_completed')
    )

    tasks_by_customer = {}
    for task in task_qs:
        cust = getattr(task.vehicle, 'customer', None)
        if not cust:
            continue
        tasks_by_customer.setdefault(cust.id, {'customer': cust, 'tasks': []})
        tasks_by_customer[cust.id]['tasks'].append(task)

    reminders_sent = 0
    booking_url = request.build_absolute_uri(reverse('accounts:public_booking')) + "?service=maintenance"
    portal_url = request.build_absolute_uri(reverse('accounts:customer_dashboard'))

    for payload in tasks_by_customer.values():
        customer = payload['customer']
        tasks = payload['tasks']
        recipient = (customer.email or None)
        if not recipient and getattr(customer, 'portal_user', None):
            recipient = customer.portal_user.email or None
        if not recipient:
            continue
        customer_cc_emails = customer.get_cc_emails()
        cc_recipients = build_cc_list(*customer_cc_emails, exclude=[recipient])

        profile = getattr(request.user, 'profile', None)
        business_name = (
            (profile.company_name if profile and profile.company_name else None)
            or request.user.get_full_name()
            or request.user.username
        )
        contact_email = (
            (profile.company_email if profile and profile.company_email else None)
            or request.user.email
            or settings.DEFAULT_FROM_EMAIL
        )

        overdue_tasks = []
        upcoming_tasks = []
        for task in sorted(tasks, key=lambda t: (t.due_date or window_end, t.priority, t.title.lower())):
            task.last_completed = last_service_map.get(task.vehicle_id)
            if task.due_date and task.due_date < today:
                overdue_tasks.append(task)
            else:
                upcoming_tasks.append(task)

        context = {
            "customer": customer,
            "profile": profile,
            "business_name": business_name,
            "contact_email": contact_email,
            "overdue_tasks": overdue_tasks,
            "upcoming_tasks": upcoming_tasks,
            "today": today,
            "window_end": window_end,
            "portal_url": portal_url,
            "booking_url": booking_url,
            "days_ahead": days_ahead,
            "last_service_map": last_service_map,
        }

        subject = f"Upcoming maintenance reminders from {business_name}"
        html_body = render_to_string("emails/maintenance_due_email.html", context)
        text_body = strip_tags(html_body)

        email = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[recipient],
            cc=cc_recipients or None,
        )
        email.attach_alternative(html_body, "text/html")

        try:
            email.send()
        except Exception:
            continue

        task_ids = [t.id for t in tasks]
        VehicleMaintenanceTask.objects.filter(id__in=task_ids).update(last_reminder_sent=now)
        reminders_sent += 1

    if reminders_sent:
        messages.success(request, f"Sent maintenance reminders to {reminders_sent} customer(s).")
    else:
        messages.warning(request, "No maintenance reminders were sent (missing emails or tasks).")
    return redirect('accounts:maintenance_center')


def custom_404(request, exception):
    return render(request, '404.html', status=404)

@activation_required
@subscription_required
@login_required
def add_estimate(request):
    template_name = 'accounts/mach/daily_workslip.html'

    if request.method == 'POST':
        grouped_estimate_form = GroupedEstimateForm(request.POST, user=request.user)
        formset = EstimateRecordFormSet(
            request.POST,
            queryset=EstimateRecord.objects.none(),
            form_kwargs={'user': request.user}
        )
        customer_form = CustomerForm(user=request.user)

        if grouped_estimate_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    grouped_estimate = grouped_estimate_form.save(commit=False)
                    grouped_estimate.user = request.user

                    # Set bill-to details from selected customer
                    customer = grouped_estimate_form.cleaned_data.get('customer')
                    if customer:
                        grouped_estimate.bill_to = customer.name
                        grouped_estimate.bill_to_email = customer.email
                        grouped_estimate.bill_to_address = customer.address

                    grouped_estimate.save()
                    formset.instance = grouped_estimate
                    formset.save()

                    grouped_estimate.recalculate_total_amount()

                    messages.success(request, f'Estimate {grouped_estimate.estimate_number} created successfully')
            except ValidationError as e:
                messages.error(request, e.message)
                return redirect('accounts:add_estimate')

            if request.POST.get('redirect_dashboard'):
                messages.success(request, "Estimate saved successfully.")
                return redirect('accounts:estimate_list')
            if 'save_continue' in request.POST:
                return redirect('accounts:add_estimate')
            if 'submit_income' in request.POST:
                return redirect('accounts:estimate_list')
            else:
                return redirect('accounts:estimate_detail', pk=grouped_estimate.pk)
        else:
            messages.error(request, 'There was an error in the form. Please correct it and try again.')
    else:
        grouped_estimate_form = GroupedEstimateForm(user=request.user)
        formset = EstimateRecordFormSet(
            queryset=EstimateRecord.objects.none(),
            form_kwargs={'user': request.user}
        )
        customer_form = CustomerForm(user=request.user)

    business_user_ids = get_customer_user_ids(request.user)
    customers = Customer.objects.filter(user__in=business_user_ids)
    existing_jobsites = EstimateRecord.objects.filter(grouped_estimate__user=request.user)\
                           .values_list('jobsite', flat=True).distinct()
    existing_trucks = EstimateRecord.objects.filter(grouped_estimate__user=request.user)\
                           .values_list('truck', flat=True).distinct()
    existing_jobs = EstimateRecord.objects.filter(grouped_estimate__user=request.user)\
                           .values_list('job', flat=True).distinct()
    job_values = list(existing_jobs.exclude(job__isnull=True).exclude(job__exact=''))
    job_catalog, service_description_strings, _job_name_choices = build_service_job_catalog(request.user)
    job_list = sorted(set(job_values) | set(service_description_strings))

    product_user_ids = get_product_user_ids(request.user)
    products = Product.objects.filter(user__in=product_user_ids)
    product_data = {
        str(product.id): {
            "name": product.name,
            "sku": product.sku or '',
            "description": product.description or '',
            "price": str(product.sale_price),
            "item_type": product.item_type,
        }
        for product in products
    }
    product_data_json = json.dumps(product_data)
    has_driver = Driver.objects.filter(user=request.user).exists()
    drivers = Driver.objects.filter(user=request.user) if has_driver else None
    fleet_vehicles = FleetVehicle.objects.filter(user=request.user)

    user_profile = getattr(request.user, 'profile', None)
    if user_profile:
        province = getattr(user_profile, 'province', None)
        tax_name = user_profile.get_tax_name() if hasattr(user_profile, 'get_tax_name') else "Tax"
    else:
        province = None
        tax_name = "Tax"
    tax_rate = PROVINCE_TAX_RATES.get(province, 0.00)
    tax_components_json = json.dumps(get_tax_component_rates(province))

    default_reasons = [
        "Indigenous Status",
        "Government Purchase",
        "Resale",
        "Export",
    ]
    saved_reasons = list(
        TaxExemptionReason.objects.filter(user=request.user).values_list('reason', flat=True)
    )
    exemption_reasons = list(dict.fromkeys(default_reasons + saved_reasons))

    document_records = []
    if getattr(grouped_estimate_form, 'instance', None) and grouped_estimate_form.instance.pk:
        document_records = list(grouped_estimate_form.instance.estimate_records.all())
    document_totals = _calculate_grouped_doc_totals(document_records)

    context = {
        'grouped_estimate_form': grouped_estimate_form,
        'grouped_invoice_form': grouped_estimate_form,
        'formset': formset,
        'customer_form': customer_form,
        'customers': customers,
        'jobsite_list': existing_jobsites,
        'truck_list': existing_trucks,
        'job_list': job_list,
        'job_catalog_json': job_catalog,
        'products': products,
        'product_data_json': product_data_json,
        'has_driver': has_driver,
        'drivers': drivers,
        'fleet_vehicles': fleet_vehicles,
        'tax_name': tax_name,
        'tax_rate': tax_rate,
        'tax_components_json': tax_components_json,
        'exemption_reasons': exemption_reasons,
        'documentType': 'estimate',  # used in your template switch
        'vehicle_form': VehicleForm(),
        'generated_invoice_number': GroupedEstimate.generate_estimate_number(request.user),
        'base_subtotal_amount': document_totals['base_subtotal'],
        'shop_supply_total_amount': document_totals['shop_supply_total'],
        'discount_total_amount': document_totals['discount_total'],
        'interest_total_amount': document_totals['interest_total'],
        'pre_tax_subtotal_amount': document_totals['pre_tax_subtotal'],
        'subtotal_after_interest_amount': document_totals['subtotal_after_interest'],
        'tax_total_amount': document_totals['tax_total'],
        'grand_total_amount': document_totals['grand_total'],
    }
    return render(request, template_name, context)


@activation_required
@subscription_required
@login_required
def edit_estimate(request, pk):
    estimate = get_object_or_404(GroupedEstimate, pk=pk, user=request.user)
    template_name = 'accounts/mach/daily_workslip.html'

    if request.method == 'POST':
        grouped_estimate_form = GroupedEstimateForm(request.POST, instance=estimate, user=request.user)
        formset = EstimateRecordFormSet(
            request.POST,
            instance=estimate,
            form_kwargs={'user': request.user}
        )

        if grouped_estimate_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    grouped_estimate = grouped_estimate_form.save(commit=False)
                    grouped_estimate.user = request.user

                    customer = grouped_estimate_form.cleaned_data.get('customer')
                    if customer:
                        grouped_estimate.bill_to = customer.name
                        grouped_estimate.bill_to_email = customer.email
                        grouped_estimate.bill_to_address = customer.address
                    else:
                        grouped_estimate.bill_to = grouped_estimate_form.cleaned_data.get('bill_to', grouped_estimate.bill_to)
                        grouped_estimate.bill_to_email = grouped_estimate_form.cleaned_data.get('bill_to_email', grouped_estimate.bill_to_email)
                        grouped_estimate.bill_to_address = grouped_estimate_form.cleaned_data.get('bill_to_address', grouped_estimate.bill_to_address)

                    grouped_estimate.save()
                    formset.save()

                    grouped_estimate.recalculate_total_amount()

                    messages.success(request, f"Estimate {grouped_estimate.estimate_number} updated successfully")
            except Exception as e:
                messages.error(request, f"An error occurred during update: {e}")
                logger.error(f"Error editing estimate (pk={pk}): {e}", exc_info=True)

            if request.POST.get('redirect_dashboard'):
                messages.success(request, f"Estimate {estimate.estimate_number} updated successfully")
                return redirect('accounts:estimate_list')
            if 'save_continue' in request.POST:
                return redirect(f"{reverse('accounts:add_dailylog')}?type=estimate")
            if 'submit_income' in request.POST:
                return redirect('accounts:estimate_list')
            return redirect('accounts:estimate_list')
        else:
            error_list = []
            if grouped_estimate_form.errors:
                error_list.extend([f"{field}: {error[0]}" for field, error in grouped_estimate_form.errors.items()])
            if formset.errors:
                for index, form_errors in enumerate(formset.errors):
                    if form_errors:
                        error_list.extend([f"Item {index + 1} - {field}: {err[0]}" for field, err in form_errors.items()])
            if formset.non_form_errors():
                error_list.extend(formset.non_form_errors())
            if error_list:
                messages.error(request, 'Please correct the errors below: ' + "; ".join(error_list))
            else:
                messages.error(request, 'Please correct the errors below.')
    else:
        grouped_estimate_form = GroupedEstimateForm(instance=estimate, user=request.user)
        formset = EstimateRecordFormSet(
            instance=estimate,
            form_kwargs={'user': request.user}
        )

    business_user_ids = get_customer_user_ids(request.user)
    customers = Customer.objects.filter(user__in=business_user_ids)
    existing_jobsites = EstimateRecord.objects.filter(grouped_estimate__user=request.user)\
                           .values_list('jobsite', flat=True).distinct()
    existing_trucks = EstimateRecord.objects.filter(grouped_estimate__user=request.user)\
                           .values_list('truck', flat=True).distinct()
    existing_jobs = EstimateRecord.objects.filter(grouped_estimate__user=request.user)\
                           .values_list('job', flat=True).distinct()
    job_values = list(existing_jobs.exclude(job__isnull=True).exclude(job__exact=''))
    job_catalog, service_description_strings, _job_name_choices = build_service_job_catalog(request.user)
    job_list = sorted(set(job_values) | set(service_description_strings))

    product_user_ids = get_product_user_ids(request.user)
    products = Product.objects.filter(user__in=product_user_ids)
    product_data = {
        str(product.id): {
            "name": product.name,
            "sku": product.sku or '',
            "description": product.description or '',
            "price": str(product.sale_price),
            "item_type": product.item_type,
        }
        for product in products
    }
    product_data_json = json.dumps(product_data)
    has_driver = Driver.objects.filter(user=request.user).exists()
    drivers = Driver.objects.filter(user=request.user) if has_driver else None
    fleet_vehicles = FleetVehicle.objects.filter(user=request.user)

    user_profile = getattr(request.user, 'profile', None)
    if user_profile:
        province = getattr(user_profile, 'province', None)
        tax_name = user_profile.get_tax_name() if hasattr(user_profile, 'get_tax_name') else "Tax"
    else:
        province = None
        tax_name = "Tax"
    tax_rate = PROVINCE_TAX_RATES.get(province, 0.00)
    tax_components_json = json.dumps(get_tax_component_rates(province))

    default_reasons = [
        "Indigenous Status",
        "Government Purchase",
        "Resale",
        "Export",
    ]
    saved_reasons = list(
        TaxExemptionReason.objects.filter(user=request.user).values_list('reason', flat=True)
    )
    exemption_reasons = list(dict.fromkeys(default_reasons + saved_reasons))

    resolved_vehicle = getattr(getattr(grouped_estimate_form, 'instance', None), 'vehicle', None)
    vehicle_history_payload = _serialize_vehicle_history(resolved_vehicle)
    vehicle_maintenance_payload = _serialize_vehicle_maintenance(resolved_vehicle)

    document_totals = _calculate_grouped_doc_totals(estimate.estimate_records.all())

    context = {
        'grouped_estimate_form': grouped_estimate_form,
        'grouped_invoice_form': grouped_estimate_form,
        'formset': formset,
        'customer_form': CustomerForm(user=request.user),
        'customers': customers,
        'jobsite_list': existing_jobsites,
        'truck_list': existing_trucks,
        'job_list': job_list,
        'job_catalog_json': job_catalog,
        'products': products,
        'product_data_json': product_data_json,
        'has_driver': has_driver,
        'drivers': drivers,
        'fleet_vehicles': fleet_vehicles,
        'tax_name': tax_name,
        'tax_rate': tax_rate,
        'tax_components_json': tax_components_json,
        'exemption_reasons': exemption_reasons,
        'documentType': 'estimate',
        'vehicle_form': VehicleForm(),
        'generated_invoice_number': estimate.estimate_number,
        'base_subtotal_amount': document_totals['base_subtotal'],
        'shop_supply_total_amount': document_totals['shop_supply_total'],
        'discount_total_amount': document_totals['discount_total'],
        'interest_total_amount': document_totals['interest_total'],
        'pre_tax_subtotal_amount': document_totals['pre_tax_subtotal'],
        'subtotal_after_interest_amount': document_totals['subtotal_after_interest'],
        'tax_total_amount': document_totals['tax_total'],
        'grand_total_amount': document_totals['grand_total'],
        'vehicle_history': vehicle_history_payload,
        'vehicle_history_json': json.dumps(vehicle_history_payload, cls=DjangoJSONEncoder),
        'vehicle_history_api': reverse('accounts:vehicle_history_summary'),
        'vehicle_maintenance': vehicle_maintenance_payload,
        'vehicle_maintenance_json': json.dumps(vehicle_maintenance_payload, cls=DjangoJSONEncoder),
        'vehicle_maintenance_api': reverse('accounts:vehicle_maintenance_summary'),
    }
    return render(request, template_name, context)


class GroupedEstimateDeleteView(LoginRequiredMixin, DeleteView):
    model = GroupedEstimate
    template_name = 'app/groupedestimate_confirm_delete.html'
    success_url = reverse_lazy('accounts:estimate_list')

    def get_queryset(self):
        return GroupedEstimate.objects.filter(user=self.request.user)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        estimate_number = self.object.estimate_number
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f"Estimate {estimate_number} deleted successfully")
        return response


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import ValidationError
from django.contrib.auth.decorators import login_required
from accounts.decorators import activation_required, subscription_required
from .models import GroupedEstimate, EstimateRecord, Product, Driver, Customer
from .forms import GroupedEstimateForm, EstimateRecordFormSet, CustomerForm
# Import the convert function from your models or define it here if needed.
# For example, if it's defined as a method on GroupedEstimate, you can call it directly.

class GroupedEstimateDetailView(LoginRequiredMixin, DetailView):
    model = GroupedEstimate
    template_name = 'app/groupedinvoice_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        estimate = self.get_object()

        records = estimate.estimate_records.all()
        total_qty = sum(record.qty or 0 for record in records)
        subtotal = sum(record.amount for record in records)
        tax = sum(record.tax_collected for record in records)
        total_amount = subtotal + tax

        context.update({
            'total_qty': total_qty,
            'subtotal': subtotal,
            'tax': tax,
            'total_amount': total_amount,
            'documentType': 'estimate',
            'show_tax': True,
            'stripe_invoice_pdf_url': None,
            'show_inperson_btn': False,
            'payment_link': None,
            'payment_provider': None,
        })

        profile = get_object_or_404(Profile, user=estimate.user)
        context['profile'] = profile
        context['note'] = profile.note if getattr(profile, 'show_note', False) else ''
        context['company_logo_url'] = resolve_company_logo_url(
            profile,
            request=self.request,
        )

        return context

@activation_required
@subscription_required
@login_required
def convert_estimate_to_invoice(request, estimate_id):
    """
    Converts the given estimate into a real invoice.
    If successful, redirects to the new invoice's detail view.
    """
    estimate = get_object_or_404(GroupedEstimate, id=estimate_id, user=request.user)
    try:
        with transaction.atomic():
            # Assumes the GroupedEstimate model has a convert_to_invoice() method.
            invoice = estimate.convert_to_invoice()
            messages.success(
                request,
                f"Estimate {estimate.estimate_number} converted to Invoice {invoice.invoice_number} successfully!"
            )
            # Redirect to the invoice detail view (adjust URL name and kwargs as needed)
            return redirect('accounts:groupedinvoice_detail', pk=invoice.id)
    except Exception as e:
        messages.error(request, f"Error converting estimate: {str(e)}")
        return redirect('accounts:estimate_detail', pk=estimate.id)

def product_low_stock_notifications(request):
    """
    View to get notifications as an HTML partial for AJAX loading.
    Fetches low stock product notifications and is structured for future notification types.
    """
    low_stock_products = apply_stock_fields(list(Product.get_low_stock_products(request.user)))

    expiring_warranty_products = Product.get_expiring_warranty_products(request.user, days_ahead=30)

    # Example: Overdue tasks (if you have a Task model and it's relevant here)
    # try:
    #     from tasks.models import Task # Assuming a 'tasks' app
    #     overdue_tasks = Task.objects.filter(
    #         assignee=request.user,
    #         due_date__lt=timezone.now(),
    #         completed=False
    #     ).order_by('due_date')
    # except ImportError:
    #     overdue_tasks = None

    context = {
        'low_stock_products': low_stock_products,
        'expiring_warranty_products': expiring_warranty_products,
        # 'overdue_tasks': overdue_tasks,
    }
    # Using the exact path you provided for the partial:
    return render(request, 'app/partials/notification_content.html', context)


@login_required
def analytics_overview(request):
    """
    Lightweight analytics cockpit surfaced inside the portal so owners can see
    tag status and embed their GA dashboards without leaving the app.
    """
    measurement_id = getattr(settings, 'GOOGLE_ANALYTICS_MEASUREMENT_ID', '').strip()
    context = {
        'analytics_enabled': bool(measurement_id),
        'google_analytics_measurement_id': measurement_id,
        'google_analytics_debug': getattr(settings, 'GOOGLE_ANALYTICS_DEBUG', False),
        'looker_studio_embed_url': getattr(settings, 'LOOKER_STUDIO_EMBED_URL', '').strip(),
        'tag_assistant_url': 'https://tagassistant.google.com/#/workspace',
        'site_url_for_testing': request.build_absolute_uri('/'),
    }
    return render(request, 'accounts/analytics_overview.html', context)


@login_required
def accountant_hub(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    portal_form = AccountantPortalForm(profile=profile)

    if request.method == 'POST':
        action = (request.POST.get("action") or "").strip()
        if action == "accountant_portal":
            portal_form = AccountantPortalForm(request.POST, profile=profile)
            if portal_form.is_valid():
                portal_form.save()
                messages.success(request, "Accountant portal access updated.")
                return redirect("accounts:accountant_hub")
            messages.error(request, "Please correct the accountant portal details below.")
        else:
            def clean_value(value):
                cleaned = (value or '').strip()
                return cleaned or None

            profile.accountant_name = clean_value(request.POST.get('accountant_name'))
            profile.accountant_firm = clean_value(request.POST.get('accountant_firm'))
            profile.accountant_email = clean_value(request.POST.get('accountant_email'))
            profile.accountant_phone = clean_value(request.POST.get('accountant_phone'))
            profile.accountant_timezone = clean_value(request.POST.get('accountant_timezone'))

            access_level = clean_value(request.POST.get('accountant_access_level'))
            if access_level:
                profile.accountant_access_level = access_level

            profile.save(update_fields=[
                'accountant_name',
                'accountant_firm',
                'accountant_email',
                'accountant_phone',
                'accountant_access_level',
                'accountant_timezone',
            ])
            messages.success(request, 'Accountant profile updated.')
            return redirect('accounts:accountant_hub')

    today = datetime.date.today()
    range_param = (request.GET.get('range') or '').strip()
    start_date_param = (request.GET.get('start_date') or '').strip()
    end_date_param = (request.GET.get('end_date') or '').strip()

    if range_param and not (start_date_param or end_date_param):
        if range_param == 'today':
            start_date_param = today.strftime('%Y-%m-%d')
            end_date_param = start_date_param
        elif range_param == 'yesterday':
            target = today - datetime.timedelta(days=1)
            start_date_param = target.strftime('%Y-%m-%d')
            end_date_param = start_date_param
        elif range_param == 'this_week':
            start_date = today - datetime.timedelta(days=today.weekday())
            end_date = start_date + datetime.timedelta(days=6)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = end_date.strftime('%Y-%m-%d')
        elif range_param == 'last_week':
            end_date = today - datetime.timedelta(days=today.weekday() + 1)
            start_date = end_date - datetime.timedelta(days=6)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = end_date.strftime('%Y-%m-%d')
        elif range_param == 'this_month':
            start_date = today.replace(day=1)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = today.strftime('%Y-%m-%d')
        elif range_param == 'last_month':
            first_day_this_month = today.replace(day=1)
            end_date = first_day_this_month - datetime.timedelta(days=1)
            start_date = end_date.replace(day=1)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = end_date.strftime('%Y-%m-%d')
        elif range_param == 'this_year':
            start_date = today.replace(month=1, day=1)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = today.strftime('%Y-%m-%d')
        elif range_param == 'last_year':
            start_date = today.replace(year=today.year - 1, month=1, day=1)
            end_date = today.replace(year=today.year - 1, month=12, day=31)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = end_date.strftime('%Y-%m-%d')

    year_param = request.GET.get('year', '')
    start_month_param = request.GET.get('start_month', '')
    end_month_param = request.GET.get('end_month', '')
    month_selected = request.GET.get('month', '')

    date_range_active = bool(start_date_param or end_date_param)
    if date_range_active:
        year_param = ''
        start_month_param = ''
        end_month_param = ''
        month_selected = ''

    if not date_range_active and not any([range_param, year_param, start_month_param, end_month_param, month_selected]):
        year_param = str(today.year)
        range_param = 'this_year'

    context = get_report_data(
        request,
        year_param=year_param,
        start_month_param=start_month_param,
        end_month_param=end_month_param,
        month_selected=month_selected,
        start_date_param=start_date_param,
        end_date_param=end_date_param,
    )

    quickbooks_settings = QuickBooksSettings.objects.filter(user=request.user).first()
    quickbooks_status_label = 'Not configured'
    if quickbooks_settings:
        if quickbooks_settings.is_configured:
            quickbooks_status_label = 'Connected' if quickbooks_settings.has_valid_access_token() else 'Action needed'
        else:
            quickbooks_status_label = 'Incomplete'

    filters_query = request.GET.copy()
    filters_query.pop('report_type', None)
    filters_query.pop('format', None)
    filters_querystring = filters_query.urlencode()

    context.update({
        'profile': profile,
        'accountant_access_choices': Profile.ACCOUNTANT_ACCESS_CHOICES,
        'quickbooks_settings': quickbooks_settings,
        'quickbooks_status_label': quickbooks_status_label,
        'filters_querystring': filters_querystring,
        'active_range': range_param,
        'selected_start_date': start_date_param,
        'selected_end_date': end_date_param,
        'selected_year': year_param,
        'selected_start_month': start_month_param,
        'selected_end_month': end_month_param,
        'month_selected': month_selected,
        'accountant_portal_form': portal_form,
        'accountant_portal_user': profile.accountant_portal_user,
        'accountant_portal_enabled': bool(getattr(profile.accountant_portal_user, 'is_active', False)),
        'accountant_portal_login_url': request.build_absolute_uri(
            reverse('accounts:accountant_portal_login')
        ),
    })

    return render(request, 'accounts/accountant_hub.html', context)


@accountant_login_required
def accountant_portal_dashboard(request):
    business_profile = request.user.accountant_portal
    business_user = business_profile.user

    today = datetime.date.today()
    range_param = (request.GET.get('range') or '').strip()
    start_date_param = (request.GET.get('start_date') or '').strip()
    end_date_param = (request.GET.get('end_date') or '').strip()

    if range_param and not (start_date_param or end_date_param):
        if range_param == 'today':
            start_date_param = today.strftime('%Y-%m-%d')
            end_date_param = start_date_param
        elif range_param == 'yesterday':
            target = today - datetime.timedelta(days=1)
            start_date_param = target.strftime('%Y-%m-%d')
            end_date_param = start_date_param
        elif range_param == 'this_week':
            start_date = today - datetime.timedelta(days=today.weekday())
            end_date = start_date + datetime.timedelta(days=6)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = end_date.strftime('%Y-%m-%d')
        elif range_param == 'last_week':
            end_date = today - datetime.timedelta(days=today.weekday() + 1)
            start_date = end_date - datetime.timedelta(days=6)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = end_date.strftime('%Y-%m-%d')
        elif range_param == 'this_month':
            start_date = today.replace(day=1)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = today.strftime('%Y-%m-%d')
        elif range_param == 'last_month':
            first_day_this_month = today.replace(day=1)
            end_date = first_day_this_month - datetime.timedelta(days=1)
            start_date = end_date.replace(day=1)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = end_date.strftime('%Y-%m-%d')
        elif range_param == 'this_year':
            start_date = today.replace(month=1, day=1)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = today.strftime('%Y-%m-%d')
        elif range_param == 'last_year':
            start_date = today.replace(year=today.year - 1, month=1, day=1)
            end_date = today.replace(year=today.year - 1, month=12, day=31)
            start_date_param = start_date.strftime('%Y-%m-%d')
            end_date_param = end_date.strftime('%Y-%m-%d')

    year_param = request.GET.get('year', '')
    start_month_param = request.GET.get('start_month', '')
    end_month_param = request.GET.get('end_month', '')
    month_selected = request.GET.get('month', '')

    date_range_active = bool(start_date_param or end_date_param)
    if date_range_active:
        year_param = ''
        start_month_param = ''
        end_month_param = ''
        month_selected = ''

    if not date_range_active and not any([range_param, year_param, start_month_param, end_month_param, month_selected]):
        year_param = str(today.year)
        range_param = 'this_year'

    context = get_report_data(
        request,
        year_param=year_param,
        start_month_param=start_month_param,
        end_month_param=end_month_param,
        month_selected=month_selected,
        start_date_param=start_date_param,
        end_date_param=end_date_param,
        report_user=business_user,
    )

    filters_query = request.GET.copy()
    filters_query.pop('report_type', None)
    filters_query.pop('format', None)
    filters_querystring = filters_query.urlencode()

    access_label = dict(Profile.ACCOUNTANT_ACCESS_CHOICES).get(
        business_profile.accountant_access_level,
        'Read only',
    )
    company_logo_url = resolve_company_logo_url(business_profile, request=request)
    business_name = (
        business_profile.company_name
        or business_user.get_full_name()
        or business_user.get_username()
    )
    accountant_display_name = (
        business_profile.accountant_name
        or request.user.get_full_name()
        or request.user.get_username()
    )
    detail_access = business_profile.accountant_access_level in ("full", "read_only")
    accountant_payroll_access = business_profile.accountant_access_level in ("full", "read_only")
    accountant_can_edit_payroll = business_profile.accountant_access_level == "full"

    context.update({
        'business_profile': business_profile,
        'business_user': business_user,
        'accountant_user': request.user,
        'accountant_display_name': accountant_display_name,
        'accountant_firm': business_profile.accountant_firm,
        'accountant_access_label': access_label,
        'accountant_timezone': business_profile.accountant_timezone,
        'filters_querystring': filters_querystring,
        'active_range': range_param,
        'selected_start_date': start_date_param,
        'selected_end_date': end_date_param,
        'selected_year': year_param,
        'selected_start_month': start_month_param,
        'selected_end_month': end_month_param,
        'month_selected': month_selected,
        'business_name': business_name,
        'company_logo_url': company_logo_url,
        'portal_support_email': business_profile.company_email or business_user.email,
        'portal_support_phone': business_profile.company_phone,
        'accountant_detail_access': detail_access,
        'accountant_payroll_access': accountant_payroll_access,
        'accountant_can_edit_payroll': accountant_can_edit_payroll,
    })

    return render(request, 'accountant/portal_dashboard.html', context)


@accountant_login_required
def accountant_portal_payroll_deductions(request):
    business_profile = request.user.accountant_portal
    business_user = business_profile.user
    payroll_access = business_profile.accountant_access_level in ("full", "read_only")
    if not payroll_access:
        messages.error(request, "Payroll deductions are available for read-only or full access accounts.")
        return redirect("accounts:accountant_portal_dashboard")

    accountant_access_label = business_profile.get_accountant_access_level_display() or "Read only"
    company_logo_url = resolve_company_logo_url(business_profile, request=request)
    business_name = (
        business_profile.company_name
        or business_user.get_full_name()
        or business_user.get_username()
    )
    accountant_display_name = (
        business_profile.accountant_name
        or request.user.get_full_name()
        or request.user.get_username()
    )
    accountant_can_edit_payroll = business_profile.accountant_access_level == "full"

    from .models import Employee

    employees = (
        Employee.objects.filter(user=business_user)
        .annotate(
            deduction_count=models.Count("recurring_deductions", distinct=True),
            active_deduction_count=models.Count(
                "recurring_deductions",
                filter=models.Q(recurring_deductions__active=True),
                distinct=True,
            ),
        )
        .order_by("first_name", "last_name")
    )

    context = {
        "business_profile": business_profile,
        "business_user": business_user,
        "accountant_user": request.user,
        "accountant_display_name": accountant_display_name,
        "accountant_firm": business_profile.accountant_firm,
        "accountant_access_label": accountant_access_label,
        "accountant_timezone": business_profile.accountant_timezone,
        "business_name": business_name,
        "company_logo_url": company_logo_url,
        "portal_support_email": business_profile.company_email or business_user.email,
        "portal_support_phone": business_profile.company_phone,
        "accountant_can_edit_payroll": accountant_can_edit_payroll,
        "employees": employees,
    }

    return render(request, "accountant/payroll_deductions.html", context)


@accountant_login_required
def accountant_portal_payroll_deductions_employee(request, employee_id):
    business_profile = request.user.accountant_portal
    business_user = business_profile.user
    payroll_access = business_profile.accountant_access_level in ("full", "read_only")
    if not payroll_access:
        messages.error(request, "Payroll deductions are available for read-only or full access accounts.")
        return redirect("accounts:accountant_portal_dashboard")

    accountant_access_label = business_profile.get_accountant_access_level_display() or "Read only"
    company_logo_url = resolve_company_logo_url(business_profile, request=request)
    business_name = (
        business_profile.company_name
        or business_user.get_full_name()
        or business_user.get_username()
    )
    accountant_display_name = (
        business_profile.accountant_name
        or request.user.get_full_name()
        or request.user.get_username()
    )
    accountant_can_edit_payroll = business_profile.accountant_access_level == "full"

    from .forms import EmployeeRecurringDeductionForm, EmployeeTaxProfileForm
    from .models import Employee, EmployeeRecurringDeduction, EmployeeTaxProfile

    employee = get_object_or_404(Employee, id=employee_id, user=business_user)
    tax_profile, _ = EmployeeTaxProfile.objects.get_or_create(
        employee=employee,
        defaults={"province": getattr(getattr(employee.user, "profile", None), "province", "ON")},
    )
    deductions = list(
        EmployeeRecurringDeduction.objects.filter(employee=employee).order_by("name", "id")
    )

    bound_deduction_form = None
    bound_deduction_id = None

    if request.method == "POST":
        if not accountant_can_edit_payroll:
            messages.error(request, "Your access level does not allow editing payroll deductions.")
            return redirect("accounts:accountant_portal_payroll_deductions_employee", employee_id=employee.id)

        if "save_tax_profile" in request.POST:
            tax_form = EmployeeTaxProfileForm(request.POST, instance=tax_profile)
            deduction_form = EmployeeRecurringDeductionForm(prefix="new_deduction")
            if tax_form.is_valid():
                tax_form.save()
                messages.success(request, "Tax profile updated.")
                return redirect("accounts:accountant_portal_payroll_deductions_employee", employee_id=employee.id)
            messages.error(request, "Please correct the errors below.")
        elif "add_deduction" in request.POST:
            tax_form = EmployeeTaxProfileForm(instance=tax_profile)
            deduction_form = EmployeeRecurringDeductionForm(request.POST, prefix="new_deduction")
            if deduction_form.is_valid():
                deduction = deduction_form.save(commit=False)
                deduction.employee = employee
                deduction.save()
                messages.success(request, "Deduction added.")
                return redirect("accounts:accountant_portal_payroll_deductions_employee", employee_id=employee.id)
            messages.error(request, "Please correct the errors below.")
        elif "update_deduction" in request.POST:
            tax_form = EmployeeTaxProfileForm(instance=tax_profile)
            deduction_form = EmployeeRecurringDeductionForm(prefix="new_deduction")
            deduction_id = request.POST.get("deduction_id")
            deduction = get_object_or_404(
                EmployeeRecurringDeduction,
                id=deduction_id,
                employee=employee,
            )
            prefix = f"deduction_{deduction.id}"
            bound_deduction_form = EmployeeRecurringDeductionForm(
                request.POST,
                instance=deduction,
                prefix=prefix,
            )
            bound_deduction_id = deduction.id
            if bound_deduction_form.is_valid():
                bound_deduction_form.save()
                messages.success(request, "Deduction updated.")
                return redirect("accounts:accountant_portal_payroll_deductions_employee", employee_id=employee.id)
            messages.error(request, "Please correct the errors below.")
        else:
            tax_form = EmployeeTaxProfileForm(instance=tax_profile)
            deduction_form = EmployeeRecurringDeductionForm(prefix="new_deduction")
    else:
        tax_form = EmployeeTaxProfileForm(instance=tax_profile)
        deduction_form = EmployeeRecurringDeductionForm(prefix="new_deduction")

    deduction_forms = []
    for deduction in deductions:
        prefix = f"deduction_{deduction.id}"
        if bound_deduction_form is not None and bound_deduction_id == deduction.id:
            form = bound_deduction_form
        else:
            form = EmployeeRecurringDeductionForm(instance=deduction, prefix=prefix)
        if not accountant_can_edit_payroll:
            for field in form.fields.values():
                field.disabled = True
        deduction_forms.append({"deduction": deduction, "form": form})

    if not accountant_can_edit_payroll:
        for field in tax_form.fields.values():
            field.disabled = True
        for field in deduction_form.fields.values():
            field.disabled = True

    context = {
        "business_profile": business_profile,
        "business_user": business_user,
        "accountant_user": request.user,
        "accountant_display_name": accountant_display_name,
        "accountant_firm": business_profile.accountant_firm,
        "accountant_access_label": accountant_access_label,
        "accountant_timezone": business_profile.accountant_timezone,
        "business_name": business_name,
        "company_logo_url": company_logo_url,
        "portal_support_email": business_profile.company_email or business_user.email,
        "portal_support_phone": business_profile.company_phone,
        "accountant_can_edit_payroll": accountant_can_edit_payroll,
        "employee": employee,
        "tax_form": tax_form,
        "deduction_form": deduction_form,
        "deduction_forms": deduction_forms,
        "deductions": deductions,
    }

    return render(request, "accountant/payroll_employee_deductions.html", context)


@accountant_login_required
def accountant_portal_download_report(request):
    business_profile = request.user.accountant_portal
    business_user = business_profile.user
    return download_report(request, report_user=business_user)


@login_required
@require_POST
def session_ping(request):
    """
    Lightweight endpoint to keep the user's session active.
    Used by the portal idle-timeout warning to prevent surprise logouts and cold-start timeouts.
    """
    # Touch the session so session backends refresh expiry
    try:
        request.session.modified = True
    except Exception:
        pass
    return JsonResponse({"ok": True})
