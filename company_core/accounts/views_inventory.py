from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.template.loader import render_to_string
from django.db import transaction
from django.db.models import Q, F, Sum, ExpressionWrapper, DecimalField, Max
from django.db.models.functions import Coalesce, TruncDate
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.urls import reverse  # for building URL for redirect
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.conf import settings
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import timedelta, datetime, date
import json
import qrcode
import os
import textwrap
import re
from io import BytesIO
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import (
    ActivityLog,
    Category,
    CategoryGroup,
    CategoryAttribute,
    CategoryAttributeOption,
    Customer,
    Supplier,
    Product,
    ProductAlternateSku,
    ProductAttributeValue,
    ProductStock,
    InventoryTransaction,
    InventoryLocation,
    Profile,
    ProductBrand,
    ProductModel,
    ProductVin,
    InventoryRoleAssignment,
    MarginGuardrailSetting,
    ReplenishmentRule,
    PurchaseOrder,
    PurchaseOrderItem,
    CycleCountSession,
    CycleCountEntry,
    ProductRMA,
    FleetVehicle,
    FleetPartList,
    FleetPartListItem,
    DispatchTicket,
    SupplierScorecardSnapshot,
)
from .utils import (
    apply_stock_fields,
    annotate_products_with_stock,
    get_business_user,
    get_product_user_ids,
    get_product_stock_user_ids,
    get_stock_owner,
    get_store_user_ids,
    upsert_product_stock,
)
from .forms import (
    CategoryForm,
    CategoryGroupForm,
    CategoryAttributeForm,
    SupplierForm,
    ProductForm,
    ProductAttributeForm,
    ProductInlineForm,
    ProductBrandForm,
    ProductModelForm,
    ProductVinForm,
    InventoryTransactionForm,
    InventoryLocationForm,
)
from .excel_formatting import apply_template_styling


PRODUCT_TEMPLATE_HEADERS = [
    "SKU",
    "OEM Part Number",
    "Barcode",
    "Alternate SKUs",
    "Name",
    "Description",
    "Fitment Notes",
    "Category",
    "Supplier",
    "Brand",
    "Model",
    "VIN",
    "Item Type",
    "Cost Price",
    "Sale Price",
    "Promotion Price",
    "Margin",
    "Quantity",
    "Reorder Level",
    "Max Stock Level",
    "Location",
    "Warranty Expiry Date",
    "Warranty Length (Days)",
    "Show on Storefront",
    "Featured",
]

OPTIONAL_PRODUCT_TEMPLATE_HEADERS = {
    "OEM Part Number",
    "Barcode",
    "Alternate SKUs",
    "Fitment Notes",
    "Max Stock Level",
}


SUPPLIER_TEMPLATE_HEADERS = [
    "Name",
    "Contact Person",
    "Email",
    "Phone Number",
    "Address",
]


CATEGORY_TEMPLATE_HEADERS = [
    "Name",
    "Description",
    "Group",
    "Parent Category",
    "Sort Order",
    "Active",
]


CATEGORY_GROUP_TEMPLATE_HEADERS = [
    "Name",
    "Description",
    "Sort Order",
    "Active",
]


BRAND_TEMPLATE_HEADERS = [
    "Name",
    "Description",
    "Sort Order",
    "Active",
]


MODEL_TEMPLATE_HEADERS = [
    "Name",
    "Brand",
    "Description",
    "Year Start",
    "Year End",
    "Sort Order",
    "Active",
]


VIN_TEMPLATE_HEADERS = [
    "VIN",
    "Description",
    "Sort Order",
    "Active",
]


LOCATION_TEMPLATE_HEADERS = [
    "Name",
]


def _is_ajax(request):
    return request.headers.get("x-requested-with") == "XMLHttpRequest"


def _form_errors_json(form):
    return {field: [str(err) for err in errs] for field, errs in form.errors.items()}


def _paginate_queryset(request, queryset, per_page=100):
    """Return a single page of objects and the current querystring sans page."""

    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    params = request.GET.copy()
    params.pop("page", None)
    return page_obj, params.urlencode()


def _get_inventory_user_ids(request):
    if not hasattr(request, "_inventory_user_ids"):
        request._inventory_user_ids = get_product_user_ids(request.user)
    return request._inventory_user_ids


def _get_inventory_stock_user_ids(request):
    if not hasattr(request, "_inventory_stock_user_ids"):
        request._inventory_stock_user_ids = get_product_stock_user_ids(request.user)
    return request._inventory_stock_user_ids


def _get_inventory_transaction_user_ids(request):
    if not hasattr(request, "_inventory_transaction_user_ids"):
        request._inventory_transaction_user_ids = get_store_user_ids(request.user)
    return request._inventory_transaction_user_ids


def _transaction_scope_filter(user_ids):
    return Q(user__in=user_ids) | Q(user__isnull=True, product__user__in=user_ids)


def _product_transaction_scope_filter(user_ids):
    return Q(transactions__user__in=user_ids) | Q(
        transactions__user__isnull=True,
        transactions__product__user__in=user_ids,
    )


def _safe_int(value, default=0, *, minimum=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    if minimum is not None and parsed < minimum:
        return minimum
    return parsed


def _safe_decimal(value, default=Decimal("0.00"), *, minimum=None):
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        parsed = default
    if minimum is not None and parsed < minimum:
        return minimum
    return parsed


def _get_inventory_business_user(request):
    if not hasattr(request, "_inventory_business_user"):
        request._inventory_business_user = get_business_user(request.user) or request.user
    return request._inventory_business_user


def _get_inventory_role(request):
    if not hasattr(request, "_inventory_role"):
        business_user = _get_inventory_business_user(request)
        request._inventory_role = InventoryRoleAssignment.resolve_role(business_user, request.user)
    return request._inventory_role


def _can_inventory(request, capability):
    role = _get_inventory_role(request)
    return InventoryRoleAssignment.role_allows(role, capability)


def _log_inventory_activity(request, *, action, object_type, description, object_id="", metadata=None):
    business_user = _get_inventory_business_user(request)
    ActivityLog.objects.create(
        business=business_user,
        actor=request.user,
        action=action,
        object_type=object_type,
        object_id=str(object_id or ""),
        description=description,
        metadata=metadata or {},
    )


def _calculate_margin_percent(cost_price, sale_price):
    if cost_price in (None, "") or sale_price in (None, ""):
        return None
    cost = _safe_decimal(cost_price, default=Decimal("0.00"))
    sale = _safe_decimal(sale_price, default=Decimal("0.00"))
    if cost <= Decimal("0.00"):
        return None
    return ((sale - cost) / cost * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _average_daily_usage(product, transaction_scope, *, lookback_days=60):
    lookback_days = max(int(lookback_days or 60), 1)
    start_date = timezone.now() - timedelta(days=lookback_days)
    sold_qty = (
        product.transactions.filter(
            transaction_scope,
            transaction_type="OUT",
            transaction_date__gte=start_date,
        ).aggregate(total=Sum("quantity")).get("total")
        or 0
    )
    avg_usage = Decimal(sold_qty) / Decimal(lookback_days)
    return sold_qty, avg_usage


def _recommended_reorder_for_product(product, *, transaction_scope, rule=None, lookback_days=60):
    stock_qty = int(getattr(product, "quantity_in_stock", 0) or 0)
    reorder_level = int(getattr(product, "reorder_level", 0) or 0)
    max_stock_level = int(getattr(product, "max_stock_level", 0) or 0)
    sold_qty, avg_usage = _average_daily_usage(product, transaction_scope, lookback_days=lookback_days)

    if rule:
        recommendation = rule.calculate_recommended_quantity(
            current_stock=stock_qty,
            avg_daily_usage=avg_usage,
            reorder_level=reorder_level,
            max_stock_level=max_stock_level,
        )
    else:
        target_stock = max(max_stock_level, reorder_level)
        recommendation = {
            "recommended_qty": max(target_stock - stock_qty, 0),
            "target_stock": target_stock,
            "projected_demand": Decimal("0.00"),
            "days_covered": 0,
        }

    recommendation["sold_qty"] = sold_qty
    recommendation["avg_daily_usage"] = avg_usage.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    recommendation["rule_applied"] = bool(rule)
    return recommendation


def _build_supplier_scorecards(*, business_user, since_date=None, period_days=90):
    if since_date is None:
        since_date = timezone.now() - timedelta(days=period_days)

    purchase_orders = (
        PurchaseOrder.objects.filter(user=business_user, created_at__gte=since_date, supplier__isnull=False)
        .select_related("supplier")
        .prefetch_related("items")
    )

    supplier_rows = {}
    for po in purchase_orders:
        supplier_id = po.supplier_id
        if supplier_id not in supplier_rows:
            supplier_rows[supplier_id] = {
                "supplier": po.supplier,
                "po_count": 0,
                "on_time_hits": 0,
                "on_time_checks": 0,
                "qty_ordered": 0,
                "qty_received": 0,
                "lead_time_total": Decimal("0.00"),
                "lead_time_count": 0,
            }
        row = supplier_rows[supplier_id]
        row["po_count"] += 1

        qty_ordered = sum((item.quantity_ordered or 0) for item in po.items.all())
        qty_received = sum((item.quantity_received or 0) for item in po.items.all())
        row["qty_ordered"] += qty_ordered
        row["qty_received"] += qty_received

        if po.expected_delivery_date and po.status in {"partially_received", "received"}:
            row["on_time_checks"] += 1
            received_date = po.updated_at.date()
            if received_date <= po.expected_delivery_date:
                row["on_time_hits"] += 1

        if po.status in {"partially_received", "received"}:
            lead_days = max((po.updated_at.date() - po.created_at.date()).days, 0)
            row["lead_time_total"] += Decimal(lead_days)
            row["lead_time_count"] += 1

    rma_counts = (
        ProductRMA.objects.filter(
            user=business_user,
            supplier__isnull=False,
            opened_at__gte=since_date,
        )
        .values("supplier_id")
        .annotate(total=Sum("quantity"))
    )
    rma_map = {row["supplier_id"]: row["total"] or 0 for row in rma_counts}

    scorecards = []
    for supplier_id, row in supplier_rows.items():
        qty_ordered = row["qty_ordered"] or 0
        qty_received = row["qty_received"] or 0

        on_time_rate = Decimal("100.00")
        if row["on_time_checks"]:
            on_time_rate = (Decimal(row["on_time_hits"]) / Decimal(row["on_time_checks"])) * Decimal("100")

        fill_rate = Decimal("0.00")
        if qty_ordered:
            fill_rate = (Decimal(qty_received) / Decimal(qty_ordered)) * Decimal("100")

        avg_lead_time = Decimal("0.00")
        if row["lead_time_count"]:
            avg_lead_time = row["lead_time_total"] / Decimal(row["lead_time_count"])

        rma_rate = Decimal("0.00")
        supplier_rma_qty = Decimal(rma_map.get(supplier_id, 0))
        if qty_ordered:
            rma_rate = (supplier_rma_qty / Decimal(qty_ordered)) * Decimal("100")

        weighted_score = (
            (on_time_rate * Decimal("0.40"))
            + (fill_rate * Decimal("0.40"))
            + (max(Decimal("0.00"), Decimal("100.00") - (rma_rate * Decimal("2.00"))) * Decimal("0.20"))
        )

        scorecards.append(
            {
                "supplier": row["supplier"],
                "po_count": row["po_count"],
                "on_time_rate": on_time_rate.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
                "fill_rate": fill_rate.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
                "avg_lead_time_days": avg_lead_time.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
                "rma_rate": rma_rate.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
                "weighted_score": weighted_score.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
            }
        )

    scorecards.sort(key=lambda entry: (-entry["weighted_score"], entry["supplier"].name if entry["supplier"] else ""))
    return scorecards


def _sync_alternate_skus(product, sku_list):
    if product is None or sku_list is None:
        return
    normalized = []
    seen = set()
    for sku in sku_list:
        if not sku:
            continue
        key = sku.casefold()
        if key in seen:
            continue
        seen.add(key)
        normalized.append(sku)

    main_sku = (product.sku or "").strip()
    if main_sku:
        main_key = main_sku.casefold()
        normalized = [sku for sku in normalized if sku.casefold() != main_key]

    existing = list(product.alternate_skus.all())
    existing_map = {alt.sku.casefold(): alt for alt in existing}
    desired_keys = {sku.casefold() for sku in normalized}

    for alt in existing:
        if alt.sku.casefold() not in desired_keys:
            alt.delete()

    for sku in normalized:
        key = sku.casefold()
        existing_alt = existing_map.get(key)
        if existing_alt:
            if existing_alt.sku != sku:
                existing_alt.sku = sku
                existing_alt.save(update_fields=["sku"])
            continue
        ProductAlternateSku.objects.create(
            product=product,
            sku=sku,
        )


def _parse_alternate_sku_input(value):
    if value in (None, ""):
        return []
    if isinstance(value, (list, tuple)):
        raw_values = value
    else:
        raw_values = re.split(r"[\n,;]+", str(value))

    parsed = []
    seen = set()
    for raw in raw_values:
        normalized = (str(raw) if raw is not None else "").strip()
        if not normalized:
            continue
        key = normalized.casefold()
        if key in seen:
            continue
        seen.add(key)
        parsed.append(normalized)
    return parsed


def _sync_product_attributes_from_payload(product, user, payload):
    if not product or not product.pk or not payload:
        return

    if hasattr(payload, "lists"):
        payload_items = []
        for key, values in payload.lists():
            if not key.startswith("attr_"):
                continue
            payload_items.append((key, values[-1] if values else ""))
    else:
        payload_items = [
            (key, value)
            for key, value in payload.items()
            if str(key).startswith("attr_")
        ]

    if not payload_items:
        return

    attribute_ids = []
    attribute_payload = {}
    for key, value in payload_items:
        _, _, raw_id = str(key).partition("_")
        try:
            attr_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        attribute_ids.append(attr_id)
        attribute_payload[attr_id] = value

    if not attribute_ids:
        return

    product_user = product.user
    attributes = {
        attribute.id: attribute
        for attribute in CategoryAttribute.objects.filter(
            user=product_user,
            id__in=attribute_ids,
            is_active=True,
        ).prefetch_related("options")
    }

    for attr_id, raw_value in attribute_payload.items():
        attribute = attributes.get(attr_id)
        if not attribute:
            continue

        existing = ProductAttributeValue.objects.filter(
            product=product,
            attribute=attribute,
        ).first()
        value = raw_value.strip() if isinstance(raw_value, str) else raw_value

        if attribute.attribute_type == "select":
            if value in ("", None):
                if existing:
                    existing.delete()
                continue
            try:
                option_id = int(value)
            except (TypeError, ValueError):
                if existing:
                    existing.delete()
                continue
            option = CategoryAttributeOption.objects.filter(
                attribute=attribute,
                id=option_id,
                is_active=True,
            ).first()
            if not option:
                if existing:
                    existing.delete()
                continue
            if not existing:
                existing = ProductAttributeValue(product=product, attribute=attribute)
            existing.option = option
            existing.value_text = ""
            existing.value_number = None
            existing.value_boolean = None
            existing.save()
            continue

        if attribute.attribute_type == "number":
            if value in ("", None):
                if existing:
                    existing.delete()
                continue
            try:
                numeric_value = Decimal(str(value))
            except (InvalidOperation, ValueError, TypeError):
                if existing:
                    existing.delete()
                continue
            if not existing:
                existing = ProductAttributeValue(product=product, attribute=attribute)
            existing.option = None
            existing.value_text = ""
            existing.value_number = numeric_value
            existing.value_boolean = None
            existing.save()
            continue

        if attribute.attribute_type == "boolean":
            if value in ("", None, "None", "unknown"):
                if existing:
                    existing.delete()
                continue
            normalized = str(value).strip().lower()
            if normalized in {"true", "1", "yes", "on"}:
                bool_value = True
            elif normalized in {"false", "0", "no", "off"}:
                bool_value = False
            else:
                if existing:
                    existing.delete()
                continue
            if not existing:
                existing = ProductAttributeValue(product=product, attribute=attribute)
            existing.option = None
            existing.value_text = ""
            existing.value_number = None
            existing.value_boolean = bool_value
            existing.save()
            continue

        text_value = str(value).strip() if value is not None else ""
        if not text_value:
            if existing:
                existing.delete()
            continue
        if not existing:
            existing = ProductAttributeValue(product=product, attribute=attribute)
        existing.option = None
        existing.value_text = text_value
        existing.value_number = None
        existing.value_boolean = None
        existing.save()


@login_required
def inventory_view(request):
    tab = request.GET.get("tab")
    if tab == "transactions":
        return redirect("accounts:inventory_transactions")
    if tab == "products":
        return redirect("accounts:inventory_products")
    if tab == "stock-orders":
        return redirect("accounts:inventory_stock_orders")
    if tab == "suppliers":
        return redirect("accounts:inventory_suppliers")
    if tab == "categories":
        return redirect("accounts:inventory_categories")
    if tab == "locations":
        return redirect("accounts:inventory_locations")
    if tab == "operations":
        return redirect("accounts:inventory_operations")
    return redirect("accounts:inventory_hub")

@login_required
def inventory_hub(request):
    user = request.user
    product_user_ids = _get_inventory_user_ids(request)
    transaction_user_ids = _get_inventory_transaction_user_ids(request)
    now = timezone.now()
    window_days = 30
    window_start_date = now.date() - timedelta(days=window_days - 1)
    recent_period_start = now - timedelta(days=window_days)

    products = (
        Product.objects.filter(user__in=product_user_ids)
        .select_related("category", "supplier")
    )
    products = annotate_products_with_stock(products, request.user)
    transactions = (
        InventoryTransaction.objects.filter(_transaction_scope_filter(transaction_user_ids))
        .select_related("product", "product__category")
    )

    price_expr = Coalesce("sale_price", "cost_price")
    value_expr = ExpressionWrapper(
        F("stock_quantity") * price_expr,
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )

    totals = products.aggregate(
        total_value=Sum(value_expr),
        total_qty=Sum("stock_quantity"),
    )
    total_inventory_value = totals.get("total_value") or Decimal("0.00")
    total_quantity = totals.get("total_qty") or 0

    low_stock_qs = products.filter(stock_quantity__lt=F("stock_reorder"))
    low_stock_count = low_stock_qs.count()
    out_of_stock_count = products.filter(stock_quantity__lte=0).count()

    top_selling = (
        products.annotate(
            qty_sold=Sum(
                "transactions__quantity",
                filter=(
                    Q(
                        transactions__transaction_type="OUT",
                        transactions__transaction_date__gte=recent_period_start,
                    )
                    & _product_transaction_scope_filter(transaction_user_ids)
                ),
            )
        )
        .filter(qty_sold__gt=0)
        .order_by("-qty_sold", "name")
        .first()
    )
    if top_selling:
        apply_stock_fields([top_selling])

    category_rows = (
        products.values("category__name")
        .annotate(
            value=Sum(value_expr),
            qty=Sum("stock_quantity"),
        )
        .order_by("-value")
    )
    category_breakdown = []
    for row in category_rows:
        category_breakdown.append(
            {
                "name": row["category__name"] or "Uncategorized",
                "value": row["value"] or Decimal("0.00"),
                "qty": row["qty"] or 0,
            }
        )

    movement_rows = (
        transactions.filter(transaction_date__date__gte=window_start_date)
        .annotate(day=TruncDate("transaction_date"))
        .values("day")
        .annotate(
            stock_in=Sum("quantity", filter=Q(transaction_type="IN")),
            stock_out=Sum("quantity", filter=Q(transaction_type="OUT")),
        )
    )
    movement_lookup = {
        row["day"]: {
            "stock_in": row["stock_in"] or 0,
            "stock_out": row["stock_out"] or 0,
        }
        for row in movement_rows
    }

    movement_labels = []
    stock_in_series = []
    stock_out_series = []
    for offset in range(window_days):
        day = window_start_date + timedelta(days=offset)
        movement_labels.append(str(day.day))
        day_data = movement_lookup.get(day, {"stock_in": 0, "stock_out": 0})
        stock_in_series.append(day_data.get("stock_in") or 0)
        stock_out_series.append(day_data.get("stock_out") or 0)

    stock_in_total = sum(stock_in_series)
    stock_out_total = sum(stock_out_series)
    net_change = stock_in_total - stock_out_total
    value_trend_percent = round((net_change / total_quantity) * 100, 2) if total_quantity else 0
    trend_direction = "up" if value_trend_percent >= 0 else "down"

    product_form = ProductForm(user=user)
    supplier_form = SupplierForm(user=user)
    category_form = CategoryForm(user=user)
    location_form = InventoryLocationForm(user=user)

    recent_transactions_qs = transactions.order_by("-transaction_date")[:6]
    recent_transactions = []
    for tx in recent_transactions_qs:
        unit_price = (
            tx.product.sale_price
            if tx.product.sale_price is not None
            else tx.product.cost_price
            or Decimal("0.00")
        )
        total_value = (unit_price or Decimal("0.00")) * Decimal(tx.quantity)
        recent_transactions.append(
            {
                "id": tx.id,
                "product": tx.product.name,
                "type": tx.get_transaction_type_display(),
                "type_code": tx.transaction_type,
                "quantity": tx.quantity,
                "date": tx.transaction_date,
                "value": total_value,
                "remarks": tx.remarks or "",
            }
        )

    category_total_value = sum((c["value"] for c in category_breakdown), Decimal("0.00"))
    if category_total_value:
        for entry in category_breakdown:
            entry["percent"] = float((entry["value"] / category_total_value) * Decimal("100"))
    else:
        for entry in category_breakdown:
            entry["percent"] = 0
    category_chart_labels = [c["name"] for c in category_breakdown]
    category_chart_values = [float(c["value"]) for c in category_breakdown]

    movement_chart = {
        "labels": movement_labels,
        "stock_in": stock_in_series,
        "stock_out": stock_out_series,
    }

    context = {
        "total_inventory_value": total_inventory_value,
        "total_quantity": total_quantity,
        "low_stock_count": low_stock_count,
        "out_of_stock_count": out_of_stock_count,
        "top_selling": top_selling,
        "category_breakdown": category_breakdown,
        "category_total_value": category_total_value,
        "category_chart_labels": category_chart_labels,
        "category_chart_values": category_chart_values,
        "movement_chart": movement_chart,
        "stock_in_total": stock_in_total,
        "stock_out_total": stock_out_total,
        "value_trend_percent": value_trend_percent,
        "trend_direction": trend_direction,
        "recent_transactions": recent_transactions,
        "window_days": window_days,
        "product_form": product_form,
        "supplier_form": supplier_form,
        "category_form": category_form,
        "location_form": location_form,
    }
    return render(request, "inventory/hub.html", context)

@login_required
def inventory_transactions_view(request):
    query = request.GET.get("q", "").strip()
    transaction_user_ids = _get_inventory_transaction_user_ids(request)
    transactions = (
        InventoryTransaction.objects.filter(_transaction_scope_filter(transaction_user_ids))
        .select_related("product", "product__supplier", "product__category")
        .order_by("-transaction_date")
    )

    if query:
        transactions = transactions.filter(
            Q(product__name__icontains=query)
            | Q(product__sku__icontains=query)
            | Q(product__oem_part_number__icontains=query)
            | Q(product__barcode_value__icontains=query)
            | Q(product__alternate_skus__sku__icontains=query)
            | Q(product__fitment_notes__icontains=query)
            | Q(remarks__icontains=query)
            | Q(product__category__name__icontains=query)
            | Q(product__supplier__name__icontains=query)
        ).distinct()

    transactions_page, querystring = _paginate_queryset(request, transactions, per_page=100)

    context = {
        "transactions_page": transactions_page,
        "query": query,
        "querystring": querystring,
    }
    return render(request, "inventory/transactions_list.html", context)

@login_required
def inventory_products_view(request):
    search_query = request.GET.get("q", "").strip()
    group_ids = _extract_ids(request.GET.getlist("group"))
    category_ids = _extract_ids(request.GET.getlist("category"))
    brand_ids = _extract_ids(request.GET.getlist("brand"))
    stock_filter = request.GET.get("stock", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    stock_user_ids = _get_inventory_stock_user_ids(request)
    products_qs = (
        Product.objects.filter(user__in=product_user_ids)
        .select_related("supplier", "category", "brand", "vehicle_model", "vin_number")
        .prefetch_related("alternate_skus")
        .order_by("name")
    )

    if search_query:
        products_qs = products_qs.filter(
            Q(name__icontains=search_query)
            | Q(sku__icontains=search_query)
            | Q(oem_part_number__icontains=search_query)
            | Q(barcode_value__icontains=search_query)
            | Q(alternate_skus__sku__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(fitment_notes__icontains=search_query)
            | Q(category__name__icontains=search_query)
            | Q(supplier__name__icontains=search_query)
            | Q(brand__name__icontains=search_query)
            | Q(vehicle_model__name__icontains=search_query)
            | Q(vin_number__vin__icontains=search_query)
            | Q(location__icontains=search_query)
        ).distinct()

    if group_ids:
        products_qs = products_qs.filter(category__group_id__in=group_ids)
    if category_ids:
        products_qs = products_qs.filter(category_id__in=category_ids)
    if brand_ids:
        products_qs = products_qs.filter(brand_id__in=brand_ids)

    products_qs = annotate_products_with_stock(products_qs, request.user)
    stock_metrics_qs = products_qs.filter(user_id__in=stock_user_ids)
    price_expr = Coalesce("sale_price", "cost_price")
    stock_value_expr = ExpressionWrapper(
        F("stock_quantity") * price_expr,
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    stock_totals = stock_metrics_qs.aggregate(
        total_stock_value=Sum(stock_value_expr),
        total_stock_units=Sum("stock_quantity"),
    )
    low_stock_count = stock_metrics_qs.filter(
        stock_reorder__gt=0,
        stock_quantity__lte=F("stock_reorder"),
        stock_quantity__gt=0,
    ).count()
    out_of_stock_count = stock_metrics_qs.filter(stock_quantity__lte=0).count()
    missing_sku_count = products_qs.filter(Q(sku__isnull=True) | Q(sku="")).count()
    missing_supplier_count = products_qs.filter(supplier__isnull=True).count()

    if stock_filter == "low":
        products_qs = products_qs.filter(
            user_id__in=stock_user_ids,
            stock_reorder__gt=0,
            stock_quantity__lte=F("stock_reorder"),
            stock_quantity__gt=0,
        )
    elif stock_filter == "out":
        products_qs = products_qs.filter(
            user_id__in=stock_user_ids,
            stock_quantity__lte=0,
        )
    elif stock_filter == "no_sku":
        products_qs = products_qs.filter(Q(sku__isnull=True) | Q(sku=""))
    elif stock_filter == "no_supplier":
        products_qs = products_qs.filter(supplier__isnull=True)
    else:
        stock_filter = ""

    products_page, querystring = _paginate_queryset(request, products_qs, per_page=100)
    apply_stock_fields(products_page.object_list)
    categories = Category.objects.filter(user__in=product_user_ids).order_by("name")
    category_groups = CategoryGroup.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    suppliers = Supplier.objects.filter(user__in=product_user_ids).order_by("name")
    brands = ProductBrand.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    models = ProductModel.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    vins = ProductVin.objects.filter(user__in=product_user_ids).order_by("sort_order", "vin")
    locations = InventoryLocation.objects.filter(user__in=product_user_ids).order_by("name")
    item_type_choices = Product._meta.get_field("item_type").choices
    attribute_types = CategoryAttribute._meta.get_field("attribute_type").choices
    return render(
        request,
        "inventory/products_list.html",
        {
            "products": products_page,
            "products_count": products_page.paginator.count,
            "categories": categories,
            "category_groups": category_groups,
            "suppliers": suppliers,
            "brands": brands,
            "models": models,
            "vins": vins,
            "locations": locations,
            "item_type_choices": item_type_choices,
            "attribute_types": attribute_types,
            "search_query": search_query,
            "selected_group_ids": [str(value) for value in group_ids],
            "selected_category_ids": [str(value) for value in category_ids],
            "selected_brand_ids": [str(value) for value in brand_ids],
            "selected_stock_filter": stock_filter,
            "low_stock_count": low_stock_count,
            "out_of_stock_count": out_of_stock_count,
            "missing_sku_count": missing_sku_count,
            "missing_supplier_count": missing_supplier_count,
            "total_stock_units": stock_totals.get("total_stock_units") or 0,
            "total_stock_value": stock_totals.get("total_stock_value") or Decimal("0.00"),
            "stock_user_ids": stock_user_ids,
            "querystring": querystring,
        },
    )

@login_required
def inventory_suppliers_view(request):
    search_query = request.GET.get("q", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    suppliers_qs = Supplier.objects.filter(user__in=product_user_ids).order_by("name")

    if search_query:
        suppliers_qs = suppliers_qs.filter(
            Q(name__icontains=search_query)
            | Q(contact_person__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(phone_number__icontains=search_query)
            | Q(address__icontains=search_query)
        )

    suppliers = list(suppliers_qs)
    return render(
        request,
        "inventory/suppliers_list.html",
        {
            "suppliers": suppliers,
            "suppliers_count": len(suppliers),
            "search_query": search_query,
        },
    )

@login_required
def inventory_categories_view(request):
    search_query = request.GET.get("q", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    categories_qs = (
        Category.objects.filter(user__in=product_user_ids)
        .select_related("group", "parent")
        .order_by("sort_order", "name")
    )

    if search_query:
        categories_qs = categories_qs.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )

    categories = list(categories_qs)
    category_groups = CategoryGroup.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    return render(
        request,
        "inventory/categories_list.html",
        {
            "categories": categories,
            "categories_count": len(categories),
            "search_query": search_query,
            "category_groups": category_groups,
        },
    )


@login_required
def inventory_category_groups_view(request):
    search_query = request.GET.get("q", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    groups_qs = CategoryGroup.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    if search_query:
        groups_qs = groups_qs.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )
    groups = list(groups_qs)
    return render(
        request,
        "inventory/category_groups_list.html",
        {
            "groups": groups,
            "groups_count": len(groups),
            "search_query": search_query,
        },
    )


@login_required
def inventory_attributes_view(request):
    search_query = request.GET.get("q", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    attributes_qs = (
        CategoryAttribute.objects.filter(user__in=product_user_ids)
        .select_related("category")
        .order_by("sort_order", "name")
    )
    if search_query:
        attributes_qs = attributes_qs.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(category__name__icontains=search_query)
            | Q(value_unit__icontains=search_query)
        )
    attributes = list(attributes_qs)
    for attribute in attributes:
        attribute.active_options = [
            option.value for option in attribute.options.all() if option.is_active
        ]
    categories = Category.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    attribute_types = CategoryAttribute._meta.get_field("attribute_type").choices
    return render(
        request,
        "inventory/attributes_list.html",
        {
            "attributes": attributes,
            "attributes_count": len(attributes),
            "categories": categories,
            "attribute_types": attribute_types,
            "search_query": search_query,
        },
    )


@login_required
def inventory_brands_view(request):
    search_query = request.GET.get("q", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    brands_qs = ProductBrand.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    if search_query:
        brands_qs = brands_qs.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )
    brands = list(brands_qs)
    return render(
        request,
        "inventory/brands_list.html",
        {
            "brands": brands,
            "brands_count": len(brands),
            "search_query": search_query,
        },
    )


@login_required
def inventory_models_view(request):
    search_query = request.GET.get("q", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    models_qs = (
        ProductModel.objects.filter(user__in=product_user_ids)
        .select_related("brand")
        .order_by("sort_order", "name")
    )
    if search_query:
        models_qs = models_qs.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(brand__name__icontains=search_query)
        )
    models = list(models_qs)
    brands = ProductBrand.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    return render(
        request,
        "inventory/models_list.html",
        {
            "models": models,
            "models_count": len(models),
            "brands": brands,
            "search_query": search_query,
        },
    )


@login_required
def inventory_vins_view(request):
    search_query = request.GET.get("q", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    vins_qs = ProductVin.objects.filter(user__in=product_user_ids).order_by("sort_order", "vin")
    if search_query:
        vins_qs = vins_qs.filter(
            Q(vin__icontains=search_query) | Q(description__icontains=search_query)
        )
    vins = list(vins_qs)
    return render(
        request,
        "inventory/vins_list.html",
        {
            "vins": vins,
            "vins_count": len(vins),
            "search_query": search_query,
        },
    )

@login_required
def inventory_locations_view(request):
    query = request.GET.get("q", "").strip()
    product_user_ids = _get_inventory_user_ids(request)
    locations = InventoryLocation.objects.filter(user__in=product_user_ids).order_by("name")

    if query:
        locations = locations.filter(name__icontains=query)

    locations_page, querystring = _paginate_queryset(request, locations, per_page=100)
    location_form = InventoryLocationForm(user=request.user)
    return render(
        request,
        "inventory/locations_list.html",
        {
            "locations_page": locations_page,
            "location_form": location_form,
            "query": query,
            "querystring": querystring,
        },
    )


##############################
# AJAX "get form" endpoints  #
##############################
@login_required
def get_transaction_form(request):
    transaction_id = request.GET.get("transaction_id")
    if not transaction_id:
        return HttpResponseBadRequest("Transaction ID not provided.")
    transaction_user_ids = _get_inventory_transaction_user_ids(request)
    transaction = get_object_or_404(
        InventoryTransaction.objects.filter(_transaction_scope_filter(transaction_user_ids)),
        id=transaction_id,
    )
    form = InventoryTransactionForm(user=request.user, instance=transaction)
    html = render_to_string(
        "inventory/partials/_transaction_form.html", {"form": form}, request=request
    )
    return JsonResponse({"html": html})


@login_required
def get_supplier_form(request):
    supplier_id = request.GET.get("supplier_id")
    if not supplier_id:
        return HttpResponseBadRequest("Supplier ID not provided.")
    product_user_ids = _get_inventory_user_ids(request)
    supplier = get_object_or_404(Supplier, id=supplier_id, user__in=product_user_ids)
    form = SupplierForm(user=request.user, instance=supplier)
    html = render_to_string(
        "inventory/partials/_supplier_form.html", {"form": form}, request=request
    )
    return JsonResponse({"html": html})


@login_required
def get_product_form(request):
    product_id = request.GET.get("product_id")
    if not product_id:
        return HttpResponseBadRequest("Product ID not provided.")
    product_user_ids = _get_inventory_user_ids(request)
    product = get_object_or_404(Product, id=product_id, user__in=product_user_ids)
    stock_user_ids = _get_inventory_stock_user_ids(request)
    stock_visible = product.user_id in stock_user_ids
    attributes_only = request.GET.get("attributes_only")
    context = {"form": None, "stock_visible": stock_visible}
    if attributes_only:
        form = ProductAttributeForm(user=request.user, instance=product)
        template_name = "inventory/partials/_product_attribute_form.html"
        context["attribute_types"] = CategoryAttribute._meta.get_field("attribute_type").choices
    else:
        form = ProductForm(user=request.user, instance=product)
        stock_owner = get_stock_owner(request.user)
        if stock_owner:
            stock_record = ProductStock.objects.filter(product=product, user=stock_owner).first()
            form.initial["quantity_in_stock"] = stock_record.quantity_in_stock if stock_record else 0
            form.initial["reorder_level"] = stock_record.reorder_level if stock_record else 0
            form.initial["max_stock_level"] = stock_record.max_stock_level if stock_record else 0
        template_name = "inventory/partials/_product_form.html"
    context["form"] = form
    html = render_to_string(template_name, context, request=request)
    return JsonResponse({"html": html})


@login_required
def get_attribute_fields(request):
    category_id = request.GET.get("category_id")
    product_id = request.GET.get("product_id")
    product = None
    if product_id:
        product_user_ids = _get_inventory_user_ids(request)
        product = get_object_or_404(Product, id=product_id, user__in=product_user_ids)
    form = ProductForm(user=request.user, instance=product, category_id=category_id)
    html = render_to_string(
        "inventory/partials/_attribute_fields.html", {"form": form}, request=request
    )
    return JsonResponse({"html": html})


@login_required
def get_category_form(request):
    category_id = request.GET.get("category_id")
    if not category_id:
        return HttpResponseBadRequest("Category ID not provided.")
    product_user_ids = _get_inventory_user_ids(request)
    category = get_object_or_404(Category, id=category_id, user__in=product_user_ids)
    form = CategoryForm(user=request.user, instance=category)
    html = render_to_string(
        "inventory/partials/_category_form.html", {"form": form}, request=request
    )
    return JsonResponse({"html": html})


@login_required
def get_location_form(request):
    location_id = request.GET.get("location_id")
    if not location_id:
        return HttpResponseBadRequest("Location ID not provided.")
    product_user_ids = _get_inventory_user_ids(request)
    location = get_object_or_404(InventoryLocation, id=location_id, user__in=product_user_ids)
    form = InventoryLocationForm(user=request.user, instance=location)
    html = render_to_string(
        "inventory/partials/_location_form.html", {"form": form}, request=request
    )
    return JsonResponse({"html": html})


##############################
# Transaction CRUD           #
##############################
@login_required
def add_transaction(request):
    if request.method == "POST":
        form = InventoryTransactionForm(user=request.user, data=request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            try:
                if not transaction.user_id:
                    transaction.user = request.user
                transaction.save()
                messages.success(request, "Transaction added successfully.")
            except ValidationError as e:
                messages.error(request, e.message)
        else:
            messages.error(
                request, "Error adding transaction. Please check form for errors."
            )
    # Redirect back to Transactions tab
    return redirect(reverse("accounts:inventory_transactions"))


@login_required
def edit_transaction(request):
    if request.method == "POST":
        transaction_id = request.POST.get("transaction_id")
        transaction_user_ids = _get_inventory_transaction_user_ids(request)
        transaction = get_object_or_404(
            InventoryTransaction.objects.filter(_transaction_scope_filter(transaction_user_ids)),
            id=transaction_id,
        )
        existing_user_id = transaction.user_id
        form = InventoryTransactionForm(
            user=request.user, data=request.POST, instance=transaction
        )
        if form.is_valid():
            try:
                transaction = form.save(commit=False)
                if not transaction.user_id:
                    transaction.user_id = existing_user_id or request.user.id
                transaction.save()
                messages.success(request, "Transaction updated successfully.")
            except ValidationError as e:
                messages.error(request, e.message)
        else:
            messages.error(
                request, "Error updating transaction. Please check form for errors."
            )
    # Return to Transactions tab
    return redirect(reverse("accounts:inventory_transactions"))


@login_required
def delete_transaction(request):
    if request.method == "POST":
        transaction_id = request.POST.get("transaction_id")
        transaction_user_ids = _get_inventory_transaction_user_ids(request)
        transaction = get_object_or_404(
            InventoryTransaction.objects.filter(_transaction_scope_filter(transaction_user_ids)),
            id=transaction_id,
        )
        transaction.delete()
        messages.success(request, "Transaction deleted successfully.")
    else:
        messages.error(request, "Invalid request method.")
    # Return to Transactions tab
    return redirect(reverse("accounts:inventory_transactions"))


##############################
# Supplier CRUD              #
##############################
@login_required
def add_supplier(request):
    next_url = request.POST.get("next") or request.GET.get("next") or (reverse("accounts:inventory_suppliers"))
    if request.method == "POST":
        form = SupplierForm(request.POST, user=request.user)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.user = request.user
            supplier.save()
            messages.success(request, "Supplier added successfully.")
            if _is_ajax(request):
                return JsonResponse(
                    {"success": True, "value": supplier.id, "label": supplier.name}
                )
        else:
            messages.error(
                request, "Error adding supplier. Please check the form for errors."
            )
            if _is_ajax(request):
                return JsonResponse(
                    {"success": False, "errors": _form_errors_json(form)}, status=400
                )
    # Return to Suppliers tab
    return redirect(next_url)


@login_required
def edit_supplier(request):
    if request.method == "POST":
        supplier_id = request.POST.get("supplier_id")
        product_user_ids = _get_inventory_user_ids(request)
        supplier = get_object_or_404(Supplier, id=supplier_id, user__in=product_user_ids)
        form = SupplierForm(request.POST, user=request.user, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier updated successfully.")
        else:
            messages.error(
                request, "Error updating supplier. Please check the form for errors."
            )
    # Return to Suppliers tab
    return redirect(reverse("accounts:inventory_suppliers"))


@login_required
def delete_supplier(request):
    if request.method == "POST":
        supplier_id = request.POST.get("supplier_id")
        product_user_ids = _get_inventory_user_ids(request)
        supplier = get_object_or_404(Supplier, id=supplier_id, user__in=product_user_ids)
        supplier.delete()
        messages.success(request, "Supplier deleted successfully.")
    else:
        messages.error(request, "Invalid request method.")
    # Return to Suppliers tab
    return redirect(reverse("accounts:inventory_suppliers"))


@login_required
@require_POST
def bulk_delete_suppliers(request):
    supplier_ids = _extract_ids(request.POST.getlist("supplier_ids"))
    if not supplier_ids:
        messages.warning(request, "No suppliers were selected for deletion.")
        return redirect(reverse("accounts:inventory_suppliers"))

    product_user_ids = _get_inventory_user_ids(request)
    queryset = Supplier.objects.filter(user__in=product_user_ids, id__in=supplier_ids)
    deleted_count = queryset.count()

    if not deleted_count:
        messages.warning(request, "No matching suppliers were found to delete.")
    else:
        queryset.delete()
        messages.success(request, f"Deleted {deleted_count} supplier(s).")

    return redirect(reverse("accounts:inventory_suppliers"))


@login_required
@require_POST
def save_supplier_inline(request):
    """Create or update a supplier via inline editing."""

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except (TypeError, ValueError, json.JSONDecodeError):
        return JsonResponse({"error": "Invalid data submitted."}, status=400)

    supplier_id_raw = payload.get("id")
    supplier_instance = None
    if supplier_id_raw:
        try:
            supplier_pk = int(supplier_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid supplier identifier provided."}, status=400)
        product_user_ids = _get_inventory_user_ids(request)
        supplier_instance = get_object_or_404(Supplier, pk=supplier_pk, user__in=product_user_ids)

    form_data = {
        "name": payload.get("name", ""),
        "contact_person": payload.get("contact_person", ""),
        "email": payload.get("email", ""),
        "phone_number": payload.get("phone_number", ""),
        "address": payload.get("address", ""),
    }
    form = SupplierForm(form_data, instance=supplier_instance, user=request.user)
    if not form.is_valid():
        return JsonResponse({"errors": form.errors}, status=400)

    supplier = form.save(commit=False)
    if not supplier.user_id:
        supplier.user = request.user
    supplier.save()

    response_data = {
        "supplier": {
            "id": supplier.pk,
            "name": supplier.name,
            "contact_person": supplier.contact_person or "",
            "email": supplier.email or "",
            "phone_number": supplier.phone_number or "",
            "address": supplier.address or "",
        }
    }
    status_code = 200 if supplier_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def delete_inventory_supplier(request, pk):
    product_user_ids = _get_inventory_user_ids(request)
    supplier = get_object_or_404(Supplier, pk=pk, user__in=product_user_ids)
    supplier.delete()

    if _is_ajax(request) or "application/json" in (request.headers.get("Accept", "") or ""):
        return JsonResponse({"deleted": True})

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_suppliers")
    messages.success(request, "Supplier deleted successfully.")
    return redirect(next_url)


@login_required
@require_POST
def save_category_inline(request):
    """Create or update a category via inline editing."""

    payload = None
    content_type = request.content_type or ""
    if "application/json" in content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)
    elif request.POST:
        payload = request.POST
    else:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)

    category_id_raw = payload.get("id")
    category_instance = None
    if category_id_raw:
        try:
            category_pk = int(category_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid category identifier provided."}, status=400)
        product_user_ids = _get_inventory_user_ids(request)
        category_instance = get_object_or_404(Category, pk=category_pk, user__in=product_user_ids)

    def _normalize(value, default=""):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip()

    def _normalize_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        normalized = str(value).strip().lower()
        if normalized in ("1", "true", "yes", "on"):
            return True
        if normalized in ("0", "false", "no", "off"):
            return False
        return default

    form_data = {
        "name": _normalize(payload.get("name", "")),
        "description": _normalize(payload.get("description", "")),
        "group": _normalize(payload.get("group_id", payload.get("group")), None),
        "parent": _normalize(payload.get("parent_id", payload.get("parent")), None),
        "sort_order": _normalize(payload.get("sort_order", 0), 0),
        "is_active": _normalize_bool(payload.get("is_active"), True),
    }
    form = CategoryForm(form_data, files=request.FILES, instance=category_instance, user=request.user)
    if not form.is_valid():
        return JsonResponse({"errors": form.errors}, status=400)

    category = form.save(commit=False)
    if not category.user_id:
        category.user = request.user
    category.save()

    response_data = {
        "category": {
            "id": category.pk,
            "name": category.name,
            "description": category.description or "",
            "group": {
                "id": category.group_id,
                "name": category.group.name if category.group else "",
            },
            "parent": {
                "id": category.parent_id,
                "name": category.parent.name if category.parent else "",
            },
            "sort_order": category.sort_order,
            "is_active": category.is_active,
            "image_url": category.image.url if category.image else "",
        }
    }
    status_code = 200 if category_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def delete_inventory_category(request, pk):
    product_user_ids = _get_inventory_user_ids(request)
    category = get_object_or_404(Category, pk=pk, user__in=product_user_ids)
    category.delete()

    if _is_ajax(request) or "application/json" in (request.headers.get("Accept", "") or ""):
        return JsonResponse({"deleted": True})

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_categories")
    messages.success(request, "Category deleted successfully.")
    return redirect(next_url)


@login_required
@require_POST
def save_category_group_inline(request):
    payload = None
    content_type = request.content_type or ""
    if "application/json" in content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)
    elif request.POST:
        payload = request.POST
    else:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)

    group_id_raw = payload.get("id")
    group_instance = None
    if group_id_raw:
        try:
            group_pk = int(group_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid group identifier provided."}, status=400)
        product_user_ids = _get_inventory_user_ids(request)
        group_instance = get_object_or_404(CategoryGroup, pk=group_pk, user__in=product_user_ids)

    def _normalize(value, default=""):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip()

    def _normalize_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        normalized = str(value).strip().lower()
        if normalized in ("1", "true", "yes", "on"):
            return True
        if normalized in ("0", "false", "no", "off"):
            return False
        return default

    form_data = {
        "name": _normalize(payload.get("name", "")),
        "description": _normalize(payload.get("description", "")),
        "sort_order": _normalize(payload.get("sort_order", 0), 0),
        "is_active": _normalize_bool(payload.get("is_active"), True),
    }
    form = CategoryGroupForm(form_data, files=request.FILES, instance=group_instance, user=request.user)
    if not form.is_valid():
        return JsonResponse({"errors": form.errors}, status=400)

    group = form.save(commit=False)
    if not group.user_id:
        group.user = request.user
    group.save()

    response_data = {
        "group": {
            "id": group.pk,
            "name": group.name,
            "description": group.description or "",
            "sort_order": group.sort_order,
            "is_active": group.is_active,
            "image_url": group.image.url if group.image else "",
        }
    }
    status_code = 200 if group_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def delete_category_group(request, pk):
    product_user_ids = _get_inventory_user_ids(request)
    group = get_object_or_404(CategoryGroup, pk=pk, user__in=product_user_ids)
    group.delete()

    if _is_ajax(request) or "application/json" in (request.headers.get("Accept", "") or ""):
        return JsonResponse({"deleted": True})

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_category_groups")
    messages.success(request, "Category group deleted successfully.")
    return redirect(next_url)


@login_required
@require_POST
def save_attribute_inline(request):
    payload = None
    content_type = request.content_type or ""
    if "application/json" in content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)
    elif request.POST:
        payload = request.POST
    else:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)

    attribute_id_raw = payload.get("id")
    attribute_instance = None
    if attribute_id_raw:
        try:
            attribute_pk = int(attribute_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid attribute identifier provided."}, status=400)
        product_user_ids = _get_inventory_user_ids(request)
        attribute_instance = get_object_or_404(CategoryAttribute, pk=attribute_pk, user__in=product_user_ids)

    def _normalize(value, default=""):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip()

    def _normalize_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        normalized = str(value).strip().lower()
        if normalized in ("1", "true", "yes", "on"):
            return True
        if normalized in ("0", "false", "no", "off"):
            return False
        return default

    form_data = {
        "name": _normalize(payload.get("name", "")),
        "description": _normalize(payload.get("description", "")),
        "value_unit": _normalize(payload.get("value_unit", "")),
        "category": _normalize(payload.get("category_id", payload.get("category")), None),
        "attribute_type": _normalize(payload.get("attribute_type", "select")),
        "is_filterable": _normalize_bool(payload.get("is_filterable"), True),
        "is_comparable": _normalize_bool(payload.get("is_comparable"), False),
        "sort_order": _normalize(payload.get("sort_order", 0), 0),
        "is_active": _normalize_bool(payload.get("is_active"), True),
    }
    form = CategoryAttributeForm(form_data, instance=attribute_instance, user=request.user)
    if not form.is_valid():
        return JsonResponse({"errors": form.errors}, status=400)

    attribute = form.save(commit=False)
    if not attribute.user_id:
        attribute.user = request.user
    attribute.save()

    options_raw = payload.get("options", "") or ""
    option_values = [opt.strip() for opt in re.split(r"[\n,]+", str(options_raw)) if opt.strip()]

    if attribute.attribute_type == "select":
        existing_options = {option.value.lower(): option for option in attribute.options.all()}
        seen_keys = set()
        for idx, value in enumerate(option_values):
            key = value.lower()
            seen_keys.add(key)
            option = existing_options.get(key)
            if option:
                option.value = value
                option.sort_order = idx
                option.is_active = True
                option.save()
            else:
                CategoryAttributeOption.objects.create(
                    attribute=attribute,
                    value=value,
                    sort_order=idx,
                    is_active=True,
                )
        for key, option in existing_options.items():
            if key not in seen_keys and option.is_active:
                option.is_active = False
                option.save(update_fields=["is_active"])
    else:
        attribute.options.filter(is_active=True).update(is_active=False)

    response_data = {
        "attribute": {
            "id": attribute.pk,
            "name": attribute.name,
            "description": attribute.description or "",
            "value_unit": attribute.value_unit or "",
            "category": {
                "id": attribute.category_id,
                "name": attribute.category.name if attribute.category else "",
            },
            "attribute_type": attribute.attribute_type,
            "is_filterable": attribute.is_filterable,
            "is_comparable": attribute.is_comparable,
            "sort_order": attribute.sort_order,
            "is_active": attribute.is_active,
            "options": [option.value for option in attribute.options.filter(is_active=True).order_by("sort_order", "value")],
        }
    }
    status_code = 200 if attribute_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def delete_attribute(request, pk):
    product_user_ids = _get_inventory_user_ids(request)
    attribute = get_object_or_404(CategoryAttribute, pk=pk, user__in=product_user_ids)
    attribute.delete()

    if _is_ajax(request) or "application/json" in (request.headers.get("Accept", "") or ""):
        return JsonResponse({"deleted": True})

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_attributes")
    messages.success(request, "Attribute deleted successfully.")
    return redirect(next_url)


@login_required
@require_POST
def save_brand_inline(request):
    payload = None
    content_type = request.content_type or ""
    if "application/json" in content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)
    elif request.POST:
        payload = request.POST
    else:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)

    brand_id_raw = payload.get("id")
    brand_instance = None
    if brand_id_raw:
        try:
            brand_pk = int(brand_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid brand identifier provided."}, status=400)
        product_user_ids = _get_inventory_user_ids(request)
        brand_instance = get_object_or_404(ProductBrand, pk=brand_pk, user__in=product_user_ids)

    def _normalize(value, default=""):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip()

    def _normalize_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        normalized = str(value).strip().lower()
        if normalized in ("1", "true", "yes", "on"):
            return True
        if normalized in ("0", "false", "no", "off"):
            return False
        return default

    form_data = {
        "name": _normalize(payload.get("name", "")),
        "description": _normalize(payload.get("description", "")),
        "sort_order": _normalize(payload.get("sort_order", 0), 0),
        "is_active": _normalize_bool(payload.get("is_active"), True),
    }
    form = ProductBrandForm(form_data, files=request.FILES, instance=brand_instance, user=request.user)
    if not form.is_valid():
        return JsonResponse({"errors": form.errors}, status=400)

    brand = form.save(commit=False)
    if not brand.user_id:
        brand.user = request.user
    brand.save()

    response_data = {
        "brand": {
            "id": brand.pk,
            "name": brand.name,
            "description": brand.description or "",
            "sort_order": brand.sort_order,
            "is_active": brand.is_active,
            "logo_url": brand.logo.url if brand.logo else "",
        }
    }
    status_code = 200 if brand_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def delete_brand(request, pk):
    product_user_ids = _get_inventory_user_ids(request)
    brand = get_object_or_404(ProductBrand, pk=pk, user__in=product_user_ids)
    brand.delete()

    if _is_ajax(request) or "application/json" in (request.headers.get("Accept", "") or ""):
        return JsonResponse({"deleted": True})

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_brands")
    messages.success(request, "Brand deleted successfully.")
    return redirect(next_url)


@login_required
@require_POST
def save_model_inline(request):
    payload = None
    content_type = request.content_type or ""
    if "application/json" in content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)
    elif request.POST:
        payload = request.POST
    else:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)

    model_id_raw = payload.get("id")
    model_instance = None
    if model_id_raw:
        try:
            model_pk = int(model_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid model identifier provided."}, status=400)
        product_user_ids = _get_inventory_user_ids(request)
        model_instance = get_object_or_404(ProductModel, pk=model_pk, user__in=product_user_ids)

    def _normalize(value, default=""):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip()

    def _normalize_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        normalized = str(value).strip().lower()
        if normalized in ("1", "true", "yes", "on"):
            return True
        if normalized in ("0", "false", "no", "off"):
            return False
        return default

    form_data = {
        "name": _normalize(payload.get("name", "")),
        "description": _normalize(payload.get("description", "")),
        "brand": _normalize(payload.get("brand_id", payload.get("brand")), None),
        "year_start": _normalize(payload.get("year_start", "")),
        "year_end": _normalize(payload.get("year_end", "")),
        "sort_order": _normalize(payload.get("sort_order", 0), 0),
        "is_active": _normalize_bool(payload.get("is_active"), True),
    }
    form = ProductModelForm(form_data, instance=model_instance, user=request.user)
    if not form.is_valid():
        return JsonResponse({"errors": form.errors}, status=400)

    model = form.save(commit=False)
    if not model.user_id:
        model.user = request.user
    model.save()

    response_data = {
        "model": {
            "id": model.pk,
            "name": model.name,
            "description": model.description or "",
            "brand": {
                "id": model.brand_id,
                "name": model.brand.name if model.brand else "",
            },
            "year_start": model.year_start,
            "year_end": model.year_end,
            "sort_order": model.sort_order,
            "is_active": model.is_active,
        }
    }
    status_code = 200 if model_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def delete_model(request, pk):
    product_user_ids = _get_inventory_user_ids(request)
    model = get_object_or_404(ProductModel, pk=pk, user__in=product_user_ids)
    model.delete()

    if _is_ajax(request) or "application/json" in (request.headers.get("Accept", "") or ""):
        return JsonResponse({"deleted": True})

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_models")
    messages.success(request, "Model deleted successfully.")
    return redirect(next_url)


@login_required
@require_POST
def save_vin_inline(request):
    payload = None
    content_type = request.content_type or ""
    if "application/json" in content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)
    elif request.POST:
        payload = request.POST
    else:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)

    vin_id_raw = payload.get("id")
    vin_instance = None
    if vin_id_raw:
        try:
            vin_pk = int(vin_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid VIN identifier provided."}, status=400)
        product_user_ids = _get_inventory_user_ids(request)
        vin_instance = get_object_or_404(ProductVin, pk=vin_pk, user__in=product_user_ids)

    def _normalize(value, default=""):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip()

    def _normalize_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        normalized = str(value).strip().lower()
        if normalized in ("1", "true", "yes", "on"):
            return True
        if normalized in ("0", "false", "no", "off"):
            return False
        return default

    form_data = {
        "vin": _normalize(payload.get("vin", "")),
        "description": _normalize(payload.get("description", "")),
        "sort_order": _normalize(payload.get("sort_order", 0), 0),
        "is_active": _normalize_bool(payload.get("is_active"), True),
    }
    form = ProductVinForm(form_data, instance=vin_instance, user=request.user)
    if not form.is_valid():
        return JsonResponse({"errors": form.errors}, status=400)

    vin = form.save(commit=False)
    if not vin.user_id:
        vin.user = request.user
    vin.save()

    response_data = {
        "vin": {
            "id": vin.pk,
            "vin": vin.vin,
            "description": vin.description or "",
            "sort_order": vin.sort_order,
            "is_active": vin.is_active,
        }
    }
    status_code = 200 if vin_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def delete_vin(request, pk):
    product_user_ids = _get_inventory_user_ids(request)
    vin = get_object_or_404(ProductVin, pk=pk, user__in=product_user_ids)
    vin.delete()

    if _is_ajax(request) or "application/json" in (request.headers.get("Accept", "") or ""):
        return JsonResponse({"deleted": True})

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_vins")
    messages.success(request, "VIN deleted successfully.")
    return redirect(next_url)


##############################
# Product CRUD               #
##############################
def _serialize_product(product, *, stock_user_ids=None, stock_owner=None):
    stock_visible = True
    quantity_in_stock = product.quantity_in_stock
    reorder_level = product.reorder_level
    max_stock_level = product.max_stock_level
    if stock_owner:
        stock_record = ProductStock.objects.filter(product=product, user=stock_owner).first()
        quantity_in_stock = stock_record.quantity_in_stock if stock_record else 0
        reorder_level = stock_record.reorder_level if stock_record else 0
        max_stock_level = stock_record.max_stock_level if stock_record else 0
    elif stock_user_ids is not None:
        stock_visible = product.user_id in stock_user_ids
        if not stock_visible:
            quantity_in_stock = ""
            reorder_level = ""
            max_stock_level = ""
    return {
        "id": product.pk,
        "name": product.name,
        "sku": product.sku or "",
        "oem_part_number": product.oem_part_number or "",
        "barcode_value": product.barcode_value or "",
        "alternate_skus": list(
            product.alternate_skus.order_by("kind", "sku").values_list("sku", flat=True)
        ),
        "description": product.description or "",
        "fitment_notes": product.fitment_notes or "",
        "category": {
            "id": product.category_id,
            "name": product.category.name if product.category else "",
        },
        "supplier": {
            "id": product.supplier_id,
            "name": product.supplier.name if product.supplier else "",
        },
        "brand": {
            "id": product.brand_id,
            "name": product.brand.name if product.brand else "",
        },
        "vehicle_model": {
            "id": product.vehicle_model_id,
            "name": product.vehicle_model.name if product.vehicle_model else "",
        },
        "vin_number": {
            "id": product.vin_number_id,
            "vin": product.vin_number.vin if product.vin_number else "",
        },
        "cost_price": str(product.cost_price) if product.cost_price is not None else "",
        "sale_price": str(product.sale_price) if product.sale_price is not None else "",
        "core_price": str(product.core_price) if product.core_price is not None else "",
        "environmental_fee": (
            str(product.environmental_fee) if product.environmental_fee is not None else ""
        ),
        "quantity_in_stock": quantity_in_stock,
        "reorder_level": reorder_level,
        "max_stock_level": max_stock_level,
        "stock_visible": stock_visible,
        "item_type": product.item_type,
        "location": product.location or "",
        "image_url": product.image.url if product.image else "",
    }


@login_required
@require_POST
def save_product_inline(request):
    """Create or update a product via inline editing."""

    payload = None
    content_type = request.content_type or ""
    if "application/json" in content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)
    elif request.POST:
        payload = request.POST
    elif request.body:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"error": "Invalid data submitted."}, status=400)
    else:
        payload = {}

    def _normalize(value, default=""):
        if value is None:
            return default
        return str(value).strip() if isinstance(value, str) else str(value)

    def _normalize_quantity(value):
        if value in (None, ""):
            return "0"
        return _normalize(value)

    product_id_raw = payload.get("id")
    product_instance = None
    if product_id_raw:
        try:
            product_pk = int(product_id_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid product identifier provided."}, status=400)
        product_user_ids = _get_inventory_user_ids(request)
        product_instance = get_object_or_404(Product, pk=product_pk, user__in=product_user_ids)
    stock_owner = get_stock_owner(request.user)
    original_quantity = product_instance.quantity_in_stock if product_instance else None
    original_reorder = product_instance.reorder_level if product_instance else None
    original_max_stock = product_instance.max_stock_level if product_instance else None

    item_type = (payload.get("item_type") or "inventory").strip()
    item_type_choices = dict(Product._meta.get_field("item_type").choices)
    if item_type not in item_type_choices:
        item_type = "inventory"

    quantity_raw = payload.get("quantity_in_stock")
    reorder_raw = payload.get("reorder_level")
    max_stock_raw = payload.get("max_stock_level")
    form_data = {
        "name": _normalize(payload.get("name")),
        "sku": _normalize(payload.get("sku")),
        "oem_part_number": _normalize(payload.get("oem_part_number")),
        "barcode_value": _normalize(payload.get("barcode_value")),
        "alternate_skus": payload.get("alternate_skus", ""),
        "description": _normalize(payload.get("description")),
        "fitment_notes": _normalize(payload.get("fitment_notes")),
        "category": _normalize(payload.get("category_id")),
        "supplier": _normalize(payload.get("supplier_id")),
        "brand": _normalize(payload.get("brand_id", payload.get("brand"))),
        "vehicle_model": _normalize(payload.get("vehicle_model_id", payload.get("vehicle_model"))),
        "vin_number": _normalize(payload.get("vin_number_id", payload.get("vin_number"))),
        "cost_price": _normalize(payload.get("cost_price")),
        "sale_price": _normalize(payload.get("sale_price")),
        "core_price": _normalize(payload.get("core_price")),
        "environmental_fee": _normalize(payload.get("environmental_fee")),
        "quantity_in_stock": _normalize_quantity(quantity_raw),
        "reorder_level": _normalize_quantity(reorder_raw),
        "max_stock_level": _normalize_quantity(max_stock_raw),
        "item_type": item_type,
        "location": _normalize(payload.get("location")),
    }

    form = ProductInlineForm(
        form_data,
        files=request.FILES,
        instance=product_instance,
        user=request.user,
    )
    if not form.is_valid():
        return JsonResponse({"errors": form.errors}, status=400)

    product = form.save(commit=False)
    if not product.user_id:
        product.user = stock_owner or request.user
    if product_instance and stock_owner and product_instance.user_id != stock_owner.id:
        product.quantity_in_stock = original_quantity or 0
        product.reorder_level = original_reorder or 0
        product.max_stock_level = original_max_stock or 0
    product.save()
    _sync_alternate_skus(product, form.cleaned_data.get("alternate_skus"))

    stock_quantity = form.cleaned_data.get("quantity_in_stock") or 0
    stock_reorder = form.cleaned_data.get("reorder_level") or 0
    stock_max = form.cleaned_data.get("max_stock_level") or 0
    if form.cleaned_data.get("item_type") != "inventory":
        stock_quantity = 0
        stock_reorder = 0
        stock_max = 0
    upsert_product_stock(
        product,
        request.user,
        quantity_in_stock=stock_quantity,
        reorder_level=stock_reorder,
        max_stock_level=stock_max,
    )
    _sync_product_attributes_from_payload(product, request.user, payload)

    response_data = {
        "product": _serialize_product(
            product,
            stock_user_ids=_get_inventory_stock_user_ids(request),
            stock_owner=stock_owner,
        )
    }
    status_code = 200 if product_instance else 201
    return JsonResponse(response_data, status=status_code)


@login_required
@require_POST
def bulk_update_products(request):
    product_ids = _extract_ids(request.POST.getlist("product_ids"))
    if not product_ids:
        return JsonResponse({"error": "Select at least one product to update."}, status=400)

    stock_owner = get_stock_owner(request.user)
    products = list(
        Product.objects.filter(
            user__in=_get_inventory_user_ids(request),
            id__in=product_ids,
        ).select_related(
            "category",
            "supplier",
            "brand",
            "vehicle_model",
            "vin_number",
        )
    )
    if not products:
        return JsonResponse({"error": "No matching products found."}, status=404)

    found_ids = {product.id for product in products}
    missing_ids = [pid for pid in product_ids if pid not in found_ids]
    if missing_ids:
        return JsonResponse({"error": "Some selected products were not found."}, status=404)

    field_map = {
        "category_id": "category",
        "supplier_id": "supplier",
        "brand_id": "brand",
        "vehicle_model_id": "vehicle_model",
        "vin_number_id": "vin_number",
        "item_type": "item_type",
        "location": "location",
        "oem_part_number": "oem_part_number",
        "barcode_value": "barcode_value",
        "description": "description",
        "fitment_notes": "fitment_notes",
        "cost_price": "cost_price",
        "sale_price": "sale_price",
        "core_price": "core_price",
        "environmental_fee": "environmental_fee",
        "quantity_in_stock": "quantity_in_stock",
        "reorder_level": "reorder_level",
        "max_stock_level": "max_stock_level",
    }

    updates = {}
    for request_key, field_name in field_map.items():
        if request_key not in request.POST:
            continue
        value = request.POST.get(request_key)
        if value == "__keep__":
            continue
        updates[field_name] = value

    if "item_type" in updates:
        item_type = updates["item_type"]
        valid_item_types = dict(Product._meta.get_field("item_type").choices)
        if item_type not in valid_item_types:
            return JsonResponse({"error": "Invalid item type selected."}, status=400)

    image_file = request.FILES.get("image")
    if not updates and not image_file:
        return JsonResponse({"error": "Choose at least one field to update."}, status=400)

    def _product_form_data(product):
        return {
            "name": product.name or "",
            "sku": product.sku or "",
            "oem_part_number": product.oem_part_number or "",
            "description": product.description or "",
            "fitment_notes": product.fitment_notes or "",
            "barcode_value": product.barcode_value or "",
            "category": product.category_id or "",
            "supplier": product.supplier_id or "",
            "brand": product.brand_id or "",
            "vehicle_model": product.vehicle_model_id or "",
            "vin_number": product.vin_number_id or "",
            "cost_price": str(product.cost_price) if product.cost_price is not None else "",
            "sale_price": str(product.sale_price) if product.sale_price is not None else "",
            "core_price": str(product.core_price) if product.core_price is not None else "",
            "environmental_fee": (
                str(product.environmental_fee) if product.environmental_fee is not None else ""
            ),
            "quantity_in_stock": (
                str(product.quantity_in_stock) if product.quantity_in_stock is not None else "0"
            ),
            "reorder_level": str(product.reorder_level) if product.reorder_level is not None else "0",
            "max_stock_level": (
                str(product.max_stock_level) if product.max_stock_level is not None else "0"
            ),
            "item_type": product.item_type or "inventory",
            "location": product.location or "",
        }

    errors = {}
    updated_products = []
    try:
        with transaction.atomic():
            for product in products:
                base_data = _product_form_data(product)
                form_data = base_data.copy()
                form_data.update(updates)
                files = None
                if image_file:
                    image_file.seek(0)
                    files = {"image": image_file}
                form = ProductInlineForm(
                    form_data,
                    files=files,
                    instance=product,
                    user=request.user,
                )
                if not form.is_valid():
                    errors[product.pk] = form.errors
                    raise ValidationError("Bulk update failed.")
                updated_product = form.save(commit=False)
                if not updated_product.user_id:
                    updated_product.user = stock_owner or request.user
                if stock_owner and product.user_id != stock_owner.id:
                    updated_product.quantity_in_stock = product.quantity_in_stock or 0
                    updated_product.reorder_level = product.reorder_level or 0
                    updated_product.max_stock_level = product.max_stock_level or 0
                updated_product.save()
                updated_products.append(updated_product)
                stock_quantity = (
                    form.cleaned_data.get("quantity_in_stock")
                    if "quantity_in_stock" in updates
                    else None
                )
                stock_reorder = (
                    form.cleaned_data.get("reorder_level")
                    if "reorder_level" in updates
                    else None
                )
                stock_max = (
                    form.cleaned_data.get("max_stock_level")
                    if "max_stock_level" in updates
                    else None
                )
                if form.cleaned_data.get("item_type") != "inventory":
                    if "quantity_in_stock" in updates:
                        stock_quantity = 0
                    if "reorder_level" in updates:
                        stock_reorder = 0
                    if "max_stock_level" in updates:
                        stock_max = 0
                upsert_product_stock(
                    updated_product,
                    request.user,
                    quantity_in_stock=stock_quantity,
                    reorder_level=stock_reorder,
                    max_stock_level=stock_max,
                )
    except ValidationError:
        error_payload = errors or {"__all__": ["Could not update selected products."]}
        return JsonResponse({"errors": error_payload}, status=400)

    return JsonResponse(
        {
            "products": [
                _serialize_product(
                    product,
                    stock_user_ids=_get_inventory_stock_user_ids(request),
                    stock_owner=stock_owner,
                )
                for product in updated_products
            ]
        }
    )

@login_required
def add_product(request):
    next_url = request.POST.get("next") or request.GET.get("next") or (reverse("accounts:inventory_products"))
    if request.method == "POST":
        form = ProductForm(user=request.user, data=request.POST, files=request.FILES)
        if form.is_valid():
            stock_owner = get_stock_owner(request.user)
            product = form.save(commit=False)
            product.user = stock_owner or request.user
            product.save()
            form.save_attribute_values(product)
            _sync_alternate_skus(product, form.cleaned_data.get("alternate_skus"))
            stock_quantity = form.cleaned_data.get("quantity_in_stock") or 0
            stock_reorder = form.cleaned_data.get("reorder_level") or 0
            stock_max = form.cleaned_data.get("max_stock_level") or 0
            if form.cleaned_data.get("item_type") != "inventory":
                stock_quantity = 0
                stock_reorder = 0
                stock_max = 0
            upsert_product_stock(
                product,
                request.user,
                quantity_in_stock=stock_quantity,
                reorder_level=stock_reorder,
                max_stock_level=stock_max,
            )
            messages.success(request, "Product added successfully.")
        else:
            messages.error(
                request, "Error adding product. Please check the form for errors."
            )
    # Return to Products tab
    return redirect(next_url)


@login_required
@require_POST
def update_inventory_margin(request):
    profile = getattr(request.user, "profile", None)
    if not profile:
        messages.error(request, "No profile found for the current user.")
        return redirect(reverse("accounts:inventory_products"))

    margin_input = request.POST.get("default_margin_percent", "").strip()
    if margin_input == "":
        messages.error(request, "Enter a margin percentage before saving.")
        return redirect(reverse("accounts:inventory_products"))

    try:
        margin_value = Decimal(margin_input)
    except (InvalidOperation, TypeError):
        messages.error(request, "Invalid margin percentage.")
        return redirect(reverse("accounts:inventory_products"))

    if margin_value < Decimal("0"):
        messages.error(request, "Margin percentage cannot be negative.")
        return redirect(reverse("accounts:inventory_products"))

    margin_value = margin_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    profile.default_inventory_margin_percent = margin_value
    profile.save(update_fields=["default_inventory_margin_percent"])
    messages.success(request, "Default margin updated successfully.")
    return redirect(reverse("accounts:inventory_products"))


@login_required
@require_POST
def apply_margin_to_products(request):
    profile = getattr(request.user, "profile", None)
    if not profile:
        messages.error(request, "No profile found for the current user.")
        return redirect(reverse("accounts:inventory_products"))

    margin_percent = profile.default_inventory_margin_percent
    if margin_percent is None:
        messages.error(request, "Set a default margin before running the update.")
        return redirect(reverse("accounts:inventory_products"))

    updated_count = 0
    products = Product.objects.filter(
        user__in=_get_inventory_user_ids(request)
    ).exclude(cost_price__isnull=True)
    for product in products:
        product.sale_price = None
        product.margin = None
        product.save()
        updated_count += 1

    messages.success(
        request,
        f"Updated sale prices for {updated_count} product(s) using the {margin_percent}% margin.",
    )
    return redirect(reverse("accounts:inventory_products"))


@login_required
def edit_product(request):
    if request.method == "POST":
        product_id = request.POST.get("product_id")
        product = get_object_or_404(
            Product,
            id=product_id,
            user__in=_get_inventory_user_ids(request),
        )
        stock_owner = get_stock_owner(request.user)
        original_quantity = product.quantity_in_stock
        original_reorder = product.reorder_level
        original_max_stock = product.max_stock_level
        form = ProductForm(user=request.user, data=request.POST, files=request.FILES, instance=product)
        if form.is_valid():
            product = form.save(commit=False)
            if stock_owner and product.user_id != stock_owner.id:
                product.quantity_in_stock = original_quantity or 0
                product.reorder_level = original_reorder or 0
                product.max_stock_level = original_max_stock or 0
            product.save()
            form.save_attribute_values(product)
            _sync_alternate_skus(product, form.cleaned_data.get("alternate_skus"))
            stock_quantity = form.cleaned_data.get("quantity_in_stock") or 0
            stock_reorder = form.cleaned_data.get("reorder_level") or 0
            stock_max = form.cleaned_data.get("max_stock_level") or 0
            if form.cleaned_data.get("item_type") != "inventory":
                stock_quantity = 0
                stock_reorder = 0
                stock_max = 0
            upsert_product_stock(
                product,
                request.user,
                quantity_in_stock=stock_quantity,
                reorder_level=stock_reorder,
                max_stock_level=stock_max,
            )
            messages.success(request, "Product updated successfully.")
        else:
            error_details = []
            for field_name, errors in form.errors.items():
                if field_name == "__all__":
                    label = "General"
                else:
                    label = form.fields.get(field_name).label or field_name.replace("_", " ").title()

                for error in errors:
                    error_details.append(f"{label}: {error}")

            if not error_details:
                error_details.append("Please check the form for errors.")

            detailed_message = "Error updating product. " + "; ".join(error_details)
            messages.error(request, detailed_message)
    # Return to Products tab
    return redirect(reverse("accounts:inventory_products"))


@login_required
@require_POST
def update_product_attributes(request):
    product_id = request.POST.get("product_id")
    if not product_id:
        messages.error(request, "Product ID not provided.")
        return redirect(reverse("accounts:inventory_products"))

    product = get_object_or_404(
        Product,
        id=product_id,
        user__in=_get_inventory_user_ids(request),
    )
    form = ProductAttributeForm(user=request.user, data=request.POST, instance=product)
    if form.is_valid():
        product = form.save()
        form.save_attribute_values(product)
        messages.success(request, "Product attributes updated successfully.")
    else:
        error_details = []
        for field_name, errors in form.errors.items():
            if field_name == "__all__":
                label = "General"
            else:
                label = form.fields.get(field_name).label or field_name.replace("_", " ").title()

            for error in errors:
                error_details.append(f"{label}: {error}")

        if not error_details:
            error_details.append("Please check the form for errors.")

        detailed_message = "Error updating product attributes. " + "; ".join(error_details)
        messages.error(request, detailed_message)

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_products")
    return redirect(next_url)


@login_required
def delete_product(request):
    if request.method == "POST":
        product_id = request.POST.get("product_id")
        product = get_object_or_404(
            Product,
            id=product_id,
            user__in=_get_inventory_user_ids(request),
        )
        product.delete()
        messages.success(request, "Product deleted successfully.")
    else:
        messages.error(request, "Invalid request method.")
    # Return to Products tab
    return redirect(next_url)


@login_required
@require_POST
def delete_inventory_product(request, pk):
    product = get_object_or_404(
        Product,
        pk=pk,
        user__in=_get_inventory_user_ids(request),
    )
    product.delete()

    if _is_ajax(request) or "application/json" in (request.headers.get("Accept", "") or ""):
        return JsonResponse({"deleted": True})

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_products")
    messages.success(request, "Product deleted successfully.")
    return redirect(next_url)


@login_required
@require_POST
def bulk_delete_products(request):
    product_ids = _extract_ids(request.POST.getlist("product_ids"))
    if not product_ids:
        messages.warning(request, "No products were selected for deletion.")
        return redirect(reverse("accounts:inventory_products"))

    queryset = Product.objects.filter(
        user__in=_get_inventory_user_ids(request),
        id__in=product_ids,
    )
    deleted_count = queryset.count()

    if not deleted_count:
        messages.warning(request, "No matching products were found to delete.")
    else:
        queryset.delete()
        messages.success(request, f"Deleted {deleted_count} product(s).")

    return redirect(reverse("accounts:inventory_products"))


@login_required
def export_products_template(request):
    """Download an Excel template populated with the user's current products."""

    product_user_ids = _get_inventory_user_ids(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"

    instruction_text = (
        "Instructions: Enter warranty expiry dates using the YYYY-MM-DD format. "
        "Leave 'Warranty Length (Days)' blank so the system calculates remaining "
        "warranty days automatically. Provide both cost and sale prices or include "
        "the margin with one of the prices so missing values can be derived. Use "
        "Item Type values of Inventory or Non-inventory. Use Yes/No for Show on "
        "Storefront and Featured. Use commas to separate Alternate SKUs."
    )
    instruction_row = [instruction_text] + [""] * (len(PRODUCT_TEMPLATE_HEADERS) - 1)
    worksheet.append(instruction_row)
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(PRODUCT_TEMPLATE_HEADERS),
    )
    worksheet.append(PRODUCT_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A3"

    item_type_labels = dict(Product._meta.get_field("item_type").choices)
    products = (
        annotate_products_with_stock(
            Product.objects.filter(user__in=product_user_ids).order_by("name"),
            request.user,
        )
        .prefetch_related("alternate_skus")
    )
    for product in products:
        quantity_value = product.stock_quantity
        reorder_value = product.stock_reorder
        max_value = product.stock_max
        if product.margin is not None:
            margin_value = str(product.margin)
        elif product.cost_price is not None and product.sale_price is not None:
            margin_value = str(product.sale_price - product.cost_price)
        else:
            margin_value = ""
        alternate_skus_value = ", ".join(
            product.alternate_skus.order_by("kind", "sku").values_list("sku", flat=True)
        )
        item_type_label = item_type_labels.get(product.item_type, product.item_type)
        worksheet.append(
            [
                product.sku or "",
                product.oem_part_number or "",
                product.barcode_value or "",
                alternate_skus_value,
                product.name or "",
                product.description or "",
                product.fitment_notes or "",
                product.category.name if product.category else "",
                product.supplier.name if product.supplier else "",
                product.brand.name if product.brand else "",
                product.vehicle_model.name if product.vehicle_model else "",
                product.vin_number.vin if product.vin_number else "",
                item_type_label or "",
                str(product.cost_price) if product.cost_price is not None else "",
                str(product.sale_price) if product.sale_price is not None else "",
                str(product.promotion_price) if product.promotion_price is not None else "",
                margin_value,
                quantity_value,
                reorder_value,
                max_value,
                product.location or "",
                product.warranty_expiry_date.isoformat()
                if product.warranty_expiry_date
                else "",
                product.warranty_length or "",
                "Yes" if product.is_published_to_store else "No",
                "Yes" if product.is_featured else "No",
            ]
        )

    categories = list(
        dict.fromkeys(
            Category.objects.filter(user__in=product_user_ids)
            .order_by("name")
            .values_list("name", flat=True)
        )
    )
    suppliers = list(
        dict.fromkeys(
            Supplier.objects.filter(user__in=product_user_ids)
            .order_by("name")
            .values_list("name", flat=True)
        )
    )
    brands = list(
        dict.fromkeys(
            ProductBrand.objects.filter(user__in=product_user_ids)
            .order_by("sort_order", "name")
            .values_list("name", flat=True)
        )
    )
    models = list(
        dict.fromkeys(
            ProductModel.objects.filter(user__in=product_user_ids)
            .order_by("sort_order", "name")
            .values_list("name", flat=True)
        )
    )
    vins = list(
        dict.fromkeys(
            ProductVin.objects.filter(user__in=product_user_ids)
            .order_by("sort_order", "vin")
            .values_list("vin", flat=True)
        )
    )

    product_locations = (
        Product.objects.filter(user__in=product_user_ids)
        .exclude(location__isnull=True)
        .exclude(location__exact="")
        .values_list("location", flat=True)
    )
    location_sources = list(
        InventoryLocation.objects.filter(user__in=product_user_ids)
        .order_by("name")
        .values_list("name", flat=True)
    )
    cleaned_product_locations = [loc.strip() for loc in product_locations if loc and loc.strip()]
    locations = list(dict.fromkeys(location_sources + cleaned_product_locations))
    item_types = [label for _value, label in Product._meta.get_field("item_type").choices]
    yes_no = ["Yes", "No"]

    options_sheet = workbook.create_sheet(title="Options")
    options_sheet.sheet_state = "hidden"

    def _write_option_column(column_index, header, values):
        options_sheet.cell(row=1, column=column_index, value=header)
        for row_index, value in enumerate(values, start=2):
            options_sheet.cell(row=row_index, column=column_index, value=value)
        last_row = max(2, len(values) + 1)
        column_letter = get_column_letter(column_index)
        return f"=Options!${column_letter}$2:${column_letter}${last_row}"

    category_formula = _write_option_column(1, "Categories", categories)
    supplier_formula = _write_option_column(2, "Suppliers", suppliers)
    location_formula = _write_option_column(3, "Locations", locations)
    brand_formula = _write_option_column(4, "Brands", brands)
    model_formula = _write_option_column(5, "Models", models)
    vin_formula = _write_option_column(6, "VINs", vins)
    item_type_formula = _write_option_column(7, "Item Types", item_types)
    yes_no_formula = _write_option_column(8, "Yes/No", yes_no)

    def _apply_validation(target_header, formula):
        target_column_index = PRODUCT_TEMPLATE_HEADERS.index(target_header) + 1
        target_column_letter = get_column_letter(target_column_index)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{target_column_letter}2:{target_column_letter}1048576")

    _apply_validation("Category", category_formula)
    _apply_validation("Supplier", supplier_formula)
    _apply_validation("Location", location_formula)
    _apply_validation("Brand", brand_formula)
    _apply_validation("Model", model_formula)
    _apply_validation("VIN", vin_formula)
    _apply_validation("Item Type", item_type_formula)
    _apply_validation("Show on Storefront", yes_no_formula)
    _apply_validation("Featured", yes_no_formula)

    apply_template_styling(
        worksheet,
        headers=PRODUCT_TEMPLATE_HEADERS,
        header_row_index=2,
        instruction_row_index=1,
        column_width_overrides={
            "SKU": 18,
            "OEM Part Number": 20,
            "Barcode": 22,
            "Alternate SKUs": 26,
            "Name": 30,
            "Description": 48,
            "Fitment Notes": 38,
            "Category": 24,
            "Supplier": 24,
            "Brand": 20,
            "Model": 20,
            "VIN": 20,
            "Item Type": 16,
            "Cost Price": 16,
            "Sale Price": 16,
            "Promotion Price": 18,
            "Margin": 14,
            "Quantity": 14,
            "Reorder Level": 18,
            "Max Stock Level": 18,
            "Location": 22,
            "Warranty Expiry Date": 24,
            "Warranty Length (Days)": 26,
            "Show on Storefront": 20,
            "Featured": 14,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="products_template.xlsx"'
    return response


@login_required
def export_suppliers_template(request):
    """Download an Excel template populated with the user's current suppliers."""

    product_user_ids = _get_inventory_user_ids(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Suppliers"
    worksheet.append(SUPPLIER_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    suppliers = Supplier.objects.filter(user__in=product_user_ids).order_by("name")
    for supplier in suppliers:
        worksheet.append(
            [
                supplier.name or "",
                supplier.contact_person or "",
                supplier.email or "",
                supplier.phone_number or "",
                supplier.address or "",
            ]
        )

    apply_template_styling(
        worksheet,
        headers=SUPPLIER_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={
            "Name": 32,
            "Contact Person": 28,
            "Email": 36,
            "Phone Number": 22,
            "Address": 42,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="suppliers_template.xlsx"'
    return response


@login_required
def export_categories_template(request):
    """Download an Excel template populated with the user's current categories."""

    product_user_ids = _get_inventory_user_ids(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Categories"
    worksheet.append(CATEGORY_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    categories = Category.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    for category in categories:
        worksheet.append([
            category.name or "",
            category.description or "",
            category.group.name if category.group else "",
            category.parent.name if category.parent else "",
            category.sort_order,
            "Active" if category.is_active else "Inactive",
        ])

    groups = list(
        dict.fromkeys(
            CategoryGroup.objects.filter(user__in=product_user_ids)
            .order_by("sort_order", "name")
            .values_list("name", flat=True)
        )
    )
    parent_categories = list(
        dict.fromkeys(
            categories.values_list("name", flat=True)
        )
    )
    active_states = ["Active", "Inactive"]

    options_sheet = workbook.create_sheet(title="Options")
    options_sheet.sheet_state = "hidden"

    def _write_option_column(column_index, header, values):
        options_sheet.cell(row=1, column=column_index, value=header)
        for row_index, value in enumerate(values, start=2):
            options_sheet.cell(row=row_index, column=column_index, value=value)
        last_row = max(2, len(values) + 1)
        column_letter = get_column_letter(column_index)
        return f"=Options!${column_letter}$2:${column_letter}${last_row}"

    group_formula = _write_option_column(1, "Groups", groups)
    parent_formula = _write_option_column(2, "Parent Categories", parent_categories)
    active_formula = _write_option_column(3, "Active", active_states)

    def _apply_validation(target_header, formula):
        target_column_index = CATEGORY_TEMPLATE_HEADERS.index(target_header) + 1
        target_column_letter = get_column_letter(target_column_index)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{target_column_letter}2:{target_column_letter}1048576")

    _apply_validation("Group", group_formula)
    _apply_validation("Parent Category", parent_formula)
    _apply_validation("Active", active_formula)

    apply_template_styling(
        worksheet,
        headers=CATEGORY_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={
            "Name": 32,
            "Description": 48,
            "Group": 24,
            "Parent Category": 24,
            "Sort Order": 14,
            "Active": 12,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="categories_template.xlsx"'
    return response


@login_required
def export_category_groups_template(request):
    """Download an Excel template populated with the user's current category groups."""

    product_user_ids = _get_inventory_user_ids(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Category Groups"
    worksheet.append(CATEGORY_GROUP_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    groups = CategoryGroup.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    for group in groups:
        worksheet.append([
            group.name or "",
            group.description or "",
            group.sort_order,
            "Active" if group.is_active else "Inactive",
        ])

    active_states = ["Active", "Inactive"]
    options_sheet = workbook.create_sheet(title="Options")
    options_sheet.sheet_state = "hidden"

    def _write_option_column(column_index, header, values):
        options_sheet.cell(row=1, column=column_index, value=header)
        for row_index, value in enumerate(values, start=2):
            options_sheet.cell(row=row_index, column=column_index, value=value)
        last_row = max(2, len(values) + 1)
        column_letter = get_column_letter(column_index)
        return f"=Options!${column_letter}$2:${column_letter}${last_row}"

    active_formula = _write_option_column(1, "Active", active_states)

    def _apply_validation(target_header, formula):
        target_column_index = CATEGORY_GROUP_TEMPLATE_HEADERS.index(target_header) + 1
        target_column_letter = get_column_letter(target_column_index)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{target_column_letter}2:{target_column_letter}1048576")

    _apply_validation("Active", active_formula)

    apply_template_styling(
        worksheet,
        headers=CATEGORY_GROUP_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={
            "Name": 32,
            "Description": 48,
            "Sort Order": 14,
            "Active": 12,
        },
    )


@login_required
def inventory_stock_orders_view(request):
    product_user_ids = _get_inventory_user_ids(request)
    transaction_scope = _transaction_scope_filter(_get_inventory_transaction_user_ids(request))
    business_user = _get_inventory_business_user(request)
    products = (
        Product.objects.filter(user__in=product_user_ids, item_type="inventory")
        .select_related("supplier")
        .order_by("supplier__name", "name")
    )
    products = annotate_products_with_stock(products, request.user)
    low_stock_products = products.filter(stock_quantity__lt=F("stock_reorder"))
    low_stock_products = apply_stock_fields(list(low_stock_products))
    rule_map = {
        rule.product_id: rule
        for rule in ReplenishmentRule.objects.filter(
            user=business_user,
            product_id__in=[product.id for product in low_stock_products],
        )
    }

    business_name = getattr(getattr(request.user, "profile", None), "company_name", None)
    if not business_name:
        business_name = getattr(settings, "DEFAULT_BUSINESS_NAME", "Truck Zone") or "Truck Zone"

    def build_email_body(supplier_label, contact_name, items):
        greeting_target = contact_name or supplier_label or "there"
        lines = [
            f"Hello {greeting_target},",
            "",
            "We need to reorder the following low-stock items:",
            "",
        ]
        for item in items:
            sku = item["sku"] or "N/A"
            lines.append(
                f"- {item['name']} (SKU: {sku}) | On hand: {item['quantity_in_stock']} | Min: {item['reorder_level']} | Max: {item['max_stock_level']} | Order qty: {item['order_qty']}"
            )
        lines.extend(
            [
                "",
                "Please confirm availability, pricing, and ETA.",
                "",
                "Thank you,",
                business_name,
            ]
        )
        return "\n".join(lines)

    supplier_groups = []
    supplier_lookup = {}
    for product in low_stock_products:
        supplier = product.supplier
        supplier_key = supplier.id if supplier else "unassigned"
        group = supplier_lookup.get(supplier_key)
        if not group:
            supplier_name = supplier.name if supplier else "Unassigned supplier"
            contact_person = supplier.contact_person if supplier else ""
            email = supplier.email if supplier else ""
            phone = supplier.phone_number if supplier else ""
            group = {
                "supplier": supplier,
                "supplier_name": supplier_name,
                "contact_person": contact_person,
                "email": email,
                "phone": phone,
                "products": [],
                "total_order_qty": 0,
            }
            supplier_lookup[supplier_key] = group
            supplier_groups.append(group)

        reorder_level = product.reorder_level or 0
        max_stock_level = product.max_stock_level or 0
        rule = rule_map.get(product.id)
        recommendation = _recommended_reorder_for_product(
            product,
            transaction_scope=transaction_scope,
            rule=rule,
        )
        target_stock = recommendation["target_stock"]
        order_qty = recommendation["recommended_qty"]
        group["products"].append(
            {
                "id": product.id,
                "sku": product.sku,
                "name": product.name,
                "quantity_in_stock": product.quantity_in_stock or 0,
                "reorder_level": reorder_level,
                "max_stock_level": max_stock_level,
                "target_stock": target_stock,
                "order_qty": order_qty,
                "location": product.location or "",
                "rule_applied": recommendation["rule_applied"],
                "avg_daily_usage": recommendation["avg_daily_usage"],
            }
        )
        group["total_order_qty"] += order_qty

    subject = f"Stock order request - {business_name}"
    for group in supplier_groups:
        group["product_count"] = len(group["products"])
        group["email_subject"] = subject
        group["email_body"] = build_email_body(
            group["supplier_name"],
            group["contact_person"],
            group["products"],
        )

    context = {
        "supplier_groups": supplier_groups,
        "low_stock_total": len(low_stock_products),
    }
    return render(request, "inventory/stock_orders.html", context)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="category_groups_template.xlsx"'
    return response


@login_required
def export_brands_template(request):
    """Download an Excel template populated with the user's current brands."""

    product_user_ids = _get_inventory_user_ids(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Brands"
    worksheet.append(BRAND_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    brands = ProductBrand.objects.filter(user__in=product_user_ids).order_by("sort_order", "name")
    for brand in brands:
        worksheet.append([
            brand.name or "",
            brand.description or "",
            brand.sort_order,
            "Active" if brand.is_active else "Inactive",
        ])

    active_states = ["Active", "Inactive"]
    options_sheet = workbook.create_sheet(title="Options")
    options_sheet.sheet_state = "hidden"

    def _write_option_column(column_index, header, values):
        options_sheet.cell(row=1, column=column_index, value=header)
        for row_index, value in enumerate(values, start=2):
            options_sheet.cell(row=row_index, column=column_index, value=value)
        last_row = max(2, len(values) + 1)
        column_letter = get_column_letter(column_index)
        return f"=Options!${column_letter}$2:${column_letter}${last_row}"

    active_formula = _write_option_column(1, "Active", active_states)

    def _apply_validation(target_header, formula):
        target_column_index = BRAND_TEMPLATE_HEADERS.index(target_header) + 1
        target_column_letter = get_column_letter(target_column_index)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{target_column_letter}2:{target_column_letter}1048576")

    _apply_validation("Active", active_formula)

    apply_template_styling(
        worksheet,
        headers=BRAND_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={
            "Name": 32,
            "Description": 48,
            "Sort Order": 14,
            "Active": 12,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="brands_template.xlsx"'
    return response


@login_required
def export_models_template(request):
    """Download an Excel template populated with the user's current models."""

    product_user_ids = _get_inventory_user_ids(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Models"
    worksheet.append(MODEL_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    models = (
        ProductModel.objects.filter(user__in=product_user_ids)
        .select_related("brand")
        .order_by("sort_order", "name")
    )
    for model in models:
        worksheet.append([
            model.name or "",
            model.brand.name if model.brand else "",
            model.description or "",
            model.year_start or "",
            model.year_end or "",
            model.sort_order,
            "Active" if model.is_active else "Inactive",
        ])

    brand_names = list(
        dict.fromkeys(
            ProductBrand.objects.filter(user__in=product_user_ids)
            .order_by("sort_order", "name")
            .values_list("name", flat=True)
        )
    )
    active_states = ["Active", "Inactive"]
    options_sheet = workbook.create_sheet(title="Options")
    options_sheet.sheet_state = "hidden"

    def _write_option_column(column_index, header, values):
        options_sheet.cell(row=1, column=column_index, value=header)
        for row_index, value in enumerate(values, start=2):
            options_sheet.cell(row=row_index, column=column_index, value=value)
        last_row = max(2, len(values) + 1)
        column_letter = get_column_letter(column_index)
        return f"=Options!${column_letter}$2:${column_letter}${last_row}"

    brand_formula = _write_option_column(1, "Brands", brand_names)
    active_formula = _write_option_column(2, "Active", active_states)

    def _apply_validation(target_header, formula):
        target_column_index = MODEL_TEMPLATE_HEADERS.index(target_header) + 1
        target_column_letter = get_column_letter(target_column_index)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{target_column_letter}2:{target_column_letter}1048576")

    _apply_validation("Brand", brand_formula)
    _apply_validation("Active", active_formula)

    apply_template_styling(
        worksheet,
        headers=MODEL_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={
            "Name": 28,
            "Brand": 24,
            "Description": 48,
            "Year Start": 14,
            "Year End": 14,
            "Sort Order": 14,
            "Active": 12,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="models_template.xlsx"'
    return response


@login_required
def export_vins_template(request):
    """Download an Excel template populated with the user's current VINs."""

    product_user_ids = _get_inventory_user_ids(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "VINs"
    worksheet.append(VIN_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    vins = ProductVin.objects.filter(user__in=product_user_ids).order_by("sort_order", "vin")
    for vin in vins:
        worksheet.append([
            vin.vin or "",
            vin.description or "",
            vin.sort_order,
            "Active" if vin.is_active else "Inactive",
        ])

    active_states = ["Active", "Inactive"]
    options_sheet = workbook.create_sheet(title="Options")
    options_sheet.sheet_state = "hidden"

    def _write_option_column(column_index, header, values):
        options_sheet.cell(row=1, column=column_index, value=header)
        for row_index, value in enumerate(values, start=2):
            options_sheet.cell(row=row_index, column=column_index, value=value)
        last_row = max(2, len(values) + 1)
        column_letter = get_column_letter(column_index)
        return f"=Options!${column_letter}$2:${column_letter}${last_row}"

    active_formula = _write_option_column(1, "Active", active_states)

    def _apply_validation(target_header, formula):
        target_column_index = VIN_TEMPLATE_HEADERS.index(target_header) + 1
        target_column_letter = get_column_letter(target_column_index)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{target_column_letter}2:{target_column_letter}1048576")

    _apply_validation("Active", active_formula)

    apply_template_styling(
        worksheet,
        headers=VIN_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={
            "VIN": 24,
            "Description": 48,
            "Sort Order": 14,
            "Active": 12,
        },
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="vins_template.xlsx"'
    return response


@login_required
def export_locations_template(request):
    """Download an Excel template populated with the user's current inventory locations."""

    product_user_ids = _get_inventory_user_ids(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Locations"
    worksheet.append(LOCATION_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"

    locations = InventoryLocation.objects.filter(user__in=product_user_ids).order_by("name")
    for location in locations:
        worksheet.append([location.name or ""])

    apply_template_styling(
        worksheet,
        headers=LOCATION_TEMPLATE_HEADERS,
        header_row_index=1,
        column_width_overrides={"Name": 34},
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="locations_template.xlsx"'
    return response


def _parse_decimal(value, field_name, allow_blank=True):
    if value in (None, ""):
        if allow_blank:
            return None
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name} must be a number.") from exc


def _parse_integer(value, field_name):
    if value in (None, ""):
        return 0
    try:
        decimal_value = Decimal(str(value))
        if decimal_value != decimal_value.to_integral_value():
            raise ValueError
        return int(decimal_value)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name} must be an integer.") from exc


def _parse_optional_integer(value, field_name):
    if value in (None, ""):
        return None
    return _parse_integer(value, field_name)


def _parse_date(value, field_name):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
        raise ValueError(
            f"{field_name} must be a valid date in YYYY-MM-DD format."
        )
    raise ValueError(f"{field_name} must be a valid date.")


def _parse_boolean(value, field_name, default=None):
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        if value == 1:
            return True
        if value == 0:
            return False
    if isinstance(value, str):
        normalized = value.strip().lower()
        truthy = {"1", "true", "yes", "y", "active", "enabled", "on"}
        falsy = {"0", "false", "no", "n", "inactive", "disabled", "off"}
        if normalized in truthy:
            return True
        if normalized in falsy:
            return False
    raise ValueError(f"{field_name} must be yes/no or true/false.")


@login_required
@require_POST
def import_products_from_excel(request):
    next_url = request.POST.get("next") or request.GET.get("next") or (reverse("accounts:inventory_products"))
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(next_url)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded file. Please upload a valid .xlsx file.")
        return redirect(next_url)

    worksheet = workbook.active

    header_cells = None
    header_row_index = None
    for row in worksheet.iter_rows(min_row=1):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        first_value = row_values[0]
        if isinstance(first_value, str) and first_value.startswith("Instructions:"):
            continue
        header_cells = row
        header_row_index = row[0].row
        break

    if header_cells is None:
        messages.error(request, "The uploaded file is empty or missing a header row.")
        return redirect(next_url)

    headers = []
    for cell in header_cells:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        headers.append(value or "")

    missing_headers = [
        h
        for h in PRODUCT_TEMPLATE_HEADERS
        if h not in headers and h not in OPTIONAL_PRODUCT_TEMPLATE_HEADERS
    ]
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: "
            + ", ".join(missing_headers),
        )
        return redirect(next_url)

    header_indexes = {header: idx for idx, header in enumerate(headers)}

    item_type_choices = Product._meta.get_field("item_type").choices
    item_type_map = {}
    for value, label in item_type_choices:
        item_type_map[str(value).strip().lower()] = value
        item_type_map[str(label).strip().lower()] = value
        item_type_map[str(label).strip().lower().replace("-", "_")] = value

    created_count = 0
    updated_count = 0
    errors = []

    data_start_row = header_row_index + 1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if not row or not any(row):
            continue

        def get_value(column_name):
            idx = header_indexes[column_name]
            return row[idx] if idx < len(row) else None

        try:
            sku = _normalize_text(get_value("SKU"))
            oem_part_number = _normalize_text(get_value("OEM Part Number"))
            barcode_value = _normalize_text(get_value("Barcode"))
            fitment_notes = _normalize_text(get_value("Fitment Notes"))
            alternate_skus = None
            if "Alternate SKUs" in header_indexes:
                alternate_skus = _parse_alternate_sku_input(get_value("Alternate SKUs"))

            raw_name = get_value("Name")
            name = _normalize_text(raw_name)
            if not name:
                raise ValueError("Name is required.")

            description = get_value("Description") or ""
            if isinstance(description, str):
                description = description.strip()

            category_name = _normalize_text(get_value("Category"))
            category = None
            if category_name:
                category = Category.objects.filter(
                    user=request.user, name__iexact=category_name
                ).first()
                if not category:
                    category = Category.objects.create(
                        user=request.user,
                        name=category_name,
                        description="",
                    )

            supplier_name = _normalize_text(get_value("Supplier"))
            supplier = None
            if supplier_name:
                supplier = Supplier.objects.filter(
                    user=request.user, name__iexact=supplier_name
                ).first()
                if not supplier:
                    supplier = Supplier.objects.create(
                        user=request.user,
                        name=supplier_name,
                    )

            brand_name = _normalize_text(get_value("Brand"))
            brand = None
            if brand_name:
                brand = ProductBrand.objects.filter(
                    user=request.user, name__iexact=brand_name
                ).first()
                if not brand:
                    brand = ProductBrand.objects.create(
                        user=request.user,
                        name=brand_name,
                    )

            model_name = _normalize_text(get_value("Model"))
            vehicle_model = None
            if model_name:
                model_qs = ProductModel.objects.filter(
                    user=request.user, name__iexact=model_name
                )
                if brand:
                    model_qs = model_qs.filter(brand=brand)
                vehicle_model = model_qs.first()
                if not vehicle_model:
                    vehicle_model = ProductModel.objects.create(
                        user=request.user,
                        name=model_name,
                        brand=brand,
                    )

            vin_value = _normalize_text(get_value("VIN"))
            vin_number = None
            if vin_value:
                vin_number = ProductVin.objects.filter(
                    user=request.user, vin__iexact=vin_value
                ).first()
                if not vin_number:
                    vin_number = ProductVin.objects.create(
                        user=request.user,
                        vin=vin_value,
                    )

            cost_price = _parse_decimal(get_value("Cost Price"), "Cost Price")
            sale_price = _parse_decimal(get_value("Sale Price"), "Sale Price")
            promotion_price = _parse_decimal(get_value("Promotion Price"), "Promotion Price")
            margin = _parse_decimal(get_value("Margin"), "Margin")

            for label, value in (("Cost Price", cost_price), ("Sale Price", sale_price)):
                if value is not None and value < Decimal("0"):
                    raise ValueError(f"{label} cannot be negative.")
            if margin is None:
                if cost_price is None and sale_price is None:
                    raise ValueError(
                        "Provide Cost Price and Sale Price, or include the Margin so missing values can be calculated."
                    )
                if cost_price is None or sale_price is None:
                    raise ValueError(
                        "Margin is required when only one of Cost Price or Sale Price is supplied."
                    )
                margin = sale_price - cost_price
            else:
                if cost_price is None and sale_price is None:
                    raise ValueError(
                        "Cost Price or Sale Price is required when Margin is provided."
                    )
                if cost_price is None:
                    cost_price = sale_price - margin
                if sale_price is None:
                    sale_price = cost_price + margin
                if cost_price is not None and sale_price is not None:
                    margin = sale_price - cost_price

            if cost_price is None or sale_price is None:
                raise ValueError("Unable to determine both Cost Price and Sale Price from the provided values.")

            if cost_price < Decimal("0"):
                raise ValueError("Cost Price cannot be negative.")
            if sale_price < Decimal("0"):
                raise ValueError("Sale Price cannot be negative.")
            if promotion_price is not None and promotion_price < Decimal("0"):
                raise ValueError("Promotion Price cannot be negative.")

            quantity = _parse_integer(get_value("Quantity"), "Quantity")
            if quantity < 0:
                raise ValueError("Quantity cannot be negative.")

            reorder_level = _parse_integer(
                get_value("Reorder Level"), "Reorder Level"
            )
            if reorder_level < 0:
                raise ValueError("Reorder Level cannot be negative.")
            max_stock_level = _parse_integer(
                get_value("Max Stock Level"), "Max Stock Level"
            )
            if max_stock_level < 0:
                raise ValueError("Max Stock Level cannot be negative.")
            if max_stock_level and max_stock_level < reorder_level:
                raise ValueError("Max Stock Level must be greater than or equal to Reorder Level.")

            location = _normalize_text(get_value("Location"))

            warranty_expiry = _parse_date(
                get_value("Warranty Expiry Date"), "Warranty Expiry Date"
            )
            warranty_length = get_value("Warranty Length (Days)")
            if warranty_length in (None, ""):
                warranty_length_value = None
            else:
                warranty_length_value = _parse_integer(
                    warranty_length, "Warranty Length (Days)"
                )
                if warranty_length_value < 0:
                    raise ValueError("Warranty Length cannot be negative.")

            item_type_raw = _normalize_text(get_value("Item Type"))
            item_type = None
            if item_type_raw:
                normalized_item_type = item_type_raw.strip().lower().replace(" ", "_").replace("-", "_")
                item_type = item_type_map.get(normalized_item_type)
                if not item_type:
                    raise ValueError("Item Type must be Inventory or Non-inventory.")

            show_on_storefront = _parse_boolean(
                get_value("Show on Storefront"), "Show on Storefront", default=None
            )
            featured = _parse_boolean(get_value("Featured"), "Featured", default=None)

            product = None
            if sku:
                product = Product.objects.filter(
                    user=request.user, sku__iexact=sku
                ).first()
            if not product:
                product = Product.objects.filter(
                    user=request.user, name__iexact=name
                ).first()

            if product:
                is_new = False
            else:
                product = Product(user=request.user)
                is_new = True

            stock_owner = get_stock_owner(request.user)
            product.user = stock_owner or request.user
            product.sku = sku or None
            product.oem_part_number = oem_part_number or None
            product.barcode_value = barcode_value or None
            product.name = name
            product.description = description or ""
            product.fitment_notes = fitment_notes or ""
            product.category = category
            product.supplier = supplier
            product.brand = brand
            product.vehicle_model = vehicle_model
            product.vin_number = vin_number
            product.cost_price = cost_price
            product.sale_price = sale_price
            product.promotion_price = promotion_price
            product.margin = margin
            product.quantity_in_stock = quantity
            product.reorder_level = reorder_level
            product.max_stock_level = max_stock_level
            product.location = location or ""
            product.warranty_expiry_date = warranty_expiry
            product.warranty_length = warranty_length_value
            if item_type:
                product.item_type = item_type
            if show_on_storefront is not None:
                product.is_published_to_store = show_on_storefront
            if featured is not None:
                product.is_featured = featured

            product.clean()
            product.save()
            _sync_alternate_skus(product, alternate_skus)
            stock_quantity = quantity if product.item_type == "inventory" else 0
            stock_reorder = reorder_level if product.item_type == "inventory" else 0
            stock_max = max_stock_level if product.item_type == "inventory" else 0
            upsert_product_stock(
                product,
                request.user,
                quantity_in_stock=stock_quantity,
                reorder_level=stock_reorder,
                max_stock_level=stock_max,
            )

            if is_new:
                created_count += 1
            else:
                updated_count += 1

        except ValidationError as exc:
            errors.append(f"Row {row_index}: {exc.message}")
        except ValueError as exc:
            errors.append(f"Row {row_index}: {exc}")

    if created_count or updated_count:
        messages.success(
            request,
            f"Imported {created_count} new product(s) and updated {updated_count} product(s).",
        )

    if errors:
        error_preview = "; ".join(errors[:10])
        if len(errors) > 10:
            error_preview += f"; and {len(errors) - 10} more issue(s)."
        messages.warning(
            request,
            f"Some rows could not be imported: {error_preview}",
        )

    return redirect(reverse("accounts:inventory_products"))


def _normalize_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ("%s" % value).strip()
    return str(value).strip()


def _extract_ids(raw_ids):
    valid_ids = []
    for value in raw_ids:
        try:
            valid_ids.append(int(value))
        except (TypeError, ValueError):
            continue
    return valid_ids


@login_required
@require_POST
def import_suppliers_from_excel(request):
    next_url = request.POST.get("next") or request.GET.get("next") or (reverse("accounts:inventory_suppliers"))
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(next_url)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded file. Please upload a valid .xlsx file.")
        return redirect(next_url)

    worksheet = workbook.active
    header_cells = None
    header_row_index = None
    for row in worksheet.iter_rows(min_row=1):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        first_value = row_values[0]
        if isinstance(first_value, str) and first_value.startswith("Instructions:"):
            continue
        header_cells = row
        header_row_index = row[0].row
        break

    if header_cells is None:
        messages.error(request, "The uploaded file is empty or missing a header row.")
        return redirect(next_url)

    headers = []
    for cell in header_cells:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        headers.append(value or "")

    missing_headers = [h for h in SUPPLIER_TEMPLATE_HEADERS if h not in headers]
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: " + ", ".join(missing_headers),
        )
        return redirect(next_url)

    header_indexes = {header: headers.index(header) for header in SUPPLIER_TEMPLATE_HEADERS}

    created_count = 0
    updated_count = 0
    errors = []

    data_start_row = header_row_index + 1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if not row or not any(row):
            continue

        def get_value(column_name):
            idx = header_indexes[column_name]
            return row[idx] if idx < len(row) else None

        try:
            raw_name = get_value("Name")
            name = _normalize_text(raw_name)
            if not name:
                raise ValueError("Name is required.")

            contact_person = _normalize_text(get_value("Contact Person"))
            email_value = _normalize_text(get_value("Email"))
            phone_number = _normalize_text(get_value("Phone Number"))
            address = _normalize_text(get_value("Address"))

            supplier = Supplier.objects.filter(user=request.user, name__iexact=name).first()
            if supplier:
                is_new = False
            else:
                supplier = Supplier(user=request.user)
                is_new = True

            supplier.name = name
            supplier.contact_person = contact_person or None
            supplier.email = email_value or None
            supplier.phone_number = phone_number or None
            supplier.address = address or ""

            supplier.full_clean()
            supplier.save()

            if is_new:
                created_count += 1
            else:
                updated_count += 1

        except ValidationError as exc:
            errors.append(f"Row {row_index}: {exc.message}")
        except ValueError as exc:
            errors.append(f"Row {row_index}: {exc}")

    if created_count or updated_count:
        messages.success(
            request,
            f"Imported {created_count} new supplier(s) and updated {updated_count} supplier(s).",
        )

    if errors:
        error_preview = "; ".join(errors[:10])
        if len(errors) > 10:
            error_preview += f"; and {len(errors) - 10} more issue(s)."
        messages.warning(
            request,
            f"Some rows could not be imported: {error_preview}",
        )

    return redirect(next_url)


@login_required
@require_POST
def import_categories_from_excel(request):
    next_url = request.POST.get("next") or request.GET.get("next") or (reverse("accounts:inventory_categories"))
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(next_url)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded file. Please upload a valid .xlsx file.")
        return redirect(next_url)

    worksheet = workbook.active
    header_cells = None
    header_row_index = None
    for row in worksheet.iter_rows(min_row=1):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        first_value = row_values[0]
        if isinstance(first_value, str) and first_value.startswith("Instructions:"):
            continue
        header_cells = row
        header_row_index = row[0].row
        break

    if header_cells is None:
        messages.error(request, "The uploaded file is empty or missing a header row.")
        return redirect(next_url)

    headers = []
    for cell in header_cells:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        headers.append(value or "")

    missing_headers = [h for h in CATEGORY_TEMPLATE_HEADERS if h not in headers]
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: "
            + ", ".join(missing_headers),
        )
        return redirect(next_url)

    header_indexes = {header: headers.index(header) for header in CATEGORY_TEMPLATE_HEADERS}

    created_count = 0
    updated_count = 0
    errors = []

    data_start_row = header_row_index + 1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if not row or not any(row):
            continue

        def get_value(column_name):
            idx = header_indexes[column_name]
            return row[idx] if idx < len(row) else None

        try:
            name = _normalize_text(get_value("Name"))
            if not name:
                raise ValueError("Name is required.")

            description = _normalize_text(get_value("Description"))
            group_name = _normalize_text(get_value("Group"))
            parent_name = _normalize_text(get_value("Parent Category"))
            sort_order = _parse_integer(get_value("Sort Order"), "Sort Order")
            is_active = _parse_boolean(get_value("Active"), "Active", default=None)

            group = None
            if group_name:
                group = CategoryGroup.objects.filter(
                    user=request.user, name__iexact=group_name
                ).first()
                if not group:
                    group = CategoryGroup.objects.create(
                        user=request.user,
                        name=group_name,
                    )

            parent = None
            if parent_name and parent_name.lower() != name.lower():
                parent = Category.objects.filter(
                    user=request.user, name__iexact=parent_name
                ).first()
                if not parent:
                    parent = Category.objects.create(
                        user=request.user,
                        name=parent_name,
                        group=group,
                    )

            category = Category.objects.filter(
                user=request.user, name__iexact=name
            ).first()
            if category:
                is_new = False
            else:
                category = Category(user=request.user)
                is_new = True

            category.name = name
            category.description = description or ""
            category.group = group
            category.parent = parent
            category.sort_order = sort_order
            if is_active is not None:
                category.is_active = is_active
            if category.parent_id and category.parent_id == category.pk:
                category.parent = None
            category.full_clean()
            category.save()

            if is_new:
                created_count += 1
            else:
                updated_count += 1

        except ValidationError as exc:
            errors.append(f"Row {row_index}: {exc.message}")
        except ValueError as exc:
            errors.append(f"Row {row_index}: {exc}")

    if created_count or updated_count:
        messages.success(
            request,
            f"Imported {created_count} new category(ies) and updated {updated_count} category(ies).",
        )

    if errors:
        error_preview = "; ".join(errors[:10])
        if len(errors) > 10:
            error_preview += f"; and {len(errors) - 10} more issue(s)."
        messages.warning(
            request,
            f"Some rows could not be imported: {error_preview}",
        )

    return redirect(next_url)


@login_required
@require_POST
def import_category_groups_from_excel(request):
    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_category_groups")
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(next_url)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded file. Please upload a valid .xlsx file.")
        return redirect(next_url)

    worksheet = workbook.active
    header_cells = None
    header_row_index = None
    for row in worksheet.iter_rows(min_row=1):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        first_value = row_values[0]
        if isinstance(first_value, str) and first_value.startswith("Instructions:"):
            continue
        header_cells = row
        header_row_index = row[0].row
        break

    if header_cells is None:
        messages.error(request, "The uploaded file is empty or missing a header row.")
        return redirect(next_url)

    headers = []
    for cell in header_cells:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        headers.append(value or "")

    missing_headers = [h for h in CATEGORY_GROUP_TEMPLATE_HEADERS if h not in headers]
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: "
            + ", ".join(missing_headers),
        )
        return redirect(next_url)

    header_indexes = {
        header: headers.index(header) for header in CATEGORY_GROUP_TEMPLATE_HEADERS
    }

    created_count = 0
    updated_count = 0
    errors = []

    data_start_row = header_row_index + 1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if not row or not any(row):
            continue

        def get_value(column_name):
            idx = header_indexes[column_name]
            return row[idx] if idx < len(row) else None

        try:
            name = _normalize_text(get_value("Name"))
            if not name:
                raise ValueError("Name is required.")

            description = _normalize_text(get_value("Description"))
            sort_order = _parse_integer(get_value("Sort Order"), "Sort Order")
            is_active = _parse_boolean(get_value("Active"), "Active", default=None)

            group = CategoryGroup.objects.filter(
                user=request.user, name__iexact=name
            ).first()
            if group:
                is_new = False
            else:
                group = CategoryGroup(user=request.user)
                is_new = True

            group.name = name
            group.description = description or ""
            group.sort_order = sort_order
            if is_active is not None:
                group.is_active = is_active

            group.full_clean()
            group.save()

            if is_new:
                created_count += 1
            else:
                updated_count += 1

        except ValidationError as exc:
            errors.append(f"Row {row_index}: {exc.message}")
        except ValueError as exc:
            errors.append(f"Row {row_index}: {exc}")

    if created_count or updated_count:
        messages.success(
            request,
            f"Imported {created_count} new group(s) and updated {updated_count} group(s).",
        )

    if errors:
        error_preview = "; ".join(errors[:10])
        if len(errors) > 10:
            error_preview += f"; and {len(errors) - 10} more issue(s)."
        messages.warning(
            request,
            f"Some rows could not be imported: {error_preview}",
        )

    return redirect(next_url)


@login_required
@require_POST
def import_brands_from_excel(request):
    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_brands")
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(next_url)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded file. Please upload a valid .xlsx file.")
        return redirect(next_url)

    worksheet = workbook.active
    header_cells = None
    header_row_index = None
    for row in worksheet.iter_rows(min_row=1):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        first_value = row_values[0]
        if isinstance(first_value, str) and first_value.startswith("Instructions:"):
            continue
        header_cells = row
        header_row_index = row[0].row
        break

    if header_cells is None:
        messages.error(request, "The uploaded file is empty or missing a header row.")
        return redirect(next_url)

    headers = []
    for cell in header_cells:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        headers.append(value or "")

    missing_headers = [h for h in BRAND_TEMPLATE_HEADERS if h not in headers]
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: "
            + ", ".join(missing_headers),
        )
        return redirect(next_url)

    header_indexes = {header: headers.index(header) for header in BRAND_TEMPLATE_HEADERS}

    created_count = 0
    updated_count = 0
    errors = []

    data_start_row = header_row_index + 1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if not row or not any(row):
            continue

        def get_value(column_name):
            idx = header_indexes[column_name]
            return row[idx] if idx < len(row) else None

        try:
            name = _normalize_text(get_value("Name"))
            if not name:
                raise ValueError("Name is required.")

            description = _normalize_text(get_value("Description"))
            sort_order = _parse_integer(get_value("Sort Order"), "Sort Order")
            is_active = _parse_boolean(get_value("Active"), "Active", default=None)

            brand = ProductBrand.objects.filter(
                user=request.user, name__iexact=name
            ).first()
            if brand:
                is_new = False
            else:
                brand = ProductBrand(user=request.user)
                is_new = True

            brand.name = name
            brand.description = description or ""
            brand.sort_order = sort_order
            if is_active is not None:
                brand.is_active = is_active

            brand.full_clean()
            brand.save()

            if is_new:
                created_count += 1
            else:
                updated_count += 1

        except ValidationError as exc:
            errors.append(f"Row {row_index}: {exc.message}")
        except ValueError as exc:
            errors.append(f"Row {row_index}: {exc}")

    if created_count or updated_count:
        messages.success(
            request,
            f"Imported {created_count} new brand(s) and updated {updated_count} brand(s).",
        )

    if errors:
        error_preview = "; ".join(errors[:10])
        if len(errors) > 10:
            error_preview += f"; and {len(errors) - 10} more issue(s)."
        messages.warning(
            request,
            f"Some rows could not be imported: {error_preview}",
        )

    return redirect(next_url)


@login_required
@require_POST
def import_models_from_excel(request):
    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_models")
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(next_url)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded file. Please upload a valid .xlsx file.")
        return redirect(next_url)

    worksheet = workbook.active
    header_cells = None
    header_row_index = None
    for row in worksheet.iter_rows(min_row=1):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        first_value = row_values[0]
        if isinstance(first_value, str) and first_value.startswith("Instructions:"):
            continue
        header_cells = row
        header_row_index = row[0].row
        break

    if header_cells is None:
        messages.error(request, "The uploaded file is empty or missing a header row.")
        return redirect(next_url)

    headers = []
    for cell in header_cells:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        headers.append(value or "")

    missing_headers = [h for h in MODEL_TEMPLATE_HEADERS if h not in headers]
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: "
            + ", ".join(missing_headers),
        )
        return redirect(next_url)

    header_indexes = {header: headers.index(header) for header in MODEL_TEMPLATE_HEADERS}

    created_count = 0
    updated_count = 0
    errors = []

    data_start_row = header_row_index + 1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if not row or not any(row):
            continue

        def get_value(column_name):
            idx = header_indexes[column_name]
            return row[idx] if idx < len(row) else None

        try:
            name = _normalize_text(get_value("Name"))
            if not name:
                raise ValueError("Name is required.")

            brand_name = _normalize_text(get_value("Brand"))
            description = _normalize_text(get_value("Description"))
            year_start = _parse_optional_integer(get_value("Year Start"), "Year Start")
            year_end = _parse_optional_integer(get_value("Year End"), "Year End")
            sort_order = _parse_integer(get_value("Sort Order"), "Sort Order")
            is_active = _parse_boolean(get_value("Active"), "Active", default=None)

            brand = None
            if brand_name:
                brand = ProductBrand.objects.filter(
                    user=request.user, name__iexact=brand_name
                ).first()
                if not brand:
                    brand = ProductBrand.objects.create(
                        user=request.user,
                        name=brand_name,
                    )

            model = ProductModel.objects.filter(
                user=request.user, name__iexact=name
            ).first()
            if model:
                is_new = False
            else:
                model = ProductModel(user=request.user)
                is_new = True

            model.name = name
            model.brand = brand
            model.description = description or ""
            model.year_start = year_start
            model.year_end = year_end
            model.sort_order = sort_order
            if is_active is not None:
                model.is_active = is_active

            model.full_clean()
            model.save()

            if is_new:
                created_count += 1
            else:
                updated_count += 1

        except ValidationError as exc:
            errors.append(f"Row {row_index}: {exc.message}")
        except ValueError as exc:
            errors.append(f"Row {row_index}: {exc}")

    if created_count or updated_count:
        messages.success(
            request,
            f"Imported {created_count} new model(s) and updated {updated_count} model(s).",
        )

    if errors:
        error_preview = "; ".join(errors[:10])
        if len(errors) > 10:
            error_preview += f"; and {len(errors) - 10} more issue(s)."
        messages.warning(
            request,
            f"Some rows could not be imported: {error_preview}",
        )

    return redirect(next_url)


@login_required
@require_POST
def import_vins_from_excel(request):
    next_url = request.POST.get("next") or request.GET.get("next") or reverse("accounts:inventory_vins")
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(next_url)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded file. Please upload a valid .xlsx file.")
        return redirect(next_url)

    worksheet = workbook.active
    header_cells = None
    header_row_index = None
    for row in worksheet.iter_rows(min_row=1):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        first_value = row_values[0]
        if isinstance(first_value, str) and first_value.startswith("Instructions:"):
            continue
        header_cells = row
        header_row_index = row[0].row
        break

    if header_cells is None:
        messages.error(request, "The uploaded file is empty or missing a header row.")
        return redirect(next_url)

    headers = []
    for cell in header_cells:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        headers.append(value or "")

    missing_headers = [h for h in VIN_TEMPLATE_HEADERS if h not in headers]
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: "
            + ", ".join(missing_headers),
        )
        return redirect(next_url)

    header_indexes = {header: headers.index(header) for header in VIN_TEMPLATE_HEADERS}

    created_count = 0
    updated_count = 0
    errors = []

    data_start_row = header_row_index + 1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if not row or not any(row):
            continue

        def get_value(column_name):
            idx = header_indexes[column_name]
            return row[idx] if idx < len(row) else None

        try:
            vin_value = _normalize_text(get_value("VIN"))
            if not vin_value:
                raise ValueError("VIN is required.")
            vin_value = vin_value.upper()

            description = _normalize_text(get_value("Description"))
            sort_order = _parse_integer(get_value("Sort Order"), "Sort Order")
            is_active = _parse_boolean(get_value("Active"), "Active", default=None)

            vin = ProductVin.objects.filter(
                user=request.user, vin__iexact=vin_value
            ).first()
            if vin:
                is_new = False
            else:
                vin = ProductVin(user=request.user)
                is_new = True

            vin.vin = vin_value
            vin.description = description or ""
            vin.sort_order = sort_order
            if is_active is not None:
                vin.is_active = is_active

            vin.full_clean()
            vin.save()

            if is_new:
                created_count += 1
            else:
                updated_count += 1

        except ValidationError as exc:
            errors.append(f"Row {row_index}: {exc.message}")
        except ValueError as exc:
            errors.append(f"Row {row_index}: {exc}")

    if created_count or updated_count:
        messages.success(
            request,
            f"Imported {created_count} new VIN(s) and updated {updated_count} VIN(s).",
        )

    if errors:
        error_preview = "; ".join(errors[:10])
        if len(errors) > 10:
            error_preview += f"; and {len(errors) - 10} more issue(s)."
        messages.warning(
            request,
            f"Some rows could not be imported: {error_preview}",
        )

    return redirect(next_url)


@login_required
@require_POST
def import_locations_from_excel(request):
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(reverse("accounts:inventory_locations"))

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded file. Please upload a valid .xlsx file.")
        return redirect(reverse("accounts:inventory_locations"))

    worksheet = workbook.active
    try:
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
    except StopIteration:
        messages.error(request, "The uploaded file is empty.")
        return redirect(reverse("accounts:inventory_locations"))

    headers = []
    for cell in header_cells:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        headers.append(value or "")

    missing_headers = [h for h in LOCATION_TEMPLATE_HEADERS if h not in headers]
    if missing_headers:
        messages.error(
            request,
            "The uploaded file is missing required columns: "
            + ", ".join(missing_headers),
        )
        return redirect(reverse("accounts:inventory_locations"))

    header_indexes = {header: headers.index(header) for header in LOCATION_TEMPLATE_HEADERS}

    created_count = 0
    updated_count = 0
    errors = []

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not row or not any(row):
            continue

        def get_value(column_name):
            idx = header_indexes[column_name]
            return row[idx] if idx < len(row) else None

        try:
            name = _normalize_text(get_value("Name"))
            if not name:
                raise ValueError("Name is required.")

            location = InventoryLocation.objects.filter(
                user=request.user, name__iexact=name
            ).first()
            if location:
                is_new = False
            else:
                location = InventoryLocation(user=request.user)
                is_new = True

            location.name = name
            location.full_clean()
            location.save()

            if is_new:
                created_count += 1
            else:
                updated_count += 1

        except ValidationError as exc:
            errors.append(f"Row {row_index}: {exc.message}")
        except ValueError as exc:
            errors.append(f"Row {row_index}: {exc}")

    if created_count or updated_count:
        messages.success(
            request,
            f"Imported {created_count} new location(s) and updated {updated_count} location(s).",
        )

    if errors:
        error_preview = "; ".join(errors[:10])
        if len(errors) > 10:
            error_preview += f"; and {len(errors) - 10} more issue(s)."
        messages.warning(
            request,
            f"Some rows could not be imported: {error_preview}",
        )

    return redirect(reverse("accounts:inventory_locations"))


##############################
# Category CRUD              #
##############################
@login_required
def add_category(request):
    next_url = request.POST.get("next") or request.GET.get("next")
    if request.method == "POST":
        form = CategoryForm(request.POST, user=request.user)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            messages.success(request, "Category added successfully.")
            if _is_ajax(request):
                return JsonResponse(
                    {"success": True, "value": category.id, "label": category.name}
                )
        else:
            messages.error(
                request, "Error adding category. Please check the form for errors."
            )
            if _is_ajax(request):
                return JsonResponse(
                    {"success": False, "errors": _form_errors_json(form)}, status=400
                )
    return redirect(next_url or reverse("accounts:inventory_categories"))


@login_required
def edit_category(request):
    if request.method == "POST":
        category_id = request.POST.get("category_id")
        category = get_object_or_404(
            Category,
            id=category_id,
            user__in=_get_inventory_user_ids(request),
        )
        form = CategoryForm(request.POST, user=request.user, instance=category)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Category updated successfully.")
            except Exception as e:
                messages.error(request, "Error updating category: " + str(e))
        else:
            messages.error(
                request, "Error updating category. Please check the form for errors."
            )
    # Return to Categories tab
    return redirect(reverse("accounts:inventory_categories"))


@login_required
def delete_category(request):
    if request.method == "POST":
        category_id = request.POST.get("category_id")
        category = get_object_or_404(
            Category,
            id=category_id,
            user__in=_get_inventory_user_ids(request),
        )
        category.delete()
        messages.success(request, "Category deleted successfully.")
    else:
        messages.error(request, "Invalid request method.")
    # Return to Categories tab
    return redirect(reverse("accounts:inventory_categories"))


@login_required
@require_POST
def bulk_delete_categories(request):
    category_ids = _extract_ids(request.POST.getlist("category_ids"))
    if not category_ids:
        messages.warning(request, "No categories were selected for deletion.")
        return redirect(reverse("accounts:inventory_categories"))

    queryset = Category.objects.filter(
        user__in=_get_inventory_user_ids(request),
        id__in=category_ids,
    )
    deleted_count = queryset.count()

    if not deleted_count:
        messages.warning(request, "No matching categories were found to delete.")
    else:
        queryset.delete()
        messages.success(request, f"Deleted {deleted_count} category(ies).")

    return redirect(reverse("accounts:inventory_categories"))


##############################
# Location CRUD              #
##############################
@login_required
def add_location(request):
    next_url = request.POST.get("next") or request.GET.get("next") or (reverse("accounts:inventory_locations"))
    if request.method == "POST":
        form = InventoryLocationForm(request.POST, user=request.user)
        if form.is_valid():
            location = form.save(commit=False)
            location.user = request.user
            location.save()
            messages.success(request, "Location added successfully.")
            if _is_ajax(request):
                return JsonResponse(
                    {"success": True, "value": location.name, "label": location.name}
                )
        else:
            messages.error(
                request, "Error adding location. Please check the form for errors."
            )
            if _is_ajax(request):
                return JsonResponse(
                    {"success": False, "errors": _form_errors_json(form)}, status=400
                )
    return redirect(next_url)


@login_required
def edit_location(request):
    if request.method == "POST":
        location_id = request.POST.get("location_id")
        location = get_object_or_404(
            InventoryLocation,
            id=location_id,
            user__in=_get_inventory_user_ids(request),
        )
        form = InventoryLocationForm(request.POST, user=request.user, instance=location)
        if form.is_valid():
            form.save()
            messages.success(request, "Location updated successfully.")
        else:
            messages.error(
                request, "Error updating location. Please check the form for errors."
            )
    return redirect(reverse("accounts:inventory_locations"))


@login_required
def delete_location(request):
    if request.method == "POST":
        location_id = request.POST.get("location_id")
        location = get_object_or_404(
            InventoryLocation,
            id=location_id,
            user__in=_get_inventory_user_ids(request),
        )
        location.delete()
        messages.success(request, "Location deleted successfully.")
    else:
        messages.error(request, "Invalid request method.")
    return redirect(reverse("accounts:inventory_locations"))


@login_required
@require_POST
def bulk_delete_locations(request):
    location_ids = _extract_ids(request.POST.getlist("location_ids"))
    if not location_ids:
        messages.warning(request, "No locations were selected for deletion.")
        return redirect(reverse("accounts:inventory_locations"))

    queryset = InventoryLocation.objects.filter(
        user__in=_get_inventory_user_ids(request),
        id__in=location_ids,
    )
    deleted_count = queryset.count()

    if not deleted_count:
        messages.warning(request, "No matching locations were found to delete.")
    else:
        queryset.delete()
        messages.success(request, f"Deleted {deleted_count} location(s).")

    return redirect(reverse("accounts:inventory_locations"))


##############################
# QR Code & Stock-In Views    #
##############################


@login_required
def product_qr_pdf(request, product_id):
    product = get_object_or_404(
        Product,
        id=product_id,
        user__in=_get_inventory_user_ids(request),
    )
    stock_url = request.build_absolute_uri(
        reverse("accounts:qr_stock_in", args=[product.id])
    )

    try:
        profile = request.user.profile
        font_scale_percent = profile.qr_code_font_scale or 100
        show_name = profile.qr_show_name
        show_description = profile.qr_show_description
        show_sku = profile.qr_show_sku
    except (Profile.DoesNotExist, AttributeError):
        font_scale_percent = 100
        show_name = True
        show_description = True
        show_sku = True

    try:
        font_scale_percent = int(font_scale_percent)
    except (TypeError, ValueError):
        font_scale_percent = 100

    font_scale_percent = max(20, min(160, font_scale_percent))
    font_scale = font_scale_percent / 100.0

    img = qrcode.make(stock_url)
    if img.mode != "RGB":
        img = img.convert("RGB")

    dpi = 300
    label_width_in = 3.5
    label_height_in = 1.125
    label_width_px = int(round(label_width_in * dpi))
    label_height_px = int(round(label_height_in * dpi))

    label_image = Image.new("RGB", (label_width_px, label_height_px), "white")
    draw = ImageDraw.Draw(label_image)

    try:
        resampling = Image.Resampling.LANCZOS  # Pillow >= 9.1
    except AttributeError:  # pragma: no cover - Pillow < 9.1
        resampling = Image.LANCZOS

    qr_target_size = int(label_height_px * 0.9)
    qr_image = img.resize((qr_target_size, qr_target_size), resampling)

    vertical_center = (label_height_px - qr_target_size) // 2
    horizontal_padding = int(label_height_px * 0.1)
    label_image.paste(qr_image, (horizontal_padding, vertical_center))

    text_start_x = horizontal_padding + qr_target_size + horizontal_padding
    text_area_width = label_width_px - text_start_x - horizontal_padding
    top_padding = int(label_height_px * 0.1)
    bottom_padding = top_padding
    text_top = top_padding
    text_bottom_limit = label_height_px - bottom_padding
    text_color = (26, 26, 26)

    bold_font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    ]
    regular_font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]

    def load_font(size, bold=False):
        paths = bold_font_paths if bold else regular_font_paths
        for path in paths:
            if os.path.exists(path):
                return ImageFont.truetype(path, size=size)
        return ImageFont.load_default()

    def line_height(font):
        ascent, descent = font.getmetrics()
        return ascent + descent

    name_font = load_font(max(int(label_height_px * 0.28 * font_scale), 10), bold=True)
    description_font = load_font(max(int(label_height_px * 0.18 * font_scale), 8))
    sku_font = load_font(max(int(label_height_px * 0.16 * font_scale), 8))

    def wrap_text(text, font):
        if not text:
            return []
        lines = []
        words = text.split()
        current_line = ""
        for word in words:
            test_line = f"{current_line} {word}".strip()
            if draw.textlength(test_line, font=font) <= text_area_width:
                current_line = test_line
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        if current_line:
            lines.append(current_line)
        return lines

    sku_reserved_height = line_height(sku_font) if show_sku and product.sku else 0
    available_text_bottom = max(text_top, text_bottom_limit - sku_reserved_height)

    current_y = text_top

    def draw_wrapped_lines(lines, font, fill, max_bottom):
        nonlocal current_y
        if not lines:
            return
        height = line_height(font)
        for line in lines:
            if current_y + height > max_bottom:
                break
            draw.text((text_start_x, current_y), line, font=font, fill=fill)
            current_y += height

    product_name = (product.name or "").upper()
    if show_name and product_name:
        name_lines = wrap_text(product_name, name_font)
        draw_wrapped_lines(name_lines, name_font, text_color, available_text_bottom)

    if show_description:
        if product.description:
            description_text = product.description.strip()
            description_fill = text_color
        else:
            description_text = "No description available"
            description_fill = (136, 136, 136)

        wrapped_description = wrap_text(description_text, description_font)
        draw_wrapped_lines(wrapped_description, description_font, description_fill, available_text_bottom)

    if show_sku and product.sku:
        sku_text = product.sku
        sku_height = line_height(sku_font)
        sku_y = max(current_y, text_bottom_limit - sku_height)
        draw.text((text_start_x, sku_y), sku_text, font=sku_font, fill=(85, 85, 85))

    response = HttpResponse(content_type="image/jpeg")
    filename = f"{product.sku or product.id}_qr.jpg"
    if "download" in request.GET:
        disposition = f"attachment; filename={filename}"
    else:
        disposition = f"inline; filename={filename}"
    response["Content-Disposition"] = disposition
    label_image.save(response, format="JPEG", dpi=(dpi, dpi), quality=95)
    return response


@login_required
def qr_stock_in(request, product_id):
    """Display a quick inventory transaction form for QR code scans."""

    product_queryset = annotate_products_with_stock(
        Product.objects.filter(user__in=_get_inventory_stock_user_ids(request)),
        request.user,
    )
    product = get_object_or_404(product_queryset, id=product_id)
    apply_stock_fields([product])
    transaction_types = InventoryTransaction.TRANSACTION_TYPES

    if request.method == "POST":
        try:
            quantity = int(request.POST.get("quantity", 0))
        except (TypeError, ValueError):
            quantity = 0

        transaction_type = request.POST.get("transaction_type", "IN")
        if transaction_type not in dict(transaction_types).keys():
            transaction_type = "IN"

        remarks = request.POST.get("remarks", "").strip()
        if not remarks:
            remarks = "QR Transaction"

        if quantity > 0:
            InventoryTransaction.objects.create(
                product=product,
                transaction_type=transaction_type,
                quantity=quantity,
                transaction_date=timezone.now(),
                remarks=remarks,
                user=request.user,
            )
            messages.success(request, "Inventory updated successfully.")
            return redirect("accounts:inventory_transactions")
        else:
            messages.error(request, "Please enter a valid quantity.")

    context = {
        "product": product,
        "transaction_types": transaction_types,
    }
    return render(request, "inventory/qr_stock_in.html", context)


@login_required
def search_inventory(request):
    """Return basic JSON results for product search autocomplete."""
    query = request.GET.get("q", "").strip()
    results = []
    if query:
        product_user_ids = _get_inventory_user_ids(request)
        products = (
            Product.objects.filter(user__in=product_user_ids)
            .filter(
                Q(name__icontains=query)
                | Q(sku__icontains=query)
                | Q(oem_part_number__icontains=query)
                | Q(barcode_value__icontains=query)
                | Q(alternate_skus__sku__icontains=query)
                | Q(description__icontains=query)
                | Q(fitment_notes__icontains=query)
                | Q(category__name__icontains=query)
                | Q(supplier__name__icontains=query)
            )
            .select_related("category", "supplier")
            .distinct()[:5]
        )

        for p in products:
            primary_code = p.sku or p.oem_part_number or p.barcode_value or "No code"
            results.append(
                {
                    "id": p.id,
                    "label": f"{p.name} ({primary_code})",
                    "type": "product",
                }
            )

    return JsonResponse({"results": results})


@login_required
def filter_options(request):
    """Return products, suppliers, and categories based on current filter selections."""
    supplier_id = request.GET.get("supplier")
    category_id = request.GET.get("category")
    product_id = request.GET.get("product")

    product_user_ids = _get_inventory_user_ids(request)
    products = Product.objects.filter(user__in=product_user_ids)
    suppliers = Supplier.objects.filter(user__in=product_user_ids)
    categories = Category.objects.filter(user__in=product_user_ids)

    if supplier_id:
        products = products.filter(supplier_id=supplier_id)
        categories = categories.filter(products__supplier_id=supplier_id).distinct()

    if category_id:
        products = products.filter(category_id=category_id)
        suppliers = suppliers.filter(products__category_id=category_id).distinct()

    if product_id:
        try:
            product = Product.objects.get(id=product_id, user=user)
            if product.supplier_id:
                suppliers = suppliers.filter(id=product.supplier_id)
            else:
                suppliers = suppliers.none()
            if product.category_id:
                categories = categories.filter(id=product.category_id)
            else:
                categories = categories.none()
        except Product.DoesNotExist:
            products = products.none()
            suppliers = suppliers.none()
            categories = categories.none()

    data = {
        "products": [{"id": p.id, "name": p.name} for p in products.distinct()],
        "suppliers": [{"id": s.id, "name": s.name} for s in suppliers.distinct()],
        "categories": [{"id": c.id, "name": c.name} for c in categories.distinct()],
    }

    return JsonResponse(data)


@login_required
def inventory_analytics(request):
    stock_user_ids = _get_inventory_stock_user_ids(request)
    transaction_user_ids = _get_inventory_transaction_user_ids(request)
    period_days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=period_days)

    products = Product.objects.filter(user__in=stock_user_ids)
    products = annotate_products_with_stock(products, request.user)
    transaction_scope = _transaction_scope_filter(transaction_user_ids)

    turnover_data = []
    for p in products:
        sold = (
            p.transactions.filter(
                transaction_scope,
                transaction_type='OUT',
                transaction_date__gte=start_date,
            )
            .aggregate(total=Sum('quantity'))['total']
            or 0
        )
        avg_stock = (p.stock_quantity + sold) / 2 if sold else p.stock_quantity
        turnover = sold / avg_stock if avg_stock else 0
        turnover_data.append({'product': p, 'qty_sold': sold, 'turnover': round(turnover, 2)})

    apply_stock_fields([entry["product"] for entry in turnover_data])
    top_sellers = sorted(turnover_data, key=lambda x: x['qty_sold'], reverse=True)[:5]
    slow_movers = sorted(turnover_data, key=lambda x: x['qty_sold'])[:5]

    value_expr_cost = ExpressionWrapper(F('stock_quantity') * F('cost_price'), output_field=DecimalField(max_digits=12, decimal_places=2))
    value_expr_sale = ExpressionWrapper(F('stock_quantity') * F('sale_price'), output_field=DecimalField(max_digits=12, decimal_places=2))

    totals = products.aggregate(
        total_cost=Sum(value_expr_cost),
        total_retail=Sum(value_expr_sale)
    )
    totals = {k: v or Decimal('0.00') for k, v in totals.items()}

    low_stock = products.filter(
        item_type='inventory',
        stock_quantity__lt=F('stock_reorder'),
    )
    unsold_days = period_days * 6
    cutoff = timezone.now() - timedelta(days=unsold_days)
    unsold = (
        products.annotate(
            last_sale=Max(
                'transactions__transaction_date',
                filter=Q(transactions__transaction_type='OUT') & _product_transaction_scope_filter(transaction_user_ids)
            )
        )
        .filter(Q(last_sale__lt=cutoff) | Q(last_sale__isnull=True))
    )

    low_stock = apply_stock_fields(list(low_stock))
    unsold = apply_stock_fields(list(unsold))

    context = {
        'turnover_data': turnover_data,
        'top_sellers': top_sellers,
        'slow_movers': slow_movers,
        'totals': totals,
        'period_days': period_days,
        'low_stock_products': low_stock,
        'unsold_products': unsold,
        'unsold_days': unsold_days,
    }

    return render(request, 'inventory/analytics_dashboard.html', context)


def _parse_date_input(raw_value):
    if not raw_value:
        return None
    try:
        return datetime.strptime(str(raw_value), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _parse_datetime_input(raw_value):
    if not raw_value:
        return None
    try:
        return datetime.strptime(str(raw_value), "%Y-%m-%dT%H:%M")
    except (TypeError, ValueError):
        return None


OPERATIONS_ACTION_CAPABILITY = {
    "create_low_stock_pos": InventoryRoleAssignment.CAP_PURCHASE_ORDERS,
    "update_po_status": InventoryRoleAssignment.CAP_PURCHASE_ORDERS,
    "receive_po_item": InventoryRoleAssignment.CAP_PURCHASE_ORDERS,
    "start_cycle_count": InventoryRoleAssignment.CAP_CYCLE_COUNTS,
    "save_cycle_counts": InventoryRoleAssignment.CAP_CYCLE_COUNTS,
    "close_cycle_count": InventoryRoleAssignment.CAP_CYCLE_COUNTS,
    "save_replenishment_rule": InventoryRoleAssignment.CAP_REPLENISHMENT,
    "create_rma": InventoryRoleAssignment.CAP_RMA,
    "update_rma_status": InventoryRoleAssignment.CAP_RMA,
    "create_fleet_list": InventoryRoleAssignment.CAP_FLEET_LISTS,
    "add_fleet_list_item": InventoryRoleAssignment.CAP_FLEET_LISTS,
    "update_margin_guardrails": InventoryRoleAssignment.CAP_MARGIN_GUARDRAILS,
    "create_dispatch": InventoryRoleAssignment.CAP_DISPATCH,
    "update_dispatch_status": InventoryRoleAssignment.CAP_DISPATCH,
    "refresh_supplier_scorecards": InventoryRoleAssignment.CAP_SUPPLIER_SCORECARDS,
    "save_role_assignment": InventoryRoleAssignment.CAP_ROLE_ADMIN,
}


@login_required
def inventory_operations_view(request):
    business_user = _get_inventory_business_user(request)
    product_user_ids = _get_inventory_user_ids(request)
    transaction_scope = _transaction_scope_filter(_get_inventory_transaction_user_ids(request))

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        capability = OPERATIONS_ACTION_CAPABILITY.get(action)
        if capability and not _can_inventory(request, capability):
            messages.error(request, "Your inventory role does not allow this operation.")
            return redirect("accounts:inventory_operations")

        if action == "save_role_assignment":
            member_id = _safe_int(request.POST.get("member_id"), minimum=1)
            role_value = (request.POST.get("role") or "").strip()
            active = request.POST.get("is_active") in {"on", "true", "1"}
            valid_roles = {choice[0] for choice in InventoryRoleAssignment._meta.get_field("role").choices}
            if role_value not in valid_roles:
                role_value = InventoryRoleAssignment.ROLE_VIEWER

            team_ids = set(get_product_user_ids(request.user))
            team_ids.add(business_user.id)
            member = User.objects.filter(id=member_id, id__in=team_ids).first()
            if not member:
                messages.error(request, "Selected team member was not found.")
                return redirect("accounts:inventory_operations")

            assignment, created = InventoryRoleAssignment.objects.update_or_create(
                business=business_user,
                member=member,
                defaults={
                    "role": role_value,
                    "is_active": active,
                },
            )
            messages.success(request, f"Role {'created' if created else 'updated'} for {member.username}.")
            _log_inventory_activity(
                request,
                action="inventory_role_assignment_saved",
                object_type="inventory_role_assignment",
                object_id=assignment.pk,
                description=f"{member.username} assigned role {role_value}",
                metadata={"active": active},
            )
            return redirect("accounts:inventory_operations")

        if action == "create_low_stock_pos":
            products = annotate_products_with_stock(
                Product.objects.filter(user__in=product_user_ids, item_type="inventory").select_related("supplier"),
                request.user,
            )
            low_stock_products = apply_stock_fields(list(products.filter(stock_quantity__lt=F("stock_reorder"))))
            rule_map = {
                rule.product_id: rule
                for rule in ReplenishmentRule.objects.filter(
                    user=business_user,
                    product_id__in=[product.id for product in low_stock_products],
                )
            }

            grouped = {}
            for product in low_stock_products:
                recommendation = _recommended_reorder_for_product(
                    product,
                    transaction_scope=transaction_scope,
                    rule=rule_map.get(product.id),
                )
                order_qty = recommendation["recommended_qty"]
                if order_qty <= 0:
                    continue

                supplier_key = product.supplier_id or 0
                if supplier_key not in grouped:
                    grouped[supplier_key] = {
                        "supplier": product.supplier,
                        "items": [],
                        "lead_days": [],
                    }
                grouped[supplier_key]["items"].append((product, recommendation))
                if rule_map.get(product.id):
                    grouped[supplier_key]["lead_days"].append(rule_map[product.id].lead_time_days or 0)

            po_created = 0
            item_created = 0
            with transaction.atomic():
                for _, payload in grouped.items():
                    lead_days = max(payload["lead_days"]) if payload["lead_days"] else 7
                    po = PurchaseOrder.objects.create(
                        user=business_user,
                        supplier=payload["supplier"],
                        status="draft",
                        expected_delivery_date=timezone.localdate() + timedelta(days=max(lead_days, 1)),
                        created_by=request.user,
                    )
                    created_any_item = False
                    for product, recommendation in payload["items"]:
                        qty = recommendation["recommended_qty"]
                        if qty <= 0:
                            continue
                        PurchaseOrderItem.objects.create(
                            purchase_order=po,
                            product=product,
                            quantity_ordered=qty,
                            recommended_quantity=qty,
                            unit_cost=product.cost_price or Decimal("0.00"),
                            notes=f"Target stock: {recommendation['target_stock']}",
                        )
                        created_any_item = True
                        item_created += 1
                    if created_any_item:
                        po_created += 1
                    else:
                        po.delete()

            if po_created:
                messages.success(request, f"Created {po_created} draft purchase order(s) with {item_created} line item(s).")
                _log_inventory_activity(
                    request,
                    action="inventory_purchase_orders_generated",
                    object_type="inventory_purchase_order",
                    description=f"Generated {po_created} draft purchase orders",
                    metadata={"po_count": po_created, "item_count": item_created},
                )
            else:
                messages.info(request, "No purchase orders were created. Review stock levels and replenishment rules.")
            return redirect("accounts:inventory_operations")

        if action == "update_po_status":
            po_id = _safe_int(request.POST.get("po_id"), minimum=1)
            status_value = (request.POST.get("status") or "").strip()
            valid_statuses = {choice[0] for choice in PurchaseOrder._meta.get_field("status").choices}
            po = PurchaseOrder.objects.filter(user=business_user, pk=po_id).first()
            if not po:
                messages.error(request, "Purchase order not found.")
                return redirect("accounts:inventory_operations")
            if status_value not in valid_statuses:
                messages.error(request, "Invalid purchase order status.")
                return redirect("accounts:inventory_operations")

            po.status = status_value
            if status_value == "ordered" and not po.ordered_date:
                po.ordered_date = timezone.localdate()
            po.save(update_fields=["status", "ordered_date", "updated_at"])
            messages.success(request, f"Purchase order {po.po_number} updated.")
            _log_inventory_activity(
                request,
                action="inventory_purchase_order_status_updated",
                object_type="inventory_purchase_order",
                object_id=po.pk,
                description=f"{po.po_number} set to {po.status}",
            )
            return redirect("accounts:inventory_operations")

        if action == "receive_po_item":
            item_id = _safe_int(request.POST.get("item_id"), minimum=1)
            receive_qty = _safe_int(request.POST.get("receive_qty"), minimum=0)
            item = PurchaseOrderItem.objects.filter(
                pk=item_id,
                purchase_order__user=business_user,
            ).select_related("purchase_order").first()
            if not item:
                messages.error(request, "Purchase order item not found.")
                return redirect("accounts:inventory_operations")
            if receive_qty <= 0:
                messages.error(request, "Enter a positive quantity to receive.")
                return redirect("accounts:inventory_operations")

            posted_qty = item.receive_stock(receive_qty, actor=request.user)
            po = item.purchase_order
            if po.total_received_qty >= po.total_ordered_qty and po.total_ordered_qty > 0:
                po.status = "received"
                po.save(update_fields=["status", "updated_at"])
            elif po.total_received_qty > 0:
                po.status = "partially_received"
                po.save(update_fields=["status", "updated_at"])

            messages.success(request, f"Received {posted_qty} unit(s) for {item.product.name}.")
            _log_inventory_activity(
                request,
                action="inventory_purchase_order_item_received",
                object_type="inventory_purchase_order_item",
                object_id=item.pk,
                description=f"Received {posted_qty} units for {item.product.name}",
                metadata={"purchase_order": po.po_number},
            )
            return redirect("accounts:inventory_operations")

        if action == "save_replenishment_rule":
            product_id = _safe_int(request.POST.get("product_id"), minimum=1)
            product = Product.objects.filter(pk=product_id, user__in=product_user_ids).first()
            if not product:
                messages.error(request, "Product not found for replenishment rule.")
                return redirect("accounts:inventory_operations")

            defaults = {
                "lead_time_days": _safe_int(request.POST.get("lead_time_days"), default=7, minimum=0),
                "coverage_days": _safe_int(request.POST.get("coverage_days"), default=30, minimum=1),
                "buffer_percent": _safe_decimal(request.POST.get("buffer_percent"), default=Decimal("10.00"), minimum=Decimal("0.00")),
                "min_order_qty": _safe_int(request.POST.get("min_order_qty"), default=0, minimum=0),
                "order_multiple": _safe_int(request.POST.get("order_multiple"), default=1, minimum=1),
                "seasonality_factor": _safe_decimal(request.POST.get("seasonality_factor"), default=Decimal("1.00"), minimum=Decimal("0.10")),
                "auto_generate_po": request.POST.get("auto_generate_po") in {"on", "true", "1"},
            }
            rule, created = ReplenishmentRule.objects.update_or_create(
                user=business_user,
                product=product,
                defaults=defaults,
            )
            messages.success(request, f"Replenishment rule {'created' if created else 'updated'} for {product.name}.")
            _log_inventory_activity(
                request,
                action="inventory_replenishment_rule_saved",
                object_type="inventory_replenishment_rule",
                object_id=rule.pk,
                description=f"Rule saved for {product.name}",
            )
            return redirect("accounts:inventory_operations")

        if action == "start_cycle_count":
            existing_open = CycleCountSession.objects.filter(user=business_user, status="open").first()
            if existing_open:
                messages.info(request, "An open cycle count session already exists.")
                return redirect("accounts:inventory_operations")

            title = (request.POST.get("cycle_title") or "").strip() or f"Cycle Count - {timezone.localdate()}"
            products = annotate_products_with_stock(
                Product.objects.filter(user__in=product_user_ids, item_type="inventory").select_related("supplier"),
                request.user,
            )
            candidates = list(products.filter(stock_reorder__gt=0).order_by("stock_quantity", "name")[:40])
            if not candidates:
                candidates = list(products.order_by("name")[:40])
            candidates = apply_stock_fields(candidates)

            with transaction.atomic():
                session = CycleCountSession.objects.create(
                    user=business_user,
                    title=title,
                    created_by=request.user,
                )
                for product in candidates:
                    CycleCountEntry.objects.create(
                        session=session,
                        product=product,
                        expected_quantity=max(int(product.quantity_in_stock or 0), 0),
                    )

            messages.success(request, f"Cycle count session started with {len(candidates)} item(s).")
            _log_inventory_activity(
                request,
                action="inventory_cycle_count_started",
                object_type="inventory_cycle_count",
                object_id=session.pk,
                description=f"Started cycle count {session.title}",
                metadata={"entry_count": len(candidates)},
            )
            return redirect("accounts:inventory_operations")

        if action == "save_cycle_counts":
            session_id = _safe_int(request.POST.get("session_id"), minimum=1)
            session = CycleCountSession.objects.filter(
                user=business_user,
                status="open",
                pk=session_id,
            ).prefetch_related("entries").first()
            if not session:
                messages.error(request, "Cycle count session not found.")
                return redirect("accounts:inventory_operations")

            updated = 0
            for entry in session.entries.all():
                count_key = f"counted_{entry.id}"
                if count_key not in request.POST:
                    continue
                raw_count = (request.POST.get(count_key) or "").strip()
                if raw_count == "":
                    continue
                counted_qty = _safe_int(raw_count, minimum=0)
                note_value = (request.POST.get(f"notes_{entry.id}") or "").strip()
                entry.apply_count(counted_qty, notes=note_value)
                updated += 1

            messages.success(request, f"Updated {updated} cycle count entry/entries.")
            _log_inventory_activity(
                request,
                action="inventory_cycle_count_saved",
                object_type="inventory_cycle_count",
                object_id=session.pk,
                description=f"Saved counts for session {session.pk}",
                metadata={"updated_entries": updated},
            )
            return redirect("accounts:inventory_operations")

        if action == "close_cycle_count":
            session_id = _safe_int(request.POST.get("session_id"), minimum=1)
            session = CycleCountSession.objects.filter(
                user=business_user,
                status="open",
                pk=session_id,
            ).first()
            if not session:
                messages.error(request, "Cycle count session not found.")
                return redirect("accounts:inventory_operations")

            session.close(actor=request.user)
            messages.success(request, "Cycle count session closed and adjustments posted.")
            _log_inventory_activity(
                request,
                action="inventory_cycle_count_closed",
                object_type="inventory_cycle_count",
                object_id=session.pk,
                description=f"Closed cycle count session {session.pk}",
            )
            return redirect("accounts:inventory_operations")

        if action == "create_rma":
            product_id = _safe_int(request.POST.get("product_id"), minimum=1)
            product = Product.objects.filter(pk=product_id, user__in=product_user_ids).first()
            if not product:
                messages.error(request, "Selected product was not found.")
                return redirect("accounts:inventory_operations")

            rma_type = (request.POST.get("rma_type") or "core").strip()
            if rma_type not in {choice[0] for choice in ProductRMA._meta.get_field("rma_type").choices}:
                rma_type = "core"
            status_value = (request.POST.get("status") or "open").strip()
            if status_value not in {choice[0] for choice in ProductRMA._meta.get_field("status").choices}:
                status_value = "open"

            supplier_id = _safe_int(request.POST.get("supplier_id"), minimum=1)
            customer_id = _safe_int(request.POST.get("customer_id"), minimum=1)
            supplier = Supplier.objects.filter(id=supplier_id, user__in=product_user_ids).first()
            customer = Customer.objects.filter(id=customer_id, user=business_user).first()

            expected_credit = None
            if request.POST.get("expected_credit"):
                expected_credit = _safe_decimal(
                    request.POST.get("expected_credit"),
                    default=Decimal("0.00"),
                    minimum=Decimal("0.00"),
                )

            rma = ProductRMA.objects.create(
                user=business_user,
                product=product,
                supplier=supplier or product.supplier,
                customer=customer,
                rma_type=rma_type,
                status=status_value,
                quantity=_safe_int(request.POST.get("quantity"), default=1, minimum=1),
                reason=(request.POST.get("reason") or "").strip(),
                supplier_reference=(request.POST.get("supplier_reference") or "").strip(),
                due_date=_parse_date_input(request.POST.get("due_date")),
                expected_credit=expected_credit,
                opened_by=request.user,
            )

            if rma.status in {"received", "credited", "closed"}:
                rma.apply_stock_adjustment(actor=request.user)

            messages.success(request, f"Created RMA {rma.rma_number}.")
            _log_inventory_activity(
                request,
                action="inventory_rma_created",
                object_type="inventory_rma",
                object_id=rma.pk,
                description=f"Created RMA {rma.rma_number} for {product.name}",
            )
            return redirect("accounts:inventory_operations")

        if action == "update_rma_status":
            rma_id = _safe_int(request.POST.get("rma_id"), minimum=1)
            rma = ProductRMA.objects.filter(user=business_user, pk=rma_id).first()
            if not rma:
                messages.error(request, "RMA not found.")
                return redirect("accounts:inventory_operations")

            status_value = (request.POST.get("status") or "").strip()
            if status_value not in {choice[0] for choice in ProductRMA._meta.get_field("status").choices}:
                messages.error(request, "Invalid RMA status.")
                return redirect("accounts:inventory_operations")

            rma.status = status_value
            if request.POST.get("actual_credit"):
                rma.actual_credit = _safe_decimal(
                    request.POST.get("actual_credit"),
                    default=Decimal("0.00"),
                    minimum=Decimal("0.00"),
                )
            rma.save(update_fields=["status", "actual_credit", "updated_at"])
            stock_adjusted = False
            if status_value in {"received", "credited", "closed"}:
                stock_adjusted = rma.apply_stock_adjustment(actor=request.user)

            messages.success(request, f"RMA {rma.rma_number} updated.")
            _log_inventory_activity(
                request,
                action="inventory_rma_updated",
                object_type="inventory_rma",
                object_id=rma.pk,
                description=f"RMA {rma.rma_number} set to {status_value}",
                metadata={"stock_adjusted": stock_adjusted},
            )
            return redirect("accounts:inventory_operations")

        if action == "create_fleet_list":
            list_name = (request.POST.get("name") or "").strip()
            if not list_name:
                messages.error(request, "Fleet list name is required.")
                return redirect("accounts:inventory_operations")
            vehicle_id = _safe_int(request.POST.get("fleet_vehicle_id"), minimum=1)
            vehicle = FleetVehicle.objects.filter(user=business_user, pk=vehicle_id).first()
            fleet_list = FleetPartList.objects.create(
                user=business_user,
                fleet_vehicle=vehicle,
                name=list_name,
                notes=(request.POST.get("notes") or "").strip(),
                created_by=request.user,
            )
            messages.success(request, f"Created fleet part list '{fleet_list.name}'.")
            _log_inventory_activity(
                request,
                action="inventory_fleet_list_created",
                object_type="inventory_fleet_list",
                object_id=fleet_list.pk,
                description=f"Created fleet list {fleet_list.name}",
            )
            return redirect("accounts:inventory_operations")

        if action == "add_fleet_list_item":
            fleet_list_id = _safe_int(request.POST.get("fleet_list_id"), minimum=1)
            product_id = _safe_int(request.POST.get("product_id"), minimum=1)
            fleet_list = FleetPartList.objects.filter(user=business_user, pk=fleet_list_id).first()
            product = Product.objects.filter(pk=product_id, user__in=product_user_ids).first()
            if not fleet_list or not product:
                messages.error(request, "Fleet list or product not found.")
                return redirect("accounts:inventory_operations")

            item, created = FleetPartListItem.objects.get_or_create(
                part_list=fleet_list,
                product=product,
                defaults={
                    "quantity": _safe_int(request.POST.get("quantity"), default=1, minimum=1),
                },
            )
            item.quantity = _safe_int(request.POST.get("quantity"), default=item.quantity or 1, minimum=1)
            item.install_interval_days = _safe_int(request.POST.get("install_interval_days"), default=0, minimum=0) or None
            item.install_interval_km = _safe_int(request.POST.get("install_interval_km"), default=0, minimum=0) or None
            item.is_required = request.POST.get("is_required") in {"on", "true", "1"}
            item.notes = (request.POST.get("notes") or "").strip()
            item.save()

            messages.success(request, f"{'Added' if created else 'Updated'} fleet list item for {product.name}.")
            _log_inventory_activity(
                request,
                action="inventory_fleet_list_item_saved",
                object_type="inventory_fleet_list_item",
                object_id=item.pk,
                description=f"Saved fleet list item {product.name}",
            )
            return redirect("accounts:inventory_operations")

        if action == "update_margin_guardrails":
            guardrail, _ = MarginGuardrailSetting.objects.get_or_create(user=business_user)
            min_margin = _safe_decimal(request.POST.get("min_margin_percent"), default=guardrail.min_margin_percent, minimum=Decimal("0.00"))
            warning_margin = _safe_decimal(request.POST.get("warning_margin_percent"), default=guardrail.warning_margin_percent, minimum=Decimal("0.00"))
            if warning_margin < min_margin:
                warning_margin = min_margin

            guardrail.min_margin_percent = min_margin
            guardrail.warning_margin_percent = warning_margin
            guardrail.enforce_min_margin = request.POST.get("enforce_min_margin") in {"on", "true", "1"}
            try:
                guardrail.full_clean()
                guardrail.save()
                messages.success(request, "Margin guardrail settings updated.")
            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))
            else:
                _log_inventory_activity(
                    request,
                    action="inventory_margin_guardrails_updated",
                    object_type="inventory_margin_guardrails",
                    object_id=guardrail.pk,
                    description="Updated margin guardrail settings",
                )
            return redirect("accounts:inventory_operations")

        if action == "create_dispatch":
            destination_name = (request.POST.get("destination_name") or "").strip()
            if not destination_name:
                messages.error(request, "Destination name is required for dispatch.")
                return redirect("accounts:inventory_operations")

            po_id = _safe_int(request.POST.get("purchase_order_id"), minimum=1)
            supplier_id = _safe_int(request.POST.get("supplier_id"), minimum=1)
            customer_id = _safe_int(request.POST.get("customer_id"), minimum=1)
            dispatch = DispatchTicket.objects.create(
                user=business_user,
                purchase_order=PurchaseOrder.objects.filter(user=business_user, pk=po_id).first(),
                supplier=Supplier.objects.filter(user__in=product_user_ids, pk=supplier_id).first(),
                customer=Customer.objects.filter(user=business_user, pk=customer_id).first(),
                destination_name=destination_name,
                destination_address=(request.POST.get("destination_address") or "").strip(),
                scheduled_at=_parse_datetime_input(request.POST.get("scheduled_at")),
                driver_name=(request.POST.get("driver_name") or "").strip(),
                vehicle_reference=(request.POST.get("vehicle_reference") or "").strip(),
                tracking_notes=(request.POST.get("tracking_notes") or "").strip(),
                created_by=request.user,
            )
            messages.success(request, f"Dispatch ticket {dispatch.reference_number} created.")
            _log_inventory_activity(
                request,
                action="inventory_dispatch_created",
                object_type="inventory_dispatch",
                object_id=dispatch.pk,
                description=f"Created dispatch ticket {dispatch.reference_number}",
            )
            return redirect("accounts:inventory_operations")

        if action == "update_dispatch_status":
            dispatch_id = _safe_int(request.POST.get("dispatch_id"), minimum=1)
            status_value = (request.POST.get("status") or "").strip()
            dispatch = DispatchTicket.objects.filter(user=business_user, pk=dispatch_id).first()
            if not dispatch:
                messages.error(request, "Dispatch ticket not found.")
                return redirect("accounts:inventory_operations")
            if status_value not in {choice[0] for choice in DispatchTicket._meta.get_field("status").choices}:
                messages.error(request, "Invalid dispatch status.")
                return redirect("accounts:inventory_operations")

            note = (request.POST.get("tracking_note") or "").strip()
            if note:
                existing_notes = dispatch.tracking_notes or ""
                dispatch.tracking_notes = f"{existing_notes}\n{note}".strip()
            dispatch.status = status_value
            dispatch.save()
            messages.success(request, f"Dispatch {dispatch.reference_number} updated.")
            _log_inventory_activity(
                request,
                action="inventory_dispatch_updated",
                object_type="inventory_dispatch",
                object_id=dispatch.pk,
                description=f"Dispatch {dispatch.reference_number} moved to {dispatch.status}",
            )
            return redirect("accounts:inventory_operations")

        if action == "refresh_supplier_scorecards":
            period_days = _safe_int(request.POST.get("period_days"), default=90, minimum=30)
            scorecards = _build_supplier_scorecards(
                business_user=business_user,
                since_date=timezone.now() - timedelta(days=period_days),
                period_days=period_days,
            )
            snapshot_date = timezone.localdate()
            saved_count = 0
            for entry in scorecards:
                SupplierScorecardSnapshot.objects.update_or_create(
                    user=business_user,
                    supplier=entry["supplier"],
                    snapshot_date=snapshot_date,
                    period_days=period_days,
                    defaults={
                        "po_count": entry["po_count"],
                        "on_time_rate": entry["on_time_rate"],
                        "fill_rate": entry["fill_rate"],
                        "avg_lead_time_days": entry["avg_lead_time_days"],
                        "rma_rate": entry["rma_rate"],
                        "weighted_score": entry["weighted_score"],
                    },
                )
                saved_count += 1
            messages.success(request, f"Refreshed supplier scorecards for {saved_count} supplier(s).")
            _log_inventory_activity(
                request,
                action="inventory_supplier_scorecards_refreshed",
                object_type="inventory_supplier_scorecard",
                description=f"Refreshed supplier scorecards ({period_days} days)",
                metadata={"supplier_count": saved_count, "period_days": period_days},
            )
            return redirect("accounts:inventory_operations")

        messages.info(request, "No operation was executed.")
        return redirect("accounts:inventory_operations")

    products = annotate_products_with_stock(
        Product.objects.filter(user__in=product_user_ids, item_type="inventory")
        .select_related("supplier")
        .order_by("name"),
        request.user,
    )
    products = apply_stock_fields(list(products))

    guardrail, _ = MarginGuardrailSetting.objects.get_or_create(user=business_user)

    rule_map = {
        rule.product_id: rule
        for rule in ReplenishmentRule.objects.filter(
            user=business_user,
            product_id__in=[product.id for product in products],
        )
    }

    low_stock_products = [product for product in products if product.quantity_in_stock < product.reorder_level]
    replenishment_rows = []
    for product in low_stock_products[:30]:
        recommendation = _recommended_reorder_for_product(
            product,
            transaction_scope=transaction_scope,
            rule=rule_map.get(product.id),
        )
        replenishment_rows.append(
            {
                "product": product,
                "rule": rule_map.get(product.id),
                "recommendation": recommendation,
            }
        )

    purchase_orders = (
        PurchaseOrder.objects.filter(user=business_user)
        .select_related("supplier")
        .prefetch_related("items__product")
        .order_by("-created_at")[:20]
    )
    open_purchase_orders = [po for po in purchase_orders if po.status not in {"received", "cancelled"}]

    open_cycle_session = (
        CycleCountSession.objects.filter(user=business_user, status="open")
        .prefetch_related("entries__product")
        .first()
    )
    recent_cycle_sessions = CycleCountSession.objects.filter(user=business_user).order_by("-started_at")[:10]

    rma_entries = (
        ProductRMA.objects.filter(user=business_user)
        .select_related("product", "supplier", "customer")
        .order_by("-opened_at")[:25]
    )

    fleet_lists = (
        FleetPartList.objects.filter(user=business_user, is_active=True)
        .select_related("fleet_vehicle")
        .prefetch_related("items__product")
        .order_by("name")[:20]
    )

    dispatch_tickets = (
        DispatchTicket.objects.filter(user=business_user)
        .select_related("purchase_order", "supplier", "customer")
        .order_by("-created_at")[:20]
    )

    scorecards = _build_supplier_scorecards(
        business_user=business_user,
        since_date=timezone.now() - timedelta(days=90),
        period_days=90,
    )
    latest_snapshots = (
        SupplierScorecardSnapshot.objects.filter(user=business_user)
        .select_related("supplier")
        .order_by("-snapshot_date", "-created_at")[:20]
    )

    team_user_ids = sorted(set(get_product_user_ids(request.user) + [business_user.id]))
    team_users = User.objects.filter(id__in=team_user_ids).order_by("username")
    role_assignments = {
        assignment.member_id: assignment
        for assignment in InventoryRoleAssignment.objects.filter(business=business_user, member_id__in=team_user_ids)
    }
    role_rows = []
    for member in team_users:
        assignment = role_assignments.get(member.id)
        resolved_role = InventoryRoleAssignment.resolve_role(business_user, member)
        role_rows.append(
            {
                "member": member,
                "assignment": assignment,
                "resolved_role": resolved_role,
            }
        )

    margin_alerts = []
    for product in products:
        margin_percent = _calculate_margin_percent(product.cost_price, product.sale_price)
        if margin_percent is None:
            continue
        if margin_percent >= guardrail.warning_margin_percent:
            continue
        margin_alerts.append(
            {
                "product": product,
                "margin_percent": margin_percent,
                "below_floor": margin_percent < guardrail.min_margin_percent,
            }
        )
    margin_alerts.sort(key=lambda entry: (entry["margin_percent"], entry["product"].name))
    margin_alerts = margin_alerts[:25]

    now_ts = timezone.now()
    dead_stock_candidates = []
    dead_stock_buckets = {
        "bucket_60_89": 0,
        "bucket_90_179": 0,
        "bucket_180_plus": 0,
    }
    dead_stock_qs = Product.objects.filter(user__in=product_user_ids, item_type="inventory").annotate(
        last_sale=Max(
            "transactions__transaction_date",
            filter=Q(transactions__transaction_type="OUT") & _product_transaction_scope_filter(_get_inventory_transaction_user_ids(request)),
        )
    )
    for product in dead_stock_qs:
        if product.last_sale:
            days_unsold = max((now_ts - product.last_sale).days, 0)
        else:
            days_unsold = 9999
        if days_unsold >= 180:
            dead_stock_buckets["bucket_180_plus"] += 1
        elif days_unsold >= 90:
            dead_stock_buckets["bucket_90_179"] += 1
        elif days_unsold >= 60:
            dead_stock_buckets["bucket_60_89"] += 1
        if days_unsold >= 60:
            dead_stock_candidates.append(
                {
                    "product": product,
                    "days_unsold": days_unsold,
                    "last_sale": product.last_sale,
                }
            )
    dead_stock_candidates.sort(key=lambda entry: (-entry["days_unsold"], entry["product"].name))
    dead_stock_candidates = dead_stock_candidates[:25]

    inventory_audit_logs = (
        ActivityLog.objects.filter(
            business=business_user,
            object_type__startswith="inventory_",
        )
        .select_related("actor")
        .order_by("-created_at")[:30]
    )

    capabilities = {
        InventoryRoleAssignment.CAP_PURCHASE_ORDERS: _can_inventory(request, InventoryRoleAssignment.CAP_PURCHASE_ORDERS),
        InventoryRoleAssignment.CAP_CYCLE_COUNTS: _can_inventory(request, InventoryRoleAssignment.CAP_CYCLE_COUNTS),
        InventoryRoleAssignment.CAP_REPLENISHMENT: _can_inventory(request, InventoryRoleAssignment.CAP_REPLENISHMENT),
        InventoryRoleAssignment.CAP_RMA: _can_inventory(request, InventoryRoleAssignment.CAP_RMA),
        InventoryRoleAssignment.CAP_FLEET_LISTS: _can_inventory(request, InventoryRoleAssignment.CAP_FLEET_LISTS),
        InventoryRoleAssignment.CAP_MARGIN_GUARDRAILS: _can_inventory(request, InventoryRoleAssignment.CAP_MARGIN_GUARDRAILS),
        InventoryRoleAssignment.CAP_DISPATCH: _can_inventory(request, InventoryRoleAssignment.CAP_DISPATCH),
        InventoryRoleAssignment.CAP_SUPPLIER_SCORECARDS: _can_inventory(request, InventoryRoleAssignment.CAP_SUPPLIER_SCORECARDS),
        InventoryRoleAssignment.CAP_ROLE_ADMIN: _can_inventory(request, InventoryRoleAssignment.CAP_ROLE_ADMIN),
    }

    context = {
        "inventory_role": _get_inventory_role(request),
        "capabilities": capabilities,
        "purchase_orders": purchase_orders,
        "open_purchase_orders": open_purchase_orders,
        "replenishment_rows": replenishment_rows,
        "open_cycle_session": open_cycle_session,
        "recent_cycle_sessions": recent_cycle_sessions,
        "rma_entries": rma_entries,
        "fleet_lists": fleet_lists,
        "dispatch_tickets": dispatch_tickets,
        "scorecards": scorecards,
        "latest_snapshots": latest_snapshots,
        "team_users": team_users,
        "role_assignments": role_assignments,
        "role_rows": role_rows,
        "guardrail": guardrail,
        "margin_alerts": margin_alerts,
        "dead_stock_buckets": dead_stock_buckets,
        "dead_stock_candidates": dead_stock_candidates,
        "inventory_audit_logs": inventory_audit_logs,
        "products_for_forms": products[:300],
        "suppliers_for_forms": Supplier.objects.filter(user__in=product_user_ids).order_by("name"),
        "customers_for_forms": Customer.objects.filter(user=business_user).order_by("name"),
        "fleet_vehicles_for_forms": FleetVehicle.objects.filter(user=business_user).order_by("truck_number", "vin_number"),
        "fleet_lists_for_forms": fleet_lists,
        "rma_type_choices": ProductRMA._meta.get_field("rma_type").choices,
        "rma_status_choices": ProductRMA._meta.get_field("status").choices,
        "dispatch_status_choices": DispatchTicket._meta.get_field("status").choices,
        "po_status_choices": PurchaseOrder._meta.get_field("status").choices,
        "role_choices": InventoryRoleAssignment._meta.get_field("role").choices,
    }
    return render(request, "inventory/operations.html", context)
