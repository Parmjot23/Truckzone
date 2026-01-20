from decimal import Decimal, InvalidOperation
from datetime import timedelta
import datetime
from io import BytesIO
import csv
import zipfile

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.db import transaction
from django.db.models import Sum, Count
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.utils.dateparse import parse_date, parse_time
from openpyxl import Workbook, load_workbook

from .forms import (
    EmployeeForm,
    EmployeeTaxProfileForm,
    EmployeeRecurringDeductionForm,
    PayrollRunForm,
    PayrollSettingsForm,
    ShiftTemplateForm,
    TimesheetForm,
    TimeEntryFormSet,
)
from .models import (
    Employee,
    EmployeeRecurringDeduction,
    EmployeeTaxProfile,
    PayStub,
    PayStubLineItem,
    PayrollRun,
    PayrollSettings,
    ShiftTemplate,
    TimeEntry,
    Timesheet,
    TimesheetSnapshot,
)
from .payroll_utils import (
    build_timesheet_snapshot_payload,
    calculate_employee_pay,
    calculate_timesheet_hours,
    compare_timesheet_entries,
    get_pay_period_for_date,
    upsert_timesheet_snapshot,
)
from .pdf_utils import render_template_to_pdf
from .utils import get_business_user, resolve_company_logo_url


def _require_payroll_admin(request):
    business_user = get_business_user(request.user)
    if request.user != business_user and not request.user.is_staff:
        return None
    return business_user


def _get_payroll_settings(user):
    settings, _ = PayrollSettings.objects.get_or_create(user=user)
    return settings


def _get_employee_tax_profile(employee):
    profile, _ = EmployeeTaxProfile.objects.get_or_create(
        employee=employee,
        defaults={
            "province": getattr(employee.user.profile, "province", "ON"),
        },
    )
    return profile


def _safe_filename(value, fallback):
    slug = slugify(value or "")
    return slug or fallback


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _parse_decimal(value):
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (TypeError, ValueError, InvalidOperation):
        return None


def _load_table_rows(upload):
    filename = (upload.name or "").lower()
    if filename.endswith(".csv"):
        data = upload.read().decode("utf-8", errors="ignore").splitlines()
        reader = csv.reader(data)
        return list(reader)

    workbook = load_workbook(upload, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _build_paystub_context(paystub, request):
    business_user = paystub.payroll_run.user
    profile = getattr(business_user, "profile", None)
    company_logo_url = resolve_company_logo_url(profile, request=request, for_pdf=True) if profile else ""
    timesheet = (
        Timesheet.objects.filter(
            employee=paystub.employee,
            period_start=paystub.payroll_run.period_start,
            period_end=paystub.payroll_run.period_end,
        )
        .prefetch_related("entries")
        .first()
    )
    time_entries = []
    if timesheet:
        time_entries = list(timesheet.entries.all().order_by("work_date", "start_time"))
    return {
        "paystub": paystub,
        "employee": paystub.employee,
        "payroll_run": paystub.payroll_run,
        "line_items": paystub.line_items.all(),
        "time_entries": time_entries,
        "business_user": business_user,
        "profile": profile,
        "company_logo_url": company_logo_url,
        "generated_on": timezone.localdate(),
        "request": request,
    }


def _render_paystub_pdf(paystub, request):
    context = _build_paystub_context(paystub, request)
    return render_template_to_pdf("payroll/paystub_pdf.html", context)


@login_required
def payroll_settings_view(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    settings = _get_payroll_settings(business_user)
    employees_qs = Employee.objects.filter(user=business_user)
    employee_count = employees_qs.count()
    active_employee_count = employees_qs.filter(status=Employee.STATUS_ACTIVE).count()

    timesheets_qs = Timesheet.objects.filter(employee__user=business_user).select_related("employee")
    timesheet_counts = {
        "draft": timesheets_qs.filter(status=Timesheet.STATUS_DRAFT).count(),
        "submitted": timesheets_qs.filter(status=Timesheet.STATUS_SUBMITTED).count(),
        "approved": timesheets_qs.filter(status=Timesheet.STATUS_APPROVED).count(),
    }
    recent_timesheets = list(timesheets_qs.order_by("-period_start")[:5])

    runs_qs = PayrollRun.objects.filter(user=business_user)
    run_counts = {
        "draft": runs_qs.filter(status=PayrollRun.STATUS_DRAFT).count(),
        "approved": runs_qs.filter(status=PayrollRun.STATUS_APPROVED).count(),
        "paid": runs_qs.filter(status=PayrollRun.STATUS_PAID).count(),
    }
    recent_runs = list(
        runs_qs.order_by("-period_start").annotate(paystub_count=Count("paystubs"))[:5]
    )

    paystubs_qs = PayStub.objects.filter(payroll_run__user=business_user).select_related(
        "employee", "payroll_run"
    )
    paystub_count = paystubs_qs.count()
    recent_paystubs = list(paystubs_qs.order_by("-payroll_run__period_end", "-id")[:5])

    last_run = recent_runs[0] if recent_runs else None
    last_run_totals = None
    if last_run:
        last_run_totals = last_run.paystubs.aggregate(
            total_gross=Sum("gross_pay"),
            total_net=Sum("net_pay"),
            total_employer=Sum("employer_total"),
        )

    attention_items = []
    if employee_count == 0:
        attention_items.append("Add at least one employee before creating timesheets.")
    if timesheet_counts["approved"] == 0:
        attention_items.append("Approve timesheets to generate paystubs.")
    run_total = sum(run_counts.values())
    if run_total == 0:
        attention_items.append("Create your first payroll run when timesheets are ready.")
    if run_total > 0 and paystub_count == 0:
        attention_items.append("Payroll runs have no paystubs yet. Recalculate or import paystubs.")
    if request.method == "POST":
        form = PayrollSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, "Payroll settings saved.")
            return redirect("accounts:payroll_settings")
        messages.error(request, "Please correct the errors below.")
    else:
        form = PayrollSettingsForm(instance=settings)
    return render(
        request,
        "payroll/settings.html",
        {
            "form": form,
            "employee_count": employee_count,
            "active_employee_count": active_employee_count,
            "timesheet_counts": timesheet_counts,
            "run_counts": run_counts,
            "paystub_count": paystub_count,
            "recent_timesheets": recent_timesheets,
            "recent_runs": recent_runs,
            "recent_paystubs": recent_paystubs,
            "last_run": last_run,
            "last_run_totals": last_run_totals,
            "attention_items": attention_items,
        },
    )


@login_required
def employee_list(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    employees = Employee.objects.filter(user=business_user).order_by("first_name", "last_name")
    return render(request, "payroll/employee_list.html", {"employees": employees})


@login_required
def employee_create(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    if request.method == "POST":
        form = EmployeeForm(request.POST, user=business_user)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.user = business_user
            employee.save()
            _get_employee_tax_profile(employee)
            messages.success(request, "Employee created.")
            return redirect("accounts:employee_detail", employee_id=employee.id)
        messages.error(request, "Please correct the errors below.")
    else:
        form = EmployeeForm(user=business_user)
    return render(request, "payroll/employee_form.html", {"form": form, "mode": "create"})


@login_required
def employee_update(request, employee_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    employee = get_object_or_404(Employee, id=employee_id, user=business_user)
    if request.method == "POST":
        form = EmployeeForm(request.POST, instance=employee, user=business_user)
        if form.is_valid():
            form.save()
            _get_employee_tax_profile(employee)
            messages.success(request, "Employee updated.")
            return redirect("accounts:employee_detail", employee_id=employee.id)
        messages.error(request, "Please correct the errors below.")
    else:
        form = EmployeeForm(instance=employee, user=business_user)
    return render(request, "payroll/employee_form.html", {"form": form, "mode": "edit", "employee": employee})


@login_required
def employee_detail(request, employee_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    employee = get_object_or_404(Employee, id=employee_id, user=business_user)
    tax_profile = _get_employee_tax_profile(employee)
    deductions = EmployeeRecurringDeduction.objects.filter(employee=employee).order_by("name")
    last_paystub = (
        PayStub.objects.filter(employee=employee)
        .select_related("payroll_run")
        .order_by("-payroll_run__period_end", "-id")
        .first()
    )
    active_deductions = list(deductions.filter(active=True))
    pre_tax_total = sum(
        (
            d.amount
            for d in active_deductions
            if d.is_employee_contribution and d.is_pre_tax
        ),
        Decimal("0.00"),
    )
    post_tax_total = sum(
        (
            d.amount
            for d in active_deductions
            if d.is_employee_contribution and not d.is_pre_tax
        ),
        Decimal("0.00"),
    )
    employer_total = sum(
        (d.amount for d in active_deductions if d.is_employer_contribution),
        Decimal("0.00"),
    )

    def _format_currency(value):
        try:
            return f"${Decimal(value):,.2f}"
        except (TypeError, ValueError, InvalidOperation):
            return "$0.00"

    def _display_amount(value, negative=False):
        amount = Decimal(value or 0)
        formatted = _format_currency(amount)
        if negative and amount > 0:
            return f"-{formatted}"
        return formatted

    def _build_line(label, value=None, negative=True, exempt=False, note=None):
        if last_paystub:
            display = _display_amount(value, negative=negative)
            is_placeholder = False
        elif exempt:
            display = _display_amount(Decimal("0.00"), negative=False)
            is_placeholder = False
        elif value is None:
            display = "Calculated at payroll run"
            is_placeholder = True
        else:
            display = _display_amount(value, negative=negative)
            is_placeholder = False
        return {
            "label": label,
            "display": display,
            "is_placeholder": is_placeholder,
            "exempt": exempt,
            "note": note,
        }

    paystub_preview_lines = [
        _build_line("Federal tax", getattr(last_paystub, "federal_tax", None)),
        _build_line("Provincial tax", getattr(last_paystub, "provincial_tax", None)),
        _build_line(
            "CPP",
            getattr(last_paystub, "cpp_employee", None),
            exempt=tax_profile.cpp_exempt,
            note="Paystub line: CPP (employee).",
        ),
        _build_line(
            "CPP2",
            getattr(last_paystub, "cpp2_employee", None),
            exempt=tax_profile.cpp2_exempt,
            note="Paystub line: CPP2 (employee).",
        ),
        _build_line(
            "EI",
            getattr(last_paystub, "ei_employee", None),
            exempt=tax_profile.ei_exempt,
            note="Paystub line: EI (employee).",
        ),
        _build_line(
            "Other deductions",
            getattr(last_paystub, "other_deductions", post_tax_total),
            note="Sum of employee recurring deductions (after tax).",
        ),
    ]

    if request.method == "POST":
        if "save_tax_profile" in request.POST:
            tax_form = EmployeeTaxProfileForm(request.POST, instance=tax_profile)
            if tax_form.is_valid():
                tax_form.save()
                messages.success(request, "Tax profile updated.")
                return redirect("accounts:employee_detail", employee_id=employee.id)
            deduction_form = EmployeeRecurringDeductionForm()
        else:
            tax_form = EmployeeTaxProfileForm(instance=tax_profile)
            deduction_form = EmployeeRecurringDeductionForm(request.POST)
            if deduction_form.is_valid():
                deduction = deduction_form.save(commit=False)
                deduction.employee = employee
                deduction.save()
                messages.success(request, "Deduction added.")
                return redirect("accounts:employee_detail", employee_id=employee.id)
    else:
        tax_form = EmployeeTaxProfileForm(instance=tax_profile)
        deduction_form = EmployeeRecurringDeductionForm()

    return render(
        request,
        "payroll/employee_detail.html",
        {
            "employee": employee,
            "tax_form": tax_form,
            "deduction_form": deduction_form,
            "deductions": deductions,
            "tax_profile": tax_profile,
            "last_paystub": last_paystub,
            "pre_tax_total": pre_tax_total,
            "post_tax_total": post_tax_total,
            "employer_total": employer_total,
            "active_deductions": active_deductions,
            "paystub_preview_lines": paystub_preview_lines,
        },
    )


@login_required
def shift_template_list(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    templates = ShiftTemplate.objects.filter(user=business_user).order_by("name")
    return render(request, "payroll/shift_template_list.html", {"templates": templates})


@login_required
def shift_template_create(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    if request.method == "POST":
        form = ShiftTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.user = business_user
            template.save()
            messages.success(request, "Shift template created.")
            return redirect("accounts:shift_template_list")
        messages.error(request, "Please correct the errors below.")
    else:
        form = ShiftTemplateForm()
    return render(request, "payroll/shift_template_form.html", {"form": form, "mode": "create"})


@login_required
def shift_template_update(request, template_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    template = get_object_or_404(ShiftTemplate, id=template_id, user=business_user)
    if request.method == "POST":
        form = ShiftTemplateForm(request.POST, instance=template)
        if form.is_valid():
            form.save()
            messages.success(request, "Shift template updated.")
            return redirect("accounts:shift_template_list")
        messages.error(request, "Please correct the errors below.")
    else:
        form = ShiftTemplateForm(instance=template)
    return render(
        request,
        "payroll/shift_template_form.html",
        {"form": form, "mode": "edit", "template": template},
    )


@login_required
def shift_template_delete(request, template_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    template = get_object_or_404(ShiftTemplate, id=template_id, user=business_user)
    if request.method == "POST":
        template.delete()
        messages.success(request, "Shift template deleted.")
        return redirect(request.POST.get("next") or "accounts:shift_template_list")
    return redirect("accounts:shift_template_list")


@login_required
def timesheet_list(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    employee_id = request.GET.get("employee_id")
    timesheets = Timesheet.objects.filter(employee__user=business_user)
    if request.method == "POST":
        action = request.POST.get("action")
        selected_ids = request.POST.getlist("timesheet_id")
        if action in ("submit", "approve") and selected_ids:
            selected_qs = Timesheet.objects.filter(
                id__in=selected_ids, employee__user=business_user
            )
            if action == "submit":
                updated = selected_qs.update(
                    status=Timesheet.STATUS_SUBMITTED, approved_at=None, approved_by=None
                )
                messages.success(request, f"Submitted {updated} timesheet(s).")
            else:
                updated = selected_qs.update(
                    status=Timesheet.STATUS_APPROVED,
                    approved_at=timezone.now(),
                    approved_by=request.user,
                )
                messages.success(request, f"Approved {updated} timesheet(s).")
        else:
            messages.error(request, "Select at least one timesheet and an action.")
        return redirect(request.POST.get("next") or "accounts:timesheet_list")

    if employee_id:
        try:
            employee_id_int = int(employee_id)
        except (TypeError, ValueError):
            employee_id_int = None
        if employee_id_int:
            timesheets = timesheets.filter(employee_id=employee_id_int)
    status_counts = {
        Timesheet.STATUS_DRAFT: 0,
        Timesheet.STATUS_SUBMITTED: 0,
        Timesheet.STATUS_APPROVED: 0,
    }
    for item in timesheets.values("status").annotate(count=Count("id")):
        status_counts[item["status"]] = item["count"]
    total_count = sum(status_counts.values())

    timesheets = list(
        timesheets.order_by("-period_start").prefetch_related("snapshots", "entries")
    )
    for timesheet in timesheets:
        snapshot_map = {snapshot.source: snapshot for snapshot in timesheet.snapshots.all()}
        mechanic_snapshot = snapshot_map.get(TimesheetSnapshot.SOURCE_MECHANIC)
        admin_snapshot = snapshot_map.get(TimesheetSnapshot.SOURCE_ADMIN)
        timesheet.has_submission = bool(mechanic_snapshot)
        timesheet.mismatch_count = 0
        timesheet.submitted_at = mechanic_snapshot.captured_at if mechanic_snapshot else None
        if mechanic_snapshot:
            if admin_snapshot:
                reference_entries = admin_snapshot.entries
            else:
                reference_entries, _ = build_timesheet_snapshot_payload(
                    timesheet, entries=timesheet.entries.all()
                )
            differences = compare_timesheet_entries(
                reference_entries, mechanic_snapshot.entries
            )
            timesheet.mismatch_count = len(differences)
    employees = Employee.objects.filter(user=business_user).order_by("first_name", "last_name")
    selected_employee_id = None
    if employee_id:
        try:
            selected_employee_id = int(employee_id)
        except (TypeError, ValueError):
            selected_employee_id = None
    return render(
        request,
        "payroll/timesheet_list.html",
        {
            "timesheets": timesheets,
            "employees": employees,
            "selected_employee_id": selected_employee_id,
            "status_counts": status_counts,
            "total_count": total_count,
        },
    )


@login_required
def timesheet_weekly_grid(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    settings = _get_payroll_settings(business_user)
    employees = Employee.objects.filter(user=business_user).order_by("first_name", "last_name")
    shift_templates = ShiftTemplate.objects.filter(user=business_user, active=True).order_by("name")

    def _normalize_week_start(date_value):
        return date_value - timedelta(days=date_value.weekday())

    def _format_time(value):
        if not value:
            return ""
        if isinstance(value, str):
            return value
        return value.strftime("%H:%M")

    raw_week_start = request.GET.get("week_start")
    if isinstance(raw_week_start, datetime.date):
        week_start = raw_week_start
    else:
        week_start = parse_date(str(raw_week_start)) if raw_week_start else None
    week_start = week_start or timezone.localdate()
    week_start = _normalize_week_start(week_start)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    copy_last_week = request.GET.get("copy") == "1"
    posted_values = {}

    if request.method == "POST":
        raw_week_start = request.POST.get("week_start")
        if isinstance(raw_week_start, datetime.date):
            week_start = raw_week_start
        else:
            week_start = parse_date(str(raw_week_start)) if raw_week_start else None
        week_start = week_start or timezone.localdate()
        week_start = _normalize_week_start(week_start)
        week_dates = [week_start + timedelta(days=i) for i in range(7)]
        week_end = week_start + timedelta(days=6)
        existing_entries = (
            TimeEntry.objects.filter(
                timesheet__employee__user=business_user,
                work_date__range=(week_start, week_end),
            )
            .select_related("timesheet", "timesheet__employee")
        )
        entry_map = {(entry.timesheet.employee_id, entry.work_date): entry for entry in existing_entries}

        errors = []
        saved_entries = 0
        deleted_entries = 0
        timesheets_touched = {}
        with transaction.atomic():
            for employee in employees:
                for work_date in week_dates:
                    start_key = f"start_time_{employee.id}_{work_date:%Y-%m-%d}"
                    end_key = f"end_time_{employee.id}_{work_date:%Y-%m-%d}"
                    start_raw = request.POST.get(start_key, "").strip()
                    end_raw = request.POST.get(end_key, "").strip()
                    posted_values[(employee.id, work_date)] = {
                        "start_time": start_raw,
                        "end_time": end_raw,
                    }

                    if not start_raw and not end_raw:
                        existing_entry = entry_map.get((employee.id, work_date))
                        if existing_entry and (existing_entry.start_time or existing_entry.end_time):
                            existing_entry.delete()
                            deleted_entries += 1
                        continue

                    if not start_raw or not end_raw:
                        errors.append(
                            f"Start and finish times are required for {employee.full_name} on {work_date}."
                        )
                        continue

                    start_time = parse_time(start_raw)
                    end_time = parse_time(end_raw)
                    if not start_time or not end_time:
                        errors.append(
                            f"Invalid time format for {employee.full_name} on {work_date}."
                        )
                        continue

                    period_start, period_end = get_pay_period_for_date(settings, work_date)
                    timesheet, created = Timesheet.objects.get_or_create(
                        employee=employee,
                        period_start=period_start,
                        period_end=period_end,
                        defaults={
                            "created_by": request.user,
                            "status": Timesheet.STATUS_APPROVED if settings.auto_approve_timesheets else Timesheet.STATUS_DRAFT,
                            "approved_at": timezone.now() if settings.auto_approve_timesheets else None,
                            "approved_by": request.user if settings.auto_approve_timesheets else None,
                        },
                    )
                    if not created and settings.auto_approve_timesheets and timesheet.status != Timesheet.STATUS_APPROVED:
                        timesheet.status = Timesheet.STATUS_APPROVED
                        timesheet.approved_at = timezone.now()
                        timesheet.approved_by = request.user
                        timesheet.save(update_fields=["status", "approved_at", "approved_by"])
                    timesheets_touched[timesheet.id] = timesheet

                    TimeEntry.objects.update_or_create(
                        timesheet=timesheet,
                        work_date=work_date,
                        defaults={
                            "start_time": start_time,
                            "end_time": end_time,
                        },
                    )
                    saved_entries += 1

        if errors:
            messages.error(request, " ".join(errors))
        else:
            for timesheet in timesheets_touched.values():
                upsert_timesheet_snapshot(
                    timesheet, TimesheetSnapshot.SOURCE_ADMIN, submitted_by=request.user
                )
            messages.success(
                request,
                f"Saved {saved_entries} entry(ies). Deleted {deleted_entries} entry(ies).",
            )
            return redirect(
                f"{reverse('accounts:timesheet_weekly_grid')}?week_start={week_start:%Y-%m-%d}"
            )

    week_end = week_start + timedelta(days=6)
    entry_map = {}
    previous_map = {}
    if request.method != "POST" or not posted_values:
        existing_entries = (
            TimeEntry.objects.filter(
                timesheet__employee__user=business_user,
                work_date__range=(week_start, week_end),
            )
            .select_related("timesheet", "timesheet__employee")
        )
        entry_map = {(entry.timesheet.employee_id, entry.work_date): entry for entry in existing_entries}

        if copy_last_week:
            previous_start = week_start - timedelta(days=7)
            previous_end = previous_start + timedelta(days=6)
            previous_entries = (
                TimeEntry.objects.filter(
                    timesheet__employee__user=business_user,
                    work_date__range=(previous_start, previous_end),
                )
                .select_related("timesheet", "timesheet__employee")
            )
            previous_map = {
                (entry.timesheet.employee_id, entry.work_date): entry for entry in previous_entries
            }

    grid_rows = []
    for employee in employees:
        days = []
        for work_date in week_dates:
            if posted_values:
                posted = posted_values.get((employee.id, work_date), {})
                start_value = posted.get("start_time", "")
                end_value = posted.get("end_time", "")
            else:
                entry = entry_map.get((employee.id, work_date))
                if not entry and copy_last_week:
                    previous_entry = previous_map.get((employee.id, work_date - timedelta(days=7)))
                    entry = previous_entry if previous_entry else None
                start_value = _format_time(entry.start_time) if entry else ""
                end_value = _format_time(entry.end_time) if entry else ""
            days.append(
                {
                    "date": work_date,
                    "start_time": start_value,
                    "end_time": end_value,
                }
            )
        grid_rows.append({"employee": employee, "days": days})

    prev_week_start = week_start - timedelta(days=7)
    next_week_start = week_start + timedelta(days=7)

    return render(
        request,
        "payroll/timesheet_weekly.html",
        {
            "week_start": week_start,
            "week_end": week_end,
            "week_dates": week_dates,
            "grid_rows": grid_rows,
            "shift_templates": shift_templates,
            "prev_week_start": prev_week_start,
            "next_week_start": next_week_start,
        },
    )


@login_required
def timesheet_bulk_entry(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    settings = _get_payroll_settings(business_user)
    employees = Employee.objects.filter(user=business_user).order_by("first_name", "last_name")
    employee_ids_set = set(employees.values_list("id", flat=True))

    period_start_value = request.POST.get("period_start") if request.method == "POST" else ""
    period_end_value = request.POST.get("period_end") if request.method == "POST" else ""
    rows = []

    if request.method == "POST":
        employee_ids = request.POST.getlist("employee_id")
        work_dates = request.POST.getlist("work_date")
        start_times = request.POST.getlist("start_time")
        end_times = request.POST.getlist("end_time")
        notes_list = request.POST.getlist("notes")

        period_start = parse_date(period_start_value)
        period_end = parse_date(period_end_value)
        errors = []

        if not period_start or not period_end:
            errors.append("Period start and end dates are required.")
        elif period_end < period_start:
            errors.append("Period end must be on or after period start.")

        max_rows = max(len(employee_ids), len(work_dates), len(start_times), len(end_times), len(notes_list))
        seen_keys = set()
        entries_to_create = []

        for idx in range(max_rows):
            employee_id = employee_ids[idx] if idx < len(employee_ids) else ""
            work_date_raw = work_dates[idx] if idx < len(work_dates) else ""
            start_raw = start_times[idx] if idx < len(start_times) else ""
            end_raw = end_times[idx] if idx < len(end_times) else ""
            notes = notes_list[idx] if idx < len(notes_list) else ""

            has_data = any([employee_id, work_date_raw, start_raw, end_raw, notes])
            if not has_data:
                continue

            rows.append(
                {
                    "employee_id": employee_id,
                    "work_date": work_date_raw,
                    "start_time": start_raw,
                    "end_time": end_raw,
                    "notes": notes,
                }
            )

            if not employee_id or not work_date_raw or not start_raw or not end_raw:
                errors.append("Each entry needs employee, date, start, and finish times.")
                continue
            try:
                employee_id_int = int(employee_id)
            except (TypeError, ValueError):
                errors.append("Invalid employee selection.")
                continue
            if employee_id_int not in employee_ids_set:
                errors.append("Invalid employee selection.")
                continue

            work_date = parse_date(work_date_raw)
            start_time = parse_time(start_raw)
            end_time = parse_time(end_raw)
            if not work_date or not start_time or not end_time:
                errors.append("Invalid date or time format in one or more entries.")
                continue

            entry_key = (employee_id_int, work_date)
            if entry_key in seen_keys:
                errors.append("Duplicate entries found for the same employee and date.")
                continue
            seen_keys.add(entry_key)

            entries_to_create.append(
                {
                    "employee_id": employee_id_int,
                    "work_date": work_date,
                    "start_time": start_time,
                    "end_time": end_time,
                    "notes": notes,
                }
            )

        if not rows:
            errors.append("Please add at least one entry.")

        if errors:
            messages.error(request, " ".join(errors))
        else:
            created_entries = 0
            employees_touched = set()
            timesheets_touched = {}
            with transaction.atomic():
                for entry in entries_to_create:
                    timesheet, created = Timesheet.objects.get_or_create(
                        employee_id=entry["employee_id"],
                        period_start=period_start,
                        period_end=period_end,
                        defaults={
                            "created_by": request.user,
                            "status": Timesheet.STATUS_APPROVED if settings.auto_approve_timesheets else Timesheet.STATUS_DRAFT,
                            "approved_at": timezone.now() if settings.auto_approve_timesheets else None,
                            "approved_by": request.user if settings.auto_approve_timesheets else None,
                        },
                    )
                    employees_touched.add(entry["employee_id"])
                    if not created and settings.auto_approve_timesheets and timesheet.status != Timesheet.STATUS_APPROVED:
                        timesheet.status = Timesheet.STATUS_APPROVED
                        timesheet.approved_at = timezone.now()
                        timesheet.approved_by = request.user
                        timesheet.save(update_fields=["status", "approved_at", "approved_by"])
                    timesheets_touched[timesheet.id] = timesheet

                    TimeEntry.objects.update_or_create(
                        timesheet=timesheet,
                        work_date=entry["work_date"],
                        defaults={
                            "start_time": entry["start_time"],
                            "end_time": entry["end_time"],
                            "notes": entry["notes"],
                        },
                    )
                    created_entries += 1

            for timesheet in timesheets_touched.values():
                upsert_timesheet_snapshot(
                    timesheet, TimesheetSnapshot.SOURCE_ADMIN, submitted_by=request.user
                )
            messages.success(
                request,
                f"Saved {created_entries} entry(ies) for {len(employees_touched)} employee(s).",
            )
            return redirect("accounts:timesheet_list")

    if not rows:
        rows = [{"employee_id": "", "work_date": "", "start_time": "", "end_time": "", "notes": ""}]

    return render(
        request,
        "payroll/timesheet_bulk.html",
        {
            "employees": employees,
            "rows": rows,
            "period_start": period_start_value,
            "period_end": period_end_value,
        },
    )


@login_required
def timesheet_export(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    export_format = request.GET.get("format", "csv").lower()
    employee_id = request.GET.get("employee_id")
    start_date = parse_date(request.GET.get("start_date"))
    end_date = parse_date(request.GET.get("end_date"))

    entries = TimeEntry.objects.filter(timesheet__employee__user=business_user).select_related(
        "timesheet", "timesheet__employee"
    )
    if employee_id:
        try:
            employee_id_int = int(employee_id)
        except (TypeError, ValueError):
            employee_id_int = None
        if employee_id_int:
            entries = entries.filter(timesheet__employee_id=employee_id_int)
    if start_date and end_date:
        entries = entries.filter(work_date__range=(start_date, end_date))
    elif start_date:
        entries = entries.filter(work_date__gte=start_date)
    elif end_date:
        entries = entries.filter(work_date__lte=end_date)

    entries = entries.order_by("timesheet__period_start", "timesheet__employee__first_name", "work_date")

    header = [
        "employee_id",
        "employee_name",
        "employee_email",
        "period_start",
        "period_end",
        "timesheet_status",
        "work_date",
        "start_time",
        "end_time",
        "hours",
        "notes",
    ]
    rows = []
    for entry in entries:
        timesheet = entry.timesheet
        employee = timesheet.employee
        rows.append(
            [
                employee.id,
                employee.full_name,
                employee.email or "",
                timesheet.period_start.isoformat(),
                timesheet.period_end.isoformat(),
                timesheet.status,
                entry.work_date.isoformat() if entry.work_date else "",
                entry.start_time.strftime("%H:%M") if entry.start_time else "",
                entry.end_time.strftime("%H:%M") if entry.end_time else "",
                f"{entry.hours}",
                entry.notes or "",
            ]
        )

    if export_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(header)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="timesheets_export.xlsx"'
        return response

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="timesheets_export.csv"'
    writer = csv.writer(response)
    writer.writerow(header)
    writer.writerows(rows)
    return response


@login_required
def timesheet_import(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    settings = _get_payroll_settings(business_user)
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Please choose a CSV or Excel file to import.")
            return redirect("accounts:timesheet_import")
        try:
            raw_rows = _load_table_rows(upload)
        except Exception:
            messages.error(request, "Unable to read the uploaded file.")
            return redirect("accounts:timesheet_import")

        if not raw_rows:
            messages.error(request, "The uploaded file is empty.")
            return redirect("accounts:timesheet_import")

        headers = [_normalize_header(value) for value in raw_rows[0]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        def row_value(row, key):
            idx = header_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return row[idx]

        created_entries = 0
        skipped_rows = 0
        error_messages = []

        timesheets_touched = {}
        with transaction.atomic():
            for row_index, row in enumerate(raw_rows[1:], start=2):
                employee_id_raw = str(row_value(row, "employee_id") or "").strip()
                employee_email_raw = str(row_value(row, "employee_email") or "").strip()
                employee_name_raw = str(row_value(row, "employee_name") or "").strip()

                employee = None
                if employee_id_raw:
                    try:
                        employee = Employee.objects.filter(
                            id=int(employee_id_raw), user=business_user
                        ).first()
                    except (TypeError, ValueError):
                        employee = None
                if not employee and employee_email_raw:
                    employee = Employee.objects.filter(
                        email__iexact=employee_email_raw, user=business_user
                    ).first()
                if not employee and employee_name_raw:
                    name_parts = employee_name_raw.split()
                    if len(name_parts) >= 2:
                        first_name = name_parts[0]
                        last_name = " ".join(name_parts[1:])
                        employee = Employee.objects.filter(
                            user=business_user,
                            first_name__iexact=first_name,
                            last_name__iexact=last_name,
                        ).first()

                if not employee:
                    skipped_rows += 1
                    if len(error_messages) < 5:
                        error_messages.append(f"Row {row_index}: employee not found.")
                    continue

                work_date = parse_date(str(row_value(row, "work_date") or ""))
                if not work_date:
                    work_date = parse_date(str(row_value(row, "date") or ""))
                if not work_date:
                    skipped_rows += 1
                    if len(error_messages) < 5:
                        error_messages.append(f"Row {row_index}: missing work_date.")
                    continue

                period_start = parse_date(str(row_value(row, "period_start") or ""))
                period_end = parse_date(str(row_value(row, "period_end") or ""))
                if not period_start or not period_end:
                    period_start, period_end = get_pay_period_for_date(settings, work_date)

                status_raw = str(row_value(row, "timesheet_status") or row_value(row, "status") or "").strip().lower()
                status = status_raw if status_raw in dict(Timesheet.STATUS_CHOICES) else None

                start_time = parse_time(str(row_value(row, "start_time") or ""))
                end_time = parse_time(str(row_value(row, "end_time") or ""))
                hours = _parse_decimal(row_value(row, "hours"))
                notes = str(row_value(row, "notes") or "").strip()

                if not (start_time and end_time) and hours is None:
                    skipped_rows += 1
                    if len(error_messages) < 5:
                        error_messages.append(f"Row {row_index}: missing times or hours.")
                    continue

                defaults = {
                    "created_by": request.user,
                    "status": Timesheet.STATUS_APPROVED if settings.auto_approve_timesheets else Timesheet.STATUS_DRAFT,
                    "approved_at": timezone.now() if settings.auto_approve_timesheets else None,
                    "approved_by": request.user if settings.auto_approve_timesheets else None,
                }
                if status:
                    defaults["status"] = status
                    defaults["approved_at"] = timezone.now() if status == Timesheet.STATUS_APPROVED else None
                    defaults["approved_by"] = request.user if status == Timesheet.STATUS_APPROVED else None

                timesheet, created = Timesheet.objects.get_or_create(
                    employee=employee,
                    period_start=period_start,
                    period_end=period_end,
                    defaults=defaults,
                )
                timesheets_touched[timesheet.id] = timesheet
                if status and timesheet.status != status:
                    timesheet.status = status
                    timesheet.approved_at = defaults["approved_at"]
                    timesheet.approved_by = defaults["approved_by"]
                    timesheet.save(update_fields=["status", "approved_at", "approved_by"])

                entry_defaults = {"notes": notes}
                if start_time and end_time:
                    entry_defaults.update({"start_time": start_time, "end_time": end_time})
                else:
                    entry_defaults.update({"start_time": None, "end_time": None, "hours": hours})

                TimeEntry.objects.update_or_create(
                    timesheet=timesheet,
                    work_date=work_date,
                    defaults=entry_defaults,
                )
                created_entries += 1

        for timesheet in timesheets_touched.values():
            upsert_timesheet_snapshot(
                timesheet, TimesheetSnapshot.SOURCE_ADMIN, submitted_by=request.user
            )
        if error_messages:
            messages.error(request, " ".join(error_messages))
        messages.success(request, f"Imported {created_entries} entries. Skipped {skipped_rows} rows.")
        return redirect("accounts:timesheet_list")

    return render(request, "payroll/timesheet_import.html")


@login_required
def timesheet_create(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    settings = _get_payroll_settings(business_user)
    employee_id = request.GET.get("employee_id")
    selected_employee_id = None
    if employee_id:
        try:
            selected_employee_id = int(employee_id)
        except (TypeError, ValueError):
            selected_employee_id = None
    if request.method == "POST":
        form = TimesheetForm(request.POST, user=business_user)
        formset = TimeEntryFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            timesheet = form.save(commit=False)
            timesheet.created_by = request.user
            if settings.auto_approve_timesheets:
                timesheet.status = Timesheet.STATUS_APPROVED
                timesheet.approved_at = timezone.now()
                timesheet.approved_by = request.user
            timesheet.save()
            formset.instance = timesheet
            formset.save()
            upsert_timesheet_snapshot(
                timesheet, TimesheetSnapshot.SOURCE_ADMIN, submitted_by=request.user
            )
            messages.success(request, "Timesheet created.")
            return redirect("accounts:timesheet_detail", timesheet_id=timesheet.id)
        messages.error(request, "Please correct the errors below.")
    else:
        form = TimesheetForm(user=business_user)
        if selected_employee_id:
            form.initial["employee"] = selected_employee_id
        formset = TimeEntryFormSet()

    return render(
        request,
        "payroll/timesheet_form.html",
        {
            "form": form,
            "formset": formset,
            "mode": "create",
            "settings": settings,
            "employees": Employee.objects.filter(user=business_user).order_by("first_name", "last_name"),
            "selected_employee_id": selected_employee_id,
            "shift_templates": ShiftTemplate.objects.filter(user=business_user, active=True).order_by("name"),
        },
    )


@login_required
def timesheet_detail(request, timesheet_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    timesheet = get_object_or_404(Timesheet, id=timesheet_id, employee__user=business_user)
    settings = _get_payroll_settings(business_user)
    if request.method == "POST":
        form = TimesheetForm(request.POST, instance=timesheet, user=business_user)
        formset = TimeEntryFormSet(request.POST, instance=timesheet)
        if form.is_valid() and formset.is_valid():
            timesheet = form.save(commit=False)
            if settings.auto_approve_timesheets:
                timesheet.status = Timesheet.STATUS_APPROVED
            if timesheet.status == Timesheet.STATUS_APPROVED and not timesheet.approved_at:
                timesheet.approved_at = timezone.now()
                timesheet.approved_by = request.user
            timesheet.save()
            formset.save()
            upsert_timesheet_snapshot(
                timesheet, TimesheetSnapshot.SOURCE_ADMIN, submitted_by=request.user
            )
            messages.success(request, "Timesheet updated.")
            return redirect("accounts:timesheet_detail", timesheet_id=timesheet.id)
        messages.error(request, "Please correct the errors below.")
    else:
        form = TimesheetForm(instance=timesheet, user=business_user)
        formset = TimeEntryFormSet(instance=timesheet)

    total_hours = (
        timesheet.entries.aggregate(total=Sum("hours")).get("total") or Decimal("0.00")
    )
    snapshot_map = {snapshot.source: snapshot for snapshot in timesheet.snapshots.all()}
    mechanic_snapshot = snapshot_map.get(TimesheetSnapshot.SOURCE_MECHANIC)
    admin_snapshot = snapshot_map.get(TimesheetSnapshot.SOURCE_ADMIN)
    comparison_label = None
    differences = []
    if mechanic_snapshot:
        if admin_snapshot:
            reference_entries = admin_snapshot.entries
            comparison_label = "Admin record"
        else:
            reference_entries, _ = build_timesheet_snapshot_payload(timesheet)
            comparison_label = "Current record"
        differences = compare_timesheet_entries(
            reference_entries, mechanic_snapshot.entries
        )

    return render(
        request,
        "payroll/timesheet_form.html",
        {
            "form": form,
            "formset": formset,
            "mode": "edit",
            "timesheet": timesheet,
            "total_hours": total_hours,
            "mechanic_snapshot": mechanic_snapshot,
            "comparison_label": comparison_label,
            "differences": differences,
            "employees": Employee.objects.filter(user=business_user).order_by("first_name", "last_name"),
            "selected_employee_id": timesheet.employee_id,
            "shift_templates": ShiftTemplate.objects.filter(user=business_user, active=True).order_by("name"),
        },
    )


@login_required
def payroll_run_list(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    runs = PayrollRun.objects.filter(user=business_user).order_by("-period_start")
    return render(request, "payroll/payroll_run_list.html", {"runs": runs})


@login_required
def payroll_run_create(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    settings = _get_payroll_settings(business_user)
    if request.method == "POST":
        form = PayrollRunForm(request.POST)
        if form.is_valid():
            payroll_run = form.save(commit=False)
            payroll_run.user = business_user
            payroll_run.created_by = request.user
            if not payroll_run.pay_date:
                payroll_run.pay_date = payroll_run.period_end + timedelta(
                    days=settings.default_pay_date_offset_days
                )
            payroll_run.save()

            try:
                _generate_paystubs_for_run(payroll_run, settings)
            except ValueError as exc:
                payroll_run.delete()
                messages.error(request, str(exc))
                return redirect("accounts:payroll_run_list")
            messages.success(request, "Payroll run created.")
            return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)
        messages.error(request, "Please correct the errors below.")
    else:
        form = PayrollRunForm()

    return render(
        request,
        "payroll/payroll_run_form.html",
        {"form": form, "settings": settings, "mode": "create"},
    )


def _generate_paystubs_for_run(payroll_run: PayrollRun, settings: PayrollSettings):
    eligible_timesheets = (
        Timesheet.objects.filter(
            employee__user=payroll_run.user,
            period_start=payroll_run.period_start,
            period_end=payroll_run.period_end,
            status__in=[Timesheet.STATUS_APPROVED, Timesheet.STATUS_SUBMITTED],
        )
        .select_related("employee", "employee__user")
        .prefetch_related("entries")
    )
    if not eligible_timesheets.exists():
        raise ValueError("No submitted or approved timesheets found for this period.")

    with transaction.atomic():
        for timesheet in eligible_timesheets:
            employee = timesheet.employee
            _get_employee_tax_profile(employee)
            entries = list(timesheet.entries.all())
            hours, regular_hours, overtime_hours = calculate_timesheet_hours(entries, settings)
            calculation = calculate_employee_pay(
                employee=employee,
                period_start=payroll_run.period_start,
                period_end=payroll_run.period_end,
                hours=hours,
                settings=settings,
                regular_hours=regular_hours,
                overtime_hours=overtime_hours,
            )

            paystub, _ = PayStub.objects.update_or_create(
                payroll_run=payroll_run,
                employee=employee,
                defaults={
                    "hours": calculation["total_hours"],
                    "regular_hours": calculation["regular_hours"],
                    "overtime_hours": calculation["overtime_hours"],
                    "regular_pay": calculation["regular_pay"],
                    "overtime_pay": calculation["overtime_pay"],
                    "gross_pay": calculation["gross_pay"],
                    "taxable_income": calculation["taxable_income"],
                    "federal_tax": calculation["federal_tax"],
                    "provincial_tax": calculation["provincial_tax"],
                    "cpp_employee": calculation["cpp_employee"],
                    "cpp_employer": calculation["cpp_employer"],
                    "cpp2_employee": calculation["cpp2_employee"],
                    "cpp2_employer": calculation["cpp2_employer"],
                    "ei_employee": calculation["ei_employee"],
                    "ei_employer": calculation["ei_employer"],
                    "other_deductions": calculation["other_deductions"],
                    "net_pay": calculation["net_pay"],
                    "employer_total": calculation["employer_total"],
                },
            )

            PayStubLineItem.objects.filter(paystub=paystub).delete()
            if calculation["other_deductions"] > 0:
                PayStubLineItem.objects.create(
                    paystub=paystub,
                    line_type=PayStubLineItem.TYPE_DEDUCTION,
                    name="Other deductions",
                    amount=calculation["other_deductions"],
                )
            for contribution in calculation.get("employer_contributions", []):
                PayStubLineItem.objects.create(
                    paystub=paystub,
                    line_type=PayStubLineItem.TYPE_BENEFIT,
                    name=contribution.name,
                    amount=Decimal("0.00"),
                    employer_amount=contribution.amount,
                )
            for tax_name, amount in calculation["employer_taxes"]:
                PayStubLineItem.objects.create(
                    paystub=paystub,
                    line_type=PayStubLineItem.TYPE_EMPLOYER_TAX,
                    name=tax_name,
                    amount=Decimal("0.00"),
                    employer_amount=amount,
                )


@login_required
def payroll_run_detail(request, run_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    payroll_run = get_object_or_404(PayrollRun, id=run_id, user=business_user)
    settings = _get_payroll_settings(business_user)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "approve":
            payroll_run.status = PayrollRun.STATUS_APPROVED
            payroll_run.approved_at = timezone.now()
            payroll_run.approved_by = request.user
            payroll_run.save(update_fields=["status", "approved_at", "approved_by"])
            messages.success(request, "Payroll run approved.")
            approve_url = reverse("accounts:payroll_run_detail", args=[payroll_run.id])
            return redirect(f"{approve_url}?paystub_actions=1")
        elif action == "mark_paid":
            payroll_run.status = PayrollRun.STATUS_PAID
            payroll_run.paid_at = timezone.now()
            payroll_run.save(update_fields=["status", "paid_at"])
            messages.success(request, "Payroll run marked as paid.")
        elif action == "recalculate":
            try:
                _generate_paystubs_for_run(payroll_run, settings)
                messages.success(request, "Payroll run recalculated.")
            except ValueError as exc:
                messages.error(request, str(exc))
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)

    paystubs = payroll_run.paystubs.select_related("employee").prefetch_related("line_items")
    totals = paystubs.aggregate(
        total_gross=Sum("gross_pay"),
        total_net=Sum("net_pay"),
        total_employer=Sum("employer_total"),
    )

    return render(
        request,
        "payroll/payroll_run_detail.html",
        {
            "payroll_run": payroll_run,
            "paystubs": paystubs,
            "totals": totals,
        },
    )


@login_required
def employee_delete(request, employee_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    employee = get_object_or_404(Employee, id=employee_id, user=business_user)
    if request.method != "POST":
        return HttpResponseForbidden()
    if employee.paystubs.exists() or employee.timesheets.exists():
        messages.error(request, "Employee has payroll history. Set status to inactive instead.")
        return redirect(request.POST.get("next") or "accounts:employee_detail", employee_id=employee.id)
    employee.delete()
    messages.success(request, "Employee deleted.")
    return redirect(request.POST.get("next") or "accounts:employee_list")


@login_required
def timesheet_delete(request, timesheet_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    timesheet = get_object_or_404(Timesheet, id=timesheet_id, employee__user=business_user)
    if request.method != "POST":
        return HttpResponseForbidden()
    has_paystub = PayStub.objects.filter(
        employee=timesheet.employee,
        payroll_run__period_start=timesheet.period_start,
        payroll_run__period_end=timesheet.period_end,
    ).exists()
    if has_paystub:
        messages.error(request, "Timesheet is already tied to a payroll run and cannot be deleted.")
        return redirect(request.POST.get("next") or "accounts:timesheet_detail", timesheet_id=timesheet.id)
    timesheet.delete()
    messages.success(request, "Timesheet deleted.")
    return redirect(request.POST.get("next") or "accounts:timesheet_list")


@login_required
def timesheet_employee_redirect(request):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    employee_id = request.GET.get("employee_id")
    if not employee_id:
        return redirect("accounts:timesheet_list")
    try:
        employee_id_int = int(employee_id)
    except (TypeError, ValueError):
        return redirect("accounts:timesheet_list")

    timesheet = (
        Timesheet.objects.filter(employee_id=employee_id_int, employee__user=business_user)
        .order_by("-period_start")
        .first()
    )
    if timesheet:
        return redirect("accounts:timesheet_detail", timesheet_id=timesheet.id)
    create_url = reverse("accounts:timesheet_create")
    return redirect(f"{create_url}?employee_id={employee_id_int}")


@login_required
def payroll_run_update(request, run_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    payroll_run = get_object_or_404(PayrollRun, id=run_id, user=business_user)
    if payroll_run.status != PayrollRun.STATUS_DRAFT:
        messages.error(request, "Only draft payroll runs can be edited.")
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)
    settings = _get_payroll_settings(business_user)
    if request.method == "POST":
        form = PayrollRunForm(request.POST, instance=payroll_run)
        if form.is_valid():
            payroll_run = form.save(commit=False)
            if not payroll_run.pay_date:
                payroll_run.pay_date = payroll_run.period_end + timedelta(
                    days=settings.default_pay_date_offset_days
                )
            payroll_run.save()
            try:
                _generate_paystubs_for_run(payroll_run, settings)
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)
            messages.success(request, "Payroll run updated.")
            return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)
        messages.error(request, "Please correct the errors below.")
    else:
        form = PayrollRunForm(instance=payroll_run)

    return render(
        request,
        "payroll/payroll_run_form.html",
        {"form": form, "settings": settings, "mode": "edit"},
    )


@login_required
def payroll_run_delete(request, run_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    payroll_run = get_object_or_404(PayrollRun, id=run_id, user=business_user)
    if request.method != "POST":
        return HttpResponseForbidden()
    if payroll_run.status != PayrollRun.STATUS_DRAFT:
        messages.error(request, "Only draft payroll runs can be deleted.")
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)
    payroll_run.delete()
    messages.success(request, "Payroll run deleted.")
    return redirect(request.POST.get("next") or "accounts:payroll_run_list")


@login_required
def payroll_run_paystubs_download(request, run_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    payroll_run = get_object_or_404(PayrollRun, id=run_id, user=business_user)
    if payroll_run.status not in (PayrollRun.STATUS_APPROVED, PayrollRun.STATUS_PAID):
        messages.error(request, "Payroll run must be approved before downloading paystubs.")
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)

    paystubs = list(
        payroll_run.paystubs.select_related("employee").prefetch_related("line_items")
    )
    if not paystubs:
        messages.error(request, "No paystubs available for this payroll run.")
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for paystub in paystubs:
            employee = paystub.employee
            employee_label = _safe_filename(employee.full_name, f"employee-{employee.id}")
            filename = f"paystub_{employee_label}_{payroll_run.period_end:%Y%m%d}.pdf"
            try:
                pdf_bytes = _render_paystub_pdf(paystub, request)
            except ImportError as exc:
                messages.error(request, str(exc))
                return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)
            archive.writestr(filename, pdf_bytes)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = (
        f'attachment; filename="paystubs_{payroll_run.period_end:%Y%m%d}.zip"'
    )
    return response


@login_required
def payroll_run_paystubs_send(request, run_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    payroll_run = get_object_or_404(PayrollRun, id=run_id, user=business_user)
    if request.method != "POST":
        return HttpResponseForbidden()
    if payroll_run.status not in (PayrollRun.STATUS_APPROVED, PayrollRun.STATUS_PAID):
        messages.error(request, "Payroll run must be approved before sending paystubs.")
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)

    paystubs = payroll_run.paystubs.select_related("employee").prefetch_related("line_items")
    if not paystubs.exists():
        messages.error(request, "No paystubs available for this payroll run.")
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)

    sent_count = 0
    missing_emails = []
    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", None)
    subject = f"Pay statement {payroll_run.period_start} - {payroll_run.period_end}"

    for paystub in paystubs:
        employee = paystub.employee
        recipient = (employee.email or "").strip()
        if not recipient:
            missing_emails.append(employee.full_name or f"Employee {employee.id}")
            continue
        try:
            pdf_bytes = _render_paystub_pdf(paystub, request)
        except ImportError as exc:
            messages.error(request, str(exc))
            return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)

        email_context = {
            "employee": employee,
            "payroll_run": payroll_run,
            "business_user": business_user,
            "profile": getattr(business_user, "profile", None),
        }
        html_content = render_to_string("emails/paystub_email.html", email_context)
        body = (
            "Please find your pay statement attached.\n"
            "If you have questions, contact your employer."
        )
        message = EmailMultiAlternatives(
            subject=subject,
            body=body,
            from_email=from_email,
            to=[recipient],
        )
        message.attach_alternative(html_content, "text/html")
        employee_label = _safe_filename(employee.full_name, f"employee-{employee.id}")
        filename = f"paystub_{employee_label}_{payroll_run.period_end:%Y%m%d}.pdf"
        message.attach(filename, pdf_bytes, "application/pdf")
        message.send()
        sent_count += 1

    if sent_count:
        messages.success(request, f"Paystubs sent to {sent_count} employee(s).")
    if missing_emails:
        missing_preview = ", ".join(missing_emails[:5])
        suffix = "..." if len(missing_emails) > 5 else ""
        messages.warning(
            request,
            f"Missing email for {len(missing_emails)} employee(s): {missing_preview}{suffix}"
        )

    return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)


@login_required
def payroll_run_paystubs_export(request, run_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    payroll_run = get_object_or_404(PayrollRun, id=run_id, user=business_user)
    export_format = request.GET.get("format", "csv").lower()

    paystubs = payroll_run.paystubs.select_related("employee").order_by("employee__first_name", "employee__last_name")
    header = [
        "employee_id",
        "employee_name",
        "employee_email",
        "period_start",
        "period_end",
        "hours",
        "regular_hours",
        "overtime_hours",
        "regular_pay",
        "overtime_pay",
        "gross_pay",
        "taxable_income",
        "federal_tax",
        "provincial_tax",
        "cpp_employee",
        "cpp2_employee",
        "ei_employee",
        "other_deductions",
        "net_pay",
        "cpp_employer",
        "cpp2_employer",
        "ei_employer",
        "employer_total",
    ]
    rows = []
    for stub in paystubs:
        employee = stub.employee
        rows.append(
            [
                employee.id,
                employee.full_name,
                employee.email or "",
                payroll_run.period_start.isoformat(),
                payroll_run.period_end.isoformat(),
                f"{stub.hours}",
                f"{stub.regular_hours}",
                f"{stub.overtime_hours}",
                f"{stub.regular_pay}",
                f"{stub.overtime_pay}",
                f"{stub.gross_pay}",
                f"{stub.taxable_income}",
                f"{stub.federal_tax}",
                f"{stub.provincial_tax}",
                f"{stub.cpp_employee}",
                f"{stub.cpp2_employee}",
                f"{stub.ei_employee}",
                f"{stub.other_deductions}",
                f"{stub.net_pay}",
                f"{stub.cpp_employer}",
                f"{stub.cpp2_employer}",
                f"{stub.ei_employer}",
                f"{stub.employer_total}",
            ]
        )

    if export_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(header)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="paystubs_{payroll_run.period_end:%Y%m%d}.xlsx"'
        )
        return response

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="paystubs_{payroll_run.period_end:%Y%m%d}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(header)
    writer.writerows(rows)
    return response


@login_required
def payroll_run_paystubs_import(request, run_id):
    business_user = _require_payroll_admin(request)
    if not business_user:
        return HttpResponseForbidden()
    payroll_run = get_object_or_404(PayrollRun, id=run_id, user=business_user)
    if payroll_run.status != PayrollRun.STATUS_DRAFT:
        messages.error(request, "Paystubs can only be imported for draft payroll runs.")
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)

    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Please choose a CSV or Excel file to import.")
            return redirect("accounts:payroll_run_paystubs_import", run_id=payroll_run.id)
        try:
            raw_rows = _load_table_rows(upload)
        except Exception:
            messages.error(request, "Unable to read the uploaded file.")
            return redirect("accounts:payroll_run_paystubs_import", run_id=payroll_run.id)

        if not raw_rows:
            messages.error(request, "The uploaded file is empty.")
            return redirect("accounts:payroll_run_paystubs_import", run_id=payroll_run.id)

        headers = [_normalize_header(value) for value in raw_rows[0]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        def row_value(row, key):
            idx = header_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return row[idx]

        updated_count = 0
        skipped_rows = 0
        error_messages = []

        with transaction.atomic():
            for row_index, row in enumerate(raw_rows[1:], start=2):
                employee_id_raw = str(row_value(row, "employee_id") or "").strip()
                employee_email_raw = str(row_value(row, "employee_email") or "").strip()
                employee_name_raw = str(row_value(row, "employee_name") or "").strip()

                employee = None
                if employee_id_raw:
                    try:
                        employee = Employee.objects.filter(
                            id=int(employee_id_raw), user=business_user
                        ).first()
                    except (TypeError, ValueError):
                        employee = None
                if not employee and employee_email_raw:
                    employee = Employee.objects.filter(
                        email__iexact=employee_email_raw, user=business_user
                    ).first()
                if not employee and employee_name_raw:
                    name_parts = employee_name_raw.split()
                    if len(name_parts) >= 2:
                        first_name = name_parts[0]
                        last_name = " ".join(name_parts[1:])
                        employee = Employee.objects.filter(
                            user=business_user,
                            first_name__iexact=first_name,
                            last_name__iexact=last_name,
                        ).first()

                if not employee:
                    skipped_rows += 1
                    if len(error_messages) < 5:
                        error_messages.append(f"Row {row_index}: employee not found.")
                    continue

                regular_hours = _parse_decimal(row_value(row, "regular_hours")) or Decimal("0.00")
                overtime_hours = _parse_decimal(row_value(row, "overtime_hours")) or Decimal("0.00")
                hours = _parse_decimal(row_value(row, "hours"))
                if hours is None:
                    hours = regular_hours + overtime_hours
                if regular_hours == Decimal("0.00") and overtime_hours == Decimal("0.00"):
                    regular_hours = hours

                regular_pay = _parse_decimal(row_value(row, "regular_pay")) or Decimal("0.00")
                overtime_pay = _parse_decimal(row_value(row, "overtime_pay")) or Decimal("0.00")
                gross_pay = _parse_decimal(row_value(row, "gross_pay"))
                if gross_pay is None:
                    gross_pay = regular_pay + overtime_pay

                taxable_income = _parse_decimal(row_value(row, "taxable_income")) or gross_pay
                federal_tax = _parse_decimal(row_value(row, "federal_tax")) or Decimal("0.00")
                provincial_tax = _parse_decimal(row_value(row, "provincial_tax")) or Decimal("0.00")
                cpp_employee = _parse_decimal(row_value(row, "cpp_employee")) or Decimal("0.00")
                cpp2_employee = _parse_decimal(row_value(row, "cpp2_employee")) or Decimal("0.00")
                ei_employee = _parse_decimal(row_value(row, "ei_employee")) or Decimal("0.00")
                other_deductions = _parse_decimal(row_value(row, "other_deductions")) or Decimal("0.00")
                net_pay = _parse_decimal(row_value(row, "net_pay"))
                if net_pay is None:
                    net_pay = gross_pay - federal_tax - provincial_tax - cpp_employee - cpp2_employee - ei_employee - other_deductions

                cpp_employer = _parse_decimal(row_value(row, "cpp_employer")) or Decimal("0.00")
                cpp2_employer = _parse_decimal(row_value(row, "cpp2_employer")) or Decimal("0.00")
                ei_employer = _parse_decimal(row_value(row, "ei_employer")) or Decimal("0.00")
                employer_total = _parse_decimal(row_value(row, "employer_total"))
                if employer_total is None:
                    employer_total = cpp_employer + cpp2_employer + ei_employer

                PayStub.objects.update_or_create(
                    payroll_run=payroll_run,
                    employee=employee,
                    defaults={
                        "hours": hours,
                        "regular_hours": regular_hours,
                        "overtime_hours": overtime_hours,
                        "regular_pay": regular_pay,
                        "overtime_pay": overtime_pay,
                        "gross_pay": gross_pay,
                        "taxable_income": taxable_income,
                        "federal_tax": federal_tax,
                        "provincial_tax": provincial_tax,
                        "cpp_employee": cpp_employee,
                        "cpp2_employee": cpp2_employee,
                        "ei_employee": ei_employee,
                        "other_deductions": other_deductions,
                        "net_pay": net_pay,
                        "cpp_employer": cpp_employer,
                        "cpp2_employer": cpp2_employer,
                        "ei_employer": ei_employer,
                        "employer_total": employer_total,
                    },
                )
                updated_count += 1

        if error_messages:
            messages.error(request, " ".join(error_messages))
        messages.success(request, f"Imported {updated_count} paystub(s). Skipped {skipped_rows} rows.")
        return redirect("accounts:payroll_run_detail", run_id=payroll_run.id)

    return render(
        request,
        "payroll/paystub_import.html",
        {"payroll_run": payroll_run},
    )
